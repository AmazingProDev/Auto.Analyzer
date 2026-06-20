"""
Minimal HTTP server for Optim Analyzer (no external frameworks).

Routes consumed by the frontend:
- POST /api/trp/import            multipart/form-data file=...
- POST /api/nemo/lte/import       multipart/form-data file=... (one or more Nemo LTE .txt exports)
- POST /api/nmfs/decode           multipart/form-data file=... (external converter bridge)
- GET  /api/nmfs/config           effective converter configuration
- POST /api/nmfs/config           save converter configuration
- POST /api/nmfs/config/test      validate converter command
- GET  /api/runs                  list runs
- GET  /api/runs/<id>             run + track + events
- GET  /api/runs/<id>/catalog     signal catalog (names)
- GET  /api/runs/<id>/sidebar     sidebar groups
- GET  /api/runs/<id>/signals     signal catalog (same as catalog.signals)
- GET  /api/runs/<id>/timeseries?signal=<name>&max_points=<int>
- GET  /api/runs/<id>/track
- GET  /api/runs/<id>/events
- GET  /api/runs/<id>/neighbors_at_time?time=<ISO>&tolMs=200&bucketMs=80
- GET  /api/runs/<id>/pilot_pollution_at_point?time=<ISO>&lat=<float>&lng=<float>&windowMs=12000
- GET  /api/runs/<id>/l1l2/capabilities
- GET  /api/runs/<id>/l1l2/at_time?time=<ISO>&windowMs=2000
- GET  /api/bdd/status             BDD load state + cell count
- GET  /api/bdd/config             saved BDD path config
- GET  /api/bdd/sectors            serve bdd_sectors.json or parse xlsx on-the-fly
- GET  /api/benchmark/status       benchmark workbook load status
- GET  /api/benchmark-nemo/status  Nemo TXT benchmark load status
- GET  /api/benchmark-mycom/status Mycom hourly benchmark-context load status
- POST /api/bdd/configure          body: {path}  — load BDD from Excel
- POST /api/bdd/upload             multipart file upload → save + write bdd_sectors.json
- POST /api/bdd/sectors-cache      body: {sectors} — write bdd_sectors.json from client parse
- POST /api/bdd/match              body: {eventLat,eventLon,earfcn,polluters,maxDistM}
- POST /api/benchmark/load         body: {path?} — load and normalize benchmark workbook
- POST /api/benchmark/upload       multipart file upload → save + normalize benchmark workbook
- POST /api/benchmark-nemo/load    body: {paths:[...]} — load and analyze Nemo TXT benchmark files
- POST /api/benchmark-nemo/upload  multipart files upload → save + analyze Nemo TXT benchmark files
- POST /api/benchmark-mycom/upload multipart file upload → save + normalize Mycom hourly stats
- POST /api/benchmark-mycom/correlate correlate current Nemo benchmark with current Mycom hourly stats
"""

from __future__ import annotations

import bisect
import json
import os
import hashlib
import math
import csv
import re
import shlex
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import traceback
import uuid
import pickle
import zlib
from datetime import datetime as _dt, timedelta as _td
from functools import lru_cache
from http.server import HTTPServer, SimpleHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs

from local_ai import LMStudioClient, analyze_uploaded_log, load_local_ai_config
from local_ai.lmstudio_client import LMStudioClientError, LMStudioModelUnavailableError
from local_ai.pilot_pollution_ai import analyze as _analyze_pilot_pollution_ai
from local_ai.pipeline import LocalAIUploadError
from trp_importer import (
    import_trp_file,
    list_runs,
    fetch_run_detail,
    fetch_kpi_series,
    fetch_neighbors_at_time,
    fetch_mrdc_cells_at_time,
    analyze_pilot_pollution_at_point,
    fetch_l1l2_scheduler_capabilities,
    fetch_l1l2_scheduler_at_time,
    fetch_run_catalog,
    fetch_run_sidebar,
    fetch_run_signals,
    fetch_timeseries_by_signal,
    fetch_run_track,
    fetch_run_events,
    fetch_signaling_window_decode,
    scan_route_pilot_pollution,
    fetch_run_pcap,
)
from lte_rrc_per_decoder import (
    decode_measurement_report_payload,
    decode_rrc_event_payload,
    decode_rrc_reconfiguration_payload,
)
import bdd_matcher as _bdd
from nemo_lte_importer import parse_nemo_lte_files, register_nemo_lte_run, parse_rrc_files_deferred

UPLOAD_DIR = os.environ.get("OPTIM_UPLOAD_DIR", "/tmp/optim_uploads")
DB_PATH = None  # kept for backward compatibility; in-memory store ignores it
NMFS_CONFIG_PATH = os.environ.get("OPTIM_NMFS_CONFIG_PATH", os.path.join(UPLOAD_DIR, "nmfs_converter_config.json"))
# BDD config lives in a persistent user-writable dir so it survives /tmp cleanup
_BDD_STORE_DIR = os.path.join(os.path.expanduser("~"), ".optim_analyzer")
BDD_CONFIG_PATH = os.environ.get("OPTIM_BDD_CONFIG_PATH", os.path.join(_BDD_STORE_DIR, "bdd_config.json"))
BENCHMARK_CONFIG_PATH = os.environ.get("OPTIM_BENCHMARK_CONFIG_PATH", os.path.join(_BDD_STORE_DIR, "benchmark_config.json"))
BENCHMARK_NEMO_CONFIG_PATH = os.environ.get("OPTIM_BENCHMARK_NEMO_CONFIG_PATH", os.path.join(_BDD_STORE_DIR, "benchmark_nemo_config.json"))
BENCHMARK_MYCOM_CONFIG_PATH = os.environ.get("OPTIM_BENCHMARK_MYCOM_CONFIG_PATH", os.path.join(_BDD_STORE_DIR, "benchmark_mycom_config.json"))
BENCHMARK_NEMO_LIBRARY_DB_PATH = os.environ.get(
    "OPTIM_BENCHMARK_NEMO_LIBRARY_DB_PATH",
    os.path.join(_BDD_STORE_DIR, "benchmark_nemo_library.sqlite3"),
)
# Bump ONLY when the row-level parser (column extraction) changes — it is part of the
# dataset cache key, so a bump misses the SQLite cache and forces a fresh re-parse of the
# TXT files. (Analysis-only changes use _BENCHMARK_NEMO_ANALYSIS_VERSION, which rebuilds KPIs
# from the already-parsed rows without re-parsing.)
_BENCHMARK_NEMO_PARSER_VERSION = 7
BENCHMARK_DEFAULT_PATH = os.environ.get(
    "OPTIM_BENCHMARK_PATH",
    "/Users/abdelilah/Desktop/AutoAnalyzer IAM/Benchmark/DR Rabat - Benchmark Avril 2026.xlsx",
)
# Pre-generated sector JSON served as a static file — persists in the app directory
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
BDD_SECTORS_JSON_PATH = os.path.join(_APP_DIR, "bdd_sectors.json")
HO_ANALYSIS_STORE = {}
HO_ANALYSIS_SEQ = 0
LTE_RRC_PRECOMPUTE_STORE = {}
LTE_RRC_PRECOMPUTE_DIR = os.path.join(UPLOAD_DIR, "lte_rrc_precompute_cache")
BENCHMARK_DATASET = {"path": "", "data": None, "loaded_at": None}
BENCHMARK_NEMO_DATASET = {
    "paths": [],
    "path_mtimes": {},
    "data": None,
    "loaded_at": None,
    "operator_files": [],
    "mode_datasets": {},
    "mode_dataset_ids": {},
    "mode_dataset_keys": {},
    "dt_datasets": {},
    "dataset_id": None,
    "dataset_key": "",
    "dl_mode": "app_rate_dl",
    "window_mode": "all_dt_session",
}
BENCHMARK_MYCOM_DATASET = {"path": "", "data": None, "loaded_at": None}

# ── BDD cells cache for serving-cell matching (loaded once in background) ─────
import threading as _threading
_BDD_SERVING_CELLS: list = []
_BDD_SERVING_CELLS_LOCK = _threading.Lock()
_BDD_SERVING_CELLS_STATE: dict = {"loaded": False, "loading": False, "error": None}
_BDD_SERVING_4G_PATH = "/Users/abdelilah/Documents/My projects/Sites and Data/Sites/BDD_Mensuel_M04.xlsx"
_BDD_SERVING_5G_PATH = "/Users/abdelilah/Documents/My projects/Sites and Data/Sites/BDD_5G.xlsx"
# Local stored copy of the 5G BDD (persists next to bdd_current.xlsx)
_BDD_STORE_5G_PATH = os.path.join(_BDD_STORE_DIR, "bdd_5g.xlsx")


def _load_serving_bdd_background():
    """Load 4G and 5G BDD cells into the module-level cache (called in a daemon thread)."""
    global _BDD_SERVING_CELLS, _BDD_SERVING_CELLS_STATE
    with _BDD_SERVING_CELLS_LOCK:
        if _BDD_SERVING_CELLS_STATE["loaded"] or _BDD_SERVING_CELLS_STATE["loading"]:
            return
        _BDD_SERVING_CELLS_STATE["loading"] = True

    def _nc(s):
        return str(s).lower().replace(" ", "").replace("_", "").replace(".", "").replace("/", "")

    def _gv(row, col_idx, names):
        for nm in names:
            i = col_idx.get(_nc(nm))
            if i is not None and i < len(row):
                v = row[i]
                if v is not None and str(v).strip() != "":
                    return v
        return None

    def _sf(v):
        try:
            return float(v) if v is not None else None
        except Exception:
            return None

    def _si(v):
        try:
            return int(v) if v is not None else None
        except Exception:
            return None

    def _load_one(path, techs):
        if not path or not os.path.isfile(path):
            return []
        cells = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return []
        for sheet_name in wb.sheetnames:
            tech = str(sheet_name).strip().upper()
            if tech not in [t.upper() for t in techs]:
                continue
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            raw_header = next(rows_iter, None)
            if raw_header is None:
                continue
            col_idx = {_nc(str(h)): i for i, h in enumerate(raw_header) if h is not None}
            for row in rows_iter:
                lat = _sf(_gv(row, col_idx, ["latitude", "lat"]))
                lon = _sf(_gv(row, col_idx, ["longitude", "long", "lon", "lng"]))
                if lat is None or lon is None:
                    continue
                if tech == "4G":
                    pci = _si(_gv(row, col_idx, ["physical cell id", "pci"]))
                    cell_name = str(_gv(row, col_idx, ["cellname", "cell name"]) or "").strip()
                    site_name = str(_gv(row, col_idx, ["basestationname", "site name", "sitename", "nodeb name"]) or "").strip()
                elif tech == "5G":
                    pci = _si(_gv(row, col_idx, ["pci", "physical cell id"]))
                    cell_name = str(_gv(row, col_idx, ["cellname", "cell name"]) or "").strip()
                    site_name = str(_gv(row, col_idx, ["site name", "sitename", "gnb name", "gnbname"]) or "").strip()
                else:
                    continue
                if not cell_name and not site_name:
                    continue
                cells.append({"tech": tech, "pci": pci, "lat": lat, "lon": lon,
                               "cell_name": cell_name, "site_name": site_name})
        return cells

    try:
        cells = _load_one(_BDD_SERVING_4G_PATH, ["4G"]) + _load_one(_BDD_SERVING_5G_PATH, ["5G"])
        with _BDD_SERVING_CELLS_LOCK:
            _BDD_SERVING_CELLS = cells
            _BDD_SERVING_CELLS_STATE["loaded"] = True
            _BDD_SERVING_CELLS_STATE["loading"] = False
            _BDD_SERVING_CELLS_STATE["error"] = None
    except Exception as exc:
        with _BDD_SERVING_CELLS_LOCK:
            _BDD_SERVING_CELLS_STATE["loading"] = False
            _BDD_SERVING_CELLS_STATE["error"] = str(exc)


def _ensure_serving_bdd_loaded():
    """Trigger background BDD load if not yet started. Non-blocking."""
    with _BDD_SERVING_CELLS_LOCK:
        if _BDD_SERVING_CELLS_STATE["loaded"] or _BDD_SERVING_CELLS_STATE["loading"]:
            return
    t = _threading.Thread(target=_load_serving_bdd_background, daemon=True, name="bdd-serving-loader")
    t.start()



def _find_bdd_folder_xlsx() -> str:
    """Return the first .xlsx file found in the BDD/ subfolder of the app directory."""
    bdd_dir = os.path.join(_APP_DIR, "BDD")
    if not os.path.isdir(bdd_dir):
        return ""
    for name in sorted(os.listdir(bdd_dir)):
        if name.lower().endswith((".xlsx", ".xls")) and not name.startswith("~$"):
            return os.path.join(bdd_dir, name)
    return ""


def _bdd_load_saved_path() -> None:
    """Auto-load BDD on startup. Priority: OPTIM_BDD_PATH env → BDD/ folder → saved config."""
    # 1. Explicit env-var override
    env_path = os.environ.get("OPTIM_BDD_PATH", "").strip()
    if env_path and os.path.isfile(env_path):
        _bdd.load_bdd(env_path)
        threading.Thread(target=_write_bdd_sectors_json, args=(env_path,), daemon=True).start()
        return

    # 2. BDD/ subfolder inside the project (place any .xlsx there)
    folder_path = _find_bdd_folder_xlsx()
    if folder_path:
        _bdd.load_bdd(folder_path)
        threading.Thread(target=_write_bdd_sectors_json, args=(folder_path,), daemon=True).start()
        return

    # 3. Saved config from a previous "Update BDD" upload
    if os.path.isfile(BDD_CONFIG_PATH):
        try:
            with open(BDD_CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            saved_path = str(cfg.get("path") or "").strip()
            if saved_path and os.path.isfile(saved_path):
                _bdd.load_bdd(saved_path)
                threading.Thread(target=_write_bdd_sectors_json, args=(saved_path,), daemon=True).start()
        except Exception:
            pass


def _write_bdd_sectors_json(path: str) -> None:
    """Parse multi-sheet BDD xlsx and write bdd_sectors.json + .gz to the app directory."""
    import gzip as _gzip
    import tempfile
    import shutil
    import os
    try:
        sectors = []
        result = _bdd.get_map_sectors(path)
        if result.get("ok") and result.get("sectors"):
            sectors.extend(result["sectors"])

        # Resolve 5G BDD path: same dir → local store → known external path
        dir_name = os.path.dirname(path)
        candidate_5g = os.path.join(dir_name, "BDD_5G.xlsx")
        if not os.path.isfile(candidate_5g) or candidate_5g == path:
            candidate_5g = _BDD_STORE_5G_PATH
        if not os.path.isfile(candidate_5g):
            candidate_5g = _BDD_SERVING_5G_PATH
        path_5g = candidate_5g if os.path.isfile(candidate_5g) and candidate_5g != path else None

        if path_5g:
            # Keep local stored copy up-to-date
            if path_5g != _BDD_STORE_5G_PATH and os.path.isfile(path_5g):
                try:
                    os.makedirs(os.path.dirname(_BDD_STORE_5G_PATH), exist_ok=True)
                    shutil.copy2(path_5g, _BDD_STORE_5G_PATH)
                except Exception:
                    pass
            result_5g = _bdd.get_map_sectors(path_5g)
            if result_5g.get("ok") and result_5g.get("sectors"):
                sectors.extend(result_5g["sectors"])
                
        if sectors:
            data = json.dumps(sectors).encode("utf-8")
            
            # Write to temporary files first to prevent corruption and race conditions
            fd_json, tmp_json = tempfile.mkstemp(dir=os.path.dirname(BDD_SECTORS_JSON_PATH), suffix=".json")
            with open(fd_json, "wb") as f:
                f.write(data)
                
            fd_gz, tmp_gz = tempfile.mkstemp(dir=os.path.dirname(BDD_SECTORS_JSON_PATH), suffix=".gz")
            with open(fd_gz, "wb") as f:
                pass
            with _gzip.open(tmp_gz, "wb", compresslevel=1) as f:
                f.write(data)
                
            os.replace(tmp_json, BDD_SECTORS_JSON_PATH)
            os.replace(tmp_gz, BDD_SECTORS_JSON_PATH + ".gz")
    except Exception as e:
        print(f"Error writing BDD sectors JSON: {e}")


def _benchmark_saved_path() -> str:
    if os.path.isfile(BENCHMARK_CONFIG_PATH):
        try:
            with open(BENCHMARK_CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            saved_path = str(cfg.get("path") or "").strip()
            if saved_path:
                return saved_path
        except Exception:
            pass
    return ""


def _benchmark_resolve_path(explicit_path: str = "") -> str:
    candidate = str(explicit_path or "").strip()
    if candidate:
        return candidate
    saved_path = _benchmark_saved_path()
    if saved_path:
        return saved_path
    return BENCHMARK_DEFAULT_PATH


def _benchmark_text(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat") and not isinstance(value, str):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value).strip()


def _benchmark_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        num = float(value)
        return num if math.isfinite(num) else None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        num = float(text)
    except Exception:
        return None
    return num if math.isfinite(num) else None


def _nemo_normalize_value(value, kpi_type: str):
    num = _benchmark_float(value)
    if num is None:
        return None, False
    kind = str(kpi_type or "").strip().lower()
    if abs(num) <= 1e10:
        return num, False

    def _pick_candidate(scales, min_value=None, max_value=None):
        for scale in scales:
            candidate = num / float(scale)
            if min_value is not None and candidate < min_value:
                continue
            if max_value is not None and candidate > max_value:
                continue
            return candidate
        return None

    normalized = None
    if kind in ("longitude", "lon", "lng"):
        normalized = _pick_candidate((1e16, 1e15), -180.0, 180.0)
    elif kind in ("latitude", "lat"):
        normalized = _pick_candidate((1e15, 1e16), -90.0, 90.0)
    elif kind in ("rsrp", "rsrq", "sinr", "bler", "retx", "retransmission"):
        normalized = _pick_candidate((1e15, 1e16, 1e14), -300.0, 300.0)

    if normalized is None:
        return num, False
    return normalized, True


def _benchmark_int(value):
    num = _benchmark_float(value)
    if num is None:
        return None
    return int(round(num))


def _benchmark_operator(dl_avg, dl_max, ul_avg, ul_max, freq):
    return {
        "dlAvg": _benchmark_float(dl_avg),
        "dlMax": _benchmark_float(dl_max),
        "ulAvg": _benchmark_float(ul_avg),
        "ulMax": _benchmark_float(ul_max),
        "freq": _benchmark_int(freq),
    }


def _derive_benchmark_dr(path: str) -> str:
    base_name = os.path.splitext(os.path.basename(path or ""))[0]
    match = re.search(r"(?i)\bdr\s+([^-_]+)", base_name)
    if not match:
        return ""
    label = re.sub(r"\s+", " ", match.group(1)).strip(" _-")
    return f"DR {label.title()}" if label else ""


def _parse_benchmark_workbook(path: str) -> dict:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Global"] if "Global" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    dr_label = _derive_benchmark_dr(path)
    points = []
    city_stats = {}

    for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        values = list(row[:27])
        if len(values) < 27:
            values.extend([None] * (27 - len(values)))
        if not any(v not in (None, "") for v in values):
            continue

        ville = _benchmark_text(values[1])
        adresse = _benchmark_text(values[3])
        logfile = _benchmark_text(values[13])
        if not (ville or adresse or logfile):
            continue

        lng = _benchmark_float(values[4])
        lat = _benchmark_float(values[5])
        has_coordinates = lat is not None and lng is not None

        operators = {
            "IAM": _benchmark_operator(values[6], values[7], values[19], values[20], values[16]),
            "ORANGE": _benchmark_operator(values[8], values[9], values[21], values[22], values[17]),
            "INWI": _benchmark_operator(values[10], values[11], values[23], values[24], values[18]),
        }
        best_dl_operator = ""
        best_dl_value = None
        for operator_name, operator_values in operators.items():
            current_dl = operator_values.get("dlAvg")
            if current_dl is None:
                continue
            if best_dl_value is None or current_dl > best_dl_value:
                best_dl_value = current_dl
                best_dl_operator = operator_name

        label_parts = [ville, adresse or logfile or f"Point {len(points) + 1}"]
        point_label = " - ".join([part for part in label_parts if part])
        point_id = "::".join([part for part in [dr_label, ville, adresse or logfile or str(row_number), str(row_number)] if part])

        points.append({
            "id": point_id,
            "rowNumber": row_number,
            "dr": dr_label,
            "vendor": _benchmark_text(values[0]),
            "ville": ville,
            "dateTest": _benchmark_text(values[2]),
            "adresse": adresse,
            "pointLabel": point_label,
            "lng": lng,
            "lat": lat,
            "hasCoordinates": has_coordinates,
            "logfile": logfile,
            "servingCell": _benchmark_text(values[14]),
            "servingFrequency": _benchmark_text(values[15]),
            "rankingDlIam": _benchmark_int(values[12]),
            "rankingUlIam": _benchmark_int(values[25]),
            "bestDlOperator": best_dl_operator,
            "operators": operators,
        })

        city_key = ville or "(blank)"
        stats = city_stats.setdefault(city_key, {"ville": city_key, "count": 0, "mappedCount": 0})
        stats["count"] += 1
        if has_coordinates:
            stats["mappedCount"] += 1

    workbook.close()

    mapped_count = sum(1 for point in points if point.get("hasCoordinates"))
    return {
        "name": os.path.basename(path),
        "path": path,
        "dr": dr_label,
        "pointCount": len(points),
        "mappedCount": mapped_count,
        "missingCoordinatesCount": max(0, len(points) - mapped_count),
        "filters": {
            "dr": sorted({point["dr"] for point in points if point.get("dr")}),
            "ville": sorted({point["ville"] for point in points if point.get("ville")}, key=str.upper),
            "adresse": sorted({point["adresse"] for point in points if point.get("adresse")}, key=str.upper),
        },
        "citiesSummary": sorted(city_stats.values(), key=lambda item: str(item.get("ville") or "").upper()),
        "points": points,
    }


def _benchmark_status_payload() -> dict:
    dataset = BENCHMARK_DATASET.get("data") or {}
    return {
        "loaded": bool(BENCHMARK_DATASET.get("data")),
        "path": BENCHMARK_DATASET.get("path") or "",
        "configuredPath": _benchmark_saved_path(),
        "defaultPath": BENCHMARK_DEFAULT_PATH,
        "pointCount": int(dataset.get("pointCount") or 0),
        "mappedCount": int(dataset.get("mappedCount") or 0),
        "missingCoordinatesCount": int(dataset.get("missingCoordinatesCount") or 0),
    }


def _load_benchmark_workbook(explicit_path: str = "") -> dict:
    target_path = _benchmark_resolve_path(explicit_path)
    if not target_path:
        return {"ok": False, "error": "No benchmark workbook path configured"}
    if not os.path.isfile(target_path):
        return {"ok": False, "error": f"File not found: {target_path}"}

    dataset = _parse_benchmark_workbook(target_path)
    BENCHMARK_DATASET["path"] = target_path
    BENCHMARK_DATASET["data"] = dataset
    BENCHMARK_DATASET["loaded_at"] = time.time()

    try:
        os.makedirs(os.path.dirname(BENCHMARK_CONFIG_PATH), exist_ok=True)
        with open(BENCHMARK_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump({"path": target_path}, f)
    except Exception:
        pass

    return {"ok": True, "path": target_path, "dataset": dataset}


def _benchmark_nemo_saved_paths() -> list[str]:
    if os.path.isfile(BENCHMARK_NEMO_CONFIG_PATH):
        try:
            with open(BENCHMARK_NEMO_CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            paths = cfg.get("paths") if isinstance(cfg, dict) else []
            if isinstance(paths, list):
                return [str(path).strip() for path in paths if str(path or "").strip()]
        except Exception:
            pass
    return []


def _benchmark_nemo_upload_dir_paths() -> list[str]:
    """Fallback source: the Nemo TXT files sitting in the upload dir. Used when the saved
    config points at stale/temp paths that no longer exist (e.g. after a reboot or a temp-file
    cleanup), so a reload still works instead of failing with 'no files found'."""
    folder = os.path.join(UPLOAD_DIR, "benchmark_nemo")
    try:
        names = sorted(os.listdir(folder))
    except OSError:
        return []
    out = []
    for name in names:
        if not name.lower().endswith((".txt", ".csv", ".tsv")):
            continue
        if _nemo_is_session_stats_file(name):
            continue
        out.append(os.path.join(folder, name))
    return out


def _benchmark_nemo_resolve_paths(explicit_paths=None) -> list[str]:
    if isinstance(explicit_paths, list):
        paths = [str(path).strip() for path in explicit_paths if str(path or "").strip()]
        if paths:
            return paths
    saved = _benchmark_nemo_saved_paths()
    existing = [p for p in saved if os.path.isfile(p)]
    if existing:
        return existing
    # Saved config is empty or stale (points at files that no longer exist) → fall back to
    # whatever Nemo TXT files are in the upload dir so a reload still succeeds.
    fallback = _benchmark_nemo_upload_dir_paths()
    return fallback or saved


# DL throughput is computed only from "App. rate DL" (instantaneous appDlMbps). The old
# "App rate DL avg" mode was removed at the user's request; any legacy value normalizes
# to app_rate_dl below.
_BENCHMARK_NEMO_DL_MODE_DEFAULT = "app_rate_dl"
_BENCHMARK_NEMO_DL_MODE_LABELS = {
    "app_rate_dl": "App. rate DL",
}
_BENCHMARK_NEMO_WINDOW_MODE_DEFAULT = "all_dt_session"
_BENCHMARK_NEMO_WINDOW_MODE_LABELS = {
    "all_dt_session": "All DT session",
    "active_dl_session": "Active DL session",
}


def _benchmark_nemo_normalize_dl_mode(dl_mode: str | None) -> str:
    value = str(dl_mode or "").strip().lower()
    if value in _BENCHMARK_NEMO_DL_MODE_LABELS:
        return value
    return _BENCHMARK_NEMO_DL_MODE_DEFAULT


def _benchmark_nemo_dl_mode_label(dl_mode: str | None) -> str:
    mode = _benchmark_nemo_normalize_dl_mode(dl_mode)
    return _BENCHMARK_NEMO_DL_MODE_LABELS.get(mode, _BENCHMARK_NEMO_DL_MODE_LABELS[_BENCHMARK_NEMO_DL_MODE_DEFAULT])


def _benchmark_nemo_normalize_window_mode(window_mode: str | None) -> str:
    value = str(window_mode or "").strip().lower()
    if value in _BENCHMARK_NEMO_WINDOW_MODE_LABELS:
        return value
    return _BENCHMARK_NEMO_WINDOW_MODE_DEFAULT


def _benchmark_nemo_window_mode_label(window_mode: str | None) -> str:
    mode = _benchmark_nemo_normalize_window_mode(window_mode)
    return _BENCHMARK_NEMO_WINDOW_MODE_LABELS.get(mode, _BENCHMARK_NEMO_WINDOW_MODE_LABELS[_BENCHMARK_NEMO_WINDOW_MODE_DEFAULT])


def _benchmark_nemo_mode_cache_key(dl_mode: str | None, window_mode: str | None) -> str:
    return f"{_benchmark_nemo_normalize_dl_mode(dl_mode)}::{_benchmark_nemo_normalize_window_mode(window_mode)}"


def _benchmark_nemo_cache_get(cache: dict | None, dl_mode: str | None, window_mode: str | None):
    if not isinstance(cache, dict):
        return None
    mode = _benchmark_nemo_normalize_dl_mode(dl_mode)
    mode_key = _benchmark_nemo_mode_cache_key(mode, window_mode)
    if mode_key in cache:
        return cache.get(mode_key)
    legacy_key = mode if _benchmark_nemo_normalize_window_mode(window_mode) == _BENCHMARK_NEMO_WINDOW_MODE_DEFAULT else None
    if legacy_key and legacy_key in cache:
        return cache.get(legacy_key)
    return None


def _benchmark_nemo_canonical_dl_metric_key(operator_file: dict, dl_mode: str | None) -> str:
    rows = (operator_file or {}).get("rows") or []
    mode = _benchmark_nemo_normalize_dl_mode(dl_mode)
    if mode == "app_rate_dl":
        return (
            (operator_file or {}).get("_dlMetricKeyOverride")
            or (operator_file or {}).get("dlMetricKey")
            or _nemo_select_dl_metric_key(rows)
        )
    return (
        (operator_file or {}).get("_benchmarkDlMetricKeyAvgOverride")
        or (operator_file or {}).get("_benchmarkDlMetricKeyOverride")
        or (operator_file or {}).get("benchmarkDlMetricKeyDefault")
        or (operator_file or {}).get("benchmarkDlMetricKey")
        or _nemo_select_benchmark_dl_metric_key(rows)
    )


def _benchmark_nemo_status_payload() -> dict:
    dataset = BENCHMARK_NEMO_DATASET.get("data") or {}
    try:
        library_count = len(_benchmark_nemo_library_list())
    except Exception:
        library_count = 0
    return {
        "loaded": bool(BENCHMARK_NEMO_DATASET.get("data")),
        "paths": list(BENCHMARK_NEMO_DATASET.get("paths") or []),
        "configuredPaths": _benchmark_nemo_saved_paths(),
        "operatorCount": int(dataset.get("operatorCount") or 0),
        "testCount": int(dataset.get("testCount") or 0),
        "datasetId": BENCHMARK_NEMO_DATASET.get("dataset_id"),
        "datasetKey": BENCHMARK_NEMO_DATASET.get("dataset_key") or "",
        "dlMode": dataset.get("dlMode") or BENCHMARK_NEMO_DATASET.get("dl_mode") or "app_rate_dl",
        "windowMode": dataset.get("windowMode") or BENCHMARK_NEMO_DATASET.get("window_mode") or _BENCHMARK_NEMO_WINDOW_MODE_DEFAULT,
        "libraryCount": library_count,
    }


def _benchmark_nemo_collect_mtimes(paths: list[str]) -> dict:
    mtimes = {}
    for path in paths or []:
        try:
            mtimes[path] = os.path.getmtime(path)
        except OSError:
            mtimes[path] = None
    return mtimes


def _benchmark_nemo_save_paths(paths: list[str]):
    try:
        os.makedirs(os.path.dirname(BENCHMARK_NEMO_CONFIG_PATH), exist_ok=True)
        with open(BENCHMARK_NEMO_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump({"paths": list(paths or [])}, f)
    except Exception:
        pass


def _benchmark_nemo_library_connect():
    os.makedirs(os.path.dirname(BENCHMARK_NEMO_LIBRARY_DB_PATH), exist_ok=True)
    conn = sqlite3.connect(BENCHMARK_NEMO_LIBRARY_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _benchmark_nemo_library_init():
    with _benchmark_nemo_library_connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS nemo_benchmark_datasets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dataset_key TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                created_at REAL NOT NULL,
                last_loaded_at REAL NOT NULL,
                operator_count INTEGER NOT NULL DEFAULT 0,
                test_count INTEGER NOT NULL DEFAULT 0,
                transfer_session_count INTEGER NOT NULL DEFAULT 0,
                dt_count INTEGER NOT NULL DEFAULT 0,
                dataset_blob BLOB NOT NULL,
                operator_files_blob BLOB NOT NULL
            );

            CREATE TABLE IF NOT EXISTS nemo_benchmark_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dataset_id INTEGER NOT NULL,
                ordinal INTEGER NOT NULL,
                operator TEXT,
                file_name TEXT,
                path TEXT,
                size_bytes INTEGER,
                mtime REAL,
                sha256 TEXT,
                FOREIGN KEY(dataset_id) REFERENCES nemo_benchmark_datasets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS nemo_benchmark_dts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dataset_id INTEGER NOT NULL,
                dt_index INTEGER NOT NULL,
                label TEXT NOT NULL,
                titles_by_operator_json TEXT NOT NULL,
                FOREIGN KEY(dataset_id) REFERENCES nemo_benchmark_datasets(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_nemo_benchmark_files_dataset_id
                ON nemo_benchmark_files(dataset_id, ordinal);
            CREATE INDEX IF NOT EXISTS idx_nemo_benchmark_dts_dataset_id
                ON nemo_benchmark_dts(dataset_id, dt_index);
            """
        )


def _benchmark_nemo_pack_blob(value) -> bytes:
    return sqlite3.Binary(zlib.compress(pickle.dumps(value, protocol=pickle.HIGHEST_PROTOCOL), 6))


def _benchmark_nemo_unpack_blob(blob: bytes):
    if blob is None:
        return None
    return pickle.loads(zlib.decompress(blob))


def _benchmark_nemo_collect_file_meta(paths: list[str], uploaded_hashes: dict | None = None) -> list[dict]:
    metas = []
    uploaded_hashes = uploaded_hashes or {}
    for ordinal, path in enumerate(paths or []):
        file_name = os.path.basename(path)
        operator = _nemo_guess_operator(path)
        try:
            stat = os.stat(path)
            size_bytes = int(stat.st_size)
            mtime = float(stat.st_mtime)
        except OSError:
            size_bytes = None
            mtime = None
        metas.append({
            "ordinal": ordinal,
            "operator": operator,
            "fileName": file_name,
            "path": path,
            "sizeBytes": size_bytes,
            "mtime": mtime,
            "sha256": str(uploaded_hashes.get(path) or ""),
        })
    return metas


def _benchmark_nemo_dataset_key(
    file_metas: list[dict],
    dl_mode: str | None = None,
    window_mode: str | None = None,
) -> str:
    payload = {
        "parserVersion": _BENCHMARK_NEMO_PARSER_VERSION,
        "dlMode": _benchmark_nemo_normalize_dl_mode(dl_mode),
        "windowMode": _benchmark_nemo_normalize_window_mode(window_mode),
        "files": [
            {
                "ordinal": int(meta.get("ordinal") or 0),
                "operator": str(meta.get("operator") or ""),
                "fileName": str(meta.get("fileName") or ""),
                "path": str(meta.get("path") or ""),
                "sizeBytes": meta.get("sizeBytes"),
                "mtime": meta.get("mtime"),
                "sha256": str(meta.get("sha256") or ""),
            }
            for meta in (file_metas or [])
        ],
    }
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _benchmark_nemo_library_store_dataset(dataset_key: str, file_metas: list[dict], dataset: dict, operator_files: list[dict]) -> int | None:
    if not dataset_key or not dataset:
        return None
    _benchmark_nemo_library_init()
    now = time.time()
    with _benchmark_nemo_library_connect() as conn:
        row = conn.execute(
            "SELECT id, created_at FROM nemo_benchmark_datasets WHERE dataset_key = ?",
            (dataset_key,),
        ).fetchone()
        created_at = float(row["created_at"]) if row else now
        if row:
            dataset_id = int(row["id"])
            conn.execute(
                """
                UPDATE nemo_benchmark_datasets
                   SET name = ?, last_loaded_at = ?, operator_count = ?, test_count = ?,
                       transfer_session_count = ?, dt_count = ?, dataset_blob = ?, operator_files_blob = ?
                 WHERE id = ?
                """,
                (
                    str(dataset.get("name") or "Nemo TXT Benchmark"),
                    now,
                    int(dataset.get("operatorCount") or 0),
                    int(dataset.get("testCount") or 0),
                    int(dataset.get("transferSessionCount") or 0),
                    len(dataset.get("dtList") or []),
                    _benchmark_nemo_pack_blob(dataset),
                    _benchmark_nemo_pack_blob(operator_files),
                    dataset_id,
                ),
            )
        else:
            cur = conn.execute(
                """
                INSERT INTO nemo_benchmark_datasets (
                    dataset_key, name, created_at, last_loaded_at, operator_count, test_count,
                    transfer_session_count, dt_count, dataset_blob, operator_files_blob
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    dataset_key,
                    str(dataset.get("name") or "Nemo TXT Benchmark"),
                    created_at,
                    now,
                    int(dataset.get("operatorCount") or 0),
                    int(dataset.get("testCount") or 0),
                    int(dataset.get("transferSessionCount") or 0),
                    len(dataset.get("dtList") or []),
                    _benchmark_nemo_pack_blob(dataset),
                    _benchmark_nemo_pack_blob(operator_files),
                ),
            )
            dataset_id = int(cur.lastrowid)

        conn.execute("DELETE FROM nemo_benchmark_files WHERE dataset_id = ?", (dataset_id,))
        conn.execute("DELETE FROM nemo_benchmark_dts WHERE dataset_id = ?", (dataset_id,))
        for meta in file_metas or []:
            conn.execute(
                """
                INSERT INTO nemo_benchmark_files (
                    dataset_id, ordinal, operator, file_name, path, size_bytes, mtime, sha256
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    dataset_id,
                    int(meta.get("ordinal") or 0),
                    str(meta.get("operator") or ""),
                    str(meta.get("fileName") or ""),
                    str(meta.get("path") or ""),
                    meta.get("sizeBytes"),
                    meta.get("mtime"),
                    str(meta.get("sha256") or ""),
                ),
            )
        for dt in dataset.get("dtList") or []:
            conn.execute(
                """
                INSERT INTO nemo_benchmark_dts (
                    dataset_id, dt_index, label, titles_by_operator_json
                ) VALUES (?, ?, ?, ?)
                """,
                (
                    dataset_id,
                    int(dt.get("index") or 0),
                    str(dt.get("label") or ""),
                    json.dumps(dt.get("titlesByOperator") or {}, ensure_ascii=False, sort_keys=True),
                ),
            )
    return dataset_id


def _benchmark_nemo_library_load_dataset_row(where_sql: str, params: tuple):
    _benchmark_nemo_library_init()
    with _benchmark_nemo_library_connect() as conn:
        row = conn.execute(
            f"""
            SELECT id, dataset_key, name, created_at, last_loaded_at, operator_count, test_count,
                   transfer_session_count, dt_count, dataset_blob, operator_files_blob
              FROM nemo_benchmark_datasets
             WHERE {where_sql}
            """,
            params,
        ).fetchone()
        if not row:
            return None
        conn.execute(
            "UPDATE nemo_benchmark_datasets SET last_loaded_at = ? WHERE id = ?",
            (time.time(), int(row["id"])),
        )
        file_rows = conn.execute(
            """
            SELECT ordinal, operator, file_name, path, size_bytes, mtime, sha256
              FROM nemo_benchmark_files
             WHERE dataset_id = ?
             ORDER BY ordinal ASC
            """,
            (int(row["id"]),),
        ).fetchall()
        dt_rows = conn.execute(
            """
            SELECT dt_index, label, titles_by_operator_json
              FROM nemo_benchmark_dts
             WHERE dataset_id = ?
             ORDER BY dt_index ASC
            """,
            (int(row["id"]),),
        ).fetchall()
    return {
        "id": int(row["id"]),
        "datasetKey": str(row["dataset_key"] or ""),
        "dataset": _benchmark_nemo_unpack_blob(row["dataset_blob"]),
        "operatorFiles": _benchmark_nemo_unpack_blob(row["operator_files_blob"]),
        "files": [
            {
                "ordinal": int(file_row["ordinal"] or 0),
                "operator": str(file_row["operator"] or ""),
                "fileName": str(file_row["file_name"] or ""),
                "path": str(file_row["path"] or ""),
                "sizeBytes": file_row["size_bytes"],
                "mtime": file_row["mtime"],
                "sha256": str(file_row["sha256"] or ""),
            }
            for file_row in file_rows
        ],
        "dtList": [
            {
                "index": int(dt_row["dt_index"] or 0),
                "label": str(dt_row["label"] or ""),
                "titlesByOperator": json.loads(str(dt_row["titles_by_operator_json"] or "{}")),
            }
            for dt_row in dt_rows
        ],
    }


def _benchmark_nemo_library_load_dataset_by_key(dataset_key: str):
    if not dataset_key:
        return None
    return _benchmark_nemo_library_load_dataset_row("dataset_key = ?", (dataset_key,))


def _benchmark_nemo_library_load_dataset_by_id(dataset_id: int):
    return _benchmark_nemo_library_load_dataset_row("id = ?", (int(dataset_id),))


def _benchmark_nemo_library_list() -> list[dict]:
    _benchmark_nemo_library_init()
    with _benchmark_nemo_library_connect() as conn:
        dataset_rows = conn.execute(
            """
            SELECT id, dataset_key, name, created_at, last_loaded_at, operator_count, test_count,
                   transfer_session_count, dt_count
              FROM nemo_benchmark_datasets
             ORDER BY last_loaded_at DESC, id DESC
            """
        ).fetchall()
        file_rows = conn.execute(
            """
            SELECT dataset_id, ordinal, operator, file_name, path, size_bytes, mtime, sha256
              FROM nemo_benchmark_files
             ORDER BY dataset_id ASC, ordinal ASC
            """
        ).fetchall()
        dt_rows = conn.execute(
            """
            SELECT dataset_id, dt_index, label, titles_by_operator_json
              FROM nemo_benchmark_dts
             ORDER BY dataset_id ASC, dt_index ASC
            """
        ).fetchall()
    files_by_dataset = {}
    for row in file_rows:
        files_by_dataset.setdefault(int(row["dataset_id"]), []).append({
            "ordinal": int(row["ordinal"] or 0),
            "operator": str(row["operator"] or ""),
            "fileName": str(row["file_name"] or ""),
            "path": str(row["path"] or ""),
            "sizeBytes": row["size_bytes"],
            "mtime": row["mtime"],
            "sha256": str(row["sha256"] or ""),
        })
    dts_by_dataset = {}
    for row in dt_rows:
        dts_by_dataset.setdefault(int(row["dataset_id"]), []).append({
            "index": int(row["dt_index"] or 0),
            "label": str(row["label"] or ""),
            "titlesByOperator": json.loads(str(row["titles_by_operator_json"] or "{}")),
        })
    datasets = []
    for row in dataset_rows:
        dataset_id = int(row["id"])
        datasets.append({
            "id": dataset_id,
            "datasetKey": str(row["dataset_key"] or ""),
            "name": str(row["name"] or ""),
            "createdAt": row["created_at"],
            "lastLoadedAt": row["last_loaded_at"],
            "operatorCount": int(row["operator_count"] or 0),
            "testCount": int(row["test_count"] or 0),
            "transferSessionCount": int(row["transfer_session_count"] or 0),
            "dtCount": int(row["dt_count"] or 0),
            "sourceFiles": files_by_dataset.get(dataset_id, []),
            "dtList": dts_by_dataset.get(dataset_id, []),
        })
    return datasets


def _benchmark_nemo_library_load_into_memory(record: dict):
    if not isinstance(record, dict):
        return None
    operator_files = record.get("operatorFiles") or []
    record_dataset = record.get("dataset") or {}
    dl_mode = _benchmark_nemo_normalize_dl_mode(record_dataset.get("dlMode"))
    window_mode = _benchmark_nemo_normalize_window_mode(record_dataset.get("windowMode"))
    # Rebuild from the already-parsed operator_files when the stored blob predates the
    # current analysis version (DT-weighted DL average, Deep Benchmark); else just
    # ensure deepBenchmark exists. No TXT re-parse either way.
    dataset, rebuilt = _benchmark_nemo_refresh_dataset(
        record_dataset,
        operator_files,
        dl_mode=dl_mode,
        window_mode=window_mode,
    )
    if not rebuilt:
        dataset = _ensure_deep_benchmark(dataset)
    files = record.get("files") or []
    paths = [str(item.get("path") or "") for item in files if str(item.get("path") or "")]
    BENCHMARK_NEMO_DATASET["paths"] = paths
    BENCHMARK_NEMO_DATASET["path_mtimes"] = _benchmark_nemo_collect_mtimes(paths)
    BENCHMARK_NEMO_DATASET["data"] = dataset
    BENCHMARK_NEMO_DATASET["loaded_at"] = time.time()
    BENCHMARK_NEMO_DATASET["operator_files"] = operator_files
    cache_key = _benchmark_nemo_mode_cache_key(dataset.get("dlMode"), dataset.get("windowMode"))
    BENCHMARK_NEMO_DATASET["mode_datasets"] = {cache_key: dataset}
    BENCHMARK_NEMO_DATASET["mode_dataset_ids"] = {cache_key: record.get("id")}
    BENCHMARK_NEMO_DATASET["mode_dataset_keys"] = {cache_key: str(record.get("datasetKey") or "")}
    BENCHMARK_NEMO_DATASET["dt_datasets"] = {}
    BENCHMARK_NEMO_DATASET["dataset_id"] = record.get("id")
    BENCHMARK_NEMO_DATASET["dataset_key"] = str(record.get("datasetKey") or "")
    BENCHMARK_NEMO_DATASET["dl_mode"] = dataset.get("dlMode") or dl_mode
    BENCHMARK_NEMO_DATASET["window_mode"] = dataset.get("windowMode") or window_mode
    if rebuilt and str(record.get("datasetKey") or ""):
        try:
            _benchmark_nemo_library_store_dataset(
                str(record.get("datasetKey")), files, dataset, operator_files
            )
        except Exception:
            pass
    if paths:
        _benchmark_nemo_save_paths(paths)
    return dataset


def _benchmark_mycom_saved_path() -> str:
    if os.path.isfile(BENCHMARK_MYCOM_CONFIG_PATH):
        try:
            with open(BENCHMARK_MYCOM_CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            return str((cfg or {}).get("path") or "").strip()
        except Exception:
            return ""
    return ""


def _benchmark_mycom_status_payload() -> dict:
    dataset = BENCHMARK_MYCOM_DATASET.get("data") or {}
    return {
        "loaded": bool(BENCHMARK_MYCOM_DATASET.get("data")),
        "path": BENCHMARK_MYCOM_DATASET.get("path") or "",
        "configuredPath": _benchmark_mycom_saved_path(),
        "cellCount": int(dataset.get("cellCount") or 0),
        "hourCount": int(dataset.get("hourCount") or 0),
        "metricCount": int(len(dataset.get("metrics") or [])),
        "dateRange": dataset.get("dateRange") or {},
    }


_MYCOM_METRIC_MAP = {
    "5G Availability": "availabilityPct",
    "5G Traffic Volume [TB]": "trafficTb",
    "5G Traffic Volume [GB]": "trafficGb",
    "5G Downlink Traffic Volume [GB]": "dlTrafficGb",
    "5G Uplink Traffic Volume [GB]": "ulTrafficGb",
    "5G Traffic Volume [MB]": "trafficMb",
    "5G Downlink Traffic Volume [MB]": "dlTrafficMb",
    "5G Uplink Traffic Volume [MB]": "ulTrafficMb",
    "5G User_DL_Throughput [Mbps]": "userDlMbps",
    "5G User_UL_Throughput [Mbps]": "userUlMbps",
    "5G Peak Users Number": "peakUsers",
    "5G Average Users Number": "avgUsers",
    "5G Cssr Tentatives": "cssrAttempts",
    "5G Cssr Success": "cssrSuccess",
    "5G Failures": "failures",
    "5G Data Cssr": "cssrPct",
    "Tx Echec Data 5G": "failurePct",
    "5G DCR Tentatives": "dcrAttempts",
    "Tx Coupure Data 5G": "dropPct",
    "5G Coupures": "drops",
    "PRB Util DL 5G": "prbUtilPct",
    "PRB Util DL 5G_BH": "prbUtilBhPct",
}


def _mycom_parse_numeric(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("%", "").replace("\u202f", "").replace(" ", "")
    text = text.replace(",", ".") if text.count(",") == 1 and text.count(".") == 0 else text
    try:
        num = float(text)
    except Exception:
        return None
    return int(num) if num.is_integer() else num


def _mycom_norm(value):
    import unicodedata

    text = str(value or "").strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


def _mycom_parse_hour(value):
    from datetime import datetime

    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%y, %H:%M", "%d/%m/%Y, %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue
    return None


def _parse_benchmark_mycom_file(path: str) -> dict:
    rows = None
    last_exc = None
    ext = os.path.splitext(str(path or ""))[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(["" if value is None else str(value) for value in row])
        except Exception as exc:
            last_exc = exc
            rows = None
    else:
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                with open(path, "r", encoding=encoding, newline="") as f:
                    rows = list(csv.reader(f))
                break
            except Exception as exc:
                last_exc = exc
                rows = None
    if rows is None:
        raise RuntimeError(f"Unable to read Mycom export: {last_exc}")
    if not rows:
        raise RuntimeError("Empty Mycom export")

    header = rows[0]
    hourly_columns = []
    for idx in range(3, len(header)):
        dt_val = _mycom_parse_hour(header[idx])
        if dt_val is not None:
            hourly_columns.append((idx, dt_val))
    if not hourly_columns:
        raise RuntimeError("No hourly columns found in Mycom export")

    cells_by_key = {}
    metrics_present = []
    current_short_name = ""
    current_gnb_key = ""
    total_values = 0
    for row in rows[1:]:
        if not row:
            continue
        short_name = str(row[0] if len(row) > 0 else "").strip()
        gnb_key = str(row[1] if len(row) > 1 else "").strip()
        metric_label = str(row[2] if len(row) > 2 else "").strip()
        if short_name:
            current_short_name = short_name
            current_gnb_key = gnb_key
        if not current_short_name or not metric_label:
            continue
        metric_key = _MYCOM_METRIC_MAP.get(metric_label)
        if not metric_key:
            continue
        if metric_label not in metrics_present:
            metrics_present.append(metric_label)
        cell_key = current_gnb_key or current_short_name
        cell_entry = cells_by_key.setdefault(cell_key, {
            "shortName": current_short_name,
            "gnbCellKey": current_gnb_key,
            "hours": {},
        })
        for idx, dt_val in hourly_columns:
            raw_value = row[idx] if idx < len(row) else ""
            num_value = _mycom_parse_numeric(raw_value)
            if num_value is None:
                continue
            ts_iso = dt_val.isoformat()
            cell_entry["hours"].setdefault(ts_iso, {})[metric_key] = num_value
            total_values += 1

    all_hours = set()
    cells_payload = []
    for cell in cells_by_key.values():
        hour_items = []
        for ts_iso, kpis in sorted(cell["hours"].items(), key=lambda item: item[0]):
            hour_items.append({"ts": ts_iso, "kpis": kpis})
            all_hours.add(ts_iso)
        cells_payload.append({
            "shortName": cell["shortName"],
            "gnbCellKey": cell["gnbCellKey"],
            "hours": hour_items,
        })
    all_hours_sorted = sorted(all_hours)
    return {
        "name": "Mycom 5G Hourly Export",
        "sourceFile": path,
        "timeGranularity": "hour",
        "dateRange": {
            "start": all_hours_sorted[0] if all_hours_sorted else "",
            "end": all_hours_sorted[-1] if all_hours_sorted else "",
        },
        "metrics": metrics_present,
        "cellCount": len(cells_payload),
        "hourCount": len(all_hours_sorted),
        "valueCount": total_values,
        "cells": cells_payload,
    }


def _mycom_site_key(short_name: str) -> str:
    text = str(short_name or "").strip()
    if not text:
        return ""
    match = re.match(r"^(.*?)(\d+)$", text)
    if match:
        prefix = re.sub(r"[_-]+$", "", match.group(1).strip())
        if prefix:
            return prefix
    return text


def _load_benchmark_mycom_file(explicit_path: str = "") -> dict:
    target_path = str(explicit_path or "").strip() or _benchmark_mycom_saved_path()
    if not target_path:
        return {"ok": False, "error": "No Mycom export path configured"}
    if not os.path.isfile(target_path):
        return {"ok": False, "error": f"File not found: {target_path}"}
    dataset = _parse_benchmark_mycom_file(target_path)
    BENCHMARK_MYCOM_DATASET["path"] = target_path
    BENCHMARK_MYCOM_DATASET["data"] = dataset
    BENCHMARK_MYCOM_DATASET["loaded_at"] = time.time()
    try:
        os.makedirs(os.path.dirname(BENCHMARK_MYCOM_CONFIG_PATH), exist_ok=True)
        with open(BENCHMARK_MYCOM_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump({"path": target_path}, f)
    except Exception:
        pass
    return {"ok": True, "path": target_path, "dataset": dataset}


# Nemo timestamps repeat heavily (dozens of rows share one value) and use non-zero-padded dates
# (e.g. "2026-5-13 14:38:04.394"), so datetime.fromisoformat always fails. Memoize the per-string
# parse and try the matching strptime format first to avoid millions of exceptions.
@lru_cache(maxsize=300000)
def _nemo_parse_time_str(text: str):
    for fmt in (
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S.%f",
        "%Y/%m/%d %H:%M:%S",
    ):
        try:
            return _dt.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return _dt.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None


def _nemo_fmt_hms_ms(dt) -> str:
    """Format a datetime as HH:MM:SS.mmm (millisecond precision), matching the Nemo
    export's timestamp style (e.g. 08:38:54.347). Returns "" for None."""
    if dt is None:
        return ""
    try:
        return dt.strftime("%H:%M:%S") + f".{dt.microsecond // 1000:03d}"
    except Exception:
        return ""


def _nemo_parse_time(value):
    if value is None or value == "":
        return None
    if isinstance(value, _dt):
        return value
    if isinstance(value, (int, float)):
        try:
            serial = float(value)
            if math.isfinite(serial) and serial > 20000:
                return _dt(1899, 12, 30) + _td(days=serial)
        except Exception:
            pass
    text = str(value).strip()
    if not text:
        return None
    return _nemo_parse_time_str(text)


def _nemo_iso(value) -> str:
    if value is None:
        return ""
    try:
        return value.isoformat()
    except Exception:
        return _benchmark_text(value)


def _nemo_guess_operator(file_name: str) -> str:
    base = os.path.basename(str(file_name or "")).lower()
    if "iam" in base:
        return "IAM"
    if "inwi" in base:
        return "INWI"
    if "orange" in base:
        return "Orange"
    return "UNKNOWN"


def _nemo_guess_delimiter(sample_text: str) -> str:
    candidates = ["\t", ",", ";"]
    best = "\t"
    best_score = -1
    head = str(sample_text or "").splitlines()[:4]
    for delimiter in candidates:
        score = sum(line.count(delimiter) for line in head)
        if score > best_score:
            best = delimiter
            best_score = score
    return best


def _nemo_percentile(values: list[float], pct: float):
    if not values:
        return None
    ordered = sorted(float(v) for v in values if v is not None and math.isfinite(float(v)))
    if not ordered:
        return None
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * pct
    lower = int(math.floor(position))
    upper = int(math.ceil(position))
    if lower == upper:
        return ordered[lower]
    weight = position - lower
    return ordered[lower] + (ordered[upper] - ordered[lower]) * weight


def _nemo_metric_stats(values: list[float]) -> dict:
    cleaned = [float(v) for v in values if v is not None and math.isfinite(float(v))]
    if not cleaned:
        return {"average": None, "median": None, "p10": None, "p90": None, "max": None, "sampleCount": 0}
    sample_count = len(cleaned)
    return {
        "average": round(sum(cleaned) / float(sample_count), 2),
        "median": round(_nemo_percentile(cleaned, 0.5), 2),
        "p10": round(_nemo_percentile(cleaned, 0.10), 2),
        "p90": round(_nemo_percentile(cleaned, 0.90), 2),
        "max": round(max(cleaned), 2),
        "sampleCount": sample_count,
    }


def _nemo_gap_pct(iam_value, best_value):
    if iam_value is None or best_value is None:
        return None
    try:
        best = float(best_value)
        current = float(iam_value)
    except Exception:
        return None
    if not math.isfinite(best) or not math.isfinite(current) or best == 0:
        return None
    return round(((current - best) / best) * 100.0, 1)


def _nemo_safe_round(value, digits=2):
    try:
        num = float(value)
    except Exception:
        return None
    return round(num, digits) if math.isfinite(num) else None


def _nemo_distribution(values: list[str]) -> list[dict]:
    counts = {}
    for value in values or []:
        text = str(value or "").strip()
        if not text:
            continue
        counts[text] = counts.get(text, 0) + 1
    total = sum(counts.values())
    ranked = sorted(counts.items(), key=lambda item: (-item[1], item[0].lower()))
    result = []
    for label, count in ranked:
        share = round((count / float(total)) * 100.0, 1) if total else 0.0
        result.append({"label": label, "count": count, "share": share})
    return result


def _nemo_forward_filled_timeshare(timeline: list[tuple]) -> dict[str, int]:
    """Forward-fill a change-event column and return per-second occupancy counts.

    The Nemo ``Serving technology`` / ``Packet technology`` columns are only logged on
    change events (≈0.3–0.9% of rows), so raw per-row counts are an unrepresentative
    sampling artifact. Given ``(event_time, value)`` pairs, carry the last non-empty
    value forward across seconds and count seconds per value, so the resulting shares are
    time-based and consistent with the time-based 5G-presence metric. Seconds before the
    first reported value are not attributed."""
    from datetime import datetime as _dt_cls

    ordered = sorted(
        (item for item in (timeline or []) if isinstance(item[0], _dt_cls)),
        key=lambda item: item[0],
    )
    per_second: dict = {}
    last = None
    for event_time, value in ordered:
        cleaned = str(value or "").strip()
        if cleaned:
            last = cleaned
        if last is None:
            continue
        per_second[event_time.replace(microsecond=0)] = last
    counts: dict = {}
    for value in per_second.values():
        counts[value] = counts.get(value, 0) + 1
    return counts


def _nemo_distribution_from_counts(counts: dict[str, int]) -> list[dict]:
    total = sum((int(count) for count in (counts or {}).values() if count), 0)
    ranked = sorted(
        ((str(label or "").strip(), int(count)) for label, count in (counts or {}).items() if str(label or "").strip() and count),
        key=lambda item: (-item[1], item[0].lower()),
    )
    result = []
    for label, count in ranked:
        share = round((count / float(total)) * 100.0, 1) if total else 0.0
        if share <= 0.0:
            # Drop negligible slivers (e.g. a 1-second forward-fill blip across a
            # multi-thousand-second window) that would render as a noisy "label (0.0%)".
            continue
        result.append({"label": label, "count": count, "share": share})
    return result


def _nemo_distribution_label(distribution: list[dict]) -> str:
    if not distribution:
        return "—"
    return ", ".join(f"{item.get('label')} ({item.get('share')}%)" for item in distribution)


def _nemo_clean_modulation(value) -> str:
    text = str(value or "").strip().upper().replace(" ", "")
    if not text:
        return ""
    if "QPSK" in text:
        return "QPSK"
    if "16QAM" in text or text == "QAM16":
        return "16QAM"
    if "64QAM" in text or text == "QAM64":
        return "64QAM"
    if "256QAM" in text or text == "QAM256":
        return "256QAM"
    return text


def _nemo_distribution_share(distribution: list[dict], label: str):
    wanted = str(label or "").strip().upper()
    for item in distribution or []:
        if str(item.get("label") or "").strip().upper() == wanted:
            return item.get("share")
    return 0.0 if distribution else None


def _nemo_pdsch_active_row(row: dict) -> bool:
    if not isinstance(row, dict):
        return False
    for key in (
        "pdschPrbs",
        "pdschSched5gMbps",
        "pdschDl5gMbps",
        "pdschBitsPerHz",
        "pdschMaxBitsPerHz",
        "pdschMcsCw0",
        "pdschMcsCw1",
        "pdschTbsCw0",
        "pdschTbsCw1",
        "scheduledRank",
        "pdschSlotPct",
    ):
        value = row.get(key)
        if value is None:
            continue
        try:
            if float(value) > 0:
                return True
        except Exception:
            pass
    return bool(_nemo_clean_modulation(row.get("pdschModulationCw0")) or _nemo_clean_modulation(row.get("pdschModulationCw1")))


def _nemo_spectral_efficiency_insight(iam_kpis: dict, cmp_kpis: dict) -> dict:
    iam_mod = iam_kpis.get("pdschModulation") or {}
    cmp_mod = cmp_kpis.get("pdschModulation") or {}
    iam_mcs = iam_kpis.get("pdschMcs") or {}
    cmp_mcs = cmp_kpis.get("pdschMcs") or {}
    iam_bits = iam_kpis.get("pdschBitPerHz") or {}
    cmp_bits = cmp_kpis.get("pdschBitPerHz") or {}
    iam_rank = iam_kpis.get("scheduledRank") or {}
    cmp_rank = cmp_kpis.get("scheduledRank") or {}

    iam_qpsk = _nemo_distribution_share(iam_mod.get("distribution") or [], "QPSK")
    iam_16 = _nemo_distribution_share(iam_mod.get("distribution") or [], "16QAM")
    iam_64 = _nemo_distribution_share(iam_mod.get("distribution") or [], "64QAM")
    iam_256 = _nemo_distribution_share(iam_mod.get("distribution") or [], "256QAM")
    cmp_qpsk = _nemo_distribution_share(cmp_mod.get("distribution") or [], "QPSK")
    cmp_16 = _nemo_distribution_share(cmp_mod.get("distribution") or [], "16QAM")
    cmp_64 = _nemo_distribution_share(cmp_mod.get("distribution") or [], "64QAM")
    cmp_256 = _nemo_distribution_share(cmp_mod.get("distribution") or [], "256QAM")

    iam_high_share = None if iam_64 is None and iam_256 is None else round(float(iam_64 or 0) + float(iam_256 or 0), 1)
    cmp_high_share = None if cmp_64 is None and cmp_256 is None else round(float(cmp_64 or 0) + float(cmp_256 or 0), 1)
    iam_low_share = None if iam_qpsk is None and iam_16 is None else round(float(iam_qpsk or 0) + float(iam_16 or 0), 1)
    cmp_low_share = None if cmp_qpsk is None and cmp_16 is None else round(float(cmp_qpsk or 0) + float(cmp_16 or 0), 1)

    iam_mcs_med = iam_mcs.get("median")
    cmp_mcs_med = cmp_mcs.get("median")
    iam_bits_med = iam_bits.get("median")
    cmp_bits_med = cmp_bits.get("median")
    iam_rank_med = iam_rank.get("median")
    cmp_rank_med = cmp_rank.get("median")

    iam_samples = max(
        int(iam_mod.get("sampleCount") or 0),
        int(iam_mcs.get("sampleCount") or 0),
        int(iam_bits.get("sampleCount") or 0),
    )
    cmp_samples = max(
        int(cmp_mod.get("sampleCount") or 0),
        int(cmp_mcs.get("sampleCount") or 0),
        int(cmp_bits.get("sampleCount") or 0),
    )
    sufficient = iam_samples >= 5 and cmp_samples >= 5

    lower_modulation = (
        iam_high_share is not None and cmp_high_share is not None and iam_high_share <= (cmp_high_share - 15)
    ) or (
        iam_low_share is not None and cmp_low_share is not None and iam_low_share >= (cmp_low_share + 15)
    )
    lower_mcs = iam_mcs_med is not None and cmp_mcs_med is not None and iam_mcs_med <= (cmp_mcs_med - 4)
    lower_bits = iam_bits_med is not None and cmp_bits_med not in (None, 0) and iam_bits_med <= (float(cmp_bits_med) * 0.75)
    lower_rank = iam_rank_med is not None and cmp_rank_med is not None and iam_rank_med <= (cmp_rank_med - 1)

    comparable_modulation = (
        iam_high_share is not None and cmp_high_share is not None and abs(iam_high_share - cmp_high_share) <= 10
    ) and (
        iam_low_share is not None and cmp_low_share is not None and abs(iam_low_share - cmp_low_share) <= 10
    )
    comparable_mcs = iam_mcs_med is not None and cmp_mcs_med is not None and abs(iam_mcs_med - cmp_mcs_med) <= 2
    comparable_bits = (
        iam_bits_med is not None and cmp_bits_med not in (None, 0)
        and abs(float(iam_bits_med) - float(cmp_bits_med)) <= (float(cmp_bits_med) * 0.2)
    )

    confirmed = bool(sufficient and (lower_modulation or lower_mcs or lower_bits))
    comparable = bool(sufficient and not confirmed and ((comparable_modulation and comparable_mcs) or (comparable_bits and comparable_mcs)))

    # IAM is strictly better than comparator on spectral efficiency
    higher_modulation = (
        iam_high_share is not None and cmp_high_share is not None and iam_high_share >= (cmp_high_share + 15)
    ) or (
        iam_qpsk is not None and cmp_qpsk is not None and float(iam_qpsk or 0) <= float(cmp_qpsk or 0) - 15
    )
    higher_mcs = iam_mcs_med is not None and cmp_mcs_med is not None and float(iam_mcs_med) >= float(cmp_mcs_med) + 2
    higher_bits = (
        iam_bits_med is not None and cmp_bits_med not in (None, 0)
        and float(iam_bits_med) >= float(cmp_bits_med) * 1.25
    )
    spectral_better_count = int(bool(higher_modulation)) + int(bool(higher_mcs)) + int(bool(higher_bits))
    iam_better = bool(sufficient and spectral_better_count >= 2 and not lower_modulation and not lower_mcs and not lower_bits)

    note_parts = []
    if not sufficient:
        note_parts.append(
            f"Modulation/MCS evidence is limited ({iam_samples} IAM samples vs {cmp_samples} comparator samples)."
        )
    elif iam_better:
        iam_dom = iam_mod.get("dominant") or "—"
        cmp_dom = cmp_mod.get("dominant") or "—"
        note_parts.append(
            f"IAM modulation, MCS and spectral efficiency are better than the comparator during the available 5G samples. "
            f"IAM dominant modulation is {iam_dom} with {iam_high_share or 0}% 64/256QAM, while the comparator is {cmp_dom}-dominant. "
            f"Modulation quality is not the main throughput limitation versus the comparator."
        )
    else:
        if lower_modulation:
            note_parts.append("IAM uses lower-order modulation more often than the comparator.")
        if lower_mcs:
            note_parts.append("IAM median PDSCH MCS is lower than the comparator.")
        if lower_bits:
            note_parts.append("IAM median PDSCH bit/s/Hz is lower than the comparator.")
        if lower_rank:
            note_parts.append("IAM scheduled rank is lower, which also reduces effective spectral efficiency.")
        if comparable and not note_parts:
            note_parts.append("Modulation/MCS/bit-per-Hz are broadly comparable; the gap is not explained by spectral efficiency.")

    return {
        "sufficient": sufficient,
        "confirmed": confirmed,
        "comparable": comparable,
        "iamBetter": iam_better,
        "iamSamples": iam_samples,
        "comparatorSamples": cmp_samples,
        "iamHighShare": iam_high_share,
        "comparatorHighShare": cmp_high_share,
        "iamLowShare": iam_low_share,
        "comparatorLowShare": cmp_low_share,
        "iamMedianMcs": iam_mcs_med,
        "comparatorMedianMcs": cmp_mcs_med,
        "iamMedianBitsPerHz": iam_bits_med,
        "comparatorMedianBitsPerHz": cmp_bits_med,
        "lowerModulation": lower_modulation,
        "lowerMcs": lower_mcs,
        "lowerBitsPerHz": lower_bits,
        "lowerRank": lower_rank,
        "higherModulation": higher_modulation,
        "higherMcs": higher_mcs,
        "higherBitsPerHz": higher_bits,
        "note": " ".join(note_parts).strip(),
    }


def _nemo_ordinal(value) -> str:
    try:
        num = int(value)
    except Exception:
        return "N/A"
    if 10 <= (num % 100) <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(num % 10, "th")
    return f"{num}{suffix}"


def _nemo_ordinal_fr(value) -> str:
    try:
        num = int(value)
    except Exception:
        return "N/A"
    return {1: "premier", 2: "deuxième", 3: "troisième", 4: "quatrième", 5: "cinquième"}.get(num, f"{num}e")


def _nemo_haversine_m(lat1, lon1, lat2, lon2) -> float | None:
    try:
        la1 = math.radians(float(lat1))
        lo1 = math.radians(float(lon1))
        la2 = math.radians(float(lat2))
        lo2 = math.radians(float(lon2))
    except Exception:
        return None
    dlat = la2 - la1
    dlon = lo2 - lo1
    a = math.sin(dlat / 2.0) ** 2 + math.cos(la1) * math.cos(la2) * math.sin(dlon / 2.0) ** 2
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
    return 6371000.0 * c


_NEMO_INVALID_BAND_TOKENS = ("undefined", "none", "null", "n/a", "-", "")


def _nemo_is_valid_band(value) -> bool:
    """A band token is meaningful only if it isn't blank or a placeholder like
    literal "undefined" (which Nemo emits when the band can't be resolved)."""
    return str(value or "").strip().lower() not in _NEMO_INVALID_BAND_TOKENS


def _nemo_share_over_threshold(values: list[float], threshold: float) -> float | None:
    if not values:
        return None
    return round((sum(1 for value in values if float(value) > threshold) / float(len(values))) * 100.0, 1)


def _nemo_active_status_share(values: list[str], active_tokens: tuple[str, ...] = ("ca", "active", "on", "enabled", "2 cc", "3 cc")) -> float | None:
    normalized = [str(value or "").strip().lower() for value in values if str(value or "").strip()]
    if not normalized:
        return None
    active_count = 0
    for value in normalized:
        if any(token in value for token in active_tokens) and all(token not in value for token in ("no ca", "inactive", "off", "disabled")):
            active_count += 1
    return round((active_count / float(len(normalized))) * 100.0, 1)


def _nemo_band_row_filter(row: dict) -> bool:
    band = str(row.get("band") or "").strip().lower()
    if not band:
        return False
    return (
        band.startswith("n")
        or row.get("nrChannelNumber") is not None
        or "en-dc" in str(row.get("servingTechnology") or "").lower()
        or "en-dc" in str(row.get("packetTechnology") or "").lower()
        or any(str(cell or "").strip().upper() == "SCG PSCELL" for cell in (row.get("cellTypes") or []))
    )


def _nemo_compute_technology_status(rows: list[dict], operator: str) -> dict:
    serving_tech_timeline = []
    packet_tech_timeline = []
    scg_pscell_samples = 0
    nr_channel_non_null_count = 0
    mac_dl_5g_positive_count = 0
    pdsch_dl_5g_positive_count = 0
    nr_bands = set()
    nr_presence_by_second = {}
    has_5g = False
    _NR_CT_UPPER = {"NR SERVING", "NR SCG PSCELL", "SCG PSCELL", "5G SERVING"}

    from datetime import datetime as _dt_class

    for payload in rows or []:
        band_text = payload.get("band")
        nr_channel_number = payload.get("nrChannelNumber")
        serving_technology = str(payload.get("servingTechnology") or "")
        packet_technology = str(payload.get("packetTechnology") or "")
        cell_types_upper = {
            str(cell or "").strip().upper()
            for cell in (payload.get("cellTypes") or [])
        }
        serving_technology_upper = serving_technology.upper()
        packet_technology_upper = packet_technology.upper()

        if payload.get("pdschDl5gMbps") is not None and float(payload.get("pdschDl5gMbps") or 0) > 0:
            pdsch_dl_5g_positive_count += 1
        if nr_channel_number is not None:
            nr_channel_non_null_count += 1
        if payload.get("macDl5gMbps") is not None and float(payload.get("macDl5gMbps") or 0) > 0:
            mac_dl_5g_positive_count += 1
        if "SCG PSCELL" in cell_types_upper:
            scg_pscell_samples += 1

        row_has_nr = (
            nr_channel_number is not None
            or "EN-DC" in serving_technology_upper
            or "5G" in serving_technology_upper
            or "EN-DC" in packet_technology_upper
            or float(payload.get("macDl5gMbps") or 0) > 0
            or float(payload.get("pdschDl5gMbps") or 0) > 0
            or bool(cell_types_upper & _NR_CT_UPPER)
        )
        if row_has_nr:
            has_5g = True
        if band_text and row_has_nr and _nemo_is_valid_band(band_text):
            nr_bands.add(str(band_text).strip())
        event_time = payload.get("_dt")
        if isinstance(event_time, _dt_class):
            second_bucket = event_time.replace(microsecond=0)
            nr_presence_by_second[second_bucket] = nr_presence_by_second.get(second_bucket, False) or row_has_nr
            serving_tech_timeline.append((event_time, serving_technology))
            packet_tech_timeline.append((event_time, packet_technology))

    nr_presence_seconds = sum(1 for present in nr_presence_by_second.values() if present)
    lte_only_seconds = sum(1 for present in nr_presence_by_second.values() if not present)
    total_presence_seconds = len(nr_presence_by_second)
    nr_presence_pct = round(nr_presence_seconds / float(total_presence_seconds) * 100.0, 1) if total_presence_seconds else None
    lte_only_presence_pct = round(lte_only_seconds / float(total_presence_seconds) * 100.0, 1) if total_presence_seconds else None
    return {
        "operator": operator,
        "has5g": has_5g,
        "fiveGStatus": "5G/EN-DC detected" if has_5g else "No 5G detected in export",
        "scgPscellSamples": scg_pscell_samples,
        "nrChannelNonNullCount": nr_channel_non_null_count,
        "macDl5gPositiveSamples": mac_dl_5g_positive_count,
        "pdschDl5gPositiveSamples": pdsch_dl_5g_positive_count,
        "fiveGThroughputSamples": mac_dl_5g_positive_count + pdsch_dl_5g_positive_count,
        "nrPresencePct": nr_presence_pct,
        "lteOnlyPresencePct": lte_only_presence_pct,
        "nrPresenceSeconds": nr_presence_seconds,
        "lteOnlySeconds": lte_only_seconds,
        "totalPresenceSeconds": total_presence_seconds,
        "servingTechnologyDistribution": _nemo_distribution_from_counts(_nemo_forward_filled_timeshare(serving_tech_timeline)),
        "packetTechnologyDistribution": _nemo_distribution_from_counts(_nemo_forward_filled_timeshare(packet_tech_timeline)),
        "nrBands": sorted(nr_bands),
        "comment": (
            "5G/EN-DC detected. Operator is eligible for 5G diagnosis."
            if has_5g
            else "No 5G NR/EN-DC detected in export. Operator remains in DL ranking but 5G-specific KPIs are N/A."
        ),
    }


def _nemo_radio_presence_breakdown_from_cells(cells_payload: list[dict]) -> dict:
    totals = {"4G": 0.0, "5G": 0.0}
    for item in cells_payload or []:
        dwell = item.get("dwellSec")
        if dwell is None:
            continue
        tech = str(item.get("tech") or "").upper()
        if tech.startswith(("5G", "NR")):
            totals["5G"] += float(dwell)
        elif tech.startswith("4G"):
            totals["4G"] += float(dwell)
    total = totals["4G"] + totals["5G"]
    if total <= 0:
        return {}
    breakdown = {}
    if totals["5G"] > 0:
        breakdown["5G"] = round(totals["5G"] / total * 100.0, 1)
    if totals["4G"] > 0:
        breakdown["4G"] = round(totals["4G"] / total * 100.0, 1)
    return breakdown


def _nemo_resolve_row_tech(row: dict) -> str | None:
    pci = row.get("pci")
    if pci is None:
        return None
    cell_types_upper = {
        str(cell or "").strip().upper()
        for cell in (row.get("cellTypes") or [])
    }
    serving_technology_upper = str(row.get("servingTechnology") or "").upper()
    packet_technology_upper = str(row.get("packetTechnology") or "").upper()
    if (
        row.get("nrChannelNumber") is not None
        or "EN-DC" in serving_technology_upper
        or "5G" in serving_technology_upper
        or "EN-DC" in packet_technology_upper
        or bool(cell_types_upper & {"NR SERVING", "NR SCG PSCELL", "SCG PSCELL", "5G SERVING"})
    ):
        return "5G"
    if row.get("lteChannelNumber") is not None or "LTE" in serving_technology_upper or "LTE" in packet_technology_upper or "LTE SERVING" in cell_types_upper:
        return "4G"
    return None


def _nemo_resolve_dominant_arfcn_by_tech_pci(rows: list[dict]) -> dict:
    from collections import Counter

    counts: dict = {}
    for row in rows or []:
        tech = _nemo_resolve_row_tech(row)
        pci = row.get("pci")
        if tech is None or pci is None:
            continue
        try:
            int_pci = int(float(pci))
        except Exception:
            continue
        arfcn = row.get("nrChannelNumber") if tech == "5G" else row.get("lteChannelNumber")
        if arfcn is None:
            continue
        try:
            int_arfcn = int(float(arfcn))
        except Exception:
            continue
        key = (tech, int_pci)
        bucket = counts.get(key)
        if bucket is None:
            bucket = Counter()
            counts[key] = bucket
        bucket[int_arfcn] += 1

    resolved = {}
    for key, bucket in counts.items():
        resolved[key] = bucket.most_common(1)[0][0]
    return resolved


def _nemo_lte_band_label(raw_band: str | None) -> str | None:
    band_text = str(raw_band or "").strip().upper()
    if not band_text:
        return None
    if band_text.startswith("L"):
        return band_text
    band_map = {
        "B1": "L2100",
        "B2": "L1900",
        "B3": "L1800",
        "B4": "L-AWS",
        "B5": "L850",
        "B7": "L2600",
        "B8": "L900",
        "B20": "L800",
        "B28": "L700",
        "B32": "L1500",
        "B38": "L2600-TDD",
        "B40": "L2300-TDD",
        "B66": "L-AWS-3",
    }
    return band_map.get(band_text, band_text)


def _nemo_presence_seconds(ts_list: list, max_gap_seconds: float = 1.5) -> float | None:
    from datetime import datetime as _dt_class

    second_buckets = sorted({
        ts.replace(microsecond=0)
        for ts in (ts_list or [])
        if isinstance(ts, _dt_class)
    })
    if not second_buckets:
        return None

    total_seconds = 1.0
    prev = second_buckets[0]
    for current in second_buckets[1:]:
        gap = (current - prev).total_seconds()
        if gap <= max_gap_seconds:
            total_seconds += max(gap, 0.0)
        else:
            total_seconds += 1.0
        prev = current
    return round(total_seconds, 0)


def _nemo_sum_interval_seconds_by_key(intervals: list[dict], fallback_end_dt=None) -> dict:
    totals: dict = {}
    for interval in intervals or []:
        key = interval.get("key")
        start = interval.get("start")
        end = interval.get("end") or fallback_end_dt
        if key is None or start is None or end is None:
            continue
        try:
            seconds = max((end - start).total_seconds(), 0.0)
        except Exception:
            continue
        totals[key] = round(totals.get(key, 0.0) + seconds, 0)
    return totals


def _nemo_build_episode_ranges(timed_records: list[dict], global_end_dt=None) -> list[dict]:
    episodes: list[dict] = []
    if not timed_records:
        return episodes

    ep_key = timed_records[0]["key"]
    ep_recs = [timed_records[0]]
    for rec in timed_records[1:]:
        if rec["key"] == ep_key:
            ep_recs.append(rec)
            continue
        episodes.append({
            "key": ep_key,
            "records": ep_recs,
            "start": ep_recs[0].get("dt"),
            "end": rec.get("dt"),
            "dwellSec": round(max((rec.get("dt") - ep_recs[0].get("dt")).total_seconds(), 0.0), 0)
            if rec.get("dt") and ep_recs[0].get("dt") else None,
        })
        ep_key = rec["key"]
        ep_recs = [rec]

    last_start = ep_recs[0].get("dt")
    last_end = global_end_dt
    dwell = None
    if last_start and last_end:
        try:
            dwell = round(max((last_end - last_start).total_seconds(), 0.0), 0)
        except Exception:
            dwell = None
    episodes.append({
        "key": ep_key,
        "records": ep_recs,
        "start": last_start,
        "end": last_end,
        "dwellSec": dwell,
    })
    return episodes


def _nemo_normalize_rrc_state_label(raw) -> str | None:
    if raw is None:
        return None
    txt = str(raw).strip()
    if not txt:
        return None
    low = txt.lower()
    compact = re.sub(r"[^a-z0-9]", "", low)
    if re.fullmatch(r"-?\d+", txt):
        num = int(txt)
        if num == 0:
            return "Idle"
        if num == 1:
            return "Connected"
        if num == 2:
            return "Inactive"
    if re.search(r"(rrc_?connected|\bconnected\b|cell_dch|cell_fach|cell_pch|ura_pch)", low):
        return "Connected"
    if re.search(r"(rrc_?idle|\bidle\b|released?|release complete|pcch.*paging|\bpaging\b)", low):
        return "Idle"
    if re.search(r"(rrc_?inactive|\binactive\b)", low):
        return "Inactive"
    if compact in {"rrcconnected", "connected"}:
        return "Connected"
    if compact in {"rrcidle", "idle"}:
        return "Idle"
    if compact in {"rrcinactive", "inactive"}:
        return "Inactive"
    return txt


def _nemo_episode_connection_state(records: list[dict]) -> str | None:
    counts: dict = {}
    for rec in records or []:
        state = _nemo_normalize_rrc_state_label(rec.get("rrcState"))
        if not state:
            continue
        counts[state] = counts.get(state, 0) + 1
    if counts:
        return max(counts, key=counts.get)
    for rec in records or []:
        if rec.get("servingMode") or rec.get("dl") or rec.get("appTs"):
            return "Connected (inferred)"
    return None


def _nemo_episode_dl_window(records: list[dict]) -> tuple[str | None, str | None]:
    from datetime import datetime as _datetime

    app_ts = sorted(
        dt
        for rec in (records or [])
        for dt in (rec.get("appTs") or [])
        if isinstance(dt, _datetime)
    )
    if not app_ts:
        return None, None
    return _nemo_fmt_hms_ms(app_ts[0]), _nemo_fmt_hms_ms(app_ts[-1])


def _nemo_episode_ul_window(records: list[dict]) -> tuple[str | None, str | None]:
    from datetime import datetime as _datetime

    app_ts = sorted(
        dt
        for rec in (records or [])
        for dt in (rec.get("appTsUl") or [])
        if isinstance(dt, _datetime)
    )
    if not app_ts:
        return None, None
    return _nemo_fmt_hms_ms(app_ts[0]), _nemo_fmt_hms_ms(app_ts[-1])


def _nemo_numeric_avg(values: list[float]) -> float | None:
    nums = [float(v) for v in (values or []) if v is not None]
    return round(sum(nums) / len(nums), 1) if nums else None


def _nemo_numeric_median(values: list[float]) -> float | None:
    nums = sorted(float(v) for v in (values or []) if v is not None)
    if not nums:
        return None
    n = len(nums)
    return round(nums[n // 2] if n % 2 else (nums[n // 2 - 1] + nums[n // 2]) / 2, 1)


def _nemo_episode_display_payload(
    episode: dict,
    idx: int | None = None,
    color: str | None = None,
    start_dt=None,
    end_dt=None,
    records: list[dict] | None = None,
) -> dict:
    from datetime import datetime as _datetime

    key = episode.get("key")
    base_display = dict(episode.get("display") or {})
    recs = list(records if records is not None else (episode.get("records") or []))
    if key is None:
        return base_display

    cn, sn, tech, band = key
    start_dt = start_dt if start_dt is not None else episode.get("start")
    end_dt = end_dt if end_dt is not None else episode.get("end")
    dwell_sec = episode.get("dwellSec")
    if start_dt is not None and end_dt is not None:
        try:
            dwell_sec = max((end_dt - start_dt).total_seconds(), 0.0)
        except Exception:
            pass

    dl_vals = []
    for rec in (episode.get("records") or []):
        dl_series = list(rec.get("dl") or [])
        ts_series = list(rec.get("appTs") or [])
        if dl_series and ts_series and len(dl_series) == len(ts_series) and start_dt is not None and end_dt is not None:
            for dl_val, ts in zip(dl_series, ts_series):
                if not isinstance(ts, _datetime):
                    continue
                if start_dt <= ts <= end_dt and dl_val is not None:
                    dl_vals.append(float(dl_val))
            continue
        if rec in recs:
            dl_vals.extend(float(v) for v in dl_series if v is not None)
    rsrp_vals = [v for rec in recs for v in (rec.get("rsrp") or []) if v is not None]
    sinr_vals = [v for rec in recs for v in (rec.get("sinr") or []) if v is not None]
    app_ts = []
    for rec in (episode.get("records") or []):
        for ts in (rec.get("appTs") or []):
            if not isinstance(ts, _datetime):
                continue
            if start_dt is not None and end_dt is not None and not (start_dt <= ts <= end_dt):
                continue
            app_ts.append(ts)
    if app_ts:
        app_ts.sort()
        dl_start_iso = _nemo_fmt_hms_ms(app_ts[0])
        dl_end_iso = _nemo_fmt_hms_ms(app_ts[-1])
    else:
        dl_start_iso, dl_end_iso = _nemo_episode_dl_window(recs)
    app_ts_ul = []
    for rec in (episode.get("records") or []):
        for ts in (rec.get("appTsUl") or []):
            if not isinstance(ts, _datetime):
                continue
            if start_dt is not None and end_dt is not None and not (start_dt <= ts <= end_dt):
                continue
            app_ts_ul.append(ts)
    if app_ts_ul:
        app_ts_ul.sort()
        ul_start_iso = _nemo_fmt_hms_ms(app_ts_ul[0])
        ul_end_iso = _nemo_fmt_hms_ms(app_ts_ul[-1])
    else:
        ul_start_iso, ul_end_iso = _nemo_episode_ul_window(recs)
    connection_state = _nemo_episode_connection_state(recs) or base_display.get("connectionState")

    mode_counts: dict = {}
    anchor_counts: dict = {}
    for rec in recs:
        m = rec.get("servingMode")
        if m:
            mode_counts[m] = mode_counts.get(m, 0) + 1
        a = rec.get("lteAnchor")
        if a:
            anchor_counts[a] = anchor_counts.get(a, 0) + 1
    ep_mode = max(mode_counts, key=mode_counts.get) if mode_counts else base_display.get("servingMode")
    ep_lte_anchor = max(anchor_counts, key=anchor_counts.get) if anchor_counts else base_display.get("lteAnchor")

    return {
        "idx": idx if idx is not None else base_display.get("idx"),
        "cellName": cn,
        "siteName": sn,
        "tech": tech,
        "band": band,
        "samples": len(recs),
        "dwellSec": dwell_sec,
        "startTime": _nemo_fmt_hms_ms(start_dt) if start_dt else "",
        "endTime": _nemo_fmt_hms_ms(end_dt) if end_dt else "",
        "dlStartTime": dl_start_iso,
        "dlEndTime": dl_end_iso,
        "ulStartTime": ul_start_iso,
        "ulEndTime": ul_end_iso,
        "avgDlMbps": _nemo_numeric_avg(dl_vals),
        "medianRsrp": _nemo_numeric_median(rsrp_vals),
        "medianSinr": _nemo_numeric_median(sinr_vals),
        "connectionState": connection_state,
        "servingMode": ep_mode,
        "lteAnchor": ep_lte_anchor,
        "color": color or base_display.get("color"),
    }


def _nemo_downlink_transfer_intervals(sessions: list[dict]) -> list[dict]:
    from datetime import datetime as _datetime

    intervals = []
    for session in sessions or []:
        direction = str(session.get("direction") or "").strip().lower()
        if not direction.startswith("down"):
            continue
        start_raw = session.get("startTime")
        end_raw = session.get("endTime")
        try:
            start_dt = _datetime.fromisoformat(str(start_raw))
            end_dt = _datetime.fromisoformat(str(end_raw))
        except Exception:
            continue
        if end_dt <= start_dt:
            continue
        intervals.append({"start": start_dt, "end": end_dt})
    intervals.sort(key=lambda item: item["start"])
    merged: list[dict] = []
    for interval in intervals:
        if not merged or interval["start"] > merged[-1]["end"]:
            merged.append(dict(interval))
            continue
        if interval["end"] > merged[-1]["end"]:
            merged[-1]["end"] = interval["end"]
    return merged


def _nemo_active_download_intervals(rows: list[dict], gap_tolerance_sec: float = 2.0) -> list[dict]:
    """Intervals during which the application-layer DL throughput was actually flowing
    (appDlMbps > 0), merged across gaps up to ``gap_tolerance_sec``.

    The transfer-*session* window (marker-to-marker) often extends past the moment the file
    finished downloading — into connection teardown / the next session boundary — so scoping
    the "during download" serving-cell sequence to it pulls in cells that only served after
    the download completed. Bounding by the active-DL seconds fixes that.
    """
    active = sorted(
        row["_dt"]
        for row in (rows or [])
        if row.get("_dt") is not None
        and row.get("appDlMbps") is not None
        and float(row.get("appDlMbps") or 0) > 0
    )
    if not active:
        return []
    intervals: list[dict] = []
    start = prev = active[0]
    for ts in active[1:]:
        if (ts - prev).total_seconds() > gap_tolerance_sec:
            intervals.append({"start": start, "end": prev})
            start = ts
        prev = ts
    intervals.append({"start": start, "end": prev})
    return intervals


def _nemo_extract_dl_events(rows: list[dict]) -> dict:
    """Extract download session events from Event ID column (DAA/DAC/DREQ/DCOMP/DAD).

    Nemo logs each event on multiple cell rows at the same millisecond (one row per
    active cell), so we deduplicate by (eventId, datetime) before grouping.

    Returns a dict with:
      sessions  — list of per-session dicts with timestamps and derived KPIs
      markers   — flat list of {type, label, ts, tSec, sessionIdx} for the timeline
      downloadIntervals — [{start, end}] keyed by DREQ→DCOMP (for scoping KPIs)
      sessionIntervals  — [{start, end}] keyed by DAA→DAD (wider session window)
      kpis      — aggregate {sessionCount, timeToConnectAvgMs, downloadDurationAvgS, …}
    """
    _EVENT_IDS = {"DAA", "DAC", "DREQ", "DCOMP", "DAD"}
    _EVENT_LABELS = {
        "DAA":   "Session start (DAA)",
        "DAC":   "Server connected (DAC)",
        "DREQ":  "DL start (DREQ)",
        "DCOMP": "DL end (DCOMP)",
        "DAD":   "Session end (DAD)",
    }

    # Collect unique events (deduplicate same event at same ms across cell rows)
    seen: set = set()
    events: list[dict] = []
    for row in (rows or []):
        eid = str(row.get("eventId") or "").strip().upper()
        if eid not in _EVENT_IDS:
            continue
        dt = row.get("_dt")
        if dt is None:
            continue
        key = (eid, dt)
        if key in seen:
            continue
        seen.add(key)
        events.append({
            "type": eid,
            "dt": dt,
            "eventText": str(row.get("eventText") or "").strip(),
        })

    events.sort(key=lambda e: e["dt"])

    # Group into sessions: DAA ... DAD boundaries
    sessions_raw: list[dict] = []
    remaining = list(events)
    while remaining:
        daa_pos = next((i for i, e in enumerate(remaining) if e["type"] == "DAA"), None)
        if daa_pos is None:
            break
        # Find next DAD after this DAA
        dad_pos = next(
            (i for i, e in enumerate(remaining) if i > daa_pos and e["type"] == "DAD"),
            None,
        )
        if dad_pos is not None:
            # DCOMP and DAD can fire at the same millisecond (transfer completes and
            # session closes simultaneously). Extend the window past DAD to capture
            # any events at the same timestamp (e.g. DCOMP logged after DAD in the file).
            dad_dt = remaining[dad_pos]["dt"]
            end_pos = dad_pos
            while end_pos + 1 < len(remaining) and remaining[end_pos + 1]["dt"] == dad_dt:
                end_pos += 1
            session_evts = remaining[daa_pos: end_pos + 1]
            remaining = remaining[end_pos + 1:]
        else:
            session_evts = remaining[daa_pos:]
            remaining = []

        def _first_dt(etype):
            return next((e["dt"] for e in session_evts if e["type"] == etype), None)

        daa_dt   = _first_dt("DAA")
        dac_dt   = _first_dt("DAC")
        dreq_dt  = _first_dt("DREQ")
        dcomp_dt = _first_dt("DCOMP")
        dad_dt   = _first_dt("DAD")

        time_to_connect_ms = (
            round((dac_dt - daa_dt).total_seconds() * 1000, 1)
            if dac_dt and daa_dt else None
        )
        # Delay to start download = session start (DAA) → download request (DREQ).
        # This is the full pre-download latency (connect + request setup), used as the
        # offset-mode lead-in so the per-operator "delay to start download" is visible.
        start_delay_s = (
            round((dreq_dt - daa_dt).total_seconds(), 3)
            if dreq_dt and daa_dt else None
        )
        download_duration_s = (
            round((dcomp_dt - dreq_dt).total_seconds(), 3)
            if dreq_dt and dcomp_dt else None
        )
        session_duration_s = (
            round((dad_dt - daa_dt).total_seconds(), 3)
            if dad_dt and daa_dt else None
        )

        sessions_raw.append({
            "_daa_dt": daa_dt, "_dac_dt": dac_dt,
            "_dreq_dt": dreq_dt, "_dcomp_dt": dcomp_dt, "_dad_dt": dad_dt,
            "daa":   _nemo_fmt_hms_ms(daa_dt)   if daa_dt   else None,
            "dac":   _nemo_fmt_hms_ms(dac_dt)   if dac_dt   else None,
            "dreq":  _nemo_fmt_hms_ms(dreq_dt)  if dreq_dt  else None,
            "dcomp": _nemo_fmt_hms_ms(dcomp_dt) if dcomp_dt else None,
            "dad":   _nemo_fmt_hms_ms(dad_dt)   if dad_dt   else None,
            "timeToConnectMs": time_to_connect_ms,
            "startDelayS": start_delay_s,
            "downloadDurationS": download_duration_s,
            "sessionDurationS": session_duration_s,
        })

    # ── Enrich each session from the time-series rows in its [DAA, DAD] window ──
    # The per-second export logs Bytes DL / Bytes UL / Download time / direction / protocol
    # / status on the rows inside each session window. Reading them here lets us classify
    # ping / upload / download and recover Nemo's own per-operation KPIs WITHOUT the
    # separate "Data transfer session statistics" file — so the same DT export that drives
    # the timeline also yields the authoritative download timing & throughput.
    def _avg(vals):
        return round(sum(vals) / len(vals), 2) if vals else None

    for sess in sessions_raw:
        win_start = sess.get("_daa_dt") or sess.get("_dreq_dt")
        win_end = sess.get("_dad_dt") or sess.get("_dcomp_dt")
        # RF/PHY is averaged over the actual transfer window (DREQ→DCOMP), which excludes the
        # brief connect (DAA→DREQ) and teardown (DCOMP→DAD) phases where SINR can be transitional.
        rf_lo = sess.get("_dreq_dt") or win_start
        rf_hi = sess.get("_dcomp_dt") or win_end
        direction = protocol = status = ""
        bytes_dl = bytes_ul = dl_time_kpi = file_size = None
        rf_rsrp = []; rf_sinr = []; rf_prb = []; nr_pdsch = []; lte_pdsch = []
        mac_total = []
        app_samples = []
        bw_samples = []
        scell_samples = []
        active_coords = []
        # Band / modulation / rank / NR-presence live on sparse change-event rows (not the
        # throughput rows), so they are accumulated over the whole transfer window below —
        # the same fix applied to RF. Throughput-row-only counting produced n78%=—,
        # mod256%=0, rank=None and an impossible 5G dwell of 125%.
        win_banded_rows = 0
        win_nr_rows = 0
        win_band_counts: dict[str, int] = {}
        win_mod256 = 0
        win_mod_total = 0
        win_rank = []
        measurement_title = None

        def _append_sample(bucket, when, value, positive=False, weight=None):
            if value is None:
                return
            try:
                num = float(value)
            except (TypeError, ValueError):
                return
            if positive and num <= 0:
                return
            w = None
            if weight is not None:
                try:
                    w = float(weight)
                except (TypeError, ValueError):
                    w = None
            bucket.append((when, num, w))

        def _avg_series(bucket, start_dt=None, min_points=2):
            if not bucket:
                return None
            chosen = bucket
            if start_dt is not None:
                scoped = [
                    (dt, val, weight)
                    for dt, val, weight in bucket
                    if dt >= start_dt
                ]
                if len(scoped) >= min_points:
                    chosen = scoped
            vals = [val for _, val, _ in chosen]
            return _avg(vals)

        def _weighted_avg_series(bucket, start_dt=None, min_points=2):
            if not bucket:
                return None
            chosen = bucket
            if start_dt is not None:
                scoped = [
                    (dt, val, weight)
                    for dt, val, weight in bucket
                    if dt >= start_dt
                ]
                if len(scoped) >= min_points:
                    chosen = scoped
            weighted = [
                (val, weight)
                for _, val, weight in chosen
                if weight is not None and weight > 0
            ]
            if weighted:
                total_weight = sum(weight for _, weight in weighted)
                if total_weight > 0:
                    return round(
                        sum(val * weight for val, weight in weighted) / total_weight,
                        2,
                    )
            vals = [val for _, val, _ in chosen]
            return _avg(vals)

        def _median_series(bucket, start_dt=None, min_points=1):
            if not bucket:
                return None
            chosen = bucket
            if start_dt is not None:
                scoped = [val for dt, val, _ in bucket if dt >= start_dt]
                if len(scoped) >= min_points:
                    chosen = [(None, val, None) for val in scoped]
            return _nemo_numeric_median([val for _, val, _ in chosen])

        if win_start and win_end:
            for row in rows:
                rdt = row.get("_dt")
                if rdt is None or rdt < win_start or rdt > win_end:
                    continue
                # SS-RSRP / SS-SINR / PRB are logged on DIFFERENT rows than the per-second
                # throughput samples (sparse change-event columns), so the two practically
                # never co-occur. Collect RF over ALL rows in the download window — not only
                # throughput-active rows — otherwise the averages come back empty. Simple
                # windowed mean (no throughput weighting; weights would need co-located
                # throughput that doesn't exist). This still excludes the pre-DREQ connect and
                # post-DCOMP idle (RF is restricted to the DREQ→DCOMP transfer window below).
                if (rf_lo is None or rdt >= rf_lo) and (rf_hi is None or rdt <= rf_hi):
                    _append_sample(rf_rsrp, rdt, row.get("rsrpNr"))
                    _append_sample(rf_sinr, rdt, row.get("sinrNr"))
                    _append_sample(rf_prb, rdt, row.get("dlPrbPct"), positive=True)
                    # Band / modulation / rank / NR-presence over the transfer window.
                    _band_text = str(row.get("band") or "").strip()
                    if _nemo_is_valid_band(_band_text):
                        win_banded_rows += 1
                        _bk = _band_text.lower()
                        _tech = (
                            str(row.get("servingTechnology") or "").upper()
                            + " "
                            + str(row.get("packetTechnology") or "").upper()
                        )
                        if _bk.startswith("n") or "5G" in _tech or "EN-DC" in _tech:
                            win_nr_rows += 1
                            if _bk.startswith("n"):
                                win_band_counts[_bk] = win_band_counts.get(_bk, 0) + 1
                    _m0 = _nemo_clean_modulation(row.get("pdschModulationCw0"))
                    _m1 = _nemo_clean_modulation(row.get("pdschModulationCw1"))
                    if _m0 or _m1:
                        win_mod_total += 1
                        if _m0 == "256QAM" or _m1 == "256QAM":
                            win_mod256 += 1
                    _rk = row.get("scheduledRank")
                    if _rk in (None, ""):
                        _rk = row.get("ri")
                    _append_sample(win_rank, rdt, _rk, positive=True)
                if not measurement_title:
                    measurement_title = str(row.get("measurementTitle") or "").strip() or None
                d = str(row.get("dataTransferDirection") or "").strip()
                if d and not direction:
                    direction = d
                p = str(row.get("applicationProtocol") or "").strip()
                if p and not protocol:
                    protocol = p
                bd = row.get("bytesDl")
                if bd is not None and (bytes_dl is None or bd > bytes_dl):
                    bytes_dl = bd
                bu = row.get("bytesUl")
                if bu is not None and (bytes_ul is None or bu > bytes_ul):
                    bytes_ul = bu
                dts = row.get("downloadTimeS")
                if dts is not None and dts > 0:
                    dl_time_kpi = dts
                fs = row.get("fileSizeBytes")
                if fs is not None and fs > 0 and file_size is None:
                    file_size = fs
                st = str(row.get("transferStatus") or "").strip()
                if st.lower() in ("success", "protocol error or timeout", "failed", "aborted", "timeout"):
                    status = st
                mac_total_val = row.get("totalMacDlMbps")
                if mac_total_val in (None, ""):
                    mac_lte = row.get("macDlLteMbps")
                    mac_nr = row.get("macDl5gMbps")
                    if mac_lte not in (None, "") or mac_nr not in (None, ""):
                        try:
                            mac_total_val = float(mac_lte or 0) + float(mac_nr or 0)
                        except (TypeError, ValueError):
                            mac_total_val = None
                app_dl = row.get("appDlMbps")
                app_dl_val = None
                if app_dl is not None:
                    try:
                        app_dl_val = float(app_dl)
                        if app_dl_val > 0:
                            app_samples.append((rdt, app_dl_val))
                    except (TypeError, ValueError):
                        app_dl_val = None
                nr_pdsch_val = row.get("pdschDl5gMbps")
                lte_pdsch_val = row.get("pdschDlLteMbps")
                nr_pdsch_num = None
                lte_pdsch_num = None
                try:
                    nr_pdsch_num = (
                        float(nr_pdsch_val)
                        if nr_pdsch_val not in (None, "")
                        else None
                    )
                except (TypeError, ValueError):
                    nr_pdsch_num = None
                try:
                    lte_pdsch_num = (
                        float(lte_pdsch_val)
                        if lte_pdsch_val not in (None, "")
                        else None
                    )
                except (TypeError, ValueError):
                    lte_pdsch_num = None
                pdsch_total = max(
                    (nr_pdsch_num or 0.0) + (lte_pdsch_num or 0.0),
                    0.0,
                )
                active_slot_weight = None
                if app_dl_val is not None and app_dl_val > 0:
                    active_slot_weight = app_dl_val
                elif pdsch_total > 0:
                    active_slot_weight = pdsch_total
                if active_slot_weight is not None:
                    _append_sample(
                        scell_samples,
                        rdt,
                        row.get("scellsCount"),
                        positive=False,
                    )
                    lat_val = row.get("lat")
                    lon_val = row.get("lon")
                    try:
                        if lat_val is not None and lon_val is not None:
                            active_coords.append(
                                (rdt, float(lat_val), float(lon_val))
                            )
                    except (TypeError, ValueError):
                        pass
                    _append_sample(
                        nr_pdsch,
                        rdt,
                        nr_pdsch_num,
                        positive=True,
                    )
                    _append_sample(
                        lte_pdsch,
                        rdt,
                        lte_pdsch_num,
                        positive=True,
                    )
                    _append_sample(
                        mac_total,
                        rdt,
                        mac_total_val,
                        positive=True,
                    )
                bw_val = row.get("caTotalBwMhz")
                if bw_val in (None, ""):
                    primary_bw = row.get("primaryBwMhz")
                    secondary_bw = row.get("sumSecondaryBwMhz")
                    if primary_bw not in (None, "") and secondary_bw not in (None, ""):
                        try:
                            bw_val = float(primary_bw) + float(secondary_bw)
                        except (TypeError, ValueError):
                            bw_val = primary_bw
                    else:
                        bw_val = primary_bw
                if bw_val is not None and active_slot_weight is not None:
                    _append_sample(bw_samples, rdt, bw_val, positive=True)
        if not bw_samples:
            for row in rows:
                if measurement_title:
                    row_title = str(row.get("measurementTitle") or "").strip()
                    if row_title != measurement_title:
                        continue
                bw_val = row.get("caTotalBwMhz")
                if bw_val in (None, ""):
                    primary_bw = row.get("primaryBwMhz")
                    secondary_bw = row.get("sumSecondaryBwMhz")
                    if primary_bw not in (None, "") and secondary_bw not in (None, ""):
                        try:
                            bw_val = float(primary_bw) + float(secondary_bw)
                        except (TypeError, ValueError):
                            bw_val = primary_bw
                    else:
                        bw_val = primary_bw
                if bw_val is None:
                    continue
                _append_sample(bw_samples, row.get("_dt"), bw_val, positive=True)
        proto_low = protocol.lower()
        dir_low = direction.lower()
        if "ping" in proto_low or "icmp" in proto_low:
            kind = "ping"
        elif dir_low.startswith("up"):
            kind = "upload"
        elif dir_low.startswith("down"):
            kind = "download"
        else:
            kind = "other"
        # Effective transfer time: Nemo "Download time" KPI when logged, else DCOMP−DREQ.
        eff_time = dl_time_kpi if dl_time_kpi else sess.get("downloadDurationS")
        bytes_xfer = bytes_ul if kind == "upload" else bytes_dl
        avg_rate_raw = (
            bytes_xfer * 8 / eff_time / 1e6
            if (bytes_xfer and eff_time and eff_time > 0) else None
        )
        avg_rate = round(avg_rate_raw, 2) if avg_rate_raw is not None else None
        peak_mbps = steady_state_mbps = ramp_up_s = slow_start_loss_pct = peak_to_avg_ratio = None
        steady_state_raw = None
        slow_start_dominated = False
        plateau_start_dt = None
        plateau_samples = []
        steady_sample_count = 0
        prb_util_mean = bw_mhz = mac_total_mean = None
        rsrp_mean = sinr_mean = nr_pdsch_mean = lte_pdsch_mean = None
        mbps_per_mhz = mbps_per_prb_pct = None
        spectral_eff_bps_hz = scheduler_yield_mbps_per_prb_pct = None
        delivery_efficiency_pct = None
        load_state = None
        confidence_class = confidence_note = None
        efficiency_class = None
        if app_samples:
            app_samples.sort(key=lambda item: item[0])
            peak_raw = max(v for _, v in app_samples)
            peak_mbps = round(peak_raw, 1)
            ramp_threshold = peak_raw * 0.9
            ramp_end_dt = next(
                (sample_dt for sample_dt, sample_val in app_samples if sample_val >= ramp_threshold),
                app_samples[0][0],
            )
            transfer_start_dt = sess.get("_dreq_dt") or app_samples[0][0]
            ramp_up_s = round((ramp_end_dt - transfer_start_dt).total_seconds(), 2)
            steady_samples = [sample_val for sample_dt, sample_val in app_samples if sample_dt >= ramp_end_dt]
            plateau_floor = peak_raw * 0.85
            plateau_start_dt = next(
                (sample_dt for sample_dt, sample_val in app_samples if sample_val >= plateau_floor),
                ramp_end_dt,
            )
            plateau_samples = [
                sample_val
                for sample_dt, sample_val in app_samples
                if sample_dt >= plateau_start_dt and sample_val >= plateau_floor
            ]
            steady_sample_count = len(plateau_samples) or len(steady_samples)
            if len(plateau_samples) >= 2:
                steady_state_raw = sum(plateau_samples) / len(plateau_samples)
                steady_state_mbps = round(steady_state_raw, 1)
            elif steady_samples:
                steady_state_raw = sum(steady_samples) / len(steady_samples)
                steady_state_mbps = round(steady_state_raw, 1)
            if steady_state_raw and avg_rate_raw and steady_state_raw > 0:
                slow_start_loss_pct = round((1 - avg_rate_raw / steady_state_raw) * 100, 1)
            if eff_time and ramp_up_s is not None and eff_time > 0:
                slow_start_dominated = ramp_up_s / eff_time >= 0.25
            if avg_rate_raw and avg_rate_raw > 0 and peak_raw > 0:
                peak_to_avg_ratio = round(peak_raw / avg_rate_raw, 2)
        rsrp_mean = _weighted_avg_series(rf_rsrp)
        sinr_mean = _weighted_avg_series(rf_sinr)
        prb_util_mean = _avg_series(rf_prb)
        nr_pdsch_mean = _avg_series(nr_pdsch)
        lte_pdsch_mean = _avg_series(lte_pdsch)
        mac_total_mean = _avg_series(mac_total, plateau_start_dt)
        bw_mhz = _median_series(bw_samples)
        agg_bw_mhz = _avg_series(bw_samples, min_points=1) or bw_mhz
        avg_rank = _avg_series(win_rank, min_points=1)
        scell_count = _avg_series(scell_samples, min_points=1)
        active_slot_count = len(app_samples) or len(
            {dt for dt, _, _ in nr_pdsch} | {dt for dt, _, _ in lte_pdsch}
        )
        rf_sample_count = max(len(rf_rsrp), len(rf_sinr))
        steady_or_avg = steady_state_mbps or avg_rate
        steady_or_avg_raw = steady_state_raw or avg_rate_raw
        # 5G dwell = NR-banded rows / all banded rows in the transfer window (≤100%).
        nr_dwell_pct = (
            round((win_nr_rows / float(win_banded_rows)) * 100.0, 1)
            if win_banded_rows
            else None
        )
        # Per-NR-band dwell = each NR band's share of NR time in the window.
        nr_band_dwell_pct = {}
        if win_nr_rows:
            nr_band_dwell_pct = {
                band: round((count / float(win_nr_rows)) * 100.0, 1)
                for band, count in sorted(win_band_counts.items())
            }
        # 256QAM share = rows at 256QAM / rows with any modulation in the window.
        mod256_pct = (
            round((win_mod256 / float(win_mod_total)) * 100.0, 1)
            if win_mod_total
            else None
        )
        throughput_spread = None
        if app_samples:
            app_vals = [value for _, value in app_samples]
            throughput_spread = {
                "min": round(min(app_vals), 1),
                "p10": round(_nemo_percentile(app_vals, 0.10), 1),
                "p50": round(_nemo_percentile(app_vals, 0.50), 1),
                "p90": round(_nemo_percentile(app_vals, 0.90), 1),
                "max": round(max(app_vals), 1),
                "n": len(app_vals),
            }
        if steady_or_avg_raw and bw_mhz and bw_mhz > 0:
            mbps_per_mhz = round(steady_or_avg_raw / bw_mhz, 2)
            spectral_eff_bps_hz = mbps_per_mhz
        spectral_eff_mbps_per_mhz = mbps_per_mhz
        if steady_or_avg_raw and prb_util_mean and prb_util_mean > 0:
            mbps_per_prb_pct = round(steady_or_avg_raw / prb_util_mean, 2)
            scheduler_yield_mbps_per_prb_pct = mbps_per_prb_pct
        dl_centroid = None
        dl_median_speed_kmh = None
        if active_coords:
            dl_centroid = {
                "lat": round(
                    sum(lat for _, lat, _ in active_coords)
                    / float(len(active_coords)),
                    6,
                ),
                "lon": round(
                    sum(lon for _, _, lon in active_coords)
                    / float(len(active_coords)),
                    6,
                ),
            }
            speed_samples = []
            ordered_coords = sorted(active_coords, key=lambda item: item[0])
            for idx in range(1, len(ordered_coords)):
                prev_dt, prev_lat, prev_lon = ordered_coords[idx - 1]
                curr_dt, curr_lat, curr_lon = ordered_coords[idx]
                delta_s = (curr_dt - prev_dt).total_seconds()
                if delta_s <= 0:
                    continue
                dist_m = _nemo_haversine_m(prev_lat, prev_lon, curr_lat, curr_lon)
                if dist_m is None:
                    continue
                speed_samples.append((dist_m / delta_s) * 3.6)
            if speed_samples:
                dl_median_speed_kmh = round(
                    _nemo_numeric_median(speed_samples),
                    1,
                )
            elif len(active_coords) >= 1:
                dl_median_speed_kmh = 0.0
        if steady_or_avg_raw and mac_total_mean and mac_total_mean > 0:
            delivery_candidate = steady_or_avg_raw / mac_total_mean * 100.0
            if 20 <= delivery_candidate <= 105:
                delivery_efficiency_pct = round(delivery_candidate, 2)
        if prb_util_mean is not None:
            if prb_util_mean < 15:
                load_state = "headroom"
            elif prb_util_mean > 70 or (prb_util_mean >= 50 and (sinr_mean or -999) >= 5):
                load_state = "loaded"
            elif (
                delivery_efficiency_pct is not None
                and delivery_efficiency_pct < 70
                and prb_util_mean < 50
                and (sinr_mean is None or sinr_mean >= 5)
            ):
                load_state = "delivery_limited"
            elif (
                ((sinr_mean is not None and sinr_mean < 0) or (rsrp_mean is not None and rsrp_mean < -105))
                and prb_util_mean < 60
            ):
                load_state = "rf_limited"
            elif (
                delivery_efficiency_pct is not None
                and delivery_efficiency_pct < 80
                and ((sinr_mean is not None and sinr_mean < 5) or prb_util_mean >= 50)
            ):
                load_state = "mixed"
            else:
                load_state = "moderate"
        elif delivery_efficiency_pct is not None and delivery_efficiency_pct < 70:
            load_state = "delivery_limited"
        if load_state in ("headroom", "loaded", "moderate"):
            efficiency_class = load_state
        elif load_state == "rf_limited":
            efficiency_class = "moderate"
        elif load_state == "delivery_limited":
            efficiency_class = "moderate"
        elif load_state == "mixed":
            efficiency_class = "loaded"
        if eff_time and ramp_up_s is not None and eff_time > 0:
            slow_start_dominated = (
                ramp_up_s / eff_time >= 0.25
                and (efficiency_class == "headroom" or efficiency_class is None)
            )
        evidence_count = sum(
            1
            for value in (bw_mhz, prb_util_mean, sinr_mean, delivery_efficiency_pct, mac_total_mean)
            if value is not None
        )
        rf_consistency_issues = []
        if rf_sample_count < 3:
            rf_consistency_issues.append("too_few_rf_samples")
        if (
            sinr_mean is not None
            and sinr_mean < 0
            and (
                (nr_pdsch_mean is not None and nr_pdsch_mean >= 120)
                or (steady_or_avg_raw is not None and steady_or_avg_raw >= 180)
            )
        ):
            rf_consistency_issues.append("sinr_vs_nr_throughput")
        if (
            rsrp_mean is not None
            and rsrp_mean <= -110
            and (
                (steady_or_avg_raw is not None and steady_or_avg_raw >= 150)
                or (nr_pdsch_mean is not None and nr_pdsch_mean >= 120)
            )
        ):
            rf_consistency_issues.append("rsrp_vs_throughput")
        rf_consistency_note = None
        if rf_consistency_issues:
            issue_labels = {
                "too_few_rf_samples": "too few active RF samples",
                "sinr_vs_nr_throughput": "negative SINR with high NR throughput",
                "rsrp_vs_throughput": "very low RSRP with high throughput",
            }
            rf_consistency_note = (
                "RF consistency warning: "
                + ", ".join(issue_labels[issue] for issue in rf_consistency_issues)
                + ". Treat the RF diagnosis as low-confidence and verify the source export."
            )
        if (eff_time and eff_time < 6) or (steady_sample_count < 3 and len(app_samples) < 6):
            confidence_class = "low"
            confidence_note = "Short transfer / sparse steady-state window — directional only."
        elif (
            eff_time and eff_time >= 10
            and (steady_sample_count >= 6 or len(app_samples) >= 8)
            and evidence_count >= 4
        ):
            confidence_class = "high"
            confidence_note = "Sustained transfer with enough RF and layer samples to support the label."
        else:
            confidence_class = "medium"
            confidence_note = "Useful directional signal, but confirm on repeated DTs."
        if rf_consistency_issues:
            confidence_class = "low"
            confidence_note = rf_consistency_note
        slow_start_note = None
        if (
            kind == "download"
            and slow_start_dominated
            and eff_time is not None
            and steady_state_mbps is not None
            and avg_rate is not None
            and slow_start_loss_pct is not None
        ):
            slow_start_note = (
                f"Short {round(eff_time, 1)}s transfer is slow-start dominated: "
                f"steady-state {steady_state_mbps:.1f} Mbps vs file-average {avg_rate:.1f} Mbps "
                f"({slow_start_loss_pct:.1f}% lower). Network capacity is better represented by "
                "the steady-state figure."
            )
        sess.update({
            "kind": kind, "direction": direction, "protocol": protocol, "status": status,
            "success": status.lower().startswith("success"),
            "bytesDl": bytes_dl, "bytesUl": bytes_ul, "fileSizeBytes": file_size,
            "downloadTimeKpiS": dl_time_kpi, "effTransferTimeS": eff_time,
            "avgRateMbps": avg_rate,
            "ssRsrpMean": rsrp_mean, "ssSinrMean": sinr_mean,
            "prbUtilMean": prb_util_mean,
            "nrPdschTput": nr_pdsch_mean, "ltePdschTput": lte_pdsch_mean,
            "macDlTput": mac_total_mean,
            "peakMbps": peak_mbps,
            "steadyStateMbps": steady_state_mbps,
            "steadyStateSampleCount": steady_sample_count or None,
            "rampUpSeconds": ramp_up_s,
            "slowStartLossPct": slow_start_loss_pct,
            "slowStartDominated": slow_start_dominated,
            "peakToAvgRatio": peak_to_avg_ratio,
            "dlSlowStartNote": slow_start_note,
            "bwMHz": bw_mhz,
            "aggBwMhz": agg_bw_mhz,
            "scellCount": scell_count,
            "mbpsPerMHz": mbps_per_mhz,
            "mbpsPerPrbPct": mbps_per_prb_pct,
            "spectralEfficiencyBpsHz": spectral_eff_bps_hz,
            "spectralEffMbpsPerMhz": spectral_eff_mbps_per_mhz,
            "schedulerYieldMbpsPerPrbPct": scheduler_yield_mbps_per_prb_pct,
            "deliveryEfficiencyPct": delivery_efficiency_pct,
            "nrDwellPct": nr_dwell_pct,
            "nrRoutePresencePct": None,
            "nrBandDwellPct": nr_band_dwell_pct,
            "mod256Pct": mod256_pct,
            "avgRank": avg_rank,
            "dlCentroid": dl_centroid,
            "dlMedianSpeedKmh": dl_median_speed_kmh,
            "loadState": load_state,
            "confidenceClass": confidence_class,
            "confidenceLevel": confidence_class,
            "confidenceNote": confidence_note,
            "confidenceReason": confidence_note,
            "efficiencyClass": efficiency_class,
            "throughputSpreadMbps": throughput_spread,
            "dlSampleSpread": throughput_spread,
            "activeSlotCount": active_slot_count or None,
            "rfSampleCount": rf_sample_count or None,
            "rfConsistencyIssues": rf_consistency_issues,
            "rfConsistencyFlags": list(rf_consistency_issues),
            "rfConsistencyNote": rf_consistency_note,
        })

    # Identify the operations (the DT runs ping ×2, one upload, one download).
    download_sess = (
        next((s for s in sessions_raw if s.get("kind") == "download" and s.get("success")), None)
        or next((s for s in sessions_raw if s.get("kind") == "download"), None)
    )
    upload_sess = (
        next((s for s in sessions_raw if s.get("kind") == "upload" and s.get("success")), None)
        or next((s for s in sessions_raw if s.get("kind") == "upload"), None)
    )
    ping_sessions = [s for s in sessions_raw if s.get("kind") == "ping"]

    # Markers + scoping windows are built for the DOWNLOAD session only — the timeline
    # shows exactly the download (ping/upload windows are excluded).
    markers: list[dict] = []
    download_intervals: list[dict] = []
    session_intervals: list[dict] = []
    download_window: list[dict] = []
    if download_sess:
        for eid, dt_key in (
            ("DAA", "_daa_dt"), ("DAC", "_dac_dt"),
            ("DREQ", "_dreq_dt"), ("DCOMP", "_dcomp_dt"), ("DAD", "_dad_dt"),
        ):
            dt_val = download_sess.get(dt_key)
            if dt_val is None:
                continue
            markers.append({
                "type": eid,
                "label": _EVENT_LABELS[eid],
                "ts": _nemo_fmt_hms_ms(dt_val),
                "tSec": dt_val.strftime("%H:%M:%S"),
                "sessionIdx": 0,
            })
        dreq_dt = download_sess.get("_dreq_dt"); dcomp_dt = download_sess.get("_dcomp_dt")
        daa_dt = download_sess.get("_daa_dt"); dad_dt = download_sess.get("_dad_dt")
        if dreq_dt and dcomp_dt and dcomp_dt > dreq_dt:
            download_intervals.append({"start": dreq_dt, "end": dcomp_dt})
        if daa_dt and dad_dt and dad_dt > daa_dt:
            session_intervals.append({"start": daa_dt, "end": dad_dt})
            download_window.append({"start": daa_dt, "end": dad_dt})

    def _clean(s):
        return {k: v for k, v in s.items() if not k.startswith("_")} if s else None

    sessions_out = [_clean(s) for s in sessions_raw]

    # KPIs from the DOWNLOAD session ONLY — no averaging across ping/upload/download (the
    # old bug that made IAM's "download" read 13.7s instead of 4.7s).
    ping_ok = sum(1 for p in ping_sessions if p.get("success"))
    kpis: dict = {}
    if download_sess:
        ds = download_sess
        kpis = {
            "sessionCount": 1,
            # DAC−DAA: data connection establishment time
            "timeToConnectAvgMs": ds.get("timeToConnectMs"),
            "timeToConnectMedianMs": ds.get("timeToConnectMs"),
            # DREQ−DAA: time from session start to download request
            "startDelayAvgS": ds.get("startDelayS"),
            "startDelayMedianS": ds.get("startDelayS"),
            # Nemo "Download time" KPI (DCOMP−DREQ): the true download duration
            "downloadDurationAvgS": ds.get("effTransferTimeS"),
            "downloadDurationMedianS": ds.get("effTransferTimeS"),
            "dlTransferTimeS": ds.get("effTransferTimeS"),
            # Byte-based average rate = Bytes DL × 8 / Download time
            "dlAppRateMbps": ds.get("avgRateMbps"),
            "dlAppTputMbps": ds.get("avgRateMbps"),  # alias for frontend cards
            "dlBytesDl": ds.get("bytesDl"),
            "dlFileSizeBytes": ds.get("fileSizeBytes"),
            "dlSessionTimeS": ds.get("sessionDurationS"),   # DAD−DAA
            "dlStatus": ds.get("status"),
            "dlSuccess": ds.get("success"),
            "dlPeakMbps": ds.get("peakMbps"),
            "dlSteadyStateMbps": ds.get("steadyStateMbps"),
            "dlRampUpSeconds": ds.get("rampUpSeconds"),
            "dlSlowStartLossPct": ds.get("slowStartLossPct"),
            "dlSlowStartDominated": ds.get("slowStartDominated"),
            "dlPeakToAvgRatio": ds.get("peakToAvgRatio"),
            "dlSlowStartNote": ds.get("dlSlowStartNote"),
            "bwMHz": ds.get("bwMHz"),
            "aggBwMhz": ds.get("aggBwMhz"),
            "scellCount": ds.get("scellCount"),
            "mbpsPerMHz": ds.get("mbpsPerMHz"),
            "mbpsPerPrbPct": ds.get("mbpsPerPrbPct"),
            "spectralEfficiencyBpsHz": ds.get("spectralEfficiencyBpsHz"),
            "spectralEffMbpsPerMhz": ds.get("spectralEffMbpsPerMhz"),
            "schedulerYieldMbpsPerPrbPct": ds.get("schedulerYieldMbpsPerPrbPct"),
            "deliveryEfficiencyPct": ds.get("deliveryEfficiencyPct"),
            "nrDwellPct": ds.get("nrDwellPct"),
            "nrRoutePresencePct": ds.get("nrRoutePresencePct"),
            "nrBandDwellPct": ds.get("nrBandDwellPct") or {},
            "mod256Pct": ds.get("mod256Pct"),
            "avgRank": ds.get("avgRank"),
            "dlCentroid": ds.get("dlCentroid"),
            "dlMedianSpeedKmh": ds.get("dlMedianSpeedKmh"),
            "loadState": ds.get("loadState"),
            "confidenceClass": ds.get("confidenceClass"),
            "confidenceLevel": ds.get("confidenceLevel"),
            "confidenceNote": ds.get("confidenceNote"),
            "confidenceReason": ds.get("confidenceReason"),
            "steadyStateSampleCount": ds.get("steadyStateSampleCount"),
            "efficiencyClass": ds.get("efficiencyClass"),
            "throughputSpreadMbps": ds.get("throughputSpreadMbps"),
            "dlSampleSpread": ds.get("dlSampleSpread"),
            "activeSlotCount": ds.get("activeSlotCount"),
            "rfSampleCount": ds.get("rfSampleCount"),
            "rfConsistencyIssues": ds.get("rfConsistencyIssues") or [],
            "rfConsistencyFlags": ds.get("rfConsistencyFlags") or [],
            "rfConsistencyNote": ds.get("rfConsistencyNote"),
            "authoritative": True,
            "source": "timeseries",
            # Upload + ping summaries (also derived from the same time series)
            "hasUpload": upload_sess is not None,
            "ulAppTputMbps": (upload_sess or {}).get("avgRateMbps"),
            "ulTransferTimeS": (upload_sess or {}).get("effTransferTimeS"),
            "ulStatus": (upload_sess or {}).get("status"),
            "ulSuccess": (upload_sess or {}).get("success"),
            "pingCount": len(ping_sessions),
            "pingSuccessCount": ping_ok,
            "pingSuccessPct": round(ping_ok / len(ping_sessions) * 100.0, 1) if ping_sessions else None,
        }

    return {
        "sessions": sessions_out,
        "markers": markers,
        "downloadIntervals": download_intervals,
        "sessionIntervals": session_intervals,
        "downloadWindow": download_window,
        "download": _clean(download_sess),
        "upload": _clean(upload_sess),
        "pings": [_clean(p) for p in ping_sessions],
        "kpis": kpis,
    }


def _nemo_uplink_transfer_intervals(sessions: list[dict]) -> list[dict]:
    from datetime import datetime as _datetime

    intervals = []
    for session in sessions or []:
        direction = str(session.get("direction") or "").strip().lower()
        if not direction.startswith("up"):
            continue
        start_raw = session.get("startTime")
        end_raw = session.get("endTime")
        try:
            start_dt = _datetime.fromisoformat(str(start_raw))
            end_dt = _datetime.fromisoformat(str(end_raw))
        except Exception:
            continue
        if end_dt <= start_dt:
            continue
        intervals.append({"start": start_dt, "end": end_dt})
    intervals.sort(key=lambda item: item["start"])
    merged: list[dict] = []
    for interval in intervals:
        if not merged or interval["start"] > merged[-1]["end"]:
            merged.append(dict(interval))
            continue
        if interval["end"] > merged[-1]["end"]:
            merged[-1]["end"] = interval["end"]
    return merged


def _nemo_rows_within_intervals(rows: list[dict], intervals: list[dict]) -> list[dict]:
    if not rows or not intervals:
        return []
    filtered = []
    for row in rows:
        row_dt = row.get("_dt")
        if row_dt is None:
            continue
        if any(interval.get("start") <= row_dt <= interval.get("end") for interval in intervals or []):
            filtered.append(row)
    return filtered


def _nemo_clip_primary_episodes_to_intervals(episodes: list[dict], intervals: list[dict]) -> list[dict]:
    clipped = []
    for episode in episodes or []:
        ep_start = episode.get("start")
        ep_end = episode.get("end")
        if ep_start is None or ep_end is None:
            continue
        for interval in intervals or []:
            clip_start = max(ep_start, interval.get("start"))
            clip_end = min(ep_end, interval.get("end"))
            if clip_end <= clip_start:
                continue
            recs = []
            for rec in (episode.get("records") or []):
                rec_dt = rec.get("dt")
                in_dt_window = (
                    rec_dt is not None
                    and clip_start <= rec_dt <= clip_end
                )
                in_app_window = any(
                    hasattr(ts, "replace")
                    and clip_start <= ts <= clip_end
                    for ts in (rec.get("appTs") or [])
                )
                if in_dt_window or in_app_window:
                    recs.append(rec)
            clipped_episode = {
                "key": episode.get("key"),
                "records": recs,
                "start": clip_start,
                "end": clip_end,
                "dwellSec": max((clip_end - clip_start).total_seconds(), 0.0),
                "display": _nemo_episode_display_payload(
                    episode,
                    color=(episode.get("display") or {}).get("color"),
                    start_dt=clip_start,
                    end_dt=clip_end,
                    records=recs,
                ),
            }
            clipped.append(clipped_episode)
    clipped.sort(key=lambda item: item.get("start"))
    return clipped


def _nemo_radio_presence_breakdown_from_episodes(episodes: list[dict]) -> dict:
    totals = {"4G": 0.0, "5G": 0.0}
    for item in episodes or []:
        dwell = item.get("dwellSec")
        if dwell is None:
            continue
        tech = str(((item.get("display") or {}).get("tech") or item.get("tech") or "")).upper()
        if tech.startswith(("5G", "NR")):
            totals["5G"] += float(dwell)
        elif tech.startswith("4G"):
            totals["4G"] += float(dwell)
    total = totals["4G"] + totals["5G"]
    if total <= 0:
        return {}
    breakdown = {}
    if totals["5G"] > 0:
        breakdown["5G"] = round(totals["5G"] / total * 100.0, 1)
    if totals["4G"] > 0:
        breakdown["4G"] = round(totals["4G"] / total * 100.0, 1)
    return breakdown


def _nemo_tech_presence_from_timeline(
    timeline: list[tuple],
    intervals: list[dict] | None = None,
) -> dict:
    """Forward-fill a (timestamp, packet/serving-technology) change-event timeline
    and return {"5G": pct, "4G": pct}.

    EN-DC / 5G / NR → "5G";  LTE* → "4G".

    If ``intervals`` ({start: datetime, end: datetime}) are supplied, only seconds
    whose floor falls within those intervals are counted — use for download-window
    breakdowns. Returns {} when no attributable seconds are found (caller should
    fall back to the BDD dwell approach).
    """
    from datetime import datetime as _dt_cls

    ordered = sorted(
        ((dt, val) for dt, val in (timeline or []) if isinstance(dt, _dt_cls)),
        key=lambda x: x[0],
    )
    per_second: dict = {}
    last = None
    for dt, val in ordered:
        cleaned = str(val or "").strip()
        if cleaned:
            last = cleaned
        if last is None:
            continue
        per_second[dt.replace(microsecond=0)] = last

    if intervals:
        def _in_ivs(sec: "_dt_cls") -> bool:
            for iv in intervals:
                s, e = iv.get("start"), iv.get("end")
                if s is not None and e is not None and s <= sec <= e:
                    return True
            return False
        per_second = {s: v for s, v in per_second.items() if _in_ivs(s)}

    g5 = g4 = 0
    for val in per_second.values():
        u = str(val).upper()
        if "EN-DC" in u or "5G" in u or ("NR" in u and "LTE" not in u):
            g5 += 1
        elif "LTE" in u:
            g4 += 1

    total = g5 + g4
    if not total:
        return {}
    result: dict = {}
    if g5:
        result["5G"] = round(g5 / total * 100.0, 1)
    if g4:
        result["4G"] = round(g4 / total * 100.0, 1)
    return result


def _nemo_episode_dwell_by_key(episodes: list[dict]) -> dict:
    totals: dict = {}
    for item in episodes or []:
        key = item.get("key")
        dwell = item.get("dwellSec")
        if key is None or dwell is None:
            continue
        totals[key] = round(totals.get(key, 0.0) + float(dwell), 0)
    return totals


def _nemo_radio_presence_totals_from_cells(
    cells_payload: list[dict],
    dwell_field: str = "dwellSec",
) -> dict:
    totals = {"4G": 0.0, "5G": 0.0}
    for item in cells_payload or []:
        dwell = item.get(dwell_field)
        if dwell is None:
            continue
        tech = str(item.get("tech") or "").upper()
        if tech.startswith(("5G", "NR")):
            totals["5G"] += float(dwell)
        elif tech.startswith("4G"):
            totals["4G"] += float(dwell)
    return {
        "4G": round(totals["4G"], 0),
        "5G": round(totals["5G"], 0),
    }


def _nemo_presence_share_from_cells(cells_payload: list[dict], dwell_field: str) -> dict:
    totals: dict = {}
    total_dwell = 0.0
    for item in cells_payload or []:
        dwell = item.get(dwell_field)
        if dwell is None:
            continue
        if any(item.get(field) is not None for field in ("siteName", "tech", "band")):
            key = (
                item.get("cellName"),
                item.get("siteName"),
                item.get("tech"),
                item.get("band"),
            )
        else:
            key = item.get("cellName")
        dwell_value = float(dwell)
        totals[key] = totals.get(key, 0.0) + dwell_value
        total_dwell += dwell_value
    if total_dwell <= 0:
        return {}
    return {
        key: round(value / total_dwell * 100.0, 1)
        for key, value in totals.items()
    }


def _nemo_read_tabular_file(path: str) -> tuple[str, list[str], list[list[str]]]:
    import io

    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        text = f.read()
    delimiter = _nemo_guess_delimiter(text[:10000])
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    if not rows:
        return delimiter, [], []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return delimiter, headers, rows[1:]


def _nemo_header_index_map(headers: list[str]) -> dict:
    mapping = {}
    for idx, header in enumerate(headers or []):
        key = str(header or "").strip()
        mapping.setdefault(key, []).append(idx)
    return mapping


def _nemo_duplicate_headers(headers: list[str]) -> list[dict]:
    duplicates = []
    for name, indexes in _nemo_header_index_map(headers).items():
        if name and len(indexes) > 1:
            duplicates.append({"name": name, "indexes": indexes})
    return duplicates


def _nemo_resolve_indices(header_map: dict, aliases: tuple) -> list:
    """Resolve an alias tuple to a flat, ordered list of column indices — once per file.

    The result is memoized on header_map (under a sentinel key) so the per-row picks avoid
    millions of dict lookups and alias iterations. Flattening in alias order preserves the
    original "first non-empty alias wins" semantics.
    """
    cache = header_map.get("\x00idx")
    if cache is None:
        cache = {}
        header_map["\x00idx"] = cache
    resolved = cache.get(aliases)
    if resolved is None:
        resolved = []
        for alias in aliases:
            resolved.extend(header_map.get(alias) or [])
        cache[aliases] = resolved
    return resolved


def _nemo_pick_text(row: list[str], header_map: dict, *aliases) -> str:
    n = len(row)
    for idx in _nemo_resolve_indices(header_map, aliases):
        if idx < n:
            value = row[idx]
            if value:
                text = value.strip() if isinstance(value, str) else _benchmark_text(value)
                if text:
                    return text
    return ""


def _nemo_pick_all_texts(row: list[str], header_map: dict, *aliases) -> list[str]:
    values = []
    seen = set()
    n = len(row)
    for idx in _nemo_resolve_indices(header_map, aliases):
        if idx < n:
            value = row[idx]
            if value:
                text = value.strip() if isinstance(value, str) else _benchmark_text(value)
                if text and text not in seen:
                    seen.add(text)
                    values.append(text)
    return values


def _nemo_pick_float(row: list[str], header_map: dict, *aliases):
    n = len(row)
    for idx in _nemo_resolve_indices(header_map, aliases):
        if idx < n:
            value = row[idx]
            if value:  # most cells are empty strings — skip coercion entirely
                num = _benchmark_float(value)
                if num is not None:
                    return num
    return None


def _nemo_pick_time(row: list[str], header_map: dict):
    for idx in header_map.get("Time") or []:
        if idx < len(row):
            parsed = _nemo_parse_time(row[idx])
            if parsed is not None:
                return parsed
    return None


def _nemo_pick_text_resolved(row: list[str], indices: list[int], n: int):
    for idx in indices:
        if idx < n:
            value = row[idx]
            if value:
                text = value.strip() if isinstance(value, str) else _benchmark_text(value)
                if text:
                    return text
    return ""


def _nemo_pick_all_texts_resolved(row: list[str], indices: list[int], n: int) -> list[str]:
    values = []
    seen = set()
    for idx in indices:
        if idx < n:
            value = row[idx]
            if value:
                text = value.strip() if isinstance(value, str) else _benchmark_text(value)
                if text and text not in seen:
                    seen.add(text)
                    values.append(text)
    return values


def _nemo_pick_float_resolved(row: list[str], indices: list[int], n: int):
    for idx in indices:
        if idx < n:
            value = row[idx]
            if value:
                num = _benchmark_float(value)
                if num is not None:
                    return num
    return None


def _nemo_pick_time_resolved(row: list[str], indices: list[int], n: int):
    for idx in indices:
        if idx < n:
            parsed = _nemo_parse_time(row[idx])
            if parsed is not None:
                return parsed
    return None


def _nemo_classify_family(payload: dict) -> str:
    if payload.get("appDlRaw") is not None or payload.get("appDlAvgRaw") is not None or payload.get("transferStatus"):
        return "throughput_app"
    if payload.get("macDl5gRaw") is not None or payload.get("totalMacDlRaw") is not None or payload.get("macDlLteRaw") is not None:
        return "throughput_mac"
    if payload.get("pdschSched5gRaw") is not None or payload.get("pdschDl5gRaw") is not None or payload.get("pdschPrbs") is not None:
        return "pdsch"
    if payload.get("rsrp") is not None or payload.get("rsrq") is not None or payload.get("sinr") is not None:
        return "rf"
    if payload.get("ri") is not None or payload.get("wbCqi") is not None or payload.get("macDlBler") is not None:
        return "mimo_cqi"
    if payload.get("tcpHandshakeMs") is not None or payload.get("lostPacket") is not None or payload.get("pingStatus"):
        return "transport_ping"
    if payload.get("servingTechnology") or payload.get("packetTechnology") or payload.get("band") or payload.get("bandwidthPrbs") is not None:
        return "serving_config"
    return "other"


def _nemo_normalize_throughput_columns(rows: list[dict]):
    throughput_fields = [
        ("appDlRaw", "appDlMbps"),
        ("appDlAvgRaw", "appDlAvgMbps"),
        ("macDlLteRaw", "macDlLteMbps"),
        ("macDl5gRaw", "macDl5gMbps"),
        ("totalMacDlRaw", "totalMacDlMbps"),
        ("pdschSched5gRaw", "pdschSched5gMbps"),
        ("pdschDl5gRaw", "pdschDl5gMbps"),
    ]
    scales = {}
    for raw_key, normalized_key in throughput_fields:
        scales[normalized_key] = 1_000_000.0
    for row in rows:
        for raw_key, normalized_key in throughput_fields:
            raw_value = row.get(raw_key)
            if raw_value is None:
                continue
            row[normalized_key] = round(float(raw_value) / float(scales[normalized_key]), 3)
    return scales


def _nemo_infer_throughput_scales(raw_rows: list[dict]) -> dict:
    """Infer whether Nemo throughput columns are already expressed in Mbps or still in bps.

    Some exports expose App/MAC/PDSCH throughput directly in Mbps (e.g. values like 107.158),
    while others use bit/s and need division by 1e6. Using a fixed scale breaks datasets like
    Settat by collapsing valid Mbps values to ~0.000. We infer per-column scale from the
    observed magnitude of positive samples.
    """
    throughput_fields = [
        ("appDlRaw", "appDlMbps"),
        ("appDlAvgRaw", "appDlAvgMbps"),
        ("appUlRaw", "appUlMbps"),
        ("macDlLteRaw", "macDlLteMbps"),
        ("macDl5gRaw", "macDl5gMbps"),
        ("totalMacDlRaw", "totalMacDlMbps"),
        ("pdschSched5gRaw", "pdschSched5gMbps"),
        ("pdschDl5gRaw", "pdschDl5gMbps"),
        ("pdschDlLteRaw", "pdschDlLteMbps"),
        ("pdschDlLteCw0Raw", "pdschDlLteCw0Mbps"),
        ("pdschDlLteCw1Raw", "pdschDlLteCw1Mbps"),
    ]
    scales = {}
    for raw_key, normalized_key in throughput_fields:
        positive_values = []
        for row in raw_rows or []:
            value = row.get(raw_key)
            if value is None:
                continue
            try:
                numeric = float(value)
            except Exception:
                continue
            if not math.isfinite(numeric) or numeric <= 0:
                continue
            positive_values.append(abs(numeric))
            if len(positive_values) >= 256:
                break
        if not positive_values:
            scales[normalized_key] = 1_000_000.0
            continue
        max_value = max(positive_values)
        # Nemo Mbps exports typically stay in the 0-1000 range. Bit/s exports are many orders
        # of magnitude larger (millions). Treat anything already below 10k as Mbps-scale.
        scales[normalized_key] = 1.0 if max_value < 10_000.0 else 1_000_000.0
    return scales


def _nemo_reapply_throughput_normalization(operator_file: dict) -> dict:
    """Recompute normalized throughput columns from preserved raw Nemo values.

    Stored benchmark-library records keep parsed rows, including the original `*Raw` fields.
    When normalization logic changes, we can repair cached rows without reparsing TXT files.
    """
    if not isinstance(operator_file, dict):
        return operator_file
    rows = operator_file.get("rows") or []
    scales = _nemo_infer_throughput_scales(rows)
    for row in rows:
        if not isinstance(row, dict):
            continue
        mapping = (
            ("appDlRaw", "appDlMbps"),
            ("appDlAvgRaw", "appDlAvgMbps"),
            ("appUlRaw", "appUlMbps"),
            ("macDlLteRaw", "macDlLteMbps"),
            ("macDl5gRaw", "macDl5gMbps"),
            ("totalMacDlRaw", "totalMacDlMbps"),
            ("pdschSched5gRaw", "pdschSched5gMbps"),
            ("pdschDl5gRaw", "pdschDl5gMbps"),
            ("pdschDlLteRaw", "pdschDlLteMbps"),
            ("pdschDlLteCw0Raw", "pdschDlLteCw0Mbps"),
            ("pdschDlLteCw1Raw", "pdschDlLteCw1Mbps"),
        )
        for raw_key, norm_key in mapping:
            raw_value = row.get(raw_key)
            if raw_value is None:
                row.pop(norm_key, None)
                continue
            try:
                row[norm_key] = round(float(raw_value) / float(scales[norm_key]), 3)
            except Exception:
                row.pop(norm_key, None)
    operator_file["throughputScales"] = scales
    return operator_file


def _nemo_classify_transfer_operation(protocol: str, direction: str) -> str:
    """Classify a transfer-session-statistics row into ping / upload / download / other
    from its Application protocol + Transfer direction."""
    p = str(protocol or "").lower()
    d = str(direction or "").lower()
    if "ping" in p or "icmp" in p:
        return "ping"
    if "up" in d:        # Uplink
        return "upload"
    if "down" in d:      # Downlink
        return "download"
    return "other"


def _nemo_norm_name(name: str) -> str:
    """Lowercase and collapse separators (spaces/underscores) so filename matching works
    whether the file keeps its original spaces or was sanitized to underscores on upload."""
    return re.sub(r"[ _]+", " ", str(name or "").lower()).strip()


def _nemo_is_session_stats_file(path: str) -> bool:
    return "session statistic" in _nemo_norm_name(os.path.basename(str(path or "")))


def _nemo_find_session_stats_path(ts_path: str):
    """Locate the 'Data transfer session statistics' sibling of a time-series export.

    Given `/dir/Mohammedia-IAM.txt`, finds the matching session-statistics file in the same
    directory — tolerant of the 'transfert'/'transfer' spelling and of spaces being
    sanitized to underscores by the upload handler."""
    if not ts_path:
        return None
    directory = os.path.dirname(ts_path) or "."
    base = os.path.splitext(os.path.basename(ts_path))[0]
    base_norm = _nemo_norm_name(base)
    try:
        for fname in os.listdir(directory):
            if not fname.lower().endswith(".txt"):
                continue
            norm = _nemo_norm_name(fname)
            if base_norm in norm and "session statistic" in norm:
                return os.path.join(directory, fname)
    except OSError:
        pass
    return None


def _nemo_parse_session_stats(path: str) -> dict:
    """Parse a Nemo 'Data transfer session statistics' export.

    One row per transfer operation (ICMP ping / HTTP upload / HTTP download) carrying
    Nemo's own AUTHORITATIVE per-session KPIs — transfer time, app throughput, service
    access + TCP handshake latency, transfer status, and per-session RF/PHY means scoped
    exactly to that operation. This is ground truth: far more accurate than reconstructing
    sessions from the per-second time series, where ping/upload/download all look alike.

    Returns {operator, fileName, sessions:[...], download, upload, pings:[...], kpis:{...}}
    or {} when the file is missing/unreadable.
    """
    if not path or not os.path.isfile(path):
        return {}
    try:
        delimiter, headers, data_rows = _nemo_read_tabular_file(path)
    except Exception:
        return {}
    if not headers:
        return {}
    header_map = _nemo_header_index_map(headers)
    resolve = _nemo_resolve_indices
    operator = _nemo_guess_operator(path)

    idx = {
        "time":       resolve(header_map, ("Time",)),
        "protocol":   resolve(header_map, ("Application protocol",)),
        "direction":  resolve(header_map, ("Transfer direction", "Data transfer direction")),
        "operation":  resolve(header_map, ("Operation",)),
        "fileSize":   resolve(header_map, ("File size",)),
        "bytesDl":    resolve(header_map, ("Bytes DL",)),
        "bytesUl":    resolve(header_map, ("Bytes UL",)),
        "status":     resolve(header_map, ("Transfer status", "Transf. status")),
        "timeout":    resolve(header_map, ("Timeout",)),
        "svcAccessMs": resolve(header_map, ("Service access time",)),
        "ipTermMs":   resolve(header_map, ("IP term. time",)),
        "tcpHsMs":    resolve(header_map, ("TCP handshake time",)),
        "appTputDl":  resolve(header_map, ("Application tput DL (Mbps)",)),
        "appTputUl":  resolve(header_map, ("Application tput UL (Mbps)",)),
        "xferTimeS":  resolve(header_map, ("Data transfer time (s)",)),
        "ltePct":     resolve(header_map, ("LTE pct.",)),
        "lteCaPct":   resolve(header_map, ("LTE CA pct.",)),
        "endcPct":    resolve(header_map, ("EN-DC pct.",)),
        "saNrPct":    resolve(header_map, ("SA NR pct.",)),
        "ssRsrpMean": resolve(header_map, ("SS-RSRP mean",)),
        "ssRsrpMin":  resolve(header_map, ("SS-RSRP min",)),
        "ssRsrqMean": resolve(header_map, ("SS-RSRQ mean",)),
        "ssSinrMean": resolve(header_map, ("SS-SINR mean",)),
        "ssSinrMin":  resolve(header_map, ("SS-SINR min",)),
        "nrPdschTput": resolve(header_map, ("NR PDSCH tput (Mbps)",)),
        "nrPuschTput": resolve(header_map, ("NR PUSCH tput (Mbps)",)),
        "ltePdschTput": resolve(header_map, ("LTE PDSCH tput (Mbps)",)),
        "ltePuschTput": resolve(header_map, ("LTE PUSCH tput (Mbps)",)),
        "lteRsrpMean": resolve(header_map, ("RSRP mean",)),
        "lteSinrMean": resolve(header_map, ("SNR mean",)),
        "prbUtilMean": resolve(header_map, ("PRB utilization mean",)),
        "wbCqi0Mode": resolve(header_map, ("WB CQI 0 mode",)),
    }

    _TEXT_FIELDS = {"protocol", "direction", "operation", "status"}
    sessions = []
    for raw in data_rows or []:
        if not raw or all((str(c).strip() == "" for c in raw)):
            continue
        n = len(raw)
        rec = {}
        for field, indices in idx.items():
            if field == "time":
                continue
            if field in _TEXT_FIELDS:
                rec[field] = _nemo_pick_text_resolved(raw, indices, n)
            else:
                rec[field] = _nemo_pick_float_resolved(raw, indices, n)
        dt = _nemo_pick_time_resolved(raw, idx["time"], n)
        rec["_dt"] = dt
        rec["time"] = _nemo_fmt_hms_ms(dt) if dt else ""
        rec["timeFull"] = dt.strftime("%Y-%m-%d %H:%M:%S.") + f"{dt.microsecond // 1000:03d}" if dt else ""
        rec["kind"] = _nemo_classify_transfer_operation(rec.get("protocol"), rec.get("direction"))
        status = str(rec.get("status") or "")
        rec["success"] = status.lower().startswith("success")
        # Total setup latency to start the transfer (service access + TCP handshake).
        sa = rec.get("svcAccessMs")
        tcp = rec.get("tcpHsMs")
        rec["setupLatencyMs"] = (
            round((sa or 0) + (tcp or 0), 1) if (sa is not None or tcp is not None) else None
        )
        sessions.append(rec)

    def _pick_session(kind, want_success=True):
        # Prefer a successful session of this kind; else the first of that kind.
        cands = [s for s in sessions if s.get("kind") == kind]
        if want_success:
            ok = [s for s in cands if s.get("success")]
            if ok:
                return ok[0]
        return cands[0] if cands else None

    download = _pick_session("download")
    upload = _pick_session("upload")
    pings = [s for s in sessions if s.get("kind") == "ping"]

    def _strip_dt(rec):
        if not rec:
            return None
        return {k: v for k, v in rec.items() if k != "_dt"}

    ping_ok = sum(1 for p in pings if p.get("success"))
    kpis = {
        "operationCount": len(sessions),
        "hasDownload": download is not None,
        "hasUpload": upload is not None,
        "pingCount": len(pings),
        "pingSuccessCount": ping_ok,
        "pingSuccessPct": round(ping_ok / len(pings) * 100.0, 1) if pings else None,
        # Authoritative download KPIs
        "dlTransferTimeS":  (download or {}).get("xferTimeS"),
        "dlAppTputMbps":    (download or {}).get("appTputDl"),
        "dlSetupLatencyMs": (download or {}).get("setupLatencyMs"),
        "dlServiceAccessMs": (download or {}).get("svcAccessMs"),
        "dlTcpHandshakeMs": (download or {}).get("tcpHsMs"),
        "dlStatus":         (download or {}).get("status"),
        "dlSuccess":        (download or {}).get("success"),
        # Authoritative upload KPIs
        "ulTransferTimeS":  (upload or {}).get("xferTimeS"),
        "ulAppTputMbps":    (upload or {}).get("appTputUl"),
        "ulSetupLatencyMs": (upload or {}).get("setupLatencyMs"),
        "ulStatus":         (upload or {}).get("status"),
        "ulSuccess":        (upload or {}).get("success"),
    }

    return {
        "operator": operator,
        "fileName": os.path.basename(path),
        "path": path,
        "sessions": [_strip_dt(s) for s in sessions],
        "download": _strip_dt(download),
        "upload": _strip_dt(upload),
        "pings": [_strip_dt(p) for p in pings],
        "kpis": kpis,
    }


_NEMO_PARSE_CACHE: dict = {}
_NEMO_PARSE_CACHE_ORDER: list = []


def _nemo_parse_operator_file(path: str) -> dict:
    """Cached wrapper keyed by (path, mtime). The benchmark upload parses the IAM file twice
    (once for the analysis, once to register the sidebar run); the cache removes the duplicate.
    Bounded to the last few files. Consumers only read the parsed rows / derived keys idempotently."""
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = None
    key = (path, mtime)
    if mtime is not None:
        cached = _NEMO_PARSE_CACHE.get(key)
        if cached is not None:
            return cached
    result = _nemo_parse_operator_file_uncached(path)
    if mtime is not None:
        _NEMO_PARSE_CACHE[key] = result
        _NEMO_PARSE_CACHE_ORDER.append(key)
        while len(_NEMO_PARSE_CACHE_ORDER) > 3:
            _NEMO_PARSE_CACHE.pop(_NEMO_PARSE_CACHE_ORDER.pop(0), None)
    return result


def _nemo_parse_operator_file_uncached(path: str) -> dict:
    delimiter, headers, data_rows = _nemo_read_tabular_file(path)
    header_map = _nemo_header_index_map(headers)
    operator = _nemo_guess_operator(path)
    file_name = os.path.basename(path)
    resolve = _nemo_resolve_indices

    time_indices = header_map.get("Time") or []
    measurement_title_indices = resolve(header_map, ("Measurement Title", "measurement"))
    serving_technology_indices = resolve(header_map, ("Serving technology",))
    packet_technology_indices = resolve(header_map, ("Packet technology",))
    app_protocol_indices = resolve(header_map, ("Application protocol",))
    transfer_status_indices = resolve(header_map, ("Transf. status", "Data Transfer Status", "Nemo status"))
    cell_type_indices = resolve(header_map, ("Cell type",))
    band_indices = resolve(header_map, ("Band", "Serving band"))
    lon_indices = resolve(header_map, ("Lon.", "Lon", "Longitude"))
    lat_indices = resolve(header_map, ("Lat.", "Lat", "Latitude"))
    system_indices = resolve(header_map, ("System",))
    bandwidth_prbs_indices = resolve(header_map, ("Bandwidth in PRBs",))
    scells_count_indices = resolve(header_map, ("#SCells",))
    lte_ca_status_indices = resolve(header_map, ("LTE CA status",))
    nr_ca_status_indices = resolve(header_map, ("NR CA status",))
    pci_indices = resolve(header_map, ("NR PCI", "LTE PCI", "PCI"))
    lte_channel_indices = resolve(header_map, ("LTE channel number", "Downlink EARFCN", "EARFCN"))
    nr_channel_indices = resolve(header_map, ("NR channel number", "NR-ARFCN"))
    rrc_state_indices = resolve(header_map, ("RRC State", "E-RRC State", "Generic RRC State"))
    app_dl_indices = resolve(header_map, ("App. rate DL",))
    app_dl_avg_indices = resolve(header_map, ("App rate DL avg",))
    app_ul_indices = resolve(header_map, ("App. rate UL",))
    download_time_indices = resolve(header_map, ("Download time",))
    bytes_dl_indices = resolve(header_map, ("Bytes DL",))
    bytes_ul_indices = resolve(header_map, ("Bytes UL",))
    file_size_indices = resolve(header_map, ("File size",))
    transfer_direction_indices = resolve(header_map, ("Data transfer direction",))
    transfer_filename_indices = resolve(header_map, ("Filename",))
    mac_dl_lte_indices = resolve(header_map, ("MAC DL throughput (LTE)",))
    mac_dl_5g_indices = resolve(header_map, ("MAC DL throughput (5G)",))
    total_mac_dl_indices = resolve(header_map, ("Total MAC DL throughput",))
    pdsch_sched_5g_indices = resolve(header_map, ("PDSCH DL scheduled throughput (5G)",))
    pdsch_dl_5g_indices = resolve(header_map, ("PDSCH DL throughput (5G)",))
    pdsch_dl_lte_indices = resolve(header_map, ("PDSCH DL throughput (LTE)",))
    pdsch_dl_lte_cw0_indices = resolve(header_map, ("PDSCH DL throughput 0 (LTE)",))
    pdsch_dl_lte_cw1_indices = resolve(header_map, ("PDSCH DL throughput 1 (LTE)",))
    pdsch_prbs_indices = resolve(header_map, ("PDSCH PRBs",))
    pdsch_mod_cw0_indices = resolve(header_map, ("PDSCH modulation codeword 0", "Modulation 0"))
    pdsch_mod_cw1_indices = resolve(header_map, ("PDSCH modulation codeword 1", "Modulation 1"))
    pdsch_mcs_cw0_indices = resolve(header_map, ("PDSCH MCS index for codeword 0",))
    pdsch_mcs_cw1_indices = resolve(header_map, ("PDSCH MCS index for codeword 1",))
    pdsch_tbs_cw0_indices = resolve(header_map, ("PDSCH transport block size for codeword 0",))
    pdsch_tbs_cw1_indices = resolve(header_map, ("PDSCH transport block size for codeword 1",))
    pdsch_bits_per_hz_indices = resolve(header_map, ("PDSCH bit/s/Hz",))
    pdsch_max_bits_per_hz_indices = resolve(header_map, ("Max PDSCH bit/s/Hz",))
    pdsch_slot_pct_indices = resolve(header_map, (
        "PDSCH scheduling ratio",
        "PDSCH slot %",
        "NR PDSCH scheduling ratio",
        "PDSCH DL scheduling ratio",
        "Scheduling ratio DL",
        "PDSCH scheduled slots %",
    ))
    rsrp_indices = resolve(header_map, ("RSRP",))
    rsrq_indices = resolve(header_map, ("RSRQ",))
    sinr_indices = resolve(header_map, ("SINR",))
    # Nemo duplicates the RF columns: the NR group (next to NR-ARFCN/NR-PCI) comes first in
    # the header, the LTE serving group second. Split by column index so the per-RAT values
    # are read directly from their own column — robust even when the sparse `band` column
    # isn't present on the exact RF row (the band-based split silently dropped LTE SINR,
    # whose rows carry no band).
    rsrp_nr_indices, rsrp_lte_indices = rsrp_indices[:1], rsrp_indices[1:2]
    rsrq_nr_indices, rsrq_lte_indices = rsrq_indices[:1], rsrq_indices[1:2]
    sinr_nr_indices, sinr_lte_indices = sinr_indices[:1], sinr_indices[1:2]
    ri_indices = resolve(header_map, ("RI",))
    wb_cqi_indices = resolve(header_map, ("WB CQI",))
    scheduled_rank_indices = resolve(header_map, ("PDSCH scheduled rank",))
    mac_dl_bler_indices = resolve(header_map, ("MAC DL BLER",))
    mac_ul_retx_5g_indices = resolve(header_map, ("MAC UL retransmission rate (5G)",))
    tcp_handshake_indices = resolve(header_map, ("TCP handshake time (ms)",))
    lost_packet_indices = resolve(header_map, ("Lost packet",))
    ping_status_indices = resolve(header_map, ("Ping status",))
    # ── Additional capacity / reliability / CA / UL metrics ──
    dl_prb_pct_indices = resolve(header_map, ("DL PRB %",))
    prbs_avg_dl_indices = resolve(header_map, ("PRBs Avg DL",))
    sch_bitrate_per_prb_indices = resolve(header_map, ("Sch bitrate/PRB",))
    pdsch_bler_lte_indices = resolve(header_map, ("PDSCH BLER",))
    mac_dl_residual_bler_indices = resolve(header_map, ("MAC DL residual BLER",))
    pdcch_bler_indices = resolve(header_map, ("PDCCH BLER est.",))
    mac_ul_retx_lte_indices = resolve(header_map, ("MAC UL retransmission rate (LTE)",))
    ca_total_bw_indices = resolve(header_map, ("CA total BW (MHz)",))
    primary_bw_indices = resolve(header_map, ("Primary BW (MHz)",))
    sum_secondary_bw_indices = resolve(header_map, ("Sum Secondary BWs (MHz)",))
    tx_power_indices = resolve(header_map, ("TX power",))
    pusch_tx_power_indices = resolve(header_map, ("PUSCH TX power",))
    wb_cqi0_indices = resolve(header_map, ("WB CQI 0",))
    wb_cqi1_indices = resolve(header_map, ("WB CQI 1",))
    ho_uplane_interruption_indices = resolve(header_map, ("HO U-plane interruption",))
    ppp_rate_dl_indices = resolve(header_map, ("PPP rate DL",))
    recv_ppp_bytes_indices = resolve(header_map, ("Recv. PPP bytes",))
    device_model_indices = resolve(header_map, (
        "Device name",
        "Device label",
        "Device",
        "Terminal name",
        "Terminal",
        "Equipment",
        "UE model",
        "Model",
        "Device model",
        "Phone model",
    ))
    # Event ID column (BF, index 57): DAA / DAC / DAD / DREQ / DCOMP markers
    event_id_indices = resolve(header_map, ("Event ID",))
    event_text_indices = resolve(header_map, ("Event",))

    normalized_rows = []
    raw_rows = []
    measurement_titles = set()
    ordered_dt_titles = []
    rows_by_measurement_title = {}
    last_title = ""
    serving_tech_timeline = []
    packet_tech_timeline = []
    coverage = {
        "rowCount": 0,
        "nonEmptyAppDl": 0,
        "nonEmptyTransferStatus": 0,
        "nonEmptyMac5g": 0,
        "nonEmptyPdsch5g": 0,
        "nonEmptyNrChannel": 0,
        "nonEmptyRsrp": 0,
        "nonEmptySinr": 0,
        "nonEmptyRi": 0,
        "nonEmptyBler": 0,
    }
    scg_pscell_samples = 0
    nr_channel_non_null_count = 0
    mac_dl_5g_positive_count = 0
    pdsch_dl_5g_positive_count = 0
    nr_bands = set()
    nr_presence_by_second = {}
    has_5g = False
    device_models_seen = []
    parsing_qa = {
        "normalizedValueCount": 0,
        "normalizedFields": {},
        "normalizedSamples": [],
    }
    throughput_scales = {
        "appDlMbps": 1_000_000.0,
        "appDlAvgMbps": 1_000_000.0,
        "appUlMbps": 1_000_000.0,
        "macDlLteMbps": 1_000_000.0,
        "macDl5gMbps": 1_000_000.0,
        "totalMacDlMbps": 1_000_000.0,
        "pdschSched5gMbps": 1_000_000.0,
        "pdschDl5gMbps": 1_000_000.0,
        "pdschDlLteMbps": 1_000_000.0,
        "pdschDlLteCw0Mbps": 1_000_000.0,
        "pdschDlLteCw1Mbps": 1_000_000.0,
    }
    _NR_CT_UPPER = {"NR SERVING", "NR SCG PSCELL", "SCG PSCELL", "5G SERVING"}

    for row_number, row in enumerate(data_rows, start=2):
        n = len(row)
        event_time = _nemo_pick_time_resolved(row, time_indices, n)
        measurement_title = _nemo_pick_text_resolved(row, measurement_title_indices, n)
        serving_technology = _nemo_pick_text_resolved(row, serving_technology_indices, n)
        packet_technology = _nemo_pick_text_resolved(row, packet_technology_indices, n)
        app_protocol = _nemo_pick_text_resolved(row, app_protocol_indices, n)
        transfer_status = _nemo_pick_text_resolved(row, transfer_status_indices, n)
        cell_types = _nemo_pick_all_texts_resolved(row, cell_type_indices, n)
        band_text = _nemo_pick_text_resolved(row, band_indices, n)
        device_model = _nemo_pick_text_resolved(row, device_model_indices, n)
        # The "Measurement Title" (DT / "BJ" name) is only populated on the radio/serving rows that
        # start each DT block; transfer rows leave it blank. Forward-fill it so every row — including
        # the data-transfer rows — carries its DT name.
        if measurement_title:
            measurement_titles.add(measurement_title)
            last_title = measurement_title
        else:
            measurement_title = last_title
        if device_model:
            device_models_seen.append(device_model)

        app_dl_raw = _nemo_pick_float_resolved(row, app_dl_indices, n)
        app_dl_avg_raw = _nemo_pick_float_resolved(row, app_dl_avg_indices, n)
        app_ul_raw = _nemo_pick_float_resolved(row, app_ul_indices, n)
        mac_dl_lte_raw = _nemo_pick_float_resolved(row, mac_dl_lte_indices, n)
        mac_dl_5g_raw = _nemo_pick_float_resolved(row, mac_dl_5g_indices, n)
        total_mac_dl_raw = _nemo_pick_float_resolved(row, total_mac_dl_indices, n)
        pdsch_sched_5g_raw = _nemo_pick_float_resolved(row, pdsch_sched_5g_indices, n)
        pdsch_dl_5g_raw = _nemo_pick_float_resolved(row, pdsch_dl_5g_indices, n)
        pdsch_dl_lte_raw = _nemo_pick_float_resolved(row, pdsch_dl_lte_indices, n)
        pdsch_dl_lte_cw0_raw = _nemo_pick_float_resolved(row, pdsch_dl_lte_cw0_indices, n)
        pdsch_dl_lte_cw1_raw = _nemo_pick_float_resolved(row, pdsch_dl_lte_cw1_indices, n)
        nr_channel_number = _nemo_pick_float_resolved(row, nr_channel_indices, n)
        serving_technology_upper = serving_technology.upper()
        packet_technology_upper = packet_technology.upper()
        cell_types_upper = {str(cell or "").strip().upper() for cell in cell_types}

        # NR/LTE RF split read directly from each RAT's own column (NR group first, LTE
        # second). Conflated value = first non-empty (NR preferred), matching prior behaviour.
        rsrp_nr_val = _nemo_pick_float_resolved(row, rsrp_nr_indices, n)
        rsrp_lte_val = _nemo_pick_float_resolved(row, rsrp_lte_indices, n)
        rsrq_nr_val = _nemo_pick_float_resolved(row, rsrq_nr_indices, n)
        rsrq_lte_val = _nemo_pick_float_resolved(row, rsrq_lte_indices, n)
        sinr_nr_val = _nemo_pick_float_resolved(row, sinr_nr_indices, n)
        sinr_lte_val = _nemo_pick_float_resolved(row, sinr_lte_indices, n)
        rsrp_val = rsrp_nr_val if rsrp_nr_val is not None else rsrp_lte_val
        rsrq_val = rsrq_nr_val if rsrq_nr_val is not None else rsrq_lte_val
        sinr_val = sinr_nr_val if sinr_nr_val is not None else sinr_lte_val

        payload = {
            "operator": operator,
            "fileName": file_name,
            "sourceRow": row_number,
            "_dt": event_time,
            "time": _nemo_iso(event_time),
            "measurementTitle": measurement_title,
            "lon": _nemo_pick_float_resolved(row, lon_indices, n),
            "lat": _nemo_pick_float_resolved(row, lat_indices, n),
            "servingTechnology": serving_technology,
            "packetTechnology": packet_technology,
            "rrcState": _nemo_pick_text_resolved(row, rrc_state_indices, n),
            "applicationProtocol": app_protocol,
            "transferStatus": transfer_status,
            "cellTypes": cell_types,
            "system": _nemo_pick_text_resolved(row, system_indices, n),
            "band": band_text,
            "bandwidthPrbs": _nemo_pick_float_resolved(row, bandwidth_prbs_indices, n),
            "scellsCount": _nemo_pick_float_resolved(row, scells_count_indices, n),
            "lteCaStatus": _nemo_pick_text_resolved(row, lte_ca_status_indices, n),
            "nrCaStatus": _nemo_pick_text_resolved(row, nr_ca_status_indices, n),
            "pci": _nemo_pick_float_resolved(row, pci_indices, n),
            "lteChannelNumber": _nemo_pick_float_resolved(row, lte_channel_indices, n),
            "nrChannelNumber": nr_channel_number,
            "appDlRaw": app_dl_raw,
            "appDlAvgRaw": app_dl_avg_raw,
            "downloadTimeS": _nemo_pick_float_resolved(row, download_time_indices, n),
            "bytesDl": _nemo_pick_float_resolved(row, bytes_dl_indices, n),
            "bytesUl": _nemo_pick_float_resolved(row, bytes_ul_indices, n),
            "fileSizeBytes": _nemo_pick_float_resolved(row, file_size_indices, n),
            "dataTransferDirection": _nemo_pick_text_resolved(row, transfer_direction_indices, n),
            "transferFilename": _nemo_pick_text_resolved(row, transfer_filename_indices, n),
            "macDlLteRaw": mac_dl_lte_raw,
            "macDl5gRaw": mac_dl_5g_raw,
            "totalMacDlRaw": total_mac_dl_raw,
            "pdschSched5gRaw": pdsch_sched_5g_raw,
            "pdschDl5gRaw": pdsch_dl_5g_raw,
            "pdschDlLteRaw": pdsch_dl_lte_raw,
            "pdschDlLteCw0Raw": pdsch_dl_lte_cw0_raw,
            "pdschDlLteCw1Raw": pdsch_dl_lte_cw1_raw,
            "pdschPrbs": _nemo_pick_float_resolved(row, pdsch_prbs_indices, n),
            "pdschModulationCw0": _nemo_pick_text_resolved(row, pdsch_mod_cw0_indices, n),
            "pdschModulationCw1": _nemo_pick_text_resolved(row, pdsch_mod_cw1_indices, n),
            "pdschMcsCw0": _nemo_pick_float_resolved(row, pdsch_mcs_cw0_indices, n),
            "pdschMcsCw1": _nemo_pick_float_resolved(row, pdsch_mcs_cw1_indices, n),
            "pdschTbsCw0": _nemo_pick_float_resolved(row, pdsch_tbs_cw0_indices, n),
            "pdschTbsCw1": _nemo_pick_float_resolved(row, pdsch_tbs_cw1_indices, n),
            "pdschBitsPerHz": _nemo_pick_float_resolved(row, pdsch_bits_per_hz_indices, n),
            "pdschMaxBitsPerHz": _nemo_pick_float_resolved(row, pdsch_max_bits_per_hz_indices, n),
            "pdschSlotPct": _nemo_pick_float_resolved(row, pdsch_slot_pct_indices, n),
            "rsrp": rsrp_val,
            "rsrq": rsrq_val,
            "sinr": sinr_val,
            "rsrpNr": rsrp_nr_val,
            "rsrpLte": rsrp_lte_val,
            "rsrqNr": rsrq_nr_val,
            "rsrqLte": rsrq_lte_val,
            "sinrNr": sinr_nr_val,
            "sinrLte": sinr_lte_val,
            "ri": _nemo_pick_float_resolved(row, ri_indices, n),
            "wbCqi": _nemo_pick_float_resolved(row, wb_cqi_indices, n),
            "scheduledRank": _nemo_pick_float_resolved(row, scheduled_rank_indices, n),
            "macDlBler": _nemo_pick_float_resolved(row, mac_dl_bler_indices, n),
            "macUlRetx5g": _nemo_pick_float_resolved(row, mac_ul_retx_5g_indices, n),
            "dlPrbPct": _nemo_pick_float_resolved(row, dl_prb_pct_indices, n),
            "prbsAvgDl": _nemo_pick_float_resolved(row, prbs_avg_dl_indices, n),
            "schBitratePerPrb": _nemo_pick_float_resolved(row, sch_bitrate_per_prb_indices, n),
            "pdschBlerLte": _nemo_pick_float_resolved(row, pdsch_bler_lte_indices, n),
            "macDlResidualBler": _nemo_pick_float_resolved(row, mac_dl_residual_bler_indices, n),
            "pdcchBlerEst": _nemo_pick_float_resolved(row, pdcch_bler_indices, n),
            "macUlRetxLte": _nemo_pick_float_resolved(row, mac_ul_retx_lte_indices, n),
            "caTotalBwMhz": _nemo_pick_float_resolved(row, ca_total_bw_indices, n),
            "primaryBwMhz": _nemo_pick_float_resolved(row, primary_bw_indices, n),
            "sumSecondaryBwMhz": _nemo_pick_float_resolved(row, sum_secondary_bw_indices, n),
            "txPower": _nemo_pick_float_resolved(row, tx_power_indices, n),
            "puschTxPower": _nemo_pick_float_resolved(row, pusch_tx_power_indices, n),
            "wbCqi0": _nemo_pick_float_resolved(row, wb_cqi0_indices, n),
            "wbCqi1": _nemo_pick_float_resolved(row, wb_cqi1_indices, n),
            "hoUplaneInterruptionMs": _nemo_pick_float_resolved(row, ho_uplane_interruption_indices, n),
            "pppRateDl": _nemo_pick_float_resolved(row, ppp_rate_dl_indices, n),
            "recvPppBytes": _nemo_pick_float_resolved(row, recv_ppp_bytes_indices, n),
            "tcpHandshakeMs": _nemo_pick_float_resolved(row, tcp_handshake_indices, n),
            "lostPacket": _nemo_pick_float_resolved(row, lost_packet_indices, n),
            "pingStatus": _nemo_pick_text_resolved(row, ping_status_indices, n),
            "eventId": _nemo_pick_text_resolved(row, event_id_indices, n),
            "eventText": _nemo_pick_text_resolved(row, event_text_indices, n),
        }
        for field_name, kpi_type in (
            ("lon", "longitude"),
            ("lat", "latitude"),
            ("rsrp", "rsrp"),
            ("rsrq", "rsrq"),
            ("sinr", "sinr"),
            ("macDlBler", "bler"),
            ("macUlRetx5g", "retx"),
        ):
            original_value = payload.get(field_name)
            normalized_value, changed = _nemo_normalize_value(original_value, kpi_type)
            if not changed:
                continue
            payload[field_name] = normalized_value
            parsing_qa["normalizedValueCount"] += 1
            parsing_qa["normalizedFields"][field_name] = parsing_qa["normalizedFields"].get(field_name, 0) + 1
            if len(parsing_qa["normalizedSamples"]) < 8:
                parsing_qa["normalizedSamples"].append({
                    "row": row_number,
                    "field": field_name,
                    "before": original_value,
                    "after": normalized_value,
                })
        raw_rows.append(payload)

    throughput_scales = _nemo_infer_throughput_scales(raw_rows)
    for payload in raw_rows:
        if payload.get("appDlRaw") is not None:
            payload["appDlMbps"] = round(float(payload.get("appDlRaw")) / throughput_scales["appDlMbps"], 3)
        if payload.get("appDlAvgRaw") is not None:
            payload["appDlAvgMbps"] = round(float(payload.get("appDlAvgRaw")) / throughput_scales["appDlAvgMbps"], 3)
        if payload.get("appUlRaw") is not None:
            payload["appUlMbps"] = round(float(payload.get("appUlRaw")) / throughput_scales["appUlMbps"], 3)
        if payload.get("macDlLteRaw") is not None:
            payload["macDlLteMbps"] = round(float(payload.get("macDlLteRaw")) / throughput_scales["macDlLteMbps"], 3)
        if payload.get("macDl5gRaw") is not None:
            payload["macDl5gMbps"] = round(float(payload.get("macDl5gRaw")) / throughput_scales["macDl5gMbps"], 3)
        if payload.get("totalMacDlRaw") is not None:
            payload["totalMacDlMbps"] = round(float(payload.get("totalMacDlRaw")) / throughput_scales["totalMacDlMbps"], 3)
        if payload.get("pdschSched5gRaw") is not None:
            payload["pdschSched5gMbps"] = round(float(payload.get("pdschSched5gRaw")) / throughput_scales["pdschSched5gMbps"], 3)
        if payload.get("pdschDl5gRaw") is not None:
            payload["pdschDl5gMbps"] = round(float(payload.get("pdschDl5gRaw")) / throughput_scales["pdschDl5gMbps"], 3)
        if payload.get("pdschDlLteRaw") is not None:
            payload["pdschDlLteMbps"] = round(float(payload.get("pdschDlLteRaw")) / throughput_scales["pdschDlLteMbps"], 3)
        if payload.get("pdschDlLteCw0Raw") is not None:
            payload["pdschDlLteCw0Mbps"] = round(float(payload.get("pdschDlLteCw0Raw")) / throughput_scales["pdschDlLteCw0Mbps"], 3)
        if payload.get("pdschDlLteCw1Raw") is not None:
            payload["pdschDlLteCw1Mbps"] = round(float(payload.get("pdschDlLteCw1Raw")) / throughput_scales["pdschDlLteCw1Mbps"], 3)
        payload["family"] = _nemo_classify_family(payload)
        normalized_rows.append(payload)

    for payload in normalized_rows:
        event_time = payload.get("_dt")
        measurement_title = payload.get("measurementTitle")
        transfer_status = payload.get("transferStatus")
        nr_channel_number = payload.get("nrChannelNumber")
        serving_technology = str(payload.get("servingTechnology") or "")
        packet_technology = str(payload.get("packetTechnology") or "")
        cell_types = payload.get("cellTypes") or []
        band_text = payload.get("band")
        serving_technology_upper = serving_technology.upper()
        packet_technology_upper = packet_technology.upper()
        cell_types_upper = {str(cell or "").strip().upper() for cell in cell_types}
        coverage["rowCount"] += 1
        if payload.get("appDlMbps") is not None:
            coverage["nonEmptyAppDl"] += 1
        if transfer_status:
            coverage["nonEmptyTransferStatus"] += 1
        if payload.get("macDl5gMbps") is not None:
            coverage["nonEmptyMac5g"] += 1
        if payload.get("pdschDl5gMbps") is not None:
            coverage["nonEmptyPdsch5g"] += 1
        if nr_channel_number is not None:
            coverage["nonEmptyNrChannel"] += 1
            nr_channel_non_null_count += 1
        if payload.get("rsrp") is not None:
            coverage["nonEmptyRsrp"] += 1
        if payload.get("sinr") is not None:
            coverage["nonEmptySinr"] += 1
        if payload.get("ri") is not None:
            coverage["nonEmptyRi"] += 1
        if payload.get("macDlBler") is not None:
            coverage["nonEmptyBler"] += 1
        if payload.get("macDl5gMbps") not in (None, 0):
            if float(payload.get("macDl5gMbps") or 0) > 0:
                mac_dl_5g_positive_count += 1
        if payload.get("pdschDl5gMbps") not in (None, 0):
            if float(payload.get("pdschDl5gMbps") or 0) > 0:
                pdsch_dl_5g_positive_count += 1
        if "SCG PSCELL" in cell_types_upper:
            scg_pscell_samples += 1
        row_has_nr = (
            nr_channel_number is not None
            or "EN-DC" in serving_technology_upper
            or "5G" in serving_technology_upper
            or "EN-DC" in packet_technology_upper
            or float(payload.get("macDl5gMbps") or 0) > 0
            or float(payload.get("pdschDl5gMbps") or 0) > 0
            or bool(cell_types_upper & _NR_CT_UPPER)
        )
        if row_has_nr:
            has_5g = True
        if band_text and row_has_nr and _nemo_is_valid_band(band_text):
            nr_bands.add(str(band_text).strip())
        if isinstance(event_time, _dt):
            second_bucket = event_time.replace(microsecond=0)
            nr_presence_by_second[second_bucket] = nr_presence_by_second.get(second_bucket, False) or row_has_nr
            serving_tech_timeline.append((event_time, serving_technology))
            packet_tech_timeline.append((event_time, packet_technology))
        if measurement_title:
            bucket = rows_by_measurement_title.get(measurement_title)
            if bucket is None:
                rows_by_measurement_title[measurement_title] = [payload]
                ordered_dt_titles.append(measurement_title)
            else:
                bucket.append(payload)

    nr_presence_seconds = sum(1 for present in nr_presence_by_second.values() if present)
    lte_only_seconds = sum(1 for present in nr_presence_by_second.values() if not present)
    total_presence_seconds = len(nr_presence_by_second)
    nr_presence_pct = round(nr_presence_seconds / float(total_presence_seconds) * 100.0, 1) if total_presence_seconds else None
    lte_only_presence_pct = round(lte_only_seconds / float(total_presence_seconds) * 100.0, 1) if total_presence_seconds else None
    technology_status = {
        "operator": operator,
        "has5g": has_5g,
        "fiveGStatus": "5G/EN-DC detected" if has_5g else "No 5G detected in export",
        "scgPscellSamples": scg_pscell_samples,
        "nrChannelNonNullCount": nr_channel_non_null_count,
        "macDl5gPositiveSamples": mac_dl_5g_positive_count,
        "pdschDl5gPositiveSamples": pdsch_dl_5g_positive_count,
        "fiveGThroughputSamples": mac_dl_5g_positive_count + pdsch_dl_5g_positive_count,
        "nrPresencePct": nr_presence_pct,
        "lteOnlyPresencePct": lte_only_presence_pct,
        "nrPresenceSeconds": nr_presence_seconds,
        "lteOnlySeconds": lte_only_seconds,
        "totalPresenceSeconds": total_presence_seconds,
        "servingTechnologyDistribution": _nemo_distribution_from_counts(_nemo_forward_filled_timeshare(serving_tech_timeline)),
        "packetTechnologyDistribution": _nemo_distribution_from_counts(_nemo_forward_filled_timeshare(packet_tech_timeline)),
        "nrBands": sorted(nr_bands),
        "comment": (
            "5G/EN-DC detected. Operator is eligible for 5G diagnosis."
            if has_5g
            else "No 5G NR/EN-DC detected in export. Operator remains in DL ranking but 5G-specific KPIs are N/A."
        ),
    }
    device_model = None
    if device_models_seen:
        device_model = max(set(device_models_seen), key=device_models_seen.count)
        technology_status["deviceModel"] = device_model

    return {
        "operator": operator,
        "path": path,
        "fileName": file_name,
        "delimiter": delimiter,
        "measurementTitles": sorted(measurement_titles),
        "duplicateHeaders": _nemo_duplicate_headers(headers),
        "throughputScales": throughput_scales,
        "has5g": has_5g,
        "fiveGStatus": "5G/EN-DC detected" if has_5g else "No 5G detected in export",
        "rows": normalized_rows,
        "orderedDtTitles": ordered_dt_titles,
        "rowsByMeasurementTitle": rows_by_measurement_title,
        "coverage": coverage,
        "technologyStatus": technology_status,
        "deviceModel": device_model,
        "parsingQa": parsing_qa,
        # Authoritative per-session statistics from the sibling "Data transfer session
        # statistics" export (ping/upload/download), when present alongside this file.
        "sessionStats": _nemo_parse_session_stats(_nemo_find_session_stats_path(path)),
    }


def _nemo_preferred_dl_value(row: dict):
    for key in ("appDlMbps", "appDlAvgMbps", "totalMacDlMbps", "macDl5gMbps", "macDlLteMbps"):
        value = row.get(key)
        if value is not None and float(value) > 0:
            return float(value)
    return None


def _nemo_select_dl_metric_key(rows: list[dict]) -> str:
    for key in ("appDlMbps", "appDlAvgMbps", "totalMacDlMbps", "macDl5gMbps", "macDlLteMbps", "pdschDlLteMbps"):
        count = 0
        for row in rows or []:
            protocol = str(row.get("applicationProtocol") or "").lower()
            if "ping" in protocol:
                continue
            value = row.get(key)
            if value is not None and float(value) > 0:
                count += 1
        if count > 0:
            return key
    return ""


def _nemo_select_benchmark_dl_metric_key(rows: list[dict]) -> str:
    """Select the DL metric best suited for benchmark ranking / test averages.

    Prefer Nemo's explicit per-download average row when available (`App rate DL avg`),
    because some exports duplicate instantaneous `App. rate DL` samples at the same timestamp.
    If no per-download average exists, fall back to the denser instantaneous/application/radio
    metrics so older datasets continue to work.
    """
    for key in ("appDlAvgMbps", "appDlMbps", "totalMacDlMbps", "macDl5gMbps", "macDlLteMbps", "pdschDlLteMbps"):
        count = 0
        for row in rows or []:
            protocol = str(row.get("applicationProtocol") or "").lower()
            if "ping" in protocol:
                continue
            value = row.get(key)
            if value is not None and float(value) > 0:
                count += 1
        if count > 0:
            return key
    return ""


def _nemo_metric_series(rows: list[dict], key: str) -> list[float]:
    if not key:
        return []
    values = []
    seen = set()
    for row in rows or []:
        protocol = str(row.get("applicationProtocol") or "").lower()
        if "ping" in protocol:
            continue
        value = row.get(key)
        if value is None or float(value) <= 0:
            continue
        signature = (
            row.get("time") or "",
            round(float(value), 6),
            key,
        )
        if signature in seen:
            continue
        seen.add(signature)
        values.append(float(value))
    return values


_TRANSFER_SIZE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*([KkMmGg])\b")


def _nemo_target_size_from_url(url: str):
    """Derive a transfer target size in bytes from a Nemo HTTP/FTP filename URL.

    The benchmark URLs embed the size in the path (e.g. ``…/50M/50M.zip`` → 50 MB,
    ``…/10M/Handy….html`` → 10 MB). Decimal multipliers (1 K = 1000) match testdebit.info /
    the values the user confirmed (DL = 50 M, UL = 10 M). Returns None when no size is present.
    """
    if not url:
        return None
    match = _TRANSFER_SIZE_RE.search(str(url))
    if not match:
        return None
    try:
        value = float(match.group(1))
    except Exception:
        return None
    unit = match.group(2).upper()
    factor = {"K": 1_000, "M": 1_000_000, "G": 1_000_000_000}.get(unit)
    if not factor:
        return None
    return value * factor


def _nemo_collect_transfer_target_hints(rows: list[dict]) -> dict:
    """Infer canonical transfer target sizes per (direction, URL) from the dataset itself.

    Nemo benchmark exports often expose cumulative transferred bytes that do not match the
    nominal size token embedded in the URL (`50M.zip` can culminate at `200000000` bytes).
    When a successful row or an explicit `File size` value exists, prefer that observed target
    size and reuse it for other sessions of the same benchmark file.
    """
    hints: dict = {}
    ordered = [row for row in (rows or []) if row.get("_dt") is not None]
    ordered.sort(key=lambda item: item.get("_dt"))

    active_direction = None
    active_url = ""
    for row in ordered:
        direction_raw = str(row.get("dataTransferDirection") or "").strip()
        if direction_raw:
            active_direction = "uplink" if direction_raw.lower().startswith("up") else "downlink"
        url = str(row.get("transferFilename") or "").strip()
        if url:
            active_url = url
        if not active_direction or not active_url:
            continue

        key = (active_direction, active_url)
        candidates = []

        file_size = row.get("fileSizeBytes")
        if file_size is not None:
            try:
                file_size_val = float(file_size)
                if math.isfinite(file_size_val) and file_size_val > 0:
                    candidates.append(file_size_val)
            except Exception:
                pass

        status = str(row.get("transferStatus") or "").strip().lower()
        if "success" in status:
            byte_key = "bytesUl" if active_direction == "uplink" else "bytesDl"
            transferred = row.get(byte_key)
            if transferred is not None:
                try:
                    transferred_val = float(transferred)
                    if math.isfinite(transferred_val) and transferred_val > 0:
                        candidates.append(transferred_val)
                except Exception:
                    pass

        if not candidates:
            continue
        hints[key] = max([hints.get(key, 0.0)] + candidates)
    return hints


def _nemo_build_transfer_sessions(rows: list[dict], operator: str) -> list[dict]:
    """Walk the (time-ordered) rows and reconstruct each data-transfer session.

    A session opens on every row that carries a non-empty ``dataTransferDirection``
    (``Downlink`` / ``Uplink``) and runs until the next such marker. Per session we compute the
    completion % (transferred bytes ÷ target file size, capped at 100) and the success flag, and
    we attach the forward-filled Measurement Title (the DT / "BJ" name).
    """
    ordered = [row for row in (rows or []) if row.get("_dt") is not None]
    ordered.sort(key=lambda item: item.get("_dt"))
    target_hints = _nemo_collect_transfer_target_hints(ordered)

    # Indices of the session-start markers.
    marker_indices = [
        i for i, row in enumerate(ordered)
        if str(row.get("dataTransferDirection") or "").strip()
    ]
    sessions: list[dict] = []
    for position, start_idx in enumerate(marker_indices):
        end_idx = marker_indices[position + 1] if position + 1 < len(marker_indices) else len(ordered)
        window = ordered[start_idx:end_idx]
        marker = ordered[start_idx]
        direction_raw = str(marker.get("dataTransferDirection") or "").strip()
        direction = direction_raw.lower()
        is_uplink = direction.startswith("up")

        # Target file URL + size (prefer URL-embedded size, fall back to the File size column).
        file_url = ""
        for row in window:
            url = str(row.get("transferFilename") or "").strip()
            if url:
                file_url = url
                break
        file_sizes = [row.get("fileSizeBytes") for row in window if row.get("fileSizeBytes")]
        target_bytes = None
        if file_sizes:
            try:
                target_bytes = max(float(v) for v in file_sizes)
            except Exception:
                target_bytes = None
        if target_bytes is None and file_url:
            target_bytes = target_hints.get(("uplink" if is_uplink else "downlink", file_url))
        if target_bytes is None:
            target_bytes = _nemo_target_size_from_url(file_url)

        # Transferred bytes — max seen in the direction-appropriate column.
        byte_key = "bytesUl" if is_uplink else "bytesDl"
        transferred_values = [
            float(row.get(byte_key)) for row in window
            if row.get(byte_key) is not None and math.isfinite(float(row.get(byte_key)))
        ]
        transferred_bytes = max(transferred_values) if transferred_values else None

        # Keep only real file transfers (HTTP/FTP). Markers with no target file URL and no
        # transferred bytes are ICMP-ping / idle session boundaries, not file downloads.
        if not file_url and transferred_bytes is None:
            continue

        completion_pct = None
        if target_bytes and target_bytes > 0 and transferred_bytes is not None:
            completion_pct = round(min(100.0, transferred_bytes / float(target_bytes) * 100.0), 1)

        status_rows = [str(row.get("transferStatus") or "").strip() for row in window if str(row.get("transferStatus") or "").strip()]
        success = any("success" in status.lower() for status in status_rows)
        if status_rows:
            status_label = next((s for s in status_rows if "success" in s.lower()), status_rows[-1])
        else:
            status_label = "Unknown"

        lat = next((row.get("lat") for row in window if row.get("lat") is not None), None)
        lon = next((row.get("lon") for row in window if row.get("lon") is not None), None)
        app_key = "appUlMbps" if is_uplink else "appDlMbps"
        marker_dt = marker.get("_dt")
        last_activity_dt = marker_dt
        first_positive_app_dt = None
        last_positive_app_dt = None
        first_byte_progress_dt = None
        last_byte_progress_dt = None
        prev_bytes = None
        for row in window:
            row_dt = row.get("_dt")
            if row_dt is None:
                continue
            status_text = str(row.get("transferStatus") or "").strip()
            has_status = bool(status_text)
            has_positive_app = False
            try:
                app_val = row.get(app_key)
                has_positive_app = app_val is not None and float(app_val) > 0
            except Exception:
                has_positive_app = False
            current_bytes = None
            has_byte_progress = False
            try:
                raw_bytes = row.get(byte_key)
                if raw_bytes is not None and math.isfinite(float(raw_bytes)):
                    current_bytes = float(raw_bytes)
                    has_byte_progress = (
                        current_bytes > 0
                        if prev_bytes is None
                        else abs(current_bytes - prev_bytes) > 1e-6
                    )
            except Exception:
                current_bytes = None
                has_byte_progress = False
            if has_positive_app:
                if first_positive_app_dt is None:
                    first_positive_app_dt = row_dt
                last_positive_app_dt = row_dt
            if has_byte_progress:
                if first_byte_progress_dt is None:
                    first_byte_progress_dt = row_dt
                last_byte_progress_dt = row_dt
            if has_status or has_positive_app or has_byte_progress:
                last_activity_dt = row_dt
            if current_bytes is not None:
                prev_bytes = current_bytes

        meaningful_rows = [
            row for row in window
            if (
                str(row.get("transferFilename") or "").strip()
                or row.get(byte_key) is not None
                or row.get("fileSizeBytes") is not None
                or str(row.get("transferStatus") or "").strip()
            )
        ]
        start_dt = first_positive_app_dt or first_byte_progress_dt or marker_dt
        end_dt = (
            last_positive_app_dt
            or last_byte_progress_dt
            or last_activity_dt
            or (meaningful_rows[-1].get("_dt") if meaningful_rows else marker_dt)
        )

        sessions.append({
            "id": f"{operator}_X{len(sessions) + 1:03d}",
            "operator": operator,
            "measurementTitle": _benchmark_text(marker.get("measurementTitle")),
            "direction": "Uplink" if is_uplink else "Downlink",
            "fileUrl": file_url,
            "targetBytes": target_bytes,
            "transferredBytes": transferred_bytes,
            "completionPct": completion_pct,
            "success": success,
            "statusLabel": status_label,
            "startTime": _nemo_iso(start_dt),
            "endTime": _nemo_iso(end_dt),
            "markerStartTime": _nemo_iso(marker_dt),
            "lat": _nemo_safe_round(lat, 6),
            "lon": _nemo_safe_round(lon, 6),
        })
    return sessions


def _nemo_build_transfer_summary(sessions: list[dict]) -> list[dict]:
    """Aggregate transfer sessions per (operator, direction): success rate + avg completion %."""
    groups: dict = {}
    for session in sessions or []:
        key = (session.get("operator") or "UNKNOWN", session.get("direction") or "Unknown")
        groups.setdefault(key, []).append(session)

    summary = []
    for (operator, direction), items in groups.items():
        total = len(items)
        successes = sum(1 for s in items if s.get("success"))
        completion_values = [s.get("completionPct") for s in items if s.get("completionPct") is not None]
        summary.append({
            "operator": operator,
            "direction": direction,
            "transferCount": total,
            "successCount": successes,
            "successRate": round(successes / float(total) * 100.0, 1) if total else None,
            "avgCompletionPct": round(sum(completion_values) / float(len(completion_values)), 1) if completion_values else None,
        })
    summary.sort(key=lambda entry: (str(entry.get("operator") or ""), str(entry.get("direction") or "")))
    return summary


def _nemo_align_benchmark_tests_with_transfer_sessions(operator_file: dict) -> dict:
    """Use reconstructed downlink transfer sessions as the canonical DT sequence.

    Some Nemo exports leave `Measurement Title` blank on throughput rows and only expose the
    DT name on the transfer-session markers. The benchmark already reconstructs those sessions,
    so reuse them to relabel / reorder the per-DT test list instead of trusting the forward-
    filled row titles.
    """
    if not isinstance(operator_file, dict):
        return operator_file

    tests = operator_file.get("tests") or []
    downlink_sessions = [
        session for session in (operator_file.get("transferSessions") or [])
        if str(session.get("direction") or "").lower() == "downlink"
    ]
    if not downlink_sessions:
        return operator_file

    ordered_titles = [
        _benchmark_text(session.get("measurementTitle"))
        for session in downlink_sessions
        if _benchmark_text(session.get("measurementTitle"))
    ]
    if ordered_titles:
        operator_file["orderedDtTitles"] = ordered_titles

    if len(tests) != len(downlink_sessions):
        return operator_file

    for test, session in zip(tests, downlink_sessions):
        session_title = _benchmark_text(session.get("measurementTitle"))
        if session_title:
            test["measurementTitle"] = session_title
    return operator_file


def _nemo_build_tests(rows: list[dict], operator: str, benchmark_dl_metric_key: str = "") -> list[dict]:
    dl_metric_key = benchmark_dl_metric_key or _nemo_select_benchmark_dl_metric_key(rows)
    ordered_rows = [row for row in (rows or []) if row.get("_dt")]
    ordered_rows.sort(key=lambda item: item.get("_dt"))
    app_rows = []
    seen = set()
    for row in ordered_rows:
        protocol = str(row.get("applicationProtocol") or "").lower()
        if "ping" in protocol:
            continue
        dl_value = row.get(dl_metric_key) if dl_metric_key else _nemo_preferred_dl_value(row)
        if dl_value is None or dl_value <= 0:
            continue
        signature = (row.get("time") or "", round(float(dl_value), 6), dl_metric_key)
        if signature in seen:
            continue
        seen.add(signature)
        app_rows.append((row, dl_value))
    if not app_rows:
        return []

    clusters = []
    current = [app_rows[0]]
    for item in app_rows[1:]:
        prev_row = current[-1][0]
        row = item[0]
        gap_s = abs((row["_dt"] - prev_row["_dt"]).total_seconds())
        same_measurement = _benchmark_text(row.get("measurementTitle")) == _benchmark_text(prev_row.get("measurementTitle"))
        if same_measurement and gap_s <= 8.0:
            current.append(item)
        else:
            clusters.append(current)
            current = [item]
    if current:
        clusters.append(current)

    tests = []
    ordered_start = 0
    ordered_end = 0
    for index, cluster in enumerate(clusters, start=1):
        cluster_rows = [item[0] for item in cluster]
        dl_values = [float(item[1]) for item in cluster if item[1] is not None]
        start_ts = cluster_rows[0]["_dt"]
        end_ts = cluster_rows[-1]["_dt"]
        window_start = start_ts
        window_end = end_ts
        if cluster_rows[-1].get("downloadTimeS") is not None:
            try:
                from datetime import timedelta
                window_end = max(window_end, end_ts + timedelta(seconds=float(cluster_rows[-1].get("downloadTimeS") or 0.0)))
            except Exception:
                pass
        while ordered_start < len(ordered_rows) and ordered_rows[ordered_start]["_dt"] < window_start:
            ordered_start += 1
        if ordered_end < ordered_start:
            ordered_end = ordered_start
        while ordered_end < len(ordered_rows) and ordered_rows[ordered_end]["_dt"] <= window_end:
            ordered_end += 1
        associated_rows = ordered_rows[ordered_start:ordered_end]

        status_rows = []
        success = False
        has5g = False
        scheduled5g_values = []
        pdsch5g_values = []
        prbs_values = []
        rsrp_values = []
        rsrq_values = []
        sinr_values = []
        cqi_values = []
        ri_values = []
        bler_values = []
        total_mac_values = []
        band_rows = []
        for assoc_row in associated_rows:
            transfer_status = assoc_row.get("transferStatus")
            if transfer_status:
                transfer_status_text = str(transfer_status)
                status_rows.append(transfer_status_text)
                if not success and "success" in transfer_status_text.lower():
                    success = True
            if (
                (assoc_row.get("macDl5gMbps") or 0) > 0
                or (assoc_row.get("pdschDl5gMbps") or 0) > 0
                or assoc_row.get("nrChannelNumber") is not None
            ):
                has5g = True
            band_text = str(assoc_row.get("band") or "").strip()
            if band_text:
                band_rows.append(band_text)
            for key, target, positive in (
                ("pdschSched5gMbps", scheduled5g_values, True),
                ("pdschDl5gMbps", pdsch5g_values, True),
                ("pdschPrbs", prbs_values, True),
                ("rsrp", rsrp_values, False),
                ("rsrq", rsrq_values, False),
                ("sinr", sinr_values, False),
                ("wbCqi", cqi_values, True),
                ("ri", ri_values, True),
                ("macDlBler", bler_values, False),
                ("totalMacDlMbps", total_mac_values, True),
            ):
                value = assoc_row.get(key)
                if value is None:
                    continue
                value_num = float(value)
                if positive and value_num <= 0:
                    continue
                target.append(value_num)
        lat_values = [float(row.get("lat")) for row in cluster_rows if row.get("lat") is not None]
        lon_values = [float(row.get("lon")) for row in cluster_rows if row.get("lon") is not None]
        dl_stats = _nemo_metric_stats(dl_values)
        test = {
            "id": f"{operator}_T{index:02d}",
            "operator": operator,
            "measurementTitle": _benchmark_text(cluster_rows[0].get("measurementTitle")),
            "anchorTime": _nemo_iso(start_ts),
            "startTime": _nemo_iso(window_start),
            "endTime": _nemo_iso(window_end),
            "anchorLat": round(sum(lat_values) / float(len(lat_values)), 6) if lat_values else None,
            "anchorLon": round(sum(lon_values) / float(len(lon_values)), 6) if lon_values else None,
            "applicationProtocol": _benchmark_text(cluster_rows[0].get("applicationProtocol")),
            "dlStats": dl_stats,
            "avgDlMbps": dl_stats.get("average"),
            "success": success,
            "statusLabel": "Success" if success else (_benchmark_text(status_rows[0]) if status_rows else "Unknown"),
            "has5g": has5g,
            "scheduled5gStats": _nemo_metric_stats(scheduled5g_values),
            "pdsch5gStats": _nemo_metric_stats(pdsch5g_values),
            "prbsStats": _nemo_metric_stats(prbs_values),
            "rsrpStats": _nemo_metric_stats(rsrp_values),
            "rsrqStats": _nemo_metric_stats(rsrq_values),
            "sinrStats": _nemo_metric_stats(sinr_values),
            "cqiStats": _nemo_metric_stats(cqi_values),
            "riStats": _nemo_metric_stats(ri_values),
            "riGe3Share": round((sum(1 for value in ri_values if value >= 3) / float(len(ri_values))) * 100.0, 1) if ri_values else None,
            "ri1Share": round((sum(1 for value in ri_values if value <= 1) / float(len(ri_values))) * 100.0, 1) if ri_values else None,
            "blerStats": _nemo_metric_stats(bler_values),
            "totalMacStats": _nemo_metric_stats(total_mac_values),
            "bandSummary": sorted(set(band_rows)),
            "sampleCount": len(associated_rows),
        }
        tests.append(test)
    return tests


def _nemo_endc_secondary_node_stats(rows: list[dict], sustain_sec: float = 3.0) -> dict:
    """Reconstruct EN-DC NR secondary-node (SgNB) addition/removal events from the NR SCG
    PSCell presence timeline. These Nemo exports carry no explicit SgNB RRC signalling
    (SgNB Addition Request/Complete/Failure, SgNB Release), so events are INFERRED from the
    NR-active (EN-DC / SCG PSCell) presence transitions:
      • addition = NR-inactive → NR-active second
      • removal  = NR-active → NR-inactive second
      • addition SUCCESS = NR stays active ≥ sustain_sec (stable EN-DC leg); an addition that
        reverts within sustain_sec is flagged a FAILURE (likely an aborted/failed SgNB add)
      • removal SUCCESS = NR stays inactive ≥ sustain_sec; a removal that re-activates almost
        immediately is flagged a FAILURE (unstable release / ping-pong)
    A proxy, not RRC ground truth — labelled as reconstructed downstream.
    """
    def _row_nr_active(r):
        cts = [str(c or "").strip().lower() for c in (r.get("cellTypes") or [])]
        return (
            any("scg pscell" in c for c in cts)
            or r.get("nrChannelNumber") is not None
            or "en-dc" in str(r.get("servingTechnology") or "").lower()
            or "en-dc" in str(r.get("packetTechnology") or "").lower()
        )

    sec_state: dict = {}
    for r in rows or []:
        dt = r.get("_dt")
        if dt is None:
            continue
        s = dt.replace(microsecond=0)
        if _row_nr_active(r):
            sec_state[s] = True
        else:
            sec_state.setdefault(s, False)
    secs = sorted(sec_state)
    if len(secs) < 2:
        return {"available": False}

    additions = removals = add_success = add_fail = removal_success = removal_fail = 0
    prev = sec_state[secs[0]]
    for i in range(1, len(secs)):
        cur = sec_state[secs[i]]
        if cur and not prev:
            additions += 1
            t0 = secs[i]
            sustained = True
            for j in range(i, len(secs)):
                if (secs[j] - t0).total_seconds() > sustain_sec:
                    break
                if not sec_state[secs[j]]:
                    sustained = False
                    break
            if sustained:
                add_success += 1
            else:
                add_fail += 1
        if (not cur) and prev:
            removals += 1
            t0 = secs[i]
            re_active = False
            for j in range(i, len(secs)):
                if (secs[j] - t0).total_seconds() > sustain_sec:
                    break
                if sec_state[secs[j]]:
                    re_active = True
                    break
            if re_active:
                removal_fail += 1
            else:
                removal_success += 1
        prev = cur

    return {
        "available": True,
        "reconstructed": True,
        "additions": additions,
        "additionSuccess": add_success,
        "additionFailure": add_fail,
        "additionSuccessRate": round(add_success / additions * 100.0, 1) if additions else None,
        "removals": removals,
        "removalSuccess": removal_success,
        "removalFailure": removal_fail,
        "removalSuccessRate": round(removal_success / removals * 100.0, 1) if removals else None,
        "nrActiveSeconds": sum(1 for s in secs if sec_state[s]),
        "totalSeconds": len(secs),
    }


def _nemo_operator_kpis(operator_data: dict) -> dict:
    rows = operator_data.get("rows") or []
    tests = operator_data.get("tests") or []
    dl_metric_key = operator_data.get("benchmarkDlMetricKey") or operator_data.get("dlMetricKey") or _nemo_select_benchmark_dl_metric_key(rows)
    dl_values = _nemo_metric_series(rows, dl_metric_key)
    app_stats = _nemo_metric_stats(dl_values)

    # Benchmark DL throughput is a per-drive-test KPI: each download session must
    # count once. Reuse the already-built benchmark tests instead of regrouping raw
    # rows by `measurementTitle`, because some Nemo exports leave transfer rows blank
    # and forward-filling then merges multiple DTs under the same title.
    _dt_dl_avgs = [
        float(test.get("avgDlMbps")) for test in tests
        if test.get("avgDlMbps") is not None and float(test.get("avgDlMbps")) > 0
    ]
    if _dt_dl_avgs:
        app_stats = dict(app_stats)
        app_stats["pooledAverage"] = app_stats.get("average")
        app_stats["average"] = round(sum(_dt_dl_avgs) / float(len(_dt_dl_avgs)), 2)
        app_stats["perDtCount"] = len(_dt_dl_avgs)

    metric = lambda key, positive=False: [
        float(row.get(key)) for row in rows
        if row.get(key) is not None and (not positive or float(row.get(key)) > 0)
    ]
    app_dl_all_values = [
        float(row.get("appDlMbps")) for row in rows
        if row.get("appDlMbps") is not None and math.isfinite(float(row.get("appDlMbps")))
    ]
    app_dl_sample_stats = _nemo_metric_stats(app_dl_all_values)
    ri_values = metric("ri", positive=True)
    band_values = [str(row.get("band") or "").strip() for row in rows if str(row.get("band") or "").strip()]
    n78_share = None
    if band_values:
        n78_share = round((sum(1 for value in band_values if value.lower() == "n78") / float(len(band_values))) * 100.0, 1)
    nr_band_values = [str(row.get("band") or "").strip() for row in rows if str(row.get("band") or "").strip() and _nemo_band_row_filter(row)]
    n78_share_nr_only = round((sum(1 for v in nr_band_values if v.lower() == "n78") / float(len(nr_band_values))) * 100.0, 1) if nr_band_values else None
    available_prbs_values = metric("bandwidthPrbs", positive=True)
    scells_values = metric("scellsCount", positive=True)
    ca_active_share = round((sum(1 for value in scells_values if value > 0) / float(len(scells_values))) * 100.0, 1) if scells_values else None
    # #SCells stats over ALL reported rows (including 0-SCell rows) — the CA-depth /
    # activation-share view the benchmark uses (e.g. Avg 0.39, SCells>0 share 23.7%),
    # NOT the active-only average that `scells_values` (positive filter) yields.
    scells_all_values = metric("scellsCount")
    scells_avg_all = round(sum(scells_all_values) / float(len(scells_all_values)), 2) if scells_all_values else None
    scells_max = round(max(scells_all_values), 1) if scells_all_values else None
    scells_active_share = round((sum(1 for value in scells_all_values if value > 0) / float(len(scells_all_values))) * 100.0, 1) if scells_all_values else None
    # NR per-band shares over NR-only rows (n78 = capacity layer, n28 = low-band coverage).
    nr_band_shares = {}
    if nr_band_values:
        _nr_total = float(len(nr_band_values))
        _nr_counts = {}
        for _b in nr_band_values:
            _nr_counts[_b.lower()] = _nr_counts.get(_b.lower(), 0) + 1
        nr_band_shares = {band: round((count / _nr_total) * 100.0, 1) for band, count in _nr_counts.items()}
    # Share of MAC DL BLER samples above 10% / 20% (retransmission-affected route portion).
    bler_all_values = metric("macDlBler")
    bler_above_10_share = (
        round((sum(1 for v in bler_all_values if v > 10.0) / float(len(bler_all_values))) * 100.0, 1)
        if bler_all_values else None
    )
    bler_above_20_share = (
        round((sum(1 for v in bler_all_values if v > 20.0) / float(len(bler_all_values))) * 100.0, 1)
        if bler_all_values else None
    )
    # PRB utilization % = allocated PDSCH PRBs / available bandwidth PRBs, per active row.
    prb_util_values = []
    for _row in rows:
        _p = _row.get("pdschPrbs")
        _bw = _row.get("bandwidthPrbs")
        if _p is None or _bw in (None, 0):
            continue
        try:
            _pf = float(_p)
            _bwf = float(_bw)
        except Exception:
            continue
        if _bwf > 0 and _pf >= 0:
            prb_util_values.append(min(_pf / _bwf * 100.0, 100.0))
    prb_util_pct = round(sum(prb_util_values) / float(len(prb_util_values)), 1) if prb_util_values else None
    # Per-codeword MCS stats (computed from all rows; modulation distributions now come from the PDSCH-active loop below).
    mcs_cw0_stats = _nemo_metric_stats(metric("pdschMcsCw0"))
    mcs_cw1_stats = _nemo_metric_stats(metric("pdschMcsCw1"))
    # Categorical shares from already-parsed text fields (RRC / app protocol / serving + packet tech).
    # Serving technology and Packet technology are sparse change-event columns (logged only on
    # transitions, ≈0.3–0.9% of rows). Raw non-empty counts badly misrepresent time-based
    # shares (e.g. LTE anchor rows swamp NR PSCell rows in EN-DC). Forward-fill to second
    # granularity — same logic used by _nemo_technology_status_from_rows and
    # servingTechnologyDistribution in the per-cell timeline.
    rrc_state_shares = _nemo_distribution([str(r.get("rrcState") or "").strip() for r in rows if str(r.get("rrcState") or "").strip()])
    app_protocol_shares = _nemo_distribution([str(r.get("applicationProtocol") or "").strip() for r in rows if str(r.get("applicationProtocol") or "").strip()])
    _serving_tech_timeline = [(r["_dt"], r.get("servingTechnology") or "") for r in rows if r.get("_dt") is not None]
    _packet_tech_timeline = [(r["_dt"], r.get("packetTechnology") or "") for r in rows if r.get("_dt") is not None]
    serving_tech_shares = _nemo_distribution_from_counts(_nemo_forward_filled_timeshare(_serving_tech_timeline))
    packet_tech_shares = _nemo_distribution_from_counts(_nemo_forward_filled_timeshare(_packet_tech_timeline))
    total_mac_stats = _nemo_metric_stats(metric("totalMacDlMbps", positive=True))
    app_avg = app_stats.get("average")
    total_mac_avg = total_mac_stats.get("average")
    transport_ratio = round((app_avg / total_mac_avg), 2) if app_avg is not None and total_mac_avg not in (None, 0) else None

    # Throughput by RAT. The export carries explicit per-RAT MAC/PDSCH columns (LTE vs 5G),
    # so the split is intrinsic — no band classification needed. Contribution % = each RAT's
    # mean MAC DL throughput over their summed means (a layer-balance view of where the bytes
    # are carried), independent of the RAT-agnostic end-to-end app rate.
    mac_lte_stats = _nemo_metric_stats(metric("macDlLteMbps", positive=True))
    mac_5g_avg = _nemo_metric_stats(metric("macDl5gMbps", positive=True)).get("average")
    mac_lte_avg = mac_lte_stats.get("average")
    _mac_sum = sum(v for v in (mac_lte_avg, mac_5g_avg) if v is not None)
    nr_throughput_contrib_pct = round(mac_5g_avg / _mac_sum * 100.0, 1) if mac_5g_avg is not None and _mac_sum else None
    lte_throughput_contrib_pct = round(mac_lte_avg / _mac_sum * 100.0, 1) if mac_lte_avg is not None and _mac_sum else None

    prbs_stats = _nemo_metric_stats(metric("pdschPrbs", positive=True))
    pdsch_slot_stats = _nemo_metric_stats(metric("pdschSlotPct", positive=True))
    pdsch5g_stats = _nemo_metric_stats(metric("pdschDl5gMbps", positive=True))
    sched5g_stats = _nemo_metric_stats(metric("pdschSched5gMbps", positive=True))
    avail_bw_stats = _nemo_metric_stats(available_prbs_values)
    prbs_avg = prbs_stats.get("average")
    pdsch_slot_avg = pdsch_slot_stats.get("average")
    pdsch5g_avg = pdsch5g_stats.get("average")
    sched5g_avg = sched5g_stats.get("average")
    avail_bw_avg = avail_bw_stats.get("average")
    prb_efficiency = round(pdsch5g_avg / prbs_avg, 3) if prbs_avg not in (None, 0) and pdsch5g_avg is not None else None
    scheduled_efficiency = round((pdsch5g_avg or 0) / sched5g_avg * 100, 1) if sched5g_avg not in (None, 0) and pdsch5g_avg is not None else None
    resource_allocation_index = round(prbs_avg / avail_bw_avg * 100, 1) if avail_bw_avg not in (None, 0) and prbs_avg is not None else None
    prbs_per_scheduled_slot = round(prbs_avg / (pdsch_slot_avg / 100.0), 1) if prbs_avg not in (None, 0) and pdsch_slot_avg not in (None, 0) else None
    scheduled_mbps_per_slot = round(sched5g_avg / (pdsch_slot_avg / 100.0), 1) if sched5g_avg not in (None, 0) and pdsch_slot_avg not in (None, 0) else None

    pdsch_rows = [row for row in rows if _nemo_pdsch_active_row(row)]
    mcs_values = []
    tbs_values = []
    pdsch_dl_lte_total_sum = 0.0
    pdsch_dl_lte_cw1_sum = 0.0
    pdsch_dl_lte_rows = 0
    for row in pdsch_rows:
        for key in ("pdschMcsCw0", "pdschMcsCw1"):
            val = row.get(key)
            if val is not None:
                try:
                    mcs_values.append(float(val))
                except Exception:
                    pass
        for key in ("pdschTbsCw0", "pdschTbsCw1"):
            val = row.get(key)
            if val is not None:
                try:
                    fval = float(val)
                    if fval > 0:
                        tbs_values.append(fval)
                except Exception:
                    pass
        # PDSCH LTE per-codeword throughput — rank-2 utilization
        lte_total = row.get("pdschDlLteMbps")
        lte_cw1 = row.get("pdschDlLteCw1Mbps")
        if lte_total is not None:
            try:
                lte_total_f = float(lte_total)
                if lte_total_f > 0:
                    pdsch_dl_lte_total_sum += lte_total_f
                    pdsch_dl_lte_rows += 1
                    if lte_cw1 is not None:
                        lte_cw1_f = float(lte_cw1)
                        if lte_cw1_f > 0:
                            pdsch_dl_lte_cw1_sum += lte_cw1_f
            except Exception:
                pass

    # Modulation is a sparse change-event column (logged only on transitions, like serving/packet
    # technology). Raw non-empty count badly misrepresents time-based shares. Forward-fill across
    # all row timestamps, then restrict to PDSCH-active seconds for CW0 and to rank-2-active
    # seconds for CW1, so idle time does not distort the distribution.
    from datetime import datetime as _dt_cls
    _pdsch_active_secs = {
        row["_dt"].replace(microsecond=0)
        for row in pdsch_rows
        if row.get("_dt") is not None
    }
    _rank2_active_secs = {
        row["_dt"].replace(microsecond=0)
        for row in pdsch_rows
        if row.get("_dt") is not None
        and (
            row.get("pdschMcsCw1") is not None
            or (row.get("scheduledRank") is not None and float(row.get("scheduledRank") or 0) >= 2)
        )
    }

    def _ff_modulation_counts(field: str, active_seconds: set) -> dict:
        """Forward-fill a sparse modulation column and return per-modulation second counts,
        restricted to the supplied active-second set."""
        timeline = sorted(
            (
                (r["_dt"], _nemo_clean_modulation(r.get(field)))
                for r in rows
                if r.get("_dt") is not None and _nemo_clean_modulation(r.get(field))
            ),
            key=lambda x: x[0],
        )
        per_sec: dict = {}
        last: str | None = None
        for dt, val in timeline:
            if val:
                last = val
            if last is None:
                continue
            per_sec[dt.replace(microsecond=0)] = last
        counts: dict = {}
        for sec, val in per_sec.items():
            if sec in active_seconds:
                counts[val] = counts.get(val, 0) + 1
        return counts

    cw0_counts = _ff_modulation_counts("pdschModulationCw0", _pdsch_active_secs)
    cw1_counts = _ff_modulation_counts("pdschModulationCw1", _rank2_active_secs)
    combined_counts: dict = {}
    for val, cnt in cw0_counts.items():
        combined_counts[val] = combined_counts.get(val, 0) + cnt
    for val, cnt in cw1_counts.items():
        combined_counts[val] = combined_counts.get(val, 0) + cnt
    modulation_distribution = _nemo_distribution_from_counts(combined_counts)
    mod_cw0_dist = _nemo_distribution_from_counts(cw0_counts)
    mod_cw1_dist = _nemo_distribution_from_counts(cw1_counts)
    rank2_util_pct = (
        round(pdsch_dl_lte_cw1_sum / pdsch_dl_lte_total_sum * 100.0, 1)
        if pdsch_dl_lte_total_sum > 0
        else None
    )
    pdsch_mcs_stats = _nemo_metric_stats(mcs_values)
    pdsch_tbs_stats = _nemo_metric_stats(tbs_values)
    pdsch_dl_lte_stats = _nemo_metric_stats(metric("pdschDlLteMbps", positive=True))
    pdsch_dl_lte_cw0_stats = _nemo_metric_stats(metric("pdschDlLteCw0Mbps", positive=True))
    pdsch_dl_lte_cw1_stats = _nemo_metric_stats(metric("pdschDlLteCw1Mbps", positive=True))
    pdsch_bits_hz_stats = _nemo_metric_stats(metric("pdschBitsPerHz", positive=True))
    pdsch_max_bits_hz_stats = _nemo_metric_stats(metric("pdschMaxBitsPerHz", positive=True))
    scheduled_rank_stats = _nemo_metric_stats(metric("scheduledRank", positive=True))
    lte_anchor_sinr_values = []
    for row in rows:
        sinr_value = row.get("sinr")
        if sinr_value is None:
            continue
        serving_upper = str(row.get("servingTechnology") or "").upper()
        packet_upper = str(row.get("packetTechnology") or "").upper()
        cell_types_upper = {
            str(cell or "").strip().upper()
            for cell in (row.get("cellTypes") or [])
        }
        if (
            row.get("nrChannelNumber") is not None
            or "EN-DC" in serving_upper
            or "EN-DC" in packet_upper
            or bool(cell_types_upper & {"NR SERVING", "NR SCG PSCELL", "SCG PSCELL", "5G SERVING"})
        ):
            try:
                lte_anchor_sinr_values.append(float(sinr_value))
            except Exception:
                pass

    # Time-based 5G/4G presence from kpis rows
    _kpi_ts = operator_data.get("technologyStatus") or {}
    kpi_nr_presence_pct = _kpi_ts.get("nrPresencePct")
    kpi_lte_only_presence_pct = _kpi_ts.get("lteOnlyPresencePct")

    # NR/LTE RF split. Prefer the per-row, column-index-based rsrpNr/rsrpLte/… fields (read
    # directly from each RAT's own column — reliable even when the sparse `band` is absent on
    # the RF row). Fall back to a band split of the conflated value only for rows restored
    # from an older cache that predate those fields. Total is the RAT-average of the NR and
    # LTE means — NOT a sample-weighted pooled mean — so it isn't dominated by whichever RAT
    # had more samples; counts are carried alongside.
    def _collect(key):
        out = []
        for r in rows:
            v = r.get(key)
            if v is None:
                continue
            try:
                out.append(float(v))
            except (TypeError, ValueError):
                pass
        return out

    def _rf_split_stats(conflated_key, nr_key, lte_key):
        nr_vals, lte_vals = _collect(nr_key), _collect(lte_key)
        if not nr_vals and not lte_vals:
            for r in rows:
                v = r.get(conflated_key)
                if v is None:
                    continue
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    continue
                b = str(r.get("band") or "").strip().lower()
                if b.startswith("n"):
                    nr_vals.append(fv)
                elif b.startswith("b"):
                    lte_vals.append(fv)
        return _nemo_metric_stats(nr_vals), _nemo_metric_stats(lte_vals)

    rsrp_nr_stats, rsrp_lte_stats = _rf_split_stats("rsrp", "rsrpNr", "rsrpLte")
    rsrq_nr_stats, rsrq_lte_stats = _rf_split_stats("rsrq", "rsrqNr", "rsrqLte")
    sinr_nr_stats, sinr_lte_stats = _rf_split_stats("sinr", "sinrNr", "sinrLte")

    def _rat_avg(nr_stats, lte_stats):
        vals = [v for v in (nr_stats.get("average"), lte_stats.get("average")) if v is not None]
        return round(sum(vals) / len(vals), 1) if vals else None

    rf_nr_lte = {
        "rsrpNrAvg": rsrp_nr_stats.get("average"), "rsrpLteAvg": rsrp_lte_stats.get("average"),
        "rsrpTotalAvg": _rat_avg(rsrp_nr_stats, rsrp_lte_stats),
        "rsrpNrSamples": rsrp_nr_stats.get("sampleCount"), "rsrpLteSamples": rsrp_lte_stats.get("sampleCount"),
        "rsrqNrAvg": rsrq_nr_stats.get("average"), "rsrqLteAvg": rsrq_lte_stats.get("average"),
        "rsrqTotalAvg": _rat_avg(rsrq_nr_stats, rsrq_lte_stats),
        "rsrqNrSamples": rsrq_nr_stats.get("sampleCount"), "rsrqLteSamples": rsrq_lte_stats.get("sampleCount"),
        "sinrNrAvg": sinr_nr_stats.get("average"), "sinrLteAvg": sinr_lte_stats.get("average"),
        "sinrTotalAvg": _rat_avg(sinr_nr_stats, sinr_lte_stats),
        "sinrNrSamples": sinr_nr_stats.get("sampleCount"), "sinrLteSamples": sinr_lte_stats.get("sampleCount"),
    }

    return {
        "dlMetricKey": dl_metric_key,
        "dl": app_stats,
        "appDl": app_dl_sample_stats,
        "successRate": round((sum(1 for test in tests if test.get("success")) / float(len(tests))) * 100.0, 1) if tests else None,
        "testCount": len(tests),
        "mac5g": _nemo_metric_stats(metric("macDl5gMbps", positive=True)),
        "macLte": mac_lte_stats,
        "nrThroughputContribPct": nr_throughput_contrib_pct,
        "lteThroughputContribPct": lte_throughput_contrib_pct,
        "scheduled5g": _nemo_metric_stats(metric("pdschSched5gMbps", positive=True)),
        "pdsch5g": _nemo_metric_stats(metric("pdschDl5gMbps", positive=True)),
        "prbs": _nemo_metric_stats(metric("pdschPrbs", positive=True)),
        "pdschSlotPct": pdsch_slot_stats,
        "scheduledRank": scheduled_rank_stats,
        "availableBandwidthPrbs": _nemo_metric_stats(available_prbs_values),
        "rsrp": _nemo_metric_stats(metric("rsrp")),
        "rsrq": _nemo_metric_stats(metric("rsrq")),
        "sinr": _nemo_metric_stats(metric("sinr")),
        "rsrpNr": rsrp_nr_stats,
        "rsrpLte": rsrp_lte_stats,
        "rsrqNr": rsrq_nr_stats,
        "rsrqLte": rsrq_lte_stats,
        "sinrNr": sinr_nr_stats,
        "sinrLte": sinr_lte_stats,
        "rfNrLte": rf_nr_lte,
        "cqi": _nemo_metric_stats(metric("wbCqi", positive=True)),
        "ri": _nemo_metric_stats(ri_values),
        "riGe3Share": round((sum(1 for value in ri_values if value >= 3) / float(len(ri_values))) * 100.0, 1) if ri_values else None,
        "ri1Share": round((sum(1 for value in ri_values if value <= 1) / float(len(ri_values))) * 100.0, 1) if ri_values else None,
        "bler": _nemo_metric_stats(metric("macDlBler")),
        "macUlRetx": _nemo_metric_stats(metric("macUlRetx5g")),
        # Additional capacity / reliability / CA / UL metrics from the richer export.
        "dlPrbUtilPct": _nemo_metric_stats(metric("dlPrbPct")),
        "prbsAvgDlAll": _nemo_metric_stats(metric("prbsAvgDl", positive=True)),
        "schBitratePerPrb": _nemo_metric_stats(metric("schBitratePerPrb", positive=True)),
        "pdschBlerLte": _nemo_metric_stats(metric("pdschBlerLte")),
        "macDlResidualBler": _nemo_metric_stats(metric("macDlResidualBler")),
        "pdcchBlerEst": _nemo_metric_stats(metric("pdcchBlerEst")),
        "macUlRetxLte": _nemo_metric_stats(metric("macUlRetxLte")),
        "caTotalBwMhz": _nemo_metric_stats(metric("caTotalBwMhz", positive=True)),
        "primaryBwMhz": _nemo_metric_stats(metric("primaryBwMhz", positive=True)),
        "sumSecondaryBwMhz": _nemo_metric_stats(metric("sumSecondaryBwMhz", positive=True)),
        "txPower": _nemo_metric_stats(metric("txPower")),
        "puschTxPower": _nemo_metric_stats(metric("puschTxPower")),
        "wbCqi0": _nemo_metric_stats(metric("wbCqi0", positive=True)),
        "wbCqi1": _nemo_metric_stats(metric("wbCqi1", positive=True)),
        "hoUplaneInterruptionMs": _nemo_metric_stats(metric("hoUplaneInterruptionMs", positive=True)),
        "pppRate": _nemo_metric_stats(metric("pppRateDl", positive=True)),
        "endcSecondaryNode": _nemo_endc_secondary_node_stats(rows),
        "totalMacDl": total_mac_stats,
        "transportRatio": transport_ratio,
        "tcpHandshake": _nemo_metric_stats(metric("tcpHandshakeMs", positive=True)),
        "lostPacket": _nemo_metric_stats(metric("lostPacket", positive=True)),
        "n78Share": n78_share,  # over all rows with band field
        "n78ShareNrOnly": n78_share_nr_only,  # over NR-only rows
        "n28ShareNrOnly": nr_band_shares.get("n28"),  # over NR-only rows
        "nrBandShares": nr_band_shares,  # {band(lower): share%} over NR-only rows
        "blerAbove10Share": bler_above_10_share,
        "blerAbove20Share": bler_above_20_share,
        "prbUtilPct": prb_util_pct,
        "pdschModulationCw0": {"distribution": mod_cw0_dist, "dominant": mod_cw0_dist[0]["label"] if mod_cw0_dist else None},
        "pdschModulationCw1": {"distribution": mod_cw1_dist, "dominant": mod_cw1_dist[0]["label"] if mod_cw1_dist else None},
        "rank2UtilPct": rank2_util_pct,
        "pdschDlLte": pdsch_dl_lte_stats,
        "pdschDlLteCw0": pdsch_dl_lte_cw0_stats,
        "pdschDlLteCw1": pdsch_dl_lte_cw1_stats,
        "pdschMcsCw0": mcs_cw0_stats,
        "pdschMcsCw1": mcs_cw1_stats,
        "rrcStateShares": rrc_state_shares,
        "applicationProtocolShares": app_protocol_shares,
        "servingTechnologyShares": serving_tech_shares,
        "packetTechnologyShares": packet_tech_shares,
        "scellsAverage": _nemo_metric_stats(scells_values).get("average") if scells_values else None,
        "scellsAvgAll": scells_avg_all,
        "scellsMax": scells_max,
        "scellsActiveShare": scells_active_share,
        "caActiveShare": ca_active_share,
        "prbEfficiency": prb_efficiency,
        "scheduledEfficiency": scheduled_efficiency,
        "resourceAllocationIndex": resource_allocation_index,
        "prbsPerScheduledSlot": prbs_per_scheduled_slot,
        "scheduledMbpsPerSlot": scheduled_mbps_per_slot,
        "lteAnchorSinr": _nemo_metric_stats(lte_anchor_sinr_values).get("median") if lte_anchor_sinr_values else None,
        "pdschModulation": {
            "dominant": modulation_distribution[0]["label"] if modulation_distribution else None,
            "distribution": modulation_distribution,
            "sampleCount": sum(combined_counts.values()),
            "qpskShare": _nemo_distribution_share(modulation_distribution, "QPSK"),
            "qam16Share": _nemo_distribution_share(modulation_distribution, "16QAM"),
            "qam64Share": _nemo_distribution_share(modulation_distribution, "64QAM"),
            "qam256Share": _nemo_distribution_share(modulation_distribution, "256QAM"),
            "cw0": {
                "dominant": mod_cw0_dist[0]["label"] if mod_cw0_dist else None,
                "distribution": mod_cw0_dist,
                "sampleCount": sum(cw0_counts.values()),
                "qpskShare": _nemo_distribution_share(mod_cw0_dist, "QPSK"),
                "qam16Share": _nemo_distribution_share(mod_cw0_dist, "16QAM"),
                "qam64Share": _nemo_distribution_share(mod_cw0_dist, "64QAM"),
                "qam256Share": _nemo_distribution_share(mod_cw0_dist, "256QAM"),
            },
            "cw1": {
                "dominant": mod_cw1_dist[0]["label"] if mod_cw1_dist else None,
                "distribution": mod_cw1_dist,
                "sampleCount": sum(cw1_counts.values()),
                "qpskShare": _nemo_distribution_share(mod_cw1_dist, "QPSK"),
                "qam16Share": _nemo_distribution_share(mod_cw1_dist, "16QAM"),
                "qam64Share": _nemo_distribution_share(mod_cw1_dist, "64QAM"),
                "qam256Share": _nemo_distribution_share(mod_cw1_dist, "256QAM"),
            },
        },
        "pdschMcs": pdsch_mcs_stats,
        "pdschBitPerHz": pdsch_bits_hz_stats,
        "pdschMaxBitPerHz": pdsch_max_bits_hz_stats,
        "pdschTbs": pdsch_tbs_stats,
        "pdschActiveSampleCount": len(pdsch_rows),
        "nrPresencePct": kpi_nr_presence_pct,
        "lteOnlyPresencePct": kpi_lte_only_presence_pct,
    }


def _nemo_build_ranking(operators: list[dict]) -> list[dict]:
    ranking = []
    for item in operators or []:
        kpis = item.get("kpis") or {}
        dl = kpis.get("dl") or {}
        app_dl = kpis.get("appDl") or {}
        ranking.append({
            "operator": item.get("operator") or "UNKNOWN",
            "avgDlMbps": dl.get("average"),
            "avgDlAppRateMbps": app_dl.get("average"),
            "medianDlMbps": dl.get("median"),
            "p10DlMbps": dl.get("p10"),
            "p90DlMbps": dl.get("p90"),
            "maxDlMbps": dl.get("max"),
            "sampleCount": dl.get("sampleCount"),
            "has5g": bool(item.get("has5g")),
            "fiveGStatus": item.get("fiveGStatus") or "",
        })
    ranking.sort(key=lambda entry: (-(entry.get("avgDlMbps") or -1), -(entry.get("medianDlMbps") or -1), str(entry.get("operator") or "")))
    for index, entry in enumerate(ranking, start=1):
        entry["rank"] = index
    return ranking


def _nemo_gap_severity(gap) -> str:
    try:
        value = abs(float(gap))
    except Exception:
        return "Low"
    if value >= 50:
        return "Critical"
    if value >= 25:
        return "High"
    if value >= 10:
        return "Medium"
    if value > 0:
        return "Low"
    return "—"


def _nemo_recommendations_for_cause(cause: str, comparator_name: str = "the best 5G comparator") -> list[str]:
    mapping = {
        "Scheduler / resource allocation": [
            f"Check IAM DL PRB utilization and active-user load during the benchmark window. If high, congestion is the likely root cause. If low but PRBs are still few, investigate scheduler policy or QoS.",
            f"Compare available NR BWP bandwidth: confirm IAM and {comparator_name} used the same NR carrier and bandwidth part. If IAM has a narrower BWP, the issue is configuration, not load.",
            f"Verify NR CA status and average #SCells. If {comparator_name} had more active NR carriers, IAM's scheduling resource pool is smaller by design.",
            f"Analyze PDSCH PRB allocation by CGPS test point to see if the gap is uniform (configuration) or location-dependent (load/coverage).",
            f"Compare PDSCH slot % for IAM vs {comparator_name}. Low slot % with low PRBs = time-domain underutilisation. High slot % with low PRBs = frequency-domain limitation.",
            f"Check scheduled rank (PDSCH scheduled rank) and RI distribution. Lower MIMO layer usage reduces effective capacity even when PRBs are allocated.",
            f"Verify MAC DL BLER and HARQ retransmission rate. If BLER is normal (< 10%), do not classify this as a radio-efficiency weakness — focus on the PRB allocation gap.",
            f"Validate SIM profile, APN, and QoS / 5QI across all operators. A test SIM with lower scheduler priority will receive fewer PRBs regardless of RF quality.",
            f"Add RTT / TCP handshake and packet loss KPIs to confirm the issue is not transport or core network. If TCP is clean, the bottleneck is radio resource allocation.",
            f"Repeat the benchmark in a low-load window and in a busy hour. If the PRB gap increases during busy hour, cell congestion is the root cause.",
        ],
        "Coverage limitation": [
            "Review weak-signal benchmark locations and compare RSRP lower-tail behavior.",
            "Check whether IAM drops to lower-capacity layers or edge coverage before transfers.",
        ],
        "Radio quality / interference": [
            "Correlate low-SINR samples with the throughput dips and high BLER windows.",
            "Check beam dominance, interference, and TDD synchronization around weak samples.",
        ],
        "Bandwidth / BWP limitation": [
            f"Compare active NR band and available bandwidth in PRBs against {comparator_name}.",
            "Check whether IAM spends less time on n78 or a smaller active bandwidth part.",
        ],
        "Carrier aggregation limitation": [
            "Review CA activation share and average #SCells during benchmark transfers.",
            f"Compare IAM CA usage against {comparator_name} at the same benchmark moments.",
        ],
        "MIMO limitation": [
            "Compare RI and scheduled rank by benchmark test window.",
            "Check beamforming and MIMO layer usage where RI stays low.",
        ],
        "Radio inefficiency / BLER": [
            "Check BLER spikes, retransmissions, CQI adaptation, and modulation/rank efficiency.",
            "Focus on cases where PRB allocation is acceptable but delivered throughput remains weak.",
        ],
        "Transport / core limitation": [
            "Compare application throughput versus total MAC throughput during healthy RF windows.",
            "Check TCP handshake, packet loss, APN/QoS, and server path consistency.",
        ],
        "No 5G detected": [
            "Confirm whether 5G/EN-DC was available and provisioned for the tested operator.",
            "If the operator was intentionally LTE-only, keep it as LTE-only benchmark context.",
        ],
    }
    return list(mapping.get(cause) or [])


def _nemo_recommendations_for_cause_fr(cause: str, comparator_name: str = "le meilleur comparateur 5G") -> list[str]:
    mapping_fr = {
        "Scheduler / resource allocation": [
            f"Vérifier le taux d'utilisation des PRBs DL d'IAM et la charge d'utilisateurs actifs pendant la fenêtre de benchmark. Si élevé, la congestion est la cause probable. Si faible mais peu de PRBs alloués, examiner la politique du scheduler ou la QoS.",
            f"Comparer la largeur de bande NR active (BWP) : confirmer qu'IAM et {comparator_name} utilisent la même porteuse NR et la même partie de bande passante. Si IAM a un BWP plus étroit, le problème est de configuration, pas de charge.",
            f"Vérifier le statut NR CA et le nombre moyen de SCells. Si {comparator_name} avait plus de porteuses NR actives, le pool de ressources du scheduler IAM est plus réduit par conception.",
            f"Analyser l'allocation des PRBs PDSCH par point de test CGPS pour voir si l'écart est uniforme (configuration) ou dépendant de la localisation (charge/couverture).",
            f"Comparer le pourcentage de slots PDSCH pour IAM vs {comparator_name}. Faible % de slots avec peu de PRBs = sous-utilisation temporelle. Élevé % de slots avec peu de PRBs = limitation fréquentielle.",
            f"Vérifier le rang planifié (PDSCH scheduled rank) et la distribution RI. Un usage MIMO multi-couches plus faible réduit la capacité effective même lorsque les PRBs sont alloués.",
            f"Vérifier le BLER MAC DL et le taux de retransmission HARQ. Si le BLER est normal (< 10 %), ne pas classifier cela comme une faiblesse radio — concentrer sur l'écart d'allocation PRBs.",
            f"Valider le profil SIM, l'APN et la QoS / 5QI pour tous les opérateurs. Une SIM de test avec une priorité scheduler inférieure recevra moins de PRBs indépendamment de la qualité RF.",
            f"Ajouter les KPIs RTT / TCP handshake et pertes de paquets pour confirmer que le problème n'est pas de transport ou de cœur de réseau. Si TCP est propre, le goulot d'étranglement est l'allocation des ressources radio.",
            f"Répéter le benchmark en période de faible charge et en heure chargée. Si l'écart PRBs augmente en heure chargée, la congestion cellulaire est la cause principale.",
        ],
        "Coverage limitation": [
            "Examiner les localisations de benchmark à signal faible et comparer le comportement de la queue basse RSRP.",
            "Vérifier si IAM bascule vers des couches de moindre capacité ou vers de la couverture marginale avant les transferts.",
        ],
        "Radio quality / interference": [
            "Corréler les échantillons SINR faible avec les creux de débit et les fenêtres de BLER élevé.",
            "Vérifier la dominance de faisceau, les interférences et la synchronisation TDD autour des échantillons faibles.",
        ],
        "Bandwidth / BWP limitation": [
            f"Comparer la bande NR active et la largeur de bande disponible en PRBs face à {comparator_name}.",
            "Vérifier si IAM passe moins de temps sur n78 ou utilise une partie de bande active plus étroite.",
        ],
        "Carrier aggregation limitation": [
            "Examiner le taux d'activation CA et le nombre moyen de SCells durant les transferts du benchmark.",
            f"Comparer l'usage CA d'IAM face à {comparator_name} aux mêmes instants du benchmark.",
        ],
        "MIMO limitation": [
            "Comparer RI et rang planifié par fenêtre de test benchmark.",
            "Vérifier la formation de faisceau et l'usage MIMO multi-couches là où le RI reste bas.",
        ],
        "Radio inefficiency / BLER": [
            "Vérifier les pics BLER, les retransmissions, l'adaptation CQI et l'efficacité de modulation/rang.",
            "Se concentrer sur les cas où l'allocation PRBs est correcte mais le débit livré reste faible.",
        ],
        "Transport / core limitation": [
            "Comparer le débit applicatif vs le débit MAC total pendant les fenêtres RF saines.",
            "Vérifier le TCP handshake, les pertes de paquets, l'APN/QoS et la cohérence du chemin serveur.",
        ],
        "No 5G detected": [
            "Confirmer si la 5G/EN-DC était disponible et provisionnée pour l'opérateur testé.",
            "Si l'opérateur était intentionnellement LTE-only, conserver ce contexte de benchmark LTE uniquement.",
        ],
    }
    return list(mapping_fr.get(cause) or [])


def _nemo_dominant_nr_serving_info(rows: list[dict]) -> dict | None:
    from datetime import datetime as _dt_class
    primary_tokens = {"SCG PSCELL", "NR SCG PSCELL", "NR SERVING", "5G SERVING"}
    counts: dict = {}
    first_last: dict = {}
    for row in rows or []:
        arfcn = row.get("nrChannelNumber")
        pci = row.get("pci")
        dt_val = row.get("_dt")
        if arfcn is None or pci is None or not isinstance(dt_val, _dt_class):
            continue
        cell_types = {str(ct or "").strip().upper() for ct in (row.get("cellTypes") or []) if str(ct or "").strip()}
        serving_tech = str(row.get("servingTechnology") or "").upper()
        if cell_types:
            if not any(token in primary_tokens for token in cell_types):
                continue
        elif "EN-DC" not in serving_tech and "5G" not in serving_tech:
            continue
        band = str(row.get("band") or "").strip() or None
        key = (int(float(pci)), int(float(arfcn)), band)
        counts[key] = counts.get(key, 0) + 1
        if key not in first_last:
            first_last[key] = [dt_val, dt_val]
        else:
            if dt_val < first_last[key][0]:
                first_last[key][0] = dt_val
            if dt_val > first_last[key][1]:
                first_last[key][1] = dt_val
    if not counts:
        return None
    best_key = max(counts.items(), key=lambda item: (item[1], item[0][2] == "n78", -item[0][0]))[0]
    pci, arfcn, band = best_key
    start_dt, end_dt = first_last[best_key]
    band_label = band or "NR"
    return {
        "pci": pci,
        "arfcn": arfcn,
        "band": band,
        "sampleCount": counts[best_key],
        "startDt": start_dt,
        "endDt": end_dt,
        "display": f"{band_label} PCI {pci} / ARFCN {arfcn}",
    }


def _nemo_build_diagnosis(operators: list[dict], ranking: list[dict]) -> dict:
    iam = next((item for item in operators if str(item.get("operator") or "").upper() == "IAM"), None)
    if not iam:
        return {"available": False, "summary": "IAM file is missing from the imported benchmark set."}
    best_dl = ranking[0] if ranking else None
    ranking_5g = [entry for entry in ranking if entry.get("has5g")]
    best_5g = ranking_5g[0] if ranking_5g else None
    comparator = None
    if iam.get("has5g"):
        comparator = next((
            entry for entry in ranking_5g
            if str(entry.get("operator") or "").upper() != "IAM"
            and entry.get("avgDlMbps") is not None
        ), None)
    if comparator is None:
        comparator = next((entry for entry in ranking if str(entry.get("operator") or "").upper() != "IAM" and entry.get("avgDlMbps") is not None), None)
    comparator_full = next((item for item in operators if comparator and item.get("operator") == comparator.get("operator")), None)
    iam_kpis = iam.get("kpis") or {}
    comparator_kpis = comparator_full.get("kpis") if comparator_full else {}
    iam_nr_info = _nemo_dominant_nr_serving_info(iam.get("rows") or [])
    comparator_nr_info = _nemo_dominant_nr_serving_info((comparator_full or {}).get("rows") or [])
    comparator_name = comparator.get("operator") if comparator else "the comparator"
    scores = {
        "Coverage limitation": 0,
        "Radio quality / interference": 0,
        "Scheduler / resource allocation": 0,
        "Bandwidth / BWP limitation": 0,
        "Carrier aggregation limitation": 0,
        "MIMO limitation": 0,
        "Radio inefficiency / BLER": 0,
        "Transport / core limitation": 0,
        "No 5G detected": 0,
    }
    evidence = []
    evidence_rows = []

    iam_rsrp_med = (iam_kpis.get("rsrp") or {}).get("median")
    iam_rsrp_p10 = (iam_kpis.get("rsrp") or {}).get("p10")
    iam_rsrq_med = (iam_kpis.get("rsrq") or {}).get("median")
    iam_sinr_med = (iam_kpis.get("sinr") or {}).get("median")
    iam_sinr_p10 = (iam_kpis.get("sinr") or {}).get("p10")
    iam_prbs_avg = (iam_kpis.get("prbs") or {}).get("average")
    cmp_prbs_avg = ((comparator_kpis or {}).get("prbs") or {}).get("average")
    iam_sched_avg = (iam_kpis.get("scheduled5g") or {}).get("average")
    cmp_sched_avg = ((comparator_kpis or {}).get("scheduled5g") or {}).get("average")
    iam_available_bw = (iam_kpis.get("availableBandwidthPrbs") or {}).get("average")
    cmp_available_bw = ((comparator_kpis or {}).get("availableBandwidthPrbs") or {}).get("average")
    iam_prb_util = iam_kpis.get("prbUtilPct")
    iam_bler_avg = (iam_kpis.get("bler") or {}).get("average")
    iam_bler_p90 = (iam_kpis.get("bler") or {}).get("p90")
    iam_transport_ratio = iam_kpis.get("transportRatio")
    iam_ri_ge3 = iam_kpis.get("riGe3Share")
    iam_ri1_share = iam_kpis.get("ri1Share")
    cmp_ri_ge3 = (comparator_kpis or {}).get("riGe3Share")
    cmp_n78 = (comparator_kpis or {}).get("n78Share")
    iam_n78 = iam_kpis.get("n78Share")
    iam_nr_band = str((iam_nr_info or {}).get("band") or "").lower()
    comparator_nr_band = str((comparator_nr_info or {}).get("band") or "").lower()
    nr_band_mismatch = bool(iam_nr_band and comparator_nr_band and iam_nr_band != comparator_nr_band)
    lower_capacity_band_gap = bool(comparator_nr_band == "n78" and iam_nr_band and iam_nr_band != "n78")
    cmp_ca_share = (comparator_kpis or {}).get("caActiveShare")
    iam_ca_share = iam_kpis.get("caActiveShare")
    cmp_scells = (comparator_kpis or {}).get("scellsAverage")
    iam_scells = iam_kpis.get("scellsAverage")
    rf_good = iam_rsrp_med is not None and iam_rsrp_med >= -95 and iam_sinr_med is not None and iam_sinr_med >= 8

    if not iam.get("has5g"):
        scores["No 5G detected"] += 100
        evidence.append("IAM export shows no 5G NR/EN-DC samples, so 5G-specific diagnosis is not applicable.")

    if iam_rsrp_med is not None and iam_rsrp_med < -105:
        scores["Coverage limitation"] += 40
        evidence.append(f"IAM median RSRP is weak at {iam_rsrp_med} dBm.")
    if iam_rsrp_p10 is not None and iam_rsrp_p10 < -110:
        scores["Coverage limitation"] += 20
        evidence.append(f"IAM lower-tail RSRP reaches {iam_rsrp_p10} dBm.")
    if iam_rsrp_med is not None and iam_sinr_med is not None and iam_rsrp_med < -105 and iam_sinr_med < 5:
        scores["Coverage limitation"] += 25
        evidence.append(
            f"IAM shows combined coverage and quality limitation: median RSRP {iam_rsrp_med} dBm with median SINR {iam_sinr_med} dB."
        )
    serving_distance_m = iam_kpis.get("servingCellDistanceM")
    if serving_distance_m is not None and iam_rsrp_med is not None and serving_distance_m > 1000 and iam_rsrp_med < -100:
        scores["Coverage limitation"] += 20
        evidence.append(
            f"IAM serving cell appears far ({serving_distance_m} m) while median RSRP is only {iam_rsrp_med} dBm, suggesting overshooting or a missing dominant sector."
        )

    if iam_sinr_med is not None and iam_sinr_med < 5:
        scores["Radio quality / interference"] += 40
        evidence.append(f"IAM median SINR is poor at {iam_sinr_med} dB.")
    if iam_sinr_p10 is not None and iam_sinr_p10 < 0:
        scores["Radio quality / interference"] += 20
        evidence.append(f"IAM SINR P10 drops to {iam_sinr_p10} dB.")
    if iam_rsrq_med is not None and iam_rsrq_med < -14:
        scores["Radio quality / interference"] += 25
        evidence.append(f"IAM median RSRQ is poor at {iam_rsrq_med} dB.")
    if iam_rsrq_med is not None and iam_rsrq_med < -17:
        scores["Radio quality / interference"] += 20
        evidence.append(f"IAM RSRQ is very poor ({iam_rsrq_med} dB), which strongly suggests interference, overlap or high load.")
    if iam_rsrp_med is not None and iam_rsrp_med >= -95 and iam_sinr_med is not None and iam_sinr_med < 8:
        scores["Radio quality / interference"] += 30
        evidence.append("IAM has acceptable coverage but weak SINR, which points more to quality/interference than pure coverage.")
    if iam_rsrp_med is not None and iam_rsrp_med > -95 and iam_sinr_med is not None and iam_sinr_med < 5:
        scores["Radio quality / interference"] += 20
        evidence.append(
            f"IAM has good coverage but poor quality (RSRP {iam_rsrp_med} dBm, SINR {iam_sinr_med} dB), which indicates interference/pollution/dominance problems."
        )

    prbs_gap = _nemo_gap_pct(iam_prbs_avg, cmp_prbs_avg)
    sched_gap = _nemo_gap_pct(iam_sched_avg, cmp_sched_avg)
    bw_gap = _nemo_gap_pct(iam_available_bw, cmp_available_bw)
    n78_gap = _nemo_gap_pct(iam_n78, cmp_n78)
    iam_pdsch5g_avg = (iam_kpis.get("pdsch5g") or {}).get("average")
    cmp_pdsch5g_avg = ((comparator_kpis or {}).get("pdsch5g") or {}).get("average")
    delivered_gap = _nemo_gap_pct(iam_pdsch5g_avg, cmp_pdsch5g_avg)
    gap_to_comparator = _nemo_gap_pct((iam_kpis.get("dl") or {}).get("average"), (comparator or {}).get("avgDlMbps"))
    iam_prb_eff = iam_kpis.get("prbEfficiency")
    cmp_prb_eff = (comparator_kpis or {}).get("prbEfficiency")
    iam_res_alloc_idx = iam_kpis.get("resourceAllocationIndex")
    cmp_res_alloc_idx = (comparator_kpis or {}).get("resourceAllocationIndex")
    # Detect when IAM slot % and allocation ratio are not worse than comparator → scheduler not confirmed
    _iam_slot_avg = (iam_kpis.get("pdschSlotPct") or {}).get("average")
    _cmp_slot_avg = ((comparator_kpis or {}).get("pdschSlotPct") or {}).get("average")
    _pdsch_slot_gap = _nemo_gap_pct(_iam_slot_avg, _cmp_slot_avg)
    _iam_better_slot = _pdsch_slot_gap is not None and _pdsch_slot_gap > -10
    _iam_better_alloc = iam_res_alloc_idx is not None and cmp_res_alloc_idx is not None and iam_res_alloc_idx >= cmp_res_alloc_idx * 0.9
    scheduler_not_confirmed = bool(_iam_better_slot and _iam_better_alloc)
    spectral = _nemo_spectral_efficiency_insight(iam_kpis, comparator_kpis or {})

    if prbs_gap is not None and prbs_gap <= -50:
        scores["Scheduler / resource allocation"] += 55
        evidence.append(f"IAM PRB allocation is {abs(prbs_gap)}% below {comparator_name} — the primary evidence of resource starvation.")
    elif prbs_gap is not None and prbs_gap <= -25:
        scores["Scheduler / resource allocation"] += 35
        evidence.append(f"IAM PRB allocation is materially below {comparator_name} ({prbs_gap}%).")
    if sched_gap is not None and sched_gap <= -25:
        scores["Scheduler / resource allocation"] += 35
        evidence.append(f"IAM scheduled 5G throughput is {abs(sched_gap)}% below {comparator_name}.")
    # RF good + low PRBs = strong scheduler signal
    if rf_good and prbs_gap is not None and prbs_gap <= -25:
        scores["Scheduler / resource allocation"] += 30
        if prbs_gap <= -40:
            evidence.append(f"IAM RF is acceptable but PRB allocation is {abs(prbs_gap)}% lower than {comparator_name} — this is a scheduler/resource-allocation pattern, not a coverage issue.")
    # PRB gap larger than delivered gap → IAM is allocation-limited, not efficiency-limited
    if prbs_gap is not None and delivered_gap is not None and abs(prbs_gap) > abs(delivered_gap) + 10:
        scores["Scheduler / resource allocation"] += 15
        evidence.append(f"PRB gap ({abs(prbs_gap)}%) exceeds delivered throughput gap ({abs(delivered_gap)}%), confirming IAM is resource-allocation limited, not radio-efficiency limited.")
    # IAM more efficient per PRB than comparator → reduce BLER score
    if iam_prb_eff is not None and cmp_prb_eff is not None and iam_prb_eff > cmp_prb_eff * 1.1:
        scores["Radio inefficiency / BLER"] = max(0, scores["Radio inefficiency / BLER"] - 20)
        evidence.append(f"IAM PRB efficiency ({round(iam_prb_eff, 2)} Mbps/PRB) exceeds {comparator_name} ({round(cmp_prb_eff, 2)} Mbps/PRB), indicating good radio efficiency — BLER is less likely the root cause.")
    # BWP/config limitation: IAM has similar available BW but still fewer allocated PRBs → stronger scheduler signal
    if iam_res_alloc_idx is not None and cmp_res_alloc_idx is not None:
        if bw_gap is not None and bw_gap >= -10 and prbs_gap is not None and prbs_gap <= -30:
            scores["Scheduler / resource allocation"] += 20
            evidence.append(f"IAM has similar available NR bandwidth to {comparator_name} but allocates significantly fewer PRBs — confirms a scheduler/load limitation rather than a bandwidth configuration issue.")
        elif bw_gap is not None and bw_gap <= -25:
            scores["Bandwidth / BWP limitation"] += 20
            evidence.append(f"IAM has materially less available NR bandwidth than {comparator_name}, which may directly limit PRB allocation.")

    if n78_gap is not None and n78_gap <= -25:
        scores["Bandwidth / BWP limitation"] += 25
        evidence.append(f"IAM spends much less time on n78 than {comparator_name}.")

    if iam_ca_share is not None and iam_ca_share < 50:
        scores["Carrier aggregation limitation"] += 25
    if iam_scells is not None and cmp_scells is not None and (iam_scells - cmp_scells) <= -1:
        scores["Carrier aggregation limitation"] += 20
        evidence.append(f"IAM uses fewer SCells on average than {comparator_name}.")

    if iam_ri_ge3 is not None and iam_ri_ge3 < 25:
        scores["MIMO limitation"] += 25
        evidence.append(f"IAM RI>=3 share is only {iam_ri_ge3}%.")
    if iam_ri1_share is not None and iam_ri1_share > 30:
        scores["MIMO limitation"] += 25
        evidence.append(f"IAM RI1 share is high at {iam_ri1_share}%.")
    if iam_ri_ge3 is not None and cmp_ri_ge3 is not None and (iam_ri_ge3 - cmp_ri_ge3) <= -15:
        scores["MIMO limitation"] += 20

    if iam_bler_avg is not None and iam_bler_avg > 10:
        scores["Radio inefficiency / BLER"] += 35
        evidence.append(f"IAM average MAC DL BLER is high at {iam_bler_avg}%.")
    if iam_bler_p90 is not None and iam_bler_p90 > 20:
        scores["Radio inefficiency / BLER"] += 25
        evidence.append(f"IAM BLER P90 reaches {iam_bler_p90}%.")
    if iam_prbs_avg is not None and cmp_prbs_avg is not None and iam_prbs_avg >= cmp_prbs_avg * 0.8:
        pdsch_gap = _nemo_gap_pct((iam_kpis.get("pdsch5g") or {}).get("average"), ((comparator_kpis or {}).get("pdsch5g") or {}).get("average"))
        if pdsch_gap is not None and pdsch_gap <= -25:
            scores["Radio inefficiency / BLER"] += 30
            evidence.append("IAM receives near-comparable PRBs but still delivers much lower 5G PDSCH throughput.")
    if spectral.get("confirmed"):
        evidence.append(
            "IAM also shows lower spectral efficiency than the 5G comparator (modulation/MCS/bit/s/Hz), so the throughput gap is not explained by PRB count alone."
        )
    elif spectral.get("iamBetter"):
        evidence.append(
            f"IAM spectral efficiency (modulation/MCS/bit/s/Hz) is better than {comparator_name} — "
            f"the throughput gap is not driven by poor radio efficiency. The main confirmed issue is the NR resource pool / BWP capacity gap."
        )
        if scheduler_not_confirmed:
            evidence.append(
                f"Scheduler/PRB allocation is not confirmed as a root cause: IAM PDSCH slot % and allocation ratio are not worse than {comparator_name}."
            )
    elif not spectral.get("sufficient"):
        evidence.append(
            f"Spectral-efficiency evidence is limited ({spectral.get('iamSamples')} IAM samples vs {spectral.get('comparatorSamples')} comparator samples), so modulation/MCS conclusions remain guarded."
        )

    if rf_good and iam_transport_ratio is not None and iam_transport_ratio < 0.5:
        scores["Transport / core limitation"] += 35
        evidence.append("IAM application throughput is much lower than total MAC throughput despite acceptable RF.")
    if iam_prb_util is not None and gap_to_comparator is not None and gap_to_comparator <= -20:
        if iam_prb_util > 90:
            scores["Scheduler / resource allocation"] += 25
            evidence.append(
                f"IAM DL PRB utilization is severe at {iam_prb_util}%, so congestion is likely contributing to the throughput gap."
            )
        elif iam_prb_util > 80:
            scores["Scheduler / resource allocation"] += 15
            evidence.append(
                f"IAM DL PRB utilization is high at {iam_prb_util}%, indicating a congestion contribution to the throughput gap."
            )
        elif iam_prb_util < 50:
            scores["Transport / core limitation"] += 10
            evidence.append(
                f"IAM DL PRB utilization is only {iam_prb_util}%, so the low throughput is unlikely to be explained by congestion alone."
            )
    if iam_sinr_med is not None and iam_sinr_med > 10 and gap_to_comparator is not None and gap_to_comparator <= -30:
        if prbs_gap is not None and prbs_gap <= -20:
            scores["Scheduler / resource allocation"] += 20
            evidence.append(
                f"IAM throughput remains {abs(gap_to_comparator)}% below {comparator_name} despite good SINR, reinforcing a scheduler/resource-allocation limitation."
            )
        else:
            scores["Transport / core limitation"] += 20
            evidence.append(
                f"IAM throughput remains {abs(gap_to_comparator)}% below {comparator_name} despite good SINR, so capacity/configuration or transport limitations should be checked ahead of RF changes."
            )
    tcp_avg = (iam_kpis.get("tcpHandshake") or {}).get("average")
    if tcp_avg is not None and tcp_avg > 200:
        scores["Transport / core limitation"] += 15
    loss_avg = (iam_kpis.get("lostPacket") or {}).get("average")
    if loss_avg is not None and loss_avg > 0:
        scores["Transport / core limitation"] += 20

    ranked_scores = sorted(
        [{"cause": cause, "score": score} for cause, score in scores.items() if score > 0],
        key=lambda item: (-item["score"], item["cause"])
    )
    main_cause = ranked_scores[0]["cause"] if ranked_scores else "No clear dominant cause"
    main_score = ranked_scores[0]["score"] if ranked_scores else 0
    if main_score >= 70:
        severity = "Critical"
    elif main_score >= 50:
        severity = "High"
    elif main_score >= 25:
        severity = "Medium"
    elif main_score > 0:
        severity = "Low"
    else:
        severity = "Not detected"

    iam_rank = next((entry.get("rank") for entry in ranking if str(entry.get("operator") or "").upper() == "IAM"), None)
    iam_rank_5g = next((index for index, entry in enumerate(ranking_5g, start=1) if str(entry.get("operator") or "").upper() == "IAM"), None)
    gap_to_best = _nemo_gap_pct((iam_kpis.get("dl") or {}).get("average"), (comparator or {}).get("avgDlMbps"))
    gap_to_best_dl = _nemo_gap_pct((iam_kpis.get("dl") or {}).get("average"), (best_dl or {}).get("avgDlMbps"))
    diagnosis_type = "5G-specific diagnosis" if iam.get("has5g") and comparator_full and comparator_full.get("has5g") else "Cross-operator diagnosis"

    ranking_interpretation = []
    if best_dl and not best_dl.get("has5g"):
        ranking_interpretation.append(
            f"{best_dl.get('operator')} ranks first in measured DL throughput, although no 5G NR/EN-DC samples were detected in its export. "
            f"Therefore, {best_dl.get('operator')} is valid for DL throughput ranking but not valid for 5G-specific radio comparison."
        )
    elif best_dl:
        ranking_interpretation.append(
            f"{best_dl.get('operator')} ranks first in measured DL throughput and remains the best overall DL benchmark performer in this export."
        )
    if comparator_full and comparator_full.get("has5g") and iam.get("has5g"):
        ranking_interpretation.append(
            f"Among operators with 5G detected, {comparator_name} outperforms IAM significantly. "
            f"IAM ranks {iam_rank or 'N/A'} overall and {iam_rank_5g or 'N/A'} among 5G-detected operators."
        )

    display_main_cause = main_cause
    if main_cause == "Scheduler / resource allocation" and lower_capacity_band_gap:
        display_main_cause = "NR band/BWP/resource allocation difference"

    if comparator and gap_to_best is not None:
        summary = (
            f"IAM ranks {iam_rank or 'N/A'} in measured DL throughput, with a {gap_to_best}% gap to {comparator_name}. "
            f"The main suspected cause is {display_main_cause}."
        )
    else:
        summary = f"IAM ranks {iam_rank or 'N/A'} in measured DL throughput. The main suspected cause is {display_main_cause}."

    technical_interpretation = ""
    technical_interpretation_fr = ""
    if main_cause == "Scheduler / resource allocation" and comparator_full and comparator_full.get("has5g"):
        if lower_capacity_band_gap:
            technical_interpretation = (
                f"IAM appears limited by a lower-capacity NR layer and weaker NR resource allocation, not primarily by RF coverage. "
                f"IAM uses {iam_nr_band or 'its active NR band'} while {comparator_name} uses {comparator_nr_band or 'its active NR band'}, "
                f"and IAM also receives fewer PRBs with lower scheduled 5G throughput. This points first to NR band/BWP/capacity difference, "
                f"then to scheduler policy, load, QoS/SIM priority, or CA configuration."
            )
            if spectral.get("confirmed"):
                technical_interpretation += " Lower modulation/MCS/bit-per-Hz indicates that weaker spectral efficiency also contributes as a secondary factor."
            elif spectral.get("iamBetter"):
                technical_interpretation += (
                    f" IAM modulation, MCS and bit/s/Hz are better than {comparator_name}: modulation quality is not a limiting factor."
                )
                if scheduler_not_confirmed:
                    technical_interpretation += (
                        f" Scheduler/PRB allocation limitation is not confirmed — IAM PDSCH slot % and allocation ratio are not worse than {comparator_name}."
                    )
            technical_interpretation_fr = (
                f"IAM apparaît limité par une couche NR de moindre capacité et par une allocation de ressources NR plus faible, "
                f"pas principalement par la couverture RF. IAM utilise {iam_nr_band or 'sa bande NR active'} alors que {comparator_name} "
                f"utilise {comparator_nr_band or 'sa bande NR active'}, et IAM reçoit aussi moins de PRBs avec un débit 5G planifié plus faible. "
                f"Cela oriente d'abord vers une différence de bande/BWP/capacité NR, puis vers la politique scheduler, la charge, la priorité QoS/SIM ou la configuration CA."
            )
            if spectral.get("confirmed"):
                technical_interpretation_fr += " Une modulation/MCS/efficacité bit/s/Hz plus faible indique aussi une limitation secondaire d'efficacité spectrale."
            elif spectral.get("iamBetter"):
                technical_interpretation_fr += (
                    f" La modulation, le MCS et le bit/s/Hz IAM sont meilleurs que {comparator_name} : la qualité de modulation n'est pas un facteur limitant."
                )
                if scheduler_not_confirmed:
                    technical_interpretation_fr += (
                        f" L'allocation PRBs / scheduler n'est pas confirmée comme cause principale — le PDSCH slot % et le taux d'allocation IAM ne sont pas inférieurs à {comparator_name}."
                    )
        else:
            technical_interpretation = (
                f"IAM appears resource-limited, not necessarily RF-limited. The network allocates significantly fewer NR downlink resources "
                f"to the IAM UE compared with {comparator_name}. Lower n78 exposure, fewer allocated PRBs, and lower scheduled 5G throughput "
                f"together explain the weaker delivered DL throughput."
            )
            if spectral.get("confirmed"):
                technical_interpretation += " Lower modulation/MCS/bit-per-Hz shows that spectral efficiency also contributes, so the gap is not purely scheduler-driven."
            elif spectral.get("iamBetter"):
                technical_interpretation += (
                    f" IAM modulation, MCS and bit/s/Hz are better than {comparator_name}: the gap is not explained by poor spectral efficiency."
                )
                if scheduler_not_confirmed:
                    technical_interpretation += (
                        f" Scheduler/PRB allocation limitation is not confirmed — IAM PDSCH slot % and allocation ratio are not worse than {comparator_name}."
                    )
            technical_interpretation_fr = (
                f"IAM apparaît limité en ressources, pas nécessairement en qualité RF. Le réseau alloue significativement moins de ressources "
                f"NR downlink à l'UE IAM par rapport à {comparator_name}. Une exposition n78 réduite, moins de PRBs alloués et un débit 5G "
                f"planifié plus faible expliquent conjointement le débit DL livré plus faible."
            )
            if spectral.get("confirmed"):
                technical_interpretation_fr += " Une modulation/MCS/efficacité bit/s/Hz plus faible montre qu'une limitation d'efficacité spectrale contribue aussi à l'écart ; ce n'est donc pas un problème purement scheduler."
            elif spectral.get("iamBetter"):
                technical_interpretation_fr += (
                    f" La modulation, le MCS et le bit/s/Hz IAM sont meilleurs que {comparator_name} : l'écart n'est pas dû à une mauvaise efficacité spectrale."
                )
                if scheduler_not_confirmed:
                    technical_interpretation_fr += (
                        f" L'allocation PRBs / scheduler n'est pas confirmée comme cause principale — le PDSCH slot % et le taux d'allocation IAM ne sont pas inférieurs à {comparator_name}."
                    )
    elif main_cause == "Coverage limitation":
        technical_interpretation = "IAM appears primarily coverage-limited, with weak serving-cell signal levels reducing throughput potential."
        technical_interpretation_fr = "IAM apparaît principalement limité par la couverture, avec des niveaux de signal en cellule servant faibles réduisant le potentiel de débit."
    elif main_cause == "Radio quality / interference":
        technical_interpretation = "IAM appears quality-limited rather than purely coverage-limited, with poor SINR or interference reducing usable 5G capacity."
        technical_interpretation_fr = "IAM apparaît limité par la qualité radio plutôt que purement par la couverture, avec un SINR faible ou des interférences réduisant la capacité 5G utilisable."
    else:
        technical_interpretation = summary
        technical_interpretation_fr = (
            f"IAM se classe {iam_rank or 'N/A'} en débit DL mesuré"
            + (f", avec un écart de {gap_to_best}% par rapport à {comparator_name}" if comparator and gap_to_best is not None else "")
            + f". La cause principale suspectée est {display_main_cause}."
        )

    if comparator and gap_to_best is not None:
        evidence_rows.append({
            "kpi": "Avg DL throughput",
            "iam": _nemo_safe_round((iam_kpis.get("dl") or {}).get("average"), 1),
            "comparator": _nemo_safe_round((comparator_kpis.get("dl") or {}).get("average"), 1),
            "gapPct": gap_to_best,
            "severity": _nemo_gap_severity(gap_to_best),
            "interpretation": "IAM user throughput is much lower" if gap_to_best <= -25 else "IAM user throughput is lower",
            "interpretation_fr": "Le débit utilisateur IAM est nettement plus faible" if gap_to_best <= -25 else "Le débit utilisateur IAM est plus faible",
            "unit": "Mbps",
        })
    if comparator_full and comparator_full.get("has5g"):
        if prbs_gap is not None:
            evidence_rows.append({
                "kpi": "PDSCH PRBs",
                "iam": _nemo_safe_round(iam_prbs_avg, 2),
                "comparator": _nemo_safe_round(cmp_prbs_avg, 2),
                "gapPct": prbs_gap,
                "severity": _nemo_gap_severity(prbs_gap),
                "interpretation": "IAM receives far fewer NR downlink PRBs" if prbs_gap <= -25 else "IAM receives fewer NR downlink PRBs",
                "interpretation_fr": "IAM reçoit nettement moins de PRBs NR downlink" if prbs_gap <= -25 else "IAM reçoit moins de PRBs NR downlink",
                "unit": "",
            })
        if sched_gap is not None:
            evidence_rows.append({
                "kpi": "Scheduled 5G throughput",
                "iam": _nemo_safe_round(iam_sched_avg, 2),
                "comparator": _nemo_safe_round(cmp_sched_avg, 2),
                "gapPct": sched_gap,
                "severity": _nemo_gap_severity(sched_gap),
                "interpretation": "IAM has lower scheduled NR capacity" if sched_gap <= -25 else "IAM has lower scheduled 5G throughput",
                "interpretation_fr": "IAM dispose d'une capacité NR planifiée plus faible" if sched_gap <= -25 else "IAM a un débit 5G planifié plus faible",
                "unit": "Mbps",
            })
        if iam_n78 is not None or cmp_n78 is not None:
            evidence_rows.append({
                "kpi": "n78 exposure",
                "iam": _nemo_safe_round(iam_n78, 1),
                "comparator": _nemo_safe_round(cmp_n78, 1),
                "gapPct": n78_gap,
                "severity": _nemo_gap_severity(n78_gap) if n78_gap is not None else "High",
                "interpretation": (
                    f"IAM did not use n78 in this export; its detected 5G layer was {iam_nr_band or 'unknown'} only, "
                    f"while {comparator_name} used {comparator_nr_band or 'n78'}. This materially reduces IAM's 5G capacity potential."
                    if lower_capacity_band_gap and (iam_n78 is None or iam_n78 == 0.0)
                    else f"IAM spends less time on high-capacity n78 than {comparator_name}"
                ) if (n78_gap is None or n78_gap <= 0) else "IAM has comparable n78 exposure",
                "interpretation_fr": (
                    f"IAM n'a pas utilisé n78 dans cet export ; sa couche 5G détectée était {iam_nr_band or 'inconnue'} uniquement, "
                    f"tandis que {comparator_name} utilisait {comparator_nr_band or 'n78'}. Cela réduit sensiblement le potentiel de capacité 5G d'IAM."
                    if lower_capacity_band_gap and (iam_n78 is None or iam_n78 == 0.0)
                    else f"IAM passe moins de temps sur n78 haute capacité que {comparator_name}"
                ) if (n78_gap is None or n78_gap <= 0) else "IAM a une exposition n78 comparable",
                "unit": "%",
            })
        evidence_rows.append({
            "kpi": "5G status",
            "iam": "Detected" if iam.get("has5g") else "Not detected",
            "comparator": "Detected" if comparator_full.get("has5g") else "Not detected",
            "gapPct": None,
            "severity": "—",
            "interpretation": "IAM and the comparator can be compared on 5G KPIs" if iam.get("has5g") and comparator_full.get("has5g") else "5G KPI comparison is not valid",
            "interpretation_fr": "IAM et le comparateur peuvent être comparés sur les KPIs 5G" if iam.get("has5g") and comparator_full.get("has5g") else "La comparaison des KPIs 5G n'est pas valide",
            "unit": "",
        })
        if spectral.get("sufficient"):
            _spec_interp = spectral.get("note") or (
                "IAM spectral efficiency is better than the comparator — modulation quality is not the throughput limitation."
                if spectral.get("iamBetter")
                else "Spectral efficiency comparable between operators"
            )
            _spec_interp_fr = spectral.get("note") or (
                "L'efficacité spectrale IAM est meilleure que le comparateur — la qualité de modulation n'est pas la limitation de débit."
                if spectral.get("iamBetter")
                else "Efficacité spectrale comparable entre opérateurs"
            )
            evidence_rows.append({
                "kpi": "PDSCH spectral efficiency",
                "iam": _nemo_safe_round(spectral.get("iamMedianBitsPerHz"), 2),
                "comparator": _nemo_safe_round(spectral.get("comparatorMedianBitsPerHz"), 2),
                "gapPct": _nemo_gap_pct(spectral.get("iamMedianBitsPerHz"), spectral.get("comparatorMedianBitsPerHz")),
                "severity": "Medium" if spectral.get("confirmed") else "—",
                "interpretation": _spec_interp,
                "interpretation_fr": _spec_interp_fr,
                "unit": "bit/s/Hz",
            })

    summary_fr = (
        f"IAM se classe {iam_rank or 'N/A'} en débit DL mesuré, avec un écart de {gap_to_best}% par rapport à {comparator_name}. "
        f"La cause principale suspectée est {display_main_cause}."
    ) if comparator and gap_to_best is not None else f"IAM se classe {iam_rank or 'N/A'} en débit DL mesuré. La cause principale suspectée est {display_main_cause}."

    ranking_interpretation_fr = []
    if best_dl and not best_dl.get("has5g"):
        ranking_interpretation_fr.append(
            f"{best_dl.get('operator')} se classe premier en débit DL mesuré, bien qu'aucun échantillon 5G NR/EN-DC n'ait été détecté dans son export. "
            f"Par conséquent, {best_dl.get('operator')} est valide pour le classement DL mais pas pour la comparaison radio 5G spécifique."
        )
    elif best_dl:
        ranking_interpretation_fr.append(
            f"{best_dl.get('operator')} se classe premier en débit DL mesuré et reste le meilleur opérateur benchmark DL global dans cet export."
        )
    if comparator_full and comparator_full.get("has5g") and iam.get("has5g"):
        ranking_interpretation_fr.append(
            f"Parmi les opérateurs avec 5G détectée, {comparator_name} surpasse IAM significativement. "
            f"IAM se classe {iam_rank or 'N/A'} au global et {iam_rank_5g or 'N/A'} parmi les opérateurs 5G détectés."
        )

    iam_avg_dl_diag = _nemo_safe_round((iam_kpis.get('dl') or {}).get('average'), 1)
    cmp_avg_dl_diag = _nemo_safe_round((comparator_kpis.get('dl') or {}).get('average'), 1) if comparator_kpis else None
    best_dl_avg_diag = _nemo_safe_round((best_dl or {}).get('avgDlMbps'), 1) if best_dl else None
    best_dl_op_diag = (best_dl or {}).get('operator') if best_dl else None

    conclusion = (
        f"IAM ranks third in measured DL throughput with an average of {iam_avg_dl_diag} Mbps"
        + (f", compared with {cmp_avg_dl_diag} Mbps for {comparator_name}" if comparator else "")
        + (f" and {best_dl_avg_diag} Mbps for {best_dl_op_diag}" if best_dl and str(best_dl_op_diag or '').upper() != 'IAM' and best_dl_op_diag != comparator_name else "")
        + "."
    )
    conclusion_fr = (
        f"IAM se classe troisième en débit DL mesuré avec une moyenne de {iam_avg_dl_diag} Mbps"
        + (f", contre {cmp_avg_dl_diag} Mbps pour {comparator_name}" if comparator else "")
        + (f" et {best_dl_avg_diag} Mbps pour {best_dl_op_diag}" if best_dl and str(best_dl_op_diag or '').upper() != 'IAM' and best_dl_op_diag != comparator_name else "")
        + "."
    )
    if best_dl and not best_dl.get("has5g"):
        conclusion += f" {best_dl_op_diag} is kept in the ranking but is marked as No 5G detected in export, so it is not used for 5G-specific radio comparison."
        conclusion_fr += f" {best_dl_op_diag} est maintenu dans le classement mais marqué sans 5G détectée dans l'export ; il n'est pas utilisé pour la comparaison radio 5G spécifique."
    if main_cause == "Scheduler / resource allocation" and comparator_full and comparator_full.get("has5g"):
        if lower_capacity_band_gap:
            conclusion += (
                f" For the 5G comparison between IAM and {comparator_name}, IAM's main weakness is a lower-capacity NR layer plus weaker NR resource allocation. "
                f"IAM uses {iam_nr_band or 'its active NR band'}"
                + (f" PCI {(iam_nr_info or {}).get('pci')} / ARFCN {(iam_nr_info or {}).get('arfcn')}" if iam_nr_info else "")
                + f", while {comparator_name} uses {comparator_nr_band or 'its active NR band'}"
                + (f" PCI {(comparator_nr_info or {}).get('pci')} / ARFCN {(comparator_nr_info or {}).get('arfcn')}" if comparator_nr_info else "")
                + f". IAM also receives {str(abs(prbs_gap)) + '% fewer' if prbs_gap is not None else 'materially fewer'} PDSCH PRBs"
                + (f" and has {abs(sched_gap)}% lower scheduled 5G throughput" if sched_gap is not None else "")
                + ". This should not be described as pure scheduler starvation; the first hypothesis is NR band/BWP/capacity difference, then scheduler policy, load, QoS/APN prioritization, or CA configuration."
            )
            if spectral.get("confirmed"):
                conclusion += " Lower modulation/MCS/bit-per-Hz confirms that weaker spectral efficiency is also contributing."
            elif spectral.get("iamBetter"):
                conclusion += (
                    f" IAM modulation, MCS and bit/s/Hz are better than {comparator_name} during the available 5G samples: modulation quality is not the limiting factor."
                )
                if scheduler_not_confirmed:
                    conclusion += (
                        f" Scheduler/PRB allocation limitation should remain a hypothesis: IAM PDSCH slot % and allocation ratio are not worse than {comparator_name}."
                    )
            conclusion_fr += (
                f" Pour la comparaison 5G entre IAM et {comparator_name}, la principale faiblesse d'IAM est une couche NR de moindre capacité combinée à une allocation de ressources NR plus faible. "
                f"IAM utilise {iam_nr_band or 'sa bande NR active'}"
                + (f" PCI {(iam_nr_info or {}).get('pci')} / ARFCN {(iam_nr_info or {}).get('arfcn')}" if iam_nr_info else "")
                + f", alors que {comparator_name} utilise {comparator_nr_band or 'sa bande NR active'}"
                + (f" PCI {(comparator_nr_info or {}).get('pci')} / ARFCN {(comparator_nr_info or {}).get('arfcn')}" if comparator_nr_info else "")
                + f". IAM reçoit aussi {str(abs(prbs_gap)) + '% moins de' if prbs_gap is not None else 'sensiblement moins de'} PRBs PDSCH"
                + (f" et dispose de {abs(sched_gap)}% de débit 5G planifié en moins" if sched_gap is not None else "")
                + ". Il ne faut pas décrire cela comme une simple famine scheduler ; la première hypothèse est une différence de bande/BWP/capacité NR, puis la politique scheduler, la charge, la priorisation QoS/APN ou la configuration CA."
            )
            if spectral.get("confirmed"):
                conclusion_fr += " Une modulation/MCS/bit/s/Hz plus faible confirme qu'une efficacité spectrale plus faible contribue aussi à l'écart."
            elif spectral.get("iamBetter"):
                conclusion_fr += (
                    f" La modulation, le MCS et le bit/s/Hz IAM sont meilleurs que {comparator_name} : la qualité de modulation n'est pas le facteur limitant."
                )
                if scheduler_not_confirmed:
                    conclusion_fr += (
                        f" L'allocation PRBs / scheduler doit rester une hypothèse : le PDSCH slot % et le taux d'allocation IAM ne sont pas inférieurs à {comparator_name}."
                    )
        else:
            conclusion += (
                f" For the 5G comparison between IAM and {comparator_name}, IAM's main weakness is scheduler / NR resource allocation. "
                f"IAM receives {str(abs(prbs_gap)) + '% fewer' if prbs_gap is not None else 'materially fewer'} PDSCH PRBs"
                + (f" and has {abs(sched_gap)}% lower scheduled 5G throughput" if sched_gap is not None else "")
                + f" than {comparator_name}. This indicates that IAM's lower DL performance is mainly caused by fewer allocated NR downlink resources, "
                  "lower n78 exposure, possible CA/resource configuration differences, cell load, or QoS/APN prioritization, rather than a simple coverage issue."
            )
            if spectral.get("confirmed"):
                conclusion += " Lower modulation/MCS/bit-per-Hz shows the gap is not purely scheduler-driven."
            elif spectral.get("iamBetter"):
                conclusion += (
                    f" IAM modulation, MCS and bit/s/Hz are better than {comparator_name}: the gap is not explained by poor spectral efficiency."
                )
                if scheduler_not_confirmed:
                    conclusion += (
                        f" Scheduler/PRB allocation should remain a hypothesis: IAM PDSCH slot % and allocation ratio are not worse than {comparator_name}."
                    )
            conclusion_fr += (
                f" Pour la comparaison 5G entre IAM et {comparator_name}, la principale faiblesse d'IAM est le scheduler / l'allocation des ressources NR. "
                f"IAM reçoit {str(abs(prbs_gap)) + '% moins de' if prbs_gap is not None else 'sensiblement moins de'} PRBs PDSCH"
                + (f" et dispose de {abs(sched_gap)}% de débit 5G planifié en moins" if sched_gap is not None else "")
                + f" que {comparator_name}. Cela indique que les performances DL plus faibles d'IAM sont principalement dues à moins de ressources NR downlink allouées, "
                  "une exposition n78 réduite, de possibles différences de configuration CA/ressources, la charge cellulaire ou la priorisation QoS/APN, plutôt qu'un simple problème de couverture."
            )
            if spectral.get("confirmed"):
                conclusion_fr += " Une modulation/MCS/bit/s/Hz plus faible montre que l'écart n'est pas dû uniquement au scheduler."
            elif spectral.get("iamBetter"):
                conclusion_fr += (
                    f" La modulation, le MCS et le bit/s/Hz IAM sont meilleurs que {comparator_name} : l'écart n'est pas dû à une mauvaise efficacité spectrale."
                )
                if scheduler_not_confirmed:
                    conclusion_fr += (
                        f" L'allocation PRBs / scheduler doit rester une hypothèse : le PDSCH slot % et le taux d'allocation IAM ne sont pas inférieurs à {comparator_name}."
                    )

    # Diagnostic confidence level
    confidence = "Low"
    confidence_reason = "Insufficient evidence for high-confidence diagnosis."
    prbs_gap_for_conf = _nemo_gap_pct(iam_prbs_avg, cmp_prbs_avg)
    confidence_label = confidence
    if (prbs_gap_for_conf is not None and prbs_gap_for_conf <= -50
            and sched_gap is not None and sched_gap <= -25
            and rf_good
            and (iam_bler_avg is None or iam_bler_avg < 15)):
        confidence = "High"
        confidence_reason = "PRB gap > 50%, scheduled throughput gap > 25%, RF not degraded, BLER not critical."
    elif (prbs_gap_for_conf is not None and prbs_gap_for_conf <= -25
          and rf_good):
        confidence = "Medium"
        confidence_reason = "Significant PRB gap with acceptable RF. Further network data needed to confirm."
    elif main_cause != "No clear dominant cause":
        confidence = "Medium"
        confidence_reason = "Evidence points to a dominant cause but additional KPIs would strengthen the diagnosis."
    if lower_capacity_band_gap and confidence == "High":
        confidence = "Medium"
        confidence_label = "Medium to High"
        confidence_reason = "Band/capacity difference plus resource gap are clear, but load, BWP width, and QoS data are still needed before calling this purely scheduler-driven."
    elif lower_capacity_band_gap and confidence == "Medium":
        confidence_label = "Medium to High"
    else:
        confidence_label = confidence
    if spectral.get("confirmed") and confidence == "High":
        confidence = "Medium"
        confidence_label = "Medium to High"
        confidence_reason = "Resource-allocation gap is clear, but lower spectral efficiency means the gap is not purely scheduler-driven."
    elif spectral.get("iamBetter") and confidence in ("High", "Medium"):
        # IAM being better on spectral efficiency strengthens the NR resource pool / BWP capacity diagnosis
        confidence_reason = (confidence_reason + " IAM spectral efficiency is better than the comparator, further confirming that the gap is not radio-efficiency driven.").strip()
    elif not spectral.get("sufficient") and confidence == "High":
        confidence = "Medium"
        confidence_label = "Medium"
        confidence_reason = "Resource gap is strong, but modulation/MCS evidence is too sparse for a high-confidence scheduler-only diagnosis."

    # IAM Weakness Index (0–100)
    weakness_index = 0
    if gap_to_best_dl is not None:
        weakness_index += min(25, round(abs(gap_to_best_dl) * 25 / 100))
    if prbs_gap_for_conf is not None:
        weakness_index += min(30, round(abs(prbs_gap_for_conf) * 30 / 100))
    if sched_gap is not None:
        weakness_index += min(20, round(abs(sched_gap) * 20 / 100))
    if delivered_gap is not None:
        weakness_index += min(15, round(abs(delivered_gap) * 15 / 100))
    if rf_good and main_cause == "Scheduler / resource allocation":
        weakness_index += 10  # RF exclusion confirmed = +10
    weakness_index = min(100, weakness_index)

    # 5-point executive summary
    best_dl_op_for_exec = best_dl.get("operator") if best_dl else "N/A"
    iam_avg_for_exec = _nemo_safe_round((iam_kpis.get("dl") or {}).get("average"), 1)
    executive_points = []
    if best_dl:
        if not best_dl.get("has5g"):
            executive_points.append(f"{best_dl_op_for_exec} ranks first in measured DL throughput, but no 5G was detected in its export — it is excluded from 5G-specific comparison.")
        else:
            executive_points.append(f"{best_dl_op_for_exec} ranks first in measured DL throughput.")
    if comparator_name and comparator_name != (best_dl.get("operator") if best_dl else ""):
        executive_points.append(f"{comparator_name} is the best valid 5G comparator for IAM.")
    executive_points.append(f"IAM ranks {iam_rank or 'last'} in DL throughput with {iam_avg_for_exec} Mbps average.")
    if rf_good:
        executive_points.append(f"IAM is not RF-limited: RSRP and SINR are comparable to or better than {comparator_name}.")
    else:
        executive_points.append(f"IAM shows RF degradation — RSRP or SINR is weaker than {comparator_name}.")
    if prbs_gap_for_conf is not None and prbs_gap_for_conf < -20:
        executive_points.append(f"IAM's main weakness is NR downlink resource allocation: {abs(prbs_gap_for_conf)}% fewer PDSCH PRBs vs {comparator_name}.")
    else:
        executive_points.append(f"IAM's main suspected root cause: {main_cause} (severity: {severity}).")

    # What the analysis proves / does not prove
    proves = []
    does_not_prove = []
    if prbs_gap_for_conf is not None and prbs_gap_for_conf < -20:
        proves.append(f"IAM receives {abs(prbs_gap_for_conf)}% fewer PDSCH PRBs than {comparator_name}.")
    if sched_gap is not None and sched_gap < -20:
        proves.append(f"IAM has {abs(sched_gap)}% lower scheduled 5G capacity than {comparator_name}.")
    if delivered_gap is not None and delivered_gap < -20:
        proves.append(f"IAM delivers {abs(delivered_gap)}% lower 5G PDSCH throughput than {comparator_name}.")
    if rf_good:
        proves.append(f"IAM RF quality (RSRP, SINR) is not worse than {comparator_name}.")
    if iam_bler_avg is not None and iam_bler_avg < 10:
        proves.append("IAM BLER is not critically high — retransmissions are not the primary root cause.")
    does_not_prove.append("Cell congestion by itself — network load counters from the eNB/gNB are needed to confirm.")
    does_not_prove.append("Scheduler policy differences — network configuration data is needed.")
    does_not_prove.append("QoS/SIM priority differences — APN and 5QI configuration data is needed.")
    if best_dl and not best_dl.get("has5g"):
        does_not_prove.append(f"{best_dl_op_for_exec} 5G performance — no 5G was detected in its export.")

    # Ranked root-cause hypotheses
    hypotheses = []
    if main_cause == "Scheduler / resource allocation":
        hypotheses = [
            {"rank": 1, "hypothesis": "Cell congestion / high load on IAM serving cell", "probability": "High", "justification": "Low PRBs allocated despite good RF quality."},
            {"rank": 2, "hypothesis": "Scheduler policy or QoS/SIM priority difference", "probability": "High", "justification": "Low allocated resources despite good PRB efficiency."},
            {"rank": 3, "hypothesis": "Lower n78 exposure or narrower NR bandwidth (BWP)", "probability": "Medium", "justification": (
                f"IAM was not observed on n78 in this export; its detected NR layer was {iam_nr_band or 'unknown'} only, "
                f"while {comparator_name} used {comparator_nr_band or 'n78'}. This materially reduces IAM's 5G capacity potential."
                if lower_capacity_band_gap and (iam_n78 is None or iam_n78 == 0.0)
                else f"IAM spends less time on n78 than {comparator_name}."
            )},
            {"rank": 4, "hypothesis": "CA / SCells / BWP configuration difference", "probability": "Medium", "justification": "Needs NR CA/BWP confirmation from network data."},
            {"rank": 5, "hypothesis": "RF coverage issue", "probability": "Low", "justification": "IAM RSRP/SINR are not materially degraded."},
            {"rank": 6, "hypothesis": "BLER / retransmission issue", "probability": "Low", "justification": "IAM BLER is within acceptable range."},
        ]
    elif main_cause == "Coverage limitation":
        hypotheses = [
            {"rank": 1, "hypothesis": "Weak RF coverage on IAM serving cells", "probability": "High", "justification": "IAM RSRP is materially lower than comparator."},
            {"rank": 2, "hypothesis": "Cell edge or coverage gap in benchmark route", "probability": "Medium", "justification": "Low RSRP lower-tail values observed."},
            {"rank": 3, "hypothesis": "Scheduler / resource allocation as secondary factor", "probability": "Low", "justification": "May coexist with coverage issue."},
        ]
    elif main_cause == "Radio quality / interference":
        hypotheses = [
            {"rank": 1, "hypothesis": "Interference on IAM NR carrier", "probability": "High", "justification": "IAM SINR is materially lower despite acceptable RSRP."},
            {"rank": 2, "hypothesis": "Beam management or dominance issue", "probability": "Medium", "justification": "Poor SINR can indicate beam handover failure."},
            {"rank": 3, "hypothesis": "Coverage as contributing factor", "probability": "Low", "justification": "RSRP is not critically degraded."},
        ]
    else:
        hypotheses = [
            {"rank": 1, "hypothesis": main_cause, "probability": "Medium", "justification": "Identified as the highest-scoring root cause."},
        ]

    # Statistical robustness
    test_counts = [len(op.get("tests") or []) for op in operators if str(op.get("operator") or "").upper() == "IAM"]
    iam_test_count = test_counts[0] if test_counts else 0
    has_gps = any(r.get("lat") is not None for r in (iam.get("rows") or []))
    has5g_all = all(op.get("has5g") for op in operators)
    robustness = {
        "testCount": iam_test_count,
        "testCountAdequate": iam_test_count >= 3,
        "operatorsCompared": len(operators),
        "all5gDetected": has5g_all,
        "gpsAvailable": has_gps,
        "diagnosisRobust": confidence in ("High", "Medium"),
        "warning": ("Only one test session detected per operator. Results are technically useful for diagnosing this scenario, "
                    "but are not statistically representative of overall network performance.") if iam_test_count <= 1 else None,
    }

    return {
        "available": True,
        "targetOperator": "IAM",
        "comparator": comparator_name if comparator else "",
        "bestDlOperator": best_dl.get("operator") if best_dl else "",
        "best5gComparator": best_5g.get("operator") if best_5g else "",
        "iamRank": iam_rank,
        "iam5gRank": iam_rank_5g,
        "gapToBestPct": gap_to_best,
        "gapToBestDlPct": gap_to_best_dl,
        "mainCause": main_cause,
        "displayMainCause": display_main_cause,
        "mainScore": main_score,
        "severity": severity,
        "diagnosisType": diagnosis_type,
        "scores": ranked_scores,
        "evidence": evidence[:8],
        "evidenceRows": evidence_rows,
        "rankingInterpretation": ranking_interpretation,
        "rankingInterpretation_fr": ranking_interpretation_fr,
        "technicalInterpretation": technical_interpretation,
        "technicalInterpretation_fr": technical_interpretation_fr,
        "chartNote": "The DL throughput ranking includes all operators, even if 5G was not detected. Operators marked as No 5G detected remain valid for throughput ranking but are excluded from 5G-specific radio comparison.",
        "recommendations": _nemo_recommendations_for_cause(main_cause, comparator_name if comparator_full and comparator_full.get('has5g') else "the best comparator"),
        "conclusion": conclusion,
        "conclusion_fr": conclusion_fr,
        "summary": summary,
        "summary_fr": summary_fr,
        "confidence": confidence,
        "confidenceLabel": confidence_label,
        "confidence_reason": confidence_reason,
        "weaknessIndex": weakness_index,
        "executivePoints": executive_points,
        "proves": proves,
        "doesNotProve": does_not_prove,
        "hypotheses": hypotheses,
        "robustness": robustness,
    }


def _nemo_build_5g_comparator_analysis(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    iam = next((item for item in operators if str(item.get("operator") or "").upper() == "IAM"), None)
    comparator = next((item for item in operators if str(item.get("operator") or "") == comparator_name), None)
    if not iam or not comparator_name or not comparator:
        return {
            "title": "IAM vs Best 5G Comparator",
            "available": False,
            "message": "No 5G comparator available.",
        }

    iam_kpis = iam.get("kpis") or {}
    cmp_kpis = comparator.get("kpis") or {}

    def _avg(section_key):
        return (iam_kpis.get(section_key) or {}).get("average"), (cmp_kpis.get(section_key) or {}).get("average")

    def _median(section_key):
        return (iam_kpis.get(section_key) or {}).get("median"), (cmp_kpis.get(section_key) or {}).get("median")

    rows = []

    def add_pct_row(area, iam_value, cmp_value, unit, interpretation_worse, interpretation_better="", higher_is_better=True):
        gap = _nemo_gap_pct(iam_value, cmp_value)
        worse = gap is not None and ((higher_is_better and gap < 0) or ((not higher_is_better) and gap > 0))
        severity = _nemo_gap_severity(gap) if worse else ("Low" if gap not in (None, 0) else "—")
        interpretation = interpretation_worse if worse else (interpretation_better or "No material weakness detected against comparator")
        rows.append({
            "area": area,
            "iamValue": _nemo_safe_round(iam_value, 2),
            "comparatorValue": _nemo_safe_round(cmp_value, 2),
            "comparatorOperator": comparator_name,
            "gap": gap,
            "gapLabel": None if gap is None else f"{round(gap, 1)}%",
            "severity": severity,
            "interpretation": interpretation,
            "unit": unit,
        })

    def add_delta_row(area, iam_value, cmp_value, unit, stronger_comment, weaker_comment):
        if iam_value is None or cmp_value is None:
            rows.append({
                "area": area,
                "iamValue": _nemo_safe_round(iam_value, 2),
                "comparatorValue": _nemo_safe_round(cmp_value, 2),
                "comparatorOperator": comparator_name,
                "gap": None,
                "gapLabel": "N/A",
                "severity": "—",
                "interpretation": "Metric unavailable for direct comparison",
                "unit": unit,
            })
            return
        delta = round(float(iam_value) - float(cmp_value), 2)
        if delta > 0:
            interpretation = stronger_comment
            severity = "—"
        elif delta < 0:
            interpretation = weaker_comment
            severity = "Medium" if abs(delta) > 2 else "Low"
        else:
            interpretation = "Comparable to comparator"
            severity = "—"
        rows.append({
            "area": area,
            "iamValue": _nemo_safe_round(iam_value, 2),
            "comparatorValue": _nemo_safe_round(cmp_value, 2),
            "comparatorOperator": comparator_name,
            "gap": delta,
            "gapLabel": f"{delta:+.2f} {unit}".strip(),
            "severity": severity,
            "interpretation": interpretation,
            "unit": unit,
        })

    prbs_iam, prbs_cmp = _avg("prbs")
    slot_pct_iam, slot_pct_cmp = _avg("pdschSlotPct")
    sched_iam, sched_cmp = _avg("scheduled5g")
    pdsch_iam, pdsch_cmp = _avg("pdsch5g")
    mac5g_iam, mac5g_cmp = _avg("mac5g")
    rsrp_iam, rsrp_cmp = _median("rsrp")
    sinr_iam, sinr_cmp = _median("sinr")
    cqi_iam, cqi_cmp = _median("cqi")
    ri_iam, ri_cmp = iam_kpis.get("riGe3Share"), cmp_kpis.get("riGe3Share")
    bler_iam, bler_cmp = _avg("bler")
    n78_iam, n78_cmp = iam_kpis.get("n78Share"), cmp_kpis.get("n78Share")

    nr_pres_iam = iam_kpis.get("nrPresencePct")
    nr_pres_cmp = cmp_kpis.get("nrPresencePct")
    lte_only_iam = iam_kpis.get("lteOnlyPresencePct")
    lte_only_cmp = cmp_kpis.get("lteOnlyPresencePct")

    add_pct_row("5G presence % (time-based)", nr_pres_iam, nr_pres_cmp, "%", f"IAM spends less time on 5G/EN-DC than {comparator_name}. A lower 5G presence % directly limits achievable throughput.", higher_is_better=True)
    add_pct_row("4G-only presence % (time-based)", lte_only_iam, lte_only_cmp, "%", f"IAM spends more time on LTE-only than {comparator_name}, indicating less 5G anchoring.", higher_is_better=False)
    add_pct_row("PDSCH PRBs avg", prbs_iam, prbs_cmp, "", f"IAM receives far fewer NR downlink PRBs than {comparator_name}.")

    # PDSCH slot % row with combined PRB+slot interpretation
    _prb_gap_pct = _nemo_gap_pct(prbs_iam, prbs_cmp)
    _slot_gap_pct = _nemo_gap_pct(slot_pct_iam, slot_pct_cmp)
    _iam_low_prbs = _prb_gap_pct is not None and _prb_gap_pct < -15
    _iam_low_slot = _slot_gap_pct is not None and _slot_gap_pct < -15
    if slot_pct_iam is None and slot_pct_cmp is None:
        _slot_interp = ("PDSCH slot % not available in this export — add 'PDSCH scheduling ratio' to the Nemo "
                        "profile for time-domain vs frequency-domain diagnosis.")
    elif _iam_low_prbs and _iam_low_slot:
        _slot_interp = (f"Low PDSCH slot % and low PRBs together suggest time-domain underutilisation — "
                        f"IAM's scheduler allocates PDSCH in fewer slots than {comparator_name}, "
                        f"compressing both time and frequency resources.")
    elif not _iam_low_slot and _iam_low_prbs:
        _slot_interp = (f"IAM PDSCH slot % is comparable to {comparator_name}, but PRBs are significantly lower — "
                        f"this points to a frequency-domain (BWP/bandwidth) limitation rather than a time-domain scheduling deficit.")
    elif _iam_low_slot and not _iam_low_prbs:
        _slot_interp = (f"IAM PDSCH slot % is lower than {comparator_name} but PRBs are comparable — "
                        f"this may indicate lower UE demand or partial-slot scheduling rather than a bandwidth bottleneck.")
    else:
        _slot_interp = (f"IAM PDSCH slot % and PRBs are both comparable to {comparator_name} — "
                        f"throughput gap is unlikely to be caused by resource allocation alone.")
    add_pct_row("PDSCH slot % (scheduling ratio) avg", slot_pct_iam, slot_pct_cmp, "%", _slot_interp, _slot_interp, higher_is_better=True)

    add_pct_row("PDSCH DL scheduled throughput (5G) avg", sched_iam, sched_cmp, "Mbps", f"IAM has lower scheduled 5G capacity than {comparator_name}.")
    add_pct_row("PDSCH DL throughput (5G) avg", pdsch_iam, pdsch_cmp, "Mbps", f"IAM delivers lower 5G PDSCH throughput than {comparator_name}.")
    add_pct_row("MAC DL throughput (5G) avg", mac5g_iam, mac5g_cmp, "Mbps", f"IAM has lower MAC DL throughput (5G) than {comparator_name}.")
    add_delta_row("RSRP median", rsrp_iam, rsrp_cmp, "dB", f"IAM serving RSRP is not worse than {comparator_name}.", f"IAM serving RSRP is weaker than {comparator_name}.")
    add_delta_row("SINR median", sinr_iam, sinr_cmp, "dB", f"IAM serving SINR is not worse than {comparator_name}.", f"IAM serving SINR is weaker than {comparator_name}.")
    add_delta_row("WB CQI median", cqi_iam, cqi_cmp, "", f"IAM CQI is comparable to or better than {comparator_name}.", f"IAM CQI is lower than {comparator_name}.")
    add_pct_row("RI >= 3 share", ri_iam, ri_cmp, "%", f"IAM uses high-rank MIMO less than {comparator_name}.", higher_is_better=True)
    add_pct_row("MAC DL BLER avg", bler_iam, bler_cmp, "%", f"IAM BLER is higher than {comparator_name}.", f"IAM BLER is not worse than {comparator_name}.", higher_is_better=False)
    add_pct_row("n78 share", n78_iam, n78_cmp, "%", f"IAM spends less time on n78 than {comparator_name}.")

    return {
        "title": "IAM vs Best 5G Comparator",
        "available": True,
        "comparatorOperator": comparator_name,
        "rows": rows,
    }


def _nemo_build_weakness_evidence_chain(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    iam = next((item for item in operators if str(item.get("operator") or "").upper() == "IAM"), None)
    comparator = next((item for item in operators if str(item.get("operator") or "") == comparator_name), None)
    if not iam or not comparator_name or not comparator:
        return {"title": "IAM Weakness Evidence Chain", "available": False, "message": "No 5G comparator available."}
    iam_kpis = iam.get("kpis") or {}
    cmp_kpis = comparator.get("kpis") or {}
    spectral = _nemo_spectral_efficiency_insight(iam_kpis, cmp_kpis)

    def chain_step(label, iam_val, cmp_val, unit="", fmt=1):
        gap = _nemo_gap_pct(iam_val, cmp_val)
        if gap is None or gap >= 0:
            return None
        iam_str = f"{_nemo_safe_round(iam_val, fmt)}{(' ' + unit) if unit else ''}" if iam_val is not None else "N/A"
        cmp_str = f"{_nemo_safe_round(cmp_val, fmt)}{(' ' + unit) if unit else ''}" if cmp_val is not None else "N/A"
        return {
            "label": label,
            "iamValue": iam_str,
            "cmpValue": cmp_str,
            "gap": round(gap, 1),
            "severity": _nemo_gap_severity(gap),
            "text": f"{label}: IAM {iam_str} vs {comparator_name} {cmp_str} ({round(gap, 1)}%)",
        }

    chain_links = [
        chain_step("Available NR bandwidth (PRBs)", (iam_kpis.get("availableBandwidthPrbs") or {}).get("average"), (cmp_kpis.get("availableBandwidthPrbs") or {}).get("average"), "", 0),
        chain_step("PDSCH PRBs allocated", (iam_kpis.get("prbs") or {}).get("average"), (cmp_kpis.get("prbs") or {}).get("average"), "", 1),
        chain_step(
            "PDSCH modulation / MCS / spectral efficiency",
            (iam_kpis.get("pdschBitPerHz") or {}).get("median") or (iam_kpis.get("pdschMcs") or {}).get("median"),
            (cmp_kpis.get("pdschBitPerHz") or {}).get("median") or (cmp_kpis.get("pdschMcs") or {}).get("median"),
            "bit/s/Hz" if (iam_kpis.get("pdschBitPerHz") or {}).get("median") is not None or (cmp_kpis.get("pdschBitPerHz") or {}).get("median") is not None else "MCS",
            1,
        ) if spectral.get("confirmed") or spectral.get("comparable") else None,
        chain_step("Scheduled 5G throughput", (iam_kpis.get("scheduled5g") or {}).get("average"), (cmp_kpis.get("scheduled5g") or {}).get("average"), "Mbps", 1),
        chain_step("Delivered 5G PDSCH throughput", (iam_kpis.get("pdsch5g") or {}).get("average"), (cmp_kpis.get("pdsch5g") or {}).get("average"), "Mbps", 1),
        chain_step("Application DL throughput", (iam_kpis.get("dl") or {}).get("average"), (cmp_kpis.get("dl") or {}).get("average"), "Mbps", 1),
    ]
    steps = [link for link in chain_links if link is not None]

    # PRB efficiency insight
    iam_prb_eff = iam_kpis.get("prbEfficiency")
    cmp_prb_eff = cmp_kpis.get("prbEfficiency")
    prb_eff_note = None
    if iam_prb_eff is not None and cmp_prb_eff is not None:
        if iam_prb_eff > cmp_prb_eff * 1.05:
            prb_eff_note = (f"Observed PDSCH Mbps per allocated PRB is higher for IAM ({round(iam_prb_eff, 2)} Mbps/PRB) "
                            f"than {comparator_name} ({round(cmp_prb_eff, 2)} Mbps/PRB), suggesting spectral efficiency "
                            f"is not the main limitation — resource allocation is more likely the constraint. "
                            f"Note: PRB efficiency depends on band, BWP width, MCS, rank, numerology, and sample alignment.")
        elif iam_prb_eff < cmp_prb_eff * 0.90:
            prb_eff_note = (f"Observed PDSCH Mbps per allocated PRB is lower for IAM ({round(iam_prb_eff, 2)} Mbps/PRB) "
                            f"than {comparator_name} ({round(cmp_prb_eff, 2)} Mbps/PRB), suggesting both allocation and "
                            f"spectral efficiency contribute to the gap. "
                            f"Note: PRB efficiency depends on band, BWP width, MCS, rank, numerology, and sample alignment.")
    if spectral.get("note"):
        prb_eff_note = ((prb_eff_note + " ") if prb_eff_note else "") + spectral.get("note")

    # Chain interpretation
    prbs_link = next((s for s in steps if "PRBs allocated" in s["label"]), None)
    delivered_link = next((s for s in steps if "PDSCH throughput" in s["label"]), None)
    chain_interpretation = ""
    if prbs_link and delivered_link and abs(prbs_link["gap"]) > abs(delivered_link["gap"]) + 10:
        chain_interpretation = (f"The PRB allocation gap ({abs(prbs_link['gap'])}%) is larger than the delivered throughput gap "
                                f"({abs(delivered_link['gap'])}%), suggesting IAM is primarily resource-allocation limited.")
    elif prbs_link and delivered_link and abs(delivered_link["gap"]) > abs(prbs_link["gap"]) + 10:
        chain_interpretation = (f"The delivered throughput gap ({abs(delivered_link['gap'])}%) exceeds the PRB allocation gap "
                                f"({abs(prbs_link['gap'])}%), suggesting radio efficiency/BLER also contributes.")

    return {
        "title": "IAM Weakness Evidence Chain",
        "available": bool(steps),
        "steps": steps,
        "chainText": " → ".join(s["text"] for s in steps) if steps else "No clear weakness chain available.",
        "prbEfficiencyNote": prb_eff_note,
        "chainInterpretation": chain_interpretation,
    }


def _nemo_build_rf_exclusion_check(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    iam = next((item for item in operators if str(item.get("operator") or "").upper() == "IAM"), None)
    comparator = next((item for item in operators if str(item.get("operator") or "") == comparator_name), None)
    if not iam or not comparator_name or not comparator:
        return {"title": "RF Exclusion Check", "available": False, "message": "No 5G comparator available."}

    iam_kpis = iam.get("kpis") or {}
    cmp_kpis = comparator.get("kpis") or {}

    def med(section):
        return (iam_kpis.get(section) or {}).get("median"), (cmp_kpis.get(section) or {}).get("median")

    rsrp_iam, rsrp_cmp = med("rsrp")
    sinr_iam, sinr_cmp = med("sinr")
    cqi_iam, cqi_cmp = med("cqi")
    mcs_iam = (iam_kpis.get("pdschMcs") or {}).get("median")
    mcs_cmp = (cmp_kpis.get("pdschMcs") or {}).get("median")
    se_iam = (iam_kpis.get("pdschBitPerHz") or {}).get("median")
    se_cmp = (cmp_kpis.get("pdschBitPerHz") or {}).get("median")
    rows = []

    def add_row(metric, iam_value, cmp_value, unit, comment_logic):
        comment, result = comment_logic(iam_value, cmp_value)
        rows.append({
            "metric": metric,
            "iam": _nemo_safe_round(iam_value, 2),
            "comparator": _nemo_safe_round(cmp_value, 2),
            "result": result,
            "comment": comment,
            "unit": unit,
        })

    add_row("Median RSRP", rsrp_iam, rsrp_cmp, "dBm", lambda a, b: (
        ("Metric unavailable", "N/A") if a is None or b is None else
        ("IAM RSRP is not weaker than comparator", "Not worse") if a >= b - 3 else
        ("IAM RSRP is materially weaker than comparator", "Worse")
    ))
    add_row("Median SINR", sinr_iam, sinr_cmp, "dB", lambda a, b: (
        ("Metric unavailable", "N/A") if a is None or b is None else
        ("IAM SINR is not weaker than comparator", "Not worse") if a >= b - 2 else
        ("IAM SINR is materially weaker than comparator", "Worse")
    ))
    add_row("Median CQI", cqi_iam, cqi_cmp, "", lambda a, b: (
        ("Metric unavailable", "N/A") if a is None or b is None else
        ("CQI is comparable between operators", "Comparable") if abs(a - b) <= 1 else
        ("IAM CQI is lower than comparator", "Lower") if a < b else
        ("IAM CQI is higher than comparator", "Higher")
    ))
    add_row("Median MCS", mcs_iam, mcs_cmp, "", lambda a, b: (
        ("Metric unavailable", "N/A") if a is None or b is None else
        ("IAM MCS is comparable to comparator", "Comparable") if abs(a - b) <= 2 else
        ("IAM MCS is lower than comparator", "Lower") if a < b else
        ("IAM MCS is higher than comparator", "Higher")
    ))
    add_row("Median PDSCH bit/s/Hz", se_iam, se_cmp, "bit/s/Hz", lambda a, b: (
        ("Metric unavailable", "N/A") if a is None or b is None else
        ("PDSCH spectral efficiency is comparable", "Comparable") if abs(a - b) <= max(0.2, abs(b) * 0.2) else
        ("IAM PDSCH spectral efficiency is lower than comparator", "Lower") if a < b else
        ("IAM PDSCH spectral efficiency is higher than comparator", "Higher")
    ))

    # Derive PRB gap for contextual conclusion
    iam_prbs_excl = (iam_kpis.get("prbs") or {}).get("average")
    cmp_prbs_excl = (cmp_kpis.get("prbs") or {}).get("average")
    prbs_gap_excl = _nemo_gap_pct(iam_prbs_excl, cmp_prbs_excl)
    iam_prb_eff_excl = iam_kpis.get("prbEfficiency")
    cmp_prb_eff_excl = cmp_kpis.get("prbEfficiency")

    conclusion = "RF may contribute but is not clearly dominant."
    if rsrp_iam is not None and rsrp_cmp is not None and sinr_iam is not None and sinr_cmp is not None:
        if rsrp_iam >= rsrp_cmp - 3 and sinr_iam >= sinr_cmp - 2:
            if prbs_gap_excl is not None and prbs_gap_excl <= -30:
                conclusion = (f"RF quality is not the primary limitation. IAM RSRP and SINR are comparable to {comparator_name}, "
                              f"yet IAM receives {abs(prbs_gap_excl)}% fewer PDSCH PRBs. "
                              f"This confirms the weakness is in scheduler / resource allocation, not coverage or signal quality.")
                if iam_prb_eff_excl is not None and cmp_prb_eff_excl is not None and iam_prb_eff_excl > cmp_prb_eff_excl:
                    conclusion += (f" IAM also has higher PRB efficiency ({round(iam_prb_eff_excl, 2)} vs "
                                   f"{round(cmp_prb_eff_excl, 2)} Mbps/PRB), meaning IAM radio decoding is good — "
                                   f"the UE is simply not allocated enough resources.")
            else:
                conclusion = f"RF is not the primary suspected limitation — IAM RSRP and SINR are comparable to {comparator_name}."
        elif rsrp_iam < rsrp_cmp - 6:
            conclusion = "Coverage appears to be a primary limitation — IAM RSRP is materially weaker."
        elif sinr_iam < sinr_cmp - 5:
            conclusion = "Radio quality / interference appears to be a primary limitation — IAM SINR is materially lower."

    return {
        "title": "RF Exclusion Check",
        "available": True,
        "comparatorOperator": comparator_name,
        "rows": rows,
        "conclusion": conclusion,
    }


def _nemo_build_scheduler_prb_deep_dive(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    iam = next((item for item in operators if str(item.get("operator") or "").upper() == "IAM"), None)
    comparator = next((item for item in operators if str(item.get("operator") or "") == comparator_name), None)
    if not iam or not comparator_name or not comparator:
        return {"title": "Scheduler / PRB Deep Dive", "available": False, "message": "No 5G comparator available."}
    iam_kpis = iam.get("kpis") or {}
    cmp_kpis = comparator.get("kpis") or {}

    def stat_row(area, iam_value, cmp_value, unit, use_pct=True, higher_is_better=True):
        gap = _nemo_gap_pct(iam_value, cmp_value) if use_pct else None
        return {
            "area": area,
            "iamValue": _nemo_safe_round(iam_value, 2),
            "comparatorValue": _nemo_safe_round(cmp_value, 2),
            "comparatorOperator": comparator_name,
            "gap": gap,
            "gapLabel": ("—" if gap is None else f"{round(gap, 1)}%"),
            "severity": _nemo_gap_severity(gap) if gap is not None and ((higher_is_better and gap < 0) or ((not higher_is_better) and gap > 0)) else "—",
            "unit": unit,
        }

    prbs_iam = iam_kpis.get("prbs") or {}
    prbs_cmp = cmp_kpis.get("prbs") or {}
    slot_iam = iam_kpis.get("pdschSlotPct") or {}
    slot_cmp = cmp_kpis.get("pdschSlotPct") or {}
    sched_iam = iam_kpis.get("scheduled5g") or {}
    sched_cmp = cmp_kpis.get("scheduled5g") or {}
    pdsch_iam = iam_kpis.get("pdsch5g") or {}
    pdsch_cmp = cmp_kpis.get("pdsch5g") or {}
    avail_iam = iam_kpis.get("availableBandwidthPrbs") or {}
    avail_cmp = cmp_kpis.get("availableBandwidthPrbs") or {}

    # Use pre-computed derived KPIs from _nemo_operator_kpis
    prb_eff_iam = iam_kpis.get("prbEfficiency")
    prb_eff_cmp = cmp_kpis.get("prbEfficiency")
    sched_eff_iam = iam_kpis.get("scheduledEfficiency")
    sched_eff_cmp = cmp_kpis.get("scheduledEfficiency")
    res_alloc_iam = iam_kpis.get("resourceAllocationIndex")
    res_alloc_cmp = cmp_kpis.get("resourceAllocationIndex")
    prbs_per_slot_iam = iam_kpis.get("prbsPerScheduledSlot")
    prbs_per_slot_cmp = cmp_kpis.get("prbsPerScheduledSlot")
    sched_per_slot_iam = iam_kpis.get("scheduledMbpsPerSlot")
    sched_per_slot_cmp = cmp_kpis.get("scheduledMbpsPerSlot")
    spectral = _nemo_spectral_efficiency_insight(iam_kpis, cmp_kpis)

    rows = [
        stat_row("Available NR bandwidth (avg PRBs)", avail_iam.get("average"), avail_cmp.get("average"), "PRBs"),
        stat_row("PDSCH slot % (avg)", slot_iam.get("average"), slot_cmp.get("average"), "%"),
        stat_row("Resource allocation index (PDSCH PRBs / available)", res_alloc_iam, res_alloc_cmp, "%"),
        stat_row("Avg PDSCH PRBs allocated", prbs_iam.get("average"), prbs_cmp.get("average"), ""),
        stat_row("PRBs per scheduled slot", prbs_per_slot_iam, prbs_per_slot_cmp, ""),
        stat_row("Median PDSCH PRBs", prbs_iam.get("median"), prbs_cmp.get("median"), ""),
        stat_row("P10 PDSCH PRBs", prbs_iam.get("p10"), prbs_cmp.get("p10"), ""),
        stat_row("P90 PDSCH PRBs", prbs_iam.get("p90"), prbs_cmp.get("p90"), ""),
        stat_row("Avg scheduled 5G throughput", sched_iam.get("average"), sched_cmp.get("average"), "Mbps"),
        stat_row("Scheduled Mbps per slot", sched_per_slot_iam, sched_per_slot_cmp, "Mbps"),
        stat_row("Median scheduled 5G throughput", sched_iam.get("median"), sched_cmp.get("median"), "Mbps"),
        stat_row("Avg delivered 5G PDSCH throughput", pdsch_iam.get("average"), pdsch_cmp.get("average"), "Mbps"),
        stat_row("Median PDSCH MCS", (iam_kpis.get("pdschMcs") or {}).get("median"), (cmp_kpis.get("pdschMcs") or {}).get("median"), ""),
        stat_row("Median PDSCH bit/s/Hz", (iam_kpis.get("pdschBitPerHz") or {}).get("median"), (cmp_kpis.get("pdschBitPerHz") or {}).get("median"), "bit/s/Hz"),
        stat_row("PRB efficiency (PDSCH Mbps / PRB)", prb_eff_iam, prb_eff_cmp, "Mbps/PRB"),
        stat_row("Scheduled-to-delivered efficiency (%)", sched_eff_iam, sched_eff_cmp, "%"),
    ]

    prbs_gap = _nemo_gap_pct(prbs_iam.get("average"), prbs_cmp.get("average"))
    slot_gap = _nemo_gap_pct(slot_iam.get("average"), slot_cmp.get("average"))
    prbs_per_slot_gap = _nemo_gap_pct(prbs_per_slot_iam, prbs_per_slot_cmp)
    sched_gap = _nemo_gap_pct(sched_iam.get("average"), sched_cmp.get("average"))
    sched_per_slot_gap = _nemo_gap_pct(sched_per_slot_iam, sched_per_slot_cmp)
    delivered_gap = _nemo_gap_pct(pdsch_iam.get("average"), pdsch_cmp.get("average"))
    avail_gap = _nemo_gap_pct(avail_iam.get("average"), avail_cmp.get("average"))
    conclusions = []

    def _is_comparable_gap(gap, threshold=10):
        return gap is not None and abs(gap) <= threshold

    step_rows = []

    def add_step_row(kpi, iam_value, cmp_value, unit, interpretation):
        step_rows.append({
            "kpi": kpi,
            "iamValue": _nemo_safe_round(iam_value, 2) if isinstance(iam_value, (int, float)) else iam_value,
            "comparatorValue": _nemo_safe_round(cmp_value, 2) if isinstance(cmp_value, (int, float)) else cmp_value,
            "unit": unit,
            "interpretation": interpretation,
        })

    if slot_iam.get("average") is None or slot_cmp.get("average") is None:
        slot_interpretation = "PDSCH slot % is not available, so time-domain scheduling cannot be isolated from frequency-domain allocation."
    elif _is_comparable_gap(slot_gap, 10):
        slot_interpretation = f"Time-domain scheduling is comparable. IAM is not mainly disadvantaged by fewer scheduled slots than {comparator_name}."
    elif slot_gap is not None and slot_gap <= -15:
        slot_interpretation = f"IAM is scheduled in a lower share of slots than {comparator_name}, pointing to a time-domain scheduling deficit."
    else:
        slot_interpretation = f"IAM slot usage is not lower than {comparator_name}; throughput gap must be explained inside the scheduled slots."
    add_step_row("PDSCH slot %", slot_iam.get("average"), slot_cmp.get("average"), "%", slot_interpretation)

    if avail_iam.get("average") is None or avail_cmp.get("average") is None:
        avail_interpretation = "Available NR PRB pool is not available in this export."
    elif avail_gap is not None and avail_gap <= -20:
        avail_interpretation = f"IAM has a much smaller NR resource pool than {comparator_name}. This points first to band/BWP/carrier-capacity limitation."
    elif avail_gap is not None and abs(avail_gap) <= 10:
        avail_interpretation = f"Available PRB pool is broadly comparable to {comparator_name}; any PRB gap is more likely a scheduler/load/QoS issue."
    else:
        avail_interpretation = f"Available PRB pool differs, but not enough on its own to explain the full gap."
    add_step_row("Available PRBs", avail_iam.get("average"), avail_cmp.get("average"), "PRBs", avail_interpretation)

    if res_alloc_iam is None or res_alloc_cmp is None:
        alloc_interpretation = "Allocation ratio cannot be computed because available PRB pool or allocated PRBs are missing."
    elif avail_gap is not None and avail_gap <= -20 and res_alloc_iam >= 80:
        alloc_interpretation = "IAM uses most of its smaller NR pool. This supports BWP/bandwidth limitation more than pure scheduler starvation."
    elif avail_gap is not None and avail_gap >= -10 and res_alloc_iam < (res_alloc_cmp - 15):
        alloc_interpretation = f"Available pool is similar, but IAM allocates a lower share of it than {comparator_name}. This points to scheduler/load/QoS limitation."
    elif res_alloc_iam >= 80:
        alloc_interpretation = "IAM uses a high share of its available pool; low absolute PRBs mainly reflect pool size rather than under-allocation."
    else:
        alloc_interpretation = "Allocation ratio does not clearly isolate pool-size versus scheduler effects."
    add_step_row("Allocation ratio", res_alloc_iam, res_alloc_cmp, "%", alloc_interpretation)

    if prbs_per_slot_iam is None or prbs_per_slot_cmp is None:
        prbs_slot_interpretation = "PRBs per scheduled slot cannot be computed because PDSCH slot % is missing."
    elif _is_comparable_gap(slot_gap, 10) and prbs_per_slot_gap is not None and prbs_per_slot_gap <= -20:
        prbs_slot_interpretation = f"Slot usage is similar, but IAM carries fewer PRBs per scheduled opportunity than {comparator_name}. This is a frequency-domain limitation."
    elif slot_gap is not None and slot_gap <= -15 and prbs_per_slot_gap is not None and abs(prbs_per_slot_gap) <= 15:
        prbs_slot_interpretation = f"PRBs per scheduled slot are comparable, but IAM is scheduled less often. This points to time-domain scheduling limitation."
    else:
        prbs_slot_interpretation = "PRBs per scheduled slot do not point to a pure time-domain deficit."
    add_step_row("PRBs / scheduled slot", prbs_per_slot_iam, prbs_per_slot_cmp, "", prbs_slot_interpretation)

    if sched_per_slot_iam is None or sched_per_slot_cmp is None:
        sched_slot_interpretation = "Scheduled Mbps per slot cannot be computed because PDSCH slot % is missing."
    elif _is_comparable_gap(slot_gap, 10) and sched_per_slot_gap is not None and sched_per_slot_gap <= -20:
        sched_slot_interpretation = "After normalizing for slot usage, IAM still has much lower scheduled capacity per slot. The deficit is inside the scheduled slots."
    else:
        sched_slot_interpretation = "Scheduled capacity per slot does not show a large deficit after slot normalization."
    add_step_row("Scheduled Mbps / slot", sched_per_slot_iam, sched_per_slot_cmp, "Mbps", sched_slot_interpretation)

    if prb_eff_iam is None or prb_eff_cmp is None:
        prb_eff_interpretation = "PDSCH Mbps per PRB cannot be computed from this export."
    elif prb_eff_iam > prb_eff_cmp * 1.05:
        prb_eff_interpretation = (
            "IAM PRB efficiency is better than the comparator: each allocated PRB delivers more throughput. "
            "The main limitation is the NR resource pool / BWP capacity, not spectral efficiency or scheduler under-allocation."
        )
    elif prb_eff_iam < prb_eff_cmp * 0.9:
        prb_eff_interpretation = "Each IAM PRB carries less delivered throughput. Check modulation, MCS, rank, BLER/HARQ, bit/s/Hz, and sample-window quality."
    else:
        prb_eff_interpretation = "PRB efficiency is comparable; the dominant issue remains PRB pool or allocation."
    add_step_row("PDSCH Mbps / PRB", prb_eff_iam, prb_eff_cmp, "Mbps/PRB", prb_eff_interpretation)

    dom_mod_iam = (iam_kpis.get("pdschModulation") or {}).get("dominant")
    dom_mod_cmp = (cmp_kpis.get("pdschModulation") or {}).get("dominant")
    if not spectral.get("sufficient"):
        mod_interpretation = spectral.get("note") or "Modulation evidence is limited in this export."
    elif spectral.get("confirmed"):
        mod_interpretation = "Lower-order modulation contributes to IAM’s lower spectral efficiency in addition to the smaller NR resource pool."
    elif spectral.get("iamBetter"):
        mod_interpretation = (
            "IAM modulation order is better than the comparator — modulation quality is not a limitation. "
            "The throughput gap is explained by the smaller NR resource pool / BWP capacity, not poor spectral efficiency."
        )
    else:
        mod_interpretation = "Modulation is broadly comparable; the main issue remains BWP/bandwidth or PRB capacity."
    add_step_row("Dominant modulation", dom_mod_iam or "—", dom_mod_cmp or "—", "", mod_interpretation)

    mcs_iam = (iam_kpis.get("pdschMcs") or {}).get("median")
    mcs_cmp = (cmp_kpis.get("pdschMcs") or {}).get("median")
    if not spectral.get("sufficient"):
        mcs_interpretation = spectral.get("note") or "MCS evidence is limited in this export."
    elif spectral.get("confirmed") and mcs_iam is not None and mcs_cmp is not None and mcs_iam <= mcs_cmp - 3:
        mcs_interpretation = "IAM median MCS is lower, confirming a secondary spectral-efficiency limitation."
    elif spectral.get("iamBetter"):
        mcs_interpretation = (
            "IAM median MCS is better than the comparator, confirming that spectral efficiency is not the root cause. "
            "The main confirmed issue is the NR bandwidth / resource pool gap."
        )
    else:
        mcs_interpretation = "MCS does not overturn the main BWP/resource-pool diagnosis."
    add_step_row("Median MCS", mcs_iam, mcs_cmp, "", mcs_interpretation)

    if prbs_gap is not None and prbs_gap <= -50:
        conclusions.append(f"Critical PRB allocation gap ({abs(prbs_gap)}%): IAM receives far fewer downlink resources than {comparator_name}.")
    elif prbs_gap is not None and prbs_gap <= -25:
        conclusions.append(f"Significant PRB allocation gap ({abs(prbs_gap)}%).")
    if sched_gap is not None and sched_gap <= -25:
        conclusions.append(f"Scheduled 5G capacity is {abs(sched_gap)}% lower — IAM starts from a weaker capacity allocation.")

    # Allocation vs efficiency diagnosis
    if prbs_gap is not None and delivered_gap is not None and abs(prbs_gap) > abs(delivered_gap) + 10:
        conclusions.append(f"PRB gap ({abs(prbs_gap)}%) > delivered gap ({abs(delivered_gap)}%): IAM is resource-allocation limited, not radio-efficiency limited.")
    elif prbs_gap is not None and delivered_gap is not None and abs(delivered_gap) > abs(prbs_gap) + 10:
        conclusions.append(f"Delivered gap ({abs(delivered_gap)}%) > PRB gap ({abs(prbs_gap)}%): radio efficiency / BLER also contributes.")

    # PRB efficiency verdict
    if prb_eff_iam is not None and prb_eff_cmp is not None:
        if prb_eff_iam > prb_eff_cmp * 1.05:
            conclusions.append(f"Observed PDSCH Mbps/PRB is higher for IAM ({round(prb_eff_iam, 2)}) than {comparator_name} ({round(prb_eff_cmp, 2)}), suggesting lower spectral efficiency is not the main cause. Confirm using MCS, rank, PDSCH slot %, and bit/s/Hz if available.")
        elif prb_eff_iam < prb_eff_cmp * 0.90:
            conclusions.append(f"Observed PDSCH Mbps/PRB is lower for IAM ({round(prb_eff_iam, 2)}) than {comparator_name} ({round(prb_eff_cmp, 2)}), suggesting both allocation and spectral efficiency contribute to the gap. Confirm using MCS, rank, PDSCH slot %, and bit/s/Hz if available.")
    if spectral.get("confirmed"):
        conclusions.append(spectral.get("note"))
    elif spectral.get("iamBetter"):
        conclusions.append(spectral.get("note"))
    elif not spectral.get("sufficient"):
        conclusions.append(spectral.get("note"))

    # BWP vs scheduler distinction
    if avail_gap is not None and avail_gap >= -10 and prbs_gap is not None and prbs_gap <= -30:
        conclusions.append(f"Available NR bandwidth is similar (gap {avail_gap}%) but PDSCH PRB allocation gap is {abs(prbs_gap)}%: issue is scheduler/load, not bandwidth configuration.")
    elif avail_gap is not None and avail_gap <= -25:
        conclusions.append(f"Available NR bandwidth gap ({abs(avail_gap)}%): IAM may have narrower BWP or lower NR carrier capacity configured.")

    primary_diagnosis = "No clear scheduler/PRB verdict available."
    if slot_gap is not None and slot_gap <= -15 and (prbs_per_slot_gap is None or prbs_per_slot_gap > -15):
        primary_diagnosis = "Primary diagnosis: time-domain scheduling limitation."
    elif avail_gap is not None and avail_gap <= -20 and res_alloc_iam is not None and res_alloc_iam >= 80:
        primary_diagnosis = "Primary diagnosis: BWP / band / NR capacity-pool limitation."
    elif avail_gap is not None and avail_gap >= -10 and res_alloc_iam is not None and res_alloc_cmp is not None and res_alloc_iam < (res_alloc_cmp - 15):
        primary_diagnosis = "Primary diagnosis: scheduler / load / QoS resource-allocation limitation."
    elif prb_eff_iam is not None and prb_eff_cmp is not None and prb_eff_iam < prb_eff_cmp * 0.9:
        primary_diagnosis = "Primary diagnosis: lower spectral efficiency also contributes and should be confirmed with modulation/MCS/bit/s/Hz."

    verdict = conclusions[-1] if conclusions else "No clear scheduler/PRB verdict available."
    if primary_diagnosis != "No clear scheduler/PRB verdict available.":
        verdict = primary_diagnosis + " " + verdict
    # Build a structured allocation chain for the frontend
    allocation_chain = [
        {"step": "Available NR bandwidth", "iamVal": _nemo_safe_round(avail_iam.get("average"), 0), "cmpVal": _nemo_safe_round(avail_cmp.get("average"), 0), "gap": _nemo_gap_pct(avail_iam.get("average"), avail_cmp.get("average")), "unit": "PRBs"},
        {"step": "PDSCH PRBs allocated", "iamVal": _nemo_safe_round(prbs_iam.get("average"), 1), "cmpVal": _nemo_safe_round(prbs_cmp.get("average"), 1), "gap": prbs_gap, "unit": ""},
        {"step": "PDSCH spectral efficiency", "iamVal": _nemo_safe_round((iam_kpis.get("pdschBitPerHz") or {}).get("median") or (iam_kpis.get("pdschMcs") or {}).get("median"), 1), "cmpVal": _nemo_safe_round((cmp_kpis.get("pdschBitPerHz") or {}).get("median") or (cmp_kpis.get("pdschMcs") or {}).get("median"), 1), "gap": _nemo_gap_pct((iam_kpis.get("pdschBitPerHz") or {}).get("median") or (iam_kpis.get("pdschMcs") or {}).get("median"), (cmp_kpis.get("pdschBitPerHz") or {}).get("median") or (cmp_kpis.get("pdschMcs") or {}).get("median")), "unit": "bit/s/Hz" if (iam_kpis.get("pdschBitPerHz") or {}).get("median") is not None or (cmp_kpis.get("pdschBitPerHz") or {}).get("median") is not None else "MCS"},
        {"step": "Scheduled 5G throughput", "iamVal": _nemo_safe_round(sched_iam.get("average"), 1), "cmpVal": _nemo_safe_round(sched_cmp.get("average"), 1), "gap": sched_gap, "unit": "Mbps"},
        {"step": "Delivered 5G PDSCH", "iamVal": _nemo_safe_round(pdsch_iam.get("average"), 1), "cmpVal": _nemo_safe_round(pdsch_cmp.get("average"), 1), "gap": delivered_gap, "unit": "Mbps"},
    ]

    return {
        "title": "Scheduler / PRB Deep Dive",
        "available": True,
        "comparatorOperator": comparator_name,
        "rows": rows,
        "stepRows": step_rows,
        "conclusions": conclusions,
        "verdict": verdict,
        "stepConclusion": primary_diagnosis,
        "allocationChain": allocation_chain,
        "prbEfficiencyIam": prb_eff_iam,
        "prbEfficiencyCmp": prb_eff_cmp,
        "scheduledEfficiencyIam": sched_eff_iam,
        "scheduledEfficiencyCmp": sched_eff_cmp,
    }


def _nemo_evaluate_pdsch_signals(iam_kpis: dict, cmp_kpis: dict) -> dict:
    """Compute the boolean signals used by the PDSCH modulation conclusion generator."""
    iam_avail_prbs = (iam_kpis.get("availableBandwidthPrbs") or {}).get("average")
    cmp_avail_prbs = (cmp_kpis.get("availableBandwidthPrbs") or {}).get("average")
    iam_slot_pct = (iam_kpis.get("pdschSlotPct") or {}).get("average")
    cmp_slot_pct = (cmp_kpis.get("pdschSlotPct") or {}).get("average")
    iam_alloc = iam_kpis.get("resourceAllocationIndex")
    cmp_alloc = cmp_kpis.get("resourceAllocationIndex")
    iam_n78 = float(iam_kpis.get("n78Share") or 0.0)
    cmp_n78 = float(cmp_kpis.get("n78Share") or 0.0)
    iam_prb_eff = iam_kpis.get("prbEfficiency")
    cmp_prb_eff = cmp_kpis.get("prbEfficiency")
    iam_mcs = (iam_kpis.get("pdschMcs") or {}).get("median")
    cmp_mcs = (cmp_kpis.get("pdschMcs") or {}).get("median")
    iam_bits = (iam_kpis.get("pdschBitPerHz") or {}).get("median")
    cmp_bits = (cmp_kpis.get("pdschBitPerHz") or {}).get("median")
    iam_mod = iam_kpis.get("pdschModulation") or {}
    cmp_mod = cmp_kpis.get("pdschModulation") or {}
    iam_qpsk = float(iam_mod.get("qpskShare") or 0.0)
    cmp_qpsk = float(cmp_mod.get("qpskShare") or 0.0)
    iam_qam64 = float(iam_mod.get("qam64Share") or 0.0)
    cmp_qam64 = float(cmp_mod.get("qam64Share") or 0.0)
    iam_qam256 = float(iam_mod.get("qam256Share") or 0.0)
    cmp_qam256 = float(cmp_mod.get("qam256Share") or 0.0)
    iam_qam16 = float(iam_mod.get("qam16Share") or 0.0)
    cmp_qam16 = float(cmp_mod.get("qam16Share") or 0.0)

    spectral_better = bool(
        iam_mcs is not None and cmp_mcs is not None and float(iam_mcs) > float(cmp_mcs)
        and iam_bits is not None and cmp_bits is not None and float(iam_bits) > float(cmp_bits)
        and iam_prb_eff is not None and cmp_prb_eff is not None and float(iam_prb_eff) > float(cmp_prb_eff)
    )
    modulation_better = bool(
        iam_qpsk < cmp_qpsk
        and (iam_qam64 > cmp_qam64 or iam_qam256 > cmp_qam256 or iam_qam16 > cmp_qam16)
    )
    smaller_nr_pool = bool(
        iam_avail_prbs is not None and cmp_avail_prbs is not None
        and float(iam_avail_prbs) < float(cmp_avail_prbs) * 0.5
    )
    missing_n78 = bool(iam_n78 == 0.0 and cmp_n78 > 20.0)
    scheduler_hypothesis = bool(
        iam_slot_pct is not None and cmp_slot_pct is not None and float(iam_slot_pct) >= float(cmp_slot_pct)
        and iam_alloc is not None and cmp_alloc is not None and float(iam_alloc) >= float(cmp_alloc)
        and iam_prb_eff is not None and cmp_prb_eff is not None and float(iam_prb_eff) >= float(cmp_prb_eff)
    )
    return {
        "spectralBetter": spectral_better,
        "modulationBetter": modulation_better,
        "smallerNrPool": smaller_nr_pool,
        "missingN78": missing_n78,
        "schedulerHypothesis": scheduler_hypothesis,
        "iamAvailPrbs": iam_avail_prbs,
        "cmpAvailPrbs": cmp_avail_prbs,
        "iamMcs": iam_mcs,
        "iamBitsHz": iam_bits,
        "iamPrbEff": iam_prb_eff,
        "iamDomMod": iam_mod.get("dominant") or "—",
        "cmpN78": cmp_n78,
    }


def _nemo_generate_pdsch_conclusion(iam_kpis: dict, cmp_kpis: dict, comparator_name: str) -> str:
    """Generate the PDSCH Modulation & Spectral Efficiency conclusion using KPI-driven logic."""
    sig = _nemo_evaluate_pdsch_signals(iam_kpis, cmp_kpis)
    spectral_better = sig["spectralBetter"]
    modulation_better = sig["modulationBetter"]
    smaller_nr_pool = sig["smallerNrPool"]
    missing_n78 = sig["missingN78"]
    scheduler_hypothesis = sig["schedulerHypothesis"]

    parts: list[str] = []

    if spectral_better and modulation_better:
        parts.append(
            f"IAM modulation, MCS and spectral efficiency are better than {comparator_name} during the available 5G samples. "
            f"IAM dominant modulation is {sig['iamDomMod']}, with median MCS {_nemo_safe_round(sig['iamMcs'], 0)}, "
            f"median spectral efficiency {_nemo_safe_round(sig['iamBitsHz'], 2)} bit/s/Hz "
            f"and PRB efficiency {_nemo_safe_round(sig['iamPrbEff'], 3)} Mbps/PRB. "
            f"Therefore, poor modulation quality is not the main limitation versus {comparator_name}."
        )
    elif spectral_better:
        parts.append(
            f"IAM spectral efficiency is better than {comparator_name}, based on higher MCS, bit/s/Hz and PRB efficiency. "
            f"Modulation quality is not the primary suspected limitation."
        )
    else:
        parts.append(
            f"IAM modulation or spectral efficiency is weaker than {comparator_name}. "
            f"Check SINR, CQI, MCS, modulation distribution, BLER and link adaptation."
        )

    if smaller_nr_pool and missing_n78:
        parts.append(
            f"The dominant confirmed issue is the NR capacity pool gap: IAM has only {_nemo_safe_round(sig['iamAvailPrbs'], 0)} available NR PRBs "
            f"versus {_nemo_safe_round(sig['cmpAvailPrbs'], 0)} for {comparator_name}, with 0% n78 contribution. "
            f"This indicates a band/BWP/carrier-capacity limitation, mainly linked to missing n78 contribution."
        )
    elif smaller_nr_pool:
        parts.append(
            f"The dominant confirmed issue is a smaller NR resource pool: IAM has {_nemo_safe_round(sig['iamAvailPrbs'], 0)} available NR PRBs "
            f"versus {_nemo_safe_round(sig['cmpAvailPrbs'], 0)} for {comparator_name}. "
            f"This points to BWP/bandwidth/carrier-capacity limitation."
        )
    elif missing_n78:
        parts.append(
            f"IAM has no n78 contribution while {comparator_name} has {_nemo_safe_round(sig['cmpN78'], 1)}% n78 share. "
            f"This indicates a missing high-capacity 5G layer."
        )

    if scheduler_hypothesis:
        parts.append(
            f"Scheduler or PRB allocation limitation is not confirmed by this table because IAM PDSCH slot usage, "
            f"allocation ratio and PRB efficiency are not worse than {comparator_name}. "
            f"Keep scheduler/PRB allocation as a hypothesis until PRB utilization, scheduler grants, "
            f"BWP configuration, cell load and QoS indicators are available."
        )
    else:
        parts.append(
            f"Scheduler/PRB allocation may require deeper verification using PRB utilization, "
            f"scheduler grants, BWP configuration, cell load and QoS indicators."
        )

    return "\n\n".join(parts)


def _nemo_build_pdsch_modulation_efficiency(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    iam = next((item for item in operators if str(item.get("operator") or "").upper() == "IAM"), None)
    comparator = next((item for item in operators if str(item.get("operator") or "") == comparator_name), None)
    if not iam or not comparator_name or not comparator:
        return {"title": "PDSCH Modulation & Spectral Efficiency", "available": False, "message": "No 5G comparator available."}

    iam_kpis = iam.get("kpis") or {}
    cmp_kpis = comparator.get("kpis") or {}
    spectral = _nemo_spectral_efficiency_insight(iam_kpis, cmp_kpis)

    def _row_for_operator(item, kpis):
        mod = kpis.get("pdschModulation") or {}
        cw0 = mod.get("cw0") or {}
        cw1 = mod.get("cw1") or {}
        return {
            "operator": item.get("operator"),
            "dominantModulation": mod.get("dominant"),
            "cw0Dominant": cw0.get("dominant"),
            "cw1Dominant": cw1.get("dominant"),
            "cw0Samples": cw0.get("sampleCount"),
            "cw1Samples": cw1.get("sampleCount"),
            "rank2UtilPct": kpis.get("rank2UtilPct"),
            "qam256Share": mod.get("qam256Share"),
            "qam64Share": mod.get("qam64Share"),
            "qam16Share": mod.get("qam16Share"),
            "qpskShare": mod.get("qpskShare"),
            "cw0Qam256": cw0.get("qam256Share"),
            "cw0Qam64": cw0.get("qam64Share"),
            "cw0Qam16": cw0.get("qam16Share"),
            "cw0Qpsk": cw0.get("qpskShare"),
            "cw1Qam256": cw1.get("qam256Share"),
            "cw1Qam64": cw1.get("qam64Share"),
            "cw1Qam16": cw1.get("qam16Share"),
            "cw1Qpsk": cw1.get("qpskShare"),
            "medianMcs": (kpis.get("pdschMcs") or {}).get("median"),
            "medianRank": (kpis.get("scheduledRank") or {}).get("median"),
            "medianBitsPerHz": (kpis.get("pdschBitPerHz") or {}).get("median"),
            "prbEfficiency": kpis.get("prbEfficiency"),
            "avgPdschDlLte": (kpis.get("pdschDlLte") or {}).get("average"),
            "avgPdschDlLteCw0": (kpis.get("pdschDlLteCw0") or {}).get("average"),
            "avgPdschDlLteCw1": (kpis.get("pdschDlLteCw1") or {}).get("average"),
            "modulationSamples": mod.get("sampleCount"),
            "mcsSamples": (kpis.get("pdschMcs") or {}).get("sampleCount"),
            "bitsHzSamples": (kpis.get("pdschBitPerHz") or {}).get("sampleCount"),
            "pdschSamples": kpis.get("pdschActiveSampleCount"),
            "interpretation": "",
        }

    iam_row = _row_for_operator(iam, iam_kpis)
    cmp_row = _row_for_operator(comparator, cmp_kpis)

    # Compute KPI-driven signals for interpretation and conclusion
    pdsch_sig = _nemo_evaluate_pdsch_signals(iam_kpis, cmp_kpis)
    _spec_better = pdsch_sig["spectralBetter"]
    _mod_better = pdsch_sig["modulationBetter"]
    _small_pool = pdsch_sig["smallerNrPool"]
    _no_n78 = pdsch_sig["missingN78"]
    _sched_hyp = pdsch_sig["schedulerHypothesis"]

    if not spectral.get("sufficient"):
        iam_row["interpretation"] = "Evidence is limited in this export."
    elif spectral.get("confirmed"):
        iam_row["interpretation"] = "IAM combines fewer PRBs with lower spectral efficiency."
    elif _spec_better and _mod_better:
        iam_row["interpretation"] = (
            f"IAM modulation, MCS and spectral efficiency are better than {comparator_name}. "
            + ("The dominant confirmed issue is the NR capacity pool gap"
               + (" and missing n78 contribution." if _no_n78 else ".")
               if _small_pool else
               ("Missing n78 high-capacity layer is the main confirmed issue." if _no_n78 else ""))
            + (" PRB allocation is not confirmed as a root cause." if _sched_hyp else "")
        ).strip()
    elif _spec_better:
        iam_row["interpretation"] = (
            f"IAM spectral efficiency is better than {comparator_name}; modulation quality is not the primary limitation."
        )
    else:
        iam_row["interpretation"] = f"IAM modulation/MCS is broadly comparable; the main issue remains NR band/BWP capacity."

    cmp_row["interpretation"] = (
        "Reference 5G comparator."
        if spectral.get("sufficient")
        else "Reference comparator, but modulation evidence is limited."
    )

    if not spectral.get("sufficient"):
        conclusion = "Insufficient PDSCH modulation/MCS evidence to confirm a spectral-efficiency limitation."
    elif spectral.get("confirmed"):
        conclusion = "IAM shows lower-order modulation and/or lower MCS/bit-per-Hz than the comparator. The throughput gap is therefore caused by both weaker NR resource allocation and lower spectral efficiency."
    else:
        conclusion = _nemo_generate_pdsch_conclusion(iam_kpis, cmp_kpis, comparator_name)
    severity = "Medium" if spectral.get("confirmed") else ("—" if spectral.get("sufficient") else "—")

    return {
        "title": "PDSCH Modulation & Spectral Efficiency",
        "available": True,
        "comparatorOperator": comparator_name,
        "rows": [iam_row, cmp_row],
        "conclusion": conclusion,
        "severity": severity,
        "sampleGuardrail": spectral.get("note"),
    }


def _nemo_build_nr_band_exposure_analysis(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    rows = []
    iam_row = None
    cmp_row = None
    for operator in operators or []:
        if not operator.get("has5g"):
            continue
        op_rows = operator.get("rows") or []
        nr_rows = [row for row in op_rows if _nemo_band_row_filter(row)]
        # Band anchors sorted by time, with a parallel time list for bisect-based nearest lookup.
        anchors = sorted(
            (
                (row.get("_dt"), str(row.get("band") or "").strip().lower())
                for row in nr_rows
                if row.get("_dt") and _nemo_is_valid_band(row.get("band"))
            ),
            key=lambda item: item[0],
        )
        anchor_times = [item[0] for item in anchors]

        def nearest_band(row_dt):
            # Nearest band anchor within 1.0 s, found via binary search over the two neighbours.
            if not row_dt or not anchors:
                return ""
            pos = bisect.bisect_left(anchor_times, row_dt)
            best_band = ""
            best_gap = None
            for j in (pos - 1, pos):
                if 0 <= j < len(anchors):
                    gap = abs((row_dt - anchor_times[j]).total_seconds())
                    if gap <= 1.0 and (best_gap is None or gap < best_gap):
                        best_gap = gap
                        best_band = anchors[j][1]
            return best_band

        band_values = [str(row.get("band") or "").strip().lower() for row in nr_rows if _nemo_is_valid_band(row.get("band"))]
        total = len(band_values)
        counts = {}
        for band in band_values:
            counts[band] = counts.get(band, 0) + 1
        n78_share = round((counts.get("n78", 0) / float(total)) * 100.0, 1) if total else None
        n1_share = round((counts.get("n1", 0) / float(total)) * 100.0, 1) if total else None
        n28_share = round((counts.get("n28", 0) / float(total)) * 100.0, 1) if total else None
        other_share = round(((sum(count for band, count in counts.items() if band not in {"n78", "n1", "n28"}) / float(total)) * 100.0), 1) if total else None

        # Single pass over op_rows: assign each row its nearest band once, then bucket the metrics.
        dl_by_band: dict = {}
        prbs_by_band: dict = {}
        sched_by_band: dict = {}
        for row in op_rows:
            band = nearest_band(row.get("_dt"))
            if not band:
                continue
            dl_value = _nemo_preferred_dl_value(row)
            if dl_value is not None and float(dl_value) > 0:
                dl_by_band.setdefault(band, []).append(float(dl_value))
            prbs = row.get("pdschPrbs")
            if prbs is not None and float(prbs) > 0:
                prbs_by_band.setdefault(band, []).append(float(prbs))
            sched = row.get("pdschSched5gMbps")
            if sched is not None and float(sched) > 0:
                sched_by_band.setdefault(band, []).append(float(sched))

        band_metrics = []
        for band in sorted(counts):
            band_metrics.append({
                "band": band.upper(),
                "share": round((counts.get(band, 0) / float(total)) * 100.0, 1) if total else None,
                "avgDlMbps": _nemo_metric_stats(dl_by_band.get(band, [])).get("average"),
                "avgPrbs": _nemo_metric_stats(prbs_by_band.get(band, [])).get("average"),
                "avgScheduled5gMbps": _nemo_metric_stats(sched_by_band.get(band, [])).get("average"),
            })
        row_payload = {
            "operator": operator.get("operator"),
            "n78Share": n78_share,
            "n1Share": n1_share,
            "n28Share": n28_share,
            "otherNrBandShare": other_share,
            "bandMetrics": band_metrics,
            "bandDistribution": _nemo_distribution([str(row.get("band") or "").strip().upper() for row in nr_rows if _nemo_is_valid_band(row.get("band"))]),
            "interpretation": "",
            "severity": "—",
        }
        rows.append(row_payload)
        if str(operator.get("operator") or "").upper() == "IAM":
            iam_row = row_payload
        if str(operator.get("operator") or "") == comparator_name:
            cmp_row = row_payload

    section_interpretation = "No 5G band exposure comparison available."
    section_severity = "—"
    if iam_row and cmp_row:
        n78_delta = None if iam_row.get("n78Share") is None or cmp_row.get("n78Share") is None else round(float(iam_row.get("n78Share")) - float(cmp_row.get("n78Share")), 1)
        if n78_delta is not None and n78_delta <= -20:
            section_severity = "High"
            section_interpretation = "IAM spends less time on high-capacity n78, reducing average DL throughput potential."
        elif (iam_row.get("n1Share") or 0) + (iam_row.get("n28Share") or 0) > max((iam_row.get("n78Share") or 0), 50):
            section_severity = "Medium"
            section_interpretation = "IAM is using lower-capacity NR layer more often."
        else:
            section_interpretation = "NR band exposure does not appear to be the primary difference."

    return {
        "title": "NR Band Exposure Analysis",
        "available": bool(rows),
        "comparatorOperator": comparator_name,
        "rows": rows,
        "severity": section_severity,
        "interpretation": section_interpretation,
    }


def _nemo_build_ca_scells_analysis(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    rows = []
    iam_row = None
    cmp_row = None
    nr_ca_missing = True
    for operator in operators or []:
        op_rows = operator.get("rows") or []
        lte_ca_values = [str(row.get("lteCaStatus") or "").strip() for row in op_rows if str(row.get("lteCaStatus") or "").strip()]
        nr_ca_values = [str(row.get("nrCaStatus") or "").strip() for row in op_rows if str(row.get("nrCaStatus") or "").strip()]
        scells_values = [float(row.get("scellsCount")) for row in op_rows if row.get("scellsCount") is not None and float(row.get("scellsCount")) >= 0]
        row_payload = {
            "operator": operator.get("operator"),
            "lteCaDistribution": _nemo_distribution(lte_ca_values),
            "nrCaDistribution": _nemo_distribution(nr_ca_values),
            "avgScells": _nemo_metric_stats(scells_values).get("average") if scells_values else None,
            "maxScells": max(scells_values) if scells_values else None,
            "scellsActiveShare": round((sum(1 for value in scells_values if value > 0) / float(len(scells_values))) * 100.0, 1) if scells_values else None,
            "lteCaActiveShare": _nemo_active_status_share(lte_ca_values),
            "nrCaActiveShare": _nemo_active_status_share(nr_ca_values),
            "interpretation": "",
            "severity": "—",
        }
        rows.append(row_payload)
        if nr_ca_values:
            nr_ca_missing = False
        if str(operator.get("operator") or "").upper() == "IAM":
            iam_row = row_payload
        if str(operator.get("operator") or "") == comparator_name:
            cmp_row = row_payload

    interpretation = "No carrier aggregation comparison available."
    severity = "—"
    if iam_row and cmp_row:
        avg_scells_gap = None if iam_row.get("avgScells") is None or cmp_row.get("avgScells") is None else float(iam_row.get("avgScells")) - float(cmp_row.get("avgScells"))
        ca_gap = None if iam_row.get("scellsActiveShare") is None or cmp_row.get("scellsActiveShare") is None else float(iam_row.get("scellsActiveShare")) - float(cmp_row.get("scellsActiveShare"))
        if avg_scells_gap is not None and avg_scells_gap <= -1:
            severity = "High"
            interpretation = "IAM may have lower carrier aggregation depth."
        elif ca_gap is not None and ca_gap <= -20:
            severity = "Medium"
            interpretation = "IAM has less CA activation."
        else:
            interpretation = "Carrier aggregation does not appear to be the main weakness with the current export."
    if nr_ca_missing:
        interpretation += " NR CA cannot be fully assessed; export NR CA status."

    return {
        "title": "Carrier Aggregation / SCells Analysis",
        "available": bool(rows),
        "rows": rows,
        "comparatorOperator": comparator_name,
        "severity": severity,
        "interpretation": interpretation,
    }


def _nemo_build_mimo_rank_analysis(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    rows = []
    iam_row = None
    cmp_row = None
    for operator in operators or []:
        if not operator.get("has5g"):
            continue
        op_rows = operator.get("rows") or []
        ri_values = [int(round(float(row.get("ri")))) for row in op_rows if row.get("ri") is not None and float(row.get("ri")) > 0]
        sched_rank_values = [int(round(float(row.get("scheduledRank")))) for row in op_rows if row.get("scheduledRank") is not None and float(row.get("scheduledRank")) > 0]
        total = len(ri_values)
        shares = {}
        for rank in (1, 2, 3, 4):
            shares[f"ri{rank}Share"] = round((sum(1 for value in ri_values if value == rank) / float(total)) * 100.0, 1) if total else None
        row_payload = {
            "operator": operator.get("operator"),
            **shares,
            "riGe3Share": round((sum(1 for value in ri_values if value >= 3) / float(total)) * 100.0, 1) if total else None,
            "medianRi": _nemo_metric_stats(ri_values).get("median") if ri_values else None,
            "averageRi": _nemo_metric_stats(ri_values).get("average") if ri_values else None,
            "scheduledRankDistribution": _nemo_distribution([str(value) for value in sched_rank_values]),
            "interpretation": "",
            "severity": "—",
        }
        rows.append(row_payload)
        if str(operator.get("operator") or "").upper() == "IAM":
            iam_row = row_payload
        if str(operator.get("operator") or "") == comparator_name:
            cmp_row = row_payload

    interpretation = "No MIMO rank comparison available."
    severity = "—"
    if iam_row and cmp_row:
        ri_gap = None if iam_row.get("riGe3Share") is None or cmp_row.get("riGe3Share") is None else float(iam_row.get("riGe3Share")) - float(cmp_row.get("riGe3Share"))
        if ri_gap is not None and ri_gap <= -15:
            severity = "Medium"
            interpretation = "IAM has lower high-rank MIMO usage."
        elif (iam_row.get("ri1Share") or 0) > 30:
            severity = "High"
            interpretation = "IAM has rank-1 limitation."
        else:
            interpretation = "MIMO rank is not the primary issue."

    return {
        "title": "MIMO Rank Analysis",
        "available": bool(rows),
        "rows": rows,
        "comparatorOperator": comparator_name,
        "severity": severity,
        "interpretation": interpretation,
    }


def _nemo_build_bler_retx_analysis(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    rows = []
    iam_row = None
    cmp_row = None
    for operator in operators or []:
        if not operator.get("has5g"):
            continue
        op_rows = operator.get("rows") or []
        bler_values = [float(row.get("macDlBler")) for row in op_rows if row.get("macDlBler") is not None]
        ul_retx_values = [float(row.get("macUlRetx5g")) for row in op_rows if row.get("macUlRetx5g") is not None]
        row_payload = {
            "operator": operator.get("operator"),
            "blerAvg": _nemo_metric_stats(bler_values).get("average") if bler_values else None,
            "blerMedian": _nemo_metric_stats(bler_values).get("median") if bler_values else None,
            "blerP90": _nemo_metric_stats(bler_values).get("p90") if bler_values else None,
            "blerGt10Share": _nemo_share_over_threshold(bler_values, 10.0),
            "blerGt20Share": _nemo_share_over_threshold(bler_values, 20.0),
            "ulRetxAvg": _nemo_metric_stats(ul_retx_values).get("average") if ul_retx_values else None,
            "ulRetxMedian": _nemo_metric_stats(ul_retx_values).get("median") if ul_retx_values else None,
            "ulRetxP90": _nemo_metric_stats(ul_retx_values).get("p90") if ul_retx_values else None,
        }
        rows.append(row_payload)
        if str(operator.get("operator") or "").upper() == "IAM":
            iam_row = row_payload
        if str(operator.get("operator") or "") == comparator_name:
            cmp_row = row_payload

    interpretation = "No BLER / retransmission comparison available."
    severity = "—"
    if iam_row:
        if (iam_row.get("blerAvg") or 0) > 10:
            severity = "High"
            interpretation = "High BLER may reduce delivered throughput."
        elif (iam_row.get("blerP90") or 0) > 20:
            severity = "High"
            interpretation = "Worst samples suffer retransmission losses."
        else:
            severity = "Low"
            interpretation = "Retransmission is not primary; scheduler/resource allocation remains main issue."

    return {
        "title": "BLER / Retransmission Analysis",
        "available": bool(rows),
        "rows": rows,
        "comparatorOperator": comparator_name,
        "severity": severity,
        "interpretation": interpretation,
    }


def _nemo_build_transport_gap_analysis(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    rows = []
    interpretation = "No transport/core comparison available."
    severity = "—"
    for operator in operators or []:
        kpis = operator.get("kpis") or {}
        op_rows = operator.get("rows") or []
        ping_values = [str(row.get("pingStatus") or "").strip().lower() for row in op_rows if str(row.get("pingStatus") or "").strip()]
        ping_success_rate = None
        if ping_values:
            ping_success_rate = round((sum(1 for value in ping_values if "success" in value) / float(len(ping_values))) * 100.0, 1)
        app_avg = (kpis.get("dl") or {}).get("average")
        total_mac_avg = (kpis.get("totalMacDl") or {}).get("average")
        mac_5g_avg = (kpis.get("mac5g") or {}).get("average")
        pdsch_avg = (kpis.get("pdsch5g") or {}).get("average")
        app_vs_total_ratio = round(app_avg / float(total_mac_avg), 2) if app_avg not in (None, 0) and total_mac_avg not in (None, 0) else None
        app_vs_pdsch_ratio = round(app_avg / float(pdsch_avg), 2) if app_avg not in (None, 0) and pdsch_avg not in (None, 0) else None
        row_payload = {
            "operator": operator.get("operator"),
            "appDlAvg": app_avg,
            "totalMacDlAvg": total_mac_avg,
            "macDl5gAvg": mac_5g_avg,
            "pdschDlAvg": pdsch_avg,
            "tcpHandshakeMedian": (kpis.get("tcpHandshake") or {}).get("median"),
            "lostPacketAvg": (kpis.get("lostPacket") or {}).get("average"),
            "pingSuccessRate": ping_success_rate,
            "appVsTotalMacRatio": app_vs_total_ratio,
            "appVsPdschRatio": app_vs_pdsch_ratio,
            "alignmentWarning": (
                "App/MAC ratio is indicative only. Application, MAC, and PDSCH KPIs must be measured on the same "
                "active test windows for a reliable transport/core diagnosis."
            ) if app_vs_total_ratio is not None else None,
        }
        rows.append(row_payload)
        if str(operator.get("operator") or "").upper() == "IAM":
            rf_good = ((kpis.get("rsrp") or {}).get("median") is not None and (kpis.get("rsrp") or {}).get("median") >= -95 and (kpis.get("sinr") or {}).get("median") is not None and (kpis.get("sinr") or {}).get("median") >= 8)
            if app_vs_total_ratio is not None and app_vs_total_ratio < 0.5 and rf_good:
                severity = "Medium"
                interpretation = "Possible transport/core/server limitation."
            elif ((row_payload.get("tcpHandshakeMedian") or 0) > 200) or ((row_payload.get("lostPacketAvg") or 0) > 0):
                severity = "Medium"
                interpretation = "Transport impairment may contribute."
            else:
                severity = "Low"
                interpretation = "Radio/scheduler limitation remains likely."
            if row_payload.get("tcpHandshakeMedian") is None and row_payload.get("pingSuccessRate") is None:
                interpretation += " Some transport metrics are sparse; export TCP/ping fields for deeper validation."

    if any((r.get("appVsTotalMacRatio") or 0) > 2.0 for r in rows):
        interpretation += " Note: App/MAC ratios above 1.0 indicate KPI window misalignment — values should be interpreted with caution."

    return {
        "title": "Transport / Core / Application Gap",
        "available": bool(rows),
        "rows": rows,
        "comparatorOperator": comparator_name,
        "severity": severity,
        "interpretation": interpretation,
    }


def _nemo_build_per_cgps_weakness_analysis(operators: list[dict], diagnosis: dict) -> dict:
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    iam = next((item for item in operators if str(item.get("operator") or "").upper() == "IAM"), None)
    comparator = next((item for item in operators if str(item.get("operator") or "") == comparator_name), None)
    if not iam:
        return {"title": "Per-CGPS Weakness Analysis", "available": False, "message": "IAM operator data is missing."}

    def build_groups(operator_data: dict):
        groups = {}
        for row in operator_data.get("rows") or []:
            lat = row.get("lat")
            lon = row.get("lon")
            if lat is None or lon is None:
                continue
            key = f"{round(float(lat), 5):.5f},{round(float(lon), 5):.5f}"
            groups.setdefault(key, {"lat": [], "lon": [], "rows": []})
            groups[key]["lat"].append(float(lat))
            groups[key]["lon"].append(float(lon))
            groups[key]["rows"].append(row)
        payload = []
        for key, group in groups.items():
            rows = group["rows"]
            dl_key = _nemo_select_dl_metric_key(rows)
            dl_values = _nemo_metric_series(rows, dl_key)
            prbs_values = [float(row.get("pdschPrbs")) for row in rows if row.get("pdschPrbs") is not None and float(row.get("pdschPrbs")) > 0]
            sched_values = [float(row.get("pdschSched5gMbps")) for row in rows if row.get("pdschSched5gMbps") is not None and float(row.get("pdschSched5gMbps")) > 0]
            pdsch_values = [float(row.get("pdschDl5gMbps")) for row in rows if row.get("pdschDl5gMbps") is not None and float(row.get("pdschDl5gMbps")) > 0]
            rsrp_values = [float(row.get("rsrp")) for row in rows if row.get("rsrp") is not None]
            sinr_values = [float(row.get("sinr")) for row in rows if row.get("sinr") is not None]
            cqi_values = [float(row.get("wbCqi")) for row in rows if row.get("wbCqi") is not None and float(row.get("wbCqi")) > 0]
            ri_values = [float(row.get("ri")) for row in rows if row.get("ri") is not None and float(row.get("ri")) > 0]
            bler_values = [float(row.get("macDlBler")) for row in rows if row.get("macDlBler") is not None]
            band_values = [str(row.get("band") or "").strip().lower() for row in rows if _nemo_band_row_filter(row) and str(row.get("band") or "").strip()]
            group_payload = {
                "cgpsKey": key,
                "lat": round(sum(group["lat"]) / float(len(group["lat"])), 6),
                "lon": round(sum(group["lon"]) / float(len(group["lon"])), 6),
                "avgDlMbps": _nemo_metric_stats(dl_values).get("average") if dl_values else None,
                "medianDlMbps": _nemo_metric_stats(dl_values).get("median") if dl_values else None,
                "prbsAvg": _nemo_metric_stats(prbs_values).get("average") if prbs_values else None,
                "scheduled5gAvg": _nemo_metric_stats(sched_values).get("average") if sched_values else None,
                "pdschDeliveredAvg": _nemo_metric_stats(pdsch_values).get("average") if pdsch_values else None,
                "rsrpMedian": _nemo_metric_stats(rsrp_values).get("median") if rsrp_values else None,
                "sinrMedian": _nemo_metric_stats(sinr_values).get("median") if sinr_values else None,
                "cqiMedian": _nemo_metric_stats(cqi_values).get("median") if cqi_values else None,
                "riGe3Share": round((sum(1 for value in ri_values if value >= 3) / float(len(ri_values))) * 100.0, 1) if ri_values else None,
                "blerAvg": _nemo_metric_stats(bler_values).get("average") if bler_values else None,
                "n78Share": round((sum(1 for value in band_values if value == "n78") / float(len(band_values))) * 100.0, 1) if band_values else None,
                "rows": rows,
            }
            meaningful = any(
                group_payload.get(key_name) is not None
                for key_name in ("avgDlMbps", "prbsAvg", "scheduled5gAvg", "pdschDeliveredAvg", "rsrpMedian", "sinrMedian", "cqiMedian", "riGe3Share", "blerAvg")
            )
            if meaningful:
                payload.append(group_payload)
        return payload

    iam_groups = build_groups(iam)
    cmp_groups = build_groups(comparator) if comparator else []
    result_rows = []
    for group in iam_groups:
        matched = next((item for item in cmp_groups if item.get("cgpsKey") == group.get("cgpsKey")), None)
        match_distance = 0.0 if matched else None
        if not matched and cmp_groups:
            nearest = None
            nearest_dist = None
            for candidate in cmp_groups:
                dist = _nemo_haversine_m(group.get("lat"), group.get("lon"), candidate.get("lat"), candidate.get("lon"))
                if dist is None:
                    continue
                if nearest_dist is None or dist < nearest_dist:
                    nearest = candidate
                    nearest_dist = dist
            if nearest is not None and nearest_dist is not None and nearest_dist <= 30.0:
                matched = nearest
                match_distance = round(nearest_dist, 1)

        diagnosis_label = "No matched competitor location."
        if matched:
            prbs_gap = _nemo_gap_pct(group.get("prbsAvg"), matched.get("prbsAvg"))
            n78_gap = _nemo_gap_pct(group.get("n78Share"), matched.get("n78Share"))
            rsrp_gap = None if group.get("rsrpMedian") is None or matched.get("rsrpMedian") is None else round(float(group.get("rsrpMedian")) - float(matched.get("rsrpMedian")), 1)
            sinr_gap = None if group.get("sinrMedian") is None or matched.get("sinrMedian") is None else round(float(group.get("sinrMedian")) - float(matched.get("sinrMedian")), 1)
            ri_gap = None if group.get("riGe3Share") is None or matched.get("riGe3Share") is None else round(float(group.get("riGe3Share")) - float(matched.get("riGe3Share")), 1)
            bler_gap = None if group.get("blerAvg") is None or matched.get("blerAvg") is None else round(float(group.get("blerAvg")) - float(matched.get("blerAvg")), 1)
            if prbs_gap is not None and prbs_gap <= -50:
                diagnosis_label = "Scheduler / PRB limitation"
            elif n78_gap is not None and n78_gap <= -20:
                diagnosis_label = "Band exposure limitation"
            elif rsrp_gap is not None and rsrp_gap < -6:
                diagnosis_label = "Coverage limitation"
            elif sinr_gap is not None and sinr_gap < -5:
                diagnosis_label = "Quality / interference limitation"
            elif ri_gap is not None and ri_gap <= -15:
                diagnosis_label = "MIMO limitation"
            elif bler_gap is not None and bler_gap > 3 and (group.get("blerAvg") or 0) > 10:
                diagnosis_label = "BLER / retransmission issue"
            elif _nemo_gap_pct(group.get("avgDlMbps"), matched.get("avgDlMbps")) is not None and _nemo_gap_pct(group.get("avgDlMbps"), matched.get("avgDlMbps")) <= -25:
                diagnosis_label = "General throughput deficit"
            else:
                diagnosis_label = "No dominant weakness against matched comparator"

        result_rows.append({
            "cgpsKey": group.get("cgpsKey"),
            "lat": group.get("lat"),
            "lon": group.get("lon"),
            "avgDlMbps": group.get("avgDlMbps"),
            "medianDlMbps": group.get("medianDlMbps"),
            "prbsAvg": group.get("prbsAvg"),
            "scheduled5gAvg": group.get("scheduled5gAvg"),
            "pdschDeliveredAvg": group.get("pdschDeliveredAvg"),
            "rsrpMedian": group.get("rsrpMedian"),
            "sinrMedian": group.get("sinrMedian"),
            "cqiMedian": group.get("cqiMedian"),
            "riGe3Share": group.get("riGe3Share"),
            "blerAvg": group.get("blerAvg"),
            "matchedComparator": comparator_name if matched else "",
            "matchDistanceM": match_distance,
            "mainDiagnosis": diagnosis_label,
        })

    return {
        "title": "Per-CGPS Weakness Analysis",
        "available": bool(result_rows),
        "comparatorOperator": comparator_name,
        "rows": result_rows,
        "message": "" if result_rows else "No CGPS rows available for IAM.",
    }


def _nemo_build_operator_serving_cells(
    operator_data: dict,
    bdd_4g_path: str | None = None,
    bdd_5g_path: str | None = None,
) -> dict:
    """
    Match IAM GPS measurement points to serving cell names using the webapp's
    bdd_sectors.json (the multi-RAT BDD already parsed by the webapp layer).

    Matching priority (per GPS row):
      1. PCI + freq match within 3 km  (best — both RF identifiers confirmed)
      2. PCI-only match within 3 km    (good — frequency not exported by Nemo)
      3. Nearest cell within 1.5 km    (fallback — distance-only)
    """
    from math import radians, sin, cos, asin

    # ── Load bdd_sectors.json (cached in memory by mtime) ─────────────────────
    sectors_path = os.path.join(_APP_DIR, "bdd_sectors.json")
    if not os.path.isfile(sectors_path):
        return {
            "available": False,
            "bddAvailable": False,
            "message": "BDD non chargé dans la webapp — configurez le BDD via 'Configurer BDD' pour activer l'identification des cellules servantes.",
        }

    # Simple mtime-based cache so we only re-parse when the file changes
    cache = getattr(_nemo_build_operator_serving_cells, "_cache", None)
    file_mtime = os.path.getmtime(sectors_path)
    if cache is None or cache.get("mtime") != file_mtime:
        try:
            with open(sectors_path, "r", encoding="utf-8") as fh:
                raw = json.load(fh)
            sectors_list = raw if isinstance(raw, list) else raw.get("sectors", [])
            # Keep only 4G/5G sectors with pci, freq, lat, lng
            filtered = [
                s for s in sectors_list
                if s.get("pci") is not None
                and s.get("lat") is not None
                and s.get("lng") is not None
                and s.get("tech") in ("4G", "5G")
            ]
            _nemo_build_operator_serving_cells._cache = {"mtime": file_mtime, "cells": filtered}
        except json.JSONDecodeError:
            return {"available": False, "bddAvailable": False,
                    "message": "Le BDD est en cours de mise à jour en arrière-plan. Veuillez patienter quelques secondes et relancer l'import."}
        except Exception as exc:
            return {"available": False, "bddAvailable": False,
                    "message": f"Erreur lecture bdd_sectors.json : {exc}"}

    bdd_cells = _nemo_build_operator_serving_cells._cache["cells"]
    if not bdd_cells:
        return {
            "available": False,
            "bddAvailable": False,
            "message": "BDD chargé mais aucune cellule 4G/5G trouvée — vérifiez les feuilles du fichier BDD.",
        }

    # ── Find Operator data ─────────────────────────────────────────────────
    if not operator_data:
        return {"available": False, "message": "Operator data not provided."}

    op_name = operator_data.get("operator", "UNKNOWN").upper()
    rows = operator_data.get("rows") or []
    gps_rows = [r for r in rows if r.get("lat") is not None and r.get("lon") is not None]
    if not gps_rows:
        return {"available": False, "message": f"No GPS coordinates found in {op_name} Nemo export."}
    dominant_arfcn_by_tech_pci = _nemo_resolve_dominant_arfcn_by_tech_pci(rows)

    # ── Haversine helper ───────────────────────────────────────────────────────
    def _hav(lat1, lon1, lat2, lon2):
        R = 6_371_000.0
        la1, lo1, la2, lo2 = map(radians, [lat1, lon1, lat2, lon2])
        a = sin((la2 - la1) / 2) ** 2 + cos(la1) * cos(la2) * sin((lo2 - lo1) / 2) ** 2
        return R * 2.0 * asin(max(0.0, min(1.0, a)) ** 0.5)

    MAX_DIST_M      = 3000.0   # PCI or PCI+freq match radius
    FALLBACK_DIST_M = 1500.0   # Nearest-only fallback radius

    # ── Calculate Bounding Box of Drive Test to Optimize Matching ──────────────
    min_lat = min(float(r["lat"]) for r in gps_rows)
    max_lat = max(float(r["lat"]) for r in gps_rows)
    min_lon = min(float(r["lon"]) for r in gps_rows)
    max_lon = max(float(r["lon"]) for r in gps_rows)
    
    # 3km is roughly 0.027 degrees. Use 0.035 for safety padding.
    PAD = 0.035
    local_bdd_cells = [
        c for c in bdd_cells
        if (min_lat - PAD) <= c["lat"] <= (max_lat + PAD) and (min_lon - PAD) <= c["lng"] <= (max_lon + PAD)
    ]
    bdd_cells_by_tech_pci: dict = {}
    bdd_cells_by_tech_pci_freq: dict = {}
    for cell in local_bdd_cells:
        tech = cell.get("tech")
        pci = cell.get("pci")
        freq = cell.get("freq")
        bdd_cells_by_tech_pci.setdefault((tech, pci), []).append(cell)
        if freq is not None:
            bdd_cells_by_tech_pci_freq.setdefault((tech, pci, freq), []).append(cell)

    # ── Group rows by exact timestamp — one match decision per Nemo sample time ─
    # App. rate DL rows are sparse and usually do not carry serving-cell identity.
    # We therefore build explicit LTE/NR serving timelines first, then attribute
    # App DL samples to the latest active serving cell at or before each sample.
    _SERVING_TYPES = {"lte serving", "nr serving", "nr scg pscell", "5g serving"}

    from collections import defaultdict
    from datetime import datetime as _datetime

    ts_buckets: dict = defaultdict(list)
    for r in rows:
        dt = r.get("_dt")
        if isinstance(dt, _datetime):
            ts_buckets[dt].append(r)

    cell_hits: dict = {}
    match_method_counts = {"pci_freq": 0, "pci_only": 0, "proximity": 0}
    unmatched_count = 0
    decision_bucket_count = 0
    matched_bucket_count = 0
    unmatched_bucket_count = 0
    unmatched_no_serving_pci_count = 0
    unmatched_bdd_miss_count = 0
    unmatched_bdd_miss_bucket_count = 0
    ts_records: list = []   # one (or two for EN-DC) records per second
    lte_timeline: list = []
    nr_timeline: list = []
    last_known_lat = None
    last_known_lon = None
    last_known_dt = None

    _LTE_SERVING_TYPES = {"lte serving"}
    _NR_SERVING_TYPES  = {"nr serving", "nr scg pscell", "5g serving"}

    # The drive test dwells on repeated coordinates, so memoize the (expensive) BDD distance
    # search by rounded coordinate + RF identifiers — most lookups hit the cache.
    _match_cache: dict = {}

    def _match_pci(lat, lon, int_pci, int_arfcn, expected_tech):
        key = (round(lat, 5), round(lon, 5), int_pci, int_arfcn, expected_tech)
        cached = _match_cache.get(key)
        if cached is not None:
            return cached
        result = _match_pci_compute(lat, lon, int_pci, int_arfcn, expected_tech)
        _match_cache[key] = result
        return result

    def _match_pci_compute(lat, lon, int_pci, int_arfcn, expected_tech):
        if op_name in ("ORANGE", "INWI"):
            if int_pci is not None:
                canonical_arfcn = int_arfcn
                if canonical_arfcn is None:
                    canonical_arfcn = dominant_arfcn_by_tech_pci.get((expected_tech, int_pci))
                cell_name = f"PCI {int_pci}"
                site_name = f"ARFCN {canonical_arfcn}" if canonical_arfcn is not None else "ARFCN Unknown"
                tech = expected_tech or ("5G" if (canonical_arfcn and canonical_arfcn > 100000) else "4G")
                band = "n78" if tech == "5G" and (canonical_arfcn and canonical_arfcn > 600000) else ("NR" if tech == "5G" else "LTE")
                return {"cellName": cell_name, "siteName": site_name, "tech": tech, "band": band}, "pci_freq"
            return None, None

        best, best_dist, best_method = None, float("inf"), None
        if int_pci is not None and int_arfcn is not None:
            for cell in bdd_cells_by_tech_pci_freq.get((expected_tech, int_pci, int_arfcn), []):
                d = _hav(lat, lon, cell["lat"], cell["lng"])
                if d < MAX_DIST_M and d < best_dist:
                    best, best_dist, best_method = cell, d, "pci_freq"
        if best is None and int_pci is not None:
            for cell in bdd_cells_by_tech_pci.get((expected_tech, int_pci), []):
                d = _hav(lat, lon, cell["lat"], cell["lng"])
                if d < MAX_DIST_M and d < best_dist:
                    best, best_dist, best_method = cell, d, "pci_only"
        if best is None:
            for cell in local_bdd_cells:
                if expected_tech and cell.get("tech") != expected_tech:
                    continue
                d = _hav(lat, lon, cell["lat"], cell["lng"])
                if d < FALLBACK_DIST_M and d < best_dist:
                    best, best_dist, best_method = cell, d, "proximity"
        return best, best_method

    def _find_serving_pci(bucket, ct_set):
        """Return (pci, arfcn, band) from the first row in bucket whose cellTypes match ct_set."""
        for r in bucket:
            ct_str = " ".join(r.get("cellTypes") or []).lower()
            if not any(s in ct_str for s in ct_set):
                continue
            pci_v = r.get("pci")
            if pci_v is not None:
                arfcn_v = r.get("nrChannelNumber")
                return int(pci_v), (int(arfcn_v) if arfcn_v is not None else None), str(r.get("band") or "").strip() or None
        return None, None, None

    def _collect_kpis(bucket):
        rsrp_vals, sinr_vals = [], []
        for r in bucket:
            rv = r.get("rsrp")
            if rv is not None:
                rsrp_vals.append(float(rv))
            sv = r.get("sinr")
            if sv is not None:
                sinr_vals.append(float(sv))
        return rsrp_vals, sinr_vals

    def _add_timeline_entry(timeline, dt_val, key, lat_val, lon_val):
        if key is None or dt_val is None:
            return
        timeline.append({
            "dt": dt_val,
            "key": key,
            "lat": lat_val,
            "lon": lon_val,
        })

    for dt_key, bucket in ts_buckets.items():
        lat_val = lon_val = None
        for r in bucket:
            if r.get("lat") is not None and r.get("lon") is not None:
                lat_val, lon_val = float(r["lat"]), float(r["lon"])
                last_known_lat, last_known_lon, last_known_dt = lat_val, lon_val, dt_key
                break
        if lat_val is None and last_known_lat is not None and last_known_lon is not None:
            if last_known_dt is None or abs((dt_key - last_known_dt).total_seconds()) <= 10:
                lat_val, lon_val = last_known_lat, last_known_lon
        if lat_val is None:
            continue
        decision_bucket_count += 1

        rsrp_bucket, sinr_bucket = _collect_kpis(bucket)

        # Serving technology label from Nemo (e.g. "LTE CA", "LTE FDD", "EN-DC")
        serv_tech = next((str(r.get("servingTechnology") or "").strip()
                          for r in bucket if r.get("servingTechnology")), None)
        rrc_state = next((str(r.get("rrcState") or "").strip()
                          for r in bucket if str(r.get("rrcState") or "").strip()), None)

        # LTE anchor (always present in LTE/EN-DC)
        lte_pci, lte_arfcn, lte_band = _find_serving_pci(bucket, _LTE_SERVING_TYPES)
        # NR PSCell (only in EN-DC / 5G sessions)
        nr_pci,  nr_arfcn, nr_band = _find_serving_pci(bucket, _NR_SERVING_TYPES)

        bucket_had_bdd_miss = False

        def _register(pci, arfcn, is_primary, expected_tech, observed_band=None, serv_mode=None, lte_anchor=None):
            nonlocal unmatched_count, unmatched_bdd_miss_count, bucket_had_bdd_miss
            best, method = _match_pci(lat_val, lon_val, pci, arfcn, expected_tech)
            if best:
                match_method_counts[method] += 1
                band_label = best.get("band", "") or ""
                if expected_tech == "4G":
                    band_label = _nemo_lte_band_label(observed_band) or band_label
                key = (best.get("cellName") or "", best.get("siteName") or "", best.get("tech", ""), band_label)
                if key not in cell_hits:
                    cell_hits[key] = {"count": 0, "dl": [], "ul": [], "rsrp": [], "sinr": [], "ts": [], "appTs": [], "appTsUl": []}
                agg = cell_hits[key]
                agg["count"] += 1
                agg["rsrp"].extend(rsrp_bucket)
                agg["sinr"].extend(sinr_bucket)
                if isinstance(dt_key, _datetime):
                    agg["ts"].append(dt_key)
                return {"dt": dt_key, "key": key, "primary": is_primary,
                        "dl": [],
                        "ul": [],
                        "appTs": [],
                        "appTsUl": [],
                        "rsrp": rsrp_bucket,
                        "sinr": sinr_bucket,
                        "rrcState": rrc_state,
                        "servingMode": serv_mode,
                        "lteAnchor": lte_anchor,
                        "lat": lat_val,
                        "lon": lon_val,
                        "cellLat": best.get("lat"),
                        "cellLon": best.get("lon") or best.get("lng")}
            else:
                unmatched_count += 1
                unmatched_bdd_miss_count += 1
                bucket_had_bdd_miss = True
                return None

        matched_any = False
        # In EN-DC: NR PSCell is primary (episode + KPIs), LTE anchor is secondary (hit count only)
        # Pure LTE: LTE serving is primary (episode + KPIs)
        endc = nr_pci is not None
        lte_rec = None
        if lte_pci is not None:
            lte_rec = _register(lte_pci, lte_arfcn, is_primary=not endc,
                                expected_tech="4G", observed_band=lte_band,
                                serv_mode=serv_tech if not endc else None)
            if lte_rec:
                ts_records.append(lte_rec)
                _add_timeline_entry(lte_timeline, dt_key, lte_rec["key"], lat_val, lon_val)
                matched_any = True
        if nr_pci is not None:
            lte_anchor_name = lte_rec["key"][0] if lte_rec and lte_rec.get("key") else None
            rec = _register(nr_pci, nr_arfcn, is_primary=True,
                            expected_tech="5G", observed_band=nr_band,
                            serv_mode="EN-DC", lte_anchor=lte_anchor_name)
            if rec:
                ts_records.append(rec)
                _add_timeline_entry(nr_timeline, dt_key, rec["key"], lat_val, lon_val)
                matched_any = True
        if not matched_any and lte_pci is None and nr_pci is None:
            unmatched_count += 1
            unmatched_bucket_count += 1
            unmatched_no_serving_pci_count += 1
        elif not matched_any:
            unmatched_bucket_count += 1
            if bucket_had_bdd_miss:
                unmatched_bdd_miss_bucket_count += 1
        else:
            matched_bucket_count += 1

    def _build_intervals(timeline):
        intervals = []
        if not timeline:
            return intervals
        current = timeline[0]
        for entry in timeline[1:]:
            if entry["key"] == current["key"]:
                continue
            intervals.append({
                "key": current["key"],
                "start": current["dt"],
                "end": entry["dt"],
            })
            current = entry
        intervals.append({
            "key": current["key"],
            "start": current["dt"],
            "end": None,
        })
        return intervals

    lte_intervals = _build_intervals(sorted(lte_timeline, key=lambda item: item["dt"]))
    nr_intervals = _build_intervals(sorted(nr_timeline, key=lambda item: item["dt"]))

    app_samples = []
    for row in rows:
        dt_val = row.get("_dt")
        dl_val = row.get("appDlMbps")
        if not isinstance(dt_val, _datetime) or dl_val is None:
            continue
        try:
            dl_num = float(dl_val)
        except Exception:
            continue
        if dl_num <= 0:
            continue
        app_samples.append({"dt": dt_val, "dl": dl_num})

    ul_samples = []
    for row in rows:
        dt_val = row.get("_dt")
        ul_val = row.get("appUlMbps")
        if not isinstance(dt_val, _datetime) or ul_val is None:
            continue
        try:
            ul_num = float(ul_val)
        except Exception:
            continue
        if ul_num <= 0:
            continue
        ul_samples.append({"dt": dt_val, "ul": ul_num})

    app_samples.sort(key=lambda item: item["dt"])
    ul_samples.sort(key=lambda item: item["dt"])
    lte_idx = 0
    nr_idx = 0
    primary_records = sorted(
        [rec for rec in ts_records if rec.get("key") is not None and rec.get("primary") and isinstance(rec.get("dt"), _datetime)],
        key=lambda rec: rec["dt"],
    )
    primary_idx = 0
    last_primary = None

    def _interval_for_sample(intervals, index, sample_dt):
        if not intervals:
            return index, None
        while index + 1 < len(intervals):
            next_start = intervals[index + 1].get("start")
            if next_start is None or next_start > sample_dt:
                break
            index += 1
        interval = intervals[index]
        start = interval.get("start")
        end = interval.get("end")
        if start is None or sample_dt < start or (end is not None and sample_dt >= end):
            return index, None
        return index, interval

    for sample in app_samples:
        dt_val = sample["dt"]
        dl_num = sample["dl"]
        lte_idx, lte_interval = _interval_for_sample(lte_intervals, lte_idx, dt_val)
        nr_idx, nr_interval = _interval_for_sample(nr_intervals, nr_idx, dt_val)
        while primary_idx < len(primary_records) and primary_records[primary_idx]["dt"] <= dt_val:
            last_primary = primary_records[primary_idx]
            primary_idx += 1
        assigned_keys = set()
        for interval in (lte_interval, nr_interval):
            if not interval:
                continue
            key = interval.get("key")
            if key is None or key in assigned_keys:
                continue
            agg = cell_hits.get(key)
            if not agg:
                continue
            agg["dl"].append(dl_num)
            agg["appTs"].append(dt_val)
            assigned_keys.add(key)

        if last_primary is not None:
            last_primary.setdefault("dl", []).append(dl_num)
            last_primary.setdefault("appTs", []).append(dt_val)

    ul_lte_idx = 0
    ul_nr_idx = 0
    ul_primary_idx = 0
    ul_last_primary = None
    for sample in ul_samples:
        dt_val = sample["dt"]
        ul_num = sample["ul"]
        ul_lte_idx, lte_interval = _interval_for_sample(lte_intervals, ul_lte_idx, dt_val)
        ul_nr_idx, nr_interval = _interval_for_sample(nr_intervals, ul_nr_idx, dt_val)
        while ul_primary_idx < len(primary_records) and primary_records[ul_primary_idx]["dt"] <= dt_val:
            ul_last_primary = primary_records[ul_primary_idx]
            ul_primary_idx += 1
        assigned_keys = set()
        for interval in (lte_interval, nr_interval):
            if not interval:
                continue
            key = interval.get("key")
            if key is None or key in assigned_keys:
                continue
            agg = cell_hits.get(key)
            if not agg:
                continue
            agg["ul"].append(ul_num)
            agg["appTsUl"].append(dt_val)
            assigned_keys.add(key)

        if ul_last_primary is not None:
            ul_last_primary.setdefault("ul", []).append(ul_num)
            ul_last_primary.setdefault("appTsUl", []).append(dt_val)

    # ── Build result payload ───────────────────────────────────────────────────
    if not cell_hits:
        return {
            "available": False,
            "bddAvailable": True,
            "gpsRows": len(gps_rows),
            "decisionBucketCount": decision_bucket_count,
            "matchedCount": matched_bucket_count,
            "matchedBucketCount": matched_bucket_count,
            "unmatchedCount": unmatched_bucket_count,
            "unmatchedBucketCount": unmatched_bucket_count,
            "unmatchedNoServingPciCount": unmatched_no_serving_pci_count,
            "unmatchedBddMissCount": unmatched_bdd_miss_count,
            "unmatchedBddMissBucketCount": unmatched_bdd_miss_bucket_count,
            "message": f"Aucune cellule {op_name} identifiée — vérifiez que le BDD chargé correspond à la zone du benchmark.",
        }

    def _median(lst):
        if not lst:
            return None
        s = sorted(lst)
        n = len(s)
        return round(s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2, 1)

    def _avg(lst):
        return round(sum(lst) / len(lst), 1) if lst else None

    def _max(lst):
        return round(max(lst), 1) if lst else None

    total_ts = max(len(ts_buckets), 1)

    # ── Chronological episode sequence (one record per timestamp) ───────────────
    def _has_dt(rec):
        from datetime import datetime
        return isinstance(rec.get("dt"), datetime)

    timed_records = sorted(
        [rec for rec in ts_records if rec.get("key") is not None and rec.get("primary") and _has_dt(rec)],
        key=lambda rec: rec["dt"]
    )

    lte_end_dt = max((entry.get("dt") for entry in lte_timeline if entry.get("dt") is not None), default=None)
    nr_end_dt = max((entry.get("dt") for entry in nr_timeline if entry.get("dt") is not None), default=None)
    primary_end_dt = max((rec.get("dt") for rec in timed_records if rec.get("dt") is not None), default=None)

    lte_dwell_by_key = _nemo_sum_interval_seconds_by_key(lte_intervals, lte_end_dt)
    nr_dwell_by_key = _nemo_sum_interval_seconds_by_key(nr_intervals, nr_end_dt)
    primary_episodes = _nemo_build_episode_ranges(timed_records, primary_end_dt)
    primary_dwell_by_key: dict = {}
    primary_count_by_key: dict = {}
    for episode in primary_episodes:
        key = episode.get("key")
        dwell = episode.get("dwellSec")
        if key is None or dwell is None:
            pass
        else:
            primary_dwell_by_key[key] = round(primary_dwell_by_key.get(key, 0.0) + dwell, 0)
        if key is None:
            continue
        primary_count_by_key[key] = primary_count_by_key.get(key, 0) + len(episode.get("records") or [])

    total_primary_hits = sum(primary_count_by_key.values()) or 0
    sorted_cells = sorted(cell_hits.items(), key=lambda item: -item[1]["count"])
    cells_payload = []
    for (cn, sn, tech, band), agg in sorted_cells:
        key = (cn, sn, tech, band)
        if str(tech).upper().startswith("5G"):
            dwell_sec = primary_dwell_by_key.get(key)
            if dwell_sec is None:
                dwell_sec = nr_dwell_by_key.get(key)
        else:
            dwell_sec = primary_dwell_by_key.get(key)
        primary_hit_count = primary_count_by_key.get(key)
        cells_payload.append({
            "cellName":     cn,
            "siteName":     sn,
            "tech":         tech,
            "band":         band,
            "hitCount":     agg["count"],
            "primaryHitCount": primary_hit_count,
            "primarySharePercent": round(primary_hit_count / float(total_primary_hits) * 100, 1) if primary_hit_count is not None and total_primary_hits else None,
            "sharePercent": round(agg["count"] / total_ts * 100, 1),
            "avgDlMbps":    _avg(agg["dl"]),
            "maxDlMbps":    _max(agg["dl"]),
            "avgUlMbps":    _avg(agg["ul"]),
            "maxUlMbps":    _max(agg["ul"]),
            "appSampleCount": len(agg["dl"]),
            "appUlSampleCount": len(agg["ul"]),
            "appFirstTime": _nemo_fmt_hms_ms(agg["appTs"][0]) if agg.get("appTs") else None,
            "appLastTime":  _nemo_fmt_hms_ms(agg["appTs"][-1]) if agg.get("appTs") else None,
            "medianRsrp":   _median(agg["rsrp"]),
            "medianSinr":   _median(agg["sinr"]),
            "dwellSec":     dwell_sec,
        })

    _EP_COLORS = ['#3b82f6', '#a855f7', '#f97316', '#10b981', '#f43f5e', '#06b6d4', '#eab308', '#8b5cf6']
    episode_sources = []
    for ep_idx, episode in enumerate(primary_episodes):
        key = episode.get("key")
        ep_recs = episode.get("records") or []
        if key is None or not ep_recs:
            continue
        color = _EP_COLORS[ep_idx % len(_EP_COLORS)]
        source = dict(episode)
        source["display"] = _nemo_episode_display_payload(
            episode,
            idx=ep_idx + 1,
            color=color,
            records=ep_recs,
        )
        episode_sources.append(source)

    # Scope "during download" to the DREQ→DCOMP event window (exact transfer boundary).
    # Fall back to appDlMbps>0 seconds, then to session-level transfer windows.
    _dl_events = _nemo_extract_dl_events(rows)
    download_intervals = (
        _dl_events.get("downloadIntervals")
        or _nemo_active_download_intervals(rows)
        or _nemo_downlink_transfer_intervals(operator_data.get("transferSessions") or [])
    )
    download_scoped = bool(download_intervals)
    all_window_episode_payload = []
    for ep_idx, episode in enumerate(episode_sources):
        display = dict(episode.get("display") or {})
        display["idx"] = ep_idx + 1
        if not display.get("color"):
            display["color"] = _EP_COLORS[ep_idx % len(_EP_COLORS)]
        all_window_episode_payload.append(display)

    display_episode_sources = (
        _nemo_clip_primary_episodes_to_intervals(episode_sources, download_intervals)
        if download_intervals
        else episode_sources
    )
    download_dwell_by_key = _nemo_episode_dwell_by_key(display_episode_sources)
    for cell in cells_payload:
        key = (
            cell.get("cellName") or "",
            cell.get("siteName") or "",
            cell.get("tech") or "",
            cell.get("band") or "",
        )
        cell["dwellSecDownload"] = download_dwell_by_key.get(key)
    dwell_share_by_key = _nemo_presence_share_from_cells(cells_payload, "dwellSec")
    download_dwell_share_by_key = _nemo_presence_share_from_cells(cells_payload, "dwellSecDownload")
    for cell in cells_payload:
        key = (
            cell.get("cellName"),
            cell.get("siteName"),
            cell.get("tech"),
            cell.get("band"),
        )
        cell["dwellSharePercent"] = dwell_share_by_key.get(key)
        cell["dwellSharePercentDownload"] = download_dwell_share_by_key.get(key)

    upload_intervals = _nemo_uplink_transfer_intervals(operator_data.get("transferSessions") or [])
    upload_scoped = bool(upload_intervals)
    upload_episode_sources = (
        _nemo_clip_primary_episodes_to_intervals(episode_sources, upload_intervals)
        if upload_intervals
        else episode_sources
    )
    upload_dwell_by_key = _nemo_episode_dwell_by_key(upload_episode_sources)
    for cell in cells_payload:
        key = (
            cell.get("cellName") or "",
            cell.get("siteName") or "",
            cell.get("tech") or "",
            cell.get("band") or "",
        )
        cell["dwellSecUpload"] = upload_dwell_by_key.get(key)
    upload_dwell_share_by_key = _nemo_presence_share_from_cells(cells_payload, "dwellSecUpload")
    for cell in cells_payload:
        key = (
            cell.get("cellName"),
            cell.get("siteName"),
            cell.get("tech"),
            cell.get("band"),
        )
        cell["dwellSharePercentUpload"] = upload_dwell_share_by_key.get(key)
    download_window_episode_payload = []
    for ep_idx, episode in enumerate(display_episode_sources):
        display = dict(episode.get("display") or {})
        display["idx"] = ep_idx + 1
        if not display.get("color"):
            display["color"] = _EP_COLORS[ep_idx % len(_EP_COLORS)]
        download_window_episode_payload.append(display)

    # Build a (timestamp, packetTechnology) timeline from the raw rows.
    # Packet technology is a sparse change-event column — forward-filling across seconds
    # gives an accurate time-based 5G/4G presence, unlike BDD dwell which can misclassify
    # cells when the database technology tag is wrong.  Fall back to the BDD dwell approach
    # only when the timeline yields no attributable seconds (e.g. old Nemo exports with no
    # packet/serving technology columns at all).
    _tech_timeline = [
        (r["_dt"], r.get("packetTechnology") or r.get("servingTechnology") or "")
        for r in rows if r.get("_dt") is not None
    ]
    radio_presence_breakdown_all = (
        _nemo_tech_presence_from_timeline(_tech_timeline)
        or _nemo_radio_presence_breakdown_from_cells(cells_payload)
    )
    radio_presence_breakdown_download = (
        (
            _nemo_tech_presence_from_timeline(_tech_timeline, download_intervals)
            or _nemo_radio_presence_breakdown_from_episodes(display_episode_sources)
        )
        if download_intervals
        else {}
    )
    download_window_start = download_intervals[0].get("start") if download_intervals else None
    download_window_end = download_intervals[-1].get("end") if download_intervals else None

    # GPS trace: per-second points tagged with episode index and color
    gps_trace: list = []
    for ep_idx, episode in enumerate(primary_episodes):
        key = episode.get("key")
        ep_recs = episode.get("records") or []
        if key is None or not ep_recs:
            continue
        cn, sn, tech, band = key
        color = _EP_COLORS[ep_idx % len(_EP_COLORS)]
        for rec in ep_recs:
            lat = rec.get("lat")
            lon = rec.get("lon")
            if lat is None or lon is None:
                continue
            rsrp_list = rec.get("rsrp") or []
            sinr_list = rec.get("sinr") or []
            dl_list   = rec.get("dl") or []
            gps_trace.append({
                "lat":        round(lat, 6),
                "lon":        round(lon, 6),
                "t":          _nemo_fmt_hms_ms(rec["dt"]) if isinstance(rec.get("dt"), _datetime) else "",
                "cellName":   cn,
                "siteName":   sn,
                "tech":       tech,
                "band":       band or "",
                "episodeIdx": ep_idx + 1,
                "color":      color,
                "rsrp":       round(_median(rsrp_list), 1) if rsrp_list else None,
                "sinr":       round(_median(sinr_list), 1) if sinr_list else None,
                "dl":         round(_avg(dl_list), 2) if dl_list else None,
                "cellLat":    rec.get("cellLat"),
                "cellLon":    rec.get("cellLon"),
            })

    unique_sites = sorted({item["siteName"] for item in cells_payload if item["siteName"]})
    unique_cells = sorted({item["cellName"] for item in cells_payload if item["cellName"]})
    tech_breakdown: dict = {}
    for (cn, sn, tech, band), count in primary_count_by_key.items():
        tech_breakdown[tech] = tech_breakdown.get(tech, 0) + count

    return {
        "available": True,
        "bddAvailable": True,
        "bddCellCount": len(bdd_cells),
        "gpsRows": len(gps_rows),
        "decisionBucketCount": decision_bucket_count,
        "unmatchedCount": unmatched_bucket_count,
        "matchedCount": matched_bucket_count,
        "matchedBucketCount": matched_bucket_count,
        "unmatchedBucketCount": unmatched_bucket_count,
        "unmatchedNoServingPciCount": unmatched_no_serving_pci_count,
        "unmatchedBddMissCount": unmatched_bdd_miss_count,
        "unmatchedBddMissBucketCount": unmatched_bdd_miss_bucket_count,
        "matchMethods": match_method_counts,
        "uniqueCellCount": len(unique_cells),
        "uniqueSiteCount": len(unique_sites),
        "uniqueSites": unique_sites,
        "uniqueCells": unique_cells,
        "techBreakdown": tech_breakdown,
        "radioPresenceBreakdown": radio_presence_breakdown_all,
        "radioPresenceBreakdownAll": radio_presence_breakdown_all,
        "radioPresenceBreakdownDownload": radio_presence_breakdown_download,
        "downloadScoped": download_scoped,
        "uploadScoped": upload_scoped,
        "downloadWindowStart": _nemo_fmt_hms_ms(download_window_start) if download_window_start else None,
        "downloadWindowEnd": _nemo_fmt_hms_ms(download_window_end) if download_window_end else None,
        "cells": cells_payload,
        "episodes": all_window_episode_payload,
        "episodesAll": all_window_episode_payload,
        "episodesDownload": download_window_episode_payload,
        "gpsTrace": gps_trace,
        "title": f"{op_name} Serving Cells (from BDD)",
        "title_fr": f"Cellules servantes {op_name} (depuis BDD)",
    }


def _nemo_attach_serving_cell_presence_metadata(
    cells: dict | None,
    technology_status: dict | None,
    dominant_nr_info: dict | None = None,
) -> dict | None:
    if not isinstance(cells, dict):
        return cells

    ts = technology_status or {}
    nr_presence_pct = ts.get("nrPresencePct")
    lte_only_presence_pct = ts.get("lteOnlyPresencePct")
    nr_presence_seconds = ts.get("nrPresenceSeconds")
    lte_only_seconds = ts.get("lteOnlySeconds")
    total_presence_seconds = ts.get("totalPresenceSeconds")

    cells["nrPresencePct"] = nr_presence_pct
    cells["lteOnlyPresencePct"] = lte_only_presence_pct
    cells["nrPresenceSeconds"] = nr_presence_seconds
    cells["lteOnlySeconds"] = lte_only_seconds
    cells["totalPresenceSeconds"] = total_presence_seconds

    matched_tech_breakdown = dict(cells.get("techBreakdown") or {})
    radio_presence_breakdown = dict(cells.get("radioPresenceBreakdown") or {})
    if not radio_presence_breakdown:
        if nr_presence_pct is not None:
            radio_presence_breakdown["5G"] = nr_presence_pct
        if lte_only_presence_pct is not None:
            radio_presence_breakdown["4G"] = lte_only_presence_pct

    has_matched_nr_cells = any(str(tech).upper().startswith(("5G", "NR")) for tech in matched_tech_breakdown)
    has_nr_presence = bool((nr_presence_seconds or 0) > 0 or (nr_presence_pct or 0) > 0)
    nr_detected_without_matched_nr_cell = bool(has_nr_presence and not has_matched_nr_cells)

    cells["matchedTechBreakdown"] = matched_tech_breakdown
    cells["radioPresenceBreakdown"] = radio_presence_breakdown
    cells["hasMatchedNrCells"] = has_matched_nr_cells
    cells["hasNrPresence"] = has_nr_presence
    cells["nrDetectedWithoutMatchedNrCell"] = nr_detected_without_matched_nr_cell
    cells["dominantNrInfo"] = dominant_nr_info

    if nr_detected_without_matched_nr_cell:
        nr_display = None
        if isinstance(dominant_nr_info, dict):
            nr_display = dominant_nr_info.get("display")
            if not nr_display:
                nr_band = str(dominant_nr_info.get("band") or "NR").strip()
                nr_pci = dominant_nr_info.get("pci")
                nr_arfcn = dominant_nr_info.get("arfcn")
                if nr_pci is not None and nr_arfcn is not None:
                    nr_display = f"{nr_band} PCI {nr_pci} / ARFCN {nr_arfcn}"
        if nr_display:
            nr_clause = f" The unmatched NR layer was observed as {nr_display}."
            nr_clause_fr = f" La couche NR non matchee a ete observee comme {nr_display}."
        else:
            nr_clause = ""
            nr_clause_fr = ""
        cells["servingTechMismatchNote"] = (
            "5G/EN-DC was detected in Nemo time-series data, but no matching 5G BDD cell was found for the serving-cell table. "
            "The table therefore lists only the matched LTE anchor cells."
            + nr_clause
        )
        cells["servingTechMismatchNoteFr"] = (
            "La 5G/EN-DC a ete detectee dans la serie temporelle Nemo, mais aucune cellule 5G du BDD n'a ete trouvee pour la table des cellules servantes. "
            "La table liste donc uniquement les ancres LTE matchees."
            + nr_clause_fr
        )

    return cells


def _nemo_merge_technology_status_with_serving_cells(
    technology_status: dict | None,
    serving_cells: dict | None,
    window_mode: str | None = None,
) -> dict:
    merged = dict(technology_status or {})
    if not isinstance(serving_cells, dict):
        return merged
    active_dl_only = _benchmark_nemo_normalize_window_mode(window_mode) == "active_dl_session"
    breakdown = dict(
        (
            serving_cells.get("radioPresenceBreakdownDownload")
            if active_dl_only
            else serving_cells.get("radioPresenceBreakdownAll")
        )
        or serving_cells.get("radioPresenceBreakdownAll")
        or serving_cells.get("radioPresenceBreakdown")
        or {}
    )
    if not breakdown:
        return merged
    nr_pct = breakdown.get("5G")
    lte_pct = breakdown.get("4G")
    if nr_pct is not None:
        merged["nrPresencePct"] = nr_pct
    if lte_pct is not None:
        merged["lteOnlyPresencePct"] = lte_pct
    exact_totals = _nemo_radio_presence_totals_from_cells(
        serving_cells.get("cells") or [],
        "dwellSecDownload" if active_dl_only else "dwellSec",
    )
    if exact_totals.get("5G") or exact_totals.get("4G"):
        merged["nrPresenceSeconds"] = exact_totals.get("5G", 0.0)
        merged["lteOnlySeconds"] = exact_totals.get("4G", 0.0)
        merged["totalPresenceSeconds"] = round(
            float(exact_totals.get("5G", 0.0)) + float(exact_totals.get("4G", 0.0)),
            0,
        )
    else:
        total_seconds = merged.get("totalPresenceSeconds")
        if total_seconds not in (None, 0):
            try:
                total_num = float(total_seconds)
                if nr_pct is not None:
                    merged["nrPresenceSeconds"] = round(total_num * float(nr_pct) / 100.0, 0)
                if lte_pct is not None:
                    merged["lteOnlySeconds"] = round(total_num * float(lte_pct) / 100.0, 0)
            except Exception:
                pass
    return merged


def _nemo_build_missing_kpi_quality(operators: list[dict]) -> dict:

    """Task 14 — per-operator KPI availability table (21 KPIs)."""
    KPI_DEFS = [
        ("App. rate DL",                    "dl",           "stat"),
        ("Transfer status",                  "transferStatus","raw_str"),
        ("RSRP",                             "rsrp",         "stat"),
        ("SINR",                             "sinr",         "stat"),
        ("RSRQ",                             "rsrq",         "stat"),
        ("WB CQI",                           "cqi",          "stat"),
        ("RI",                               "ri",           "stat"),
        ("PDSCH PRBs",                       "prbs",         "stat"),
        ("PDSCH DL scheduled throughput (5G)","scheduled5g", "stat"),
        ("PDSCH DL throughput (5G)",          "pdsch5g",     "stat"),
        ("MAC DL BLER",                       "bler",        "stat"),
        ("MAC UL retransmission rate (5G)",   "macUlRetx",   "stat"),
        ("NR CA status",                      "nrCaStatus",  "raw_str"),
        ("#SCells",                           "scellsCount", "raw_num"),
        ("PDSCH scheduled rank",              "scheduledRank","raw_num"),
        ("PDSCH MCS index 0/1",               "pdschMcs",    "stat"),
        ("PDSCH modulation 0/1",              "pdschModulation", "distribution"),
        ("PDSCH bit/s/Hz",                    "pdschBitPerHz","stat"),
        ("TCP handshake time",                "tcpHandshake","stat"),
        ("Lost packet",                       "lostPacket",  "stat"),
        ("Ping status",                       "pingStatus",  "raw_str"),
    ]
    REQUIRES_5G = {"PDSCH DL scheduled throughput (5G)", "PDSCH DL throughput (5G)",
                   "MAC DL BLER", "MAC UL retransmission rate (5G)", "NR CA status",
                   "#SCells", "PDSCH scheduled rank"}
    rows = []
    for item in operators or []:
        op_name = item.get("operator") or "UNKNOWN"
        kpis = item.get("kpis") or {}
        raw_rows = item.get("rows") or []
        has5g = bool(item.get("has5g"))
        for (label, key, kind) in KPI_DEFS:
            if kind == "stat":
                stat = kpis.get(key) or {}
                count = int(stat.get("sampleCount") or 0)
                avail = count > 0
                if not avail and label in REQUIRES_5G and not has5g:
                    comment = "Not applicable — no 5G detected in this operator's export."
                else:
                    comment = "" if avail else "Export this KPI from Nemo for stronger diagnosis."
                rows.append({"operator": op_name, "kpi": label, "available": avail, "sampleCount": count, "comment": comment})
            elif kind == "distribution":
                dist = kpis.get(key) or {}
                count = int(dist.get("sampleCount") or 0)
                avail = count > 0
                if not avail and label in REQUIRES_5G and not has5g:
                    comment = "Not applicable — no 5G detected in this operator's export."
                else:
                    comment = "" if avail else "Export more PDSCH samples to confirm modulation behavior."
                rows.append({"operator": op_name, "kpi": label, "available": avail, "sampleCount": count, "comment": comment})
            elif kind == "raw_str":
                count = sum(1 for r in raw_rows if r.get(key) not in (None, "", "N/A"))
                avail = count > 0
                if not avail and label in REQUIRES_5G and not has5g:
                    comment = "Not applicable — no 5G detected in this operator's export."
                else:
                    comment = "" if avail else "Export this KPI from Nemo for stronger diagnosis."
                rows.append({"operator": op_name, "kpi": label, "available": avail, "sampleCount": count, "comment": comment})
            elif kind == "raw_num":
                count = sum(1 for r in raw_rows if r.get(key) is not None and float(r.get(key) or 0) > 0)
                avail = count > 0
                if not avail and label in REQUIRES_5G and not has5g:
                    comment = "Not applicable — no 5G detected in this operator's export."
                else:
                    comment = "" if avail else "Export this KPI from Nemo for stronger diagnosis."
                rows.append({"operator": op_name, "kpi": label, "available": avail, "sampleCount": count, "comment": comment})
    return {
        "title": "Missing KPI / Export Quality",
        "available": bool(rows),
        "rows": rows,
    }


def _nemo_build_recommendations_by_priority(operators: list[dict], diagnosis: dict) -> dict:
    """Task 15 — ranked recommendations: P1 = top root cause, P2 = secondary, P3 = data gaps."""
    scores = (diagnosis or {}).get("scores") or []
    comparator_name = (diagnosis or {}).get("comparator") or "the best comparator"
    priorities = []

    cmp_fr = comparator_name  # French comparator name (same token)
    if scores:
        top = scores[0] if isinstance(scores[0], dict) else {}
        cause1 = top.get("cause") or ""
        steps1 = _nemo_recommendations_for_cause(cause1, comparator_name)
        steps1_fr = _nemo_recommendations_for_cause_fr(cause1, cmp_fr)
        priorities.append({
            "priority": 1,
            "label": "Highest root cause — primary action",
            "label_fr": "Cause principale — action prioritaire",
            "cause": cause1,
            "score": top.get("score"),
            "steps": steps1,
            "steps_fr": steps1_fr,
        })
        if len(scores) > 1:
            sec = scores[1] if isinstance(scores[1], dict) else {}
            cause2 = sec.get("cause") or ""
            steps2 = _nemo_recommendations_for_cause(cause2, comparator_name)
            steps2_fr = _nemo_recommendations_for_cause_fr(cause2, cmp_fr)
            if steps2:
                priorities.append({
                    "priority": 2,
                    "label": "Secondary root cause",
                    "label_fr": "Cause secondaire",
                    "cause": cause2,
                    "score": sec.get("score"),
                    "steps": steps2,
                    "steps_fr": steps2_fr,
                })

    # P3 — data gaps (missing high-value KPIs across all operators)
    missing_gap_kpis = []
    stat_checks = [
        ("PDSCH DL scheduled throughput (5G)", "scheduled5g"),
        ("PDSCH DL throughput (5G)",           "pdsch5g"),
        ("PDSCH MCS index 0/1",               "pdschMcs"),
        ("PDSCH bit/s/Hz",                    "pdschBitPerHz"),
        ("MAC UL retransmission rate (5G)",    "macUlRetx"),
        ("TCP handshake time",                 "tcpHandshake"),
        ("Lost packet",                        "lostPacket"),
    ]
    for item in operators or []:
        kpis = item.get("kpis") or {}
        for label, key in stat_checks:
            if int((kpis.get(key) or {}).get("sampleCount") or 0) == 0:
                if label not in missing_gap_kpis:
                    missing_gap_kpis.append(label)
    if missing_gap_kpis:
        priorities.append({
            "priority": 3,
            "label": "Data gaps — enrich export",
            "label_fr": "Lacunes de données — enrichir l'export",
            "cause": "Missing data",
            "score": None,
            "steps": [f"Export '{kpi}' from Nemo to enable deeper diagnosis." for kpi in missing_gap_kpis],
            "steps_fr": [f"Exporter '{kpi}' depuis Nemo pour approfondir le diagnostic." for kpi in missing_gap_kpis],
        })

    return {
        "title": "Recommendations by Priority",
        "available": bool(priorities),
        "comparator": comparator_name,
        "priorities": priorities,
    }


def _nemo_detect_test_type(rows: list[dict]) -> str:
    """Classify the DT as quasi_static or drive from GPS spread in the raw rows."""
    lats = [float(r["lat"]) for r in rows if r.get("lat") is not None]
    lons = [float(r["lon"]) for r in rows if r.get("lon") is not None]
    if not lats or not lons:
        return "unknown"
    lat_spread = max(lats) - min(lats)
    lon_spread = max(lons) - min(lons)
    # ~0.003 deg ≈ 330 m; below this threshold → quasi-static
    return "quasi_static" if lat_spread < 0.003 and lon_spread < 0.003 else "drive"


def _nemo_build_professional_executive_summary(
    operators: list[dict], diagnosis: dict, dt_name: str = ""
) -> dict:
    """Generate a professional 10-section executive summary for IAM benchmark DT analysis."""
    comparator_name = str((diagnosis or {}).get("best5gComparator") or "").strip()
    iam = next((item for item in operators if str(item.get("operator") or "").upper() == "IAM"), None)
    comparator = next((item for item in operators if str(item.get("operator") or "") == comparator_name), None)
    if not iam or not comparator_name or not comparator:
        return {
            "title": "Professional Executive Summary",
            "available": False,
            "message": "IAM operator or 5G comparator data is missing.",
        }

    iam_kpis = iam.get("kpis") or {}
    cmp_kpis = comparator.get("kpis") or {}
    iam_rows = iam.get("rows") or []

    # ── Input values ──────────────────────────────────────────────────────────
    test_type = _nemo_detect_test_type(iam_rows)

    iam_dl_tp    = _nemo_safe_round((iam_kpis.get("dl") or {}).get("average"), 1)
    cmp_dl_tp    = _nemo_safe_round((cmp_kpis.get("dl") or {}).get("average"), 1)
    dl_success   = iam_kpis.get("successRate")

    iam_5g_pres  = iam_kpis.get("nrPresencePct")
    cmp_5g_pres  = cmp_kpis.get("nrPresencePct")
    iam_4g_only  = iam_kpis.get("lteOnlyPresencePct")
    iam_n78      = float(iam_kpis.get("n78Share") or 0.0)
    cmp_n78      = float(cmp_kpis.get("n78Share") or 0.0)
    iam_avail    = (iam_kpis.get("availableBandwidthPrbs") or {}).get("average")
    cmp_avail    = (cmp_kpis.get("availableBandwidthPrbs") or {}).get("average")

    iam_rsrp     = (iam_kpis.get("rsrp") or {}).get("median")
    cmp_rsrp     = (cmp_kpis.get("rsrp") or {}).get("median")
    iam_sinr     = (iam_kpis.get("sinr") or {}).get("median")
    cmp_sinr     = (cmp_kpis.get("sinr") or {}).get("median")
    iam_cqi      = (iam_kpis.get("cqi") or {}).get("median")
    cmp_cqi      = (cmp_kpis.get("cqi") or {}).get("median")
    iam_mcs      = (iam_kpis.get("pdschMcs") or {}).get("median")
    cmp_mcs      = (cmp_kpis.get("pdschMcs") or {}).get("median")
    iam_dom_mod  = (iam_kpis.get("pdschModulation") or {}).get("dominant") or "—"
    cmp_dom_mod  = (cmp_kpis.get("pdschModulation") or {}).get("dominant") or "—"
    iam_bits_hz  = (iam_kpis.get("pdschBitPerHz") or {}).get("median")
    cmp_bits_hz  = (cmp_kpis.get("pdschBitPerHz") or {}).get("median")
    iam_prb_eff  = iam_kpis.get("prbEfficiency")
    cmp_prb_eff  = cmp_kpis.get("prbEfficiency")
    iam_slot     = (iam_kpis.get("pdschSlotPct") or {}).get("average")
    cmp_slot     = (cmp_kpis.get("pdschSlotPct") or {}).get("average")
    iam_alloc    = iam_kpis.get("resourceAllocationIndex")
    cmp_alloc    = cmp_kpis.get("resourceAllocationIndex")
    iam_bler_avg = (iam_kpis.get("bler") or {}).get("average")
    iam_bler_p90 = (iam_kpis.get("bler") or {}).get("p90")
    iam_bler_a10 = iam_kpis.get("blerAbove10Share")
    tcp_iam      = (iam_kpis.get("tcpHandshake") or {}).get("average")
    tcp_cmp      = (cmp_kpis.get("tcpHandshake") or {}).get("average")

    # ── Derived signals ───────────────────────────────────────────────────────
    dl_gap_pct = _nemo_gap_pct(iam_dl_tp, cmp_dl_tp)
    rf_not_worse = bool(
        iam_rsrp is not None and cmp_rsrp is not None and float(iam_rsrp) >= float(cmp_rsrp)
        and iam_sinr is not None and cmp_sinr is not None and float(iam_sinr) >= float(cmp_sinr)
        and iam_cqi is not None and cmp_cqi is not None and float(iam_cqi) >= float(cmp_cqi)
    )
    spectral_better = bool(
        iam_mcs is not None and cmp_mcs is not None and float(iam_mcs) > float(cmp_mcs)
        and iam_bits_hz is not None and cmp_bits_hz is not None and float(iam_bits_hz) > float(cmp_bits_hz)
    )
    low_5g_pres = bool(
        (iam_5g_pres is not None and cmp_5g_pres is not None and float(iam_5g_pres) < float(cmp_5g_pres) - 20)
        or (iam_4g_only is not None and float(iam_4g_only) > 70)
    )
    missing_n78  = bool(iam_n78 == 0.0 and cmp_n78 > 0)
    smaller_pool = bool(
        iam_avail is not None and cmp_avail is not None
        and float(iam_avail) < float(cmp_avail) * 0.5
    )
    sched_hyp = bool(
        iam_slot is not None and cmp_slot is not None and float(iam_slot) >= float(cmp_slot)
        and iam_alloc is not None and cmp_alloc is not None and float(iam_alloc) >= float(cmp_alloc)
        and iam_prb_eff is not None and cmp_prb_eff is not None and float(iam_prb_eff) >= float(cmp_prb_eff)
    )
    bler_critical = bool(
        (iam_bler_avg is not None and iam_bler_avg > 5)
        or (iam_bler_p90 is not None and iam_bler_p90 > 15)
        or (iam_bler_a10 is not None and iam_bler_a10 > 10)
    )
    bler_moderate = bool(
        not bler_critical
        and ((iam_bler_avg is not None and iam_bler_avg > 2)
             or (iam_bler_a10 is not None and iam_bler_a10 > 5))
    )
    transport_issue = bool(
        (tcp_iam is not None and tcp_iam > 70)
        or (tcp_iam is not None and tcp_cmp is not None and float(tcp_iam) > float(tcp_cmp) * 1.2)
    )
    static_test = test_type in ("static", "quasi_static")

    def _f(v, d=1):
        return str(_nemo_safe_round(v, d)) if v is not None else "N/A"

    # ── Sections ─────────────────────────────────────────────────────────────
    sections: list[dict] = []

    # 1 — Test context
    if static_test:
        ctx = (
            "The benchmark test was performed in an almost static condition, with nearly the same "
            "CGPS/location for the compared operators. This makes the comparison more reliable because "
            "the results are less affected by route variation, mobility, handover behavior, or changing "
            "radio environment. Therefore, the observed gaps are more likely related to network configuration, "
            "serving-layer selection, EN-DC behavior, NR band/BWP capacity, scheduler/resource configuration, "
            "or site-level configuration at the tested location."
        )
    elif test_type == "drive":
        ctx = (
            "The benchmark was performed in mobility conditions. The analysis should therefore consider "
            "route segments, serving-cell changes, handovers, coverage dominance, and time spent on each "
            "radio layer."
        )
    else:
        ctx = "Test conditions could not be automatically classified from GPS data."
    sections.append({"id": "testContext", "title": "Test Context", "text": ctx})

    # 2 — Accessibility & throughput
    ap: list[str] = []
    if dl_success is not None and dl_success >= 98:
        ap.append(
            f"IAM shows good DL service accessibility, with {_f(dl_success, 1)}% DL success rate. "
            "Therefore, the issue is not related to basic download accessibility or service availability."
        )
    if dl_gap_pct is not None and dl_gap_pct <= -30:
        ap.append(
            f"However, IAM DL throughput remains significantly below the best competitor: "
            f"{_f(iam_dl_tp, 1)} Mbps versus {_f(cmp_dl_tp, 1)} Mbps for {comparator_name}."
        )
    elif dl_gap_pct is not None and dl_gap_pct > -10:
        ap.append(
            f"IAM DL throughput is globally close to the best competitor: "
            f"{_f(iam_dl_tp, 1)} Mbps versus {_f(cmp_dl_tp, 1)} Mbps for {comparator_name}."
        )
    elif iam_dl_tp is not None and cmp_dl_tp is not None:
        ap.append(
            f"IAM DL throughput ({_f(iam_dl_tp, 1)} Mbps) is below {comparator_name} ({_f(cmp_dl_tp, 1)} Mbps)."
        )
    if ap:
        sections.append({"id": "accessibility", "title": "DL Accessibility & Throughput", "text": " ".join(ap)})

    # 3 — Effective 5G presence
    if iam_5g_pres is not None or low_5g_pres:
        if low_5g_pres:
            pres = (
                f"The detailed DL-session analysis shows limited effective 5G contribution. "
                f"IAM spends only {_f(iam_5g_pres, 1)}% of the time on 5G/EN-DC, compared with "
                f"{_f(cmp_5g_pres, 1)}% for {comparator_name}"
                + (f", while IAM remains {_f(iam_4g_only, 1)}% of the time on 4G-only." if iam_4g_only is not None else ".")
            )
            if static_test:
                pres += (
                    " Since the test location is almost static, this strongly indicates a local "
                    "5G layer selection, EN-DC anchoring, or NR eligibility issue at the tested point."
                )
        else:
            pres = (
                f"IAM 5G/EN-DC presence is {_f(iam_5g_pres, 1)}%, "
                f"compared with {_f(cmp_5g_pres, 1)}% for {comparator_name}."
            )
        sections.append({"id": "fivegPresence", "title": "Effective 5G Presence", "text": pres})

    # 4 — NR capacity layer
    nr_parts: list[str] = []
    if missing_n78:
        nr_parts.append(
            f"IAM shows 0% n78 contribution, while {comparator_name} benefits from n78 "
            f"({_f(cmp_n78, 1)}% share). This confirms a missing high-capacity 5G layer."
        )
    if smaller_pool:
        nr_parts.append(
            f"IAM has only {_f(iam_avail, 0)} available NR PRBs, while {comparator_name} has "
            f"{_f(cmp_avail, 0)} PRBs. This indicates a smaller NR BWP/carrier capacity and "
            "supports the NR resource-pool limitation diagnosis."
        )
    if nr_parts:
        sections.append({"id": "nrCapacity", "title": "NR Capacity Layer", "text": " ".join(nr_parts)})

    # 5 — RF quality & modulation
    rf_parts: list[str] = []
    if rf_not_worse:
        rf_parts.append(
            f"From a radio-quality perspective, IAM is not worse than {comparator_name}. "
            f"IAM RSRP {_f(iam_rsrp, 1)} dBm vs {_f(cmp_rsrp, 1)} dBm, "
            f"SINR {_f(iam_sinr, 1)} dB vs {_f(cmp_sinr, 1)} dB, "
            f"CQI {_f(iam_cqi, 1)} vs {_f(cmp_cqi, 1)}."
        )
    if spectral_better:
        rf_parts.append(
            f"IAM also shows better modulation and spectral efficiency: dominant modulation is "
            f"{iam_dom_mod}, compared with {cmp_dom_mod} for {comparator_name}, median MCS is "
            f"{_f(iam_mcs, 0)} versus {_f(cmp_mcs, 0)}, and median PDSCH spectral efficiency is "
            f"{_f(iam_bits_hz, 2)} bit/s/Hz versus {_f(cmp_bits_hz, 2)} bit/s/Hz."
        )
    if rf_not_worse or spectral_better:
        rf_parts.append(
            "Therefore, poor RF quality, poor modulation, or MIMO rank limitation should not be "
            "considered the primary root cause."
        )
    if rf_parts:
        sections.append({"id": "rfModulation", "title": "RF Quality & Modulation", "text": " ".join(rf_parts)})

    # 6 — Scheduler / PRB allocation
    if sched_hyp:
        sched_t = (
            f"PRB allocation or scheduler limitation is not confirmed at this stage. "
            f"IAM PDSCH slot usage ({_f(iam_slot, 1)}%), allocation ratio ({_f(iam_alloc, 1)}%), "
            f"and PRB efficiency ({_f(iam_prb_eff, 3)} Mbps/PRB) are not worse than {comparator_name} "
            f"({_f(cmp_slot, 1)}%, {_f(cmp_alloc, 1)}%, {_f(cmp_prb_eff, 3)} Mbps/PRB). "
            "Scheduler/PRB limitation should remain a hypothesis until additional data such as PRB "
            "utilization, cell load, scheduler grants, QoS profile, BWP configuration, OLLA, CQI aging, "
            "and PDSCH power are available."
        )
    else:
        sched_t = (
            "Scheduler/resource allocation requires deeper verification because IAM scheduling "
            f"indicators are weaker than {comparator_name}."
        )
    sections.append({"id": "scheduler", "title": "Scheduler & PRB Allocation", "text": sched_t})

    # 7 — BLER
    if bler_critical:
        sections.append({
            "id": "bler", "title": "BLER / Retransmissions",
            "text": (
                f"A secondary point of attention is BLER. IAM BLER average is {_f(iam_bler_avg, 1)}%, "
                f"P90 is {_f(iam_bler_p90, 1)}%"
                + (f", and BLER >10% share is {_f(iam_bler_a10, 1)}%" if iam_bler_a10 is not None else "")
                + ". This indicates retransmission zones that should be localized."
            ),
        })
    elif bler_moderate:
        sections.append({
            "id": "bler", "title": "BLER / Retransmissions",
            "text": (
                f"A secondary point of attention is BLER. IAM BLER average is {_f(iam_bler_avg, 1)}%"
                + (f" with BLER >10% share around {_f(iam_bler_a10, 1)}%" if iam_bler_a10 is not None else "")
                + ". This does not indicate a critical route-wide BLER issue, but it suggests localized "
                "or short-duration retransmission zones."
            ),
        })

    # ── Root cause table ──────────────────────────────────────────────────────
    root_causes: list[dict] = []
    if low_5g_pres:
        root_causes.append({
            "priority": "P1",
            "rootCause": "Low effective 5G/EN-DC presence",
            "evidence": (
                f"IAM {_f(iam_5g_pres, 1)}% 5G time vs {comparator_name} {_f(cmp_5g_pres, 1)}%"
                + (f" — 4G-only {_f(iam_4g_only, 1)}%" if iam_4g_only is not None else "")
            ),
            "status": "Confirmed" if static_test else "Suspected",
        })
    if missing_n78:
        root_causes.append({
            "priority": "P1",
            "rootCause": "Missing n78 high-capacity NR layer",
            "evidence": f"IAM 0% n78 vs {comparator_name} {_f(cmp_n78, 1)}% n78",
            "status": "Confirmed",
        })
    if smaller_pool:
        root_causes.append({
            "priority": "P1",
            "rootCause": "Smaller NR resource pool / BWP capacity",
            "evidence": f"IAM {_f(iam_avail, 0)} PRBs available vs {comparator_name} {_f(cmp_avail, 0)} PRBs",
            "status": "Confirmed",
        })
    if transport_issue:
        root_causes.append({
            "priority": "P2",
            "rootCause": "Transport / core network latency",
            "evidence": f"TCP handshake IAM {_f(tcp_iam, 0)} ms" + (f" vs {comparator_name} {_f(tcp_cmp, 0)} ms" if tcp_cmp is not None else ""),
            "status": "Suspected",
        })
    if bler_critical or bler_moderate:
        root_causes.append({
            "priority": "P2",
            "rootCause": "BLER / HARQ retransmissions",
            "evidence": (
                f"BLER avg {_f(iam_bler_avg, 1)}%"
                + (f", P90 {_f(iam_bler_p90, 1)}%" if iam_bler_p90 is not None else "")
                + (f", >10% share {_f(iam_bler_a10, 1)}%" if iam_bler_a10 is not None else "")
            ),
            "status": "Secondary — localize by route segment",
        })
    root_causes.append({
        "priority": "P3",
        "rootCause": "Scheduler / PRB allocation limitation",
        "evidence": (
            f"IAM slot% {_f(iam_slot, 1)}% ≥ {comparator_name} {_f(cmp_slot, 1)}%; "
            f"alloc ratio {_f(iam_alloc, 1)}% ≥ {comparator_name} {_f(cmp_alloc, 1)}%"
        ) if sched_hyp else f"Scheduling indicators weaker than {comparator_name}",
        "status": "Hypothesis — not confirmed" if sched_hyp else "Suspected",
    })

    # ── Recommended actions ───────────────────────────────────────────────────
    actions: list[dict] = []
    if low_5g_pres:
        actions.append({
            "priority": "P1",
            "domain": "EN-DC / 5G anchoring",
            "action": "Verify EN-DC addition/release, LTE anchor eligibility, NR PSCell selection, and NSA thresholds.",
        })
    if missing_n78:
        actions.append({
            "priority": "P1",
            "domain": "NR band configuration",
            "action": "Check n78 deployment, availability, barring status, priority, NR neighbors, B1/B2 events, and LTE anchor relation.",
        })
    if smaller_pool:
        actions.append({
            "priority": "P1",
            "domain": "NR BWP / carrier capacity",
            "action": "Verify active NR band, BWP size, configured/active PRBs, and carrier capacity versus competitors.",
        })
    if bler_critical or bler_moderate:
        actions.append({
            "priority": "P2",
            "domain": "Link adaptation / HARQ",
            "action": "Correlate BLER >10% samples with SINR, CQI, MCS, modulation, HARQ, serving cell, NR PSCell, and PDSCH power.",
        })
    if transport_issue:
        actions.append({
            "priority": "P2",
            "domain": "Transport / core",
            "action": "Retest with same UE, SIM, server, and script. If confirmed, check backhaul, APN/DNS, firewall/NAT, and core routing.",
        })
    actions.append({
        "priority": "P3",
        "domain": "Scheduler / RAN configuration",
        "action": "Add PRB utilization, scheduler grants, cell load, QoS profile, OLLA, CQI aging, and PDSCH power before confirming scheduler action.",
    })
    actions.append({
        "priority": "—",
        "domain": "Retest governance",
        "action": "Repeat under busy-hour and off-peak conditions using the same UE, SIM profile, server, script, and location/route.",
    })

    # ── Final diagnosis ───────────────────────────────────────────────────────
    p1_triggered = low_5g_pres or missing_n78 or smaller_pool
    fd_parts: list[str] = []
    if static_test and p1_triggered:
        fd_parts.append(
            "Because the benchmark was almost static, the observed differences are unlikely to be mainly "
            "caused by mobility or route variation."
            + (" IAM has acceptable accessibility and RF/modulation efficiency at the tested point, but it"
               if (rf_not_worse or spectral_better) else " IAM")
            + " does not benefit from the same effective high-capacity 5G layer as the competitor. "
            "The dominant issue is therefore a local 5G capacity-layer limitation: "
            + (f"low effective EN-DC presence ({_f(iam_5g_pres, 1)}%), " if low_5g_pres else "")
            + ("missing n78 contribution, " if missing_n78 else "")
            + (f"and smaller NR resource pool/BWP capacity ({_f(iam_avail, 0)} vs {_f(cmp_avail, 0)} available PRBs)." if smaller_pool else ".")
        )
    elif p1_triggered:
        fd_parts.append(
            "The analysis identifies a 5G capacity-layer limitation as the primary driver of IAM's lower "
            "DL throughput. "
            + (f"IAM spends only {_f(iam_5g_pres, 1)}% of the time on 5G/EN-DC. " if low_5g_pres else "")
            + (f"Missing n78 contribution and a smaller NR resource pool ({_f(iam_avail, 0)} vs "
               f"{_f(cmp_avail, 0)} PRBs) are confirmed root causes versus {comparator_name}."
               if missing_n78 and smaller_pool else
               ("Missing n78 contribution is a confirmed root cause." if missing_n78 else
                f"Smaller NR resource pool ({_f(iam_avail, 0)} vs {_f(cmp_avail, 0)} PRBs) is a confirmed root cause versus {comparator_name}."
                if smaller_pool else ""))
        )
    else:
        fd_parts.append(
            "No dominant NR capacity-layer limitation was confirmed from the available KPIs. "
            "Deeper investigation using RAN counters, cell configuration, and scheduled carrier data is required."
        )
    if sched_hyp:
        fd_parts.append(
            "Scheduler limitation is not confirmed by the current indicators and should remain secondary "
            "until additional RAN scheduling counters are available."
        )
    sections.append({
        "id": "finalDiagnosis",
        "title": "Final Diagnosis",
        "text": " ".join(fd_parts),
    })

    return {
        "title": "Professional Executive Summary",
        "available": True,
        "dtName": dt_name,
        "testType": test_type,
        "comparatorOperator": comparator_name,
        "sections": sections,
        "rootCauses": root_causes,
        "recommendedActions": actions,
    }


def _nemo_build_ema_executive_summary(
    operators: list[dict],
    pes: dict,
    transfer_summary: list[dict],
) -> dict:
    """Dynamic 7-paragraph EMA executive summary template."""
    try:
        if not operators:
            return {"available": False}

        transfer_lookup: dict = {}
        for entry in (transfer_summary or []):
            op = str(entry.get("operator") or "").upper()
            direction = str(entry.get("direction") or "")
            if "down" in direction.lower() or direction.upper() in ("DL", "DOWNLINK"):
                dnorm = "DL"
            elif "up" in direction.lower() or direction.upper() in ("UL", "UPLINK"):
                dnorm = "UL"
            else:
                dnorm = direction.upper()
            transfer_lookup[(op, dnorm)] = entry

        iam_entry = next(
            (op for op in operators if str(op.get("operator") or "").upper() in _DEEP_IAM_ALIASES),
            None,
        )
        if not iam_entry:
            return {"available": False}

        iam = _deep_extract(iam_entry.get("kpis"), transfer_lookup, iam_entry.get("operator"))

        comps = []
        for op in operators:
            if str(op.get("operator") or "").upper() in _DEEP_IAM_ALIASES:
                continue
            flat = _deep_extract(op.get("kpis"), transfer_lookup, op.get("operator"))
            flat["_name"] = op.get("operator") or "Competitor"
            comps.append(flat)

        if not comps:
            return {"available": False}

        best_cmp = max(comps, key=lambda c: float(c.get("dlThroughput") or 0))
        best_name = best_cmp["_name"]

        test_type = (pes or {}).get("testType", "unknown")
        same_cgps = test_type in ("quasi_static", "static")

        def fmt(v, digits=1):
            if v is None:
                return "N/A"
            n = float(v)
            if abs(n - round(n)) < 1e-9 and digits == 0:
                return str(int(round(n)))
            if abs(n - round(n)) < 1e-9:
                return str(int(round(n)))
            return str(round(n, digits))

        def pct_gap(a, b):
            if a is None or b in (None, 0):
                return None
            return round((float(a) - float(b)) / abs(float(b)) * 100.0, 1)

        def dominant_mod(flat):
            mods = {
                "256QAM": float(flat.get("qam256") or 0),
                "64QAM": float(flat.get("qam64") or 0),
                "16QAM": float(flat.get("qam16") or 0),
                "QPSK": float(flat.get("qpsk") or 0),
            }
            best = max(mods, key=mods.get)
            return best if mods[best] > 0 else "Unknown"

        iam_dl = iam.get("dlThroughput")
        cmp_dl = best_cmp.get("dlThroughput")
        dl_gap = pct_gap(iam_dl, cmp_dl)

        # ── Paragraph 1: Benchmark context ──
        if same_cgps:
            p1 = (
                "The benchmark test was performed in an almost static condition, with nearly the same "
                "CGPS/location for the compared operators. This makes the comparison more reliable because "
                "the results are less affected by route variation, mobility, handover behavior, or changing "
                "radio environment. Therefore, the observed performance gaps are more likely related to "
                "network configuration, serving-layer selection, 5G/EN-DC anchoring, NR band/BWP capacity, "
                "and resource availability at the tested location."
            )
        elif test_type == "drive":
            p1 = (
                "The benchmark test was performed in mobility conditions. Therefore, the comparison must "
                "consider route variation, serving-cell changes, handovers, coverage dominance, and time "
                "spent on each radio layer. Performance gaps may be influenced by both radio conditions and "
                "mobility behavior along the route."
            )
        else:
            p1 = (
                "The benchmark context is not fully classified as static or mobility-based. Therefore, the "
                "conclusions should be interpreted with care and validated using GPS stability, serving-cell "
                "sequence, and active DL-window analysis."
            )

        # ── Paragraph 2: Accessibility and throughput ──
        iam_dl_success = iam.get("dlSuccess")
        iam_dl_completion = iam.get("dlCompletion")
        p2_parts = []
        if (
            iam_dl_success is not None and float(iam_dl_success) >= 98
            and iam_dl_completion is not None and float(iam_dl_completion) >= 98
        ):
            p2_parts.append(
                f"IAM shows good DL service accessibility, with {fmt(iam_dl_success)}% DL success "
                f"and {fmt(iam_dl_completion)}% DL completion."
            )
        if dl_gap is not None:
            if abs(dl_gap) <= 10:
                p2_parts.append(
                    f"IAM DL throughput is globally close to the best competitor in this DT, with IAM at "
                    f"{fmt(iam_dl)} Mbps versus {best_name} at {fmt(cmp_dl)} Mbps. Therefore, the issue is "
                    f"not related to basic download accessibility or service availability."
                )
            elif dl_gap < -30:
                p2_parts.append(
                    f"However, IAM DL throughput is significantly below the best competitor, with IAM at "
                    f"{fmt(iam_dl)} Mbps versus {best_name} at {fmt(cmp_dl)} Mbps, corresponding to a gap "
                    f"of {fmt(dl_gap)}%. The issue is therefore linked to performance efficiency rather than "
                    f"service accessibility."
                )
            else:
                p2_parts.append(
                    f"IAM DL throughput shows a moderate gap versus the best competitor, with IAM at "
                    f"{fmt(iam_dl)} Mbps versus {best_name} at {fmt(cmp_dl)} Mbps, corresponding to a gap "
                    f"of {fmt(dl_gap)}%."
                )
        p2 = " ".join(p2_parts) if p2_parts else None

        # ── Paragraph 3: 5G/EN-DC presence ──
        iam_5g = iam.get("fivegPresence")
        cmp_5g = best_cmp.get("fivegPresence")
        iam_4g = iam.get("fourgOnly")
        iam_n78 = iam.get("n78")
        cmp_n78 = best_cmp.get("n78")
        show_5g_issue = (
            (iam_5g is not None and cmp_5g is not None and float(iam_5g) < float(cmp_5g) - 20)
            or (iam_4g is not None and float(iam_4g) > 70)
            or (iam_n78 is not None and float(iam_n78) == 0)
        )
        p3 = None
        if show_5g_issue:
            parts3 = [
                f"However, the detailed DL-session analysis shows that IAM performance is not driven by an "
                f"effective high-capacity 5G layer. IAM spends only {fmt(iam_5g)}% of the time on 5G/EN-DC, "
                f"compared with {fmt(cmp_5g)}% for {best_name}, while IAM remains {fmt(iam_4g)}% of the time "
                f"on 4G-only."
            ]
            if iam_n78 is not None and float(iam_n78) == 0 and cmp_n78 is not None and float(cmp_n78) > 0:
                parts3.append(
                    f"In addition, IAM shows {fmt(iam_n78)}% n78 contribution, while {best_name} benefits "
                    f"from {fmt(cmp_n78)}% n78 contribution."
                )
            if same_cgps:
                parts3.append(
                    "Since the test location is almost static, this strongly indicates a local 5G layer "
                    "selection, EN-DC anchoring, or n78 eligibility issue at the tested point, rather than "
                    "a mobility-driven limitation."
                )
            p3 = " ".join(parts3)

        # ── Paragraph 4: RF quality and modulation ──
        iam_rsrp = iam.get("rsrp")
        cmp_rsrp = best_cmp.get("rsrp")
        iam_sinr = iam.get("sinr")
        cmp_sinr = best_cmp.get("sinr")
        iam_cqi = iam.get("cqi")
        cmp_cqi = best_cmp.get("cqi")
        iam_mcs = iam.get("mcs")
        cmp_mcs = best_cmp.get("mcs")
        iam_bits = iam.get("pdschBitPerHz")
        cmp_bits = best_cmp.get("pdschBitPerHz")
        iam_dom = dominant_mod(iam)
        cmp_dom = dominant_mod(best_cmp)
        rf_not_worse = (
            iam_rsrp is not None and cmp_rsrp is not None and float(iam_rsrp) >= float(cmp_rsrp)
            and iam_sinr is not None and cmp_sinr is not None and float(iam_sinr) >= float(cmp_sinr)
            and iam_cqi is not None and cmp_cqi is not None and float(iam_cqi) >= float(cmp_cqi)
        )
        mod_better = (
            iam_mcs is not None and cmp_mcs is not None and float(iam_mcs) > float(cmp_mcs)
            and iam_bits is not None and cmp_bits is not None and float(iam_bits) > float(cmp_bits)
        )
        if rf_not_worse and mod_better:
            p4 = (
                f"From a radio-quality perspective, IAM is not worse than {best_name}. IAM has stronger "
                f"serving coverage and quality, with RSRP around {fmt(iam_rsrp, 2)} dBm versus "
                f"{fmt(cmp_rsrp, 2)} dBm, SINR {fmt(iam_sinr)} dB versus {fmt(cmp_sinr)} dB, and CQI "
                f"{fmt(iam_cqi)} versus {fmt(cmp_cqi)}. IAM also shows better modulation and spectral "
                f"efficiency: dominant modulation is {iam_dom}, compared with {cmp_dom} for {best_name}, "
                f"median MCS is {fmt(iam_mcs)} versus {fmt(cmp_mcs)}, and median PDSCH spectral efficiency "
                f"is {fmt(iam_bits, 2)} bit/s/Hz versus {fmt(cmp_bits, 2)} bit/s/Hz. Therefore, poor RF "
                f"quality, poor modulation, or MIMO rank limitation should not be considered the primary "
                f"root cause."
            )
        elif not rf_not_worse:
            p4 = (
                f"Radio-quality indicators require attention. IAM RSRP is {fmt(iam_rsrp, 2)} dBm versus "
                f"{fmt(cmp_rsrp, 2)} dBm, SINR is {fmt(iam_sinr)} dB versus {fmt(cmp_sinr)} dB, and CQI "
                f"is {fmt(iam_cqi)} versus {fmt(cmp_cqi)}. The RF layer should be checked before confirming "
                f"capacity or scheduler root causes."
            )
        else:
            p4 = (
                f"IAM RF quality is acceptable versus {best_name}, but modulation/spectral efficiency "
                f"requires deeper review. IAM dominant modulation is {iam_dom}, median MCS is "
                f"{fmt(iam_mcs)}, and spectral efficiency is {fmt(iam_bits, 2)} bit/s/Hz."
            )

        # ── Paragraph 5: NR resource pool / BWP ──
        iam_prbs = iam.get("availableBandwidthPrbs")
        cmp_prbs = best_cmp.get("availableBandwidthPrbs")
        p5 = None
        if (
            iam_prbs is not None and cmp_prbs is not None
            and float(cmp_prbs) > 0 and float(iam_prbs) < float(cmp_prbs) * 0.5
        ):
            parts5 = [
                f"The strongest technical root cause is the NR resource pool limitation. IAM has only "
                f"{fmt(iam_prbs)} available NR PRBs, while {best_name} has {fmt(cmp_prbs)} PRBs. "
                f"This indicates a smaller NR BWP/carrier capacity, most likely due to the absence of "
                f"n78 contribution."
            ]
            if same_cgps:
                parts5.append(
                    "In a static benchmark, this is a strong indicator that IAM is not selecting or not "
                    "eligible for the same high-capacity NR layer at that CGPS."
                )
            p5 = " ".join(parts5)

        # ── Paragraph 6: Scheduler / PRB allocation confidence ──
        iam_slot = iam.get("pdschSlotPct")
        cmp_slot = best_cmp.get("pdschSlotPct")
        iam_alloc = iam.get("resourceAllocationIndex")
        cmp_alloc = best_cmp.get("resourceAllocationIndex")
        iam_prb_eff = iam.get("prbEfficiency")
        cmp_prb_eff = best_cmp.get("prbEfficiency")
        scheduler_not_confirmed = (
            iam_slot is not None and cmp_slot is not None and float(iam_slot) >= float(cmp_slot)
            and iam_alloc is not None and cmp_alloc is not None and float(iam_alloc) >= float(cmp_alloc)
            and iam_prb_eff is not None and cmp_prb_eff is not None and float(iam_prb_eff) >= float(cmp_prb_eff)
        )
        if scheduler_not_confirmed:
            p6 = (
                f"PRB allocation or scheduler limitation is not confirmed at this stage. IAM has PDSCH "
                f"slot usage of {fmt(iam_slot)}% versus {fmt(cmp_slot)}% for {best_name}, and an "
                f"allocation ratio of {fmt(iam_alloc)}% versus {fmt(cmp_alloc)}%. IAM PRB efficiency is "
                f"also higher, at {fmt(iam_prb_eff, 2)} Mbps/PRB versus {fmt(cmp_prb_eff, 2)} Mbps/PRB. "
                f"These indicators do not support a pure scheduler allocation issue. Scheduler/PRB "
                f"limitation should remain a hypothesis until additional data such as PRB utilization, "
                f"cell load, scheduler grants, QoS profile, BWP configuration, OLLA, CQI aging, and "
                f"PDSCH power are available."
            )
        else:
            p6 = (
                f"Scheduler and PRB allocation require deeper verification. IAM PDSCH slot usage is "
                f"{fmt(iam_slot)}% versus {fmt(cmp_slot)}%, allocation ratio is {fmt(iam_alloc)}% versus "
                f"{fmt(cmp_alloc)}%, and PRB efficiency is {fmt(iam_prb_eff, 2)} Mbps/PRB versus "
                f"{fmt(cmp_prb_eff, 2)} Mbps/PRB. Additional PRB utilization, scheduler grants, QoS "
                f"profile, cell load, and PDSCH power data are required before confirming the exact "
                f"scheduler contribution."
            )

        # ── Paragraph 7: BLER ──
        iam_bler_avg = iam.get("blerAvg")
        cmp_bler_avg = best_cmp.get("blerAvg")
        iam_bler_p90 = iam.get("blerP90")
        iam_bler_above10 = iam.get("blerAbove10")
        p7 = None
        if iam_bler_avg is not None:
            ba, bp90, bab10 = float(iam_bler_avg), float(iam_bler_p90 or 0), float(iam_bler_above10 or 0)
            if ba > 5 or bp90 > 15 or bab10 > 10:
                p7 = (
                    f"A secondary point of attention is BLER. IAM BLER average is {fmt(ba, 2)}%, higher "
                    f"than {best_name} at {fmt(cmp_bler_avg, 2)}%, with BLER P90 at {fmt(bp90)}% and "
                    f"BLER >10% share around {fmt(bab10)}%. This indicates retransmission zones that "
                    f"should be checked by correlating BLER with SINR, CQI, MCS, modulation, HARQ, "
                    f"serving cell, NR PSCell, and PDSCH power."
                )
            elif ba > 2 or bab10 > 5:
                p7 = (
                    f"A secondary point of attention is BLER. IAM BLER average is {fmt(ba, 2)}%, higher "
                    f"than {best_name} at {fmt(cmp_bler_avg, 2)}%, with BLER >10% share around "
                    f"{fmt(bab10)}%. This does not indicate a critical route-wide BLER issue, but it "
                    f"suggests localized or short-duration retransmission zones that should be checked "
                    f"at the same CGPS by correlating BLER with SINR, CQI, MCS, modulation, HARQ, "
                    f"serving cell, and NR PSCell."
                )

        if test_type in ("static", "quasi_static"):
            type_label = "Static"
        elif test_type == "drive":
            type_label = "Drive"
        else:
            type_label = "Benchmark"

        title = f"Executive Summary — IAM {type_label} Benchmark DL Session Analysis"
        paragraphs = [p for p in [p1, p2, p3, p4, p5, p6, p7] if p]
        return {
            "available": True,
            "title": title,
            "testType": test_type,
            "comparatorName": best_name,
            "paragraphs": paragraphs,
        }
    except Exception:
        import traceback as _tb
        _tb.print_exc()
        return {"available": False}


def _nemo_build_executive_conclusion(operators: list[dict], ranking: list[dict], diagnosis: dict) -> dict:
    """Task 16 — templated final executive conclusion."""
    iam = next((item for item in operators if str(item.get("operator") or "").upper() == "IAM"), None)
    iam_kpis = (iam.get("kpis") or {}) if iam else {}
    iam_entry = next((e for e in ranking if str(e.get("operator") or "").upper() == "IAM"), None)
    best_dl_entry = ranking[0] if ranking else None
    best_dl_op = (best_dl_entry or {}).get("operator") or "—"
    iam_rank = (iam_entry or {}).get("rank")
    iam_avg = (iam_entry or {}).get("avgDlMbps")
    gap_pct = (diagnosis or {}).get("gapToBestDlPct")
    main_cause = (diagnosis or {}).get("mainCause") or "—"
    display_main_cause = (diagnosis or {}).get("displayMainCause") or main_cause
    severity = (diagnosis or {}).get("severity") or "—"
    best5g = (diagnosis or {}).get("best5gComparator") or ""
    comparator_name = (diagnosis or {}).get("comparator") or best5g or "the best comparator"
    iam_nr_info = _nemo_dominant_nr_serving_info((iam or {}).get("rows") or [])
    has_5g = bool(iam.get("has5g")) if iam else False

    # PRB, scheduled, delivered, and efficiency gaps
    cmp_item = next((item for item in operators if item.get("operator") == comparator_name), None)
    cmp_kpis = (cmp_item.get("kpis") or {}) if cmp_item else {}
    cmp_nr_info = _nemo_dominant_nr_serving_info((cmp_item or {}).get("rows") or [])
    prb_gap = _nemo_gap_pct(
        (iam_kpis.get("prbs") or {}).get("average"),
        (cmp_kpis.get("prbs") or {}).get("average"),
    )
    sched_gap = _nemo_gap_pct(
        (iam_kpis.get("scheduled5g") or {}).get("average"),
        (cmp_kpis.get("scheduled5g") or {}).get("average"),
    )
    delivered_gap = _nemo_gap_pct(
        (iam_kpis.get("pdsch5g") or {}).get("average"),
        (cmp_kpis.get("pdsch5g") or {}).get("average"),
    )
    iam_prb_eff = iam_kpis.get("prbEfficiency")
    cmp_prb_eff = cmp_kpis.get("prbEfficiency")
    iam_rsrp = (iam_kpis.get("rsrp") or {}).get("median")
    iam_sinr = (iam_kpis.get("sinr") or {}).get("median")
    rf_acceptable = (iam_rsrp is not None and iam_rsrp >= -100 and iam_sinr is not None and iam_sinr >= 3)
    iam_nr_band = str((iam_nr_info or {}).get("band") or "").lower()
    cmp_nr_band = str((cmp_nr_info or {}).get("band") or "").lower()
    lower_capacity_band_gap = bool(cmp_nr_band == "n78" and iam_nr_band and iam_nr_band != "n78")

    # PDSCH sample count — used to flag low-evidence caveats
    iam_rows = (iam or {}).get("rows") or []
    pdsch_sched_samples = sum(
        1 for r in iam_rows
        if r.get("pdschSched5gMbps") is not None and float(r.get("pdschSched5gMbps") or 0) > 0
    )
    low_pdsch_samples = has_5g and pdsch_sched_samples < 15
    sample_caveat = (
        f" Note: only {pdsch_sched_samples} PDSCH scheduled sample(s) detected for IAM in this export — "
        "5G radio KPIs have low statistical weight. This result should be treated as a scenario diagnosis, "
        "not a statistically representative network conclusion."
        if low_pdsch_samples and pdsch_sched_samples > 0
        else (
            " Note: no PDSCH scheduled samples were detected for IAM — 5G radio layer analysis is not possible from this export."
            if low_pdsch_samples and pdsch_sched_samples == 0 else ""
        )
    )
    sample_caveat_fr = (
        f" Note : seulement {pdsch_sched_samples} échantillon(s) PDSCH planifié(s) détecté(s) pour IAM dans cet export — "
        "les KPIs radio 5G ont un poids statistique faible. Ce résultat doit être traité comme un diagnostic de scénario, "
        "pas une conclusion réseau statistiquement représentative."
        if low_pdsch_samples and pdsch_sched_samples > 0
        else (
            " Note : aucun échantillon PDSCH planifié détecté pour IAM — l'analyse radio 5G n'est pas possible depuis cet export."
            if low_pdsch_samples and pdsch_sched_samples == 0 else ""
        )
    )

    # Compose the conclusion using the resource-allocation chain framing
    rank_str = _nemo_ordinal(iam_rank) if iam_rank else "—"
    gap_abs = abs(gap_pct) if gap_pct is not None else None
    gap_display = f"{gap_abs}%" if gap_abs is not None else "N/A"

    five_g_caveat = ""
    if not has_5g:
        five_g_caveat = " IAM's 5G availability was not detected in this export, so 5G-specific KPIs cannot be compared."
    elif not best5g:
        five_g_caveat = " No other operator had confirmed 5G in this export, so the 5G comparison is limited."

    # Resource-allocation chain framing (from deep analysis)
    if main_cause == "Scheduler / resource allocation" and prb_gap is not None and prb_gap < -20:
        prb_txt = f"{abs(prb_gap)}% fewer PDSCH PRBs" if prb_gap is not None else ""
        sched_txt = f", {abs(sched_gap)}% lower scheduled 5G capacity" if sched_gap is not None and sched_gap < 0 else ""
        deliv_txt = f", and {abs(delivered_gap)}% lower delivered 5G PDSCH throughput" if delivered_gap is not None and delivered_gap < 0 else ""
        rf_note = (" Since IAM RF indicators (RSRP, SINR) are not materially worse than the comparator,"
                   " the issue is more likely cell load, scheduler policy, NR bandwidth/BWP configuration,"
                   " carrier aggregation activation, or QoS/SIM priority rather than coverage or signal quality."
                   if rf_acceptable else "")
        eff_note = (f" Observed PDSCH Mbps per allocated PRB is higher for IAM ({round(iam_prb_eff, 2)} vs "
                    f"{round(cmp_prb_eff, 2)} Mbps/PRB for {comparator_name}), suggesting spectral efficiency is not "
                    f"the main limitation — though PRB efficiency depends on band, BWP, MCS, rank, and sample alignment."
                    if iam_prb_eff is not None and cmp_prb_eff is not None and iam_prb_eff > cmp_prb_eff else "")
        if lower_capacity_band_gap:
            conclusion = (
                f"IAM ranks {rank_str} in DL throughput with an average of {_nemo_safe_round(iam_avg, 1)} Mbps, "
                f"a {gap_display} gap versus {best_dl_op}. "
                f"The main technical difference is that IAM's detected 5G layer is {iam_nr_band or 'its active NR band'}"
                + (f" (PCI {(iam_nr_info or {}).get('pci')} / ARFCN {(iam_nr_info or {}).get('arfcn')})" if iam_nr_info else "")
                + f", while {comparator_name} uses {cmp_nr_band or 'its active NR band'}"
                + (f" (PCI {(cmp_nr_info or {}).get('pci')} / ARFCN {(cmp_nr_info or {}).get('arfcn')})" if cmp_nr_info else "")
                + f". This gives {comparator_name} a much larger NR capacity pool: IAM receives {prb_txt}{sched_txt}{deliv_txt} compared with {comparator_name}."
                + (f" IAM RF quality does not appear to be the primary limitation: RSRP, SINR, CQI and BLER are acceptable relative to {comparator_name}." if rf_acceptable else "")
                + " The most likely explanation is lower NR capacity due to band/BWP configuration, combined with lower NR resource allocation."
                + " Scheduler policy, cell load, QoS/SIM priority, and CA/BWP configuration remain hypotheses that require network counters to confirm."
                + eff_note + five_g_caveat + sample_caveat
            )
        else:
            conclusion = (
                f"IAM ranks {rank_str} in DL throughput with an average of {_nemo_safe_round(iam_avg, 1)} Mbps, "
                f"a {gap_display} gap versus {best_dl_op}. "
                f"IAM receives {prb_txt}{sched_txt}{deliv_txt} compared with {comparator_name}."
                + (f" IAM RF quality does not appear to be the primary limitation: RSRP, SINR, CQI and BLER are acceptable relative to {comparator_name}." if rf_acceptable else "")
                + " The most likely explanation is lower NR resource allocation."
                + " Scheduler policy, cell load, QoS/SIM priority, and CA/BWP configuration remain hypotheses that require network counters to confirm."
                + eff_note + five_g_caveat + sample_caveat
            )
    else:
        prb_note = f", with {abs(prb_gap)}% fewer PDSCH PRBs allocated" if prb_gap is not None and prb_gap < 0 else ""
        sched_note = f" and {abs(sched_gap)}% lower scheduled 5G throughput" if sched_gap is not None and sched_gap < 0 else ""
        conclusion = (
            f"IAM ranks {rank_str} in DL throughput with an average of {_nemo_safe_round(iam_avg, 1)} Mbps, "
            f"representing a {gap_display} gap versus {best_dl_op} (best DL operator). "
            f"The primary root cause is '{main_cause}' (severity: {severity}){prb_note}{sched_note}.{five_g_caveat}"
        )

    # French conclusion
    if main_cause == "Scheduler / resource allocation" and prb_gap is not None and prb_gap < -20:
        prb_txt_fr = f"{abs(prb_gap)}% moins de PRBs PDSCH" if prb_gap is not None else ""
        sched_txt_fr = f", {abs(sched_gap)}% moins de capacité 5G planifiée" if sched_gap is not None and sched_gap < 0 else ""
        deliv_txt_fr = f", et {abs(delivered_gap)}% moins de débit PDSCH 5G livré" if delivered_gap is not None and delivered_gap < 0 else ""
        rf_note_fr = (" Les indicateurs RF IAM (RSRP, SINR) n'étant pas significativement moins bons que le comparateur,"
                      " la cause est probablement liée à la charge cellule, la politique scheduler, la config NR/BWP,"
                      " l'activation CA ou la priorité QoS/SIM plutôt qu'à la couverture ou la qualité signal."
                      if rf_acceptable else "")
        eff_note_fr = (f" Le PDSCH Mbps par PRB alloué est plus élevé pour IAM ({round(iam_prb_eff, 2)} vs "
                       f"{round(cmp_prb_eff, 2)} Mbps/PRB pour {comparator_name}), ce qui suggère que l'efficacité spectrale "
                       f"n'est pas la principale limitation — mais l'efficacité PRB dépend de la bande, BWP, MCS, rang et alignement des échantillons."
                       if iam_prb_eff is not None and cmp_prb_eff is not None and iam_prb_eff > cmp_prb_eff else "")
        if lower_capacity_band_gap:
            conclusion_fr = (
                f"IAM se classe {rank_str} en débit DL avec une moyenne de {_nemo_safe_round(iam_avg, 1)} Mbps, "
                f"soit un écart de {gap_display} par rapport à {best_dl_op}. "
                f"La principale différence technique est que la couche 5G détectée pour IAM est {iam_nr_band or 'sa bande NR active'}"
                + (f" (PCI {(iam_nr_info or {}).get('pci')} / ARFCN {(iam_nr_info or {}).get('arfcn')})" if iam_nr_info else "")
                + f", tandis que {comparator_name} utilise {cmp_nr_band or 'sa bande NR active'}"
                + (f" (PCI {(cmp_nr_info or {}).get('pci')} / ARFCN {(cmp_nr_info or {}).get('arfcn')})" if cmp_nr_info else "")
                + f". Cela confère à {comparator_name} un pool de capacité NR bien plus grand : IAM reçoit {prb_txt_fr}{sched_txt_fr}{deliv_txt_fr} par rapport à {comparator_name}."
                + (f" La qualité RF d'IAM ne semble pas être la principale limitation : RSRP, SINR, CQI et BLER sont acceptables par rapport à {comparator_name}." if rf_acceptable else "")
                + " L'explication la plus probable est une capacité NR plus faible due à la configuration bande/BWP, combinée à une allocation de ressources NR plus faible."
                + " La politique scheduler, la charge cellulaire, la priorité QoS/SIM et la configuration CA/BWP restent des hypothèses qui nécessitent des compteurs réseau pour être confirmées."
                + eff_note_fr + sample_caveat_fr
            )
        else:
            conclusion_fr = (
                f"IAM se classe {rank_str} en débit DL avec une moyenne de {_nemo_safe_round(iam_avg, 1)} Mbps, "
                f"soit un écart de {gap_display} par rapport à {best_dl_op}. "
                f"IAM reçoit {prb_txt_fr}{sched_txt_fr}{deliv_txt_fr} par rapport à {comparator_name}."
                + (f" La qualité RF d'IAM ne semble pas être la principale limitation : RSRP, SINR, CQI et BLER sont acceptables par rapport à {comparator_name}." if rf_acceptable else "")
                + " L'explication la plus probable est une allocation de ressources NR plus faible."
                + " La politique scheduler, la charge cellulaire, la priorité QoS/SIM et la configuration CA/BWP restent des hypothèses qui nécessitent des compteurs réseau pour être confirmées."
                + eff_note_fr + sample_caveat_fr
            )
    else:
        prb_note_fr = f", avec {abs(prb_gap)}% moins de PRBs PDSCH alloués" if prb_gap is not None and prb_gap < 0 else ""
        sched_note_fr = f" et {abs(sched_gap)}% moins de débit 5G planifié" if sched_gap is not None and sched_gap < 0 else ""
        _cause_fr_map = {
            "Scheduler / resource allocation": "Planification radio / allocation des ressources",
            "Coverage limitation": "Limitation de couverture",
            "Radio quality / interference": "Qualité radio / interférences",
            "Bandwidth / BWP limitation": "Limitation de bande passante / BWP",
            "Carrier aggregation limitation": "Limitation d'agrégation de porteuses",
            "MIMO limitation": "Limitation MIMO",
            "Radio inefficiency / BLER": "Inefficacité radio / BLER élevé",
            "Transport / core limitation": "Limitation transport / cœur réseau",
        }
        _severity_fr_map = {"Critical": "Critique", "High": "Élevée", "Medium": "Moyenne", "Low": "Faible"}
        main_cause_fr = _cause_fr_map.get(main_cause, main_cause)
        severity_fr = _severity_fr_map.get(severity, severity)
        conclusion_fr = (
            f"IAM se classe {rank_str} en débit DL avec une moyenne de {_nemo_safe_round(iam_avg, 1)} Mbps, "
            f"soit un écart de {gap_display} par rapport à {best_dl_op} (meilleur opérateur DL). "
            f"La cause principale est '{main_cause_fr}' (sévérité : {severity_fr}){prb_note_fr}{sched_note_fr}."
        )

    top_steps = _nemo_recommendations_for_cause(main_cause, comparator_name)
    top_steps_fr = _nemo_recommendations_for_cause_fr(main_cause, comparator_name)
    return {
        "title": "Final Executive Conclusion",
        "available": bool(iam_entry),
        "bestDlOperator": best_dl_op,
        "iamRank": iam_rank,
        "iamAvgMbps": _nemo_safe_round(iam_avg, 1),
        "gapVsBestDlPct": gap_pct,
        "best5gComparator": best5g,
        "comparator": comparator_name,
        "mainWeakness": main_cause,
        "displayMainWeakness": display_main_cause,
        "severity": severity,
        "prbGapPct": prb_gap,
        "scheduledGapPct": sched_gap,
        "deliveredGapPct": delivered_gap,
        "prbEfficiencyIam": _nemo_safe_round(iam_prb_eff, 2),
        "prbEfficiencyCmp": _nemo_safe_round(cmp_prb_eff, 2),
        "rfAcceptable": rf_acceptable,
        "has5g": has_5g,
        "pdschSchedSampleCount": pdsch_sched_samples,
        "lowPdschSamples": low_pdsch_samples,
        "conclusion": conclusion,
        "conclusion_fr": conclusion_fr,
        "topRecommendation": top_steps[0] if top_steps else "",
        "allRecommendations": top_steps,
        "allRecommendations_fr": top_steps_fr,
    }


def _nemo_build_validation_warnings(operators: list[dict], diagnosis: dict) -> dict:
    """Task 18 — validation warnings shown at the top of the report."""
    warnings = []
    device_models = [
        {
            "operator": item.get("operator") or "UNKNOWN",
            "deviceModel": str(item.get("deviceModel") or "").strip(),
        }
        for item in (operators or [])
        if str(item.get("deviceModel") or "").strip()
    ]
    unique_device_models = sorted({item["deviceModel"] for item in device_models})

    if len(unique_device_models) > 1:
        warnings.append({
            "type": "device_model_mismatch",
            "severity": "warning",
            "operator": "",
            "message": "Different device models were used across operators: "
            + "; ".join(
                f"{item['operator']}: {item['deviceModel']}"
                for item in device_models
            )
            + ". Confirm device parity before drawing strong conclusions from a single DT.",
        })

    # No 5G detected for any operator
    any_5g = any(item.get("has5g") for item in operators or [])
    if not any_5g:
        warnings.append({
            "type": "no_5g",
            "severity": "warning",
            "operator": "",
            "message": "No 5G/EN-DC was detected for any operator in this export. 5G-specific KPIs (PRBs, scheduled 5G, MIMO rank) are unavailable. Results reflect LTE performance only.",
        })
    else:
        no5g_ops = [item.get("operator") for item in operators if not item.get("has5g")]
        for op in no5g_ops:
            warnings.append({
                "type": "no_5g",
                "severity": "info",
                "operator": op or "",
                "message": f"{op} has no 5G detected in export — excluded from 5G KPI comparison.",
            })

    # Fewer than 2 tests per operator
    for item in operators or []:
        tests = item.get("tests") or []
        op = item.get("operator") or "UNKNOWN"
        if len(tests) <= 1:
            warnings.append({
                "type": "insufficient_tests",
                "severity": "warning",
                "operator": op,
                "message": f"{op} has only {len(tests)} test session(s). Results may not be statistically representative.",
            })

    # App DL missing (using MAC-level as fallback)
    for item in operators or []:
        op = item.get("operator") or "UNKNOWN"
        dl_key = item.get("dlMetricKey") or ""
        if dl_key not in ("appDlMbps", "appDlAvgMbps"):
            warnings.append({
                "type": "missing_app_dl",
                "severity": "warning",
                "operator": op,
                "message": f"{op}: Application-layer DL rate not found. Using MAC-layer throughput ({dl_key or 'N/A'}) as fallback. Export 'App. rate DL' from Nemo for accurate benchmark.",
            })

    # GPS missing
    for item in operators or []:
        op = item.get("operator") or "UNKNOWN"
        raw_rows = item.get("rows") or []
        gps_count = sum(1 for r in raw_rows if r.get("lat") is not None and r.get("lon") is not None)
        if gps_count == 0:
            warnings.append({
                "type": "missing_gps",
                "severity": "warning",
                "operator": op,
                "message": f"{op}: No GPS coordinates found. Per-CGPS analysis is unavailable.",
            })

    # TCP/ping missing
    for item in operators or []:
        op = item.get("operator") or "UNKNOWN"
        kpis = item.get("kpis") or {}
        tcp_count = int((kpis.get("tcpHandshake") or {}).get("sampleCount") or 0)
        ping_rows = sum(1 for r in (item.get("rows") or []) if r.get("pingStatus") not in (None, "", "N/A"))
        if tcp_count == 0 and ping_rows == 0:
            warnings.append({
                "type": "missing_transport",
                "severity": "info",
                "operator": op,
                "message": f"{op}: TCP handshake time and ping status are missing. Transport / core diagnosis is limited.",
            })

    # Low test count
    for item in operators or []:
        op = item.get("operator") or "UNKNOWN"
        tests = item.get("tests") or []
        if len(tests) == 1:
            warnings.append({
                "type": "low_test_count",
                "severity": "warning",
                "operator": op,
                "message": f"{op}: Only 1 test session detected. Results are not statistically representative. Repeat with ≥ 3 sessions for a reliable benchmark.",
            })

    return {
        "title": "Validation Warnings",
        "title_fr": "Alertes de validation",
        "available": True,
        "warnings": warnings,
        "warningCount": len([w for w in warnings if w["severity"] == "warning"]),
        "infoCount": len([w for w in warnings if w["severity"] == "info"]),
    }


def _nemo_build_rules_applied(operators: list[dict], diagnosis: dict) -> dict:
    """Task 19 — auditability: 8 rules used in the analysis."""
    comparator_name = (diagnosis or {}).get("comparator") or "the best comparator"
    best5g = (diagnosis or {}).get("best5gComparator") or ""
    iam = next((item for item in operators if str(item.get("operator") or "").upper() == "IAM"), None)
    iam_dl_key = (iam.get("dlMetricKey") or "") if iam else ""
    all_avg = [(op.get("operator"), (op.get("kpis") or {}).get("dl", {}).get("average")) for op in operators or [] if (op.get("kpis") or {}).get("dl", {}).get("average") is not None]
    best_dl_op_name = max(all_avg, key=lambda x: x[1], default=(comparator_name, 0))[0] if all_avg else comparator_name
    rules = [
        {
            "id": 1,
            "rule": "DL metric selection",
            "detail": (
                f"Preferred order: appDlMbps → appDlAvgMbps → totalMacDlMbps → macDl5gMbps → macDlLteMbps. "
                f"IAM resolved to: '{iam_dl_key or 'N/A'}'."
            ),
        },
        {
            "id": 2,
            "rule": "5G detection",
            "detail": (
                "An operator is flagged as 5G if any row has EN-DC/5G in serving or packet technology, "
                "non-zero MAC DL 5G or PDSCH 5G throughput, or a valid NR channel number."
            ),
        },
        {
            "id": 3,
            "rule": "Best global DL operator",
            "detail": (
                "Highest average DL throughput operator regardless of 5G status. "
                f"Resolved to: '{best_dl_op_name or comparator_name}'."
            ),
        },
        {
            "id": 4,
            "rule": "Best 5G comparator",
            "detail": (
                "Highest average DL throughput operator with confirmed 5G that is not IAM. "
                f"Resolved to: '{best5g or 'None — no other operator had confirmed 5G'}'."
            ),
        },
        {
            "id": 5,
            "rule": "Gap formula",
            "detail": "gap% = ((IAM − Comparator) / Comparator) × 100. Negative = IAM below comparator.",
        },
        {
            "id": 6,
            "rule": "Severity thresholds",
            "detail": "Critical ≥ 50% | High ≥ 25% | Medium ≥ 10% | Low > 0% | — = no gap.",
        },
        {
            "id": 7,
            "rule": "Root cause scoring",
            "detail": (
                "Each KPI gap (RSRP, SINR, CQI, PRBs, CA, MIMO RI, BLER, TCP, scheduled 5G) contributes to a "
                "named cause bucket. The highest-scoring bucket becomes mainCause."
            ),
        },
        {
            "id": 8,
            "rule": "N/A handling for no-5G operators",
            "detail": (
                "Operators without detected 5G are excluded from 5G-specific radio comparison tables "
                "(PRBs, scheduled rank, MIMO RI, BLER sections) but remain in the DL throughput ranking."
            ),
        },
    ]
    _rules_fr = [
        {
            "id": 1,
            "rule": "Sélection de la métrique DL",
            "detail": (
                f"Ordre de préférence : appDlMbps → appDlAvgMbps → totalMacDlMbps → macDl5gMbps → macDlLteMbps. "
                f"IAM résolu en : '{iam_dl_key or 'N/A'}'."
            ),
        },
        {
            "id": 2,
            "rule": "Détection 5G",
            "detail": (
                "Un opérateur est flaggué 5G si au moins une ligne présente EN-DC/5G dans la technologie servante ou paquet, "
                "un débit MAC DL 5G ou PDSCH 5G non nul, ou un numéro de canal NR valide."
            ),
        },
        {
            "id": 3,
            "rule": "Meilleur opérateur DL global",
            "detail": (
                "Opérateur avec le débit DL moyen le plus élevé, quel que soit le statut 5G. "
                f"Résolu en : '{best_dl_op_name or comparator_name}'."
            ),
        },
        {
            "id": 4,
            "rule": "Meilleur comparateur 5G",
            "detail": (
                "Opérateur avec la 5G confirmée et le meilleur débit DL moyen (hors IAM). "
                f"Résolu en : '{best5g or 'Aucun — aucun autre opérateur avec 5G confirmée'}'."
            ),
        },
        {
            "id": 5,
            "rule": "Formule d'écart",
            "detail": "Écart% = ((IAM − Comparateur) / Comparateur) × 100. Négatif = IAM en dessous du comparateur.",
        },
        {
            "id": 6,
            "rule": "Seuils de sévérité",
            "detail": "Critique ≥ 50% | Élevée ≥ 25% | Moyenne ≥ 10% | Faible > 0% | — = pas d'écart.",
        },
        {
            "id": 7,
            "rule": "Score causes racines",
            "detail": (
                "Chaque écart de KPI (RSRP, SINR, CQI, PRBs, CA, MIMO RI, BLER, TCP, débit 5G planifié) contribue à un "
                "compartiment de cause nommé. Le compartiment avec le score le plus élevé devient la mainCause."
            ),
        },
        {
            "id": 8,
            "rule": "Gestion N/A pour opérateurs sans 5G",
            "detail": (
                "Les opérateurs sans 5G détectée sont exclus des tableaux de comparaison radio 5G spécifiques "
                "(PRBs, rang planifié, MIMO RI, BLER) mais restent dans le classement DL."
            ),
        },
    ]
    return {
        "title": "Rules Applied",
        "title_fr": "Règles appliquées",
        "available": True,
        "rules": rules,
        "rules_fr": _rules_fr,
    }


def _nemo_build_qa_checklist(operators: list[dict], ranking: list[dict], diagnosis: dict) -> dict:
    """Task 20 — 7 automated QA checks."""
    comparator_name = (diagnosis or {}).get("comparator") or ""
    best5g = (diagnosis or {}).get("best5gComparator") or ""
    iam_entry = next((e for e in ranking if str(e.get("operator") or "").upper() == "IAM"), None)
    op_names_in_ranking = {str(e.get("operator") or "").upper() for e in ranking}
    op_names_in_files = {str(item.get("operator") or "").upper() for item in operators}
    checks = []

    # 1 — All operators present in ranking
    all_in_ranking = op_names_in_files.issubset(op_names_in_ranking)
    checks.append({
        "id": 1,
        "check": "All operators present in DL ranking",
        "passed": all_in_ranking,
        "detail": "OK" if all_in_ranking else f"Missing from ranking: {op_names_in_files - op_names_in_ranking}",
    })

    # 2 — No-5G operator not used as 5G comparator
    no5g_ops = {str(item.get("operator") or "").upper() for item in operators if not item.get("has5g")}
    bad_comparator = best5g.upper() in no5g_ops if best5g else False
    checks.append({
        "id": 2,
        "check": "No-5G operator not used as 5G comparator",
        "passed": not bad_comparator,
        "detail": f"OK — 5G comparator is '{best5g or 'None'}'" if not bad_comparator else f"ERROR: '{best5g}' has no 5G but is set as best5gComparator",
    })

    # 3 — IAM uses correct comparator (not itself)
    iam_as_comparator = comparator_name.upper() == "IAM"
    checks.append({
        "id": 3,
        "check": "IAM comparator is not IAM itself",
        "passed": not iam_as_comparator,
        "detail": f"OK — comparator is '{comparator_name}'" if not iam_as_comparator else "ERROR: IAM is compared against itself",
    })

    # 4 — Gap values are non-null when IAM and comparator both have data
    iam_avg = (iam_entry or {}).get("avgDlMbps")
    cmp_entry = next((e for e in ranking if e.get("operator") == comparator_name), None)
    cmp_avg = (cmp_entry or {}).get("avgDlMbps")
    gap_ok = (iam_avg is not None and cmp_avg is not None and (diagnosis or {}).get("gapToBestDlPct") is not None) or (iam_avg is None or cmp_avg is None)
    checks.append({
        "id": 4,
        "check": "Gap values are non-null when data is available",
        "passed": gap_ok,
        "detail": f"IAM avg: {_nemo_safe_round(iam_avg, 1)} Mbps, comparator avg: {_nemo_safe_round(cmp_avg, 1)} Mbps, gap: {(diagnosis or {}).get('gapToBestDlPct')}%",
    })

    # 5 — Severity label matches threshold
    gap_pct = (diagnosis or {}).get("gapToBestDlPct")
    severity = (diagnosis or {}).get("severity") or ""
    expected_severity = _nemo_gap_severity(gap_pct) if gap_pct is not None else "—"
    severity_ok = severity == expected_severity or gap_pct is None
    checks.append({
        "id": 5,
        "check": "Severity label matches threshold",
        "passed": severity_ok,
        "detail": f"gap={gap_pct}% → expected '{expected_severity}', got '{severity}'" if not severity_ok else f"OK — severity '{severity}' is correct for {gap_pct}% gap",
    })

    # 6 — N/A not shown as zero for no-5G operators in 5G sections
    no5g_shown_as_zero = False
    for item in operators:
        if item.get("has5g"):
            continue
        kpis = item.get("kpis") or {}
        sched = (kpis.get("scheduled5g") or {}).get("average")
        if sched is not None and float(sched) == 0:
            no5g_shown_as_zero = True
            break
    checks.append({
        "id": 6,
        "check": "N/A not shown as zero for no-5G operators",
        "passed": not no5g_shown_as_zero,
        "detail": "OK — no-5G operators have null 5G stats" if not no5g_shown_as_zero else "WARNING: a no-5G operator has scheduled5g average = 0 (may display as zero instead of N/A)",
    })

    # 7 — Export completes (operators list is non-empty)
    export_ok = bool(operators)
    checks.append({
        "id": 7,
        "check": "Export completes (operator list is non-empty)",
        "passed": export_ok,
        "detail": f"OK — {len(operators)} operator(s) loaded" if export_ok else "ERROR: no operators were parsed",
    })

    all_passed = all(c["passed"] for c in checks)
    return {
        "title": "QA Checklist",
        "available": True,
        "checks": checks,
        "allPassed": all_passed,
        "passCount": sum(1 for c in checks if c["passed"]),
        "failCount": sum(1 for c in checks if not c["passed"]),
    }


def _nemo_build_time_grid(rows: list[dict]) -> dict:
    """
    Collapses per-row Nemo data into 1-second canonical buckets and returns
    pre-aggregated statistics for three views:
      wholeTest  — every bucket that has any data
      appActive  — buckets where App DL > 0
      nrActive   — buckets where any row has NR / EN-DC active

    Using max() for throughput fields prevents double-counting in EN-DC seconds
    where two rows exist (LTE Anchor + NR SCG PSCell).
    """
    from collections import defaultdict as _tg_dd
    from datetime import datetime as _tg_dt

    _NR_CT = {"nr serving", "nr scg pscell", "5g serving"}
    app_key = "appDlMbps" if any(r.get("appDlMbps") for r in rows) else "appDlAvgMbps"

    # ── Step A: group rows into 1-second buckets ──────────────────────────
    raw_buckets: dict = _tg_dd(list)
    for row in rows:
        dt = row.get("_dt")
        if isinstance(dt, _tg_dt):
            raw_buckets[dt.replace(microsecond=0)].append(row)

    # ── Step B: aggregate each bucket into one canonical record ───────────
    def _safe_max(vals):
        v = [x for x in vals if x is not None]
        return max(v) if v else None

    def _safe_mean(vals):
        v = [x for x in vals if x is not None]
        return sum(v) / len(v) if v else None

    def _safe_median(vals):
        v = sorted(x for x in vals if x is not None)
        return v[len(v) // 2] if v else None

    ri_key = "mimoRi" if any(r.get("mimoRi") for r in rows) else "ri"

    processed = {}
    for sec_key, brows in raw_buckets.items():
        app_dl = _safe_max(r.get(app_key) for r in brows)
        if app_dl is not None:
            app_dl = app_dl if app_dl > 0 else None

        # MAC total: prefer totalMacDlMbps field; fallback to lte+5g sum
        mac_total_candidates = [r.get("totalMacDlMbps") for r in brows if r.get("totalMacDlMbps") is not None]
        if mac_total_candidates:
            mac_total = max(mac_total_candidates)
        else:
            lte_parts = [r.get("macDlLteMbps") or 0 for r in brows if r.get("macDlLteMbps") is not None]
            g5_parts  = [r.get("macDl5gMbps")  or 0 for r in brows if r.get("macDl5gMbps")  is not None]
            if lte_parts or g5_parts:
                mac_total = sum(lte_parts) + sum(g5_parts)
            else:
                mac_total = None
        if mac_total is not None and mac_total <= 0:
            mac_total = None

        mac_dl_5g  = _safe_max(r.get("macDl5gMbps")     for r in brows)
        pdsch_act  = _safe_max(r.get("pdschDl5gMbps")   for r in brows)
        pdsch_sched = _safe_max(r.get("pdschSched5gMbps") for r in brows)
        if mac_dl_5g  is not None and mac_dl_5g  <= 0: mac_dl_5g  = None
        if pdsch_act  is not None and pdsch_act  <= 0: pdsch_act  = None
        if pdsch_sched is not None and pdsch_sched <= 0: pdsch_sched = None

        pdsch_eff = (pdsch_act / pdsch_sched * 100.0) if (pdsch_act and pdsch_sched) else None

        prb  = _safe_mean(r.get("pdschPrbs")   for r in brows)
        cqi  = _safe_mean(r.get("wbCqi")        for r in brows)
        bler = _safe_mean(r.get("macDlBler")    for r in brows)
        ri   = _safe_mean(r.get(ri_key)         for r in brows)

        has_nr = any(
            r.get("nrChannelNumber") is not None
            or any(str(ct or "").strip().lower() in _NR_CT for ct in (r.get("cellTypes") or []))
            or "EN-DC" in str(r.get("servingTechnology") or "").upper()
            for r in brows
        )
        is_app_active = bool(app_dl is not None and app_dl > 0)

        processed[sec_key] = {
            "app_dl": app_dl, "mac_total": mac_total, "mac_dl_5g": mac_dl_5g,
            "pdsch_act": pdsch_act, "pdsch_sched": pdsch_sched, "pdsch_eff": pdsch_eff,
            "prb": prb, "cqi": cqi, "bler": bler, "ri": ri,
            "has_nr": has_nr, "is_app_active": is_app_active,
        }

    # ── Step C: three filtered views ──────────────────────────────────────
    whole      = list(processed.values())
    app_active = [b for b in whole if b["is_app_active"]]
    nr_active  = [b for b in whole if b["has_nr"]]

    # ── Step D: aggregate a view list to stats ────────────────────────────
    def _agg_view(buckets: list) -> dict:
        n = len(buckets)
        if n == 0:
            return {
                "nSeconds": 0, "appDlAvg": None, "macTotalAvg": None,
                "macDl5gAvg": None, "pdschActAvg": None, "pdschSchedAvg": None,
                "pdschEfficiency": None, "prbAvg": None, "cqiMedian": None,
                "blerAvg": None, "riMedian": None,
            }

        def _r1(x):
            return round(x, 1) if x is not None else None

        return {
            "nSeconds":        n,
            "appDlAvg":        _r1(_safe_mean(b["app_dl"]     for b in buckets)),
            "macTotalAvg":     _r1(_safe_mean(b["mac_total"]  for b in buckets)),
            "macDl5gAvg":      _r1(_safe_mean(b["mac_dl_5g"]  for b in buckets)),
            "pdschActAvg":     _r1(_safe_mean(b["pdsch_act"]  for b in buckets)),
            "pdschSchedAvg":   _r1(_safe_mean(b["pdsch_sched"] for b in buckets)),
            "pdschEfficiency": _r1(_safe_mean(b["pdsch_eff"]  for b in buckets)),
            "prbAvg":          _r1(_safe_mean(b["prb"]        for b in buckets)),
            "cqiMedian":       _r1(_safe_median(b["cqi"]      for b in buckets)),
            "blerAvg":         _r1(_safe_mean(b["bler"]       for b in buckets)),
            "riMedian":        _r1(_safe_median(b["ri"]       for b in buckets)),
        }

    return {
        "buckets":   processed,            # internal — not serialised to JSON
        "wholeTest": _agg_view(whole),
        "appActive": _agg_view(app_active),
        "nrActive":  _agg_view(nr_active),
    }


def _nemo_build_layer_throughput_analysis(operator_files: list[dict], iam_serving_cells: dict) -> dict:
    """
    Builds the Multi-Layer DL Throughput Analysis using a common 1-second time grid
    so that all KPI layers (App / MAC / PDSCH) are computed over the same seconds.
    This eliminates the App/MAC > 1 anomaly caused by mixing different row sets.
    """
    from datetime import datetime as _dt_class
    analysis = []
    iam_breakdown = None

    for op_file in operator_files:
        op_name = op_file.get("operator", "UNKNOWN")
        rows = op_file.get("rows") or []

        top_nr_info = _nemo_dominant_nr_serving_info(rows)
        window_rows = rows
        top_cell_display = "Unknown"
        top_band = None
        top_pci = None
        top_arfcn = None

        if top_nr_info:
            top_cell_display = top_nr_info.get("display") or "Unknown"
            top_band = top_nr_info.get("band")
            top_pci = top_nr_info.get("pci")
            top_arfcn = top_nr_info.get("arfcn")
            start_dt = top_nr_info.get("startDt")
            end_dt = top_nr_info.get("endDt")
            if isinstance(start_dt, _dt_class) and isinstance(end_dt, _dt_class):
                window_rows = [r for r in rows if r.get("_dt") and start_dt <= r["_dt"] <= end_dt]

        # 1) Build common time grid — all KPIs aligned to the same 1-second windows
        tg = _nemo_build_time_grid(rows)
        whole_view = tg["wholeTest"]
        app_view   = tg["appActive"]
        nr_view    = tg["nrActive"]

        if whole_view["nSeconds"] == 0:
            # No rows have a valid timestamp — analysis not possible for this operator
            continue

        # Source all throughput KPIs from the app-active view (backward-compatible semantics)
        app_pipe_avg        = app_view["appDlAvg"]
        mac_pipe_avg        = app_view["macTotalAvg"]
        pdsch_act_pipe_avg  = app_view["pdschActAvg"]
        pdsch_sched_pipe_avg = app_view["pdschSchedAvg"]
        pdsch_efficiency    = app_view["pdschEfficiency"]

        # Sample counts now in seconds (same window for all layers — no more mismatch)
        app_sample_count         = app_view["nSeconds"]
        mac_sample_count         = app_view["nSeconds"]
        pdsch_act_sample_count   = app_view["nSeconds"]
        pdsch_sched_sample_count = app_view["nSeconds"]

        # 5G presence from operator-level technologyStatus (time-based, already computed)
        op_ts = op_file.get("technologyStatus") or {}
        nr_presence_pct = op_ts.get("nrPresencePct")

        # 2) Compute radio explanation KPIs inside the same 5G window
        rsrps = sorted([float(r["rsrp"]) for r in window_rows if r.get("rsrp") is not None])
        sinrs = sorted([float(r["sinr"]) for r in window_rows if r.get("sinr") is not None])
        cqis = sorted([float(r["wbCqi"] if r.get("wbCqi") is not None else r.get("cqi")) for r in window_rows if r.get("wbCqi") is not None or r.get("cqi") is not None])
        mcss = sorted([float(r["mcs"]) for r in window_rows if r.get("mcs") is not None])

        ri_key = "mimoRi" if any(r.get("mimoRi") for r in window_rows) else "ri"
        ris = sorted([float(r[ri_key]) for r in window_rows if r.get(ri_key) is not None])
        sched_ris = sorted([float(r["scheduledRank"]) for r in window_rows if r.get("scheduledRank") is not None])

        blers = [float(r["macDlBler"]) for r in window_rows if r.get("macDlBler") is not None]
        prbs = [float(r["pdschPrbs"]) for r in window_rows if r.get("pdschPrbs") is not None]

        rsrp_median = rsrps[len(rsrps)//2] if rsrps else None
        sinr_median = sinrs[len(sinrs)//2] if sinrs else None
        mcs_median = mcss[len(mcss)//2] if mcss else None
        sched_ri_median = sched_ris[len(sched_ris)//2] if sched_ris else None
        bler_max = max(blers) if blers else None

        # CQI, RI, BLER, PRB sourced from the NR-active time-grid view
        # (broader and more representative than the old ARFCN-trimmed window_rows)
        cqi_median = nr_view["cqiMedian"]
        ri_median  = nr_view["riMedian"]
        bler_avg   = nr_view["blerAvg"]
        prb_avg    = nr_view["prbAvg"]

        # pdsch_sample_count: number of 5G-active seconds (used for confidence note)
        pdsch_sample_count = nr_view["nSeconds"]

        # 3) Advanced diagnosis
        case = "Case D"
        diagnosis_title = "Healthy Radio-to-App Conversion"
        diagnosis_text = "Radio capacity is effectively converted into user throughput."

        if top_cell_display == "Unknown":
            case = "N/A"
            diagnosis_title = "LTE-only / no 5G analysis"
            diagnosis_text = "This file does not contain 5G serving-cell or PDSCH data. Do not perform 5G scheduler analysis."
            mac_pipe_avg = None
            pdsch_act_pipe_avg = None
            pdsch_sched_pipe_avg = None
        elif not pdsch_act_pipe_avg and not pdsch_sched_pipe_avg:
            case = "N/A"
            diagnosis_title = "No 5G Layer Data"
            diagnosis_text = "Cannot perform 5G scheduler analysis (no PDSCH throughput samples detected)."
            if app_pipe_avg and not pdsch_act_pipe_avg:
                diagnosis_text += " App result is high but not tied to 5G PDSCH."
        elif pdsch_efficiency is not None and pdsch_efficiency < 70:
            if sinr_median is not None and sinr_median < 8:
                case = "Case A"
                diagnosis_title = "Radio quality limitation, likely SINR issue"
                diagnosis_text = "Radio efficiency is low, supported by poor median SINR."
            elif bler_avg is not None and bler_avg > 10:
                case = "Case A"
                diagnosis_title = "Radio reliability issue, likely BLER/retransmissions"
                diagnosis_text = "Radio efficiency is low, supported by high average BLER."
            elif mcs_median is not None and mcs_median < 10:
                case = "Case A"
                diagnosis_title = "Weak modulation/coding, radio adaptation issue"
                diagnosis_text = "Radio efficiency is low, supported by poor radio adaptation (weak MCS)."
            elif ri_median is not None and ri_median <= 1:
                case = "Case A"
                diagnosis_title = "Low MIMO rank limitation"
                diagnosis_text = "Radio efficiency is low, limited by poor MIMO rank."
            elif app_pipe_avg is not None and pdsch_act_pipe_avg > app_pipe_avg * 1.3:
                case = "Case C"
                diagnosis_title = "App/Transport Bottleneck, not clear 5G radio issue"
                diagnosis_text = "Although scheduled PDSCH is higher than actual PDSCH, the gap is not supported by poor SINR, high BLER, weak MCS, or low rank. The bigger issue is that App DL throughput is much lower than the radio-layer throughput. This suggests the bottleneck is more likely above the radio layer: TCP/application behavior, server limitation, test duration, packet loss outside radio, or sampling mismatch between layers."
            else:
                case = "Case B"
                diagnosis_title = "Normal radio-layer gap, no clear bottleneck"
                diagnosis_text = "Some efficiency gap exists, but radio quality is acceptable and App throughput converts decently."
        elif pdsch_act_pipe_avg and app_pipe_avg and pdsch_act_pipe_avg > app_pipe_avg * 1.3:
            case = "Case C"
            diagnosis_title = "App/Transport Bottleneck"
            diagnosis_text = "Radio throughput is much higher than App throughput. The radio delivers well, but TCP, server, packet loss outside radio, or test duration limits the final throughput."

        band_prefix = f"{top_band} " if top_band else ""
        mcs_note = ""
        if mcs_median is not None:
            if mcs_median < 10:
                mcs_note = f" Observed PDSCH MCS is low (median ~{round(mcs_median, 1)})."
            else:
                mcs_note = f" Observed PDSCH MCS is moderate-to-good (median ~{round(mcs_median, 1)})."
        sample_note = f" PDSCH sample count is small ({pdsch_sample_count})." if 0 < pdsch_sample_count < 10 else ""

        if op_name.upper() == "IAM":
            if top_cell_display == "Unknown":
                case = "N/A"
                diagnosis_title = "LTE-only / no 5G analysis"
                diagnosis_text = "This file does not contain 5G serving-cell or PDSCH data. IAM should be treated as LTE-only. Do not perform 5G scheduler analysis."
            elif top_band and str(top_band).lower() != "n78":
                case = "Case B"
                diagnosis_title = "Lower-capacity NR layer / resource limitation more likely"
                diagnosis_text = (
                    f"IAM 5G is detected on {band_prefix}PCI {top_pci} / ARFCN {top_arfcn}, not n78. "
                    "RSRP, SINR, and CQI are acceptable in this window, so simple RF coverage is probably not the primary issue. "
                    "The first hypothesis is lower NR capacity or lower NR resource allocation on this active layer."
                    + mcs_note + sample_note
                )
            else:
                diagnosis_title = "5G present, but capacity/resource gap remains the main issue"
                diagnosis_text = (
                    f"IAM 5G is detected on {band_prefix}PCI {top_pci} / ARFCN {top_arfcn}. "
                    "RF quality is acceptable enough for 5G analysis, but the main weakness remains limited NR resource allocation and lower effective capacity in the active window."
                    + mcs_note + sample_note
                )
        elif op_name.upper() == "ORANGE":
            if top_cell_display != "Unknown":
                diagnosis_title = "5G present with good burst capacity"
                diagnosis_text = (
                    f"Orange 5G is detected on {band_prefix}PCI {top_pci} / ARFCN {top_arfcn}. "
                    "The 5G layer contributes additional burst capacity in this file."
                    + mcs_note + sample_note
                )
        elif op_name.upper() == "INWI":
            if top_cell_display != "Unknown":
                diagnosis_title = "Strong 5G burst, no major radio limitation"
                diagnosis_text = (
                    f"INWI 5G is detected on {band_prefix}PCI {top_pci} / ARFCN {top_arfcn}. "
                    "High App DL is supported by strong scheduled and delivered 5G throughput in this window."
                    + mcs_note + sample_note
                )

        # Scope note: explain why App DL can exceed PDSCH values in EN-DC
        scope_note = None
        if app_pipe_avg and pdsch_sched_pipe_avg and app_pipe_avg > pdsch_sched_pipe_avg:
            if nr_presence_pct is not None and nr_presence_pct < 50:
                scope_note = (
                    f"App DL avg ({round(app_pipe_avg,1)} Mbps) exceeds PDSCH scheduled avg "
                    f"({round(pdsch_sched_pipe_avg,1)} Mbps). This is expected: App DL measures "
                    f"end-to-end user throughput (LTE + NR combined), while PDSCH scheduled covers "
                    f"the 5G NR component only. With {round(nr_presence_pct,1)}% 5G presence, "
                    f"most of the App DL comes from the LTE anchor, not the NR cell."
                )
            else:
                scope_note = (
                    f"App DL avg ({round(app_pipe_avg,1)} Mbps) exceeds PDSCH scheduled avg "
                    f"({round(pdsch_sched_pipe_avg,1)} Mbps). These metrics are computed from "
                    "different sample sets and physical layer scopes — App DL is end-to-end "
                    "(LTE + NR), PDSCH is 5G-only. Direct comparison is not meaningful."
                )

        # ── Layer funnel + inter-layer efficiencies (where DL throughput is lost) ──
        _op_kpis = op_file.get("kpis") or {}
        mac_nr_avg = (_op_kpis.get("mac5g") or {}).get("average")
        mac_lte_avg = (_op_kpis.get("macLte") or {}).get("average")

        def _eff(num, den):
            return round(num / den * 100.0, 1) if (num is not None and den not in (None, 0)) else None

        delivery_eff = _eff(pdsch_act_pipe_avg, pdsch_sched_pipe_avg)   # PDSCH delivered / scheduled (HARQ/BLER loss)
        transport_eff = _eff(app_pipe_avg, mac_pipe_avg)               # App / MAC total (transport / TCP overhead)
        layer_funnel = {
            "pdschScheduled": round(pdsch_sched_pipe_avg, 1) if pdsch_sched_pipe_avg else None,
            "pdschDelivered": round(pdsch_act_pipe_avg, 1) if pdsch_act_pipe_avg else None,
            "macNr": round(mac_nr_avg, 1) if mac_nr_avg else None,
            "macLte": round(mac_lte_avg, 1) if mac_lte_avg else None,
            "macTotal": round(mac_pipe_avg, 1) if mac_pipe_avg else None,
            "app": round(app_pipe_avg, 1) if app_pipe_avg else None,
            "deliveryEfficiencyPct": delivery_eff,    # scheduled → delivered
            "transportEfficiencyPct": transport_eff,  # MAC → App
        }

        op_data = {
            "operator": op_name,
            # Throughput KPIs — all from the app-active time-grid view (same window)
            "appDlAvg":    round(app_pipe_avg,         1) if app_pipe_avg         else None,
            "macTotalAvg": round(mac_pipe_avg,         1) if mac_pipe_avg         else None,
            "macNrAvg":    round(mac_nr_avg,           1) if mac_nr_avg           else None,
            "macLteAvg":   round(mac_lte_avg,          1) if mac_lte_avg          else None,
            "pdschActAvg": round(pdsch_act_pipe_avg,   1) if pdsch_act_pipe_avg   else None,
            "pdschSchedAvg": round(pdsch_sched_pipe_avg, 1) if pdsch_sched_pipe_avg else None,
            "layerFunnel": layer_funnel,
            # Sample counts — now in seconds (all layers share the same window)
            "appDlSampleCount":      app_sample_count,
            "macSampleCount":        mac_sample_count,
            "pdschActSampleCount":   pdsch_act_sample_count,
            "pdschSchedSampleCount": pdsch_sched_sample_count,
            # PDSCH delivery efficiency (mean of per-second ratios — more robust)
            "pdschDeliveryEfficiency": round(pdsch_efficiency, 1) if pdsch_efficiency else None,
            # Full time-grid views (all three windows)
            "timeGrid": {
                "wholeTest": whole_view,
                "appActive": app_view,
                "nrActive":  nr_view,
            },
            # Time-based 5G presence
            "nrPresencePct": nr_presence_pct,
            "topCellDisplay": top_cell_display,
            "topCellBand": top_band,
            "topCellPci": top_pci,
            "topCellArfcn": top_arfcn,
            "rsrpMedian": round(rsrp_median, 1) if rsrp_median else None,
            "sinrMedian": round(sinr_median, 1) if sinr_median else None,
            "cqiMedian": round(cqi_median, 1) if cqi_median else None,
            "mcsMedian": round(mcs_median, 1) if mcs_median else None,
            "riMedian": round(ri_median, 1) if ri_median else None,
            "schedRiMedian": round(sched_ri_median, 1) if sched_ri_median else None,
            "blerAvg": round(bler_avg, 1) if bler_avg else None,
            "blerMax": round(bler_max, 1) if bler_max else None,
            "prbAvg": round(prb_avg, 1) if prb_avg else None,
            "pdschSampleCount": pdsch_sample_count,
            "scopeNote": scope_note,
            "case": case,
            "diagnosisTitle": diagnosis_title,
            "diagnosisText": diagnosis_text,
        }
        analysis.append(op_data)

        if op_name.upper() == "IAM":
            iam_breakdown = op_data

    # Generate dynamic layer conclusion
    layer_conclusion_parts = []

    for op in analysis:
        op_name = op["operator"]
        top_cell = op["topCellDisplay"]
        top_band = op.get("topCellBand")
        top_pci = op.get("topCellPci")
        top_arfcn = op.get("topCellArfcn")

        if op_name.upper() == "INWI":
            if top_cell == "Unknown":
                layer_conclusion_parts.append("For INWI:\nINWI has high App DL throughput, but no 5G PDSCH/RSRP/SINR data is present in the file. Avoid giving a 5G-radio diagnosis for INWI.")
            else:
                layer_conclusion_parts.append(
                    f"For INWI:\n5G analysis is possible. INWI uses {top_band or 'NR'} PCI {top_pci} / ARFCN {top_arfcn}. "
                    "App throughput is high and supported by strong scheduled and delivered 5G throughput. No primary RF-limitation is obvious from this file."
                )
        elif op_name.upper() == "IAM":
            if top_cell == "Unknown":
                layer_conclusion_parts.append("For IAM:\nIAM P06 is LTE-only in this file. No 5G serving cell, no NR RSRP/SINR, and no PDSCH throughput samples were detected. Do not perform 5G scheduler analysis for IAM.")
            else:
                layer_conclusion_parts.append(
                    f"For IAM:\n5G analysis is possible. IAM uses {top_band or 'NR'} PCI {top_pci} / ARFCN {top_arfcn}."
                    + (" This is not n78, so the active 5G layer is lower-capacity by design." if top_band and str(top_band).lower() != "n78" else "")
                    + " RSRP/SINR are acceptable enough that simple coverage is probably not the main issue. "
                    + ("Observed PDSCH MCS is low, which reduces confidence in a pure scheduler-only diagnosis. " if op.get("mcsMedian") is not None and op.get("mcsMedian") < 10 else "")
                    + "The likely explanation is lower NR capacity and/or lower NR resource allocation in the active window."
                )
        elif op_name.upper() == "ORANGE":
            if top_cell == "Unknown":
                layer_conclusion_parts.append(f"For Orange:\n{op['diagnosisText']}")
            else:
                layer_conclusion_parts.append(
                    f"For Orange:\n5G analysis is possible. Orange uses {top_band or 'NR'} PCI {top_pci} / ARFCN {top_arfcn}. "
                    "The 5G layer contributes additional burst capacity in this file."
                )
        else:
            layer_conclusion_parts.append(f"For {op_name}:\n{op['diagnosisText']}")

    conclusion_text = "\n\n".join(layer_conclusion_parts)

    return {
        "title": "DL Layer KPIs — App · MAC · PDSCH (not a strict pipeline)",
        "available": bool(analysis),
        "analysis": analysis,
        "iamBreakdown": iam_breakdown,
        "conclusionText": conclusion_text
    }


def _parse_benchmark_nemo_files(paths: list[str]) -> dict:
    operator_files = _benchmark_nemo_parse_operator_files(paths)
    return _benchmark_nemo_build_dataset(operator_files)


def _nemo_ordered_dt_titles(rows: list[dict]) -> list[str]:
    """Unique Measurement Titles (DT / BJ names) in first-appearance (chronological) order."""
    seen = set()
    ordered = []
    for row in rows or []:
        title = row.get("measurementTitle")
        if title and title not in seen:
            seen.add(title)
            ordered.append(title)
    return ordered


def _nemo_group_rows_by_measurement_title(rows: list[dict]) -> tuple[list[str], dict]:
    ordered = []
    grouped = {}
    for row in rows or []:
        title = row.get("measurementTitle")
        if not title:
            continue
        if title not in grouped:
            grouped[title] = []
            ordered.append(title)
        grouped[title].append(row)
    return ordered, grouped


def _benchmark_nemo_parse_operator_files(paths: list[str]) -> list[dict]:
    operator_files = []
    for path in paths or []:
        if not os.path.isfile(path):
            continue
        ext = os.path.splitext(path)[1].lower()
        if ext not in (".txt", ".csv", ".tsv"):
            continue
        # Session-statistics exports are consumed as a sibling of the time-series file,
        # not parsed as a standalone operator time-series.
        if _nemo_is_session_stats_file(path):
            continue
        operator_files.append(_nemo_parse_operator_file(path))
    return operator_files


def _nemo_build_dt_list(operator_files: list[dict]) -> list[dict]:
    """Pair DTs positionally across operators: DT #i = each operator's i-th drive test.

    Returns one entry per DT index with a label and the per-operator title at that index.
    """
    per_op = []  # [(operator, [titles...]) ...]
    for of in operator_files or []:
        per_op.append((
            str(of.get("operator") or "UNKNOWN"),
            list(of.get("orderedDtTitles") or _nemo_ordered_dt_titles(of.get("rows") or [])),
        ))
    max_len = max((len(titles) for _, titles in per_op), default=0)
    dt_list = []
    for i in range(max_len):
        titles_by_op = {}
        for operator, titles in per_op:
            if i < len(titles):
                titles_by_op[operator] = titles[i]
        # Label from the shared timestamp prefix of the i-th titles (strip operator suffix).
        sample = next(iter(titles_by_op.values()), "")
        prefix = re.split(r"[ .]", sample)[0] if sample else ""
        dt_list.append({
            "index": i,
            "label": f"DT {i + 1}" + (f" — {prefix}" if prefix else ""),
            "titlesByOperator": titles_by_op,
        })
    return dt_list


def _nemo_clone_operator_file_for_dt_index(operator_file: dict, index: int):
    """Clone one operator dataset for a single positional downlink transfer session.

    DT-scoped benchmark analysis must slice by the reconstructed download session window, not by
    `measurementTitle` alone. Some Nemo exports leave transfer rows untitled, and the parser's
    forward-fill can therefore merge multiple DTs under the same title (e.g. Settat DT1 absorbing
    DT2..DT9). For a single selected file, use the application DL sample stream as the KPI source.
    """
    if not isinstance(operator_file, dict) or index < 0:
        return None

    rows = list(operator_file.get("rows") or [])
    transfer_sessions = list(operator_file.get("transferSessions") or [])
    if not transfer_sessions:
        transfer_sessions = _nemo_build_transfer_sessions(rows, operator_file.get("operator") or "UNKNOWN")
    downlink_sessions = [
        session for session in transfer_sessions
        if str(session.get("direction") or "").lower() == "downlink"
    ]
    if index >= len(downlink_sessions):
        return None

    target_session = downlink_sessions[index]
    target_title = _benchmark_text(target_session.get("measurementTitle"))
    sub_rows = []
    try:
        start_dt = _dt.fromisoformat(str(target_session.get("startTime") or ""))
        end_dt = _dt.fromisoformat(str(target_session.get("endTime") or ""))
    except Exception:
        start_dt = None
        end_dt = None
    if start_dt is not None and end_dt is not None:
        sub_rows = [
            row for row in rows
            if row.get("_dt") is not None and start_dt <= row.get("_dt") <= end_dt
        ]
    if not sub_rows and target_title:
        rows_by_title = operator_file.get("rowsByMeasurementTitle") or {}
        sub_rows = list(rows_by_title.get(target_title) or [])
    if not sub_rows and target_title:
        sub_rows = [row for row in rows if row.get("measurementTitle") == target_title]
    if not sub_rows:
        return None

    clone = dict(operator_file)
    clone["rows"] = sub_rows
    clone["measurementTitles"] = [target_title] if target_title else []
    clone["orderedDtTitles"] = [target_title] if target_title else []
    clone["rowsByMeasurementTitle"] = {target_title: sub_rows} if target_title else {}
    clone["transferSessions"] = [target_session]
    clone["technologyStatus"] = _nemo_compute_technology_status(sub_rows, str(clone.get("operator") or "UNKNOWN"))
    clone["has5g"] = bool((clone.get("technologyStatus") or {}).get("has5g"))
    clone["fiveGStatus"] = (clone.get("technologyStatus") or {}).get("fiveGStatus")
    clone["_dlMetricKeyOverride"] = _nemo_select_dl_metric_key(sub_rows)
    clone["_benchmarkDlMetricKeyAvgOverride"] = _nemo_select_benchmark_dl_metric_key(sub_rows)
    clone["_benchmarkDlMetricKeyOverride"] = clone.get("_dlMetricKeyOverride") or _nemo_select_benchmark_dl_metric_key(sub_rows)
    return clone


def _nemo_clone_operator_file_for_dt_index_with_window(
    operator_file: dict,
    index: int,
    window_mode: str | None = None,
):
    # Always produce the SAME per-DT base window (the downlink-session-anchored span that
    # actually contains the transfer) for both window modes. Active-DL narrowing is applied
    # downstream by `_benchmark_nemo_scope_operator_file_to_window` (the single authority),
    # which reads the `transferSessions` this clone carries. `window_mode` is accepted for
    # call-site compatibility but intentionally does not change the base window here.
    _ = _benchmark_nemo_normalize_window_mode(window_mode)
    if not isinstance(operator_file, dict) or index < 0:
        return None
    rows = list(operator_file.get("rows") or [])
    transfer_sessions = list(operator_file.get("transferSessions") or [])
    if not transfer_sessions:
        transfer_sessions = _nemo_build_transfer_sessions(rows, operator_file.get("operator") or "UNKNOWN")
    downlink_sessions = [
        session for session in transfer_sessions
        if str(session.get("direction") or "").strip().lower().startswith("down")
    ]
    if index >= len(downlink_sessions):
        return None

    target_session = downlink_sessions[index]
    try:
        target_start = _dt.fromisoformat(str(target_session.get("startTime") or ""))
        target_end = _dt.fromisoformat(str(target_session.get("endTime") or ""))
    except Exception:
        return _nemo_clone_operator_file_for_dt_index(operator_file, index)
    if target_end <= target_start:
        return _nemo_clone_operator_file_for_dt_index(operator_file, index)

    row_times = [row.get("_dt") for row in rows if row.get("_dt") is not None]
    if not row_times:
        return _nemo_clone_operator_file_for_dt_index(operator_file, index)
    row_times.sort()
    prev_session = downlink_sessions[index - 1] if index > 0 else None
    next_session = downlink_sessions[index + 1] if index + 1 < len(downlink_sessions) else None

    window_start = row_times[0]
    if prev_session:
        try:
            prev_end = _dt.fromisoformat(str(prev_session.get("endTime") or ""))
            if prev_end < target_start:
                window_start = prev_end + (target_start - prev_end) / 2
        except Exception:
            pass

    window_end = row_times[-1]
    if next_session:
        try:
            next_start = _dt.fromisoformat(str(next_session.get("startTime") or ""))
            if target_end < next_start:
                window_end = target_end + (next_start - target_end) / 2
        except Exception:
            pass

    sub_rows = [
        row for row in rows
        if row.get("_dt") is not None and window_start <= row.get("_dt") <= window_end
    ]
    if not sub_rows:
        return _nemo_clone_operator_file_for_dt_index(operator_file, index)

    clone = dict(operator_file)
    clone["rows"] = sub_rows
    target_title = _benchmark_text(target_session.get("measurementTitle"))
    clone["measurementTitles"] = [target_title] if target_title else []
    clone["orderedDtTitles"] = [target_title] if target_title else []
    clone["rowsByMeasurementTitle"] = {target_title: sub_rows} if target_title else {}
    scoped_sessions = []
    for session in transfer_sessions:
        try:
            session_start = _dt.fromisoformat(str(session.get("startTime") or ""))
            session_end = _dt.fromisoformat(str(session.get("endTime") or ""))
        except Exception:
            continue
        if session_end < window_start or session_start > window_end:
            continue
        scoped_sessions.append(session)
    clone["transferSessions"] = scoped_sessions or [target_session]
    clone["technologyStatus"] = _nemo_compute_technology_status(sub_rows, str(clone.get("operator") or "UNKNOWN"))
    clone["has5g"] = bool((clone.get("technologyStatus") or {}).get("has5g"))
    clone["fiveGStatus"] = (clone.get("technologyStatus") or {}).get("fiveGStatus")
    clone["_dlMetricKeyOverride"] = _nemo_select_dl_metric_key(sub_rows)
    clone["_benchmarkDlMetricKeyAvgOverride"] = _nemo_select_benchmark_dl_metric_key(sub_rows)
    clone["_benchmarkDlMetricKeyOverride"] = clone.get("_dlMetricKeyOverride") or _nemo_select_benchmark_dl_metric_key(sub_rows)
    return clone


def _benchmark_nemo_scope_operator_file_to_window(
    operator_file: dict,
    window_mode: str | None = None,
):
    if not isinstance(operator_file, dict):
        return operator_file
    mode = _benchmark_nemo_normalize_window_mode(window_mode)
    scoped = dict(operator_file)
    rows = list(operator_file.get("rows") or [])
    scoped["rows"] = rows
    scoped["_sessionStatsRows"] = list(operator_file.get("_sessionStatsRows") or rows)
    if mode != "active_dl_session":
        return scoped

    transfer_sessions = list(operator_file.get("transferSessions") or [])
    if not transfer_sessions:
        transfer_sessions = _nemo_build_transfer_sessions(rows, operator_file.get("operator") or "UNKNOWN")
    intervals = _nemo_downlink_transfer_intervals(transfer_sessions)
    scoped_rows = _nemo_rows_within_intervals(rows, intervals)
    # Never-empty guard: when no downlink transfer window is detectable for this scope,
    # keep ALL rows (and flag the fallback) so KPIs are never blanked out. Without this a
    # DT lacking a clean DL session returns zero rows → every KPI null.
    if not scoped_rows:
        scoped["rows"] = rows
        scoped["_windowFallback"] = True
        scoped["transferSessions"] = [
            session for session in transfer_sessions
            if str(session.get("direction") or "").strip().lower().startswith("down")
        ]
        return scoped
    scoped["rows"] = scoped_rows
    scoped["transferSessions"] = [
        session for session in transfer_sessions
        if str(session.get("direction") or "").strip().lower().startswith("down")
    ]
    scoped["technologyStatus"] = _nemo_compute_technology_status(scoped_rows, str(scoped.get("operator") or "UNKNOWN"))
    scoped["has5g"] = bool((scoped.get("technologyStatus") or {}).get("has5g"))
    scoped["fiveGStatus"] = (scoped.get("technologyStatus") or {}).get("fiveGStatus")
    scoped["_dlMetricKeyOverride"] = _nemo_select_dl_metric_key(scoped_rows)
    scoped["_benchmarkDlMetricKeyAvgOverride"] = _nemo_select_benchmark_dl_metric_key(scoped_rows)
    scoped["_benchmarkDlMetricKeyOverride"] = scoped.get("_dlMetricKeyOverride") or _nemo_select_benchmark_dl_metric_key(scoped_rows)
    return scoped


def _benchmark_nemo_dt_dataset(
    index: int,
    dl_mode: str | None = None,
    window_mode: str | None = None,
) -> dict:
    """Run the full benchmark analysis on a single positional DT across all operators.

    For DT #index, filter each operator's rows to that operator's i-th drive test, then reuse
    `_benchmark_nemo_build_dataset`. Parsing is served from the per-file cache, so this is fast."""
    if index < 0:
        return {}
    dl_mode = _benchmark_nemo_normalize_dl_mode(dl_mode)
    window_mode = _benchmark_nemo_normalize_window_mode(window_mode)
    cached_dt_datasets = BENCHMARK_NEMO_DATASET.get("dt_datasets") or {}
    mode_cache = cached_dt_datasets.get(_benchmark_nemo_mode_cache_key(dl_mode, window_mode)) or {}
    cached_dataset = mode_cache.get(index)
    if cached_dataset:
        return cached_dataset

    paths = _benchmark_nemo_resolve_paths()
    if not paths:
        return {}
    valid_paths = [path for path in paths if os.path.isfile(path)]
    valid_mtimes = _benchmark_nemo_collect_mtimes(valid_paths)
    use_cached_operator_files = (
        valid_paths
        and list(BENCHMARK_NEMO_DATASET.get("paths") or []) == valid_paths
        and (BENCHMARK_NEMO_DATASET.get("path_mtimes") or {}) == valid_mtimes
        and isinstance(BENCHMARK_NEMO_DATASET.get("operator_files"), list)
        and BENCHMARK_NEMO_DATASET.get("operator_files")
    )
    operator_files = (
        BENCHMARK_NEMO_DATASET.get("operator_files") or []
        if use_cached_operator_files
        else _benchmark_nemo_parse_operator_files(valid_paths)
    )
    if not use_cached_operator_files:
        BENCHMARK_NEMO_DATASET["paths"] = valid_paths
        BENCHMARK_NEMO_DATASET["path_mtimes"] = valid_mtimes
        BENCHMARK_NEMO_DATASET["operator_files"] = operator_files
        BENCHMARK_NEMO_DATASET["dt_datasets"] = {}

    filtered = []
    for of in operator_files:
        clone = _nemo_clone_operator_file_for_dt_index_with_window(of, index, window_mode=window_mode)
        if clone:
            filtered.append(clone)
    if not filtered:
        return {}
    dataset = _benchmark_nemo_build_dataset(filtered, dl_mode=dl_mode, window_mode=window_mode)
    mode_key = _benchmark_nemo_mode_cache_key(dl_mode, window_mode)
    BENCHMARK_NEMO_DATASET.setdefault("dt_datasets", {}).setdefault(mode_key, {})[index] = dataset
    return dataset


# ── Benchmark-Optim: per-DT/per-operator KPI workbook export ───────────────────
BENCHMARK_OPTIM_TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "benchmark_optim_template.xlsx")
_BENCHMARK_OPTIM_OPERATOR_ORDER = ("IAM", "ORANGE", "INWI")
# Per-DT "Conclusion" columns merged across the 3 operator rows (mirrors the template).
_BENCHMARK_OPTIM_MERGED_COLS = ("A", "K", "AD", "AM", "AU", "BD", "BN")


def _optim_round(value, ndigits=1):
    try:
        if value is None:
            return None
        return round(float(value), ndigits)
    except Exception:
        return None


def _optim_dist_text(distribution, top=3):
    """Render a [{label, share}] distribution as 'EN-DC (58.5%), LTE FDD (30%)'."""
    items = [d for d in (distribution or []) if isinstance(d, dict) and d.get("label")]
    if not items:
        return None
    parts = []
    for d in items[:top]:
        share = d.get("share")
        parts.append(f"{d.get('label')} ({share}%)" if share is not None else str(d.get("label")))
    return ", ".join(parts)


def _optim_serving_sequence(serving_cells, op_rows=None):
    """Compact chronological serving-cell sequence, e.g. '5G_Cell_A → 4G_Cell_B'.

    Primary source is the BDD-named episode sequence from the serving-cell analysis. That
    needs GPS (for the spatial BDD match); some operators/DTs have no GPS in the export
    (IAM in particular), so fall back to the serving-PCI chronology from the raw rows so the
    column is still filled (matches the PCI-labelled sequence other operators show)."""
    seq = []
    for ep in ((serving_cells or {}).get("episodesAll") or (serving_cells or {}).get("episodes") or []):
        name = (ep or {}).get("cellName")
        if name and (not seq or seq[-1] != name):
            seq.append(name)
    if seq:
        return " → ".join(seq)

    from datetime import datetime as _dt
    ordered = sorted([r for r in (op_rows or []) if isinstance(r.get("_dt"), _dt)], key=lambda r: r["_dt"])
    pseq = []
    last_key = None
    for row in ordered:
        pci = row.get("pci")
        if pci is None:
            continue
        try:
            pci_i = int(round(float(pci)))
        except Exception:
            continue
        is_nr = (row.get("nrChannelNumber") is not None
                 or "EN-DC" in str(row.get("servingTechnology") or "").upper())
        key = (is_nr, pci_i)
        if key == last_key:  # dedup by cell identity (ignore intermittent blank band)
            continue
        band = str(row.get("band") or "").strip()
        pseq.append(("5G" if is_nr else "4G") + f" PCI {pci_i}" + (f" ({band})" if band else ""))
        last_key = key
    if not pseq:
        return None
    if len(pseq) > 30:
        pseq = pseq[:30] + ["…"]
    return " → ".join(pseq)


def _optim_operator_row_values(op, dataset):
    """Build {column_letter: value} for one operator within one DT dataset.

    Reuses the same analyses the web UI shows — operator['kpis'] (_nemo_operator_kpis),
    operator['technologyStatus'], and the per-operator rows of the section analyses."""
    kpis = op.get("kpis") or {}
    ts = op.get("technologyStatus") or {}
    op_name = str(op.get("operator") or "").upper()

    def med(section):
        return (kpis.get(section) or {}).get("median")

    def avg(section):
        return (kpis.get(section) or {}).get("average")

    def p90(section):
        return (kpis.get(section) or {}).get("p90")

    def _find(analysis_key, default=None):
        rows = (dataset.get(analysis_key) or {}).get("rows") or []
        for r in rows:
            if str((r or {}).get("operator") or "").upper() == op_name:
                return r
        return default or {}

    band = _find("nrBandExposureAnalysis")
    ca = _find("caScellsAnalysis")
    mimo = _find("mimoRankAnalysis")
    bler = _find("blerRetxAnalysis")
    transport = _find("transportGapAnalysis")
    mod = kpis.get("pdschModulation") or {}
    serving_key = {"IAM": "iamServingCells", "ORANGE": "orangeServingCells", "INWI": "inwiServingCells"}.get(op_name)
    serving = dataset.get(serving_key) if serving_key else None

    vals = {
        "D": _optim_round(avg("dl")),
        "E": _optim_serving_sequence(serving, op.get("rows")),
        # RF Exclusion (per-operator medians)
        "F": _optim_round(med("rsrp")), "G": _optim_round(med("sinr")), "H": _optim_round(med("cqi")),
        "I": _optim_round(med("pdschMcs")), "J": _optim_round(med("pdschBitPerHz"), 2),
        # 5G Presence
        "L": _optim_round(kpis.get("nrPresencePct")), "M": _optim_round(kpis.get("lteOnlyPresencePct")),
        "N": _optim_dist_text(ts.get("packetTechnologyDistribution")),
        # Band distribution
        "O": _optim_round(band.get("n78Share")), "P": _optim_round(band.get("n1Share")),
        "Q": _optim_round(band.get("n28Share")), "R": _optim_round(band.get("otherNrBandShare")),
        # PDSCH Modulation & Spectral Efficiency
        "S": mod.get("dominant"), "T": _optim_round(mod.get("qam256Share")), "U": _optim_round(mod.get("qam64Share")),
        "V": _optim_round(mod.get("qam16Share")), "W": _optim_round(mod.get("qpskShare")),
        "X": _optim_round(med("pdschMcs")), "Y": _optim_round(med("scheduledRank")),
        "Z": _optim_round(med("pdschBitPerHz"), 2), "AA": _optim_round(kpis.get("prbEfficiency"), 3),
        "AB": kpis.get("pdschActiveSampleCount"),
        # MIMO Rank
        "AE": _optim_round(mimo.get("ri1Share")), "AF": _optim_round(mimo.get("ri2Share")),
        "AG": _optim_round(mimo.get("ri3Share")), "AH": _optim_round(mimo.get("ri4Share")),
        "AI": _optim_round(mimo.get("riGe3Share")), "AJ": _optim_round(mimo.get("medianRi")),
        "AK": _optim_round(mimo.get("averageRi")), "AL": _optim_dist_text(mimo.get("scheduledRankDistribution")),
        # Carrier Aggregation / SCells
        "AN": _optim_dist_text(ca.get("lteCaDistribution"), top=1), "AO": _optim_dist_text(ca.get("nrCaDistribution"), top=1),
        "AP": _optim_round(ca.get("avgScells"), 2), "AQ": _optim_round(ca.get("maxScells"), 0),
        "AR": _optim_round(ca.get("scellsActiveShare")), "AS": _optim_round(ca.get("lteCaActiveShare")),
        "AT": _optim_round(ca.get("nrCaActiveShare")),
        # BLER / Retransmission
        "AV": _optim_round(bler.get("blerAvg")), "AW": _optim_round(bler.get("blerMedian")),
        "AX": _optim_round(bler.get("blerP90")), "AY": _optim_round(bler.get("blerGt10Share")),
        "AZ": _optim_round(bler.get("blerGt20Share")), "BA": _optim_round(bler.get("ulRetxAvg")),
        "BB": _optim_round(bler.get("ulRetxMedian")), "BC": _optim_round(bler.get("ulRetxP90")),
        # Transport / Core / Application Gap
        "BE": _optim_round(transport.get("appDlAvg")), "BF": _optim_round(transport.get("totalMacDlAvg")),
        "BG": _optim_round(transport.get("macDl5gAvg")), "BH": _optim_round(transport.get("pdschDlAvg")),
        "BI": _optim_round(transport.get("tcpHandshakeMedian")), "BJ": _optim_round(transport.get("lostPacketAvg"), 2),
        "BK": _optim_round(transport.get("pingSuccessRate")), "BL": _optim_round(transport.get("appVsTotalMacRatio"), 2),
        "BM": _optim_round(transport.get("appVsPdschRatio"), 2),
    }
    return vals


def _optim_dt_conclusions(dataset):
    """Per-DT text for the merged Conclusion columns (English)."""
    return {
        "K": (dataset.get("rfExclusionCheck") or {}).get("conclusion"),
        "AD": (dataset.get("pdschModulationEfficiencyAnalysis") or {}).get("conclusion"),
        "AM": (dataset.get("mimoRankAnalysis") or {}).get("interpretation"),
        "AU": (dataset.get("caScellsAnalysis") or {}).get("interpretation"),
        "BD": (dataset.get("blerRetxAnalysis") or {}).get("interpretation"),
        "BN": (dataset.get("transportGapAnalysis") or {}).get("interpretation"),
    }


def _benchmark_optim_build_dt_analysis(operator_files: list[dict], dl_mode: str | None = None) -> dict:
    """Trimmed per-DT analysis: only what the Benchmark-Optim workbook needs.

    Same results as `_benchmark_nemo_build_dataset` for the columns we export (operator
    KPIs with dwell-based presence, the 7 section analyses, serving cells), but skips the
    heavy unused builders (GPS trace, recommendations, weakness chains, layer throughput,
    scheduler deep-dive, executive conclusion, charts, …)."""
    dl_mode = _benchmark_nemo_normalize_dl_mode(dl_mode)
    for of in operator_files:
        _nemo_reapply_throughput_normalization(of)
        of["dlMetricKey"] = _nemo_select_dl_metric_key(of.get("rows") or [])
        of["benchmarkDlMetricKeyDefault"] = _nemo_select_benchmark_dl_metric_key(of.get("rows") or [])
        of["benchmarkDlMetricKey"] = _benchmark_nemo_canonical_dl_metric_key(of, dl_mode)
        of["tests"] = _nemo_build_tests(
            of.get("rows") or [],
            of.get("operator") or "UNKNOWN",
            of.get("benchmarkDlMetricKey") or "",
        )
        of["transferSessions"] = _nemo_build_transfer_sessions(of.get("rows") or [], of.get("operator") or "UNKNOWN")
        of["kpis"] = _nemo_operator_kpis(of)

    ranking = _nemo_build_ranking(operator_files)
    diagnosis = _nemo_build_diagnosis(operator_files, ranking)

    serving_by_op = {}
    for of in operator_files:
        cells = _nemo_build_operator_serving_cells(of)
        if cells and cells.get("available"):
            cells = _nemo_attach_serving_cell_presence_metadata(
                cells, of.get("technologyStatus"), _nemo_dominant_nr_serving_info(of.get("rows") or [])
            )
        serving_by_op[str(of.get("operator") or "").upper()] = cells
        # Align 5G/4G presence KPIs with the dwell-based serving-cell breakdown (matches the UI).
        breakdown = (cells or {}).get("radioPresenceBreakdownAll") or (cells or {}).get("radioPresenceBreakdown") or {}
        kpis = of.get("kpis")
        if isinstance(breakdown, dict) and breakdown and isinstance(kpis, dict):
            if breakdown.get("5G") is not None:
                kpis["nrPresencePct"] = breakdown.get("5G")
            if breakdown.get("4G") is not None:
                kpis["lteOnlyPresencePct"] = breakdown.get("4G")

    return {
        "rfExclusionCheck": _nemo_build_rf_exclusion_check(operator_files, diagnosis),
        "pdschModulationEfficiencyAnalysis": _nemo_build_pdsch_modulation_efficiency(operator_files, diagnosis),
        "nrBandExposureAnalysis": _nemo_build_nr_band_exposure_analysis(operator_files, diagnosis),
        "caScellsAnalysis": _nemo_build_ca_scells_analysis(operator_files, diagnosis),
        "mimoRankAnalysis": _nemo_build_mimo_rank_analysis(operator_files, diagnosis),
        "blerRetxAnalysis": _nemo_build_bler_retx_analysis(operator_files, diagnosis),
        "transportGapAnalysis": _nemo_build_transport_gap_analysis(operator_files, diagnosis),
        "iamServingCells": serving_by_op.get("IAM"),
        "orangeServingCells": serving_by_op.get("ORANGE"),
        "inwiServingCells": serving_by_op.get("INWI"),
        "ranking": ranking,
        "diagnosis": diagnosis,
    }


def _optim_download_metrics(transfer_sessions: list[dict]) -> dict:
    """Downlink completion % and success rate for one operator/DT, from transfer sessions."""
    summary = _nemo_build_transfer_summary(transfer_sessions or [])
    dl = next((s for s in summary if str(s.get("direction") or "").lower().startswith("down")), {})
    return {
        "completionPct": dl.get("avgCompletionPct"),
        "successRate": dl.get("successRate"),
        "transferCount": dl.get("transferCount"),
    }


def _optim_iam_download_summary(operator_files: list[dict]) -> dict:
    iam = next((o for o in operator_files if str(o.get("operator") or "").upper() == "IAM"), None)
    if not iam:
        return {}
    sessions = iam.get("transferSessions") or _nemo_build_transfer_sessions(iam.get("rows") or [], "IAM")
    return _optim_download_metrics(sessions)


def _optim_aggregate_deep_analysis(dt_diags: list[dict]) -> dict:
    """Aggregate the per-DT IAM diagnoses into an honest, actionable picture.

    A single cumulative average hides where IAM actually trails competitors (and degenerates
    to an IAM-vs-IAM comparison when IAM leads on average), so we summarise the per-DT
    diagnoses: how often IAM is behind the best operator, by how much, and the dominant cause."""
    total = len(dt_diags)
    behind_rows, gaps = [], []
    cause_counts: dict = {}
    iam_best_count = 0
    all_rows = []
    for d in dt_diags:
        diag = d.get("diagnosis") or {}
        ranking = d.get("ranking") or []
        iam = next((e for e in ranking if str(e.get("operator") or "").upper() == "IAM"), {})
        best = ranking[0] if ranking else {}
        iam_dl = iam.get("avgDlMbps")
        best_dl = best.get("avgDlMbps")
        best_op = best.get("operator")
        cause = diag.get("displayMainCause") or diag.get("mainCause")
        behind = (
            str(best_op or "").upper() != "IAM"
            and iam_dl is not None and best_dl not in (None, 0) and iam_dl < best_dl
        )
        gap = round((iam_dl - best_dl) / float(best_dl) * 100.0, 1) if behind else None
        row = {"dt": d.get("dtLabel"), "iamDl": iam_dl, "bestOp": best_op, "bestDl": best_dl,
                "gap": gap, "cause": cause if behind else None}
        all_rows.append(row)
        if behind:
            behind_rows.append(row)
            gaps.append(gap)
            if cause:
                cause_counts[cause] = cause_counts.get(cause, 0) + 1
        elif str(best_op or "").upper() == "IAM":
            iam_best_count += 1
    dominant = max(cause_counts, key=cause_counts.get) if cause_counts else None
    worst = sorted(behind_rows, key=lambda r: (r.get("gap") if r.get("gap") is not None else 0))[:12]
    return {
        "totalDts": total,
        "iamBestCount": iam_best_count,
        "behindCount": len(behind_rows),
        "avgGapPct": round(sum(gaps) / len(gaps), 1) if gaps else None,
        "worstGapPct": min(gaps) if gaps else None,
        "dominantCause": dominant,
        "dominantCount": cause_counts.get(dominant, 0) if dominant else 0,
        "causeCounts": dict(sorted(cause_counts.items(), key=lambda kv: -kv[1])),
        "worst": worst,
    }


def _optim_presence_table(presence_acc: dict, global_serving: dict = None) -> list:
    """Cumulative radio-presence % per operator (all-window + download-window).

    Prefers the loaded global serving-cells breakdowns (identical to what the webapp shows,
    e.g. IAM 5G 4.5%); falls back to dwell-second accumulation across the per-DT serving cells
    when the global serving cells aren't available (upload path)."""
    names = {"IAM": "IAM", "ORANGE": "Orange", "INWI": "INWI"}
    rows = []
    for op in ("IAM", "ORANGE", "INWI"):
        sc = (global_serving or {}).get(op) if isinstance(global_serving, dict) else None
        ba = (sc or {}).get("radioPresenceBreakdownAll") or {}
        bd = (sc or {}).get("radioPresenceBreakdownDownload") or {}
        if ba or bd:
            rows.append({
                "operator": names[op],
                "all5g": ba.get("5G") or 0.0, "all4g": ba.get("4G") or 0.0,
                "dl5g": bd.get("5G") or 0.0, "dl4g": bd.get("4G") or 0.0,
            })
            continue
        a5, a4, d5, d4 = presence_acc.get(op, [0.0, 0.0, 0.0, 0.0])
        all_tot = a5 + a4
        dl_tot = d5 + d4
        rows.append({
            "operator": names[op],
            "all5g": round(a5 / all_tot * 100.0, 1) if all_tot else 0.0,
            "all4g": round(a4 / all_tot * 100.0, 1) if all_tot else 0.0,
            "dl5g": round(d5 / dl_tot * 100.0, 1) if dl_tot else 0.0,
            "dl4g": round(d4 / dl_tot * 100.0, 1) if dl_tot else 0.0,
        })
    return rows


def _optim_write_deep_analysis_sheet(wb, agg: dict, iam_download: dict, dl_timeline=None, presence_table=None):
    """Render a professional 'IAM Deep Analysis' sheet from the per-DT aggregation: verdict,
    executive summary, per-DT under-performance table, root-cause breakdown, prioritized
    actions, and download QoS."""
    import openpyxl
    if not agg or not agg.get("totalDts"):
        return
    total = int(agg.get("totalDts") or 0)
    behind = int(agg.get("behindCount") or 0)
    best_cnt = int(agg.get("iamBestCount") or 0)
    dominant = agg.get("dominantCause")
    avg_gap = agg.get("avgGapPct")
    worst_gap = agg.get("worstGapPct")

    ws = wb.create_sheet("IAM Deep Analysis", 0)
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 30), ("B", 16), ("C", 14), ("D", 14), ("E", 60)):
        ws.column_dimensions[col].width = w

    TITLE = openpyxl.styles.Font(bold=True, size=16, color="1F4E78")
    SECTION = openpyxl.styles.Font(bold=True, size=12, color="FFFFFF")
    SECTION_FILL = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    KEY = openpyxl.styles.Font(bold=True, color="333333")
    HDR = openpyxl.styles.Font(bold=True, color="FFFFFF")
    HDR_FILL = openpyxl.styles.PatternFill("solid", fgColor="4472C4")
    wrap = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    r = [1]

    def newline(n=1):
        r[0] += n

    def section(title):
        ws.cell(r[0], 1, title).font = SECTION
        for c in range(1, 6):
            ws.cell(r[0], c).fill = SECTION_FILL
        ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=5)
        newline(2)

    def kv(key, value):
        ws.cell(r[0], 1, key).font = KEY
        vc = ws.cell(r[0], 2, "" if value is None else str(value))
        vc.alignment = wrap
        ws.merge_cells(start_row=r[0], start_column=2, end_row=r[0], end_column=5)
        newline()

    def para(text):
        if not text:
            return
        c = ws.cell(r[0], 1, str(text))
        c.alignment = wrap
        ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=5)
        ws.row_dimensions[r[0]].height = max(15, min(120, 15 * (1 + len(str(text)) // 95)))
        newline()

    tc = ws.cell(r[0], 1, "IAM \u2014 Benchmark Deep Analysis (Root Cause & Actions)")
    tc.value = "IAM — Benchmark Deep Analysis (Root Cause & Actions)"
    tc.font = TITLE
    ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=5)
    newline(2)

    # Verdict
    section("Verdict")
    kv("Drive tests analysed", total)
    kv("IAM is the best operator in", f"{best_cnt} / {total} DTs")
    kv("IAM trails the best competitor in", f"{behind} / {total} DTs"
        + (f" ({round(behind/total*100)}%)" if total else ""))
    if avg_gap is not None:
        kv("Average DL gap when behind", f"{avg_gap}%")
    if worst_gap is not None:
        kv("Worst DL gap", f"{worst_gap}%")
    rc_row = r[0]
    ws.cell(rc_row, 1, "Dominant root cause").font = KEY
    rc = ws.cell(rc_row, 2, (f"{dominant} (in {agg.get('dominantCount')} DTs)" if dominant else "IAM leads — no dominant weakness"))
    rc.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    rc.fill = openpyxl.styles.PatternFill("solid", fgColor=("C00000" if behind > total * 0.4 else "E06C00" if dominant else "548235"))
    ws.merge_cells(start_row=rc_row, start_column=2, end_row=rc_row, end_column=5)
    newline(2)

    # Executive summary (synthesised)
    section("Executive Summary")
    if behind == 0:
        para("• IAM is the strongest operator for measured DL throughput across all analysed drive tests; no recurring competitive weakness was found.")
    else:
        para(f"• IAM trails the best competitor in {behind} of {total} drive tests"
             + (f", with an average DL throughput gap of {avg_gap}% (worst {worst_gap}%)." if avg_gap is not None else "."))
        if dominant:
            para(f"• The dominant root cause across under-performing drive tests is “{dominant}” ({agg.get('dominantCount')} DTs).")
        causes = agg.get("causeCounts") or {}
        others = [f"{c} ({n})" for c, n in list(causes.items())[1:4]]
        if others:
            para("• Secondary causes: " + ", ".join(others) + ".")
        para("• See the per-DT table below for the specific drive tests, competitors and gaps, and the prioritised actions for the dominant cause.")
    newline()

    # Per-DT under-performance table
    worst = agg.get("worst") or []
    if worst:
        section("Worst Drive Tests (IAM behind best competitor)")
        for col, label in enumerate(["DT", "IAM DL (Mbps)", "Best op", "Best DL (Mbps)", "Gap % / Root cause"], start=1):
            hc = ws.cell(r[0], col, label)
            hc.font = HDR
            hc.fill = HDR_FILL
        newline()
        for row in worst:
            ws.cell(r[0], 1, row.get("dt"))
            ws.cell(r[0], 2, row.get("iamDl"))
            ws.cell(r[0], 3, row.get("bestOp"))
            ws.cell(r[0], 4, row.get("bestDl"))
            ec = ws.cell(r[0], 5, (f"{row.get('gap')}%  —  {row.get('cause') or ''}").strip())
            ec.alignment = wrap
            newline()
        newline()

    # Root-cause breakdown
    causes = agg.get("causeCounts") or {}
    if causes:
        section("Root-Cause Breakdown (across under-performing DTs)")
        for col, label in enumerate(["Root cause", "# DTs", "Share"], start=1):
            hc = ws.cell(r[0], col, label)
            hc.font = HDR
            hc.fill = HDR_FILL
        newline()
        for cause, n in causes.items():
            ws.cell(r[0], 1, cause).alignment = wrap
            ws.cell(r[0], 2, n)
            ws.cell(r[0], 3, f"{round(n / behind * 100)}%" if behind else "")
            newline()
        newline()

    # Prioritised actions for the dominant cause
    if dominant:
        section(f"Recommended Actions — {dominant}")
        steps = _nemo_recommendations_for_cause(dominant)
        if not steps:
            para("No predefined action list for this cause — review the per-DT evidence above.")
        for i, step in enumerate(steps, start=1):
            para(f"   {i}. {step}")
        newline()

    # Download / QoS
    dl = iam_download or {}
    section("Download / QoS Performance (IAM, all DTs)")
    kv("Download success rate", None if dl.get("successRate") is None else f"{dl.get('successRate')}%")
    kv("Avg file download completion", None if dl.get("completionPct") is None else f"{dl.get('completionPct')}%")
    kv("Download transfers analysed", dl.get("transferCount"))
    newline()

    from openpyxl.chart import BarChart, LineChart, Reference

    # Radio presence (time-based) per operator — table + two bar charts.
    presence_table = presence_table or []
    if presence_table:
        section("Radio Presence (time-based) — IAM vs Orange vs INWI")
        hdr_row = r[0]
        for col, label in enumerate(["Operator", "5G % (all-window)", "4G % (all-window)", "5G % (download)", "4G % (download)"], start=1):
            hc = ws.cell(hdr_row, col, label)
            hc.font = HDR
            hc.fill = HDR_FILL
        newline()
        for prow in presence_table:
            ws.cell(r[0], 1, prow.get("operator"))
            ws.cell(r[0], 2, prow.get("all5g"))
            ws.cell(r[0], 3, prow.get("all4g"))
            ws.cell(r[0], 4, prow.get("dl5g"))
            ws.cell(r[0], 5, prow.get("dl4g"))
            newline()
        data_last = r[0] - 1
        cats = Reference(ws, min_col=1, min_row=hdr_row + 1, max_row=data_last)

        bar_all = BarChart()
        bar_all.type = "col"
        bar_all.title = "Radio presence (time-based, all-window)"
        bar_all.y_axis.title = "%"
        bar_all.height = 7.5
        bar_all.width = 13
        bar_all.add_data(Reference(ws, min_col=2, max_col=3, min_row=hdr_row, max_row=data_last), titles_from_data=True)
        bar_all.set_categories(cats)
        ws.add_chart(bar_all, f"G{hdr_row}")

        bar_dl = BarChart()
        bar_dl.type = "col"
        bar_dl.title = "Radio presence (time-based, download)"
        bar_dl.y_axis.title = "%"
        bar_dl.height = 7.5
        bar_dl.width = 13
        bar_dl.add_data(Reference(ws, min_col=4, max_col=5, min_row=hdr_row, max_row=data_last), titles_from_data=True)
        bar_dl.set_categories(cats)
        ws.add_chart(bar_dl, f"G{hdr_row + 16}")
        newline(16)

    # DL throughput timeline (App layer) — full per-DT ranking table + line chart.
    timeline = [t for t in (dl_timeline or []) if any(t.get(k) is not None for k in ("IAM", "ORANGE", "INWI"))]
    if timeline:
        section("DL Throughput by Drive Test (App Layer) — IAM vs Orange vs INWI")
        hdr_row = r[0]
        for col, label in enumerate(["DT", "IAM (Mbps)", "Orange (Mbps)", "INWI (Mbps)"], start=1):
            hc = ws.cell(hdr_row, col, label)
            hc.font = HDR
            hc.fill = HDR_FILL
        newline()
        for t in timeline:
            ws.cell(r[0], 1, t.get("dt"))
            ws.cell(r[0], 2, t.get("IAM"))
            ws.cell(r[0], 3, t.get("ORANGE"))
            ws.cell(r[0], 4, t.get("INWI"))
            newline()
        data_last = r[0] - 1
        line = LineChart()
        line.title = "DL Throughput Timeline — IAM vs INWI vs Orange (App Layer)"
        line.y_axis.title = "DL throughput (Mbps)"
        line.x_axis.title = "Drive test"
        line.height = 9
        line.width = 26
        line.add_data(Reference(ws, min_col=2, max_col=4, min_row=hdr_row, max_row=data_last), titles_from_data=True)
        line.set_categories(Reference(ws, min_col=1, min_row=hdr_row + 1, max_row=data_last))
        ws.add_chart(line, f"F{hdr_row}")


def generate_benchmark_optim_xlsx(txt_paths: list[str], dl_mode: str | None = None) -> bytes:
    """Parse the Nemo TXT files then fill the workbook. Thin wrapper kept for the upload path."""
    operator_files = _benchmark_nemo_parse_operator_files(txt_paths)
    if not operator_files:
        raise ValueError("No usable Nemo TXT files were parsed.")
    return generate_benchmark_optim_xlsx_from_operator_files(operator_files, dl_mode=dl_mode)


def generate_benchmark_optim_xlsx_from_operator_files(operator_files: list[dict], global_serving: dict = None, dl_mode: str | None = None) -> bytes:
    """Run the per-DT benchmark analysis for every DT and fill the Output.xlsx template
    (one 3-row block per DT: IAM/Orange/INWI). Accepts already-parsed operator files so a
    benchmark loaded in the app (in-memory / SQLite library) can be reused without re-parsing.
    ``global_serving`` (optional, keyed by operator) carries the loaded all-DT serving cells so
    the radio-presence charts match the webapp's cumulative panel exactly."""
    import io
    import openpyxl
    from openpyxl.utils import column_index_from_string
    from copy import copy as _copy

    dl_mode = _benchmark_nemo_normalize_dl_mode(dl_mode)
    operator_files = list(operator_files or [])
    if not operator_files:
        raise ValueError("No operator data available.")

    # Map ordered titles per operator and the DT count (max titles across operators).
    titles_by_op = {}
    for of in operator_files:
        op_name = str(of.get("operator") or "UNKNOWN").upper()
        titles_by_op[op_name] = list(of.get("orderedDtTitles") or _nemo_ordered_dt_titles(of.get("rows") or []))
    dt_count = max((len(t) for t in titles_by_op.values()), default=0)
    if dt_count == 0:
        raise ValueError("No drive tests (measurement titles) found in the TXT files.")

    if not os.path.isfile(BENCHMARK_OPTIM_TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {BENCHMARK_OPTIM_TEMPLATE_PATH}")
    wb = openpyxl.load_workbook(BENCHMARK_OPTIM_TEMPLATE_PATH)
    ws = wb.worksheets[0]
    # Base style copied from the template's header row so data cells match.
    body_font = _copy(ws.cell(row=2, column=1).font)
    thin = openpyxl.styles.Side(style="thin", color="DDDDDD")
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_left = openpyxl.styles.Alignment(horizontal="left", vertical="top", wrap_text=True)
    center = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Extend the matrix with a "Download QoS" section (cols BO, BP): % of file downloaded + success rate.
    bo = column_index_from_string("BO")
    bp = column_index_from_string("BP")
    ws.cell(row=1, column=bo, value="Download QoS").font = _copy(ws.cell(row=1, column=6).font)
    ws.cell(row=1, column=bo).fill = _copy(ws.cell(row=1, column=6).fill)
    ws.merge_cells(start_row=1, start_column=bo, end_row=1, end_column=bp)
    ws.cell(row=2, column=bo, value="DL completion %").font = _copy(ws.cell(row=2, column=1).font)
    ws.cell(row=2, column=bp, value="DL success %").font = _copy(ws.cell(row=2, column=1).font)

    # Extend with an "IAM Root-Cause" section (cols BQ, BR): per-DT IAM root cause + actions.
    bq = column_index_from_string("BQ")
    br = column_index_from_string("BR")
    ws.cell(row=1, column=bq, value="IAM Root-Cause Analysis").font = _copy(ws.cell(row=1, column=6).font)
    ws.cell(row=1, column=bq).fill = _copy(ws.cell(row=1, column=6).fill)
    ws.merge_cells(start_row=1, start_column=bq, end_row=1, end_column=br)
    ws.cell(row=2, column=bq, value="Root cause (IAM)").font = _copy(ws.cell(row=2, column=1).font)
    ws.cell(row=2, column=br, value="Recommended actions (IAM)").font = _copy(ws.cell(row=2, column=1).font)
    ws.column_dimensions["BQ"].width = 28
    ws.column_dimensions["BR"].width = 60

    row_cursor = 3
    dt_diags = []  # per-DT IAM diagnosis + ranking, for the deep-analysis aggregation
    dl_timeline = []  # per-DT App-layer DL Mbps for IAM/Orange/INWI (line chart + ranking table)
    presence_acc = {}  # per-operator cumulative dwell seconds: {op: [all5g, all4g, dl5g, dl4g]}
    for dt_index in range(dt_count):
        # Build the per-DT dataset by filtering each operator to its i-th drive test.
        filtered = []
        for of in operator_files:
            titles = list(of.get("orderedDtTitles") or _nemo_ordered_dt_titles(of.get("rows") or []))
            if dt_index >= len(titles):
                continue
            target = titles[dt_index]
            rows_by_title = of.get("rowsByMeasurementTitle") or {}
            sub_rows = list(rows_by_title.get(target) or [])
            if not sub_rows:
                sub_rows = [r for r in (of.get("rows") or []) if r.get("measurementTitle") == target]
            if not sub_rows:
                continue
            clone = dict(of)
            clone["rows"] = sub_rows
            clone["measurementTitles"] = [target]
            clone["orderedDtTitles"] = [target]
            clone["rowsByMeasurementTitle"] = {target: sub_rows}
            clone["technologyStatus"] = _nemo_compute_technology_status(sub_rows, str(clone.get("operator") or "UNKNOWN"))
            clone["has5g"] = bool((clone.get("technologyStatus") or {}).get("has5g"))
            clone["fiveGStatus"] = (clone.get("technologyStatus") or {}).get("fiveGStatus")
            clone["_dtTitle"] = target
            filtered.append(clone)
        if not filtered:
            continue
        dataset = _benchmark_optim_build_dt_analysis(filtered, dl_mode=dl_mode)
        by_op = {str(o.get("operator") or "").upper(): o for o in filtered}
        conclusions = _optim_dt_conclusions(dataset)
        dt_diags.append({
            "dtLabel": f"DT{dt_index + 1}",
            "diagnosis": dataset.get("diagnosis"),
            "ranking": dataset.get("ranking"),
        })
        _rank = {str(e.get("operator") or "").upper(): e.get("avgDlMbps") for e in (dataset.get("ranking") or [])}
        dl_timeline.append({
            "dt": f"DT{dt_index + 1}",
            "IAM": _rank.get("IAM"), "ORANGE": _rank.get("ORANGE"), "INWI": _rank.get("INWI"),
        })

        block_start = row_cursor
        present_ops = [op for op in _BENCHMARK_OPTIM_OPERATOR_ORDER if op in by_op]
        # Keep any operator whose name isn't in the canonical order at the end.
        present_ops += [op for op in by_op if op not in present_ops]
        for op_name in present_ops:
            op = by_op[op_name]
            ws.cell(row=row_cursor, column=2, value=op.get("_dtTitle"))  # B File name
            ws.cell(row=row_cursor, column=3, value=op.get("operator"))  # C Operator
            for col_letter, value in _optim_operator_row_values(op, dataset).items():
                if value is not None and value != "":
                    ws.cell(row=row_cursor, column=column_index_from_string(col_letter), value=value)
            _dlm = _optim_download_metrics(op.get("transferSessions"))
            if _dlm.get("completionPct") is not None:
                ws.cell(row=row_cursor, column=bo, value=_dlm["completionPct"])
            if _dlm.get("successRate") is not None:
                ws.cell(row=row_cursor, column=bp, value=_dlm["successRate"])
            if op_name == "IAM":
                _diag = dataset.get("diagnosis") or {}
                _rc = _diag.get("displayMainCause") or _diag.get("mainCause")
                if _rc:
                    ws.cell(row=row_cursor, column=bq, value=_rc)
                _acts = _diag.get("recommendations") or []
                if _acts:
                    ws.cell(row=row_cursor, column=br,
                            value="\n".join(f"{i}. {a}" for i, a in enumerate(_acts[:4], 1)))
            # Accumulate cumulative presence dwell-seconds (all-window + download) per operator.
            _sc_key = {"IAM": "iamServingCells", "ORANGE": "orangeServingCells", "INWI": "inwiServingCells"}.get(op_name)
            _sc = dataset.get(_sc_key) if _sc_key else None
            _acc = presence_acc.setdefault(op_name, [0.0, 0.0, 0.0, 0.0])
            for _cell in ((_sc or {}).get("cells") or []):
                _tech = str(_cell.get("tech") or "").upper()
                _ds = float(_cell.get("dwellSec") or 0); _dd = float(_cell.get("dwellSecDownload") or 0)
                if _tech.startswith(("5G", "NR")):
                    _acc[0] += _ds; _acc[2] += _dd
                elif _tech.startswith("4G"):
                    _acc[1] += _ds; _acc[3] += _dd
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_cursor, column=c)
                cell.font = _copy(body_font)
                cell.border = border
                cell.alignment = wrap_left
            row_cursor += 1
        block_end = row_cursor - 1
        if block_end < block_start:
            continue

        # DT label (col A) + per-DT conclusions, written on the first row then merged down.
        ws.cell(row=block_start, column=1, value=f"DT{dt_index + 1}")
        for col_letter, text in conclusions.items():
            if text:
                ws.cell(row=block_start, column=column_index_from_string(col_letter), value=text)
        if block_end > block_start:
            for col_letter in _BENCHMARK_OPTIM_MERGED_COLS:
                ci = column_index_from_string(col_letter)
                ws.merge_cells(start_row=block_start, start_column=ci, end_row=block_end, end_column=ci)
        for col_letter in _BENCHMARK_OPTIM_MERGED_COLS:
            ws.cell(row=block_start, column=column_index_from_string(col_letter)).alignment = center

    # Professional IAM deep-analysis sheet (root cause + actions), aggregated from the per-DT
    # diagnoses (honest: shows where IAM actually trails competitors, not a masking average).
    try:
        agg = _optim_aggregate_deep_analysis(dt_diags)
        iam_download = _optim_iam_download_summary(operator_files)
        presence_table = _optim_presence_table(presence_acc, global_serving)
        _optim_write_deep_analysis_sheet(wb, agg, iam_download, dl_timeline, presence_table)
    except Exception:
        import traceback
        traceback.print_exc()

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Deep Benchmark analysis (IAM-focused deterministic rule engine) ─────────────
# Layer-1 engine: KPI gap → severity → root cause → actions → priority, producing
# the 3-section output (Executive Summary, IAM KPI Benchmark, IAM Action Plan)
# that mirrors the user's manual workbook. It consumes the already-computed
# per-operator `kpis` for whatever scope (single DT or cumulative) the dataset was
# built on, so no KPI logic is duplicated. No AI: all text is rule-generated.

_DEEP_IAM_ALIASES = {"IAM", "MAROC TELECOM", "MAROCTELECOM", "IAM_MA"}
_DEEP_SEVERITY_WEIGHT = {"Critical": 5, "High": 4, "Medium": 3, "Low": 1}
_DEEP_IMPACT_WEIGHT = {"Throughput": 5, "Coverage": 4, "Stability": 4, "Latency": 3, "QoS": 3}
_DEEP_CONFIDENCE_WEIGHT = {"Direct KPI evidence": 3, "Benchmark evidence": 2, "Possible cause": 1}


def _deep_classify_gap(gap_pct, higher_is_better=True):
    """Severity from a fractional gap vs best competitor (spec §3). gap_pct e.g. -0.54."""
    if gap_pct is None:
        return "Low"
    if higher_is_better:
        if gap_pct <= -0.40:
            return "Critical"
        if gap_pct <= -0.20:
            return "High"
        if gap_pct <= -0.10:
            return "Medium"
        return "Low"
    if gap_pct >= 0.40:
        return "Critical"
    if gap_pct >= 0.20:
        return "High"
    if gap_pct >= 0.10:
        return "Medium"
    return "Low"


def _deep_priority_from_score(score):
    if score >= 11:
        return "P1"
    if score >= 8:
        return "P2"
    if score >= 5:
        return "P3"
    return "P4"


def _deep_confidence_label(confidence):
    return {
        "Direct KPI evidence": "High",
        "Benchmark evidence": "Medium",
        "Possible cause": "Low",
    }.get(confidence, "Low")


def _deep_benchmark_relevance(domain, gap_pct):
    if gap_pct is not None and gap_pct <= -0.30:
        return "Primary"
    if domain in {
        "Coverage / dominance",
        "Radio quality / interference",
        "SINR / interference",
        "Modulation profile",
        "5G capacity layer",
        "Scheduler / PRB efficiency",
        "Capacity / configuration",
        "Bandwidth / spectrum",
        "Carrier aggregation",
    }:
        return "Secondary"
    return "Context"


# Per-domain Action-Plan metadata (Owner / Expected impact / Validation target / impact class).
_DEEP_DOMAIN_META = {
    "5G capacity layer": {
        "owner": "RF Optimization + RAN Planning", "impact": "Throughput",
        "expectedImpact": "Unlock n78 5G capacity and lift peak/sustained DL throughput.",
        "validationTarget": "NR n78 share >50% and DL throughput gap vs best competitor <20%."},
    "Coverage / dominance": {
        "owner": "RF Optimization + RAN Planning", "impact": "Coverage",
        "expectedImpact": "Restore stronger serving-cell dominance and reduce overshooting / missing-sector exposure.",
        "validationTarget": "Median RSRP >-100 dBm, median SINR >5 dB, and route served by nearby dominant sectors."},
    "Radio quality / interference": {
        "owner": "RF Optimization", "impact": "Coverage",
        "expectedImpact": "Reduce interference / pollution and improve usable spectral efficiency.",
        "validationTarget": "Median RSRQ >-14 dB and median SINR >5 dB."},
    "MIMO / RI": {
        "owner": "RF Optimization + RAN Vendor", "impact": "Throughput",
        "expectedImpact": "Enable higher spatial multiplexing (RI>=2) and raise DL throughput.",
        "validationTarget": "Median RI>=2 and RI2 share >80%."},
    "SINR / interference": {
        "owner": "RF Optimization", "impact": "Coverage",
        "expectedImpact": "Improve SINR/CQI to unlock higher MCS and modulation.",
        "validationTarget": "Median SINR >=8 dB and CQI >=10."},
    "Modulation profile": {
        "owner": "RF Optimization", "impact": "Throughput",
        "expectedImpact": "Shift modulation toward 64QAM/256QAM for higher spectral efficiency.",
        "validationTarget": "256QAM share >10% and 16QAM share <30%."},
    "Scheduler / PRB efficiency": {
        "owner": "RAN Optimization + Vendor", "impact": "Throughput",
        "expectedImpact": "Raise spectral efficiency and PDSCH delivered throughput.",
        "validationTarget": "PRB efficiency gap vs best competitor <20%."},
    "Load / congestion": {
        "owner": "RAN Optimization + Capacity Planning", "impact": "Throughput",
        "expectedImpact": "Reduce congestion-driven throughput collapse during busy load windows.",
        "validationTarget": "DL PRB utilization <80% at tested hour or throughput preserved under high load."},
    "Capacity / configuration": {
        "owner": "RAN Optimization + Transport/Core", "impact": "Throughput",
        "expectedImpact": "Lift throughput when radio quality is already good by fixing scheduler, CA, QoS, bandwidth or transport limits.",
        "validationTarget": "DL throughput gap vs best competitor <20% under good radio quality."},
    "Bandwidth / spectrum": {
        "owner": "RAN Planning + Optimization", "impact": "Throughput",
        "expectedImpact": "Remove bandwidth disadvantage versus competitors through spectrum, CA or NR-layer optimization.",
        "validationTarget": "Available serving bandwidth comparable to best competitor and all intended carriers active."},
    "Carrier aggregation": {
        "owner": "Optimization + RAN Vendor", "impact": "Throughput",
        "expectedImpact": "Boost 4G/NSA DL capacity through persistent CA.",
        "validationTarget": "SCells>0 share >50% and Avg #SCells >1."},
    "LTE anchor / NSA dependency": {
        "owner": "RF Optimization", "impact": "Stability",
        "expectedImpact": "Improve LTE anchor quality so NSA 5G can be added and retained more consistently.",
        "validationTarget": "Median LTE-anchor SINR >5 dB while 5G presence stays >40%."},
    "EN-DC stability": {
        "owner": "Optimization + Transmission + Vendor", "impact": "Stability",
        "expectedImpact": "Improve NSA addition success and reduce EN-DC drops/resets.",
        "validationTarget": "EN-DC setup success >=95% and EN-DC drop rate <=2%."},
    "BLER / retransmissions": {
        "owner": "Optimization + Vendor", "impact": "Stability",
        "expectedImpact": "Stabilize throughput and reduce retransmission overhead.",
        "validationTarget": "BLER avg <2% and BLER>10% share <5%."},
    "UL quality": {
        "owner": "NOC + Optimization", "impact": "Stability",
        "expectedImpact": "Better TCP stability and faster session ramp-up.",
        "validationTarget": "UL Retx avg <1% and TCP handshake median <70 ms."},
    "Transport / core": {
        "owner": "Transport/Core + Optimization", "impact": "Latency",
        "expectedImpact": "Improve app-layer responsiveness and throughput ramp-up.",
        "validationTarget": "TCP handshake median <70 ms."},
    "Mobility / serving sequence": {
        "owner": "RF Optimization", "impact": "Stability",
        "expectedImpact": "Smoother 5G retention and fewer LTE-anchor fallbacks.",
        "validationTarget": "Reduced LTE-only time and stable EN-DC retention along the route."},
    "Retest governance": {
        "owner": "Optimization QA", "impact": "QoS",
        "expectedImpact": "Avoid wrong actions based on a single sample.",
        "validationTarget": "Consistent results across 3+ retests (busy hour + off-peak)."},
}


def _deep_num(value):
    try:
        if value is None:
            return None
        f = float(value)
        return f if math.isfinite(f) else None
    except Exception:
        return None


def _deep_best_competitor(values, higher_is_better=True):
    vals = [v for v in (values or []) if _deep_num(v) is not None]
    if not vals:
        return None
    return max(vals) if higher_is_better else min(vals)


def _deep_gap_fraction(iam, best, higher_is_better=True):
    iam_n, best_n = _deep_num(iam), _deep_num(best)
    if iam_n is None or best_n is None or best_n == 0:
        return None
    return (iam_n - best_n) / abs(best_n)


def _deep_extract(kpis, transfer_lookup, operator):
    """Flatten the per-operator kpis dict into the scalar metrics the engine needs."""
    kpis = kpis or {}
    def stat(key, field):
        return (kpis.get(key) or {}).get(field)
    def _rat_med_avg(nr_key, lte_key):
        # RAT-average of the two displayed medians, so "Total" stays consistent with the
        # NR/LTE rows (and equals the NR median when there is no LTE leg).
        vals = [v for v in (stat(nr_key, "median"), stat(lte_key, "median")) if v is not None]
        return round(sum(vals) / len(vals), 1) if vals else None
    mod = kpis.get("pdschModulation") or {}
    ri1 = _deep_num(kpis.get("ri1Share"))
    ri_ge3 = _deep_num(kpis.get("riGe3Share"))
    ri2 = round(100.0 - ri1 - ri_ge3, 1) if ri1 is not None and ri_ge3 is not None else None
    dl_tr = transfer_lookup.get((str(operator or "").upper(), "DL")) or {}
    return {
        "dlThroughput": stat("dl", "average"),
        "rsrp": stat("rsrp", "median"),
        "rsrq": stat("rsrq", "median"),
        "sinr": stat("sinr", "median"),
        "rsrpNr": stat("rsrpNr", "median"),
        "rsrpLte": stat("rsrpLte", "median"),
        "rsrpTotal": _rat_med_avg("rsrpNr", "rsrpLte"),
        "sinrNr": stat("sinrNr", "median"),
        "sinrLte": stat("sinrLte", "median"),
        "sinrTotal": _rat_med_avg("sinrNr", "sinrLte"),
        "rsrqNr": stat("rsrqNr", "median"),
        "rsrqLte": stat("rsrqLte", "median"),
        "cqi": stat("cqi", "median"),
        "mcs": stat("pdschMcs", "median"),
        "fivegPresence": _deep_num(kpis.get("nrPresencePct")),
        "fourgOnly": _deep_num(kpis.get("lteOnlyPresencePct")),
        "n78": _deep_num(kpis.get("n78ShareNrOnly")),
        "n1": _deep_num((kpis.get("nrBandShares") or {}).get("n1")),
        "n28": _deep_num(kpis.get("n28ShareNrOnly")),
        "qam256": _deep_num(mod.get("qam256Share")),
        "qam64": _deep_num(mod.get("qam64Share")),
        "qam16": _deep_num(mod.get("qam16Share")),
        "qpsk": _deep_num(mod.get("qpskShare")),
        "medianRank": stat("ri", "median"),
        "ri1": ri1,
        "ri2": ri2,
        "riGe3": ri_ge3,
        "scellsAvg": _deep_num(kpis.get("scellsAvgAll")),
        "scellsMax": _deep_num(kpis.get("scellsMax")),
        "scellsActive": _deep_num(kpis.get("scellsActiveShare")),
        "caActive": _deep_num(
            kpis.get("caActiveShare")
            if kpis.get("caActiveShare") is not None
            else kpis.get("lteCaActiveShare")
        ),
        "blerAvg": stat("bler", "average"),
        "blerP90": stat("bler", "p90"),
        "blerAbove10": _deep_num(kpis.get("blerAbove10Share")),
        "ulRetx": stat("macUlRetx", "average"),
        "pdschDlAvg": stat("pdsch5g", "average"),
        "prbEfficiency": _deep_num(kpis.get("prbEfficiency")),
        "prbUtilPct": _deep_num(kpis.get("prbUtilPct")),
        "availableBandwidthPrbs": stat("availableBandwidthPrbs", "average"),
        "resourceAllocationIndex": _deep_num(kpis.get("resourceAllocationIndex")),
        "prbsAvg": stat("prbs", "average"),
        "scheduled5gAvg": stat("scheduled5g", "average"),
        # Throughput by RAT (intrinsic LTE vs 5G columns).
        "macDlLte": stat("macLte", "average"),
        "macDl5g": stat("mac5g", "average"),
        "macDlTotal": stat("totalMacDl", "average"),
        "pdschDlLteMbps": stat("pdschDlLte", "average"),
        "pdschDl5gMbps": stat("pdsch5g", "average"),
        "nrThroughputContrib": _deep_num(kpis.get("nrThroughputContribPct")),
        "lteThroughputContrib": _deep_num(kpis.get("lteThroughputContribPct")),
        # Diagnosis-confidence inputs (sample sizes behind the conclusions).
        "testCount": kpis.get("testCount"),
        "nrRfSamples": (kpis.get("rfNrLte") or {}).get("rsrpNrSamples"),
        "lteRfSamples": (kpis.get("rfNrLte") or {}).get("rsrpLteSamples"),
        # Capacity / reliability / CA / UL additions from the richer export.
        "dlPrbUtilPct": stat("dlPrbUtilPct", "average"),
        "prbsAvgDl": stat("prbsAvgDlAll", "average"),
        "schBitratePerPrb": stat("schBitratePerPrb", "average"),
        "pdschBlerLte": stat("pdschBlerLte", "average"),
        "macDlResidualBler": stat("macDlResidualBler", "average"),
        "pdcchBlerEst": stat("pdcchBlerEst", "average"),
        "ulRetxLte": stat("macUlRetxLte", "average"),
        "caTotalBwMhz": stat("caTotalBwMhz", "average"),
        "primaryBwMhz": stat("primaryBwMhz", "average"),
        "sumSecondaryBwMhz": stat("sumSecondaryBwMhz", "average"),
        "txPower": stat("txPower", "average"),
        "puschTxPower": stat("puschTxPower", "average"),
        "wbCqi0": stat("wbCqi0", "median"),
        "wbCqi1": stat("wbCqi1", "median"),
        "hoUplaneInterruptionMs": stat("hoUplaneInterruptionMs", "average"),
        # EN-DC NR secondary-node (SgNB) mobility — reconstructed from NR-SCG presence.
        "sgnbAdditions": (kpis.get("endcSecondaryNode") or {}).get("additions"),
        "sgnbAdditionSuccess": (kpis.get("endcSecondaryNode") or {}).get("additionSuccess"),
        "sgnbAdditionFailure": (kpis.get("endcSecondaryNode") or {}).get("additionFailure"),
        "sgnbAdditionSuccessRate": (kpis.get("endcSecondaryNode") or {}).get("additionSuccessRate"),
        "sgnbRemovals": (kpis.get("endcSecondaryNode") or {}).get("removals"),
        "sgnbRemovalSuccess": (kpis.get("endcSecondaryNode") or {}).get("removalSuccess"),
        "sgnbRemovalFailure": (kpis.get("endcSecondaryNode") or {}).get("removalFailure"),
        "sgnbRemovalSuccessRate": (kpis.get("endcSecondaryNode") or {}).get("removalSuccessRate"),
        # PPP-layer throughput + comparison vs application throughput.
        "pppRateAvg": stat("pppRate", "average"),
        "pppRatePeak": stat("pppRate", "max"),
        "pppVsAppPct": (
            round(stat("pppRate", "average") / stat("dl", "average") * 100.0, 1)
            if stat("pppRate", "average") is not None and stat("dl", "average") not in (None, 0)
            else None
        ),
        "lteAnchorSinr": _deep_num(kpis.get("lteAnchorSinr")),
        "servingCellDistanceM": _deep_num(kpis.get("servingCellDistanceM")),
        "endcSetupSuccessRate": _deep_num(kpis.get("endcSetupSuccessRate")),
        "endcDropRate": _deep_num(kpis.get("endcDropRate")),
        "tcpHandshake": stat("tcpHandshake", "median"),
        "lostPacket": stat("lostPacket", "average"),
        "dlCompletion": _deep_num(dl_tr.get("avgCompletionPct")),
        "dlSuccess": _deep_num(dl_tr.get("successRate")),
        # Extended metrics for the Detailed Analysis (full-KPI surface).
        "rsrpMedian": stat("rsrp", "median"),
        "rsrqMedian": stat("rsrq", "median"),
        "sinrP10": stat("sinr", "p10"),
        "cqiMedian": stat("cqi", "median"),
        "blerAbove20": _deep_num(kpis.get("blerAbove20Share")),
        "totalMacDlAvg": stat("totalMacDl", "average"),
        "mac5gAvg": stat("mac5g", "average"),
        "transportRatio": _deep_num(kpis.get("transportRatio")),
        "scheduledRankMedian": stat("scheduledRank", "median"),
        "scheduledEfficiency": _deep_num(kpis.get("scheduledEfficiency")),
        "pdschBitPerHz": stat("pdschBitPerHz", "median"),
        "pdschMaxBitPerHz": stat("pdschMaxBitPerHz", "median"),
        "mcsCw0Median": stat("pdschMcsCw0", "median"),
        "mcsCw1Median": stat("pdschMcsCw1", "median"),
        "pdschSlotPct": stat("pdschSlotPct", "average"),
        "nrBandShares": kpis.get("nrBandShares") or {},
        "rrcStateShares": kpis.get("rrcStateShares") or [],
        "servingTechnologyShares": kpis.get("servingTechnologyShares") or [],
    }


def _deep_fmt(value, ndigits=1):
    n = _deep_num(value)
    if n is None:
        return ""
    if abs(n - round(n)) < 1e-9:
        return str(int(round(n)))
    return str(round(n, ndigits))


def _deep_delta(iam, other, unit):
    """Formatted delta IAM vs one competitor. unit in {'%','dB','pp',''}."""
    iam_n, other_n = _deep_num(iam), _deep_num(other)
    if iam_n is None or other_n is None:
        return ""
    if unit == "%":
        if other_n == 0:
            return ""
        d = (iam_n - other_n) / abs(other_n) * 100.0
        return f"{'+' if d >= 0 else ''}{round(d, 1)}%"
    d = iam_n - other_n
    sign = "+" if d >= 0 else ""
    if unit == "dB":
        return f"{sign}{round(d, 1)} dB"
    if unit == "pp":
        return f"{sign}{round(d, 1)} pp"
    return f"{sign}{_deep_fmt(d)}"


# KPI Benchmark table spec: (label, key, unit, higher_is_better, domain)
_DEEP_KPI_TABLE = [
    ("DL Throughput (Mbps)",      "dlThroughput",  "%",  True,  "Throughput"),
    ("PDSCH DL Avg (Mbps)",       "pdschDlAvg",    "%",  True,  "Throughput"),
    ("MAC DL Total (Mbps)",       "macDlTotal",    "%",  True,  "Throughput by RAT"),
    ("MAC DL NR/5G (Mbps)",       "macDl5g",       "%",  True,  "Throughput by RAT"),
    ("MAC DL LTE (Mbps)",         "macDlLte",      "%",  True,  "Throughput by RAT"),
    ("NR throughput contrib %",   "nrThroughputContrib", "pp", True,  "Throughput by RAT"),
    ("LTE throughput contrib %",  "lteThroughputContrib","pp", True,  "Throughput by RAT"),
    ("PDSCH DL NR/5G (Mbps)",     "pdschDl5gMbps", "%",  True,  "Throughput by RAT"),
    ("PDSCH DL LTE (Mbps)",       "pdschDlLteMbps","%",  True,  "Throughput by RAT"),
    ("LTE RSRP Median (dBm)",     "rsrpLte",       "dB", True,  "RF Quality (LTE)"),
    ("LTE SINR Median (dB)",      "sinrLte",       "dB", True,  "RF Quality (LTE)"),
    ("LTE RSRQ Median (dB)",      "rsrqLte",       "dB", True,  "RF Quality (LTE)"),
    ("NR RSRP Median (dBm)",      "rsrpNr",        "dB", True,  "NR RF Quality"),
    ("NR SINR Median (dB)",       "sinrNr",        "dB", True,  "NR RF Quality"),
    ("NR RSRQ Median (dB)",       "rsrqNr",        "dB", True,  "NR RF Quality"),
    ("NR WB CQI Median",          "cqi",           "",   True,  "NR RF Quality"),
    ("RSRP Total (RAT-avg dBm)",  "rsrpTotal",     "dB", True,  "RF Quality (Total)"),
    ("SINR Total (RAT-avg dB)",   "sinrTotal",     "dB", True,  "RF Quality (Total)"),
    ("5G Presence %",             "fivegPresence", "pp", True,  "5G Presence"),
    ("4G Only %",                 "fourgOnly",     "pp", False, "5G Presence"),
    ("NR n78 share %",            "n78",           "pp", True,  "5G Presence"),
    ("NR n28 share %",            "n28",           "pp", True,  "5G Presence"),
    ("NR PDSCH MCS Median",       "mcs",           "",   True,  "NR Modulation / MCS"),
    ("NR 256QAM share %",         "qam256",        "pp", True,  "NR Modulation / MCS"),
    ("NR 64QAM share %",          "qam64",         "pp", True,  "NR Modulation / MCS"),
    ("NR 16QAM share %",          "qam16",         "pp", False, "NR Modulation / MCS"),
    ("NR QPSK share %",           "qpsk",          "pp", False, "NR Modulation / MCS"),
    ("NR PDSCH Rank Median",      "medianRank",    "",   True,  "NR MIMO"),
    ("NR RI1 share %",            "ri1",           "pp", False, "NR MIMO"),
    ("NR RI2 share %",            "ri2",           "pp", True,  "NR MIMO"),
    ("NR RI>=3 share %",          "riGe3",         "pp", True,  "NR MIMO"),
    ("Avg # SCells (LTE CA)",     "scellsAvg",     "",   True,  "Carrier Aggregation"),
    ("Max # SCells (LTE CA)",     "scellsMax",     "",   True,  "Carrier Aggregation"),
    ("SCells >0 share %",         "scellsActive",  "pp", True,  "Carrier Aggregation"),
    ("LTE CA active share %",     "caActive",      "pp", True,  "Carrier Aggregation"),
    ("CA Total BW (MHz)",         "caTotalBwMhz",     "",   True,  "Carrier Aggregation"),
    ("Primary BW (MHz)",          "primaryBwMhz",     "",   True,  "Carrier Aggregation"),
    ("Secondary BW sum (MHz)",    "sumSecondaryBwMhz","",   True,  "Carrier Aggregation"),
    ("NR MAC DL BLER Avg %",      "blerAvg",       "pp", False, "NR BLER / Retx"),
    ("NR MAC DL BLER P90 %",      "blerP90",       "pp", False, "NR BLER / Retx"),
    ("NR BLER >10% share %",      "blerAbove10",   "pp", False, "NR BLER / Retx"),
    ("NR UL Retx Avg %",          "ulRetx",        "pp", False, "NR BLER / Retx"),
    ("LTE PDSCH BLER Avg %",      "pdschBlerLte",     "pp", False, "LTE BLER / Retx"),
    ("MAC DL Residual BLER %",    "macDlResidualBler","pp", False, "LTE BLER / Retx"),
    ("PDCCH BLER est. %",         "pdcchBlerEst",     "pp", False, "LTE BLER / Retx"),
    ("LTE UL Retx Avg %",         "ulRetxLte",        "pp", False, "LTE BLER / Retx"),
    ("TCP Handshake Median (ms)", "tcpHandshake",          "",   False, "Accessibility"),
    ("DL completion %",           "dlCompletion",          "pp", True,  "Accessibility"),
    ("DL success %",              "dlSuccess",             "pp", True,  "Accessibility"),
    ("Available NR PRBs",         "availableBandwidthPrbs","",   True,  "Scheduler & PRB Allocation"),
    ("Allocated PRBs (avg)",      "prbsAvg",               "",   True,  "Scheduler & PRB Allocation"),
    ("Allocation ratio %",        "resourceAllocationIndex","pp", True, "Scheduler & PRB Allocation"),
    ("PDSCH slot %",              "pdschSlotPct",          "pp", True,  "Scheduler & PRB Allocation"),
    ("PRB efficiency (Mbps/PRB)", "prbEfficiency",         "",   True,  "Scheduler & PRB Allocation"),
    ("Scheduled 5G avg (Mbps)",   "scheduled5gAvg",        "%",  True,  "Scheduler & PRB Allocation"),
    ("DL PRB Utilization %",      "dlPrbUtilPct",          "pp", True,  "Scheduler & PRB Allocation"),
    ("Avg DL PRBs",               "prbsAvgDl",             "",   True,  "Scheduler & PRB Allocation"),
    ("Sched. bitrate/PRB",        "schBitratePerPrb",      "",   True,  "Scheduler & PRB Allocation"),
    ("WB CQI CW0 Median",         "wbCqi0",                "",   True,  "CQI (per codeword)"),
    ("WB CQI CW1 Median",         "wbCqi1",                "",   True,  "CQI (per codeword)"),
    ("UL TX Power Avg (dBm)",     "txPower",               "",   False, "UL Power"),
    ("PUSCH TX Power Avg (dBm)",  "puschTxPower",          "",   False, "UL Power"),
    ("HO U-plane Interruption (ms)", "hoUplaneInterruptionMs", "", False, "Mobility"),
    ("SgNB Additions",            "sgnbAdditions",          "",   True,  "EN-DC Secondary Node (SgNB)"),
    ("SgNB Add Success",          "sgnbAdditionSuccess",    "",   True,  "EN-DC Secondary Node (SgNB)"),
    ("SgNB Add Failure",          "sgnbAdditionFailure",    "",   False, "EN-DC Secondary Node (SgNB)"),
    ("SgNB Add Success %",        "sgnbAdditionSuccessRate","pp", True,  "EN-DC Secondary Node (SgNB)"),
    ("SgNB Removals",             "sgnbRemovals",           "",   True,  "EN-DC Secondary Node (SgNB)"),
    ("SgNB Removal Success",      "sgnbRemovalSuccess",     "",   True,  "EN-DC Secondary Node (SgNB)"),
    ("SgNB Removal Failure",      "sgnbRemovalFailure",     "",   False, "EN-DC Secondary Node (SgNB)"),
    ("SgNB Removal Success %",    "sgnbRemovalSuccessRate", "pp", True,  "EN-DC Secondary Node (SgNB)"),
    ("PPP DL Avg (Mbps)",         "pppRateAvg",             "%",  True,  "PPP Throughput"),
    ("PPP DL Peak (Mbps)",        "pppRatePeak",            "%",  True,  "PPP Throughput"),
]


_SCHED_INTERP = {
    "Available NR PRBs": (
        "IAM NR resource pool is comparable to competitors.",
        "IAM has fewer available NR PRBs — indicates a smaller NR BWP or missing n78 contribution.",
    ),
    "Allocated PRBs (avg)": (
        "IAM receives comparable PRB allocation.",
        "IAM receives fewer allocated PRBs on average — investigate scheduler grants and cell load.",
    ),
    "Allocation ratio %": (
        "IAM allocation ratio is comparable — scheduler is not under-allocating.",
        "IAM allocation ratio is lower — scheduler grants a smaller share of available PRBs to IAM UEs.",
    ),
    "PDSCH slot %": (
        "IAM PDSCH time-domain utilization is comparable.",
        "IAM uses fewer PDSCH slots — time-domain under-utilization; check scheduling activity and load.",
    ),
    "PRB efficiency (Mbps/PRB)": (
        "IAM PRB efficiency is good — each allocated PRB delivers comparable throughput.",
        "IAM delivers less throughput per allocated PRB — check modulation, MCS, rank, BLER, and link adaptation.",
    ),
    "Scheduled 5G avg (Mbps)": (
        "IAM scheduled 5G throughput is comparable.",
        "IAM scheduled 5G throughput is lower — the NR PDSCH layer delivers less capacity.",
    ),
}


def _deep_kpi_interpretation(label, iam, orange, inwi, higher):
    """Short rule-based interpretation per KPI row (best-effort, never raises)."""
    iam_n = _deep_num(iam)
    comps = {"Orange": _deep_num(orange), "INWI": _deep_num(inwi)}
    comp_vals = [v for v in comps.values() if v is not None]
    if iam_n is None:
        is_nr_metric = label.startswith("NR ") or label.startswith("Available NR") or label.startswith("PDSCH slot") or label.startswith("Allocation ratio") or label.startswith("PRB eff") or label.startswith("Scheduled 5G")
        if is_nr_metric:
            return "Métrique NR — non disponible (IAM en LTE only pour ce DT)."
        return "IAM value not available for this KPI in the current scope."
    if not comp_vals:
        return "No competitor reference available for this KPI."
    best = max(comp_vals) if higher else min(comp_vals)
    best_op = max(comps, key=lambda k: (comps[k] if comps[k] is not None else -1e9)) if higher \
        else min(comps, key=lambda k: (comps[k] if comps[k] is not None else 1e9))
    # Scheduler-specific interpretations
    if label in _SCHED_INTERP:
        ok_text, bad_text = _SCHED_INTERP[label]
        gap = (iam_n - best) / abs(best) * 100.0 if best not in (None, 0) else 0
        return ok_text if gap >= -10 else bad_text
    if higher:
        if iam_n >= best:
            return f"IAM leads or matches competitors on this KPI."
        worst = min(comp_vals)
        mid = " but above the weaker competitor" if iam_n > worst else ""
        return f"IAM trails {best_op}{mid}; improving this KPI supports higher throughput/quality."
    else:
        if iam_n <= best:
            return f"IAM is at or better than the best competitor (lower is better)."
        return f"IAM is worse than {best_op} on this KPI (lower is better); investigate and reduce it."


def _deep_kpi_severity(iam_v, or_v, in_v, higher: bool) -> str:
    """Severity of the IAM gap vs best competitor."""
    iam_n = _deep_num(iam_v)
    comps = [_deep_num(v) for v in (or_v, in_v) if v is not None]
    comps = [c for c in comps if c is not None]
    if iam_n is None or not comps:
        return "—"
    best = max(comps) if higher else min(comps)
    if best == 0:
        return "—"
    gap_pct = (iam_n - best) / abs(best) * 100.0  # negative = IAM trails (higher) or is better (lower)
    if higher:
        if gap_pct >= -3:
            return "—"
        if gap_pct >= -10:
            return "Low"
        if gap_pct >= -20:
            return "Medium"
        if gap_pct >= -35:
            return "High"
        return "Critical"
    else:
        if gap_pct <= 3:
            return "—"
        if gap_pct <= 15:
            return "Low"
        if gap_pct <= 30:
            return "Medium"
        if gap_pct <= 50:
            return "High"
        return "Critical"


def _benchmark_deep_kpi_rows(iam, orange, inwi):
    rows = []
    for entry in _DEEP_KPI_TABLE:
        label, key, unit, higher = entry[0], entry[1], entry[2], entry[3]
        domain = entry[4] if len(entry) > 4 else "—"
        iam_v, or_v, in_v = iam.get(key), orange.get(key) if orange else None, inwi.get(key) if inwi else None
        rows.append({
            "domain": domain,
            "kpi": label,
            "iam": _deep_num(iam_v),
            "orange": _deep_num(or_v),
            "inwi": _deep_num(in_v),
            "vsOrange": _deep_delta(iam_v, or_v, unit),
            "vsInwi": _deep_delta(iam_v, in_v, unit),
            "severity": _deep_kpi_severity(iam_v, or_v, in_v, higher),
            "interpretation": _deep_kpi_interpretation(label, iam_v, or_v, in_v, higher),
        })
    return rows


def _deep_make_finding(domain, kpi, iam_value, benchmark_value, severity, finding, root_cause,
                       actions, confidence="Direct KPI evidence", gap_pct=None):
    meta = _DEEP_DOMAIN_META.get(domain, {})
    score = (
        _DEEP_SEVERITY_WEIGHT.get(severity, 1)
        + _DEEP_IMPACT_WEIGHT.get(meta.get("impact", "QoS"), 3)
        + _DEEP_CONFIDENCE_WEIGHT.get(confidence, 1)
    )
    return {
        "domain": domain,
        "kpi": kpi,
        "iamValue": _deep_num(iam_value),
        "benchmarkValue": _deep_num(benchmark_value),
        "gapPercent": round(gap_pct * 100.0, 1) if gap_pct is not None else None,
        "severity": severity,
        "finding": finding,
        "rootCause": root_cause,
        "recommendedActions": actions,
        "owner": meta.get("owner", "Optimization"),
        "expectedImpact": meta.get("expectedImpact", ""),
        "validationTarget": meta.get("validationTarget", ""),
        "confidence": _deep_confidence_label(confidence),
        "confidenceEvidence": confidence,
        "benchmarkRelevance": _deep_benchmark_relevance(domain, gap_pct),
        "priorityScore": score,
        "priority": _deep_priority_from_score(score),
    }


def _deep_merge_same_domain(findings):
    """Collapse multiple findings that share a domain into one professional finding: the
    highest-severity/priority finding leads, the rest become supporting evidence, and the
    recommended actions are unioned. Keeps the report to one narrative per domain."""
    items = list(findings or [])
    severity_rank = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}
    by_domain: dict = {}
    order: list = []
    for it in items:
        d = str(it.get("domain") or "")
        if d not in by_domain:
            by_domain[d] = []
            order.append(d)
        by_domain[d].append(it)
    out = []
    for d in order:
        grp = by_domain[d]
        if len(grp) == 1:
            out.append(grp[0])
            continue
        lead = max(grp, key=lambda it: (severity_rank.get(str(it.get("severity") or ""), 0), it.get("priorityScore") or 0))
        merged = dict(lead)
        support = [f"{it.get('kpi')}: {it.get('finding')}" for it in grp if it is not lead and it.get("finding")]
        merged["supportingEvidence"] = (merged.get("supportingEvidence") or []) + support
        if support:
            merged["finding"] = f"{lead.get('finding')} Also: " + " ".join(support)
        actions, seen = [], set()
        for it in grp:
            for a in it.get("recommendedActions") or []:
                t = str(a or "").strip()
                if t and t not in seen:
                    seen.add(t)
                    actions.append(t)
        merged["recommendedActions"] = actions
        merged["severity"] = max(grp, key=lambda it: severity_rank.get(str(it.get("severity") or ""), 0)).get("severity")
        merged["priorityScore"] = max(it.get("priorityScore") or 0 for it in grp)
        merged["priority"] = _deep_priority_from_score(merged["priorityScore"])
        out.append(merged)
    return out


def _deep_consolidate_findings(findings):
    items = list(findings or [])
    radio_domains = {"Radio quality / interference", "SINR / interference", "Modulation profile"}
    radio = [item for item in items if item.get("domain") in radio_domains]
    if len(radio) <= 1:
        return items

    other = [item for item in items if item.get("domain") not in radio_domains]
    severity_rank = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}
    preferred_kpis = {
        "RSRP + SINR": 0,
        "Median RSRQ": 1,
        "Median SINR": 2,
        "Median CQI": 3,
        "256QAM / 16QAM share": 4,
    }
    lead = sorted(
        radio,
        key=lambda item: (
            preferred_kpis.get(str(item.get("kpi") or ""), 99),
            -severity_rank.get(str(item.get("severity") or ""), 0),
            -(item.get("priorityScore") or 0),
        ),
    )[0]

    merged = dict(lead)
    merged["domain"] = "Radio quality / interference"
    merged["severity"] = max(radio, key=lambda item: severity_rank.get(str(item.get("severity") or ""), 0)).get("severity")
    merged["supportingEvidence"] = [
        f"{item.get('kpi')}: {item.get('finding')}"
        for item in radio
        if item is not lead and item.get("finding")
    ]
    if merged["supportingEvidence"]:
        merged["finding"] = (
            f"{lead.get('finding')} Supporting evidence: "
            + " ".join(merged["supportingEvidence"])
        )

    merged_actions = []
    seen_actions = set()
    for item in radio:
        for action in item.get("recommendedActions") or []:
            text = str(action or "").strip()
            if not text or text in seen_actions:
                continue
            seen_actions.add(text)
            merged_actions.append(text)
    merged["recommendedActions"] = merged_actions
    merged["priorityScore"] = max(item.get("priorityScore") or 0 for item in radio)
    merged["priority"] = _deep_priority_from_score(merged["priorityScore"])
    merged["confidence"] = max(radio, key=lambda item: _DEEP_CONFIDENCE_WEIGHT.get(item.get("confidenceEvidence"), 0)).get("confidence")
    merged["confidenceEvidence"] = max(
        radio,
        key=lambda item: _DEEP_CONFIDENCE_WEIGHT.get(item.get("confidenceEvidence"), 0),
    ).get("confidenceEvidence")
    merged["benchmarkRelevance"] = max(
        radio,
        key=lambda item: {"Primary": 3, "Secondary": 2, "Context": 1}.get(str(item.get("benchmarkRelevance") or ""), 0),
    ).get("benchmarkRelevance")
    return other + [merged]


def _benchmark_deep_findings(iam, competitors):
    """Run the rule decision tree (spec §4–§12) on IAM vs best competitor."""
    findings = []
    comp = competitors or []

    def best_of(key, higher=True):
        return _deep_best_competitor([c.get(key) for c in comp], higher_is_better=higher)

    dl = iam.get("dlThroughput")
    best_dl = best_of("dlThroughput", True)
    dl_gap = _deep_gap_fraction(dl, best_dl, True)
    dl_success = iam.get("dlSuccess")
    dl_completion = iam.get("dlCompletion")
    fiveg = iam.get("fivegPresence")
    rsrp = iam.get("rsrp")
    rsrq = iam.get("rsrq")
    sinr = iam.get("sinr")
    serving_distance = iam.get("servingCellDistanceM")
    lte_anchor_sinr = iam.get("lteAnchorSinr")
    prb_util_pct = iam.get("prbUtilPct")
    bandwidth_prbs = iam.get("availableBandwidthPrbs")
    best_bandwidth_prbs = best_of("availableBandwidthPrbs", True)
    endc_setup_success = iam.get("endcSetupSuccessRate")
    endc_drop_rate = iam.get("endcDropRate")
    mimo_ok = (
        (iam.get("medianRank") is not None and iam.get("medianRank") >= 2)
        and (iam.get("ri1") is not None and iam.get("ri1") <= 10)
        and (iam.get("ri2") is not None and iam.get("ri2") >= 80)
    )
    n1_share = iam.get("n1")

    # Coverage / dominance — RAT-aware: localize the weak leg using the NR/LTE RSRP split so a
    # weak NR layer over a healthy LTE anchor is not reported as a generic "coverage hole".
    rsrp_nr_v = iam.get("rsrpNr")
    rsrp_lte_v = iam.get("rsrpLte")
    nr_cov_weak = rsrp_nr_v is not None and rsrp_nr_v < -105
    lte_cov_weak = rsrp_lte_v is not None and rsrp_lte_v < -105
    conflated_weak = rsrp is not None and rsrp < -105
    if conflated_weak or nr_cov_weak or lte_cov_weak:
        poor_quality_too = sinr is not None and sinr < 5
        lte_anchor_ok = rsrp_lte_v is not None and rsrp_lte_v >= -100
        if nr_cov_weak and lte_anchor_ok and not lte_cov_weak:
            # NR leg weak while the LTE anchor is healthy → NR coverage/dominance issue.
            kpi_label = "NR RSRP (LTE anchor OK)"
            kpi_value, bench_value = rsrp_nr_v, best_of("rsrpNr", True)
            severity = "Critical" if (rsrp_nr_v < -112 or poor_quality_too) else "High"
            finding = (
                f"NR layer is coverage-limited: IAM NR RSRP is {_deep_fmt(rsrp_nr_v)} dBm while the LTE anchor "
                f"is adequate (LTE RSRP {_deep_fmt(rsrp_lte_v)} dBm)"
                + (f"; NR SINR is {_deep_fmt(sinr)} dB" if poor_quality_too else "") + "."
            )
            root_cause = (
                "The LTE anchor is healthy, so the throughput limitation is on the NR layer — weak n78 "
                "dominance/coverage rather than a site-level coverage hole."
            )
            actions = [
                "Audit n78 dominance/footprint along the route (SSB RSRP/SINR, beam coverage).",
                "Check NR overshoot and PCI dominance; densify or re-point n78 where weak.",
                "Verify NR is co-sited and antenna-aligned with the LTE anchor.",
                "Review EN-DC addition thresholds (B1/A4) so NR is added/retained earlier.",
            ]
        else:
            # Both legs weak (or only one RAT present) → genuine coverage/dominance limitation.
            kpi_label = "Median RSRP"
            kpi_value, bench_value = rsrp, best_of("rsrp", True)
            severity = "Critical" if ((rsrp is not None and rsrp < -110) or poor_quality_too) else "High"
            leg_detail = (
                f" (NR {_deep_fmt(rsrp_nr_v)} / LTE {_deep_fmt(rsrp_lte_v)} dBm)"
                if rsrp_nr_v is not None and rsrp_lte_v is not None else ""
            )
            finding = (
                f"Coverage and quality limitation: IAM median RSRP is {_deep_fmt(rsrp)} dBm{leg_detail} and SINR is {_deep_fmt(sinr)} dB."
                if poor_quality_too
                else f"Weak serving coverage: IAM median RSRP is {_deep_fmt(rsrp)} dBm{leg_detail}."
            )
            root_cause = (
                "Serving coverage is weak and radio quality is also poor, indicating a combined coverage/dominance limitation."
                if poor_quality_too
                else "Serving coverage is weak and reduces throughput potential."
            )
            actions = [
                "Check nearest IAM serving site distance.",
                "Check azimuth and tilt.",
                "Check overshooting from far cells.",
                "Check missing nearby site / sector.",
                "Verify antenna height and mechanical tilt.",
                "Check indoor/outdoor route classification.",
            ]
            if serving_distance is not None and serving_distance > 1000 and rsrp is not None and rsrp < -100:
                finding += f" Serving cell distance is {_deep_fmt(serving_distance, 0)} m, suggesting overshooting or a missing dominant sector."
                root_cause = "Possible overshooting or missing dominant cell."
        findings.append(_deep_make_finding(
            "Coverage / dominance",
            kpi_label,
            kpi_value,
            bench_value,
            severity,
            finding,
            root_cause,
            actions,
            confidence="Direct KPI evidence",
        ))

    if serving_distance is not None and serving_distance > 1000 and rsrp is not None and rsrp < -100 and not any(
        item.get("domain") == "Coverage / dominance" and "overshooting or a missing dominant sector" in str(item.get("finding") or "")
        for item in findings
    ):
        findings.append(_deep_make_finding(
            "Coverage / dominance",
            "Serving distance + RSRP",
            rsrp,
            best_of("rsrp", True),
            "High",
            f"Possible overshooting or a missing dominant sector: serving-cell distance is {_deep_fmt(serving_distance, 0)} m while IAM median RSRP is only {_deep_fmt(rsrp)} dBm.",
            "The route appears served by a far cell instead of a nearby dominant sector.",
            [
                "Check nearest IAM serving site distance.",
                "Check azimuth and tilt.",
                "Check overshooting from far cells.",
                "Check missing nearby site / sector.",
                "Verify antenna height and mechanical tilt.",
                "Check indoor/outdoor route classification.",
            ],
            confidence="Benchmark evidence",
        ))

    if rsrq is not None and rsrq < -14:
        findings.append(_deep_make_finding(
            "Radio quality / interference",
            "Median RSRQ",
            rsrq,
            best_of("rsrq", True),
            "Critical" if rsrq < -17 else "High",
            f"{'Very poor' if rsrq < -17 else 'Poor'} RSRQ: IAM median RSRQ is {_deep_fmt(rsrq)} dB, consistent with interference, overlap or high load.",
            "Poor RF quality, interference or load is degrading channel quality.",
            [
                "Check DL interference.",
                "Check high PRB utilization.",
                "Check overlapping cells.",
                "Check pilot pollution.",
                "Check PCI planning.",
                "Check load balancing.",
            ],
            confidence="Direct KPI evidence",
        ))

    if rsrp is not None and rsrp > -95 and sinr is not None and sinr < 5:
        findings.append(_deep_make_finding(
            "Radio quality / interference",
            "RSRP + SINR",
            sinr,
            best_of("sinr", True),
            "High",
            f"Good coverage but poor quality: IAM median RSRP is {_deep_fmt(rsrp)} dBm while median SINR is only {_deep_fmt(sinr)} dB.",
            "Interference / pollution / wrong dominance is limiting throughput despite acceptable signal level.",
            [
                "Optimize azimuth and tilt.",
                "Check overshooting cells.",
                "Check PCI collision/confusion.",
                "Check inter-frequency interference.",
                "Check neighboring relation.",
                "Review power balancing between sectors.",
            ],
            confidence="Direct KPI evidence",
        ))

    if sinr is not None and sinr > 10 and dl_gap is not None and dl_gap <= -0.30:
        findings.append(_deep_make_finding(
            "Capacity / configuration",
            "DL throughput vs radio quality",
            dl,
            best_dl,
            "High",
            f"Throughput is low despite good radio quality: IAM SINR is {_deep_fmt(sinr)} dB but DL throughput is only {_deep_fmt(dl)} Mbps versus {_deep_fmt(best_dl)} Mbps.",
            "Capacity, scheduler, CA, MIMO, bandwidth or transport limitations are more likely than RF coverage.",
            [
                "Check PRB allocation.",
                "Check bandwidth configuration.",
                "Check CA activation.",
                "Check MIMO rank.",
                "Check scheduler profile.",
                "Check QoS limitation.",
                "Check TCP/server limitation.",
                "Check backhaul congestion.",
            ],
            confidence="Benchmark evidence",
            gap_pct=dl_gap,
        ))

    if prb_util_pct is not None and prb_util_pct > 80 and dl_gap is not None and dl_gap <= -0.20:
        findings.append(_deep_make_finding(
            "Load / congestion",
            "DL PRB utilization",
            prb_util_pct,
            80.0,
            "Critical" if prb_util_pct > 90 else "High",
            f"{'Severe DL congestion' if prb_util_pct > 90 else 'DL congestion'}: IAM DL PRB utilization is {_deep_fmt(prb_util_pct)}% while throughput remains below benchmark.",
            "Cell load is likely starving user throughput resources during the test window.",
            [
                "Add capacity carrier.",
                "Activate CA.",
                "Activate/load-balance 5G n78.",
                "Optimize load balancing.",
                "Review scheduler fairness.",
                "Add sector split or new site.",
                "Check traffic hotspot.",
            ],
            confidence="Direct KPI evidence",
        ))
    elif prb_util_pct is not None and prb_util_pct < 50 and dl_gap is not None and dl_gap <= -0.30:
        findings.append(_deep_make_finding(
            "Capacity / configuration",
            "DL PRB utilization",
            prb_util_pct,
            50.0,
            "High",
            f"Low throughput is not caused by congestion: IAM DL PRB utilization is only {_deep_fmt(prb_util_pct)}%.",
            "Throughput limitation is more likely due to radio quality, MIMO, CA, scheduler or transport than pure congestion.",
            [
                "Check radio quality, MIMO and CA activation together.",
                "Check scheduler policy and QoS limitation.",
                "Check transport/backhaul limitation.",
                "Compare PRB allocation and bandwidth configuration.",
            ],
            confidence="Direct KPI evidence",
        ))

    bw_gap = _deep_gap_fraction(bandwidth_prbs, best_bandwidth_prbs, True)
    if bw_gap is not None and bw_gap <= -0.20:
        ca_present = (iam.get("scellsAvg") or 0) >= 1 or (iam.get("caActive") or 0) >= 30
        n1_bandwidth_profile = (
            n1_share is not None
            and n1_share >= 50
            and bandwidth_prbs is not None
            and abs(float(bandwidth_prbs) - 79.0) <= 1.0
            and ca_present
        )
        findings.append(_deep_make_finding(
            "Bandwidth / spectrum",
            "Available bandwidth (PRBs)",
            bandwidth_prbs,
            best_bandwidth_prbs,
            "High",
            (
                f"IAM NR layer is mainly n1 ({_deep_fmt(n1_share)}% NR share, {_deep_fmt(bandwidth_prbs)} PRBs) and LTE CA is active, but the best competitor reaches {_deep_fmt(best_bandwidth_prbs)} PRBs with n78 capacity."
                if n1_bandwidth_profile
                else f"IAM has lower available serving bandwidth ({_deep_fmt(bandwidth_prbs)} PRBs) than the best competitor ({_deep_fmt(best_bandwidth_prbs)} PRBs)."
            ),
            (
                "Capacity gap is mainly driven by the missing n78 high-capacity layer, not by absence of LTE CA."
                if n1_bandwidth_profile
                else "Spectrum bandwidth disadvantage is likely contributing to the throughput gap."
            ),
            [
                "Check available LTE/NR bandwidth.",
                "Check NR n78 deployment.",
                "Check CA combinations.",
                "Check whether all carriers are active.",
                "Verify spectrum refarming opportunity.",
            ],
            confidence="Benchmark evidence",
            gap_pct=bw_gap,
        ))

    if fiveg is not None and fiveg > 40 and lte_anchor_sinr is not None and lte_anchor_sinr < 5:
        findings.append(_deep_make_finding(
            "LTE anchor / NSA dependency",
            "LTE-anchor SINR",
            lte_anchor_sinr,
            5.0,
            "High",
            f"Weak LTE anchor limiting NSA 5G performance: 5G presence is {_deep_fmt(fiveg)}% but LTE-anchor SINR is only {_deep_fmt(lte_anchor_sinr)} dB.",
            "The LTE anchor quality is too weak to support stable EN-DC performance.",
            [
                "Optimize LTE anchor cell.",
                "Check LTE anchor band selection.",
                "Review EN-DC anchor configuration.",
                "Check B1/B3/B7 anchor priorities.",
                "Verify NR addition success rate.",
            ],
            confidence="Direct KPI evidence",
        ))

    if endc_setup_success is not None and endc_setup_success < 95:
        findings.append(_deep_make_finding(
            "EN-DC stability",
            "EN-DC setup success",
            endc_setup_success,
            95.0,
            "High",
            f"EN-DC setup failure risk: setup success rate is only {_deep_fmt(endc_setup_success)}%.",
            "NSA addition success is below target.",
            [
                "Check NR neighbor configuration.",
                "Check X2/Xn interface.",
                "Check NSA addition thresholds.",
                "Check SSB RSRP/SINR.",
                "Check LTE anchor coverage.",
                "Check gNB/eNB alarms.",
            ],
            confidence="Direct KPI evidence",
        ))

    if endc_drop_rate is not None and endc_drop_rate > 2:
        findings.append(_deep_make_finding(
            "EN-DC stability",
            "EN-DC drop rate",
            endc_drop_rate,
            2.0,
            "High",
            f"EN-DC instability: drop rate reaches {_deep_fmt(endc_drop_rate)}%.",
            "NSA retention is unstable and likely causing 5G interruptions or resets.",
            [
                "Check NR neighbor configuration.",
                "Check X2/Xn interface.",
                "Check NSA addition thresholds.",
                "Check SSB RSRP/SINR.",
                "Check LTE anchor coverage.",
                "Check gNB/eNB alarms.",
            ],
            confidence="Direct KPI evidence",
        ))

    # §4.3 / §12 — 5G capacity layer (n78 missing while competitor uses n78)
    n78 = iam.get("n78")
    comp_n78 = best_of("n78", True)
    if (n78 is not None and (n78 < 30) and ((comp_n78 or 0) >= 50)) or \
       (fiveg is not None and fiveg >= 40 and (n78 is not None and n78 < 30)):
        low_band_text = (
            f"mainly served by NR n1 ({_deep_fmt(n1_share)}% of NR samples)"
            if n1_share is not None and n1_share >= max((iam.get("n28") or 0), 1)
            else (
                f"mainly served by NR n28 ({_deep_fmt(iam.get('n28'))}% of NR samples)"
                if iam.get("n28") is not None
                else "served by low-band NR"
            )
        )
        findings.append(_deep_make_finding(
            "5G capacity layer", "NR n78 share", n78, comp_n78, "Critical",
            f"IAM 5G presence is {_deep_fmt(fiveg)}% but NR n78 share is only {_deep_fmt(n78)}%; the route is {low_band_text}, while the best competitor reaches {_deep_fmt(comp_n78)}% n78. IAM has 5G availability, but the high-capacity n78 layer is missing in this test.",
            "5G continuity exists, but the route lacks the high-capacity n78 layer needed to compete on throughput.",
            [
                "Audit n78 deployment and eligibility along the tested route.",
                "Check whether n78 is deployed, barred, disabled, weak or not selected on serving sites.",
                "Verify LTE anchor EN-DC relations and NSA addition thresholds (A3/A5/B1/B2) for n78.",
                "Check NR neighbors, SSB footprint/beam coverage and add or retune n78 where available.",
            ],
            confidence="Benchmark evidence", gap_pct=_deep_gap_fraction(n78, comp_n78, True)))

    # §7 — MIMO / Rank Indicator
    rank = iam.get("medianRank")
    ri1 = iam.get("ri1")
    ri2 = iam.get("ri2")
    comp_ri2 = best_of("ri2", True)
    if (rank is not None and rank <= 1) or (ri1 is not None and ri1 > 40) or \
       (ri2 is not None and comp_ri2 is not None and ri2 < 70 and comp_ri2 > 85):
        sev = "Critical" if (rank is not None and rank <= 1) or (ri2 is not None and ri2 < 70 and (comp_ri2 or 0) > 85) else "High"
        findings.append(_deep_make_finding(
            "MIMO / RI", "Median Rank / RI", rank, best_of("medianRank", True), sev,
            f"IAM median rank is {_deep_fmt(rank)}, RI1={_deep_fmt(ri1)}%, RI2={_deep_fmt(ri2)}%, RI>=3={_deep_fmt(iam.get('riGe3'))}%; best competitor RI2={_deep_fmt(comp_ri2)}%. Weak spatial multiplexing.",
            "MIMO not delivering rank-2+; spatial multiplexing limited.",
            [
                "Check antenna ports and 4T4R/2T2R configuration.",
                "Check RRU branch alarms, VSWR and feeder issues.",
                "Check cross-polar imbalance and calibration alarms.",
                "Review azimuth and mechanical/electrical tilt.",
                "Verify RI/PMI/CQI reporting and rank adaptation parameters.",
                "Audit cell coverage overlap and scattering environment.",
            ]))

    # §5.1 — SINR / interference
    sinr = iam.get("sinr")
    if sinr is not None and sinr < 5:
        findings.append(_deep_make_finding(
            "SINR / interference", "Median SINR", sinr, best_of("sinr", True),
            "High" if sinr >= 0 else "Critical",
            f"IAM median SINR is only {_deep_fmt(sinr)} dB; CQI={_deep_fmt(iam.get('cqi'))} and MCS={_deep_fmt(iam.get('mcs'))} keep modulation mainly low.",
            "Poor radio quality limiting spectral efficiency (CQI/MCS/modulation).",
            [
                "Run SINR/RSRP grid on the DT route.",
                "Check overshooting cells and pilot pollution / PCI confusion.",
                "Check missing neighbors and overlapping sectors.",
                "Review azimuth and mechanical/electrical tilt.",
                "Check DL interference; apply tilt/azimuth corrections and neighbor cleanup first.",
            ]))

    # §5.2 — Low CQI
    cqi = iam.get("cqi")
    if cqi is not None and cqi < 9:
        sev = "Critical" if cqi < 7 else "High"
        findings.append(_deep_make_finding(
            "SINR / interference", "Median CQI", cqi, best_of("cqi", True), sev,
            f"IAM median CQI is {_deep_fmt(cqi)}, restricting MCS and modulation.",
            "Low CQI limiting link adaptation and modulation order.",
            [
                "Improve SINR through RF cleanup (interference, overshoot, tilt).",
                "Review CQI reporting, OLLA and CQI aging.",
                "Check PDSCH power allocation and MCS table configuration.",
            ], confidence="Direct KPI evidence"))

    # §6 — Modulation profile
    qam256 = iam.get("qam256")
    qam16 = iam.get("qam16")
    if (qam256 is not None and qam256 == 0) and (qam16 is not None and qam16 > 50):
        findings.append(_deep_make_finding(
            "Modulation profile", "256QAM / 16QAM share", qam256, None, "High",
            f"No 256QAM observed and 16QAM share is {_deep_fmt(qam16)}%; throughput limited by radio quality / scheduler.",
            "Low modulation profile (heavy 16QAM, no 256QAM).",
            [
                "Verify 256QAM activation, UE capability and vendor feature status.",
                "Check CQI distribution, SINR and MCS thresholds.",
                "Improve SINR/CQI before expecting higher 64QAM/256QAM usage.",
            ], confidence="Benchmark evidence"))

    # §10 — Scheduler / PRB efficiency
    prb = iam.get("prbEfficiency")
    comp_prb = best_of("prbEfficiency", True)
    prb_gap = _deep_gap_fraction(prb, comp_prb, True)
    if prb_gap is not None and prb_gap <= -0.20:
        findings.append(_deep_make_finding(
            "Scheduler / PRB efficiency", "PRB efficiency", prb, comp_prb, "High",
            f"IAM PRB efficiency is {_deep_fmt(prb, 3)} versus best competitor {_deep_fmt(comp_prb, 3)}; PDSCH average {_deep_fmt(iam.get('pdschDlAvg'))} Mbps.",
            "Scheduler / spectral efficiency below competitor.",
            [
                "Review scheduler weights and proportional-fair / QoS configuration.",
                "Check PRB utilization versus throughput and CQI aging / OLLA.",
                "Review MCS selection and PDSCH power.",
                "Check CA and MIMO contribution; compare PRB share vs Orange/INWI.",
            ], confidence="Benchmark evidence", gap_pct=prb_gap))

    # §8 — Carrier aggregation
    ca_active = iam.get("caActive")
    scells_avg = iam.get("scellsAvg")
    if (ca_active is not None and ca_active < 40) or (scells_avg is not None and scells_avg < 0.5):
        findings.append(_deep_make_finding(
            "Carrier aggregation", "Avg #SCells / CA active", scells_avg, best_of("scellsAvg", True), "High",
            f"IAM Avg #SCells={_deep_fmt(scells_avg, 2)} and SCells>0 share {_deep_fmt(iam.get('scellsActive'))}%; secondary carriers rarely activated.",
            "Low CA persistence / SCell activation.",
            [
                "Audit CA combinations configured on serving cells and UE capability matching.",
                "Check SCell activation thresholds and timers (A2/A4/A6 events).",
                "Check SCell coverage and RSRP; PCell/SCell load balancing.",
                "Check CA license / feature activation.",
            ]))

    # §9 — BLER / retransmissions
    bler_avg = iam.get("blerAvg")
    bler_above10 = iam.get("blerAbove10")
    bler_p90 = iam.get("blerP90")
    if (bler_avg is not None and bler_avg > 5) or (bler_above10 is not None and bler_above10 > 20) or (bler_p90 is not None and bler_p90 > 15):
        sev = "Critical" if (bler_avg is not None and bler_avg > 10) else "High"
        localized_bler = (bler_avg is not None and bler_avg <= 6) and (bler_p90 is not None and bler_p90 > 20)
        findings.append(_deep_make_finding(
            "BLER / retransmissions", "BLER avg / P90 / >10%", bler_avg, best_of("blerAvg", False), sev,
            (
                f"Localized high-BLER peaks: IAM BLER average={_deep_fmt(bler_avg)}%, P90={_deep_fmt(bler_p90)}%, BLER>10% share={_deep_fmt(bler_above10)}%."
                if localized_bler
                else f"IAM BLER average={_deep_fmt(bler_avg)}%, P90={_deep_fmt(bler_p90)}%, BLER>10% share={_deep_fmt(bler_above10)}%."
            ),
            (
                "Retransmission issues are bursty/localized rather than constant across the whole route; map the exact zones and serving context."
                if localized_bler
                else "High BLER peaks / retransmission zones destabilizing throughput."
            ),
            [
                "Map BLER>10% samples by GPS/time/serving cell.",
                "Check SINR and interference at high-BLER locations.",
                "Review link adaptation / OLLA and MCS aggressiveness.",
                "Check PDSCH power allocation and HARQ retransmissions.",
                "Check antenna/RF hardware issues.",
            ]))

    # §11 / UL quality — UL retransmission
    ul_retx = iam.get("ulRetx")
    comp_ul = best_of("ulRetx", False)
    if ul_retx is not None and ul_retx > 1 and (comp_ul is None or ul_retx > comp_ul):
        findings.append(_deep_make_finding(
            "UL quality", "UL Retx", ul_retx, comp_ul, "Medium" if ul_retx <= 3 else "High",
            f"IAM UL retransmission avg={_deep_fmt(ul_retx)}% versus best competitor {_deep_fmt(comp_ul)}%.",
            "UL quality / interference / power-control issue.",
            [
                "Check UL interference/RTWP on serving LTE anchors.",
                "Check PUSCH power control and UL pathloss.",
                "Check antenna branch health and UL scheduler.",
                "Correlate with TCP handshake and packet loss.",
            ], confidence="Benchmark evidence"))

    # §11 — Transport / TCP
    tcp = iam.get("tcpHandshake")
    comp_tcp = best_of("tcpHandshake", False)
    if (tcp is not None and tcp > 80) or (tcp is not None and comp_tcp is not None and tcp > comp_tcp * 1.20):
        findings.append(_deep_make_finding(
            "Transport / core", "TCP handshake", tcp, comp_tcp, "Medium",
            f"IAM TCP handshake median is {_deep_fmt(tcp)} ms versus best competitor {_deep_fmt(comp_tcp)} ms.",
            "Transport / core path slower than benchmark.",
            [
                "Validate backhaul latency / jitter / loss on involved eNB/gNB.",
                "Check S1/N3 path, DNS/APN path, firewall/NAT.",
                "Retest with a controlled server to exclude test-tool artifact.",
                "Compare same device and same server across operators.",
            ], confidence="Benchmark evidence"))

    # §12 — Mobility / serving sequence (LTE-anchor fallback while 5G is present)
    fourg_only = iam.get("fourgOnly")
    if fiveg is not None and fourg_only is not None and fiveg >= 20 and fourg_only >= 30:
        findings.append(_deep_make_finding(
            "Mobility / serving sequence", "4G-only time", fourg_only, None, "Low",
            f"IAM spends {_deep_fmt(fourg_only)}% of time 4G-only despite {_deep_fmt(fiveg)}% 5G presence; serving alternates between LTE anchor and NR.",
            "Frequent returns to LTE anchor reduce 5G retention and dip throughput.",
            [
                "Review handover and EN-DC addition/release events along the segment.",
                "Check if repeated LTE-anchor returns cause NR drops or SCell reset.",
                "Optimize neighbor priorities and EN-DC addition thresholds.",
            ], confidence="Possible cause"))

    # Always-on §15 governance row.
    findings.append(_deep_make_finding(
        "Retest governance", "Sampling", None, None, "Low",
        "Conclusions should be validated on repeated DTs rather than a single sample.",
        "Single-sample dataset for this scope.",
        [
            "Repeat the same route in busy hour and off-peak with the same UE, SIM plan and server.",
            "Use locked/controlled test scripts and add geo-location bins to localize bad segments.",
        ], confidence="Possible cause"))

    # Headline DL-throughput finding context (§4.1) — annotate accessibility vs capacity.
    if dl_gap is not None and dl_gap <= -0.30:
        access_ok = (dl_success is None or dl_success >= 98) and (dl_completion is None or dl_completion >= 98)
        dl_actions = [
            "Audit n78 availability and EN-DC configuration.",
            "Review scheduler and PRB efficiency.",
            "Validate active download-window serving path versus global 5G presence.",
        ]
        if not mimo_ok:
            dl_actions.insert(1, "Check MIMO rank limitation and improve SINR/CQI through RF optimization.")
        findings.insert(0, _deep_make_finding(
            "5G capacity layer", "DL Throughput", dl, best_dl, _deep_classify_gap(dl_gap, True),
            f"IAM DL throughput is {_deep_fmt(dl)} Mbps versus best competitor {_deep_fmt(best_dl)} Mbps"
            + (" despite 100% DL completion/success" if access_ok else "") + ".",
            "Capacity / radio-efficiency limitation, not an accessibility issue." if access_ok else "Throughput gap with possible accessibility component.",
            dl_actions, confidence="Direct KPI evidence", gap_pct=dl_gap))

    # ── Per-RAT diagnosis (throughput layer balance + RF leg localization) ──
    nr_contrib = iam.get("nrThroughputContrib")
    lte_contrib = iam.get("lteThroughputContrib")
    sinr_nr = iam.get("sinrNr")
    sch_eff = iam.get("schBitratePerPrb")
    if nr_contrib is not None and fiveg is not None and fiveg >= 50 and nr_contrib < 50:
        findings.append(_deep_make_finding(
            "LTE anchor / NSA dependency", "NR throughput contribution", nr_contrib,
            best_of("nrThroughputContrib", True), "High",
            f"Despite {_deep_fmt(fiveg)}% 5G presence, only {_deep_fmt(nr_contrib)}% of IAM MAC DL throughput is carried on NR — {_deep_fmt(lte_contrib)}% rides the LTE anchor.",
            "EN-DC is established but the NR layer is underused — NR scheduling/quality is limited, so the LTE anchor carries most of the payload.",
            [
                "Audit NR scheduler activity and n78 availability during the download window.",
                "Check EN-DC SCG activation latency and NR BWP/PRB allocation.",
                "Verify NR SINR/CQI is high enough for the scheduler to load NR.",
            ], confidence="Direct KPI evidence"))
    elif nr_contrib is not None and nr_contrib >= 85:
        prb_util = iam.get("dlPrbUtilPct")
        nr_quality_bad = sinr_nr is not None and sinr_nr < 5
        eff_low = sch_eff is not None and sch_eff < 0.5
        util_low = prb_util is not None and prb_util < 5
        if nr_quality_bad:
            findings.append(_deep_make_finding(
                "5G capacity layer", "NR throughput contribution", nr_contrib, None, "High",
                f"IAM carries {_deep_fmt(nr_contrib)}% of traffic on NR, but NR radio quality is weak (NR SINR {_deep_fmt(sinr_nr)} dB, spectral efficiency {_deep_fmt(sch_eff)} b/PRB).",
                "Traffic is NR-dominant, so throughput is capped directly by poor NR radio quality — not by the LTE anchor.",
                [
                    "Improve NR SINR via RF optimization (azimuth / tilt / overshoot).",
                    "Raise modulation/MCS by improving NR CQI.",
                    "Confirm n78 dominance on the route.",
                ], confidence="Direct KPI evidence"))
        elif eff_low or util_low:
            findings.append(_deep_make_finding(
                "Scheduler / PRB efficiency", "NR throughput contribution", nr_contrib, None, "High",
                f"IAM carries {_deep_fmt(nr_contrib)}% of traffic on NR with adequate NR SINR ({_deep_fmt(sinr_nr)} dB) but low spectral efficiency ({_deep_fmt(sch_eff)} b/PRB)"
                + (f" and very low DL PRB utilization ({_deep_fmt(prb_util)}%)" if util_low else "") + ".",
                "NR is the dominant layer but is under-scheduled / running at low MCS despite good SINR — throughput is limited by scheduler utilization and spectral efficiency, not by coverage or quality.",
                [
                    "Check NR PRB allocation and scheduler load during the active download window.",
                    "Investigate low MCS/rank despite good SINR (CQI reporting, BWP/numerology config).",
                    "Confirm the test file size/duration is large enough to fully load NR.",
                ], confidence="Direct KPI evidence"))

    # (NR-weak / LTE-anchor-OK coverage is now handled RAT-aware in the Coverage rule above.)
    pdsch_bler_lte = iam.get("pdschBlerLte")
    if pdsch_bler_lte is not None and pdsch_bler_lte > 8:
        findings.append(_deep_make_finding(
            "BLER / retransmissions", "LTE PDSCH BLER", pdsch_bler_lte,
            best_of("pdschBlerLte", False), "Medium",
            f"IAM LTE PDSCH BLER averages {_deep_fmt(pdsch_bler_lte)}% — elevated retransmissions on the LTE leg.",
            "High LTE BLER triggers HARQ retransmissions that cut effective LTE-anchor throughput and slow EN-DC aggregation.",
            [
                "Check LTE interference / overshoot on the anchor band.",
                "Review CQI-to-MCS link adaptation on LTE.",
                "Inspect UL power-limited conditions affecting HARQ.",
            ], confidence="Direct KPI evidence"))

    # Carrier-aggregation bandwidth ceiling (peak-rate cap independent of radio quality).
    ca_bw = iam.get("caTotalBwMhz")
    best_ca_bw = best_of("caTotalBwMhz", True)
    ca_bw_gap = _deep_gap_fraction(ca_bw, best_ca_bw, True)
    if ca_bw_gap is not None and ca_bw_gap <= -0.20:
        findings.append(_deep_make_finding(
            "Bandwidth / spectrum", "CA aggregated bandwidth", ca_bw, best_ca_bw, "High",
            f"IAM aggregates {_deep_fmt(ca_bw)} MHz (primary {_deep_fmt(iam.get('primaryBwMhz'))} MHz + {_deep_fmt(iam.get('sumSecondaryBwMhz'))} MHz secondary) versus {_deep_fmt(best_ca_bw)} MHz for the best competitor.",
            "A narrower aggregated bandwidth caps achievable peak throughput regardless of radio quality.",
            [
                "Add/activate eligible CA SCells (LTE) and NR SCC where configured.",
                "Verify CA combinations match UE capability and serving-cell configuration.",
                "Check spectrum refarming / additional-carrier availability on the route.",
            ], confidence="Benchmark evidence", gap_pct=ca_bw_gap))

    # UL power-limited at cell edge (uses the explicit TX-power + per-RAT UL retx).
    tx_power = iam.get("txPower")
    ul_retx_lte = iam.get("ulRetxLte")
    ul_retx_nr = iam.get("ulRetx")
    if tx_power is not None and tx_power >= 22 and (
        (ul_retx_lte is not None and ul_retx_lte > 5) or (ul_retx_nr is not None and ul_retx_nr > 5)
    ):
        findings.append(_deep_make_finding(
            "UL quality", "UL TX power / retransmissions", tx_power, None, "Medium",
            f"IAM UE transmits near maximum power ({_deep_fmt(tx_power)} dBm) with elevated UL retransmissions (LTE {_deep_fmt(ul_retx_lte)}%, NR {_deep_fmt(ul_retx_nr)}%).",
            "The UE is UL power-limited at the cell edge — UL HARQ retransmissions slow TCP ramp-up and EN-DC signalling.",
            [
                "Check UL link budget / pathloss and serving-cell distance.",
                "Review PUSCH power control (P0, alpha) and PHR reporting.",
                "Check UL interference / RTWP on the serving anchor.",
            ], confidence="Direct KPI evidence"))

    # EN-DC NR secondary-node (SgNB) mobility instability — reconstructed from NR-SCG presence.
    sn_add_rate = iam.get("sgnbAdditionSuccessRate")
    sn_add_fail = iam.get("sgnbAdditionFailure") or 0
    sn_rem_fail = iam.get("sgnbRemovalFailure") or 0
    if (sn_add_rate is not None and sn_add_rate < 90) or sn_add_fail > 0 or sn_rem_fail > 0:
        findings.append(_deep_make_finding(
            "EN-DC stability", "SgNB add/remove (reconstructed)", sn_add_rate, None, "Medium",
            f"EN-DC secondary-node (SgNB) mobility: {_deep_fmt(iam.get('sgnbAdditions'))} additions "
            f"({_deep_fmt(iam.get('sgnbAdditionSuccess'))} sustained / {_deep_fmt(sn_add_fail)} aborted), "
            f"{_deep_fmt(iam.get('sgnbRemovals'))} removals ({_deep_fmt(sn_rem_fail)} re-activated within 3 s). "
            "Reconstructed from NR-SCG presence — the export carries no explicit SgNB RRC signalling.",
            "Unstable EN-DC secondary-node retention (aborted additions or removal ping-pong) interrupts the NR leg and dips throughput.",
            [
                "Review EN-DC addition/release thresholds (B1 / A2) and SCG-failure handling.",
                "Check NR SSB RSRP/SINR at the add/drop locations along the route.",
                "Inspect SgNB-change / X2-Xn events for ping-pong; confirm with an RRC trace.",
            ], confidence="Possible cause"))

    findings = _deep_consolidate_findings(findings)
    findings = _deep_merge_same_domain(findings)
    findings.sort(key=lambda f: (-(f.get("priorityScore") or 0), f.get("domain") or ""))
    return findings


def _deep_throughput_gap_attribution(iam, competitors):
    """Attribute IAM's DL-throughput deficit (vs the throughput leader) across the
    competitive drivers, ranked by relative shortfall. This is a relative-deficit
    decomposition (contributions sum to 100%), not a strict multiplicative throughput
    identity — the drivers are measured on different bases — so it answers 'where are the
    largest competitive gaps' rather than claiming an exact causal split."""
    iam_dl = _deep_num(iam.get("dlThroughput"))
    comp = [c for c in (competitors or []) if _deep_num(c.get("dlThroughput")) is not None]
    if iam_dl is None or not comp:
        return None
    leader = max(comp, key=lambda c: _deep_num(c.get("dlThroughput")))
    leader_dl = _deep_num(leader.get("dlThroughput"))
    if leader_dl is None or leader_dl <= iam_dl:
        return None  # IAM leads or ties → no gap to attribute
    # Independent multiplicative throughput factors: throughput ≈ bandwidth × utilization ×
    # spectral-efficiency × (1 - BLER). Modulation and MIMO live INSIDE spectral efficiency
    # (bits/PRB), so they are not counted separately (avoids double-counting). Each driver's
    # shortfall is throughput-proportional: a fractional deficit for the higher-is-better
    # factors, and the absolute pp loss for BLER (so a 1.6% BLER doesn't masquerade as a large
    # relative gap against a 0.5% competitor).
    raw = []
    for label, key in (
        ("Aggregated bandwidth", "availableBandwidthPrbs"),
        ("PRB utilization / scheduling", "dlPrbUtilPct"),
        ("Spectral efficiency (modulation × MIMO)", "schBitratePerPrb"),
    ):
        iv, lv = _deep_num(iam.get(key)), _deep_num(leader.get(key))
        if iv is None or lv is None or lv <= 0:
            continue
        shortfall = max(0.0, (lv - iv) / lv)
        if shortfall > 0.02:
            raw.append({"driver": label, "iam": round(iv, 2), "leader": round(lv, 2), "_s": shortfall})
    iam_bler, leader_bler = _deep_num(iam.get("blerAvg")), _deep_num(leader.get("blerAvg"))
    if iam_bler is not None and leader_bler is not None and iam_bler > leader_bler:
        shortfall = (iam_bler - leader_bler) / 100.0  # ~fractional throughput lost to retx
        if shortfall > 0.02:
            raw.append({"driver": "Reliability (BLER)", "iam": round(iam_bler, 2), "leader": round(leader_bler, 2), "_s": shortfall})
    total = sum(r["_s"] for r in raw)
    if not raw or total <= 0:
        return None
    for r in raw:
        r["contributionPct"] = round(r.pop("_s") / total * 100.0)
    raw.sort(key=lambda r: -r["contributionPct"])
    gap_pct = round((leader_dl - iam_dl) / leader_dl * 100.0, 1)
    summary = (
        f"The {gap_pct}% DL throughput deficit vs {leader.get('_operator') or 'the leader'} is driven mainly by "
        + ", ".join(f"{r['driver']} ({int(r['contributionPct'])}%)" for r in raw[:3]) + "."
    )
    return {"leader": leader.get("_operator"), "gapPercent": gap_pct, "drivers": raw, "summary": summary}


def _deep_causal_chain(iam, competitors):
    """Walk the DL-throughput causal pipeline (RF quality → channel feedback → spectral
    efficiency → resource scheduling → bandwidth → throughput) and find where it BREAKS:
    the first stage that is materially degraded while the upstream stages are healthy. A
    stage is 'weak' if it fails an absolute engineering threshold OR sits well below the
    throughput leader. The break point is the causal root cause; downstream stages are
    consequences. Returns ordered stages + a narrative that localizes the break."""
    comp = [c for c in (competitors or []) if _deep_num(c.get("dlThroughput")) is not None]
    leader = max(comp, key=lambda c: _deep_num(c.get("dlThroughput"))) if comp else None
    # (stage, key, unit, decimals, absolute-good threshold or None, leader-ratio floor)
    specs = [
        ("NR RF quality (SINR)", "sinrNr", "dB", 1, 7.0, 0.70),
        ("Channel feedback (CQI)", "cqi", "", 0, 9.0, 0.75),
        ("Spectral efficiency (bits/PRB)", "schBitratePerPrb", "b/PRB", 2, 0.60, 0.80),
        ("Resource scheduling (PRB util)", "dlPrbUtilPct", "%", 1, 5.0, 0.50),
        ("Aggregated bandwidth", "availableBandwidthPrbs", "PRBs", 0, None, 0.80),
    ]
    steps = []
    for stage, key, unit, dec, absth, ratfloor in specs:
        iv = _deep_num(iam.get(key))
        if key == "sinrNr" and iv is None:
            iv = _deep_num(iam.get("sinr"))
        if iv is None:
            continue
        lv = _deep_num((leader or {}).get(key)) if leader else None
        weak_abs = absth is not None and iv < absth
        weak_rel = lv is not None and lv > 0 and (iv / lv) < ratfloor
        steps.append({
            "stage": stage, "unit": unit,
            "iam": round(iv, dec) if dec else round(iv),
            "leader": (round(lv, dec) if dec else round(lv)) if lv is not None else None,
            "status": "weak" if (weak_abs or weak_rel) else "ok",
        })
    if not steps:
        return None
    dl = _deep_num(iam.get("dlThroughput"))
    ldl = _deep_num((leader or {}).get("dlThroughput")) if leader else None
    gap_pct = round((1.0 - dl / ldl) * 100.0) if (dl and ldl and ldl > 0 and dl < ldl) else None

    def short(s):
        return s["stage"].split(" (")[0]

    healthy_prefix = []
    for s in steps:
        if s["status"] == "ok":
            healthy_prefix.append(s)
        else:
            break
    break_idx = len(healthy_prefix)
    break_step = steps[break_idx] if break_idx < len(steps) else None
    weak_downstream = [s for s in steps[break_idx:] if s["status"] == "weak"]

    def val(s):
        return f"{s['iam']}{(' ' + s['unit']) if s['unit'] else ''}" + (
            f" vs {s['leader']} leader" if s.get("leader") is not None else "")

    if break_step:
        up = " and ".join(short(s) for s in healthy_prefix) if healthy_prefix else "the upstream radio layer"
        weak_txt = "; ".join(f"{short(s)} {val(s)}" for s in weak_downstream)
        narrative = (
            f"Causal chain: {up} {'are' if len(healthy_prefix) != 1 else 'is'} healthy, "
            f"but the chain breaks at {weak_txt}"
            + (f", cascading to the {gap_pct}% DL throughput gap" if gap_pct else "")
            + ". The limitation is "
            + ("downstream of the radio (scheduling / efficiency), not coverage or quality."
               if healthy_prefix and break_step["stage"].startswith(("Spectral", "Resource", "Aggregated"))
               else "in the radio layer itself.")
        )
    else:
        narrative = ("Causal chain: every stage from RF quality through scheduling and bandwidth is healthy; "
                     "the throughput gap is not explained by the standard DL pipeline — check transport/server or test-condition parity.")
    chain_text = " → ".join(f"{short(s)} {s['iam']}{(' ' + s['unit']) if s['unit'] else ''} [{s['status']}]" for s in steps)
    return {"steps": steps, "breakPoint": break_step["stage"] if break_step else None,
            "narrative": narrative, "chainText": chain_text}


def _benchmark_deep_exec_summary(iam, competitors, findings, scope_label):
    best_dl = _deep_best_competitor([c.get("dlThroughput") for c in competitors], True)
    best_dl_op = ""
    for c in competitors:
        if _deep_num(c.get("dlThroughput")) == best_dl and best_dl is not None:
            best_dl_op = c.get("_operator") or ""
            break
    dl = iam.get("dlThroughput")
    dl_gap = _deep_gap_fraction(dl, best_dl, True)
    crit_high = [f for f in findings if f.get("severity") in ("Critical", "High")]
    domains = []
    for f in crit_high:
        d = f.get("domain")
        if d and d not in domains:
            domains.append(d)
    access_ok = (iam.get("dlSuccess") is None or iam.get("dlSuccess") >= 98) and \
                (iam.get("dlCompletion") is None or iam.get("dlCompletion") >= 98)
    # Professional narrative conclusion: position → accessibility → layer balance → primary
    # root cause (from the top-ranked finding) → secondary gaps.
    primary = crit_high[0] if crit_high else None
    parts = []
    if dl is not None and best_dl is not None and dl_gap is not None:
        if dl_gap <= -0.05:
            parts.append(f"IAM DL throughput is {_deep_fmt(dl)} Mbps — {abs(round(dl_gap * 100, 1))}% below the benchmark leader ({(best_dl_op + ' ').strip()}{_deep_fmt(best_dl)} Mbps).")
        elif dl_gap >= 0.05:
            parts.append(f"IAM DL throughput is {_deep_fmt(dl)} Mbps, leading the benchmark ({(best_dl_op + ' ').strip()}{_deep_fmt(best_dl)} Mbps).")
        else:
            parts.append(f"IAM DL throughput is {_deep_fmt(dl)} Mbps, on par with the benchmark.")
    if access_ok:
        parts.append("Accessibility is not the constraint (DL completion and success ≥ 98%).")
    nr_contrib = iam.get("nrThroughputContrib")
    if nr_contrib is not None:
        parts.append(f"Traffic is carried {_deep_fmt(nr_contrib)}% on NR and {_deep_fmt(iam.get('lteThroughputContrib'))}% on the LTE anchor.")
    gap_attribution = _deep_throughput_gap_attribution(iam, competitors)
    if gap_attribution:
        parts.append(gap_attribution.get("summary") or "")
    # Causal chain is surfaced as its own structured section (execSummary.causalChain),
    # not appended to the conclusion, to avoid overlapping with the gap attribution above.
    causal_chain = _deep_causal_chain(iam, competitors)
    if primary:
        parts.append(f"Primary limitation — {primary.get('domain')}: {primary.get('rootCause')}")
        secondary = [d for d in domains if d != primary.get('domain')][:4]
        if secondary:
            parts.append("Secondary gaps: " + ", ".join(secondary) + ".")
    elif not domains:
        parts.append("No critical IAM gaps detected in this scope.")
    main_conclusion = " ".join(p for p in parts if p)
    top_kpis = []
    if dl is not None and best_dl is not None:
        top_kpis.append({"kpi": "DL Throughput", "iam": f"{_deep_fmt(dl)} Mbps",
                         "bestCompetitor": f"{best_dl_op} {_deep_fmt(best_dl)} Mbps".strip(),
                         "gap": f"IAM is {abs(round((dl_gap or 0) * 100, 1))}% {'lower' if (dl_gap or 0) < 0 else 'higher'}"})
    if iam.get("fivegPresence") is not None:
        top_kpis.append({"kpi": "5G layer", "iam": f"{_deep_fmt(iam.get('fivegPresence'))}% 5G, n78 {_deep_fmt(iam.get('n78'))}%",
                         "bestCompetitor": f"best n78 {_deep_fmt(_deep_best_competitor([c.get('n78') for c in competitors], True))}%",
                         "gap": "n78 capacity layer gap" if (iam.get("n78") or 0) < 30 else "comparable"})
    if iam.get("medianRank") is not None:
        top_kpis.append({"kpi": "MIMO rank", "iam": f"Median RI={_deep_fmt(iam.get('medianRank'))}, RI1={_deep_fmt(iam.get('ri1'))}%",
                         "bestCompetitor": f"best RI2={_deep_fmt(_deep_best_competitor([c.get('ri2') for c in competitors], True))}%",
                         "gap": "Weak spatial multiplexing" if (iam.get("medianRank") or 9) <= 1 else "OK"})
    if iam.get("rsrpNr") is not None or iam.get("rsrpLte") is not None:
        top_kpis.append({"kpi": "RF by RAT", "iam": f"NR RSRP {_deep_fmt(iam.get('rsrpNr'))} / LTE RSRP {_deep_fmt(iam.get('rsrpLte'))} dBm; NR SINR {_deep_fmt(iam.get('sinrNr'))} dB",
                         "bestCompetitor": f"best NR RSRP {_deep_fmt(_deep_best_competitor([c.get('rsrpNr') for c in competitors], True))} dBm",
                         "gap": "NR layer weak vs healthy LTE anchor" if (iam.get('rsrpNr') is not None and iam.get('rsrpNr') < -105 and (iam.get('rsrpLte') is None or iam.get('rsrpLte') >= -100)) else ("NR coverage limited" if (iam.get('rsrpNr') or 0) < -105 else "OK")})
    if iam.get("nrThroughputContrib") is not None:
        top_kpis.append({"kpi": "Layer balance", "iam": f"NR {_deep_fmt(iam.get('nrThroughputContrib'))}% / LTE {_deep_fmt(iam.get('lteThroughputContrib'))}% of MAC DL bytes",
                         "bestCompetitor": "—",
                         "gap": "NR underused despite EN-DC" if (iam.get('nrThroughputContrib') or 100) < 50 else "NR-dominant"})
    if iam.get("sinr") is not None:
        top_kpis.append({"kpi": "SINR/CQI/MCS (NR)", "iam": f"SINR {_deep_fmt(iam.get('sinr'))} dB, CQI {_deep_fmt(iam.get('cqi'))}, MCS {_deep_fmt(iam.get('mcs'))}",
                         "bestCompetitor": f"best CQI {_deep_fmt(_deep_best_competitor([c.get('cqi') for c in competitors], True))}",
                         "gap": "Low quality limits modulation" if (iam.get("sinr") or 99) < 5 else "Acceptable"})
    if iam.get("blerAvg") is not None:
        top_kpis.append({"kpi": "BLER", "iam": f"Avg {_deep_fmt(iam.get('blerAvg'))}%, P90 {_deep_fmt(iam.get('blerP90'))}%, >10%={_deep_fmt(iam.get('blerAbove10'))}%",
                         "bestCompetitor": f"best avg {_deep_fmt(_deep_best_competitor([c.get('blerAvg') for c in competitors], False))}%",
                         "gap": "High retransmission zones" if (iam.get("blerAvg") or 0) > 5 else "OK"})
    if iam.get("tcpHandshake") is not None:
        top_kpis.append({"kpi": "Transport", "iam": f"TCP handshake {_deep_fmt(iam.get('tcpHandshake'))} ms",
                         "bestCompetitor": f"best {_deep_fmt(_deep_best_competitor([c.get('tcpHandshake') for c in competitors], False))} ms",
                         "gap": "Core/transport path to verify" if (iam.get("tcpHandshake") or 0) > 80 else "OK"})
    p1p2 = [f.get("domain") for f in findings if f.get("priority") in ("P1", "P2")]
    seen = []
    for d in p1p2:
        if d not in seen:
            seen.append(d)
    immediate = "; ".join(f"{i+1}) {d}" for i, d in enumerate(seen[:5])) if seen else "No immediate P1/P2 actions."

    # Diagnosis confidence scales with the evidence behind the conclusions — number of
    # download sessions and RF samples. A single-session scope is exploratory, not conclusive.
    test_count = int(iam.get("testCount") or 0)
    nr_rf = int(iam.get("nrRfSamples") or 0)
    lte_rf = int(iam.get("lteRfSamples") or 0)
    rf_samples = nr_rf + lte_rf
    if test_count >= 5 and rf_samples >= 300:
        conf_level = "High"
    elif test_count >= 2 and rf_samples >= 100:
        conf_level = "Medium"
    else:
        conf_level = "Low"
    conf_note = (
        f"{conf_level} confidence — {test_count} download session(s), {rf_samples} RF samples "
        f"({nr_rf} NR / {lte_rf} LTE)."
        + (" Single/low-sample scope: treat as directional and validate on repeated drive tests before field action."
           if conf_level == "Low" else
           (" Limited sample: confirm trends on additional drive tests." if conf_level == "Medium" else ""))
    )
    if conf_level == "Low":
        main_conclusion = main_conclusion + " [Low-confidence scope: directional only — validate on repeated DTs.]"
    return {
        "title": f"{scope_label} - IAM Professional Analysis",
        "scope": "IAM only, benchmarked against Orange and INWI where useful",
        "mainConclusion": main_conclusion,
        "confidence": conf_level,
        "confidenceNote": conf_note,
        "gapAttribution": gap_attribution,
        "causalChain": causal_chain,
        "topKpis": top_kpis,
        "immediatePriorities": immediate,
    }


def _benchmark_raw_parsing_qa(operator_files: list[dict] | None) -> dict:
    entries = []
    total_normalized = 0
    for item in operator_files or []:
        qa = (item or {}).get("parsingQa") or {}
        normalized_count = int(qa.get("normalizedValueCount") or 0)
        total_normalized += normalized_count
        field_counts = qa.get("normalizedFields") or {}
        field_summary = ", ".join(
            f"{field}={count}" for field, count in sorted(field_counts.items())
        ) or "none"
        samples = []
        for sample in list(qa.get("normalizedSamples") or [])[:3]:
            samples.append(
                f"row {sample.get('row')}: {sample.get('field')} {sample.get('before')} -> {sample.get('after')}"
            )
        entries.append({
            "operator": item.get("operator") or "UNKNOWN",
            "normalizedValueCount": normalized_count,
            "normalizedFieldSummary": field_summary,
            "samples": samples,
            "throughputScales": item.get("throughputScales") or {},
        })
    summary = (
        f"Normalized {total_normalized} suspicious raw Nemo numeric values before KPI analysis."
        if total_normalized
        else "No suspicious raw Nemo numeric values required normalization in the current scope."
    )
    return {
        "title": "Raw Nemo parsing QA",
        "summary": summary,
        "normalizedValueCount": total_normalized,
        "operators": entries,
    }


def _deep_supporting_cells(serving_cells: dict | None, max_items: int = 3) -> list[str]:
    if not isinstance(serving_cells, dict):
        return []
    cells = list(serving_cells.get("cells") or [])
    if not cells:
        return []

    def _sort_value(cell, key):
        value = _deep_num((cell or {}).get(key))
        return value if value is not None else -1.0

    cells.sort(
        key=lambda cell: (
            -_sort_value(cell, "dwellSecDownload"),
            -_sort_value(cell, "dwellSec"),
            -_sort_value(cell, "primaryHitCount"),
            -_sort_value(cell, "hitCount"),
        )
    )
    names = []
    for cell in cells:
        name = str((cell or {}).get("cellName") or "").strip()
        if name and name not in names:
            names.append(name)
        if len(names) >= max_items:
            break
    return names


def _deep_supporting_segments(serving_cells: dict | None, max_items: int = 3) -> list[str]:
    if not isinstance(serving_cells, dict):
        return []
    episodes = list(serving_cells.get("episodesDownload") or serving_cells.get("episodesAll") or [])
    if not episodes:
        return []
    segments = []
    for episode in episodes:
        name = str((episode or {}).get("cellName") or "").strip() or "Unknown cell"
        start = str((episode or {}).get("startTime") or (episode or {}).get("start") or "").strip()
        end = str((episode or {}).get("endTime") or (episode or {}).get("end") or "").strip()
        dwell = _deep_num((episode or {}).get("dwellSec"))
        if start or end:
            label = f"{name}: {start or '—'} -> {end or '—'}"
        else:
            label = name
        if dwell is not None:
            label += f" ({_deep_fmt(dwell)} s)"
        if label not in segments:
            segments.append(label)
        if len(segments) >= max_items:
            break
    return segments


def _deep_download_window_nr_share(serving_cells: dict | None) -> tuple[float | None, float | None, float | None]:
    if not isinstance(serving_cells, dict):
        return None, None, None
    cells = list(serving_cells.get("cells") or [])
    tech_by_name = {}
    effective_target_total = 0.0
    for cell in cells:
        name = str((cell or {}).get("cellName") or "").strip()
        tech = str((cell or {}).get("tech") or "").strip().upper()
        if name and tech:
            tech_by_name[name] = tech
        app_sample_count = _deep_num((cell or {}).get("appSampleCount"))
        dwell_download = _deep_num((cell or {}).get("dwellSecDownload"))
        if app_sample_count is not None and app_sample_count > 0 and dwell_download is not None and dwell_download > 0:
            effective_target_total += float(dwell_download)
    episodes = list(serving_cells.get("episodesDownload") or [])
    if effective_target_total > 0 and episodes:
        effective_episodes = []
        running_total = 0.0
        threshold = max(0.0, effective_target_total - 0.25)
        for episode in episodes:
            dwell = _deep_num((episode or {}).get("dwellSec"))
            if dwell is None or dwell <= 0:
                continue
            effective_episodes.append(episode)
            running_total += float(dwell)
            if running_total >= threshold:
                break
        episodes = effective_episodes
    nr_dwell = 0.0
    total_dwell = 0.0
    for episode in episodes:
        dwell = _deep_num((episode or {}).get("dwellSec"))
        if dwell is None or dwell <= 0:
            continue
        total_dwell += float(dwell)
        name = str((episode or {}).get("cellName") or "").strip()
        tech = tech_by_name.get(name)
        if not tech:
            if name.startswith("5G_"):
                tech = "5G"
            elif name.startswith("4G_"):
                tech = "4G"
        if tech == "5G":
            nr_dwell += float(dwell)
    if total_dwell <= 0:
        return None, None, None
    return round((nr_dwell / total_dwell) * 100.0, 1), round(nr_dwell, 1), round(total_dwell, 1)


def _deep_collect_action_evidence(finding: dict, iam: dict, serving_cells: dict | None = None) -> tuple[list[str], list[str], list[str], list[str]]:
    domain = str((finding or {}).get("domain") or "")
    metrics: list[str] = []
    txt_evidence: list[str] = []
    supporting_cells = _deep_supporting_cells(serving_cells)
    supporting_segments = _deep_supporting_segments(serving_cells)

    def add_metric(label: str, value, unit: str = ""):
        num = _deep_num(value)
        if num is None:
            return
        suffix = f" {unit}" if unit else ""
        metrics.append(f"{label} = {_deep_fmt(num)}{suffix}")

    if domain == "5G capacity layer":
        add_metric("DL Throughput", iam.get("dlThroughput"), "Mbps")
        add_metric("5G presence", iam.get("fivegPresence"), "%")
        add_metric("4G-only time", iam.get("fourgOnly"), "%")
        add_metric("NR n78 share", iam.get("n78"), "%")
        add_metric("NR n28 share", iam.get("n28"), "%")
        add_metric("PDSCH DL Avg", iam.get("pdschDlAvg"), "Mbps")
    elif domain == "MIMO / RI":
        add_metric("Median Rank", iam.get("medianRank"))
        add_metric("RI1 share", iam.get("ri1"), "%")
        add_metric("RI2 share", iam.get("ri2"), "%")
        add_metric("RI>=3 share", iam.get("riGe3"), "%")
    elif domain == "Coverage / dominance":
        add_metric("Median RSRP", iam.get("rsrp"), "dBm")
        add_metric("Median SINR", iam.get("sinr"), "dB")
        add_metric("Serving-cell distance", iam.get("servingCellDistanceM"), "m")
    elif domain in ("Radio quality / interference", "SINR / interference", "Modulation profile"):
        add_metric("Median RSRQ", iam.get("rsrq"), "dB")
        add_metric("Median SINR", iam.get("sinr"), "dB")
        add_metric("Median CQI", iam.get("cqi"))
        add_metric("Median MCS", iam.get("mcs"))
        add_metric("256QAM share", iam.get("qam256"), "%")
        add_metric("16QAM share", iam.get("qam16"), "%")
        add_metric("BLER Avg", iam.get("blerAvg"), "%")
    elif domain == "Load / congestion":
        add_metric("DL PRB utilization", iam.get("prbUtilPct"), "%")
        add_metric("DL Throughput", iam.get("dlThroughput"), "Mbps")
        add_metric("PRB efficiency", iam.get("prbEfficiency"))
    elif domain == "Bandwidth / spectrum":
        add_metric("Available bandwidth", iam.get("availableBandwidthPrbs"), "PRBs")
        add_metric("NR n78 share", iam.get("n78"), "%")
        add_metric("Avg #SCells", iam.get("scellsAvg"))
    elif domain == "LTE anchor / NSA dependency":
        add_metric("5G presence", iam.get("fivegPresence"), "%")
        add_metric("LTE-anchor SINR", iam.get("lteAnchorSinr"), "dB")
        add_metric("4G-only time", iam.get("fourgOnly"), "%")
    elif domain == "EN-DC stability":
        add_metric("EN-DC setup success", iam.get("endcSetupSuccessRate"), "%")
        add_metric("EN-DC drop rate", iam.get("endcDropRate"), "%")
        add_metric("5G presence", iam.get("fivegPresence"), "%")
        add_metric("LTE-anchor SINR", iam.get("lteAnchorSinr"), "dB")
    elif domain == "Carrier aggregation":
        add_metric("Avg #SCells", iam.get("scellsAvg"))
        add_metric("SCells >0 share", iam.get("scellsActive"), "%")
        add_metric("LTE CA active share", iam.get("caActive"), "%")
    elif domain == "BLER / retransmissions":
        add_metric("BLER Avg", iam.get("blerAvg"), "%")
        add_metric("BLER P90", iam.get("blerP90"), "%")
        add_metric("BLER >10% share", iam.get("blerAbove10"), "%")
    elif domain == "UL quality":
        add_metric("UL Retx Avg", iam.get("ulRetx"), "%")
        add_metric("TCP Handshake", iam.get("tcpHandshake"), "ms")
    elif domain == "Transport / core":
        add_metric("TCP Handshake", iam.get("tcpHandshake"), "ms")
        add_metric("DL completion", iam.get("dlCompletion"), "%")
        add_metric("DL success", iam.get("dlSuccess"), "%")
    elif domain == "Mobility / serving sequence":
        add_metric("4G-only time", iam.get("fourgOnly"), "%")
        add_metric("5G presence", iam.get("fivegPresence"), "%")

    if supporting_cells:
        txt_evidence.append("IAM TXT serving cells: " + ", ".join(supporting_cells))
    if supporting_segments:
        txt_evidence.append("IAM TXT download-window sequence: " + "; ".join(supporting_segments))
    return metrics, txt_evidence, supporting_cells, supporting_segments


def _deep_attach_capacity_context(row: dict, iam: dict, raw_kpis: dict | None = None) -> dict:
    clone = dict(row or {})
    if str(clone.get("domain") or "") != "5G capacity layer":
        return clone

    raw_kpis = raw_kpis or {}
    sub_causes = list(clone.get("subCauses") or [])
    excluded_checks = list(clone.get("excludedChecks") or [])
    band_shares = raw_kpis.get("nrBandShares") or {}

    def add_sub_cause(label: str, value, unit: str = "%"):
        num = _deep_num(value)
        if num is None:
            return
        suffix = f" {unit}" if unit else ""
        text = f"{label} = {_deep_fmt(num)}{suffix}"
        if text not in sub_causes:
            sub_causes.append(text)

    def add_excluded_check(text: str):
        clean = str(text or "").strip()
        if clean and clean not in excluded_checks:
            excluded_checks.append(clean)

    add_sub_cause("NR n78 share", iam.get("n78"))
    add_sub_cause("NR n1 share", iam.get("n1") if iam.get("n1") is not None else band_shares.get("n1"), "of NR samples")
    add_sub_cause("5G presence", iam.get("fivegPresence"))
    add_sub_cause("4G-only time", iam.get("fourgOnly"))

    median_rank = _deep_num(iam.get("medianRank"))
    if median_rank is None:
        median_rank = _deep_num((raw_kpis.get("rankIndicator") or {}).get("median"))
    ri1 = _deep_num(iam.get("ri1"))
    if ri1 is None:
        ri1 = _deep_num(raw_kpis.get("ri1Share"))
    ri2 = _deep_num(iam.get("ri2"))
    if ri2 is None:
        ri2 = _deep_num(raw_kpis.get("ri2Share"))
    if (
        median_rank is not None and median_rank >= 2
        and ri1 is not None and ri1 <= 10
        and ri2 is not None and ri2 >= 80
    ):
        add_excluded_check(
            "MIMO rank is acceptable for this DT: "
            f"median RI={_deep_fmt(median_rank)}, RI1={_deep_fmt(ri1)}%, RI2={_deep_fmt(ri2)}%."
        )

    clone["subCauses"] = _deep_deduplicate_text_list(sub_causes)
    clone["excludedChecks"] = _deep_deduplicate_text_list(excluded_checks)
    return clone


def _deep_action_missing_data(action_text: str) -> list[str]:
    text = str(action_text or "").lower()
    missing = []

    def add(msg: str):
        if msg not in missing:
            missing.append(msg)

    if any(marker in text for marker in ("x2/xn", "gnb/enb alarms", "alarms")):
        add("OSS interface/alarm data are not present in the current benchmark TXT.")
    if any(marker in text for marker in ("neighbor configuration", "neighbor priorities", "anchor priorities", "a3", "a5", "b1", "b2", "thresholds", "timers")):
        add("RRC/neighbor configuration exports are not present in the current benchmark TXT.")
    if any(marker in text for marker in ("ssb", "beam", "beamforming")):
        add("SSB/beam metrics are not present in the current benchmark TXT.")
    if any(marker in text for marker in ("dns/apn", "firewall/nat", "server route")):
        add("Core-network path details are not present in the current benchmark TXT.")
    if any(marker in text for marker in ("rru", "vswr", "feeder", "cross-polar", "calibration alarms", "antenna ports")):
        add("Hardware/alarm/field-measurement data are not present in the current benchmark TXT.")
    if any(marker in text for marker in ("qos", "scheduler profile", "vendor feature", "license", "barred")):
        add("Vendor configuration/policy data are not present in the current benchmark TXT.")
    return missing


def _deep_action_status(action_text: str, finding: dict, evidence_metrics: list[str]) -> tuple[str, str, list[str]]:
    text = str(action_text or "").lower()
    missing = _deep_action_missing_data(action_text)
    confidence = str((finding or {}).get("confidence") or "")
    relevance = str((finding or {}).get("benchmarkRelevance") or "")

    if any(marker in text for marker in ("x2/xn", "gnb/enb alarms", "alarms", "ssb", "beam", "neighbor configuration", "anchor priorities", "b1", "b2", "vendor feature", "license", "barred", "dns/apn", "firewall/nat")):
        return (
            "Hypothesis",
            "The current benchmark/TXT suggests this area, but this specific check needs external network/configuration data.",
            missing,
        )
    if any(marker in text for marker in ("azimuth", "tilt", "scheduler", "backhaul", "transport", "load balancing", "power balancing", "new site", "sector split", "traffic hotspot", "pathloss", "pusch power control", "en-dc configuration")):
        return (
            "Partial",
            "Current benchmark/TXT evidence supports this action area, but the exact root cause still needs deeper network verification.",
            missing,
        )
    if evidence_metrics and confidence == "High" and relevance in ("Primary", "Secondary"):
        return (
            "Confirmed",
            "Current benchmark KPIs and IAM TXT symptoms directly support this recommendation.",
            missing,
        )
    if evidence_metrics:
        return (
            "Partial",
            "Current benchmark/TXT data support the symptom, but not the full root cause behind this action.",
            missing,
        )
    return (
        "Hypothesis",
        "Current data are insufficient to confirm this recommendation.",
        missing,
    )


def _deep_validation_text(
    validation_status: str,
    domain: str,
    finding: str,
    evidence_metrics: list[str],
    evidence_from_txt: list[str],
) -> tuple[str, str]:
    domain_text = str(domain or "this area")
    finding_text = str(finding or "").strip()
    priority_map = {
        "5G capacity layer": [
            "NR n78 share",
            "5G presence",
            "PDSCH DL Avg",
            "DL Throughput",
            "4G-only time",
            "NR n28 share",
        ],
        "Bandwidth / spectrum": [
            "Available bandwidth",
            "Avg #SCells",
            "NR n78 share",
        ],
        "SINR / interference": [
            "Median SINR",
            "Median RSRQ",
            "BLER Avg",
            "Median CQI",
        ],
        "BLER / retransmissions": [
            "BLER Avg",
            "BLER P90",
            "BLER >10% share",
        ],
    }
    preferred = priority_map.get(domain_text, [])
    ranked_metrics = []
    remaining_metrics = [str(item) for item in list(evidence_metrics or []) if item]
    for marker in preferred:
        for item in list(remaining_metrics):
            if item.startswith(marker) and item not in ranked_metrics:
                ranked_metrics.append(item)
                remaining_metrics.remove(item)
                break
    ranked_metrics.extend(remaining_metrics)
    top_evidence = ranked_metrics[:3]
    if len(top_evidence) < 3:
        for item in list(evidence_from_txt or []):
            if item and item not in top_evidence:
                top_evidence.append(str(item))
            if len(top_evidence) >= 3:
                break
    evidence_summary = (
        "Evidence: " + "; ".join(top_evidence)
        if top_evidence
        else "Evidence: no direct KPI/TXT proof is available in the current scope."
    )
    if validation_status == "Confirmed":
        summary = (
            f"Confirmed for {domain_text}: "
            f"{finding_text or 'the current benchmark/TXT data directly support this issue.'}"
        )
    elif validation_status == "Partial":
        summary = (
            f"Partial for {domain_text}: "
            f"{finding_text or 'the current benchmark/TXT data support the symptom,'} "
            "but it still needs deeper network-side confirmation for the exact root cause."
        )
    else:
        summary = (
            f"Hypothesis for {domain_text}: "
            f"{finding_text or 'the current benchmark/TXT data only suggest this issue so far.'}"
        )
    return summary, evidence_summary


def _deep_deduplicate_text_list(values) -> list[str]:
    items: list[str] = []
    seen: set[str] = set()
    for value in values or []:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        items.append(text)
    return items


def _deep_merge_capacity_action_plan_rows(rows: list[dict]) -> list[dict]:
    if not rows:
        return []

    capacity_domains = {"5G capacity layer", "Bandwidth / spectrum"}
    items = [dict(row or {}) for row in rows]
    capacity_rows = [row for row in items if str((row or {}).get("domain") or "") in capacity_domains]
    fiveg_rows = [row for row in capacity_rows if str((row or {}).get("domain") or "") == "5G capacity layer"]
    bandwidth_rows = [row for row in capacity_rows if str((row or {}).get("domain") or "") == "Bandwidth / spectrum"]
    if len(fiveg_rows) <= 1 and (not fiveg_rows or not bandwidth_rows):
        return items

    other = [row for row in items if str((row or {}).get("domain") or "") not in capacity_domains]
    severity_rank = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}
    validation_rank = {"Confirmed": 3, "Partial": 2, "Hypothesis": 1}
    confidence_rank = {"High": 3, "Medium": 2, "Low": 1}
    relevance_rank = {"Primary": 3, "Secondary": 2, "Context": 1}

    def row_sort_key(row: dict) -> tuple[int, int, int, int, int, int]:
        domain = str((row or {}).get("domain") or "")
        finding_text = str((row or {}).get("finding") or "").lower()
        root_cause_text = str((row or {}).get("rootCause") or "").lower()
        text_blob = " ".join((finding_text, root_cause_text))
        headline_capacity_row = (
            domain == "5G capacity layer"
            and ("throughput" in text_blob or "accessibility" in text_blob)
        )
        return (
            1 if headline_capacity_row else 0,
            1 if domain == "5G capacity layer" else 0,
            validation_rank.get(str((row or {}).get("validationStatus") or ""), 0),
            confidence_rank.get(str((row or {}).get("confidence") or ""), 0),
            relevance_rank.get(str((row or {}).get("benchmarkRelevance") or ""), 0),
            int(_deep_num((row or {}).get("priorityScore")) or 0),
        )

    ordered_rows = sorted(capacity_rows, key=row_sort_key, reverse=True)
    lead = ordered_rows[0]
    merged = dict(lead)
    merged["domain"] = "5G capacity layer"
    merged["mergedDomains"] = _deep_deduplicate_text_list(
        [row.get("domain") for row in ordered_rows] + list(lead.get("mergedDomains") or [])
    )
    merged["recommendedActions"] = _deep_deduplicate_text_list(
        action
        for row in ordered_rows
        for action in list((row or {}).get("recommendedActions") or [])
    )
    merged["evidenceMetrics"] = _deep_deduplicate_text_list(
        metric
        for row in ordered_rows
        for metric in list((row or {}).get("evidenceMetrics") or [])
    )
    merged["evidenceFromTxt"] = _deep_deduplicate_text_list(
        item
        for row in ordered_rows
        for item in list((row or {}).get("evidenceFromTxt") or [])
    )
    merged["supportingCells"] = _deep_deduplicate_text_list(
        item
        for row in ordered_rows
        for item in list((row or {}).get("supportingCells") or [])
    )
    merged["supportingSegments"] = _deep_deduplicate_text_list(
        item
        for row in ordered_rows
        for item in list((row or {}).get("supportingSegments") or [])
    )
    merged["excludedChecks"] = _deep_deduplicate_text_list(
        item
        for row in ordered_rows
        for item in list((row or {}).get("excludedChecks") or [])
    )
    merged["missingData"] = _deep_deduplicate_text_list(
        item
        for row in ordered_rows
        for item in list((row or {}).get("missingData") or [])
    )

    merged_details = []
    seen_actions = set()
    for row in ordered_rows:
        for detail in list((row or {}).get("recommendedActionsDetailed") or []):
            text = str((detail or {}).get("text") or "").strip()
            if not text or text in seen_actions:
                continue
            seen_actions.add(text)
            merged_details.append(dict(detail or {}))
    if merged_details:
        merged["recommendedActionsDetailed"] = merged_details

    sub_causes = []
    for row in ordered_rows:
        sub_causes.extend(list((row or {}).get("subCauses") or []))
        sub_causes.extend(
            metric
            for metric in list((row or {}).get("evidenceMetrics") or [])
            if str(metric or "").startswith("NR n78 share")
        )
    merged["subCauses"] = _deep_deduplicate_text_list(sub_causes)
    merged["severity"] = max(
        ordered_rows,
        key=lambda row: severity_rank.get(str((row or {}).get("severity") or ""), 0),
    ).get("severity")
    merged["confidence"] = max(
        ordered_rows,
        key=lambda row: confidence_rank.get(str((row or {}).get("confidence") or ""), 0),
    ).get("confidence")
    merged["benchmarkRelevance"] = max(
        ordered_rows,
        key=lambda row: relevance_rank.get(str((row or {}).get("benchmarkRelevance") or ""), 0),
    ).get("benchmarkRelevance")
    merged["priorityScore"] = max(int(_deep_num((row or {}).get("priorityScore")) or 0) for row in ordered_rows)
    merged["priority"] = _deep_priority_from_score(merged["priorityScore"])
    merged["validationStatus"] = max(
        ordered_rows,
        key=lambda row: validation_rank.get(str((row or {}).get("validationStatus") or ""), 0),
    ).get("validationStatus")
    validation_summary, validation_evidence_summary = _deep_validation_text(
        merged.get("validationStatus"),
        merged.get("domain"),
        merged.get("finding"),
        merged.get("evidenceMetrics") or [],
        merged.get("evidenceFromTxt") or [],
    )
    merged["validationSummary"] = validation_summary
    merged["validationEvidenceSummary"] = validation_evidence_summary
    return other + [merged]


_DEEP_PRIORITY_SCORES = {
    "P1": 12,
    "P2": 8,
    "P3": 5,
    "P4": 2,
}


def _deep_set_priority(row: dict, priority: str, score: int | None = None) -> dict:
    clone = dict(row or {})
    normalized = str(priority or "P4").upper()
    clone["priority"] = normalized
    clone["priorityScore"] = int(score if score is not None else _DEEP_PRIORITY_SCORES.get(normalized, 2))
    return clone


def _deep_rebalance_action_plan_priorities(rows: list[dict]) -> list[dict]:
    if not rows:
        return []

    adjusted = [dict(row or {}) for row in rows]
    mobility_idx = None
    fiveg_rows: list[tuple[int, dict]] = []

    for index, row in enumerate(adjusted):
        domain = str((row or {}).get("domain") or "")
        finding_text = str((row or {}).get("finding") or "").lower()
        root_cause_text = str((row or {}).get("rootCause") or "").lower()
        action_text = " ".join(str(item or "") for item in list((row or {}).get("recommendedActions") or [])).lower()
        text_blob = " ".join((finding_text, root_cause_text, action_text))

        if domain == "Mobility / serving sequence" and (
            "not effectively carried by 5g" in text_blob or "download-window" in text_blob or "active dl" in text_blob
        ):
            mobility_idx = index
            adjusted[index] = _deep_set_priority(row, "P1", 13)
            continue

        if domain == "5G capacity layer":
            fiveg_rows.append((index, row))
            continue

        if domain == "Scheduler / PRB efficiency":
            if str((row or {}).get("validationStatus") or "") == "Hypothesis" or str((row or {}).get("confidence") or "") == "Low":
                adjusted[index] = _deep_set_priority(row, "P3")
            else:
                adjusted[index] = _deep_set_priority(row, "P2")
            continue

        if domain in ("BLER / retransmissions", "Bandwidth / spectrum", "Transport / core", "Capacity / configuration"):
            adjusted[index] = _deep_set_priority(row, "P2")
            continue

        if domain == "Retest governance":
            adjusted[index] = _deep_set_priority(row, "P3")

    if fiveg_rows:
        def fiveg_sort_key(item: tuple[int, dict]) -> tuple[int, int, int, int]:
            _, row = item
            finding_text = str((row or {}).get("finding") or "").lower()
            root_cause_text = str((row or {}).get("rootCause") or "").lower()
            action_text = " ".join(str(value or "") for value in list((row or {}).get("recommendedActions") or [])).lower()
            text_blob = " ".join((finding_text, root_cause_text, action_text))
            validation_status = str((row or {}).get("validationStatus") or "")
            confidence = str((row or {}).get("confidence") or "")
            relevance = str((row or {}).get("benchmarkRelevance") or "")
            current_score = int(_deep_num((row or {}).get("priorityScore")) or 0)
            return (
                1 if "n78" in text_blob else 0,
                1 if "high-capacity" in text_blob or "capacity layer" in text_blob or "missing" in text_blob else 0,
                1 if validation_status == "Confirmed" or confidence == "High" or relevance == "Primary" else 0,
                current_score,
            )

        fiveg_rows.sort(key=fiveg_sort_key, reverse=True)
        primary_index, primary_row = fiveg_rows[0]
        adjusted[primary_index] = _deep_set_priority(primary_row, "P1", 12 if mobility_idx is not None else 13)
        for secondary_index, secondary_row in fiveg_rows[1:]:
            adjusted[secondary_index] = _deep_set_priority(secondary_row, "P2", 9)

    adjusted.sort(
        key=lambda row: (
            -int(_deep_num((row or {}).get("priorityScore")) or 0),
            str((row or {}).get("domain") or ""),
            str((row or {}).get("finding") or ""),
        )
    )
    return adjusted


def _deep_bler_localization_rows(operator_file: dict | None, threshold: float = 10.0, limit: int = 25) -> list[dict]:
    """Localize BLER>threshold events for the current DT scope: one row per sample with
    time/GPS/serving cell/PSCell + radio context (SINR/MCS/modulation/rank/throughput).
    Generic over TXT exports — fields absent in a given file simply come back empty/None,
    so the table never blocks. Sorted worst-BLER first, capped at `limit`."""
    rows = []
    for raw in (operator_file or {}).get("rows") or []:
        bler = _deep_num(raw.get("macDlBler"))
        if bler is None or bler <= threshold:
            continue
        rows.append({
            "time": str(raw.get("time") or ""),
            "lat": _deep_num(raw.get("lat")),
            "lon": _deep_num(raw.get("lon")),
            "servingCell": str(raw.get("servingCellName") or raw.get("servingCell") or ""),
            "nrPscell": str(raw.get("nrPscellName") or raw.get("psCellName") or ""),
            "bler": bler,
            "sinr": _deep_num(raw.get("sinr")),
            "mcs": _deep_num(raw.get("pdschMcs")),
            "modulation": str(raw.get("pdschModulation") or ""),
            "rank": _deep_num(raw.get("rankIndicator")),
            "throughput": _deep_num(raw.get("appDlMbps")),
        })
    rows.sort(key=lambda item: (-(item.get("bler") or 0), str(item.get("time") or "")))
    return rows[:limit]


def _deep_simple_row(priority, domain, finding, action, confidence):
    return {
        "priority": priority,
        "domain": domain,
        "finding": finding,
        "action": action,
        "confidence": confidence,
    }


def _deep_domain_targeted_actions(domain, iam, best):
    """Targeted actions (each with a rationale tying it to the weak sub-metric) for one
    domain, selected from the specific IAM metrics. `best(key, higher)` returns the best
    competitor value. Returns [{action, rationale}] (possibly empty when the domain is OK)."""
    n = _deep_num
    f = _deep_fmt
    out = []
    if domain == "RF quality":
        sinr = n(iam.get("sinr")); rsrp = n(iam.get("rsrpMedian")); cqi = n(iam.get("cqiMedian"))
        if sinr is not None and sinr < 8:
            if rsrp is not None and rsrp >= -100:
                out.append({"action": "Audit overshooting cells, pilot pollution and PCI confusion; clean missing neighbors; correct mechanical/electrical tilt and azimuth.",
                            "rationale": f"RSRP is adequate ({f(rsrp)} dBm) but SINR is low ({f(sinr)} dB) → interference-dominated, not coverage."})
            else:
                out.append({"action": "Improve dominance/coverage first (tilt/azimuth/power), then re-check interference and neighbors.",
                            "rationale": f"Both RSRP ({f(rsrp)} dBm) and SINR ({f(sinr)} dB) are weak → coverage + interference."})
        if cqi is not None and cqi < 9:
            out.append({"action": "Review CQI reporting, OLLA and CQI aging; verify PDSCH power allocation.",
                        "rationale": f"Median CQI {f(cqi)} caps link adaptation and modulation order."})
    elif domain == "MIMO / spatial multiplexing":
        rank = n(iam.get("medianRank")); ri1 = n(iam.get("ri1")); ri2 = n(iam.get("ri2"))
        if (rank is not None and rank <= 1) or (ri1 is not None and ri1 > 40) or (ri2 is not None and ri2 < 70):
            out.append({"action": "Verify 4T4R/antenna-port configuration, RRU branch alarms, VSWR/feeder, cross-polar imbalance and calibration.",
                        "rationale": f"Median rank {f(rank)}, RI1={f(ri1)}%, RI2={f(ri2)}% → spatial multiplexing under-performs."})
            out.append({"action": "Review RI/PMI/CQI reporting and rank-adaptation parameters; audit overlap/scattering on the serving sequence.",
                        "rationale": "Persistent low rank despite usable SINR points to MIMO config, not pure RF."})
    elif domain == "Modulation & MCS":
        qam256 = n(iam.get("qam256")); qam16 = n(iam.get("qam16"))
        if (qam256 is not None and qam256 == 0) or (qam16 is not None and qam16 > 50):
            out.append({"action": "Verify 256QAM activation/UE capability/vendor feature; check MCS table and PDSCH power.",
                        "rationale": f"256QAM share {f(qam256)}% and 16QAM share {f(qam16)}% → modulation profile too low."})
            out.append({"action": "Improve SINR/CQI before expecting higher 64QAM/256QAM usage.",
                        "rationale": "Modulation order is gated by radio quality."})
    elif domain == "PRB / scheduler efficiency":
        prb_eff = n(iam.get("prbEfficiency")); util = n(iam.get("prbUtilPct"))
        best_eff = best("prbEfficiency", True)
        if (prb_eff is not None and best_eff is not None and prb_eff < best_eff) or (util is not None and util < 60):
            out.append({"action": "Review scheduler weights / proportional-fair / QoS profile, MCS table, OLLA, CQI aging and PDSCH power.",
                        "rationale": f"PRB efficiency {f(prb_eff)} vs best {f(best_eff)}, PRB utilization {f(util)}% → spectral efficiency below benchmark."})
            out.append({"action": "Compare PRB utilization vs delivered throughput and CA/MIMO contribution against the best-performing cells.",
                        "rationale": "Separates 'not enough PRBs' from 'PRBs used inefficiently'."})
    elif domain == "Carrier aggregation":
        scells = n(iam.get("scellsAvg")); active = n(iam.get("scellsActive"))
        if (scells is not None and scells < 1) or (active is not None and active < 50):
            out.append({"action": "Audit CA combinations and UE capability matching, SCell activation thresholds/timers (A2/A4/A6), SCell coverage and load balancing.",
                        "rationale": f"Avg #SCells {f(scells)}, SCell-active share {f(active)}% → CA rarely persistent."})
    elif domain == "BLER / retransmissions":
        p90 = n(iam.get("blerP90")); above10 = n(iam.get("blerAbove10")); ul = n(iam.get("ulRetx"))
        if (p90 is not None and p90 > 15) or (above10 is not None and above10 > 10):
            out.append({"action": "Map BLER>10% samples by GPS/time/serving-cell and correlate with SINR, MCS, modulation, HARQ and PDSCH power.",
                        "rationale": f"BLER P90 {f(p90)}% and >10% share {f(above10)}% → localized retransmission peaks, not a route-wide floor."})
            out.append({"action": "Tune OLLA/MCS aggressiveness and PDSCH power at the worst segments after RF cleanup.",
                        "rationale": "Peaks usually map to specific interference/overshoot spots."})
        if ul is not None and ul > 1:
            out.append({"action": "Check UL interference/RTWP on the LTE anchor, PUSCH power control and antenna-branch health.",
                        "rationale": f"UL retransmission {f(ul)}% can throttle TCP ramp-up and ACK feedback."})
    elif domain == "Transport / core":
        tcp = n(iam.get("tcpHandshake")); best_tcp = best("tcpHandshake", False); lost = n(iam.get("lostPacket"))
        if (tcp is not None and tcp > 70) or (lost is not None and lost > 0):
            out.append({"action": "Retest with the same UE/SIM/server/script; if confirmed, check backhaul latency/jitter/loss, APN/DNS path, firewall/NAT and core routing.",
                        "rationale": f"TCP handshake {f(tcp)} ms (best competitor {f(best_tcp)} ms), lost packets {f(lost)} → transport/core path to validate."})
    elif domain == "5G / EN-DC capacity layer":
        n78 = n(iam.get("n78")); fiveg = n(iam.get("fivegPresence"))
        if (n78 is not None and n78 < 30) or (fiveg is not None and fiveg < 40):
            out.append({"action": "Analyze EN-DC addition/release during the active DL window; verify n78 deployment/eligibility, NR neighbors, NSA addition thresholds and LTE-anchor relation.",
                        "rationale": f"5G presence {f(fiveg)}%, n78 share {f(n78)}% → the high-capacity NR layer is missing or ineffective during transfer."})
    elif domain == "Throughput delivery chain":
        ratio = n(iam.get("transportRatio"))
        if ratio is not None and ratio < 0.8:
            out.append({"action": "Localize where throughput is lost (App vs MAC vs PDSCH) on aligned active windows before assigning RAN vs transport actions.",
                        "rationale": f"App/MAC ratio {f(ratio)} indicates delivery loss above the MAC layer."})
    return out


def _deep_build_detailed_analysis(dataset, iam_operator_file=None, scope_label=None):
    """Per-domain professional diagnostics for IAM (benchmarked vs Orange/INWI) using the
    full KPI surface, the already-computed dataset builders, and raw-TXT localization.
    Returns a list of domain blocks: {domain, severity, summary, metrics, explanation,
    evidence, blerEvents, perBand, targetedActions}. Returns [] if IAM is absent."""
    if not isinstance(dataset, dict):
        return []
    operators = dataset.get("operators") or []
    transfer_lookup = {}
    for entry in dataset.get("transferSummary") or []:
        op = str(entry.get("operator") or "").upper()
        direction = str(entry.get("direction") or "")
        dnorm = "DL" if ("down" in direction.lower() or direction.upper() in ("DL", "DOWNLINK")) else \
                ("UL" if ("up" in direction.lower() or direction.upper() in ("UL", "UPLINK")) else direction.upper())
        transfer_lookup[(op, dnorm)] = entry
    iam_entry = next((o for o in operators if str(o.get("operator") or "").upper() in _DEEP_IAM_ALIASES), None)
    if not iam_entry:
        return []
    iam = _deep_extract(iam_entry.get("kpis"), transfer_lookup, iam_entry.get("operator"))
    by_op = {}
    competitors = []
    for o in operators:
        name = str(o.get("operator") or "").upper()
        if name in _DEEP_IAM_ALIASES:
            continue
        ex = _deep_extract(o.get("kpis"), transfer_lookup, o.get("operator"))
        by_op[name] = ex
        competitors.append(ex)
    orange = by_op.get("ORANGE")
    inwi = by_op.get("INWI")
    serving = dataset.get("iamServingCells")

    f = _deep_fmt
    n = _deep_num

    def best(key, higher=True):
        return _deep_best_competitor([c.get(key) for c in competitors], higher_is_better=higher)

    def M(label, key, unit="", interp=""):
        return {
            "label": label, "unit": unit, "interpretation": interp,
            "iam": n(iam.get(key)),
            "orange": n((orange or {}).get(key)),
            "inwi": n((inwi or {}).get(key)),
        }

    def gap_sev(key, higher=True, floor=0.10):
        g = _deep_gap_fraction(iam.get(key), best(key, higher), higher)
        if g is None:
            return "OK"
        s = _deep_classify_gap(g, higher)
        return s

    blocks = []
    best_dl = best("dlThroughput", True)
    dl = n(iam.get("dlThroughput"))
    dl_gap = _deep_gap_fraction(dl, best_dl, True)

    # 1. Throughput delivery chain
    ratio = n(iam.get("transportRatio"))
    sev = _deep_classify_gap(dl_gap, True)
    expl = (
        f"IAM delivers {f(dl)} Mbps versus {f(best_dl)} Mbps for the best competitor"
        + (f" ({f(abs(round((dl_gap or 0)*100,1)))}% gap)" if dl_gap is not None else "")
        + ". App/MAC ratio is "
        + (f"{f(ratio)}" if ratio is not None else "n/a")
        + (" — delivery loss sits above the MAC layer (transport/app)." if (ratio is not None and ratio < 0.8) else " — MAC delivery roughly matches the application layer.")
    )
    blocks.append({
        "domain": "Throughput delivery chain", "severity": sev, "summary": "Where the DL throughput is won or lost (App → MAC → PDSCH).",
        "metrics": [
            M("App DL throughput", "dlThroughput", "Mbps"),
            M("Total MAC DL", "totalMacDlAvg", "Mbps"),
            M("MAC DL (5G)", "mac5gAvg", "Mbps"),
            M("PDSCH DL (5G)", "pdschDlAvg", "Mbps"),
            M("App/MAC ratio", "transportRatio", ""),
        ],
        "explanation": expl, "evidence": [], "targetedActions": _deep_domain_targeted_actions("Throughput delivery chain", iam, best),
    })

    # 2. 5G / EN-DC capacity layer
    dlw_share, dlw_nr_sec, dlw_total = _deep_download_window_nr_share(serving)
    n78 = n(iam.get("n78"))
    fiveg = n(iam.get("fivegPresence"))
    cap_sev = "Critical" if (n78 is not None and n78 < 5 and (dl_gap or 0) > 0.3) else ("High" if (n78 is not None and n78 < 30) else "Low")
    per_band = []
    nbe = dataset.get("nrBandExposureAnalysis") or {}
    for r in (nbe.get("rows") or []):
        if str(r.get("operator") or "").upper() in _DEEP_IAM_ALIASES:
            per_band = r.get("bands") or r.get("perBand") or []
            break
    cap_expl = (
        f"Global 5G presence {f(fiveg)}%, n78 share {f(n78)}% (NR-only), n1 {f((iam.get('nrBandShares') or {}).get('n1'))}%. "
        + (f"During the active DL window only {f(dlw_share)}% was on 5G ({f(dlw_nr_sec)}s/{f(dlw_total)}s). " if dlw_share is not None else "")
        + ("The 5G coverage layer exists but the high-capacity n78 layer is missing/ineffective." if (n78 is not None and n78 < 30) else "5G capacity layer is present.")
    )
    cap_evidence = []
    seq = _deep_supporting_cells(serving)
    if seq:
        cap_evidence.append("Serving cells (download): " + " → ".join(seq))
    blocks.append({
        "domain": "5G / EN-DC capacity layer", "severity": cap_sev, "summary": "Effective 5G contribution and high-capacity n78 availability during the transfer.",
        "metrics": [
            M("5G presence", "fivegPresence", "%"),
            M("4G-only time", "fourgOnly", "%"),
            M("NR n78 share", "n78", "%"),
            M("NR n1 share", "n1", "%"),
            M("NR n28 share", "n28", "%"),
        ],
        "explanation": cap_expl, "evidence": cap_evidence, "perBand": per_band,
        "downloadWindow": {"share": dlw_share, "nrSec": dlw_nr_sec, "totalSec": dlw_total},
        "targetedActions": _deep_domain_targeted_actions("5G / EN-DC capacity layer", iam, best),
    })

    # 3. RF quality
    sinr = n(iam.get("sinr")); rsrp = n(iam.get("rsrpMedian"))
    rf_sev = "Critical" if (sinr is not None and sinr < 0) else ("High" if (sinr is not None and sinr < 5) else ("Medium" if (sinr is not None and sinr < 10) else "OK"))
    rf_expl = (
        f"RSRP {f(rsrp)} dBm, SINR {f(sinr)} dB (p10 {f(iam.get('sinrP10'))}), CQI {f(iam.get('cqiMedian'))}. "
        + ("Coverage is adequate but quality is interference-limited." if (rsrp is not None and rsrp >= -100 and sinr is not None and sinr < 8) else
           ("Both coverage and quality need work." if (rsrp is not None and rsrp < -105) else "RF quality is acceptable."))
        + (f" LTE-anchor SINR {f(iam.get('lteAnchorSinr'))} dB." if iam.get("lteAnchorSinr") is not None else "")
    )
    blocks.append({
        "domain": "RF quality", "severity": rf_sev, "summary": "Coverage vs interference (RSRP/RSRQ/SINR/CQI + LTE anchor).",
        "metrics": [
            M("Median RSRP", "rsrpMedian", "dBm"),
            M("Median RSRQ", "rsrqMedian", "dB"),
            M("Median SINR", "sinr", "dB"),
            M("SINR p10", "sinrP10", "dB"),
            M("Median CQI", "cqiMedian", ""),
            M("LTE-anchor SINR", "lteAnchorSinr", "dB"),
        ],
        "explanation": rf_expl, "evidence": [], "targetedActions": _deep_domain_targeted_actions("RF quality", iam, best),
    })

    # 4. MIMO / spatial multiplexing
    rank = n(iam.get("medianRank")); ri1 = n(iam.get("ri1")); ri2 = n(iam.get("ri2"))
    mimo_sev = "Critical" if (rank is not None and rank <= 1) else ("High" if (ri1 is not None and ri1 > 40) else ("Medium" if (ri2 is not None and ri2 < 70) else "OK"))
    mimo_expl = (
        f"Median rank {f(rank)}, RI1={f(ri1)}%, RI2={f(ri2)}%, RI≥3={f(iam.get('riGe3'))}%, scheduled rank {f(iam.get('scheduledRankMedian'))}. "
        + ("Spatial multiplexing is weak — mostly single-stream." if (rank is not None and rank <= 1) or (ri1 is not None and ri1 > 40) else "MIMO rank is acceptable.")
    )
    blocks.append({
        "domain": "MIMO / spatial multiplexing", "severity": mimo_sev, "summary": "Rank distribution and spatial-stream usage.",
        "metrics": [
            M("Median rank", "medianRank", ""),
            M("RI1 share", "ri1", "%"),
            M("RI2 share", "ri2", "%"),
            M("RI≥3 share", "riGe3", "%"),
            M("Scheduled rank", "scheduledRankMedian", ""),
        ],
        "explanation": mimo_expl, "evidence": [], "targetedActions": _deep_domain_targeted_actions("MIMO / spatial multiplexing", iam, best),
    })

    # 5. Modulation & MCS
    qam256 = n(iam.get("qam256")); qam16 = n(iam.get("qam16"))
    mod_sev = "High" if (qam256 is not None and qam256 == 0 and qam16 is not None and qam16 > 50) else ("Medium" if (qam16 is not None and qam16 > 40) else "OK")
    mod_expl = (
        f"Modulation mix QPSK {f(iam.get('qpsk'))}% / 16QAM {f(qam16)}% / 64QAM {f(iam.get('qam64'))}% / 256QAM {f(qam256)}%. "
        f"MCS cw0 {f(iam.get('mcsCw0Median'))}, cw1 {f(iam.get('mcsCw1Median'))}; bit/s/Hz {f(iam.get('pdschBitPerHz'))} of max {f(iam.get('pdschMaxBitPerHz'))}. "
        + ("Throughput is modulation-limited by radio quality/scheduler." if (qam256 is not None and qam256 == 0) else "Modulation profile is reasonable.")
    )
    blocks.append({
        "domain": "Modulation & MCS", "severity": mod_sev, "summary": "Modulation distribution, per-codeword MCS and spectral efficiency.",
        "metrics": [
            M("256QAM share", "qam256", "%"),
            M("64QAM share", "qam64", "%"),
            M("16QAM share", "qam16", "%"),
            M("QPSK share", "qpsk", "%"),
            M("MCS cw0", "mcsCw0Median", ""),
            M("PDSCH bit/s/Hz", "pdschBitPerHz", ""),
        ],
        "explanation": mod_expl, "evidence": [], "targetedActions": _deep_domain_targeted_actions("Modulation & MCS", iam, best),
    })

    # 6. PRB / scheduler efficiency
    prb_eff = n(iam.get("prbEfficiency")); util = n(iam.get("prbUtilPct"))
    prb_sev = gap_sev("prbEfficiency", True)
    if util is not None and util < 50 and (dl_gap or 0) > 0.3:
        prb_sev = "High"
    prb_expl = (
        f"PRB utilization {f(util)}%, PRBs {f(iam.get('prbsAvg'))}/{f(iam.get('availableBandwidthPrbs'))} available, "
        f"PRB efficiency {f(prb_eff)} (best {f(best('prbEfficiency', True))}), slot {f(iam.get('pdschSlotPct'))}%, "
        f"scheduled→delivered {f(iam.get('scheduledEfficiency'))}%. "
        + ("Spectral/scheduler efficiency is below the benchmark." if (prb_eff is not None and (best('prbEfficiency', True) or 0) and prb_eff < (best('prbEfficiency', True) or 0)) else "Scheduler efficiency is in line.")
    )
    blocks.append({
        "domain": "PRB / scheduler efficiency", "severity": prb_sev, "summary": "Available → allocated → delivered PRB chain and spectral efficiency.",
        "metrics": [
            M("PRB utilization", "prbUtilPct", "%"),
            M("Avg PDSCH PRBs", "prbsAvg", ""),
            M("Available BW PRBs", "availableBandwidthPrbs", ""),
            M("PRB efficiency", "prbEfficiency", "Mbps/PRB"),
            M("Resource alloc index", "resourceAllocationIndex", "%"),
            M("Scheduled efficiency", "scheduledEfficiency", "%"),
        ],
        "explanation": prb_expl, "evidence": [], "targetedActions": _deep_domain_targeted_actions("PRB / scheduler efficiency", iam, best),
    })

    # 7. Carrier aggregation
    scells = n(iam.get("scellsAvg")); active = n(iam.get("scellsActive"))
    ca_sev = "High" if (scells is not None and scells < 0.5) else ("Medium" if (active is not None and active < 50) else "OK")
    ca_expl = (
        f"Avg #SCells {f(scells)} (max {f(iam.get('scellsMax'))}), SCell-active share {f(active)}%, LTE CA active {f(iam.get('caActive'))}%. "
        + ("Carrier aggregation is rarely persistent." if (scells is not None and scells < 1) else "CA depth is reasonable.")
    )
    blocks.append({
        "domain": "Carrier aggregation", "severity": ca_sev, "summary": "Secondary-carrier depth and activation persistence.",
        "metrics": [
            M("Avg #SCells", "scellsAvg", ""),
            M("Max #SCells", "scellsMax", ""),
            M("SCell-active share", "scellsActive", "%"),
            M("LTE CA active", "caActive", "%"),
        ],
        "explanation": ca_expl, "evidence": [], "targetedActions": _deep_domain_targeted_actions("Carrier aggregation", iam, best),
    })

    # 8. BLER / retransmissions (with raw-TXT localization)
    p90 = n(iam.get("blerP90")); above10 = n(iam.get("blerAbove10"))
    bler_sev = "Critical" if (n(iam.get("blerAvg")) or 0) > 10 else ("High" if ((p90 or 0) > 15 or (above10 or 0) > 10) else ("Medium" if (above10 or 0) > 5 else "OK"))
    bler_events = _deep_bler_localization_rows(iam_operator_file) if iam_operator_file else []
    bler_expl = (
        f"BLER avg {f(iam.get('blerAvg'))}%, P90 {f(p90)}%, >10% share {f(above10)}%, >20% share {f(iam.get('blerAbove20'))}%, UL retx {f(iam.get('ulRetx'))}%. "
        + (f"{len(bler_events)} BLER>10% sample(s) localized by time/GPS/cell. " if bler_events else "")
        + ("Moderate average with high P90 → localized retransmission peaks, not a route-wide floor." if ((p90 or 0) > 15 and (n(iam.get('blerAvg')) or 0) <= 8) else "")
    )
    blocks.append({
        "domain": "BLER / retransmissions", "severity": bler_sev, "summary": "Retransmission peaks and their localization.",
        "metrics": [
            M("BLER avg", "blerAvg", "%"),
            M("BLER P90", "blerP90", "%"),
            M("BLER >10% share", "blerAbove10", "%"),
            M("BLER >20% share", "blerAbove20", "%"),
            M("UL retx", "ulRetx", "%"),
        ],
        "explanation": bler_expl, "evidence": [], "blerEvents": bler_events,
        "targetedActions": _deep_domain_targeted_actions("BLER / retransmissions", iam, best),
    })

    # 9. Transport / core
    tcp = n(iam.get("tcpHandshake"))
    tr_sev = "Medium" if ((tcp or 0) > 70 or (n(iam.get("lostPacket")) or 0) > 0) else "OK"
    tr_expl = (
        f"TCP handshake {f(tcp)} ms (best competitor {f(best('tcpHandshake', False))} ms), lost packets {f(iam.get('lostPacket'))}. "
        + ("Validate transport/core only after RAN causes are separated." if (tcp or 0) > 70 else "Transport latency is in line.")
    )
    blocks.append({
        "domain": "Transport / core", "severity": tr_sev, "summary": "App-layer responsiveness vs RAN.",
        "metrics": [
            M("TCP handshake", "tcpHandshake", "ms"),
            M("Lost packets", "lostPacket", ""),
        ],
        "explanation": tr_expl, "evidence": [], "targetedActions": _deep_domain_targeted_actions("Transport / core", iam, best),
    })

    return blocks


def _deep_build_simplified_action_plan(iam, competitors, iam_serving_cells, scope_label=None):
    """Concise, execution-oriented IAM Action Plan: at most 5 rows, columns
    Priority | Domain | Finding | Action | Confidence. Each row is emitted only when its
    trigger fires. Duplicate 5G topics (Mobility / 5G capacity / Bandwidth-spectrum /
    Capacity-configuration / n78-missing / LTE-dominated DL window) are folded into the
    single P1 "Active DL / 5G capacity layer" row; MIMO is never a separate row (only a
    note when rank is healthy). Driven by the IAM tuning spec — no evidence/owner/
    validation blocks, no repetition."""
    iam = iam or {}
    competitors = competitors or []

    def best(key, higher=True):
        return _deep_best_competitor([c.get(key) for c in competitors], higher_is_better=higher)

    def pct_or_default(value, default="100"):
        txt = _deep_fmt(value)
        return txt if txt != "" else default

    rows = []

    dl = _deep_num(iam.get("dlThroughput"))
    best_dl = best("dlThroughput", True)
    dl_gap_pct = ((best_dl - dl) / best_dl * 100.0) if (dl is not None and best_dl not in (None, 0)) else None
    fiveg = _deep_num(iam.get("fivegPresence"))
    fourg_only = _deep_num(iam.get("fourgOnly"))
    n78 = _deep_num(iam.get("n78"))
    sinr = _deep_num(iam.get("sinr"))
    rank = _deep_num(iam.get("medianRank"))
    ri1 = _deep_num(iam.get("ri1"))
    dl_success = _deep_num(iam.get("dlSuccess"))
    dl_completion = _deep_num(iam.get("dlCompletion"))
    dlw_share, dlw_nr_sec, dlw_total_sec = _deep_download_window_nr_share(iam_serving_cells)
    mimo_ok = rank is not None and rank >= 2 and ri1 is not None and ri1 <= 10

    # ── P1 — Active DL / 5G capacity layer (merges Mobility / n78 / bandwidth / capacity) ──
    # Only flag when IAM is genuinely behind the best competitor (a capacity action makes
    # no sense when IAM already leads), AND a 5G-weakness signal is present. The download
    # window is authoritative: "LTE-dominated" is only asserted when that window is really
    # low-5G — avoids contradictory findings on DTs where the active DL was carried by 5G.
    dl_deficit = dl_gap_pct is not None and dl_gap_pct > 30
    low_5g_window = (
        (dlw_share is not None and dlw_share < 30)
        or (dlw_share is None and fiveg is not None and fiveg == 0)
    )
    has_5g_weakness = (
        low_5g_window
        or (fiveg is not None and fiveg == 0)
        or (fourg_only is not None and fourg_only > 70)
        or (n78 is not None and n78 == 0)
    )
    trigger_p1 = dl_deficit and has_5g_weakness
    if trigger_p1:
        if dlw_share is not None and dlw_total_sec is not None:
            window_phrase = (
                f"{_deep_fmt(dlw_share)}% 5G during the active DL window "
                f"({_deep_fmt(dlw_nr_sec if dlw_nr_sec is not None else 0)}s over {_deep_fmt(dlw_total_sec)}s)"
            )
        elif fiveg is not None and fiveg == 0:
            window_phrase = "0% 5G during the active DL window"
        else:
            window_phrase = "low 5G contribution during the active DL window"
        if low_5g_window:
            lead = f"IAM DL session is LTE-dominated: {window_phrase}."
        else:
            lead = f"IAM active DL window shows {window_phrase}, but DL throughput stays capacity-limited."
        finding = (
            f"{lead} "
            f"DL throughput is {_deep_fmt(dl)} Mbps versus {_deep_fmt(best_dl)} Mbps for the best competitor, "
            f"despite {pct_or_default(dl_success)}% DL success and {pct_or_default(dl_completion)}% DL completion."
        )
        if fiveg is not None and fiveg == 0:
            finding += " No 5G was observed during the test."
        if n78 is not None and n78 == 0:
            finding += " No n78 contribution was observed."
        if mimo_ok:
            finding += " MIMO rank is acceptable and not a primary limitation."
        action = (
            "Analyze EN-DC addition/release during the DL transfer. Verify LTE anchor behavior, "
            "NR PSCell availability, n78 deployment/eligibility, NR neighbors, NSA thresholds, and "
            "serving-cell sequence during the active DL window."
        )
        rows.append(_deep_simple_row(
            "P1", "Active DL / 5G capacity layer", finding, action,
            "Confirmed symptom / Partial root cause",
        ))

    # ── P2 — BLER / retransmissions ──
    bler_avg = _deep_num(iam.get("blerAvg"))
    bler_p90 = _deep_num(iam.get("blerP90"))
    bler_above10 = _deep_num(iam.get("blerAbove10"))
    trigger_bler = (
        (bler_p90 is not None and bler_p90 > 15)
        or (bler_above10 is not None and bler_above10 > 10)
        or (bler_avg is not None and bler_avg > 5)
    )
    if trigger_bler:
        finding = (
            f"IAM shows localized retransmission peaks: BLER Avg={_deep_fmt(bler_avg)}%, "
            f"P90={_deep_fmt(bler_p90)}%, BLER>10% share={_deep_fmt(bler_above10)}%."
        )
        action = (
            "Map BLER>10% samples by GPS, time, serving cell, SINR, CQI, MCS, modulation, rank, "
            "HARQ and throughput. Decide RF/interference/link-adaptation actions after localization."
        )
        rows.append(_deep_simple_row(
            "P2", "BLER / retransmissions", finding, action,
            "Confirmed symptom / Partial root cause",
        ))

    # ── P2 — Transport / core ──
    tcp = _deep_num(iam.get("tcpHandshake"))
    best_tcp = best("tcpHandshake", False)
    trigger_tcp = tcp is not None and (tcp > 70 or (best_tcp is not None and tcp > best_tcp * 1.2))
    if trigger_tcp:
        finding = f"IAM TCP handshake is {_deep_fmt(tcp)} ms versus {_deep_fmt(best_tcp)} ms for the best competitor."
        action = (
            "Retest with same UE, SIM, server and script. If confirmed, check backhaul latency, "
            "jitter, packet loss, APN/DNS path, firewall/NAT and core routing."
        )
        rows.append(_deep_simple_row("P2", "Transport / core", finding, action, "Partial"))

    # ── P3 — Scheduler / PRB efficiency (suspected, missing data) ──
    rf_acceptable = sinr is not None and sinr >= 5
    throughput_low = dl_gap_pct is not None and dl_gap_pct > 30
    prb_incomplete = (
        iam.get("prbUtilPct") is None
        or iam.get("prbsAvg") is None
        or _deep_num(iam.get("prbEfficiency")) is None
    )
    if throughput_low and rf_acceptable and prb_incomplete:
        finding = (
            "Scheduler/PRB efficiency issue is suspected but not confirmed because PRB utilization, "
            "allocated PRBs, scheduler grants, OLLA, CQI aging, QoS profile and PDSCH power are missing or incomplete."
        )
        action = (
            "Add PRB utilization, allocated PRBs, scheduler grants, OLLA, CQI aging, QoS profile and "
            "PDSCH power to the next export before confirming scheduler action."
        )
        rows.append(_deep_simple_row("P3", "Scheduler / PRB efficiency", finding, action, "Hypothesis"))

    # ── P3 — Retest governance (single DT scope only) ──
    is_single_dt = bool(scope_label) and "All DTs" not in str(scope_label)
    if is_single_dt:
        finding = "The conclusion is based on a single DT session."
        action = (
            "Repeat the same route in busy hour and off-peak with the same UE, SIM profile, server, "
            "test script and route direction. Confirm issue stability across repeated DTs."
        )
        rows.append(_deep_simple_row("P3", "Retest governance", finding, action, "Good practice"))

    return rows[:5]


def _deep_download_context(serving_cells: dict | None) -> dict | None:
    """Localize the analysis: the dominant serving cell(s), band/tech, LTE anchor, time
    window and GPS during the active download — so findings can say *where* the benchmarked
    throughput was actually measured instead of being route-anonymous."""
    sc = serving_cells or {}
    eps = sc.get("episodesDownload") or []
    if not eps:
        return None
    top = max(eps, key=lambda e: (e.get("dwellSec") or e.get("samples") or 0))
    cell = top.get("cellName") or "?"
    lon = lat = None
    for c in sc.get("cells") or []:
        if c.get("cellName") == cell:
            lon, lat = _deep_num(c.get("lon")), _deep_num(c.get("lat"))
            break
    tech, band, anchor = top.get("tech"), top.get("band"), top.get("lteAnchor")
    win_s, win_e = sc.get("downloadWindowStart"), sc.get("downloadWindowEnd")
    cells_list = []
    for e in eps:
        cn = e.get("cellName")
        if cn and cn not in cells_list:
            cells_list.append(cn)
    label = cell + (f" ({(tech or '').strip()} {(band or '').strip()})".rstrip() if (tech or band) else "")
    if anchor:
        label += f", LTE anchor {anchor}"
    text = f"Active download served by {label}"
    if win_s and win_e:
        text += f" over {win_s}–{win_e}"
    if lon is not None and lat is not None:
        text += f" near ({lat:.5f}, {lon:.5f})"
    if len(cells_list) > 1:
        text += f"; serving sequence touched {len(cells_list)} cells"
    text += "."
    short = f"on {cell}" + (f" ({(band or tech or '').strip()})" if (band or tech) else "") + (f", {win_s}–{win_e}" if win_s and win_e else "")
    return {
        "cell": cell, "tech": tech, "band": band, "lteAnchor": anchor,
        "windowStart": win_s, "windowEnd": win_e, "lon": lon, "lat": lat,
        "cells": cells_list, "text": text, "short": short,
    }


def _deep_enrich_action_plan_with_current_data(
    deep: dict | None,
    operators_payload: list[dict] | None,
    iam_serving_cells: dict | None,
    transfer_summary: list[dict] | None,
    scope_label: str | None = None,
    iam_operator_file: dict | None = None,
) -> dict | None:
    if not isinstance(deep, dict) or not isinstance(deep.get("actionPlan"), list):
        return deep
    operators_payload = operators_payload or []
    transfer_lookup = {}
    for entry in transfer_summary or []:
        op = str(entry.get("operator") or "").upper()
        direction = str(entry.get("direction") or "")
        dnorm = "DL" if ("down" in direction.lower() or direction.upper() in ("DL", "DOWNLINK")) else \
                ("UL" if ("up" in direction.lower() or direction.upper() in ("UL", "UPLINK")) else direction.upper())
        transfer_lookup[(op, dnorm)] = entry
    iam_entry = next((o for o in operators_payload if str(o.get("operator") or "").upper() in _DEEP_IAM_ALIASES), None)
    if not iam_entry:
        return deep
    iam = _deep_extract((iam_entry or {}).get("kpis"), transfer_lookup, (iam_entry or {}).get("operator"))

    # Execution-oriented IAM Action Plan rebuilt fresh from the current-scope metrics using
    # the (improved, RAT-aware) rule engine, then rendered into the 5-column shape
    # (Priority | Domain | Finding | Action | Confidence) the UI/XLSX expect. The exec
    # summary is recomputed from the SAME findings so the headline conclusion and the action
    # plan always agree.
    competitors = []
    for o in operators_payload:
        if str(o.get("operator") or "").upper() in _DEEP_IAM_ALIASES:
            continue
        cv = _deep_extract(o.get("kpis"), transfer_lookup, o.get("operator"))
        cv["_operator"] = o.get("operator")
        competitors.append(cv)
    findings = _benchmark_deep_findings(iam, competitors)
    download_ctx = _deep_download_context(iam_serving_cells)
    # Findings that describe the download experience get localized to the active-download
    # serving cell / band / time window.
    localize_domains = {
        "Scheduler / PRB efficiency", "5G capacity layer", "Bandwidth / spectrum",
        "LTE anchor / NSA dependency", "Coverage / dominance", "Modulation profile",
        "MIMO / RI", "Radio quality / interference", "SINR / interference",
    }
    plan = []
    for f in findings:
        finding_text = str(f.get("finding") or "").strip()
        root = str(f.get("rootCause") or "").strip()
        if root and root not in finding_text:
            finding_text = (finding_text + " Root cause: " + root).strip()
        if download_ctx and f.get("domain") in localize_domains and download_ctx.get("short"):
            finding_text = finding_text + f" Observed {download_ctx['short']}."
        acts = f.get("recommendedActions")
        action_text = (
            "\n".join(f"• {str(a).strip()}" for a in acts if str(a or "").strip())
            if isinstance(acts, (list, tuple)) else (acts or "")
        )
        plan.append({
            "priority": f.get("priority") or "",
            "domain": f.get("domain") or "",
            "finding": finding_text,
            "action": action_text,
            "confidence": f.get("confidence") or "",
            "severity": f.get("severity"),
            "rootCause": f.get("rootCause"),
            "recommendedActions": f.get("recommendedActions"),
            "kpi": f.get("kpi"),
            "location": download_ctx if f.get("domain") in localize_domains else None,
        })
    clone_deep = dict(deep)
    clone_deep["actionPlan"] = plan
    exec_summary = _benchmark_deep_exec_summary(
        iam, competitors, findings, scope_label or deep.get("scopeLabel") or "Benchmark"
    )
    if download_ctx:
        exec_summary["downloadContext"] = download_ctx
        exec_summary["mainConclusion"] = (download_ctx.get("text") or "") + " " + (exec_summary.get("mainConclusion") or "")
    clone_deep["execSummary"] = exec_summary
    return clone_deep


def _benchmark_deep_analysis(operators_payload, transfer_summary, scope_label="Benchmark"):
    """Build the Deep Benchmark analysis (exec summary + KPI table + action plan) for IAM.

    `operators_payload` = the dataset's per-operator entries (each with `kpis`).
    Returns None when IAM is absent so callers can skip gracefully."""
    try:
        transfer_lookup = {}
        for entry in transfer_summary or []:
            op = str(entry.get("operator") or "").upper()
            direction = str(entry.get("direction") or "")
            dnorm = "DL" if ("down" in direction.lower() or direction.upper() in ("DL", "DOWNLINK")) else \
                    ("UL" if ("up" in direction.lower() or direction.upper() in ("UL", "UPLINK")) else direction.upper())
            transfer_lookup[(op, dnorm)] = entry

        iam_entry = next((o for o in operators_payload if str(o.get("operator") or "").upper() in _DEEP_IAM_ALIASES), None)
        if not iam_entry:
            return None
        iam = _deep_extract(iam_entry.get("kpis"), transfer_lookup, iam_entry.get("operator"))
        competitors = []
        orange = inwi = None
        for o in operators_payload:
            name = str(o.get("operator") or "").upper()
            if name in _DEEP_IAM_ALIASES:
                continue
            vals = _deep_extract(o.get("kpis"), transfer_lookup, o.get("operator"))
            vals["_operator"] = o.get("operator")
            competitors.append(vals)
            if name == "ORANGE":
                orange = vals
            elif name == "INWI":
                inwi = vals
        if not competitors:
            return None

        kpi_rows = _benchmark_deep_kpi_rows(iam, orange, inwi)
        findings = _benchmark_deep_findings(iam, competitors)
        exec_summary = _benchmark_deep_exec_summary(iam, competitors, findings, scope_label)
        return {
            "scopeLabel": scope_label,
            "target": iam_entry.get("operator") or "IAM",
            "generatedAt": _nemo_iso(_dt.now()),
            "execSummary": exec_summary,
            "kpiBenchmark": kpi_rows,
            "actionPlan": findings,
        }
    except Exception:
        import traceback as _tb
        _tb.print_exc()
        return None


def _deep_export_analysis_row(dataset: dict, analysis_key: str, operator_name: str) -> dict:
    rows = ((dataset or {}).get(analysis_key) or {}).get("rows") or []
    wanted = str(operator_name or "").upper()
    for row in rows:
        if str((row or {}).get("operator") or "").upper() == wanted:
            return row or {}
    return {}


def _deep_export_delta(iam, other, unit="", digits: int = 1):
    iam_n, other_n = _deep_num(iam), _deep_num(other)
    if iam_n is None or other_n is None:
        return ""
    if unit == "%":
        if other_n == 0:
            return ""
        value = round((iam_n - other_n) / abs(other_n) * 100.0, 1)
        text = f"{value:+.1f}%"
        return text.replace("+0.0%", "0%").replace("-0.0%", "0%")
    diff = round(iam_n - other_n, digits)
    if abs(diff - round(diff)) < 1e-9:
        diff_text = str(int(round(diff)))
    else:
        diff_text = str(diff)
    if not diff_text.startswith("-"):
        diff_text = "+" + diff_text
    if unit:
        diff_text += f" {unit}"
    return diff_text.replace("+0", "0", 1) if diff_text in ("+0", "+0 pp", "+0 dB", "+0 ms") else diff_text


def _deep_export_extract_operator(dataset: dict, operator_entry: dict, transfer_lookup: dict) -> dict:
    op_name = str((operator_entry or {}).get("operator") or "").upper()
    kpis = (operator_entry or {}).get("kpis") or {}
    mod = kpis.get("pdschModulation") or {}
    band = _deep_export_analysis_row(dataset, "nrBandExposureAnalysis", op_name)
    mimo = _deep_export_analysis_row(dataset, "mimoRankAnalysis", op_name)
    bler = _deep_export_analysis_row(dataset, "blerRetxAnalysis", op_name)
    transport = _deep_export_analysis_row(dataset, "transportGapAnalysis", op_name)
    ca = _deep_export_analysis_row(dataset, "caScellsAnalysis", op_name)
    dl_tr = transfer_lookup.get((op_name, "DL")) or {}
    ping_success = transport.get("pingSuccessRate")
    return {
        "_operator": operator_entry.get("operator") or op_name,
        "dlThroughput": _deep_num((kpis.get("dl") or {}).get("average")),
        "rsrp": _deep_num((kpis.get("rsrp") or {}).get("median")),
        "sinr": _deep_num((kpis.get("sinr") or {}).get("median")),
        "cqi": _deep_num((kpis.get("cqi") or {}).get("median")),
        "fivegPresence": _deep_num(kpis.get("nrPresencePct")),
        "fourgOnly": _deep_num(kpis.get("lteOnlyPresencePct")),
        "n78": _deep_num(band.get("n78Share")),
        "n28": _deep_num(band.get("n28Share")),
        "qam256": _deep_num(mod.get("qam256Share")),
        "qam64": _deep_num(mod.get("qam64Share")),
        "qam16": _deep_num(mod.get("qam16Share")),
        "qpsk": _deep_num(mod.get("qpskShare")),
        "medianRank": _deep_num((kpis.get("scheduledRank") or {}).get("median")),
        "ri1": _deep_num(mimo.get("ri1Share")),
        "ri2": _deep_num(mimo.get("ri2Share")),
        "riGe3": _deep_num(mimo.get("riGe3Share")),
        "scellsAvg": _deep_num(ca.get("avgScells")),
        "scellsMax": _deep_num(ca.get("maxScells")),
        "scellsActive": _deep_num(ca.get("scellsActiveShare")),
        "lteCaActive": _deep_num(ca.get("lteCaActiveShare")),
        "prbEfficiency": _deep_num((operator_entry.get("kpis") or {}).get("prbEfficiency")),
        "blerAvg": _deep_num(bler.get("blerAvg")),
        "blerP90": _deep_num(bler.get("blerP90")),
        "blerAbove10": _deep_num(bler.get("blerGt10Share")),
        "ulRetxAvg": (
            _deep_num(transport.get("ulRetxAvg"))
            if transport.get("ulRetxAvg") is not None
            else _deep_num((operator_entry.get("kpis") or {}).get("macUlRetx", {}).get("average"))
        ),
        "pdschDlAvg": _deep_num(transport.get("pdschDlAvg")),
        "tcpHandshake": _deep_num(transport.get("tcpHandshakeMedian")),
        "pingSuccessRate": _deep_num(ping_success),
        "dlCompletion": _deep_num(dl_tr.get("avgCompletionPct")),
        "dlSuccess": _deep_num(dl_tr.get("successRate")),
    }


def _deep_export_quantize(value, digits: int | None):
    num = _deep_num(value)
    if num is None:
        return None
    if digits is None:
        return num
    if digits <= 0:
        return float(int(round(num)))
    return round(num, digits)


def _deep_export_sequence_names(sequence_cells: list[str]) -> tuple[str, str]:
    if not sequence_cells:
        return ("serving LTE anchor and NR", "serving LTE anchor and NR")
    compact = []
    previous = None
    for name in sequence_cells[:3]:
        text = str(name or "").strip()
        if not text:
            continue
        if previous:
            prev_match = re.match(r"^(.*?)(\d+)$", previous)
            cur_match = re.match(r"^(.*?)(\d+)$", text)
            if prev_match and cur_match and prev_match.group(1) == cur_match.group(1):
                compact.append(cur_match.group(2))
            else:
                compact.append(text)
        else:
            compact.append(text)
        previous = text
    slash_text = "/".join(compact)
    if len(compact) >= 3:
        slash_text = "/".join(compact[:2]) + " and " + compact[2]
    return (", ".join(compact), slash_text)


def _deep_export_title(dataset: dict, deep: dict | None = None) -> str:
    dt_list = (dataset or {}).get("dtList") or []
    source_files = (dataset or {}).get("sourceFiles") or []
    city = ""
    prefixes = []
    for item in source_files:
        file_name = str((item or {}).get("fileName") or "")
        if "-" not in file_name:
            continue
        prefixes.append(file_name.split("-", 1)[0].strip())
    if prefixes and len({p.upper() for p in prefixes if p}) == 1:
        city = prefixes[0]
    if len(dt_list) == 1:
        label = str((dt_list[0] or {}).get("label") or "")
        match = re.search(r"DT\s*(\d+)", label, re.IGNORECASE)
        if city and match:
            return f"{city} DT{match.group(1)} - IAM Professional Analysis"
    if deep and (deep.get("execSummary") or {}).get("title"):
        return (deep.get("execSummary") or {}).get("title")
    return "IAM Professional Analysis"


def _deep_export_top_kpis(iam: dict, orange: dict | None, inwi: dict | None) -> list[dict]:
    top_kpis = []
    if iam.get("dlThroughput") is not None and orange and orange.get("dlThroughput") is not None:
        gap = _deep_export_delta(iam.get("dlThroughput"), orange.get("dlThroughput"), "%")
        gap = gap.replace("+", "")
        if gap.startswith("-"):
            gap = f"IAM is {gap[1:]} lower"
        else:
            gap = f"IAM is {gap} higher"
        top_kpis.append({
            "kpi": "DL Throughput",
            "iam": f"{_deep_fmt(iam.get('dlThroughput'))} Mbps",
            "bestCompetitor": f"Orange {_deep_fmt(orange.get('dlThroughput'))} Mbps",
            "gap": gap,
        })
    top_kpis.append({
        "kpi": "5G layer",
        "iam": f"{_deep_fmt(iam.get('fivegPresence'))}% 5G presence, {_deep_fmt(iam.get('n28'))}% n28",
        "bestCompetitor": f"Orange: {_deep_fmt((orange or {}).get('fivegPresence'))}% 5G, {_deep_fmt((orange or {}).get('n78'))}% n78",
        "gap": "IAM lacks n78 capacity on this DT",
    })
    top_kpis.append({
        "kpi": "MIMO rank",
        "iam": f"Median RI={_deep_fmt(iam.get('medianRank'))}, RI1={_deep_fmt(iam.get('ri1'))}%",
        "bestCompetitor": f"Orange: Median RI={_deep_fmt((orange or {}).get('medianRank'))}, RI2={_deep_fmt((orange or {}).get('ri2'))}%",
        "gap": "Weak spatial multiplexing",
    })
    top_kpis.append({
        "kpi": "SINR/CQI/MCS",
        "iam": f"SINR {_deep_fmt(iam.get('sinr'))} dB, CQI {_deep_fmt(iam.get('cqi'))}, MCS {_deep_fmt((iam or {}).get('mcs'))}",
        "bestCompetitor": f"Orange CQI {_deep_fmt((orange or {}).get('cqi'))}",
        "gap": "Low quality keeps modulation mainly 16QAM",
    })
    top_kpis.append({
        "kpi": "BLER",
        "iam": f"Avg {_deep_fmt(iam.get('blerAvg'))}%, P90 {_deep_fmt(iam.get('blerP90'))}%, >10%={_deep_fmt(iam.get('blerAbove10'))}%",
        "bestCompetitor": f"Orange avg {_deep_fmt((orange or {}).get('blerAvg'))}%, P90 {_deep_fmt((orange or {}).get('blerP90'))}%",
        "gap": "High retransmission zones",
    })
    top_kpis.append({
        "kpi": "Transport",
        "iam": f"TCP handshake {_deep_fmt(iam.get('tcpHandshake'))} ms",
        "bestCompetitor": f"INWI {_deep_fmt((inwi or {}).get('tcpHandshake'))} ms / Orange {_deep_fmt((orange or {}).get('tcpHandshake'))} ms",
        "gap": "Core/transport path to verify",
    })
    return top_kpis


def _deep_export_normalize_action_plan(rows):
    """Carry the action-plan rows into the export, guaranteeing the five canonical columns
    (priority / domain / finding / action / confidence) are present so the XLSX matches the
    live webapp. Bridges the rule-engine finding shape (rootCause + recommendedActions list)
    into the simplified shape (finding incl. root cause, action as newline-joined text)."""
    normalized = []
    for row in rows or []:
        clone = dict(row or {})
        clone["priority"] = row.get("priority") or ""
        clone["domain"] = row.get("domain") or ""
        finding_text = str(row.get("finding") or "").strip()
        root = str(row.get("rootCause") or "").strip()
        if root and root not in finding_text:
            finding_text = (finding_text + " Root cause: " + root).strip()
        clone["finding"] = finding_text
        action = row.get("action")
        if not action:
            acts = row.get("recommendedActions")
            if isinstance(acts, (list, tuple)):
                action = "\n".join(f"• {str(a).strip()}" for a in acts if str(a or "").strip())
            else:
                action = acts or ""
        clone["action"] = action
        clone["confidence"] = row.get("confidence") or ""
        normalized.append(clone)
    return normalized


def _benchmark_deep_export_model(dataset: dict, deep: dict | None = None) -> dict | None:
    if not isinstance(dataset, dict):
        return None
    if deep:
        exec_summary = dict(deep.get("execSummary") or {})
        if not exec_summary.get("title"):
            exec_summary["title"] = _deep_export_title(dataset, deep)
        return {
            "execSummary": exec_summary,
            "kpiBenchmark": list(deep.get("kpiBenchmark") or []),
            "actionPlan": _deep_export_normalize_action_plan(deep.get("actionPlan") or []),
            "rawParsingQa": (dataset or {}).get("rawParsingQa") or deep.get("rawParsingQa"),
        }
    operators = dataset.get("operators") or []
    if not operators:
        return None
    transfer_lookup = {}
    for entry in dataset.get("transferSummary") or []:
        op = str(entry.get("operator") or "").upper()
        direction = str(entry.get("direction") or "")
        dnorm = "DL" if ("down" in direction.lower() or direction.upper() in ("DL", "DOWNLINK")) else \
                ("UL" if ("up" in direction.lower() or direction.upper() in ("UL", "UPLINK")) else direction.upper())
        transfer_lookup[(op, dnorm)] = entry
    operator_map = {str(o.get("operator") or "").upper(): o for o in operators}
    iam_entry = next((o for o in operators if str(o.get("operator") or "").upper() in _DEEP_IAM_ALIASES), None)
    if not iam_entry:
        return None
    iam = _deep_export_extract_operator(dataset, iam_entry, transfer_lookup)
    iam["mcs"] = _deep_num(((iam_entry.get("kpis") or {}).get("pdschMcs") or {}).get("median"))
    orange = _deep_export_extract_operator(dataset, operator_map.get("ORANGE"), transfer_lookup) if operator_map.get("ORANGE") else None
    orange["mcs"] = _deep_num(((operator_map.get("ORANGE").get("kpis") or {}).get("pdschMcs") or {}).get("median")) if orange else None
    inwi = _deep_export_extract_operator(dataset, operator_map.get("INWI"), transfer_lookup) if operator_map.get("INWI") else None
    inwi["mcs"] = _deep_num(((operator_map.get("INWI").get("kpis") or {}).get("pdschMcs") or {}).get("median")) if inwi else None

    kpi_specs = [
        ("DL Throughput (Mbps)", "dlThroughput", 1, "%", "IAM is far behind Orange but above INWI."),
        ("Median RSRP (dBm)", "rsrp", 1, "dB", "IAM RSRP is slightly better than Orange; coverage is not the main issue."),
        ("Median SINR (dB)", "sinr", 1, "dB", "Both are low; IAM needs SINR improvement to unlock higher modulation/MCS."),
        ("Median CQI", "cqi", 0, "", "IAM radio quality feedback is weaker despite slightly better RSRP/SINR."),
        ("5G Presence %", "fivegPresence", 1, "pp", "IAM has more 5G time than Orange, but on low band only."),
        ("4G Only %", "fourgOnly", 1, "pp", f"{_deep_fmt(iam.get('fourgOnly'))}% 4G-only time still limits peak DL and user experience."),
        ("NR n78 share %", "n78", 1, "pp", "Key strategic gap: IAM 5G observed only on n28; Orange uses n78."),
        ("NR n28 share %", "n28", 1, "pp", "IAM 5G low-band gives coverage but not comparable capacity."),
        ("256QAM share %", "qam256", 1, "pp", "No 256QAM observed; radio conditions/scheduler/MCS must be improved."),
        ("64QAM share %", "qam64", 1, "pp", "Moderate modulation share; most samples still in 16QAM."),
        ("16QAM share %", "qam16", 1, "pp", "IAM is heavily stuck in 16QAM."),
        ("QPSK share %", "qpsk", 1, "pp", "Orange has lower modulation distribution but higher throughput due to n78/capacity."),
        ("Median Rank", "medianRank", 0, "", "Critical IAM MIMO limitation: mostly RI1/RI2, no RI>=3."),
        ("RI1 share %", "ri1", 1, "pp", "High RI1 share indicates poor spatial multiplexing."),
        ("RI2 share %", "ri2", 1, "pp", "IAM needs stronger RI2+ persistence."),
        ("RI>=3 share %", "riGe3", 1, "pp", "No high-rank usage; validate 4T4R/NR MIMO configuration."),
        ("Avg # SCells", "scellsAvg", 2, "", "IAM CA depth is weak versus INWI; similar to Orange LTE CA only."),
        ("Max # SCells", "scellsMax", 0, "", "IAM maximum aggregation is limited."),
        ("SCells >0 share %", "scellsActive", 1, "pp", "CA is not persistent enough; target higher SCell activation."),
        ("LTE CA active share %", "lteCaActive", 1, "pp", "LTE CA use exists but remains insufficient for throughput ambition."),
        ("BLER Avg %", "blerAvg", 1, "pp", "IAM BLER is acceptable but worse than Orange."),
        ("BLER P90 %", "blerP90", 1, "pp", "Peak BLER periods affect throughput stability."),
        ("BLER >10% share %", "blerAbove10", 1, "pp", "IAM has much more high-BLER samples."),
        ("UL Retx Avg %", "ulRetxAvg", 1, "pp", "UL retransmission indicates possible UL quality/interference/power control issue."),
        ("PDSCH DL Avg (Mbps)", "pdschDlAvg", 1, "%", "Large scheduler/NR capacity gap versus Orange."),
        ("TCP Handshake Median (ms)", "tcpHandshake", 0, "ms", "IAM transport latency is worse than both competitors."),
        ("Ping Success Rate %", "pingSuccessRate", 0, "pp", "Low ping success in both IAM/Orange; verify test setup and packet-loss path."),
        ("DL completion %", "dlCompletion", 0, "pp", "Download sessions complete; issue is performance, not session reliability."),
        ("DL success %", "dlSuccess", 0, "pp", "No basic accessibility issue visible in this sample."),
    ]
    kpi_rows = []
    for label, key, digits, unit, interpretation in kpi_specs:
        iam_v = _deep_export_quantize(iam.get(key), digits)
        orange_v = _deep_export_quantize((orange or {}).get(key) if orange else None, digits)
        inwi_v = _deep_export_quantize((inwi or {}).get(key) if inwi else None, digits)
        kpi_rows.append({
            "kpi": label,
            "iam": iam_v,
            "orange": orange_v,
            "inwi": inwi_v,
            "vsOrange": _deep_export_delta(iam_v, orange_v, unit, digits),
            "vsInwi": _deep_export_delta(iam_v, inwi_v, unit, digits),
            "interpretation": interpretation,
        })

    episodes = ((dataset.get("iamServingCells") or {}).get("episodesAll") or [])
    sequence_cells = []
    for episode in episodes:
        name = str((episode or {}).get("cellName") or "").strip()
        if name and name not in sequence_cells:
            sequence_cells.append(name)
    sequence_list_text, sequence_text = _deep_export_sequence_names(sequence_cells)

    action_plan = [
        {
            "priority": "P1",
            "domain": "5G capacity layer",
            "finding": f"IAM 5G presence is {_deep_fmt(iam.get('fivegPresence'))}% but {_deep_fmt(iam.get('n28'))}% of NR band is n28; Orange uses {_deep_fmt((orange or {}).get('n78'))}% n78 and reaches {_deep_fmt((orange or {}).get('dlThroughput'))} Mbps.",
            "recommendedActions": "Audit n78 availability around the DT path, check if n78 cells exist/are barred/not selected, verify PCI/SSB footprint, EN-DC anchor relations, NSA addition thresholds, A3/A5/B1/B2 events, SSB beam coverage, and NR neighbor configuration. Add or retune n78 layer where available.",
            "owner": "RF Optimization + 5G Planning",
            "expectedImpact": "Biggest DL gain: more mid-band capacity and higher PDSCH throughput.",
            "validationTarget": "n78 share >70%, 5G presence >70%, DL throughput >120 Mbps on same route.",
        },
        {
            "priority": "P1",
            "domain": "MIMO / RI",
            "finding": f"IAM median rank is {_deep_fmt(iam.get('medianRank'))}, RI1={_deep_fmt(iam.get('ri1'))}%, RI2={_deep_fmt(iam.get('ri2'))}%, RI>=3={_deep_fmt(iam.get('riGe3'))}; Orange median rank is {_deep_fmt((orange or {}).get('medianRank'))} with RI2={_deep_fmt((orange or {}).get('ri2'))}%.",
            "recommendedActions": "Verify antenna ports, 4T4R/2T2R configuration, TM/PMI/RI reporting, cross-polar imbalance, azimuth/tilt, calibration alarms, VSWR, feeder/RRU branch health, and rank adaptation parameters. Prioritize cells in the serving sequence: " + sequence_list_text + ".",
            "owner": "RF Optimization + RAN Vendor",
            "expectedImpact": "Improve spatial multiplexing and spectral efficiency.",
            "validationTarget": "Median RI >=2, RI1 <20%, RI2+ >80%.",
        },
        {
            "priority": "P1",
            "domain": "SINR / interference",
            "finding": f"IAM median SINR is only {_deep_fmt(iam.get('sinr'))} dB; CQI={_deep_fmt(iam.get('cqi'))} and MCS={_deep_fmt(iam.get('mcs'))} keep modulation mainly at 16QAM.",
            "recommendedActions": "Run SINR/RSRP grid on DT route, check overshooting cells, pilot pollution/PCI confusion, neighbor missing, overlapping sectors, mechanical/electrical tilt, azimuth alignment, and DL interference. Apply tilt/azimuth corrections and neighbor cleanup before parameter-only changes.",
            "owner": "RF Optimization",
            "expectedImpact": "Higher CQI/MCS and more 64QAM/256QAM.",
            "validationTarget": "Median SINR >8 dB, CQI >10, 64QAM+ share >50%.",
        },
        {
            "priority": "P1",
            "domain": "Scheduler / PRB efficiency",
            "finding": f"IAM PRB efficiency is {_deep_fmt(iam.get('prbEfficiency'), 3)} versus Orange {_deep_fmt((orange or {}).get('prbEfficiency'), 3)}; PDSCH average is {_deep_fmt(iam.get('pdschDlAvg'))} Mbps vs {_deep_fmt(_deep_export_quantize((orange or {}).get('pdschDlAvg'), 1))} Mbps.",
            "recommendedActions": "Review scheduler configuration: DL PRB allocation, proportional fair/QoS weights, MCS table, HARQ process, link adaptation, CQI aging, outer-loop link adaptation, PDSCH power allocation, and EN-DC split bearer configuration. Compare IAM parameters with best-performing cells on same vendor.",
            "owner": "RAN Vendor + Optimization",
            "expectedImpact": "Higher throughput without only adding sites.",
            "validationTarget": "PRB efficiency >0.65 and PDSCH average gap vs Orange reduced by 50%.",
        },
        {
            "priority": "P2",
            "domain": "Carrier aggregation",
            "finding": f"IAM Avg #SCells={_deep_fmt(iam.get('scellsAvg'), 2)} and SCells>0 only {_deep_fmt(iam.get('scellsActive'))}%; INWI is {_deep_fmt((inwi or {}).get('scellsAvg'), 2)} and {_deep_fmt((inwi or {}).get('scellsActive'))}%.",
            "recommendedActions": "Audit CA combinations and UE capability matching, SCell addition thresholds, SCell activation timers, PCell/SCell load balance, inter-frequency coverage, and allowed CA combos for the used terminal. Increase CA persistence where RF quality supports it.",
            "owner": "Optimization + RAN Vendor",
            "expectedImpact": "Boost 4G/NSA DL capacity, especially during 4G-only periods.",
            "validationTarget": "SCells>0 >50%, Avg #SCells >1, LTE CA active share >50%.",
        },
        {
            "priority": "P2",
            "domain": "BLER / retransmissions",
            "finding": f"IAM BLER average={_deep_fmt(iam.get('blerAvg'))}%, P90={_deep_fmt(iam.get('blerP90'))}%, BLER>10%={_deep_fmt(iam.get('blerAbove10'))}%, worse than Orange.",
            "recommendedActions": "Map BLER spikes by serving cell and location; check link adaptation aggressiveness, PDSCH power, MCS backoff, HARQ retransmission, interference, and overshooting. Correct RF first; then tune OLLA and MCS thresholds.",
            "owner": "Optimization + Vendor",
            "expectedImpact": "Stabilize throughput and reduce retransmission overhead.",
            "validationTarget": "BLER avg <3%, P90 <10%, BLER>10% <10%.",
        },
        {
            "priority": "P2",
            "domain": "UL quality",
            "finding": f"IAM UL retransmission avg={_deep_fmt(iam.get('ulRetxAvg'))}% versus Orange {_deep_fmt((orange or {}).get('ulRetxAvg'))}%.",
            "recommendedActions": "Check UL interference/RTWP on serving LTE anchors, PUSCH power control, UL pathloss, antenna branch health, and UL scheduler. Correlate with poor TCP handshake and packet loss.",
            "owner": "NOC + Optimization",
            "expectedImpact": "Better TCP stability and faster session ramp-up.",
            "validationTarget": "UL Retx avg <1%, TCP handshake median <70 ms.",
        },
        {
            "priority": "P2",
            "domain": "Transport / core",
            "finding": f"IAM TCP handshake median is {_deep_fmt(iam.get('tcpHandshake'))} ms versus Orange {_deep_fmt((orange or {}).get('tcpHandshake'))} ms and INWI {_deep_fmt((inwi or {}).get('tcpHandshake'))} ms; ping success is only {_deep_fmt(iam.get('pingSuccessRate'))}%.",
            "recommendedActions": "Validate backhaul latency/jitter/loss on involved eNB/gNB, S1/N3 path, DNS/APN path, firewall/NAT, and test server route. Repeat ping with controlled server to exclude test-tool artifact.",
            "owner": "Transport/Core + Optimization",
            "expectedImpact": "Improve app-layer responsiveness and throughput ramp-up.",
            "validationTarget": "TCP handshake <65 ms, packet loss <1%, ping success >98% in controlled retest.",
        },
        {
            "priority": "P3",
            "domain": "Mobility / serving sequence",
            "finding": f"Serving sequence alternates between {sequence_text}.",
            "recommendedActions": "Review handover and EN-DC addition/release events along the segment. Check if repeated returns to LTE anchor cause NR drops, throughput dips, or SCell reset. Optimize neighbor priorities and thresholds.",
            "owner": "RF Optimization",
            "expectedImpact": "Smoother 5G retention and less throughput fluctuation.",
            "validationTarget": "Reduce unnecessary serving changes; 5G presence >70%.",
        },
        {
            "priority": "P3",
            "domain": "Retest governance",
            "finding": "Dataset has only one row per operator/sample file; conclusions should be validated on repeated DTs.",
            "recommendedActions": "Repeat the same route in busy hour and off-peak with same UE, same SIM plan, same server, and locked/controlled test scripts. Add geo-location bins to localize bad segments.",
            "owner": "Optimization QA",
            "expectedImpact": "Avoid wrong actions based on one sample.",
            "validationTarget": "Consistent improvement across 3+ retests.",
        },
    ]

    main_conclusion = (
        f"IAM issue is not basic accessibility: DL completion and success are {_deep_fmt(iam.get('dlCompletion'))}%. "
        "The key gaps are 5G capacity layer, MIMO rank, SINR/CQI/MCS, PRB efficiency, CA persistence, BLER peaks, and transport latency."
    )
    return {
        "execSummary": {
            "title": _deep_export_title(dataset, deep),
            "scope": "IAM only, benchmarked against Orange and INWI where useful",
            "mainConclusion": main_conclusion,
            "topKpis": _deep_export_top_kpis(iam, orange, inwi),
            "immediatePriorities": "1) n78/EN-DC audit, 2) MIMO/RI audit, 3) SINR & RF cleanup, 4) scheduler/PRB efficiency tuning, 5) CA activation and BLER/TCP validation.",
        },
        "kpiBenchmark": kpi_rows,
        "actionPlan": action_plan,
    }


def _deep_scope_label_from_operators(operators) -> str:
    """Single-DT scope label = the IAM operator's drive-test title when every operator
    carries exactly one measurement title; otherwise the cumulative label. Operators'
    per-DT titles differ (different timestamps), so count titles per operator."""
    ops = operators or []
    counts = [len(o.get("measurementTitles") or []) for o in ops]
    if counts and max(counts) <= 1:
        iam_titles = next(
            (o.get("measurementTitles") for o in ops
             if str(o.get("operator") or "").upper() in _DEEP_IAM_ALIASES and o.get("measurementTitles")),
            None,
        )
        title = (iam_titles or next((o.get("measurementTitles") for o in ops if o.get("measurementTitles")), None) or [None])[0]
        return str(title) if title else "Single DT"
    return "All DTs (combined)"


def _ensure_deep_benchmark(dataset: dict) -> dict:
    """Compute `deepBenchmark` on an already-built dataset if absent (e.g. a dataset
    restored from the SQLite library cache that predates this feature). Cheap: runs the
    rule engine on the stored per-operator `kpis` — no TXT re-parse, no serving-cell build."""
    if not isinstance(dataset, dict):
        return dataset
    operators = dataset.get("operators")
    if not operators:
        return dataset
    scope_label = _deep_scope_label_from_operators(operators)
    try:
        raw_parsing_qa = dataset.get("rawParsingQa")
        deep = dataset.get("deepBenchmark")
        if not deep:
            deep = _benchmark_deep_analysis(
                operators, dataset.get("transferSummary"), scope_label
            )
        dataset["deepBenchmark"] = _deep_enrich_action_plan_with_current_data(
            deep,
            operators,
            dataset.get("iamServingCells"),
            dataset.get("transferSummary"),
            scope_label,
        )
        if dataset.get("deepBenchmark") is not None and raw_parsing_qa:
            dataset["deepBenchmark"]["rawParsingQa"] = raw_parsing_qa
    except Exception:
        import traceback as _tb
        _tb.print_exc()
    return dataset


def generate_benchmark_deep_xlsx(deep: dict, dataset: dict | None = None) -> bytes:
    """Render the Deep Benchmark analysis as the 3-sheet workbook (Executive Summary,
    IAM KPI Benchmark, IAM Action Plan) matching the user's manual template."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if dataset:
        export_model = _benchmark_deep_export_model(dataset, deep)
        if export_model:
            deep = export_model

    if not deep:
        raise ValueError("No deep benchmark analysis available for this scope.")

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(bold=True, size=14, color="1F4E78")
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    label_font = Font(bold=True)
    sev_fill = {
        "Critical": PatternFill("solid", fgColor="F8CBAD"),
        "High": PatternFill("solid", fgColor="FCE4D6"),
        "Medium": PatternFill("solid", fgColor="FFF2CC"),
        "Low": PatternFill("solid", fgColor="E2EFDA"),
    }

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = border

    wb = openpyxl.Workbook()

    # ── Sheet 1: Executive Summary ─────────────────────────────────────────────
    es = deep.get("execSummary") or {}
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.cell(row=1, column=1, value=es.get("title") or "IAM Professional Analysis").font = title_font
    r = 2
    ws1.cell(row=r, column=1, value="Scope").font = label_font
    ws1.cell(row=r, column=2, value=es.get("scope") or "")
    r += 1
    ws1.cell(row=r, column=1, value="Main conclusion").font = label_font
    ws1.cell(row=r, column=2, value=es.get("mainConclusion") or "")
    r += 1
    cc = es.get("causalChain") or {}
    if cc.get("chainText") or cc.get("narrative"):
        ws1.cell(row=r, column=1, value="Causal chain").font = label_font
        ws1.cell(row=r, column=2, value=((cc.get("chainText") or "") + ("\n" + cc.get("narrative") if cc.get("narrative") else "")).strip())
        r += 1
    ga = es.get("gapAttribution") or {}
    if ga.get("drivers"):
        ws1.cell(row=r, column=1, value="Gap attribution").font = label_font
        ws1.cell(row=r, column=2, value="; ".join(f"{d.get('driver')} {int(d.get('contributionPct') or 0)}%" for d in ga.get("drivers")))
        r += 1
    if es.get("confidence") or es.get("confidenceNote"):
        ws1.cell(row=r, column=1, value="Confidence").font = label_font
        ws1.cell(row=r, column=2, value=es.get("confidenceNote") or es.get("confidence") or "")
        r += 1
    r += 1
    for ci, h in enumerate(("Top KPI", "IAM", "Best competitor", "Gap"), start=1):
        ws1.cell(row=r, column=ci, value=h)
    style_header(ws1, r, 4)
    r += 1
    for k in es.get("topKpis") or []:
        ws1.cell(row=r, column=1, value=k.get("kpi"))
        ws1.cell(row=r, column=2, value=k.get("iam"))
        ws1.cell(row=r, column=3, value=k.get("bestCompetitor"))
        ws1.cell(row=r, column=4, value=k.get("gap"))
        for c in range(1, 5):
            ws1.cell(row=r, column=c).border = border
            ws1.cell(row=r, column=c).alignment = wrap_left
        r += 1
    r += 1
    ws1.cell(row=r, column=1, value="Immediate priorities").font = label_font
    ws1.cell(row=r, column=2, value=es.get("immediatePriorities") or "")
    r += 1
    for col, w in (("A", 24), ("B", 60), ("C", 28), ("D", 28)):
        ws1.column_dimensions[col].width = w

    # ── Sheet 2: IAM KPI Benchmark ─────────────────────────────────────────────
    ws2 = wb.create_sheet("IAM KPI Benchmark")
    headers2 = ("KPI", "IAM", "Orange", "INWI", "IAM vs Orange", "IAM vs INWI", "Interpretation")
    for ci, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=ci, value=h)
    style_header(ws2, 1, len(headers2))
    rr = 2
    for row in deep.get("kpiBenchmark") or []:
        ws2.cell(row=rr, column=1, value=row.get("kpi"))
        ws2.cell(row=rr, column=2, value=row.get("iam"))
        ws2.cell(row=rr, column=3, value=row.get("orange"))
        ws2.cell(row=rr, column=4, value=row.get("inwi"))
        ws2.cell(row=rr, column=5, value=row.get("vsOrange"))
        ws2.cell(row=rr, column=6, value=row.get("vsInwi"))
        ws2.cell(row=rr, column=7, value=row.get("interpretation"))
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=rr, column=c)
            cell.border = border
            cell.alignment = wrap_left if c in (1, 7) else center
        rr += 1
    for col, w in (("A", 26), ("B", 11), ("C", 11), ("D", 11), ("E", 14), ("F", 14), ("G", 62)):
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: IAM Action Plan (simplified 5-column) ─────────────────────────
    ws3 = wb.create_sheet("IAM Action Plan")
    headers3 = ("Priority", "Domain", "Finding", "Action", "Confidence")
    for ci, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=ci, value=h)
    style_header(ws3, 1, len(headers3))
    prio_fill = {
        "P1": PatternFill("solid", fgColor="F8CBAD"),
        "P2": PatternFill("solid", fgColor="FFE699"),
        "P3": PatternFill("solid", fgColor="E2EFDA"),
    }
    rr = 2
    for f in deep.get("actionPlan") or []:
        action_value = f.get("action")
        if isinstance(action_value, (list, tuple)):
            action_value = "\n".join(action_value)
        ws3.cell(row=rr, column=1, value=f.get("priority"))
        ws3.cell(row=rr, column=2, value=f.get("domain"))
        ws3.cell(row=rr, column=3, value=f.get("finding"))
        ws3.cell(row=rr, column=4, value=action_value)
        ws3.cell(row=rr, column=5, value=f.get("confidence"))
        fill = prio_fill.get(f.get("priority"))
        for c in range(1, len(headers3) + 1):
            cell = ws3.cell(row=rr, column=c)
            cell.border = border
            cell.alignment = center if c in (1, 5) else wrap_left
            if fill and c == 1:
                cell.fill = fill
        rr += 1
    for col, w in (("A", 9), ("B", 26), ("C", 70), ("D", 70), ("E", 26)):
        ws3.column_dimensions[col].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# Bump whenever _benchmark_nemo_build_dataset / _nemo_operator_kpis analysis logic
# changes (e.g. DT-weighted cumulative DL average, Deep Benchmark). Stale cached
# dataset blobs (in-memory + SQLite library) are then rebuilt from the already-parsed
# operator_files — no TXT re-parse — instead of being served as-is.
_BENCHMARK_NEMO_ANALYSIS_VERSION = 52


def _benchmark_nemo_dataset_current(dataset) -> bool:
    return isinstance(dataset, dict) and dataset.get("analysisVersion") == _BENCHMARK_NEMO_ANALYSIS_VERSION


def _benchmark_nemo_refresh_dataset(
    dataset,
    operator_files,
    dl_mode: str | None = None,
    window_mode: str | None = None,
):
    """Return an up-to-date dataset, rebuilding from already-parsed operator_files when
    the cached blob predates the current analysis version. Returns (dataset, rebuilt)."""
    dl_mode = _benchmark_nemo_normalize_dl_mode(dl_mode)
    window_mode = _benchmark_nemo_normalize_window_mode(window_mode)
    if (
        _benchmark_nemo_dataset_current(dataset)
        and _benchmark_nemo_normalize_dl_mode((dataset or {}).get("dlMode")) == dl_mode
        and _benchmark_nemo_normalize_window_mode((dataset or {}).get("windowMode")) == window_mode
    ):
        return dataset, False
    if not operator_files:
        return _ensure_deep_benchmark(dataset), False
    return _benchmark_nemo_build_dataset(operator_files, dl_mode=dl_mode, window_mode=window_mode), True


def _benchmark_nemo_build_dataset(
    operator_files: list[dict],
    dl_mode: str | None = None,
    window_mode: str | None = None,
) -> dict:
    dl_mode = _benchmark_nemo_normalize_dl_mode(dl_mode)
    dl_mode_label = _benchmark_nemo_dl_mode_label(dl_mode)
    window_mode = _benchmark_nemo_normalize_window_mode(window_mode)
    window_mode_label = _benchmark_nemo_window_mode_label(window_mode)
    prepared_operator_files = [
        _benchmark_nemo_scope_operator_file_to_window(operator_file, window_mode=window_mode)
        for operator_file in (operator_files or [])
    ]
    window_fallback = window_mode == "active_dl_session" and any(
        of.get("_windowFallback") for of in prepared_operator_files
    )
    for operator_file in prepared_operator_files:
        _nemo_reapply_throughput_normalization(operator_file)
        operator_file["dlMetricKey"] = operator_file.get("_dlMetricKeyOverride") or _nemo_select_dl_metric_key(operator_file.get("rows") or [])
        operator_file["benchmarkDlMetricKeyDefault"] = (
            operator_file.get("_benchmarkDlMetricKeyAvgOverride")
            or operator_file.get("_benchmarkDlMetricKeyOverride")
            or _nemo_select_benchmark_dl_metric_key(operator_file.get("rows") or [])
        )
        operator_file["benchmarkDlMetricKey"] = _benchmark_nemo_canonical_dl_metric_key(operator_file, dl_mode)
        operator_file["benchmarkDlMode"] = dl_mode
        operator_file["transferSessions"] = list(operator_file.get("transferSessions") or []) or _nemo_build_transfer_sessions(
            operator_file.get("rows") or [],
            operator_file.get("operator") or "UNKNOWN",
        )
        operator_file["tests"] = _nemo_build_tests(
            operator_file.get("rows") or [],
            operator_file.get("operator") or "UNKNOWN",
            operator_file.get("benchmarkDlMetricKey") or "",
        )
        _nemo_align_benchmark_tests_with_transfer_sessions(operator_file)
        operator_file["kpis"] = _nemo_operator_kpis(operator_file)

    ranking = _nemo_build_ranking(prepared_operator_files)
    diagnosis = _nemo_build_diagnosis(prepared_operator_files, ranking)
    weakness_chain = _nemo_build_weakness_evidence_chain(prepared_operator_files, diagnosis)
    rf_exclusion = _nemo_build_rf_exclusion_check(prepared_operator_files, diagnosis)
    scheduler_prb_deep_dive = _nemo_build_scheduler_prb_deep_dive(prepared_operator_files, diagnosis)
    pdsch_modulation_efficiency = _nemo_build_pdsch_modulation_efficiency(prepared_operator_files, diagnosis)
    nr_band_exposure = _nemo_build_nr_band_exposure_analysis(prepared_operator_files, diagnosis)
    ca_scells_analysis = _nemo_build_ca_scells_analysis(prepared_operator_files, diagnosis)
    mimo_rank_analysis = _nemo_build_mimo_rank_analysis(prepared_operator_files, diagnosis)
    bler_retx_analysis = _nemo_build_bler_retx_analysis(prepared_operator_files, diagnosis)
    transport_gap_analysis = _nemo_build_transport_gap_analysis(prepared_operator_files, diagnosis)
    per_cgps_weakness_analysis = _nemo_build_per_cgps_weakness_analysis(prepared_operator_files, diagnosis)
    missing_kpi_quality = _nemo_build_missing_kpi_quality(prepared_operator_files)
    iam_serving_cells = None
    inwi_serving_cells = None
    orange_serving_cells = None
    
    for op_data in prepared_operator_files:
        op_name = str(op_data.get("operator") or "").upper()
        cells = _nemo_build_operator_serving_cells(op_data)
        if cells and cells.get("available"):
            cells = _nemo_attach_serving_cell_presence_metadata(
                cells,
                op_data.get("technologyStatus"),
                _nemo_dominant_nr_serving_info(op_data.get("rows") or []),
            )
        if op_name == "IAM":
            iam_serving_cells = cells
        elif op_name == "INWI":
            inwi_serving_cells = cells
        elif op_name == "ORANGE":
            orange_serving_cells = cells

    # Align operator 5G/4G presence KPIs with the dwell-based serving-cell breakdown so the
    # comparator's "5G presence % (time-based)" rows match the Technology Status table.
    # In Active-DL scope this must use the download-window breakdown; otherwise it uses
    # the all-window breakdown.
    _serving_cells_by_op = {"IAM": iam_serving_cells, "INWI": inwi_serving_cells, "ORANGE": orange_serving_cells}
    for op_data in prepared_operator_files:
        sc = _serving_cells_by_op.get(str(op_data.get("operator") or "").upper())
        breakdown = (
            ((sc or {}).get("radioPresenceBreakdownDownload") if window_mode == "active_dl_session" else None)
            or (sc or {}).get("radioPresenceBreakdownAll")
            or (sc or {}).get("radioPresenceBreakdown")
            or {}
        )
        kpis = op_data.get("kpis")
        if not isinstance(breakdown, dict) or not breakdown or not isinstance(kpis, dict):
            continue
        if breakdown.get("5G") is not None:
            kpis["nrPresencePct"] = breakdown.get("5G")
        if breakdown.get("4G") is not None:
            kpis["lteOnlyPresencePct"] = breakdown.get("4G")
    iam_vs_best_5g = _nemo_build_5g_comparator_analysis(prepared_operator_files, diagnosis)

    layer_throughput_analysis = _nemo_build_layer_throughput_analysis(prepared_operator_files, iam_serving_cells)
    recommendations_by_priority = _nemo_build_recommendations_by_priority(prepared_operator_files, diagnosis)
    executive_conclusion = _nemo_build_executive_conclusion(prepared_operator_files, ranking, diagnosis)
    professional_executive_summary = _nemo_build_professional_executive_summary(prepared_operator_files, diagnosis)
    validation_warnings = _nemo_build_validation_warnings(prepared_operator_files, diagnosis)
    rules_applied = _nemo_build_rules_applied(prepared_operator_files, diagnosis)
    qa_checklist = _nemo_build_qa_checklist(prepared_operator_files, ranking, diagnosis)
    iam_entry = next((entry for entry in ranking if str(entry.get("operator") or "").upper() == "IAM"), None)
    best_dl_entry = ranking[0] if ranking else None
    no_5g_note = ""
    no_5g_note_fr = ""
    if best_dl_entry and not best_dl_entry.get("has5g"):
        no_5g_note = "The best DL operator has no 5G detected in the export, so this ranking reflects measured DL throughput, not 5G-specific superiority."
        no_5g_note_fr = "Le meilleur opérateur DL n'a pas de 5G détectée dans l'export ; ce classement reflète le débit DL mesuré, pas la supériorité 5G spécifique."
    _iam_rank_ord = _nemo_ordinal((iam_entry or {}).get('rank'))
    _best_op = (best_dl_entry or {}).get('operator') or '—'
    _best_mbps = _nemo_safe_round((best_dl_entry or {}).get('avgDlMbps'), 1)
    _iam_mbps = _nemo_safe_round((iam_entry or {}).get('avgDlMbps'), 1)
    _gap_pct = abs(diagnosis.get('gapToBestDlPct') or 0)
    _rank_ord_fr = {"1st": "1er", "2nd": "2ème", "3rd": "3ème"}.get(_iam_rank_ord, (_iam_rank_ord or 'N/A') + "ème")
    ranking_summary = (
        f"DL throughput ranking includes all operators, even if 5G was not detected. "
        f"{_best_op} ranks first with {_best_mbps} Mbps average DL throughput. "
        f"IAM ranks {_iam_rank_ord} with {_iam_mbps} Mbps, "
        f"representing a {_gap_pct}% gap versus the best DL operator. "
        f"{no_5g_note}".strip()
    ) if best_dl_entry and iam_entry else "DL throughput ranking includes all operators, even if 5G was not detected."
    ranking_summary_fr = (
        f"Le classement DL inclut tous les opérateurs, même sans 5G détectée. "
        f"{_best_op} se classe premier avec {_best_mbps} Mbps de débit DL moyen. "
        f"IAM se classe {_rank_ord_fr} avec {_iam_mbps} Mbps, "
        f"soit un écart de {_gap_pct}% par rapport au meilleur opérateur DL. "
        f"{no_5g_note_fr}".strip()
    ) if best_dl_entry and iam_entry else "Le classement DL inclut tous les opérateurs, même sans 5G détectée."

    all_tests = []
    for item in prepared_operator_files:
        all_tests.extend(item.get("tests") or [])
    all_tests.sort(key=lambda entry: (str(entry.get("anchorTime") or ""), str(entry.get("operator") or ""), str(entry.get("id") or "")))

    all_transfer_sessions = []
    for item in prepared_operator_files:
        all_transfer_sessions.extend(item.get("transferSessions") or [])
    all_transfer_sessions.sort(key=lambda entry: (str(entry.get("startTime") or ""), str(entry.get("operator") or ""), str(entry.get("id") or "")))
    transfer_summary = _nemo_build_transfer_summary(all_transfer_sessions)

    # Build per-operator DL time-series — four explicit metric categories.
    # Each category lists keys in strict priority order; the first key that has
    # ANY positive data is used (not the one with the most rows — that caused
    # inconsistent operator comparisons for the MAC category).
    _TL_CATS = {
        "app":         ["appDlMbps", "appDlAvgMbps"],
        "mac":         ["totalMacDlMbps", "macDl5gMbps", "macDlLteMbps"],
        "mac_nr":      ["macDl5gMbps"],
        "mac_lte":     ["macDlLteMbps"],
        "pdsch_dl":    ["pdschDl5gMbps"],
        "pdsch_lte":   ["pdschDlLteMbps"],
        "pdsch_sched": ["pdschSched5gMbps"],
        "ppp":         ["pppRateDl"],
    }
    _METRIC_LABELS = {
        "appDlMbps":        "App DL",
        "appDlAvgMbps":     "App DL (avg)",
        "totalMacDlMbps":   "MAC Total DL",
        "macDl5gMbps":      "MAC NR (5G) DL",
        "macDlLteMbps":     "MAC LTE DL",
        "pdschDl5gMbps":    "PDSCH NR delivered",
        "pdschDlLteMbps":   "PDSCH LTE delivered",
        "pdschSched5gMbps": "PDSCH NR scheduled",
        "pppRateDl":        "PPP rate DL",
    }

    def _tl_build_series(op_rows, key, session_intervals=None):
        # Bucket to 1-second resolution (one point per second) so the operators share a
        # common x-axis. `t` stays whole-second for that alignment; `tFull` carries the
        # real millisecond timestamp of the chosen sample for the tooltip (Nemo style).
        #
        # When session_intervals (DAA→DAD windows) are provided, only samples inside an
        # active download session are kept — the timeline then shows EXACTLY the
        # "Session DL active" span (Data server connection Attempt → Data server
        # Disconnect), excluding idle time between sessions.
        def _in_session(dt_val):
            if not session_intervals:
                return True
            for iv in session_intervals:
                if iv["start"] <= dt_val <= iv["end"]:
                    return True
            return False

        buckets: dict = {}
        for row in op_rows:
            dt_val = row.get("_dt")
            if dt_val is None:
                continue
            if not _in_session(dt_val):
                continue
            v = row.get(key)
            if v is None:
                continue
            try:
                fv = float(v)
                if fv <= 0:
                    continue
            except (TypeError, ValueError):
                continue
            sec_key = dt_val.replace(microsecond=0)
            if sec_key not in buckets or fv > buckets[sec_key][1]:
                buckets[sec_key] = (dt_val, fv)
        return [
            {"t": k.strftime("%H:%M:%S"), "tFull": _nemo_fmt_hms_ms(dt), "dl": round(v, 2)}
            for k, (dt, v) in sorted(buckets.items())
        ]

    dl_timeline_by_metric: dict = {}
    for item in prepared_operator_files:
        op_name  = item.get("operator") or "UNKNOWN"
        op_rows  = item.get("rows") or []
        op_entry: dict = {"selectedKey": item.get("dlMetricKey") or ""}
        session_rows = item.get("_sessionStatsRows") or op_rows
        # Reconstruct the download / upload / ping operations from the time-series Event IDs
        # + per-row transfer KPIs (Bytes DL, Download time, direction, protocol, status).
        # This isolates the real HTTP download session and yields its authoritative timing &
        # throughput from the SAME export that drives the timeline — no separate
        # session-statistics file needed.
        dl_events = _nemo_extract_dl_events(session_rows)
        # Scope the timeline to the download session window (DAA→DAD), +1s tail so the last
        # partial-second bucket isn't clipped.
        scope_intervals = []
        for iv in (dl_events.get("downloadWindow") or dl_events.get("sessionIntervals") or []):
            scope_intervals.append({"start": iv["start"], "end": iv["end"] + _td(seconds=1.0)})
        for cat, keys in _TL_CATS.items():
            # Use the first key in priority order that has any positive data
            best_key = None
            for k in keys:
                for r in op_rows:
                    v = r.get(k)
                    if v is not None:
                        try:
                            if float(v) > 0:
                                best_key = k
                                break
                        except (TypeError, ValueError):
                            pass
                if best_key:
                    break
            if best_key:
                pts = _tl_build_series(op_rows, best_key, scope_intervals)
                if pts:
                    op_entry[cat] = {
                        "key":   best_key,
                        "label": _METRIC_LABELS.get(best_key, best_key),
                        "points": pts,
                    }
        if dl_events.get("markers"):
            op_entry["downloadEventMarkers"] = dl_events["markers"]
        if dl_events.get("kpis"):
            op_entry["downloadEventKpis"] = dl_events["kpis"]
        # Per-operation session summary (download / upload / pings) for the upload, latency
        # and RF cards — all derived from the same time series.
        op_entry["sessionStats"] = {
            "kpis": dl_events.get("kpis") or {},
            "download": dl_events.get("download"),
            "upload": dl_events.get("upload"),
            "pings": dl_events.get("pings") or [],
        }
        if any(cat in op_entry for cat in _TL_CATS):
            dl_timeline_by_metric[op_name] = op_entry
        elif dl_events.get("markers"):
            # Even if no throughput series, keep the entry for marker-only display
            dl_timeline_by_metric[op_name] = op_entry

    operators_payload = []
    for item in prepared_operator_files:
        op_name = str(item.get("operator") or "").upper()
        serving_cells = (
            iam_serving_cells if op_name == "IAM"
            else inwi_serving_cells if op_name == "INWI"
            else orange_serving_cells if op_name == "ORANGE"
            else None
        )
        merged_technology_status = _nemo_merge_technology_status_with_serving_cells(
            item.get("technologyStatus") or {},
            serving_cells,
            window_mode=window_mode,
        )
        item["technologyStatus"] = merged_technology_status
        timeline_entry = dl_timeline_by_metric.get(item.get("operator") or "")
        nr_route_presence_pct = (
            merged_technology_status.get("nrPresencePct")
            if isinstance(merged_technology_status, dict)
            else None
        )
        if isinstance(timeline_entry, dict):
            session_stats = timeline_entry.setdefault("sessionStats", {})
            download_stats = session_stats.get("download") or {}
            if download_stats:
                download_stats["nrRoutePresencePct"] = nr_route_presence_pct
                kpis_block = session_stats.get("kpis") or {}
                kpis_block["nrRoutePresencePct"] = nr_route_presence_pct
                session_stats["kpis"] = kpis_block
                session_stats["download"] = download_stats
        operators_payload.append({
            "operator": item.get("operator"),
            "fileName": item.get("fileName"),
            "path": item.get("path"),
            "delimiter": item.get("delimiter"),
            "deviceModel": item.get("deviceModel"),
            "has5g": bool(item.get("has5g")),
            "fiveGStatus": item.get("fiveGStatus"),
            "measurementTitles": item.get("measurementTitles") or [],
            "coverage": item.get("coverage") or {},
            "duplicateHeaders": item.get("duplicateHeaders") or [],
            "throughputScales": item.get("throughputScales") or {},
            "dlMetricKey": item.get("dlMetricKey") or "",
            "benchmarkDlMetricKey": item.get("benchmarkDlMetricKey") or "",
            "benchmarkDlMetricKeyDefault": item.get("benchmarkDlMetricKeyDefault") or "",
            "benchmarkDlMode": item.get("benchmarkDlMode") or dl_mode,
            "kpis": item.get("kpis") or {},
            "tests": item.get("tests") or [],
            "transferSessions": item.get("transferSessions") or [],
            "technologyStatus": merged_technology_status,
        })

    raw_parsing_qa = _benchmark_raw_parsing_qa(prepared_operator_files)
    dt_list = _nemo_build_dt_list(prepared_operator_files)
    device_parity_models = [
        {
            "operator": item.get("operator"),
            "deviceModel": str(item.get("deviceModel") or "").strip(),
        }
        for item in prepared_operator_files
        if str(item.get("deviceModel") or "").strip()
    ]
    device_parity_unique = sorted({item["deviceModel"] for item in device_parity_models})
    device_by_operator = {
        str(item.get("operator") or "UNKNOWN"): (
            str(item.get("deviceModel") or "").strip() or None
        )
        for item in prepared_operator_files
    }
    devices_comparable = len({value for value in device_by_operator.values() if value}) <= 1
    device_parity_warning = ""
    if len(device_parity_unique) > 1:
        device_parity_warning = (
            "Different device models were detected across operators: "
            + "; ".join(
                f"{item['operator']}: {item['deviceModel']}"
                for item in device_parity_models
            )
            + ". Confirm device parity before drawing strong single-DT conclusions."
        )

    active_samples_by_operator = {}
    low_sample_ops = []
    rf_consistency_ops = []
    for op_name, op_entry in dl_timeline_by_metric.items():
        evt_kpis = (op_entry or {}).get("downloadEventKpis") or {}
        active_count = evt_kpis.get("activeSlotCount")
        active_samples_by_operator[op_name] = active_count
        if active_count is not None and active_count < 5:
            low_sample_ops.append((op_name, int(active_count)))
        if evt_kpis.get("rfConsistencyIssues") or evt_kpis.get("rfConsistencyFlags"):
            rf_consistency_ops.append(op_name)

    dt_count = len(dt_list) if dt_list else 0
    scorecard_confidence_level = "low"
    if dt_count >= 4:
        scorecard_confidence_level = "high"
    elif dt_count >= 2:
        scorecard_confidence_level = "medium"
    if dt_count <= 1 or low_sample_ops or rf_consistency_ops:
        scorecard_confidence_level = "low"

    confidence_reason_parts = []
    if dt_count <= 1:
        confidence_reason_parts.append("n=1 drive test")
    if low_sample_ops:
        confidence_reason_parts.append(
            "active download samples: "
            + ", ".join(f"{op} {count}" for op, count in low_sample_ops)
        )
    if rf_consistency_ops:
        confidence_reason_parts.append(
            "RF consistency validator flagged " + ", ".join(rf_consistency_ops)
        )
    if device_parity_warning:
        confidence_reason_parts.append("device models differ across operators")
    scorecard_confidence_reason = (
        "; ".join(confidence_reason_parts)
        if confidence_reason_parts
        else (
            "Multiple drive tests with enough active download samples were available."
            if scorecard_confidence_level == "high"
            else "Directional result with moderate supporting sample depth."
        )
    )
    methodology_note = (
        f"Methodology note: n={dt_count or 1} DT"
        + ("" if (dt_count or 1) == 1 else "s")
        + ". Download-session RF is aggregated only on active slots (App DL or PDSCH > 0), "
        + "SS-SINR and SS-RSRP are throughput-weighted, and active-bandwidth / load metrics use "
        + "the same active-slot scope. Active download samples: "
        + ", ".join(
            f"{op} {active_samples_by_operator.get(op) or 0}"
            for op in ("IAM", "Orange", "INWI")
            if op in device_by_operator or op in active_samples_by_operator
        )
        + ". Devices: "
        + ", ".join(
            f"{op} {device_by_operator.get(op) or 'unknown'}"
            for op in ("IAM", "Orange", "INWI")
            if op in device_by_operator
        )
        + ". Confidence: "
        + scorecard_confidence_reason
        + "."
    )
    benchmark_validity = {
        "deviceByOperator": device_by_operator,
        "devicesComparable": devices_comparable,
        "dtCount": dt_count or 1,
        "confidenceLevel": (
            scorecard_confidence_level.capitalize()
            if scorecard_confidence_level
            else "Low"
        ),
        "confidenceReason": scorecard_confidence_reason,
        "activeSamplesByOperator": active_samples_by_operator,
        "rfConsistencyOperators": rf_consistency_ops,
    }

    # Deep Benchmark analysis (IAM rule engine). Single-DT scope = each operator has
    # exactly one measurement title (their own DT timestamp, which differs across
    # operators), else cumulative. Flows to the frontend panel and the
    # /api/benchmark-deep export via dataset["deepBenchmark"].
    _deep_scope_label = _deep_scope_label_from_operators(operators_payload)
    deep_benchmark = _benchmark_deep_analysis(operators_payload, transfer_summary, _deep_scope_label)
    _iam_operator_file = next(
        (of for of in prepared_operator_files if str(of.get("operator") or "").upper() in _DEEP_IAM_ALIASES),
        None,
    )
    deep_benchmark = _deep_enrich_action_plan_with_current_data(
        deep_benchmark,
        operators_payload,
        iam_serving_cells,
        transfer_summary,
        _deep_scope_label,
        iam_operator_file=_iam_operator_file,
    )
    if deep_benchmark is not None:
        deep_benchmark["rawParsingQa"] = raw_parsing_qa

    ema_executive_summary = _nemo_build_ema_executive_summary(
        operators_payload, professional_executive_summary, all_transfer_sessions
    )

    return {
        "name": "Nemo TXT Benchmark",
        "parserVersion": _BENCHMARK_NEMO_PARSER_VERSION,
        "analysisVersion": _BENCHMARK_NEMO_ANALYSIS_VERSION,
        "dlMode": dl_mode,
        "dlModeLabel": dl_mode_label,
        "windowMode": window_mode,
        "windowModeLabel": window_mode_label,
        "windowFallback": window_fallback,
        "windowFallbackNote": (
            "No active-DL session detected in this scope — showing all-session KPIs."
            if window_fallback else ""
        ),
        "scorecardConfidence": {
            "confidenceLevel": scorecard_confidence_level,
            "reason": scorecard_confidence_reason,
        },
        "deviceParity": {
            "available": bool(device_parity_models),
            "allSame": bool(device_parity_models) and len(device_parity_unique) <= 1,
            "models": device_parity_models,
            "warning": device_parity_warning,
        },
        "benchmarkValidity": benchmark_validity,
        "methodologyNote": methodology_note,
        "rawParsingQa": raw_parsing_qa,
        "sourceFiles": [{"operator": item.get("operator"), "fileName": item.get("fileName"), "path": item.get("path")} for item in prepared_operator_files],
        "deepBenchmark": deep_benchmark,
        "macroContext": {
            "causalChain": (
                ((deep_benchmark or {}).get("execSummary") or {}).get("causalChain")
                or {}
            ),
            "deviceByOperator": device_by_operator,
        },
        "operatorCount": len(operators_payload),
        "testCount": len(all_tests),
        "transferSessionCount": len(all_transfer_sessions),
        "dtList": dt_list,
        "operators": operators_payload,
        "ranking": ranking,
        "tests": all_tests,
        "transferSessions": all_transfer_sessions,
        "transferSummary": transfer_summary,
        "diagnosis": diagnosis,
        "bestDlOperator": best_dl_entry.get("operator") if best_dl_entry else "",
        "best5gComparator": diagnosis.get("best5gComparator") or "",
        "rankingInterpretation": {
            "title": "DL Throughput Ranking Interpretation",
            "title_fr": "Interprétation du classement DL",
            "summary": ranking_summary,
            "summary_fr": ranking_summary_fr,
            "dlMode": dl_mode,
            "dlModeLabel": dl_mode_label,
            "windowMode": window_mode,
            "windowModeLabel": window_mode_label,
            "bestDlOperator": best_dl_entry.get("operator") if best_dl_entry else "",
            "bestDlAvgMbps": (best_dl_entry or {}).get("avgDlMbps"),
            "iamRank": (iam_entry or {}).get("rank"),
            "iamAvgMbps": (iam_entry or {}).get("avgDlMbps"),
            "iamGapVsBestDlPct": diagnosis.get("gapToBestDlPct"),
            "no5gNote": no_5g_note,
            "no5gNote_fr": no_5g_note_fr,
        },
        "technologyStatus": {
            "title": "5G Availability / Technology Status",
            "operators": [item.get("technologyStatus") or {} for item in prepared_operator_files],
        },
        "iamVsBest5gComparator": iam_vs_best_5g,
        "iamWeaknessEvidenceChain": weakness_chain,
        "rfExclusionCheck": rf_exclusion,
        "schedulerPrbDeepDive": scheduler_prb_deep_dive,
        "pdschModulationEfficiencyAnalysis": pdsch_modulation_efficiency,
        "nrBandExposureAnalysis": nr_band_exposure,
        "caScellsAnalysis": ca_scells_analysis,
        "mimoRankAnalysis": mimo_rank_analysis,
        "blerRetxAnalysis": bler_retx_analysis,
        "transportGapAnalysis": transport_gap_analysis,
        "perCgpsWeaknessAnalysis": per_cgps_weakness_analysis,
        "missingKpiQuality": missing_kpi_quality,
        "iamServingCells": iam_serving_cells,
        "inwiServingCells": inwi_serving_cells,
        "orangeServingCells": orange_serving_cells,
        "layerThroughputAnalysis": layer_throughput_analysis,
        "recommendationsByPriority": recommendations_by_priority,
        "executiveConclusion": executive_conclusion,
        "professionalExecutiveSummary": professional_executive_summary,
        "emaExecutiveSummary": ema_executive_summary,
        "validationWarnings": validation_warnings,
        "rulesApplied": rules_applied,
        "qaChecklist": qa_checklist,
        "charts": {
            "ranking": [{"operator": entry.get("operator"), "value": entry.get("avgDlMbps"), "has5g": entry.get("has5g")} for entry in ranking],
            "tests": [{"operator": entry.get("operator"), "label": entry.get("id"), "value": entry.get("avgDlMbps"), "time": entry.get("anchorTime")} for entry in all_tests],
            "dlRankingBar": [{"operator": e.get("operator"), "avg": e.get("avgDlMbps"), "median": e.get("medianDlMbps"), "has5g": e.get("has5g")} for e in ranking],
            "avgVsMedianDl": [{"operator": e.get("operator"), "avg": e.get("avgDlMbps"), "median": e.get("medianDlMbps")} for e in ranking],
            "pdschPrbs": [
                {
                    "operator": item.get("operator"),
                    "avg": (item.get("kpis") or {}).get("prbs", {}).get("average"),
                    "has5g": item.get("has5g"),
                }
                for item in prepared_operator_files if item.get("has5g")
            ],
            "scheduled5gVsDelivered": [
                {
                    "operator": item.get("operator"),
                    "scheduled": (item.get("kpis") or {}).get("scheduled5g", {}).get("average"),
                    "delivered": (item.get("kpis") or {}).get("pdsch5g", {}).get("average"),
                }
                for item in prepared_operator_files if item.get("has5g")
            ],
            "rfComparison": [
                {
                    "operator": item.get("operator"),
                    "rsrp": (item.get("kpis") or {}).get("rsrp", {}).get("median"),
                    "sinr": (item.get("kpis") or {}).get("sinr", {}).get("median"),
                    "cqi": (item.get("kpis") or {}).get("cqi", {}).get("median"),
                }
                for item in prepared_operator_files if item.get("has5g")
            ],
            "mimoRankDist": [
                {
                    "operator": item.get("operator"),
                    "ri1": (item.get("kpis") or {}).get("ri1Share"),
                    "ri2": None,
                    "ri3plus": (item.get("kpis") or {}).get("riGe3Share"),
                }
                for item in prepared_operator_files if item.get("has5g")
            ],
            "rootCauseScore": [
                {"cause": s.get("cause"), "score": s.get("score")}
                for s in (diagnosis.get("scores") or [])
            ],
            "dlTimelineByMetric": dl_timeline_by_metric,
        },
    }


def _benchmark_nemo_path_to_run_parsed(path: str) -> dict:
    """Convert a single combined Nemo benchmark TXT file to the parsed dict
    expected by register_nemo_lte_run().

    The benchmark TXT format has MULTIPLE rows per second — one per cell type
    (LTE Anchor, NR SCG PSCell, etc.).  We must bucket rows by 1-s timestamp
    and extract LTE / NR metrics from the appropriate typed rows only, exactly
    as _nemo_build_serving_cells does, to avoid metric collisions."""
    from datetime import timezone as _tz

    op_data = _nemo_parse_operator_file(path)
    rows = op_data.get("rows") or []
    titles = op_data.get("measurementTitles") or []
    name = titles[0] if titles else os.path.basename(path)
    has_nr = op_data.get("has5g", False)

    # Metric name constants — same strings as nemo_lte_importer.py
    _LTE_RSRP   = "Radio.Lte.ServingCell[8].Rsrp"
    _LTE_RSRQ   = "Radio.Lte.ServingCell[8].Rsrq"
    _LTE_SINR   = "Radio.Lte.ServingCell[8].RsSinr"
    _LTE_PCI    = "Radio.Lte.ServingCell[8].Pci"
    _LTE_EARFCN = "Radio.Lte.ServingCell[8].Downlink.Earfcn"
    _LTE_BAND   = "Radio.Lte.ServingCell[8].Band"
    _NR_RSRP    = "Radio.Nr.ServingCell[16].SsRsrp"
    _NR_RSRQ    = "Radio.Nr.ServingCell[16].SsRsrq"
    _NR_SINR    = "Radio.Nr.ServingCell[16].SsSinr"
    _NR_PCI     = "Radio.Nr.ServingCell[16].Pci"
    _NR_ARFCN   = "Radio.Nr.ServingCell[16].Downlink.NrArfcn"
    _APP_DL     = "App.DL.Throughput.Mbps"
    _MAC_DL     = "MAC.DL.Throughput.Mbps"

    _LTE_TYPES = {"lte serving", "lte anchor"}
    _NR_TYPES  = {"nr serving", "nr scg pscell", "scg pscell", "5g serving"}

    def _is_type(row_cell_types, type_set):
        return any(str(ct or "").strip().lower() in type_set for ct in (row_cell_types or []))

    def _metric_key(row, is_nr=False):
        pci = row.get("pci")
        arfcn = row.get("nrChannelNumber") if is_nr else row.get("lteChannelNumber")
        band = str(row.get("band") or "").strip().lower()
        return (
            int(round(float(pci))) if pci is not None else None,
            int(round(float(arfcn))) if arfcn is not None else None,
            band or None,
        )

    def _merge_snapshot(rows_subset, is_nr=False):
        if not rows_subset:
            return None
        grouped = {}
        order = []
        for row in rows_subset:
            key = _metric_key(row, is_nr=is_nr)
            if key not in grouped:
                grouped[key] = []
                order.append(key)
            grouped[key].append(row)
        best_key = None
        best_score = None
        for key in order:
            key_rows = grouped.get(key) or []
            metric_hits = 0
            for row in key_rows:
                if row.get("rsrp") is not None:
                    metric_hits += 1
                if row.get("rsrq") is not None:
                    metric_hits += 1
                if row.get("sinr") is not None:
                    metric_hits += 1
            populated_identity = sum(1 for value in key if value not in (None, ""))
            exact_primary_hits = 0
            if is_nr:
                exact_primary_hits = sum(
                    1
                    for row in key_rows
                    if any(str(ct or "").strip().upper() == "SCG PSCELL" for ct in (row.get("cellTypes") or []))
                )
            has_useful_metrics = 1 if metric_hits > 0 else 0
            score = (has_useful_metrics, metric_hits, populated_identity, exact_primary_hits, len(key_rows))
            if best_score is None or score > best_score:
                best_score = score
                best_key = key
        selected = grouped.get(best_key) or rows_subset
        snapshot = {
            "pci": best_key[0] if best_key else None,
            "arfcn": best_key[1] if best_key else None,
            "band": best_key[2].upper() if best_key and best_key[2] else None,
            "rsrp": None,
            "rsrq": None,
            "sinr": None,
        }
        for row in selected:
            if snapshot["pci"] is None and row.get("pci") is not None:
                snapshot["pci"] = int(round(float(row.get("pci"))))
            raw_arfcn = row.get("nrChannelNumber") if is_nr else row.get("lteChannelNumber")
            if snapshot["arfcn"] is None and raw_arfcn is not None:
                snapshot["arfcn"] = int(round(float(raw_arfcn)))
            if not snapshot["band"] and row.get("band"):
                snapshot["band"] = str(row.get("band")).strip().upper()
            if snapshot["rsrp"] is None and row.get("rsrp") is not None:
                snapshot["rsrp"] = float(row.get("rsrp"))
            if snapshot["rsrq"] is None and row.get("rsrq") is not None:
                snapshot["rsrq"] = float(row.get("rsrq"))
            if snapshot["sinr"] is None and row.get("sinr") is not None:
                snapshot["sinr"] = float(row.get("sinr"))
        return snapshot

    # Group rows by exact timestamp key. Static Nemo benchmarks can contain
    # several distinct RF snapshots inside the same second, and they must stay
    # separate (for example 14:17:53.208 vs 14:17:53.723).
    from collections import defaultdict as _dd
    buckets: dict = _dd(list)
    for row in rows:
        dt = row.get("_dt")
        if dt is None:
            continue
        buckets[dt].append(row)

    kpi_samples: list = []
    track_points: list = []
    lte_serving_snapshots: list = []
    nr_serving_snapshots: list = []
    min_t = max_t = None
    last_known_lat = None
    last_known_lon = None
    last_known_t_ms = None

    for sec_key in sorted(buckets):
        bucket = buckets[sec_key]
        try:
            dt = sec_key
            t_ms = int(dt.replace(tzinfo=_tz.utc).timestamp() * 1000) if dt.tzinfo is None else int(dt.timestamp() * 1000)
        except Exception:
            continue

        if min_t is None or t_ms < min_t:
            min_t = t_ms
        if max_t is None or t_ms > max_t:
            max_t = t_ms

        t_iso = bucket[0].get("time") or sec_key.isoformat()

        def _emit(metric_name, value_num=None, value_str=None, _t=t_ms, _iso=t_iso):
            if value_num is not None or value_str is not None:
                kpi_samples.append({"name": metric_name, "value_num": value_num,
                                    "value_str": value_str, "t_ms": _t, "time": _iso, "idx": 0})

        # GPS — use first row in bucket that has coordinates. If the bucket has no
        # fresh GPS sample, carry forward the last known coordinate for a short
        # window so static benchmark timestamps at the same point still get their
        # own map/detail sample instead of borrowing RF from a nearby timestamp.
        bucket_lat = None
        bucket_lon = None
        for row in bucket:
            lat, lon = row.get("lat"), row.get("lon")
            if lat is not None and lon is not None:
                bucket_lat = lat
                bucket_lon = lon
                last_known_lat = lat
                last_known_lon = lon
                last_known_t_ms = t_ms
                break
        if bucket_lat is None and bucket_lon is None and last_known_lat is not None and last_known_lon is not None:
            if last_known_t_ms is None or abs(int(t_ms) - int(last_known_t_ms)) <= 10000:
                bucket_lat = last_known_lat
                bucket_lon = last_known_lon
        if bucket_lat is not None and bucket_lon is not None:
            track_points.append({"time": t_iso, "lat": bucket_lat, "lon": bucket_lon})

        # Merge all same-timestamp rows belonging to the LTE anchor / NR PSCell so
        # split KPI rows (RSRP on one row, RSRQ/SINR on others) become one serving snapshot.
        lte_candidates = [r for r in bucket if _is_type(r.get("cellTypes"), _LTE_TYPES)]
        if not lte_candidates:
            lte_candidates = [r for r in bucket if r.get("lteChannelNumber") is not None or (r.get("pci") is not None and r.get("nrChannelNumber") is None)]
        lte_snapshot = _merge_snapshot(lte_candidates, is_nr=False)

        nr_candidates = [r for r in bucket if _is_type(r.get("cellTypes"), _NR_TYPES)]
        if not nr_candidates and has_nr:
            nr_candidates = [r for r in bucket if r.get("nrChannelNumber") is not None]
        nr_snapshot = _merge_snapshot(nr_candidates, is_nr=True)

        if lte_snapshot is not None:
            lte_serving_snapshots.append({"t_ms": t_ms, **lte_snapshot})
            _emit(_LTE_RSRP, lte_snapshot.get("rsrp"))
            _emit(_LTE_RSRQ, lte_snapshot.get("rsrq"))
            _emit(_LTE_SINR, lte_snapshot.get("sinr"))
            if lte_snapshot.get("pci") is not None:
                _emit(_LTE_PCI, lte_snapshot.get("pci"))
            if lte_snapshot.get("arfcn") is not None:
                _emit(_LTE_EARFCN, lte_snapshot.get("arfcn"))
            if lte_snapshot.get("band"):
                _emit(_LTE_BAND, value_str=str(lte_snapshot["band"]))

        if nr_snapshot is not None:
            nr_serving_snapshots.append({"t_ms": t_ms, **nr_snapshot})
            _emit(_NR_RSRP, nr_snapshot.get("rsrp"))
            _emit(_NR_RSRQ, nr_snapshot.get("rsrq"))
            _emit(_NR_SINR, nr_snapshot.get("sinr"))
            if nr_snapshot.get("pci") is not None:
                _emit(_NR_PCI, nr_snapshot.get("pci"))
            if nr_snapshot.get("arfcn") is not None:
                _emit(_NR_ARFCN, nr_snapshot.get("arfcn"))

        # Throughput — prefer primary/LTE anchor row; fall back to first row with data
        for row in bucket:
            dl = row.get("appDlMbps") or row.get("totalMacDlMbps") or row.get("macDlLteMbps")
            if dl is not None:
                _emit(_APP_DL, dl)
                break
        for row in bucket:
            mac_dl = row.get("totalMacDlMbps") or row.get("macDlLteMbps")
            if mac_dl is not None:
                _emit(_MAC_DL, mac_dl)
                break

        # Preserve already-joined serving snapshots on each track point so Point Details can
        # use real serving identities and avoid falling back to the default chart metric.
        if track_points:
            current_track = track_points[-1]
            if current_track.get("time") == t_iso:
                parsed = current_track.setdefault("parsed", {})
                props = current_track.setdefault("properties", {})
                if lte_snapshot:
                    parsed["serving_lte"] = {
                        "pci": lte_snapshot.get("pci"),
                        "sc": lte_snapshot.get("pci"),
                        "earfcn": lte_snapshot.get("arfcn"),
                        "freq": lte_snapshot.get("arfcn"),
                        "band": lte_snapshot.get("band"),
                        "rsrp": lte_snapshot.get("rsrp"),
                        "rsrq": lte_snapshot.get("rsrq"),
                        "sinr": lte_snapshot.get("sinr"),
                    }
                    if lte_snapshot.get("pci") is not None:
                        current_track["pci"] = lte_snapshot.get("pci")
                        props["Serving PCI"] = lte_snapshot.get("pci")
                    if lte_snapshot.get("arfcn") is not None:
                        current_track["earfcn"] = lte_snapshot.get("arfcn")
                        current_track["freq"] = lte_snapshot.get("arfcn")
                        props["Serving EARFCN"] = lte_snapshot.get("arfcn")
                    if lte_snapshot.get("rsrp") is not None:
                        current_track["rsrp"] = lte_snapshot.get("rsrp")
                        props["Serving RSRP"] = lte_snapshot.get("rsrp")
                    if lte_snapshot.get("rsrq") is not None:
                        current_track["rsrq"] = lte_snapshot.get("rsrq")
                        props["Serving RSRQ"] = lte_snapshot.get("rsrq")
                    if lte_snapshot.get("sinr") is not None:
                        current_track["sinr"] = lte_snapshot.get("sinr")
                        props["Serving SINR"] = lte_snapshot.get("sinr")
                if nr_snapshot:
                    parsed["serving_nr"] = {
                        "pci": nr_snapshot.get("pci"),
                        "arfcn": nr_snapshot.get("arfcn"),
                        "band": nr_snapshot.get("band"),
                        "rsrp": nr_snapshot.get("rsrp"),
                        "rsrq": nr_snapshot.get("rsrq"),
                        "sinr": nr_snapshot.get("sinr"),
                    }
                    if nr_snapshot.get("pci") is not None:
                        current_track[_NR_PCI] = nr_snapshot.get("pci")
                    if nr_snapshot.get("arfcn") is not None:
                        current_track[_NR_ARFCN] = nr_snapshot.get("arfcn")
                    if nr_snapshot.get("rsrp") is not None:
                        current_track[_NR_RSRP] = nr_snapshot.get("rsrp")
                    if nr_snapshot.get("rsrq") is not None:
                        current_track[_NR_RSRQ] = nr_snapshot.get("rsrq")
                    if nr_snapshot.get("sinr") is not None:
                        current_track[_NR_SINR] = nr_snapshot.get("sinr")

    def _nearest_snapshot(snapshots, target_ms, max_delta_ms=4000):
        best = None
        best_delta = None
        for snap in snapshots or []:
            snap_ms = snap.get("t_ms")
            if snap_ms is None:
                continue
            delta = abs(int(snap_ms) - int(target_ms))
            if delta > max_delta_ms:
                continue
            if best_delta is None or delta < best_delta:
                best = snap
                best_delta = delta
        return best

    def _apply_snapshot_to_track(track_point, snapshot, is_nr=False):
        if not track_point or not snapshot:
            return
        parsed = track_point.setdefault("parsed", {})
        props = track_point.setdefault("properties", {})
        if is_nr:
            parsed["serving_nr"] = {
                "pci": snapshot.get("pci"),
                "arfcn": snapshot.get("arfcn"),
                "band": snapshot.get("band"),
                "rsrp": snapshot.get("rsrp"),
                "rsrq": snapshot.get("rsrq"),
                "sinr": snapshot.get("sinr"),
            }
            if snapshot.get("pci") is not None:
                track_point[_NR_PCI] = snapshot.get("pci")
            if snapshot.get("arfcn") is not None:
                track_point[_NR_ARFCN] = snapshot.get("arfcn")
            if snapshot.get("rsrp") is not None:
                track_point[_NR_RSRP] = snapshot.get("rsrp")
            if snapshot.get("rsrq") is not None:
                track_point[_NR_RSRQ] = snapshot.get("rsrq")
            if snapshot.get("sinr") is not None:
                track_point[_NR_SINR] = snapshot.get("sinr")
            return

        parsed["serving_lte"] = {
            "pci": snapshot.get("pci"),
            "sc": snapshot.get("pci"),
            "earfcn": snapshot.get("arfcn"),
            "freq": snapshot.get("arfcn"),
            "band": snapshot.get("band"),
            "rsrp": snapshot.get("rsrp"),
            "rsrq": snapshot.get("rsrq"),
            "sinr": snapshot.get("sinr"),
        }
        if snapshot.get("pci") is not None:
            track_point["pci"] = snapshot.get("pci")
            props["Serving PCI"] = snapshot.get("pci")
        if snapshot.get("arfcn") is not None:
            track_point["earfcn"] = snapshot.get("arfcn")
            track_point["freq"] = snapshot.get("arfcn")
            props["Serving EARFCN"] = snapshot.get("arfcn")
        if snapshot.get("rsrp") is not None:
            track_point["rsrp"] = snapshot.get("rsrp")
            props["Serving RSRP"] = snapshot.get("rsrp")
        if snapshot.get("rsrq") is not None:
            track_point["rsrq"] = snapshot.get("rsrq")
            props["Serving RSRQ"] = snapshot.get("rsrq")
        if snapshot.get("sinr") is not None:
            track_point["sinr"] = snapshot.get("sinr")
            props["Serving SINR"] = snapshot.get("sinr")

    for track_point in track_points:
        try:
            tp_dt = _nemo_parse_time(track_point.get("time"))
            tp_ms = int(tp_dt.replace(tzinfo=_tz.utc).timestamp() * 1000) if tp_dt and tp_dt.tzinfo is None else int(tp_dt.timestamp() * 1000)
        except Exception:
            continue
        lte_snap = _nearest_snapshot(lte_serving_snapshots, tp_ms, max_delta_ms=5000)
        nr_snap = _nearest_snapshot(nr_serving_snapshots, tp_ms, max_delta_ms=0)
        if lte_snap:
            _apply_snapshot_to_track(track_point, lte_snap, is_nr=False)
        if nr_snap:
            _apply_snapshot_to_track(track_point, nr_snap, is_nr=True)

    return {
        "name":         name,
        "kpi_samples":  kpi_samples,
        "track_points": track_points,
        "has_nr":       has_nr,
        "start_time":   min_t,
        "end_time":     max_t,
        "rrc_files":    [],
    }


def _benchmark_nemo_register_iam_run_background(iam_path: str):
    if not iam_path:
        return
    try:
        parsed_iam = _benchmark_nemo_path_to_run_parsed(iam_path)
        if parsed_iam.get("kpi_samples"):
            register_nemo_lte_run(parsed_iam)
    except Exception:
        pass


def _benchmark_nemo_precompute_alt_window_background(dl_mode, window_mode):
    """Build + cache the OTHER window mode (cumulative scope) from the already-parsed
    operator_files so the user's first window-filter toggle is instant instead of paying a
    full rebuild. Best-effort, guarded; only touches the mode_datasets cache dict."""
    try:
        dl_mode = _benchmark_nemo_normalize_dl_mode(dl_mode)
        window_mode = _benchmark_nemo_normalize_window_mode(window_mode)
        alt = "active_dl_session" if window_mode == "all_dt_session" else "all_dt_session"
        operator_files = BENCHMARK_NEMO_DATASET.get("operator_files") or []
        if not operator_files:
            return
        mode_datasets = BENCHMARK_NEMO_DATASET.setdefault("mode_datasets", {})
        alt_key = _benchmark_nemo_mode_cache_key(dl_mode, alt)
        if _benchmark_nemo_dataset_current(mode_datasets.get(alt_key)):
            return
        ds = _benchmark_nemo_build_dataset(operator_files, dl_mode=dl_mode, window_mode=alt)
        if ds.get("operators"):
            mode_datasets[alt_key] = ds
    except Exception:
        pass


def _load_benchmark_nemo_files(
    explicit_paths=None,
    uploaded_hashes: dict | None = None,
    dl_mode: str | None = None,
    window_mode: str | None = None,
) -> dict:
    dl_mode = _benchmark_nemo_normalize_dl_mode(dl_mode)
    window_mode = _benchmark_nemo_normalize_window_mode(window_mode)
    paths = _benchmark_nemo_resolve_paths(explicit_paths)
    if not paths:
        return {"ok": False, "error": "No Nemo TXT benchmark files configured"}
    valid_paths = [path for path in paths if os.path.isfile(path)]
    if not valid_paths:
        return {"ok": False, "error": "No valid Nemo TXT benchmark files were found"}
    valid_mtimes = _benchmark_nemo_collect_mtimes(valid_paths)
    file_metas = _benchmark_nemo_collect_file_meta(valid_paths, uploaded_hashes=uploaded_hashes)
    dataset_key = _benchmark_nemo_dataset_key(file_metas, dl_mode=dl_mode, window_mode=window_mode)

    cached_paths = list(BENCHMARK_NEMO_DATASET.get("paths") or [])
    cached_mtimes = BENCHMARK_NEMO_DATASET.get("path_mtimes") or {}
    cached_dataset = BENCHMARK_NEMO_DATASET.get("data")
    if cached_dataset and cached_paths == valid_paths and cached_mtimes == valid_mtimes:
        mode_datasets = BENCHMARK_NEMO_DATASET.setdefault("mode_datasets", {})
        mode_dataset_ids = BENCHMARK_NEMO_DATASET.setdefault("mode_dataset_ids", {})
        mode_dataset_keys = BENCHMARK_NEMO_DATASET.setdefault("mode_dataset_keys", {})
        cache_key = _benchmark_nemo_mode_cache_key(dl_mode, window_mode)
        cached_mode_dataset = _benchmark_nemo_cache_get(mode_datasets, dl_mode, window_mode)
        if _benchmark_nemo_dataset_current(cached_mode_dataset):
            BENCHMARK_NEMO_DATASET["data"] = cached_mode_dataset
            BENCHMARK_NEMO_DATASET["loaded_at"] = time.time()
            BENCHMARK_NEMO_DATASET["dataset_id"] = _benchmark_nemo_cache_get(mode_dataset_ids, dl_mode, window_mode)
            BENCHMARK_NEMO_DATASET["dataset_key"] = _benchmark_nemo_cache_get(mode_dataset_keys, dl_mode, window_mode) or dataset_key
            BENCHMARK_NEMO_DATASET["dl_mode"] = dl_mode
            BENCHMARK_NEMO_DATASET["window_mode"] = window_mode
            BENCHMARK_NEMO_DATASET.setdefault("dt_datasets", {})
            _benchmark_nemo_save_paths(valid_paths)
            return {
                "ok": True,
                "paths": valid_paths,
                "dataset": cached_mode_dataset,
                "cached": True,
                "persistent": False,
                "datasetId": BENCHMARK_NEMO_DATASET.get("dataset_id"),
                "datasetKey": BENCHMARK_NEMO_DATASET.get("dataset_key") or dataset_key,
            }
        if not BENCHMARK_NEMO_DATASET.get("operator_files"):
            BENCHMARK_NEMO_DATASET["operator_files"] = _benchmark_nemo_parse_operator_files(valid_paths)
        cached_dataset, _rebuilt = _benchmark_nemo_refresh_dataset(
            cached_dataset,
            BENCHMARK_NEMO_DATASET.get("operator_files"),
            dl_mode=dl_mode,
            window_mode=window_mode,
        )
        if _rebuilt:
            BENCHMARK_NEMO_DATASET["data"] = cached_dataset
            BENCHMARK_NEMO_DATASET.setdefault("mode_datasets", {})[cache_key] = cached_dataset
            BENCHMARK_NEMO_DATASET["dt_datasets"] = {}
            try:
                dataset_id = _benchmark_nemo_library_store_dataset(
                    dataset_key, file_metas, cached_dataset, BENCHMARK_NEMO_DATASET.get("operator_files")
                )
                BENCHMARK_NEMO_DATASET.setdefault("mode_dataset_ids", {})[cache_key] = dataset_id
            except Exception:
                pass
        else:
            BENCHMARK_NEMO_DATASET.setdefault("mode_datasets", {})[cache_key] = cached_dataset
        BENCHMARK_NEMO_DATASET.setdefault("dt_datasets", {})
        BENCHMARK_NEMO_DATASET.setdefault("mode_dataset_keys", {})[cache_key] = dataset_key
        BENCHMARK_NEMO_DATASET["dataset_key"] = dataset_key
        BENCHMARK_NEMO_DATASET["dataset_id"] = _benchmark_nemo_cache_get(BENCHMARK_NEMO_DATASET.get("mode_dataset_ids"), dl_mode, window_mode)
        BENCHMARK_NEMO_DATASET["dl_mode"] = dl_mode
        BENCHMARK_NEMO_DATASET["window_mode"] = window_mode
        _benchmark_nemo_save_paths(valid_paths)
        return {
            "ok": True,
            "paths": valid_paths,
            "dataset": cached_dataset,
            "cached": True,
            "persistent": False,
            "datasetId": BENCHMARK_NEMO_DATASET.get("dataset_id"),
            "datasetKey": dataset_key,
        }

    persisted = _benchmark_nemo_library_load_dataset_by_key(dataset_key)
    if persisted:
        _benchmark_nemo_library_load_into_memory(persisted)
        return {
            "ok": True,
            "paths": valid_paths,
            "dataset": BENCHMARK_NEMO_DATASET.get("data"),
            "cached": True,
            "persistent": True,
            "datasetId": persisted.get("id"),
            "datasetKey": dataset_key,
        }

    operator_files = _benchmark_nemo_parse_operator_files(valid_paths)
    dataset = _benchmark_nemo_build_dataset(operator_files, dl_mode=dl_mode, window_mode=window_mode)
    if not dataset.get("operators"):
        return {"ok": False, "error": "No usable Nemo TXT benchmark files were parsed"}

    BENCHMARK_NEMO_DATASET["paths"] = valid_paths
    BENCHMARK_NEMO_DATASET["path_mtimes"] = valid_mtimes
    BENCHMARK_NEMO_DATASET["data"] = dataset
    BENCHMARK_NEMO_DATASET["loaded_at"] = time.time()
    BENCHMARK_NEMO_DATASET["operator_files"] = operator_files
    cache_key = _benchmark_nemo_mode_cache_key(dl_mode, window_mode)
    BENCHMARK_NEMO_DATASET["mode_datasets"] = {cache_key: dataset}
    BENCHMARK_NEMO_DATASET["mode_dataset_ids"] = {}
    BENCHMARK_NEMO_DATASET["mode_dataset_keys"] = {cache_key: dataset_key}
    BENCHMARK_NEMO_DATASET["dt_datasets"] = {}
    BENCHMARK_NEMO_DATASET["dataset_key"] = dataset_key
    BENCHMARK_NEMO_DATASET["dl_mode"] = dl_mode
    BENCHMARK_NEMO_DATASET["window_mode"] = window_mode
    dataset_id = _benchmark_nemo_library_store_dataset(dataset_key, file_metas, dataset, operator_files)
    BENCHMARK_NEMO_DATASET["dataset_id"] = dataset_id
    BENCHMARK_NEMO_DATASET["mode_dataset_ids"][cache_key] = dataset_id
    _benchmark_nemo_save_paths(valid_paths)

    return {
        "ok": True,
        "paths": valid_paths,
        "dataset": dataset,
        "cached": False,
        "persistent": False,
        "datasetId": dataset_id,
        "datasetKey": dataset_key,
    }


def _correlate_benchmark_mycom_context() -> dict:
    from datetime import datetime

    nemo_data = BENCHMARK_NEMO_DATASET.get("data") or {}
    mycom_data = BENCHMARK_MYCOM_DATASET.get("data") or {}
    if not nemo_data:
        return {"ok": False, "error": "No Nemo benchmark dataset loaded"}
    if not mycom_data:
        return {"ok": False, "error": "No Mycom hourly dataset loaded"}

    cells = mycom_data.get("cells") or []
    if not cells:
        return {"ok": True, "context": {"available": False, "message": "Mycom dataset contains no cells."}}

    cell_exact = {}
    for cell in cells:
        snorm = _mycom_norm(cell.get("shortName"))
        if snorm and snorm not in cell_exact:
            cell_exact[snorm] = cell

    iam_tests = [item for item in (nemo_data.get("tests") or []) if str(item.get("operator") or "").upper() == "IAM"]
    iam_serving = nemo_data.get("iamServingCells") or {}
    episodes = iam_serving.get("episodes") or []
    cells_payload = iam_serving.get("cells") or []
    top_cell_display = str((((nemo_data.get("layerThroughputAnalysis") or {}).get("iamBreakdown") or {}).get("topCellDisplay")) or "").strip()

    def _hms_to_seconds(text):
        try:
            hh, mm, ss = [int(part) for part in str(text or "").split(":")]
            return hh * 3600 + mm * 60 + ss
        except Exception:
            return None

    def _candidate_episode(anchor_dt):
        if not isinstance(anchor_dt, datetime):
            return None
        anchor_sec = anchor_dt.hour * 3600 + anchor_dt.minute * 60 + anchor_dt.second
        chosen = None
        nearest_gap = None
        for ep in episodes:
            start_sec = _hms_to_seconds(ep.get("startTime"))
            end_sec = _hms_to_seconds(ep.get("endTime"))
            if start_sec is None:
                continue
            if end_sec is None:
                end_sec = start_sec
            if start_sec <= anchor_sec <= end_sec:
                return ep
            gap = min(abs(anchor_sec - start_sec), abs(anchor_sec - end_sec))
            if nearest_gap is None or gap < nearest_gap:
                nearest_gap = gap
                chosen = ep
        return chosen

    def _pick_cell_by_candidates(candidates):
        cleaned = []
        for value in candidates:
            text = str(value or "").strip()
            if text and text not in cleaned:
                cleaned.append(text)
        if not cleaned:
            return None, "", "", []
        for candidate in cleaned:
            norm = _mycom_norm(candidate)
            if norm and norm in cell_exact:
                return cell_exact[norm], "short_name_exact", "high", cleaned
        partials = []
        for candidate in cleaned:
            cnorm = _mycom_norm(candidate)
            if not cnorm:
                continue
            for cell in cells:
                snorm = _mycom_norm(cell.get("shortName"))
                if not snorm:
                    continue
                if cnorm in snorm or snorm in cnorm:
                    partials.append((len(cnorm), cell))
        if partials:
            partials.sort(key=lambda item: item[0], reverse=True)
            return partials[0][1], "short_name_partial", "medium", cleaned
        return None, "", "", cleaned

    matches = []
    unmatched = []
    for test in iam_tests:
        anchor_time = str(test.get("anchorTime") or "").strip()
        try:
            anchor_dt = datetime.fromisoformat(anchor_time)
        except Exception:
            anchor_dt = None
        if anchor_dt is None:
            unmatched.append({
                "testId": test.get("id") or "",
                "anchorTime": anchor_time,
                "reason": "Invalid benchmark anchor time.",
            })
            continue
        anchor_hour = anchor_dt.replace(minute=0, second=0, microsecond=0)
        episode = _candidate_episode(anchor_dt)
        candidate_names = []
        if episode:
            candidate_names.extend([episode.get("cellName"), episode.get("siteName"), episode.get("lteAnchor")])
        if top_cell_display:
            candidate_names.append(top_cell_display)
        candidate_names.extend([item.get("cellName") for item in cells_payload[:5]])
        candidate_names.extend([item.get("siteName") for item in cells_payload[:5]])
        mycom_cell, match_method, confidence, final_candidates = _pick_cell_by_candidates(candidate_names)
        if not mycom_cell:
            unmatched.append({
                "testId": test.get("id") or "",
                "anchorTime": anchor_time,
                "reason": "No Mycom cell name matched the current IAM serving-cell candidates.",
                "candidateNames": final_candidates,
            })
            continue
        hour_iso = anchor_hour.isoformat()
        hour_map = {item.get("ts"): item.get("kpis") or {} for item in (mycom_cell.get("hours") or [])}
        kpis = hour_map.get(hour_iso)
        if not kpis:
            unmatched.append({
                "testId": test.get("id") or "",
                "anchorTime": anchor_time,
                "reason": f"No Mycom hourly data found for {hour_iso}.",
                "candidateNames": final_candidates,
                "matchedCell": {
                    "shortName": mycom_cell.get("shortName") or "",
                    "gnbCellKey": mycom_cell.get("gnbCellKey") or "",
                },
            })
            continue
        matches.append({
            "testId": test.get("id") or "",
            "anchorTime": anchor_time,
            "matchedHourTs": hour_iso,
            "avgDlMbps": test.get("avgDlMbps"),
            "servingCellName": episode.get("cellName") if episode else "",
            "servingSiteName": episode.get("siteName") if episode else "",
            "servingBand": episode.get("band") if episode else "",
            "matchMethod": match_method,
            "matchConfidence": confidence,
            "candidateNames": final_candidates,
            "matchedCell": {
                "shortName": mycom_cell.get("shortName") or "",
                "gnbCellKey": mycom_cell.get("gnbCellKey") or "",
            },
            "kpis": kpis,
        })

    context = {
        "available": bool(matches),
        "title": "IAM Cell-Hour Context (Mycom)",
        "sourceFile": mycom_data.get("sourceFile") or "",
        "dateRange": mycom_data.get("dateRange") or {},
        "matchedCount": len(matches),
        "totalIamTests": len(iam_tests),
        "matches": matches,
        "unmatched": unmatched,
        "metrics": mycom_data.get("metrics") or [],
    }
    if not matches:
        context["message"] = "No IAM benchmark test could be correlated with the imported Mycom hourly export."
    return {"ok": True, "context": context}


def _default_statistics_mycom_thresholds() -> dict:
    return {
        "lowAvailabilityPct": 95.0,
        "lowCssrPct": 98.0,
        "highDropPct": 1.0,
        "highPrbUtilPct": 70.0,
        "highPrbUtilBhPct": 85.0,
        "highAvgUsers": 8.0,
        "highPeakUsers": 15.0,
        "lowUserDlMbps": 50.0,
    }


def _merge_statistics_mycom_thresholds(raw) -> dict:
    defaults = _default_statistics_mycom_thresholds()
    merged = dict(defaults)
    if isinstance(raw, dict):
        for key in defaults:
            try:
                value = float(raw.get(key))
                if math.isfinite(value):
                    merged[key] = value
            except Exception:
                continue
    return merged


def _mycom_detect_anomalies(kpis: dict, thresholds: dict) -> list[dict]:
    items = []
    availability = _mycom_parse_numeric((kpis or {}).get("availabilityPct"))
    cssr = _mycom_parse_numeric((kpis or {}).get("cssrPct"))
    drop = _mycom_parse_numeric((kpis or {}).get("dropPct"))
    prb = _mycom_parse_numeric((kpis or {}).get("prbUtilPct"))
    prb_bh = _mycom_parse_numeric((kpis or {}).get("prbUtilBhPct"))
    avg_users = _mycom_parse_numeric((kpis or {}).get("avgUsers"))
    peak_users = _mycom_parse_numeric((kpis or {}).get("peakUsers"))
    user_dl = _mycom_parse_numeric((kpis or {}).get("userDlMbps"))

    if availability is not None and availability < thresholds["lowAvailabilityPct"]:
        items.append({
            "key": "availability_low",
            "label": "Low 5G availability",
            "severity": "Critical" if availability < (thresholds["lowAvailabilityPct"] - 3) else "High",
            "value": availability,
            "threshold": thresholds["lowAvailabilityPct"],
        })
    if cssr is not None and cssr < thresholds["lowCssrPct"]:
        items.append({
            "key": "cssr_low",
            "label": "Low 5G Data CSSR",
            "severity": "High" if cssr < (thresholds["lowCssrPct"] - 1) else "Medium",
            "value": cssr,
            "threshold": thresholds["lowCssrPct"],
        })
    if drop is not None and drop > thresholds["highDropPct"]:
        items.append({
            "key": "drop_high",
            "label": "High 5G drop rate",
            "severity": "Critical" if drop > (thresholds["highDropPct"] * 2) else "High",
            "value": drop,
            "threshold": thresholds["highDropPct"],
        })
    if prb is not None and prb > thresholds["highPrbUtilPct"]:
        items.append({
            "key": "prb_high",
            "label": "High PRB utilization",
            "severity": "High" if prb > (thresholds["highPrbUtilPct"] + 10) else "Medium",
            "value": prb,
            "threshold": thresholds["highPrbUtilPct"],
        })
    if prb_bh is not None and prb_bh > thresholds["highPrbUtilBhPct"]:
        items.append({
            "key": "prb_bh_high",
            "label": "High PRB utilization (BH)",
            "severity": "High" if prb_bh > (thresholds["highPrbUtilBhPct"] + 10) else "Medium",
            "value": prb_bh,
            "threshold": thresholds["highPrbUtilBhPct"],
        })
    if avg_users is not None and avg_users > thresholds["highAvgUsers"]:
        items.append({
            "key": "avg_users_high",
            "label": "High average users",
            "severity": "Medium",
            "value": avg_users,
            "threshold": thresholds["highAvgUsers"],
        })
    if peak_users is not None and peak_users > thresholds["highPeakUsers"]:
        items.append({
            "key": "peak_users_high",
            "label": "High peak users",
            "severity": "Medium",
            "value": peak_users,
            "threshold": thresholds["highPeakUsers"],
        })
    if user_dl is not None and user_dl < thresholds["lowUserDlMbps"]:
        items.append({
            "key": "user_dl_low",
            "label": "Low user DL throughput",
            "severity": "High" if user_dl < (thresholds["lowUserDlMbps"] * 0.5) else "Medium",
            "value": user_dl,
            "threshold": thresholds["lowUserDlMbps"],
        })
    return items


def _mycom_anomaly_bucket(anomaly_key: str) -> str:
    if anomaly_key in {"prb_high", "prb_bh_high", "avg_users_high", "peak_users_high"}:
        return "Load / Congestion"
    if anomaly_key in {"availability_low", "cssr_low", "drop_high"}:
        return "Service Stability"
    if anomaly_key in {"user_dl_low"}:
        return "Capacity / Throughput"
    return "Other"


def _build_statistics_mycom_analysis(thresholds_raw=None, query: str = "") -> dict:
    dataset = BENCHMARK_MYCOM_DATASET.get("data") or {}
    if not dataset:
        return {"ok": False, "error": "No Mycom export loaded"}

    thresholds = _merge_statistics_mycom_thresholds(thresholds_raw)
    cells = dataset.get("cells") or []
    query_text = str(query or "").strip()
    qnorm = _mycom_norm(query_text)

    all_rows = []
    summary_by_type = {}
    category_summary = {}
    cells_ranked = []
    matched_cells = []
    site_candidates = {}

    def _touch_summary(target, anomaly, short_name, hour_ts):
        entry = target.setdefault(anomaly["key"], {
            "key": anomaly["key"],
            "label": anomaly["label"],
            "severity": anomaly["severity"],
            "cells": set(),
            "hours": 0,
        })
        entry["cells"].add(short_name)
        entry["hours"] += 1

    for cell in cells:
        short_name = str(cell.get("shortName") or "").strip()
        gnb_key = str(cell.get("gnbCellKey") or "").strip()
        site_key = _mycom_site_key(short_name)
        site_candidates.setdefault(site_key, []).append(short_name)
        hours = cell.get("hours") or []
        cell_rows = []
        anomaly_counter = {}
        worst_hour = ""
        for hour in hours:
            ts = str(hour.get("ts") or "")
            kpis = hour.get("kpis") or {}
            anomalies = _mycom_detect_anomalies(kpis, thresholds)
            row = {
                "shortName": short_name,
                "gnbCellKey": gnb_key,
                "siteName": site_key,
                "ts": ts,
                "kpis": kpis,
                "anomalies": anomalies,
            }
            all_rows.append(row)
            cell_rows.append(row)
            if anomalies and not worst_hour:
                worst_hour = ts
            for anomaly in anomalies:
                anomaly_counter[anomaly["key"]] = anomaly_counter.get(anomaly["key"], 0) + 1
                _touch_summary(summary_by_type, anomaly, short_name, ts)
                bucket = _mycom_anomaly_bucket(anomaly["key"])
                bucket_entry = category_summary.setdefault(bucket, {"category": bucket, "cells": set(), "hours": 0})
                bucket_entry["cells"].add(short_name)
                bucket_entry["hours"] += 1
        if qnorm:
            snorm = _mycom_norm(short_name)
            site_norm = _mycom_norm(site_key)
            gnorm = _mycom_norm(gnb_key)
            if qnorm in snorm or qnorm in site_norm or qnorm in gnorm:
                matched_cells.append({
                    "shortName": short_name,
                    "gnbCellKey": gnb_key,
                    "siteName": site_key,
                    "rows": cell_rows,
                    "anomalyCounter": anomaly_counter,
                    "worstHour": worst_hour,
                })
        cells_ranked.append({
            "shortName": short_name,
            "gnbCellKey": gnb_key,
            "siteName": site_key,
            "anomalyHours": sum(anomaly_counter.values()),
            "anomalyTypes": len(anomaly_counter),
            "worstHour": worst_hour,
            "anomalyBreakdown": anomaly_counter,
            "availabilityMin": min(
                [_mycom_parse_numeric((item.get("kpis") or {}).get("availabilityPct")) for item in cell_rows if _mycom_parse_numeric((item.get("kpis") or {}).get("availabilityPct")) is not None] or [None]
            ),
            "userDlMin": min(
                [_mycom_parse_numeric((item.get("kpis") or {}).get("userDlMbps")) for item in cell_rows if _mycom_parse_numeric((item.get("kpis") or {}).get("userDlMbps")) is not None] or [None]
            ),
            "prbUtilMax": max(
                [_mycom_parse_numeric((item.get("kpis") or {}).get("prbUtilPct")) for item in cell_rows if _mycom_parse_numeric((item.get("kpis") or {}).get("prbUtilPct")) is not None] or [None]
            ),
        })

    summary_rows = sorted(
        [
            {
                "key": item["key"],
                "label": item["label"],
                "severity": item["severity"],
                "cellsAffected": len(item["cells"]),
                "anomalyHours": item["hours"],
            }
            for item in summary_by_type.values()
        ],
        key=lambda row: (-row["cellsAffected"], -row["anomalyHours"], row["label"]),
    )
    category_rows = sorted(
        [
            {
                "category": item["category"],
                "cellsAffected": len(item["cells"]),
                "anomalyHours": item["hours"],
            }
            for item in category_summary.values()
        ],
        key=lambda row: (-row["cellsAffected"], -row["anomalyHours"], row["category"]),
    )
    ranked_cells = sorted(
        cells_ranked,
        key=lambda row: (-int(row["anomalyHours"] or 0), -int(row["anomalyTypes"] or 0), str(row["shortName"] or "")),
    )[:50]

    search_result = {
        "query": query_text,
        "matchedCount": len(matched_cells),
        "mode": "none",
        "siteName": "",
        "cells": [],
        "selectedCell": None,
    }
    if matched_cells:
        unique_sites = {item["siteName"] for item in matched_cells if item["siteName"]}
        site_hour_map = {}
        for item in matched_cells:
            for row in item["rows"]:
                bucket = site_hour_map.setdefault(row["ts"], {
                    "ts": row["ts"],
                    "availabilityVals": [],
                    "userDlVals": [],
                    "prbVals": [],
                    "prbBhVals": [],
                    "avgUserVals": [],
                    "peakUserVals": [],
                    "cssrVals": [],
                    "dropVals": [],
                    "anomalyLabels": set(),
                    "anomalyKeys": set(),
                })
                kpis = row.get("kpis") or {}
                for key, target in (
                    ("availabilityPct", "availabilityVals"),
                    ("userDlMbps", "userDlVals"),
                    ("prbUtilPct", "prbVals"),
                    ("prbUtilBhPct", "prbBhVals"),
                    ("avgUsers", "avgUserVals"),
                    ("peakUsers", "peakUserVals"),
                    ("cssrPct", "cssrVals"),
                    ("dropPct", "dropVals"),
                ):
                    num = _mycom_parse_numeric(kpis.get(key))
                    if num is not None:
                        bucket[target].append(float(num))
                for anomaly in row["anomalies"]:
                    bucket["anomalyLabels"].add(str(anomaly["label"]))
                    bucket["anomalyKeys"].add(str(anomaly["key"]))

        def _avg(values):
            return round(sum(values) / len(values), 2) if values else None

        site_hours = [
            {
                "ts": ts,
                "kpis": {
                    "availabilityPct": _avg(item["availabilityVals"]),
                    "userDlMbps": _avg(item["userDlVals"]),
                    "prbUtilPct": _avg(item["prbVals"]),
                    "prbUtilBhPct": _avg(item["prbBhVals"]),
                    "avgUsers": _avg(item["avgUserVals"]),
                    "peakUsers": max(item["peakUserVals"]) if item["peakUserVals"] else None,
                    "cssrPct": _avg(item["cssrVals"]),
                    "dropPct": _avg(item["dropVals"]),
                },
                "anomalyLabels": sorted(item["anomalyLabels"]),
                "anomalyKeys": sorted(item["anomalyKeys"]),
            }
            for ts, item in sorted(site_hour_map.items(), key=lambda row: row[0])
        ]
        exact_short = next((item for item in matched_cells if _mycom_norm(item["shortName"]) == qnorm), None)
        if exact_short:
            exact_short["rows"] = sorted(exact_short["rows"], key=lambda row: row["ts"])
            search_result = {
                "query": query_text,
                "matchedCount": len(matched_cells),
                "mode": "cell",
                "siteName": exact_short["siteName"],
                "cells": [{
                    "shortName": exact_short["shortName"],
                    "gnbCellKey": exact_short["gnbCellKey"],
                    "siteName": exact_short["siteName"],
                    "anomalyHours": sum(exact_short["anomalyCounter"].values()),
                    "anomalyTypes": len(exact_short["anomalyCounter"]),
                    "worstHour": exact_short["worstHour"],
                    "anomalyBreakdown": exact_short["anomalyCounter"],
                }],
                "selectedCell": {
                    "shortName": exact_short["shortName"],
                    "gnbCellKey": exact_short["gnbCellKey"],
                    "siteName": exact_short["siteName"],
                    "hours": [
                        {
                            "ts": row["ts"],
                            "kpis": row["kpis"],
                            "anomalyLabels": [item["label"] for item in row["anomalies"]],
                            "anomalyKeys": [item["key"] for item in row["anomalies"]],
                        }
                        for row in exact_short["rows"]
                    ],
                },
            }
        else:
            site_name = next(iter(unique_sites)) if len(unique_sites) == 1 else query_text
            search_result = {
                "query": query_text,
                "matchedCount": len(matched_cells),
                "mode": "site" if len(unique_sites) <= 1 else "multi",
                "siteName": site_name,
                "cells": sorted(
                    [
                        {
                            "shortName": item["shortName"],
                            "gnbCellKey": item["gnbCellKey"],
                            "siteName": item["siteName"],
                            "anomalyHours": sum(item["anomalyCounter"].values()),
                            "anomalyTypes": len(item["anomalyCounter"]),
                            "worstHour": item["worstHour"],
                            "anomalyBreakdown": item["anomalyCounter"],
                        }
                        for item in matched_cells
                    ],
                    key=lambda row: (-int(row["anomalyHours"] or 0), str(row["shortName"] or "")),
                ),
                "selectedCell": {
                    "shortName": site_name,
                    "gnbCellKey": "",
                    "siteName": site_name,
                    "hours": site_hours,
                },
            }

    total_cells_with_anomalies = len({row["shortName"] for row in all_rows if row["anomalies"]})
    total_anomaly_hours = sum(1 for row in all_rows if row["anomalies"])

    return {
        "ok": True,
        "analysis": {
            "name": "Mycom Statistic Analysis",
            "sourceFile": dataset.get("sourceFile") or "",
            "dateRange": dataset.get("dateRange") or {},
            "cellCount": int(dataset.get("cellCount") or 0),
            "hourCount": int(dataset.get("hourCount") or 0),
            "metricCount": int(len(dataset.get("metrics") or [])),
            "thresholds": thresholds,
            "summary": {
                "cellsWithAnomalies": total_cells_with_anomalies,
                "anomalyHours": total_anomaly_hours,
                "rowsAnalyzed": len(all_rows),
            },
            "anomalyTypes": summary_rows,
            "anomalyCategories": category_rows,
            "topAffectedCells": ranked_cells,
            "search": search_result,
        },
    }


def _json_safe(value):
    from datetime import date as _date_class
    from datetime import datetime as _datetime_class

    if isinstance(value, _datetime_class):
        return value.isoformat()
    if isinstance(value, _date_class):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in list(value.items())}
    if isinstance(value, list):
        return [_json_safe(v) for v in list(value)]
    if isinstance(value, tuple):
        return [_json_safe(v) for v in value]
    if isinstance(value, set):
        return [_json_safe(v) for v in list(value)]
    if isinstance(value, (bytes, bytearray)):
        return bytes(value).hex()
    return value


def _json(handler: SimpleHTTPRequestHandler, obj, status: int = 200):
    body = json.dumps(_json_safe(obj)).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json")
    handler.send_header("Content-Length", str(len(body)))
    handler.send_header("Cache-Control", "no-store")
    request_id = getattr(handler, "request_id", None)
    if request_id:
        handler.send_header("X-Request-Id", str(request_id))
    handler.end_headers()
    handler.wfile.write(body)


def _read_body(handler: SimpleHTTPRequestHandler) -> bytes:
    clen = handler.headers.get("Content-Length")
    if not clen:
        return b""
    try:
        n = int(clen)
    except Exception:
        n = 0
    if n <= 0:
        return b""
    return handler.rfile.read(n)


def _parse_multipart_file_info(body: bytes, content_type: str) -> tuple[str, bytes, dict]:
    """
    Extremely small multipart/form-data parser.
    Assumes a single file part with a filename.
    """
    # content-type: multipart/form-data; boundary=----WebKitFormBoundary...
    m = None
    for part in content_type.split(";"):
        part = part.strip()
        if part.startswith("boundary="):
            m = part.split("=", 1)[1]
            break
    if not m:
        raise ValueError("Missing multipart boundary")

    boundary = ("--" + m).encode("utf-8")
    sections = body.split(boundary)
    for sec in sections:
        sec = sec.strip()
        if not sec or sec == b"--":
            continue
        # headers/body split
        if b"\r\n\r\n" not in sec:
            continue
        head, data = sec.split(b"\r\n\r\n", 1)
        head_txt = head.decode("utf-8", errors="replace")
        if "filename=" not in head_txt:
            continue
        # filename
        fnm = ""
        for line in head_txt.split("\r\n"):
            if line.lower().startswith("content-disposition:"):
                # Content-Disposition: form-data; name="file"; filename="x.trp"
                mm = line.split("filename=", 1)[1].strip()
                if mm.startswith('"') and '"' in mm[1:]:
                    fnm = mm.split('"', 2)[1]
                else:
                    fnm = mm.split(";", 1)[0].strip()
        # strip last CRLF and possible trailing --
        if data.endswith(b"\r\n"):
            data = data[:-2]
        if data.endswith(b"--"):
            data = data[:-2]
        part_headers = {}
        for line in head_txt.split("\r\n"):
            if ":" not in line:
                continue
            key, value = line.split(":", 1)
            part_headers[key.strip().lower()] = value.strip()
        return fnm or "upload.trp", data, part_headers

    raise ValueError("No file part found")


def _parse_multipart_file(body: bytes, content_type: str) -> tuple[str, bytes]:
    filename, data, _headers = _parse_multipart_file_info(body, content_type)
    return filename, data


def _parse_multipart_all_files(body: bytes, content_type: str) -> list[tuple[str, bytes]]:
    """Return list of (filename, data) for every file part in a multipart body."""
    m = None
    for part in content_type.split(";"):
        part = part.strip()
        if part.startswith("boundary="):
            m = part.split("=", 1)[1]
            break
    if not m:
        raise ValueError("Missing multipart boundary")
    boundary = ("--" + m).encode("utf-8")
    results = []
    for sec in body.split(boundary):
        sec = sec.strip()
        if not sec or sec == b"--":
            continue
        if b"\r\n\r\n" not in sec:
            continue
        head, data = sec.split(b"\r\n\r\n", 1)
        head_txt = head.decode("utf-8", errors="replace")
        if "filename=" not in head_txt:
            continue
        fnm = ""
        for line in head_txt.split("\r\n"):
            if line.lower().startswith("content-disposition:"):
                mm = line.split("filename=", 1)[1].strip()
                fnm = mm.split('"', 2)[1] if mm.startswith('"') and '"' in mm[1:] else mm.split(";", 1)[0].strip()
        if data.endswith(b"\r\n"):
            data = data[:-2]
        if data.endswith(b"--"):
            data = data[:-2]
        results.append((fnm or "upload.txt", data))
    return results


def _parse_json_body(handler: SimpleHTTPRequestHandler) -> dict:
    raw = _read_body(handler)
    if not raw:
        return {}
    try:
        data = json.loads(raw.decode("utf-8", errors="replace"))
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}


def _read_nmfs_config_file() -> dict:
    try:
        if not os.path.isfile(NMFS_CONFIG_PATH):
            return {}
        with open(NMFS_CONFIG_PATH, "r", encoding="utf-8", errors="replace") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _make_request_id() -> str:
    return f"req-{uuid.uuid4().hex[:12]}"


def _request_log(handler: SimpleHTTPRequestHandler, message: str, **fields):
    request_id = getattr(handler, "request_id", "-")
    parts = [f"request_id={request_id}", f"event={message}"]
    for key, value in fields.items():
        if value is None:
            continue
        text = str(value)
        if len(text) > 240:
            text = text[:237] + "..."
        parts.append(f"{key}={text}")
    sys.stderr.write("[optim-local-ai] " + " ".join(parts) + "\n")


def _to_num_if_finite(value):
    if value in (None, "", "N/A", "n/a", "-"):
        return None
    try:
        number = float(value)
    except Exception:
        return None
    if number != number or number in (float("inf"), float("-inf")):
        return None
    return number


def _to_int(value):
    number = _to_num_if_finite(value)
    return int(round(number)) if number is not None else None


def _sanitize_lte_earfcn(value):
    number = _to_int(value)
    return number if number is not None and number >= 0 else None


def _parse_event_ts_ms(value):
    if value in (None, "", "N/A", "n/a", "-"):
        return None
    number = _to_num_if_finite(value)
    if number is not None:
        return int(round(number))
    text = str(value).strip()
    try:
        from datetime import datetime
        return int(datetime.fromisoformat(text.replace("Z", "+00:00")).timestamp() * 1000)
    except Exception:
        pass
    import re
    match = re.match(r"^(\d{1,2}):(\d{2}):(\d{2})(?:[.,](\d{1,3}))?$", text)
    if not match:
        return None
    hh = int(match.group(1))
    mm = int(match.group(2))
    ss = int(match.group(3))
    ms = int((match.group(4) or "0").ljust(3, "0")[:3])
    return (((hh * 60 + mm) * 60 + ss) * 1000) + ms


def _format_ms_of_day(value):
    ts = _parse_event_ts_ms(value)
    if ts is None:
        return "N/A"
    ms_of_day = ((ts % 86400000) + 86400000) % 86400000
    hh = str(ms_of_day // 3600000).zfill(2)
    mm = str((ms_of_day % 3600000) // 60000).zfill(2)
    ss = str((ms_of_day % 60000) // 1000).zfill(2)
    ms = str(ms_of_day % 1000).zfill(3)
    return f"{hh}:{mm}:{ss}.{ms}"


def _render_a3_resolver_summary(resolver):
    if not isinstance(resolver, dict):
        return ""
    rows = resolver.get("a3Resolvers")
    if not isinstance(rows, list) or not rows:
        return ""
    rendered = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        report_cfg = row.get("reportConfig") if isinstance(row.get("reportConfig"), dict) else {}
        meas_object = row.get("measObject") if isinstance(row.get("measObject"), dict) else {}
        cell = None
        if isinstance(meas_object.get("cells"), list) and meas_object["cells"]:
            first = meas_object["cells"][0]
            if isinstance(first, dict):
                cell = first
        bits = []
        meas_id = _to_int(row.get("measId"))
        report_config_id = _to_int(row.get("reportConfigId"))
        meas_object_id = _to_int(row.get("measObjectId"))
        carrier_freq = _to_int(meas_object.get("carrierFreq"))
        a3_offset = _to_num_if_finite(report_cfg.get("a3OffsetDb"))
        hysteresis = _to_num_if_finite(report_cfg.get("hysteresisDb"))
        ttt_ms = _to_int(report_cfg.get("timeToTriggerMs"))
        offset_freq = _to_num_if_finite(meas_object.get("offsetFreqDb"))
        if meas_id is not None:
            bits.append(f"measId {meas_id}")
        if report_config_id is not None:
            bits.append(f"reportConfig {report_config_id}")
        if meas_object_id is not None:
            bits.append(f"measObject {meas_object_id}")
        if carrier_freq is not None:
            bits.append(f"EARFCN {carrier_freq}")
        if a3_offset is not None:
            bits.append(f"A3 offset {a3_offset} dB")
        if hysteresis is not None:
            bits.append(f"Hys {hysteresis} dB")
        if ttt_ms is not None:
            bits.append(f"TTT {ttt_ms} ms")
        if offset_freq is not None:
            bits.append(f"offsetFreq {offset_freq} dB")
        if isinstance(cell, dict):
            pci = _to_int(cell.get("physCellId"))
            cio = _to_num_if_finite(cell.get("cellIndividualOffsetDb"))
            if pci is not None:
                bits.append(f"PCI {pci} CIO {cio if cio is not None else 0} dB")
        if bits:
            rendered.append(" | ".join(bits))
    return " || ".join(rendered)


def _resolve_exact_a3_for_measurement_report(mr_row, recfg_rows):
    if not isinstance(mr_row, dict):
        return None
    decoded_mr = mr_row.get("decoded")
    if not isinstance(decoded_mr, dict):
        return None
    summary = decoded_mr.get("summary") if isinstance(decoded_mr.get("summary"), dict) else {}
    meas_id = _to_int(summary.get("measId"))
    if meas_id is None:
        return None
    mr_ts = _parse_event_ts_ms(mr_row.get("ts"))
    matching_recfg = None
    best_rank = None
    for row in recfg_rows or []:
        if not isinstance(row, dict):
            continue
        resolver = row.get("resolver") if isinstance(row.get("resolver"), dict) else {}
        a3_rows = resolver.get("a3Resolvers") if isinstance(resolver.get("a3Resolvers"), list) else []
        if not any(_to_int(item.get("measId")) == meas_id for item in a3_rows if isinstance(item, dict)):
            continue
        row_ts = _parse_event_ts_ms(row.get("ts"))
        if mr_ts is not None and row_ts is not None:
            rank = (0 if row_ts <= mr_ts else 1, abs(mr_ts - row_ts))
        elif row_ts is not None:
            rank = (0, row_ts)
        else:
            rank = (2, 10**15)
        if best_rank is None or rank < best_rank:
            best_rank = rank
            matching_recfg = row
    if not matching_recfg:
        return None
    resolver = matching_recfg.get("resolver") if isinstance(matching_recfg.get("resolver"), dict) else {}
    a3_resolver = next(
        (item for item in resolver.get("a3Resolvers", []) if isinstance(item, dict) and _to_int(item.get("measId")) == meas_id),
        None
    )
    if not isinstance(a3_resolver, dict):
        return None
    report_cfg = a3_resolver.get("reportConfig") if isinstance(a3_resolver.get("reportConfig"), dict) else {}
    meas_object = a3_resolver.get("measObject") if isinstance(a3_resolver.get("measObject"), dict) else {}
    trigger_quantity = str(report_cfg.get("triggerQuantity") or "RSRP").upper()
    measurement_key = "rsrq_db" if trigger_quantity == "RSRQ" else "rsrp_dbm"
    serving = decoded_mr.get("serving") if isinstance(decoded_mr.get("serving"), dict) else {}
    neighbor_rows = decoded_mr.get("neighbors_lte") if isinstance(decoded_mr.get("neighbors_lte"), list) else []
    source_pci = _to_int(mr_row.get("servingPci"))
    source_earfcn = _sanitize_lte_earfcn(mr_row.get("servingEarfcn"))
    Ms = _to_num_if_finite(serving.get(measurement_key))
    Ofn = _to_num_if_finite(meas_object.get("offsetFreqDb"))
    inferred_serving_freq = _sanitize_lte_earfcn(meas_object.get("carrierFreq"))
    same_freq_assumption = True if source_earfcn is None or inferred_serving_freq is None else source_earfcn == inferred_serving_freq
    Ofs = Ofn if same_freq_assumption else 0
    cells = meas_object.get("cells") if isinstance(meas_object.get("cells"), list) else []
    serving_cfg = next((cell for cell in cells if isinstance(cell, dict) and _to_int(cell.get("physCellId")) == source_pci), None)
    Ocs = _to_num_if_finite(serving_cfg.get("cellIndividualOffsetDb")) if isinstance(serving_cfg, dict) else None
    Off = _to_num_if_finite(report_cfg.get("a3OffsetDb"))
    Hys = _to_num_if_finite(report_cfg.get("hysteresisDb"))
    assumptions = []
    if Ofn is None:
        assumptions.append("Neighbor offsetFreq not configured; assumed 0 dB.")
    if Ofs is None:
        assumptions.append("Serving offsetFreq unavailable; assumed 0 dB.")
    elif same_freq_assumption and source_earfcn is not None and inferred_serving_freq is not None and source_earfcn == inferred_serving_freq:
        assumptions.append("Serving offsetFreq assumed same as target measObject (same-frequency).")
    if Ocs is None:
        assumptions.append("Serving CIO not configured in measObject; assumed 0 dB.")
    if Off is None:
        assumptions.append("A3 offset unavailable in reportConfig.")
    if Hys is None:
        assumptions.append("A3 hysteresis unavailable in reportConfig.")
    if Ms is None:
        assumptions.append(f"Serving {trigger_quantity} unavailable in MeasurementReport.")
    evaluated_neighbors = []
    for row in neighbor_rows:
        if not isinstance(row, dict):
            continue
        pci = _to_int(row.get("pci"))
        Mn = _to_num_if_finite(row.get(measurement_key))
        cell_cfg = next((cell for cell in cells if isinstance(cell, dict) and _to_int(cell.get("physCellId")) == pci), None)
        Ocn = _to_num_if_finite(cell_cfg.get("cellIndividualOffsetDb")) if isinstance(cell_cfg, dict) else None
        lhs_enter = None if Mn is None else Mn + (Ofn if Ofn is not None else 0) + (Ocn if Ocn is not None else 0) - (Hys if Hys is not None else 0)
        rhs = None if Ms is None else Ms + (Ofs if Ofs is not None else 0) + (Ocs if Ocs is not None else 0) + (Off if Off is not None else 0)
        lhs_leave = None if Mn is None else Mn + (Ofn if Ofn is not None else 0) + (Ocn if Ocn is not None else 0) + (Hys if Hys is not None else 0)
        evaluated_neighbors.append({
            "pci": pci,
            "Mn": Mn,
            "Ocn": Ocn if Ocn is not None else 0,
            "lhsEnter": lhs_enter,
            "rhs": rhs,
            "lhsLeave": lhs_leave,
            "enterSatisfied": (lhs_enter > rhs) if lhs_enter is not None and rhs is not None else None,
            "leaveSatisfied": (lhs_leave < rhs) if lhs_leave is not None and rhs is not None else None,
            "deltaVsThreshold": (lhs_enter - rhs) if lhs_enter is not None and rhs is not None else None,
        })
    evaluated_neighbors.sort(key=lambda item: item.get("deltaVsThreshold") if item.get("deltaVsThreshold") is not None else float("-inf"), reverse=True)
    best = evaluated_neighbors[0] if evaluated_neighbors else None
    summary_bits = [f"measId {meas_id}"]
    report_config_id = _to_int(a3_resolver.get("reportConfigId"))
    meas_object_id = _to_int(a3_resolver.get("measObjectId"))
    if report_config_id is not None:
        summary_bits.append(f"reportConfig {report_config_id}")
    if meas_object_id is not None:
        summary_bits.append(f"measObject {meas_object_id}")
    if inferred_serving_freq is not None:
        summary_bits.append(f"EARFCN {inferred_serving_freq}")
    if Off is not None:
        summary_bits.append(f"Off {Off:.1f} dB")
    if Hys is not None:
        summary_bits.append(f"Hys {Hys:.1f} dB")
    ttt_ms = _to_int(report_cfg.get("timeToTriggerMs"))
    if ttt_ms is not None:
        summary_bits.append(f"TTT {ttt_ms} ms")
    unit = "dB" if trigger_quantity == "RSRQ" else "dBm"
    evaluation_summary = (
        f"PCI {best.get('pci') if best and best.get('pci') is not None else '?'}: "
        f"LHSenter {best.get('lhsEnter'):.1f} {unit} vs RHS {best.get('rhs'):.1f} {unit} => "
        f"enter {'true' if best.get('enterSatisfied') else 'false'}"
    ) if best and best.get("lhsEnter") is not None and best.get("rhs") is not None else "No LTE neighbors available for A3 evaluation."
    source_pci_from_recfg = _to_int((matching_recfg.get("properties") or {}).get("HO target PCI") or (matching_recfg.get("properties") or {}).get("rrc_recfg_tgt_pci"))
    return {
        "measId": meas_id,
        "triggerQuantity": trigger_quantity,
        "sourceTimeMs": _parse_event_ts_ms(matching_recfg.get("ts")),
        "sourceTimeLabel": _format_ms_of_day(matching_recfg.get("ts")),
        "sourcePci": source_pci_from_recfg,
        "reportConfigId": report_config_id,
        "measObjectId": meas_object_id,
        "carrierFreq": inferred_serving_freq,
        "servingPci": source_pci,
        "servingEarfcn": source_earfcn,
        "servingMetric": Ms,
        "neighborOffsetFreqDb": Ofn if Ofn is not None else 0,
        "servingOffsetFreqDb": Ofs if Ofs is not None else 0,
        "servingCioDb": Ocs if Ocs is not None else 0,
        "a3OffsetDb": Off,
        "hysteresisDb": Hys,
        "timeToTriggerMs": ttt_ms,
        "mappingSummary": " | ".join(summary_bits),
        "evaluationSummary": evaluation_summary,
        "assumptions": assumptions,
        "bestNeighbor": best,
        "neighbors": evaluated_neighbors,
    }


def _lte_rrc_precompute_cache_path(cache_key: str) -> str:
    safe_key = "".join(ch for ch in str(cache_key or "").strip() if ch.isalnum() or ch in ("-", "_"))
    return os.path.join(LTE_RRC_PRECOMPUTE_DIR, f"{safe_key}.json")


def _load_lte_rrc_precompute(cache_key: str):
    if not cache_key:
        return None
    if cache_key in LTE_RRC_PRECOMPUTE_STORE:
        return LTE_RRC_PRECOMPUTE_STORE[cache_key]
    path = _lte_rrc_precompute_cache_path(cache_key)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        LTE_RRC_PRECOMPUTE_STORE[cache_key] = payload
        return payload
    except Exception:
        return None


def _store_lte_rrc_precompute(cache_key: str, payload):
    if not cache_key:
        return
    LTE_RRC_PRECOMPUTE_STORE[cache_key] = payload
    try:
        os.makedirs(LTE_RRC_PRECOMPUTE_DIR, exist_ok=True)
        with open(_lte_rrc_precompute_cache_path(cache_key), "w", encoding="utf-8") as handle:
            json.dump(payload, handle)
    except Exception:
        pass


def _write_nmfs_config_file(data: dict):
    cfg = data if isinstance(data, dict) else {}
    os.makedirs(os.path.dirname(NMFS_CONFIG_PATH), exist_ok=True)
    with open(NMFS_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


def _get_nmfs_effective_config() -> dict:
    file_cfg = _read_nmfs_config_file()
    cmd = os.environ.get("OPTIM_NMFS_CONVERTER_CMD", "").strip() or str(file_cfg.get("converterCmd") or "").strip()
    timeout_env = os.environ.get("OPTIM_NMFS_TIMEOUT_SEC", "").strip()
    keep_env = os.environ.get("OPTIM_NMFS_KEEP_TEMP", "").strip().lower()
    timeout_file = file_cfg.get("timeoutSec")
    keep_file = file_cfg.get("keepTemp")

    try:
        timeout_sec = int(timeout_env) if timeout_env else int(timeout_file if timeout_file is not None else 180)
    except Exception:
        timeout_sec = 180
    timeout_sec = max(10, timeout_sec)

    if keep_env in {"1", "true", "yes"}:
        keep_temp = True
    elif keep_env in {"0", "false", "no"}:
        keep_temp = False
    else:
        keep_temp = bool(keep_file) if keep_file is not None else False

    return {
        "converterCmd": cmd,
        "timeoutSec": timeout_sec,
        "keepTemp": keep_temp,
        "source": {
            "converterCmd": "env" if bool(os.environ.get("OPTIM_NMFS_CONVERTER_CMD", "").strip()) else ("file" if bool(file_cfg.get("converterCmd")) else "default"),
            "timeoutSec": "env" if bool(timeout_env) else ("file" if timeout_file is not None else "default"),
            "keepTemp": "env" if bool(keep_env) else ("file" if keep_file is not None else "default"),
        },
    }


def _validate_nmfs_converter_cfg(cfg: dict) -> dict:
    cmd_tpl = str((cfg or {}).get("converterCmd") or "").strip()
    timeout_sec = int((cfg or {}).get("timeoutSec") or 180)
    keep_temp = bool((cfg or {}).get("keepTemp"))
    issues = []
    warnings = []
    first_bin = None
    executable_found = False

    if not cmd_tpl:
        issues.append("converterCmd is empty.")
    if "{input}" not in cmd_tpl:
        issues.append("converterCmd must include {input} placeholder.")
    if "{output}" not in cmd_tpl:
        warnings.append("converterCmd does not include {output}; converter output discovery will rely on stdout/scan.")

    args = shlex.split(cmd_tpl) if cmd_tpl else []
    if not args:
        issues.append("converterCmd resolves to empty command.")
    else:
        first_bin = args[0]
        if os.path.isabs(first_bin):
            executable_found = os.path.exists(first_bin)
        else:
            executable_found = shutil.which(first_bin) is not None
        if not executable_found:
            warnings.append(f"Command binary not found in PATH: {first_bin}")

    return {
        "ok": len(issues) == 0,
        "issues": issues,
        "warnings": warnings,
        "converterCmd": cmd_tpl,
        "timeoutSec": max(10, timeout_sec),
        "keepTemp": keep_temp,
        "resolvedBinary": first_bin,
        "binaryFound": executable_found,
    }


def _run_nmfs_converter(input_path: str) -> dict:
    """
    Run external NMFS converter command configured by env var:
      OPTIM_NMFS_CONVERTER_CMD
    The command can use placeholders:
      {input}  -> absolute input .nmfs path
      {output} -> suggested output text path (.nmf)
    Example:
      wine /path/AnalyzeParser.exe -i "{input}" -o "{output}"
    """
    eff = _get_nmfs_effective_config()
    cmd_tpl = str(eff.get("converterCmd") or "").strip()
    if not cmd_tpl:
        raise RuntimeError(
            "NMFS converter is not configured. Set OPTIM_NMFS_CONVERTER_CMD "
            "with placeholders {input} and {output}."
        )

    try:
        timeout_sec = max(10, int(eff.get("timeoutSec") or 180))
    except Exception:
        timeout_sec = 180

    keep_temp = bool(eff.get("keepTemp"))
    tmp_dir = tempfile.mkdtemp(prefix="optim_nmfs_")
    out_name = os.path.basename(input_path) + ".nmf"
    output_path = os.path.join(tmp_dir, out_name)

    cmd_text = cmd_tpl.format(input=input_path, output=output_path)
    # Use Windows-compatible argument splitting when server runs on Windows.
    cmd_args = shlex.split(cmd_text, posix=(os.name != "nt"))
    if not cmd_args:
        raise RuntimeError("OPTIM_NMFS_CONVERTER_CMD resolved to an empty command.")

    proc = subprocess.run(
        cmd_args,
        capture_output=True,
        text=True,
        timeout=timeout_sec
    )

    decoded_text = ""
    chosen_output_path = None
    if os.path.isfile(output_path):
        try:
            with open(output_path, "r", encoding="utf-8", errors="replace") as f:
                decoded_text = f.read()
            chosen_output_path = output_path
        except Exception:
            decoded_text = ""

    # If converter ignored {output}, try to discover generated text files.
    if not decoded_text:
        candidates = []
        for root, _dirs, files in os.walk(tmp_dir):
            for fn in files:
                p = os.path.join(root, fn)
                ext = os.path.splitext(fn)[1].lower()
                if ext not in {".nmf", ".txt", ".csv", ".log"}:
                    continue
                try:
                    sz = os.path.getsize(p)
                except Exception:
                    continue
                if sz <= 0:
                    continue
                candidates.append((sz, p))
        if candidates:
            candidates.sort(reverse=True)
            chosen_output_path = candidates[0][1]
            try:
                with open(chosen_output_path, "r", encoding="utf-8", errors="replace") as f:
                    decoded_text = f.read()
            except Exception:
                decoded_text = ""

    # Last resort: some converters print plain text to stdout.
    if not decoded_text and (proc.stdout or "").strip():
        decoded_text = proc.stdout

    result = {
        "command": cmd_text,
        "returncode": int(proc.returncode),
        "stdout": proc.stdout[-4000:] if proc.stdout else "",
        "stderr": proc.stderr[-4000:] if proc.stderr else "",
        "output_path": chosen_output_path,
        "text_len": len(decoded_text),
        "temp_dir": tmp_dir if keep_temp else None,
    }

    if not keep_temp:
        try:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        except Exception:
            pass

    if not decoded_text.strip():
        raise RuntimeError(
            "NMFS converter produced no decodable text output. "
            f"returncode={proc.returncode}. stderr={result['stderr'][:500]}"
        )

    result["text"] = decoded_text
    return result


def _run_ho_analysis(payload: dict) -> dict:
    cli_path = os.path.join(os.path.dirname(__file__), "ho_analysis_cli.js")
    if not os.path.isfile(cli_path):
        raise RuntimeError("ho_analysis_cli.js is missing")
    proc = subprocess.run(
        ["node", cli_path],
        input=json.dumps(payload).encode("utf-8"),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    stdout = (proc.stdout or b"").decode("utf-8", errors="replace")
    stderr = (proc.stderr or b"").decode("utf-8", errors="replace")
    if proc.returncode != 0:
        raise RuntimeError(f"HO analysis failed: {stderr or stdout or proc.returncode}")
    try:
        parsed = json.loads(stdout)
    except Exception as exc:
        raise RuntimeError(f"Invalid HO analysis response: {exc}") from exc
    if not parsed.get("ok"):
        raise RuntimeError(parsed.get("error") or "HO analysis failed")
    return parsed["result"]


def _store_ho_analysis(result: dict, source: dict | None = None) -> str:
    global HO_ANALYSIS_SEQ
    HO_ANALYSIS_SEQ += 1
    analysis_id = f"ho-analysis-{HO_ANALYSIS_SEQ:05d}"
    HO_ANALYSIS_STORE[analysis_id] = {
        "id": analysis_id,
        "createdAt": result.get("generatedAt"),
        "result": result,
        "source": source or {},
    }
    return analysis_id


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=_APP_DIR, **kwargs)

    def log_message(self, format, *args):
        # quieter logs
        request_id = getattr(self, "request_id", "-")
        sys.stderr.write("%s - - [%s] [%s] %s\n" % (self.address_string(), self.log_date_time_string(), request_id, format % args))

    def end_headers(self):
        # Allow frontend and backend on different origins/ports.
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization")
        # Disable caching for all static and API responses to ensure fresh content
        self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(204)
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        self.request_id = _make_request_id()
        started = time.perf_counter()
        _request_log(self, "request_started", method="GET", path=path)

        try:
            # API routes
            if path == "/api/local-ai/health":
                cfg = load_local_ai_config(UPLOAD_DIR)
                client = LMStudioClient(cfg, logger=lambda message, **fields: _request_log(self, message, **fields))
                health = client.health_check()
                _json(self, {"status": "success", "requestId": self.request_id, "health": health})
                return

            if path == "/api/local-ai/model-check":
                cfg = load_local_ai_config(UPLOAD_DIR)
                client = LMStudioClient(cfg, logger=lambda message, **fields: _request_log(self, message, **fields))
                model = client.model_check()
                _json(self, {"status": "success", "requestId": self.request_id, "model": model})
                return

            if path == "/api/nmfs/config":
                cfg = _get_nmfs_effective_config()
                _json(self, {"status": "success", "config": cfg, "configPath": NMFS_CONFIG_PATH})
                return

            if path.startswith("/api/ho-analysis/") or path.startswith("/api/interfreq-ho-analysis/"):
                parts = path.strip("/").split("/")
                if len(parts) < 3:
                    _json(self, {"status": "error", "message": "Bad request"}, 400)
                    return
                analysis_id = parts[2]
                record = HO_ANALYSIS_STORE.get(analysis_id)
                if not record:
                    _json(self, {"status": "error", "message": "HO analysis not found"}, 404)
                    return
                result = record["result"]
                if len(parts) == 4 and parts[3] == "events":
                    qs = parse_qs(parsed.query or "")
                    page = max(1, int((qs.get("page") or ["1"])[0]))
                    page_size = max(1, min(500, int((qs.get("pageSize") or ["100"])[0])))
                    events = list(result.get("events") or [])
                    start = (page - 1) * page_size
                    end = start + page_size
                    _json(self, {
                        "status": "success",
                        "analysisId": analysis_id,
                        "page": page,
                        "pageSize": page_size,
                        "total": len(events),
                        "events": events[start:end],
                    })
                    return
                if len(parts) == 5 and parts[3] == "events":
                    event_id = parts[4]
                    event = next((ev for ev in result.get("events") or [] if str(ev.get("id")) == event_id), None)
                    if not event:
                        _json(self, {"status": "error", "message": "HO event not found"}, 404)
                        return
                    _json(self, {"status": "success", "analysisId": analysis_id, "event": event})
                    return
                if len(parts) == 4 and parts[3] == "kpis":
                    _json(self, {"status": "success", "analysisId": analysis_id, "kpis": result.get("kpis")})
                    return
                if len(parts) == 4 and parts[3] == "export":
                    _json(self, {"status": "success", "analysisId": analysis_id, "result": result})
                    return
                _json(self, {
                    "status": "success",
                    "analysisId": analysis_id,
                    "summary": result.get("kpis", {}).get("summary"),
                    "normalization": result.get("normalization"),
                    "debug": result.get("debug"),
                })
                return

            if path == "/api/runs":
                _json(self, {"status": "success", "runs": list_runs(DB_PATH)})
                return

            if path.startswith("/api/runs/"):
                parts = path.strip("/").split("/")  # ["api","runs","<id>", ...]
                if len(parts) < 3:
                    _json(self, {"status": "error", "message": "Bad request"}, 400)
                    return

                run_id = parts[2]

                # Sub-routes
                if len(parts) == 4 and parts[3] == "catalog":
                    _json(self, fetch_run_catalog(DB_PATH, run_id))
                    return
                if len(parts) == 4 and parts[3] == "sidebar":
                    _json(self, fetch_run_sidebar(DB_PATH, run_id))
                    return
                if len(parts) == 4 and parts[3] == "signals":
                    _json(self, fetch_run_signals(DB_PATH, run_id))
                    return
                if len(parts) == 4 and parts[3] == "track":
                    _json(self, fetch_run_track(DB_PATH, run_id))
                    return
                if len(parts) == 4 and parts[3] == "events":
                    _json(self, fetch_run_events(DB_PATH, run_id))
                    return
                if len(parts) == 4 and parts[3] == "signaling_window_decode":
                    qs = parse_qs(parsed.query or "")
                    time_iso = (qs.get("time") or [""])[0]
                    window_ms = (qs.get("windowMs") or ["30000"])[0]
                    try:
                        window_ms_i = int(window_ms)
                    except Exception:
                        window_ms_i = 30000
                    _json(self, fetch_signaling_window_decode(DB_PATH, run_id, time_iso, window_ms=window_ms_i))
                    return
                if len(parts) == 4 and parts[3] == "kpi":
                    qs = parse_qs(parsed.query or "")
                    name = (qs.get("name") or [""])[0]
                    max_points = (qs.get("max_points") or ["50000"])[0]
                    idx_raw = (qs.get("idx") or [None])[0]
                    try:
                        max_points_i = int(max_points)
                    except Exception:
                        max_points_i = 50000
                    try:
                        idx_i = int(idx_raw) if idx_raw not in (None, "") else None
                    except Exception:
                        idx_i = None
                    _json(self, fetch_kpi_series(DB_PATH, run_id, name, max_points=max_points_i, idx=idx_i))
                    return
                if len(parts) == 4 and parts[3] == "neighbors_at_time":
                    qs = parse_qs(parsed.query or "")
                    time_iso = (qs.get("time") or [""])[0]
                    tol_ms = (qs.get("tolMs") or ["200"])[0]
                    bucket_ms = (qs.get("bucketMs") or ["80"])[0]
                    try:
                        tol_ms_i = int(tol_ms)
                    except Exception:
                        tol_ms_i = 200
                    try:
                        bucket_ms_i = int(bucket_ms)
                    except Exception:
                        bucket_ms_i = 80
                    _json(self, fetch_neighbors_at_time(DB_PATH, run_id, time_iso, tol_ms=tol_ms_i, bucket_ms=bucket_ms_i))
                    return
                if len(parts) == 4 and parts[3] == "mrdc_cells_at_time":
                    qs = parse_qs(parsed.query or "")
                    time_iso = (qs.get("time") or [""])[0]
                    try:
                        tol_ms_i = int((qs.get("tolMs") or ["1500"])[0])
                    except Exception:
                        tol_ms_i = 1500
                    _json(self, fetch_mrdc_cells_at_time(DB_PATH, run_id, time_iso, tol_ms=tol_ms_i))
                    return
                if len(parts) == 4 and parts[3] == "pilot_pollution_at_point":
                    qs = parse_qs(parsed.query or "")
                    time_iso = (qs.get("time") or [""])[0]
                    lat_raw = (qs.get("lat") or [None])[0]
                    lng_raw = (qs.get("lng") or [None])[0]
                    window_ms = (qs.get("windowMs") or ["12000"])[0]
                    rf_expiry_ms = (qs.get("rfExpiryMs") or ["3000"])[0]
                    identity_expiry_ms = (qs.get("identityExpiryMs") or ["10000"])[0]
                    cluster_gap_ms = (qs.get("clusterGapMs") or ["3000"])[0]
                    cluster_gap_m = (qs.get("clusterGapM") or ["100"])[0]
                    try:
                        lat_i = float(lat_raw) if lat_raw not in (None, "") else None
                    except Exception:
                        lat_i = None
                    try:
                        lng_i = float(lng_raw) if lng_raw not in (None, "") else None
                    except Exception:
                        lng_i = None
                    try:
                        window_ms_i = int(window_ms)
                    except Exception:
                        window_ms_i = 12000
                    try:
                        rf_expiry_ms_i = int(rf_expiry_ms)
                    except Exception:
                        rf_expiry_ms_i = 3000
                    try:
                        identity_expiry_ms_i = int(identity_expiry_ms)
                    except Exception:
                        identity_expiry_ms_i = 10000
                    try:
                        cluster_gap_ms_i = int(cluster_gap_ms)
                    except Exception:
                        cluster_gap_ms_i = 3000
                    try:
                        cluster_gap_m_i = float(cluster_gap_m)
                    except Exception:
                        cluster_gap_m_i = 100.0
                    _json(
                        self,
                        analyze_pilot_pollution_at_point(
                            DB_PATH,
                            run_id,
                            time_iso,
                            lat=lat_i,
                            lon=lng_i,
                            window_ms=window_ms_i,
                            rf_expiry_ms=rf_expiry_ms_i,
                            identity_expiry_ms=identity_expiry_ms_i,
                            cluster_gap_ms=cluster_gap_ms_i,
                            cluster_gap_m=cluster_gap_m_i,
                        ),
                    )
                    return
                if len(parts) == 4 and parts[3] == "pcap":
                    try:
                        pcap_data = fetch_run_pcap(DB_PATH, run_id)
                    except KeyError:
                        _json(self, {"error": "Run not found"}, 404)
                        return
                    if not pcap_data:
                        _json(self, {"error": "No IP-sniffer (ch7) data in this TRP file"}, 404)
                        return
                    filename = f"run_{run_id}_ip.pcap"
                    self.send_response(200)
                    self.send_header("Content-Type", "application/vnd.tcpdump.pcap")
                    self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                    self.send_header("Content-Length", str(len(pcap_data)))
                    self.send_header("Access-Control-Allow-Origin", "*")
                    self.end_headers()
                    self.wfile.write(pcap_data)
                    return
                if len(parts) == 4 and parts[3] == "scan_pollution":
                    _json(self, scan_route_pilot_pollution(run_id))
                    return
                if len(parts) == 5 and parts[3] == "l1l2" and parts[4] == "capabilities":
                    _json(self, fetch_l1l2_scheduler_capabilities(DB_PATH, run_id))
                    return
                if len(parts) == 5 and parts[3] == "l1l2" and parts[4] == "at_time":
                    qs = parse_qs(parsed.query or "")
                    time_iso = (qs.get("time") or [""])[0]
                    window_ms = (qs.get("windowMs") or ["2000"])[0]
                    try:
                        window_ms_i = int(window_ms)
                    except Exception:
                        window_ms_i = 2000
                    _json(self, fetch_l1l2_scheduler_at_time(DB_PATH, run_id, time_iso, window_ms=window_ms_i))
                    return
                if len(parts) == 4 and parts[3] == "timeseries":
                    qs = parse_qs(parsed.query or "")
                    signal = (qs.get("signal") or [""])[0]
                    max_points = (qs.get("max_points") or ["50000"])[0]
                    idx_raw = (qs.get("idx") or [None])[0]
                    try:
                        max_points_i = int(max_points)
                    except Exception:
                        max_points_i = 50000
                    try:
                        idx_i = int(idx_raw) if idx_raw not in (None, "") else None
                    except Exception:
                        idx_i = None
                    _json(self, fetch_timeseries_by_signal(DB_PATH, run_id, signal, max_points=max_points_i, idx=idx_i))
                    return

                # Default: run detail
                if len(parts) == 3:
                    run, track, events = fetch_run_detail(DB_PATH, run_id)
                    _json(self, {"status": "success", "run": run, "track_points": track, "events": events})
                    return

                _json(self, {"status": "error", "message": "Not found"}, 404)
                return

            # BDD status / config
            if path == "/api/bdd/status":
                _json(self, {"status": "success", **_bdd.bdd_status()})
                return

            if path == "/api/bdd/config":
                cfg = {}
                if os.path.isfile(BDD_CONFIG_PATH):
                    try:
                        with open(BDD_CONFIG_PATH, "r", encoding="utf-8") as f:
                            cfg = json.load(f)
                    except Exception:
                        pass
                _json(self, {"status": "success", "config": cfg, "bdd": _bdd.bdd_status()})
                return

            if path == "/api/benchmark/status":
                _json(self, {"status": "success", **_benchmark_status_payload()})
                return

            if path == "/api/benchmark-nemo/status":
                _json(self, {"status": "success", **_benchmark_nemo_status_payload()})
                return

            if path == "/api/benchmark-nemo/library":
                _json(self, {"status": "success", "datasets": _benchmark_nemo_library_list()})
                return

            if path.startswith("/api/benchmark-nemo/library/"):
                tail = path[len("/api/benchmark-nemo/library/"):].strip("/")
                if not tail:
                    _json(self, {"status": "error", "message": "Missing dataset id"}, 400)
                    return
                try:
                    dataset_id = int(tail)
                except ValueError:
                    _json(self, {"status": "error", "message": "Invalid dataset id"}, 400)
                    return
                record = _benchmark_nemo_library_load_dataset_by_id(dataset_id)
                if not record:
                    _json(self, {"status": "error", "message": "Dataset not found"}, 404)
                    return
                _json(
                    self,
                    {
                        "status": "success",
                        "dataset": {
                            "id": record.get("id"),
                            "datasetKey": record.get("datasetKey") or "",
                            "name": ((record.get("dataset") or {}).get("name") or "Nemo TXT Benchmark"),
                            "operatorCount": int(((record.get("dataset") or {}).get("operatorCount") or 0)),
                            "testCount": int(((record.get("dataset") or {}).get("testCount") or 0)),
                            "transferSessionCount": int(((record.get("dataset") or {}).get("transferSessionCount") or 0)),
                            "dtList": record.get("dtList") or [],
                            "sourceFiles": record.get("files") or [],
                        },
                    },
                )
                return

            if path == "/api/benchmark-mycom/status":
                _json(self, {"status": "success", **_benchmark_mycom_status_payload()})
                return

            if path == "/api/statistics-mycom/status":
                _json(self, {"status": "success", **_benchmark_mycom_status_payload()})
                return

            if path == "/bdd_sectors.json":
                import gzip as _gzip
                # Regenerate if missing
                if not os.path.isfile(BDD_SECTORS_JSON_PATH):
                    fp = _find_bdd_folder_xlsx()
                    if fp:
                        _write_bdd_sectors_json(fp)
                gz_path = BDD_SECTORS_JSON_PATH + ".gz"
                accepts_gzip = "gzip" in self.headers.get("Accept-Encoding", "")
                serve_path = gz_path if (accepts_gzip and os.path.isfile(gz_path)) else BDD_SECTORS_JSON_PATH
                if os.path.isfile(serve_path):
                    with open(serve_path, "rb") as f:
                        body = f.read()
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.send_header("Content-Length", str(len(body)))
                    self.send_header("Cache-Control", "public, max-age=3600")
                    if serve_path == gz_path:
                        self.send_header("Content-Encoding", "gzip")
                    self.end_headers()
                    self.wfile.write(body)
                    return
                self.send_response(404)
                self.end_headers()
                return

            if path == "/api/bdd/sectors":
                # Regenerate from BDD/ folder if JSON is missing
                if not os.path.isfile(BDD_SECTORS_JSON_PATH):
                    fp = _find_bdd_folder_xlsx()
                    if fp:
                        _write_bdd_sectors_json(fp)
                if os.path.isfile(BDD_SECTORS_JSON_PATH):
                    try:
                        with open(BDD_SECTORS_JSON_PATH, "r", encoding="utf-8") as f:
                            sectors = json.load(f)
                        _json(self, {"status": "success", "ok": True, "sectors": sectors, "count": len(sectors)})
                        return
                    except Exception:
                        pass
                result = _bdd.get_map_sectors()
                _json(self, {"status": "success" if result.get("ok") else "error", **result})
                return

            # Static files
            return super().do_GET()

        except Exception as e:
            traceback.print_exc()
            _json(self, {"status": "error", "message": str(e)}, 500)
        finally:
            _request_log(self, "request_finished", method="GET", path=path, duration_ms=round((time.perf_counter() - started) * 1000, 1))

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        self.request_id = _make_request_id()
        started = time.perf_counter()
        _request_log(self, "request_started", method="POST", path=path)

        try:
            if path == "/api/local-ai/rsrp-comment":
                cfg = load_local_ai_config(UPLOAD_DIR)
                body = _parse_json_body(self)
                rsrp_raw = body.get("rsrp") or []
                rsrp = [round(float(v), 1) for v in rsrp_raw if isinstance(v, (int, float)) and -150 <= float(v) <= 0]
                if len(rsrp) < 2:
                    _json(self, {"status": "error", "requestId": self.request_id, "message": "Need at least 2 valid RSRP values."}, 400)
                    return
                pci = body.get("pci")
                user_msg = f"RSRP sequence ({len(rsrp)} samples, dBm): {rsrp}\n"
                if pci is not None:
                    user_msg += f"Serving PCI: {pci}\n"
                user_msg += "In 1-2 sentences: describe the trend and any coverage concern."
                payload = {
                    "model": cfg.model,
                    "temperature": 0.1,
                    "messages": [
                        {"role": "system", "content": "You are a concise LTE RF signal analysis assistant. Given RSRP measurements, provide exactly 1-2 sentences describing the trend and any coverage concern. Be direct and technical. No padding."},
                        {"role": "user", "content": user_msg},
                    ],
                }
                if cfg.is_ollama:
                    payload["keep_alive"] = -1
                    payload["max_tokens"] = 80
                    payload["options"] = {"num_ctx": 512}
                    if "qwen3" in cfg.model.lower():
                        payload["think"] = False
                client = LMStudioClient(cfg)
                started = time.perf_counter()
                status_code, _h, resp_text = client._request("POST", cfg.chat_completions_url, payload=payload)
                if status_code >= 400:
                    raise LMStudioClientError(f"AI server returned HTTP {status_code}: {resp_text[:200]}")
                import json as _json_mod
                envelope = _json_mod.loads(resp_text)
                choices = envelope.get("choices") if isinstance(envelope, dict) else None
                message = (choices[0].get("message") if isinstance(choices, list) and choices else None) or {}
                comment = str(message.get("content") or message.get("reasoning_content") or "").strip()
                if "<think>" in comment:
                    end = comment.find("</think>")
                    comment = comment[end + len("</think>"):].strip() if end >= 0 else comment
                _json(self, {"status": "success", "requestId": self.request_id, "comment": comment or "No response.", "latencyMs": round((time.perf_counter() - started) * 1000, 1), "model": cfg.model})
                return

            if path == "/api/local-ai/pilot-pollution-insights":
                cfg = load_local_ai_config(UPLOAD_DIR)
                body = _parse_json_body(self)
                analysis = body.get("analysis")
                if not isinstance(analysis, dict) or "verdict" not in analysis:
                    _json(
                        self,
                        {
                            "status": "error",
                            "requestId": self.request_id,
                            "message": "Request body must be JSON with an 'analysis' key containing a valid pilot pollution analysis object.",
                        },
                        400,
                    )
                    return
                result = _analyze_pilot_pollution_ai(analysis, cfg)
                _json(self, {"status": "success", "requestId": self.request_id, **result})
                return

            if path == "/api/local-ai/analyze-log":
                cfg = load_local_ai_config(UPLOAD_DIR)
                try:
                    content_length = int(self.headers.get("Content-Length") or "0")
                except Exception:
                    content_length = 0
                if content_length > (cfg.max_upload_bytes + 1024 * 1024):
                    _json(
                        self,
                        {
                            "status": "error",
                            "requestId": self.request_id,
                            "message": (
                                f"Upload rejected before parsing because Content-Length exceeded the configured limit of "
                                f"{round(cfg.max_upload_bytes / (1024 * 1024), 2)} MB."
                            ),
                        },
                        413,
                    )
                    return
                ctype = self.headers.get("Content-Type", "")
                if "multipart/form-data" not in ctype.lower():
                    _json(
                        self,
                        {
                            "status": "error",
                            "requestId": self.request_id,
                            "message": "Expected multipart/form-data upload with a single file field named file.",
                        },
                        400,
                    )
                    return
                body = _read_body(self)
                filename, data, part_headers = _parse_multipart_file_info(body, ctype)
                _request_log(
                    self,
                    "local_ai_upload_received",
                    file_name=os.path.basename(filename or ""),
                    part_content_type=part_headers.get("content-type"),
                    size_bytes=len(data or b""),
                )
                result = analyze_uploaded_log(
                    filename=filename,
                    data=data,
                    config=cfg,
                    request_id=self.request_id,
                    logger=lambda message, **fields: _request_log(self, message, **fields),
                )
                status_code = 200 if result.get("status") in {"success", "partial_success"} else 502
                _json(self, result, status_code)
                return

            if path == "/api/nmfs/config":
                payload = _parse_json_body(self)
                timeout_raw = payload.get("timeoutSec")
                try:
                    timeout_val = int(timeout_raw)
                except Exception:
                    timeout_val = 180
                update = {
                    "converterCmd": str(payload.get("converterCmd") or "").strip(),
                    "timeoutSec": timeout_val,
                    "keepTemp": bool(payload.get("keepTemp")),
                }
                if update["timeoutSec"] < 10:
                    update["timeoutSec"] = 10
                _write_nmfs_config_file(update)
                eff = _get_nmfs_effective_config()
                _json(self, {"status": "success", "config": eff, "configPath": NMFS_CONFIG_PATH})
                return

            if path == "/api/nmfs/config/test":
                cfg = _get_nmfs_effective_config()
                report = _validate_nmfs_converter_cfg(cfg)
                _json(self, {"status": "success", "report": report, "configPath": NMFS_CONFIG_PATH})
                return

            if path == "/api/nmfs/decode":
                os.makedirs(UPLOAD_DIR, exist_ok=True)
                ctype = self.headers.get("Content-Type", "")
                body = _read_body(self)
                filename, data = _parse_multipart_file(body, ctype)
                safe_name = os.path.basename(filename or "upload.nmfs")
                save_path = os.path.join(UPLOAD_DIR, safe_name)
                with open(save_path, "wb") as f:
                    f.write(data)

                conv = _run_nmfs_converter(save_path)
                _json(
                    self,
                    {
                        "status": "success",
                        "filename": safe_name,
                        "converter": {
                            "returncode": conv.get("returncode"),
                            "output_path": conv.get("output_path"),
                            "text_len": conv.get("text_len"),
                            "stdout": conv.get("stdout", ""),
                            "stderr": conv.get("stderr", ""),
                        },
                        "text": conv.get("text", ""),
                    },
                )
                return

            if path == "/api/lte_rrc/decode":
                payload = _parse_json_body(self)
                event_name = str(payload.get("eventName") or payload.get("event_name") or "").strip()
                payload_hex = str(payload.get("payloadHex") or payload.get("payload_hex") or "").strip()
                if not event_name or not payload_hex:
                    _json(self, {"status": "error", "message": "eventName and payloadHex are required"}, 400)
                    return
                try:
                    payload_bytes = bytes.fromhex(payload_hex)
                except Exception:
                    _json(self, {"status": "error", "message": "Invalid payloadHex"}, 400)
                    return
                name_lc = event_name.lower()
                if "measurementreport" in name_lc:
                    decoded = decode_measurement_report_payload(payload_bytes)
                elif "rrcconnectionreconfiguration" in name_lc and "complete" not in name_lc:
                    decoded = decode_rrc_reconfiguration_payload(payload_bytes)
                else:
                    decoded = decode_rrc_event_payload(payload_bytes, event_name)
                _json(self, {"status": "success", "decoded": decoded})
                return

            if path == "/api/lte_rrc/decode_batch":
                payload = _parse_json_body(self)
                items = payload.get("items")
                if not isinstance(items, list) or not items:
                    _json(self, {"status": "error", "message": "items array is required"}, 400)
                    return
                decoded_items = []
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    event_name = str(item.get("eventName") or item.get("event_name") or "").strip()
                    payload_hex = str(item.get("payloadHex") or item.get("payload_hex") or "").strip()
                    if not event_name or not payload_hex:
                        continue
                    try:
                        payload_bytes = bytes.fromhex(payload_hex)
                    except Exception:
                        decoded_items.append({
                            "eventName": event_name,
                            "payloadHex": payload_hex,
                            "decoded": None,
                            "error": "Invalid payloadHex",
                        })
                        continue
                    name_lc = event_name.lower()
                    if "measurementreport" in name_lc:
                        decoded = decode_measurement_report_payload(payload_bytes)
                    elif "rrcconnectionreconfiguration" in name_lc and "complete" not in name_lc:
                        decoded = decode_rrc_reconfiguration_payload(payload_bytes)
                    else:
                        decoded = decode_rrc_event_payload(payload_bytes, event_name)
                    decoded_items.append({
                        "eventName": event_name,
                        "payloadHex": payload_hex,
                        "decoded": decoded,
                    })
                _json(self, {"status": "success", "items": decoded_items})
                return

            if path == "/api/lte_rrc/precompute":
                payload = _parse_json_body(self)
                items = payload.get("items")
                provided_cache_key = str(payload.get("cacheKey") or payload.get("cache_key") or "").strip()
                if provided_cache_key:
                    cached = _load_lte_rrc_precompute(provided_cache_key)
                    if cached is not None:
                        _json(self, {"status": "success", "cacheKey": provided_cache_key, "cached": True, **cached})
                        return
                if not isinstance(items, list) or not items:
                    _json(self, {"status": "error", "message": "items array is required when cache is missing"}, 400)
                    return
                if not provided_cache_key:
                    fingerprint_rows = []
                    for item in items:
                        if not isinstance(item, dict):
                            continue
                        fingerprint_rows.append({
                            "eventName": str(item.get("eventName") or item.get("event_name") or "").strip(),
                            "payloadHex": str(item.get("payloadHex") or item.get("payload_hex") or "").strip(),
                            "time": str(item.get("time") or "").strip(),
                            "servingPci": item.get("servingPci"),
                            "servingEarfcn": item.get("servingEarfcn"),
                        })
                    provided_cache_key = hashlib.sha1(
                        json.dumps(fingerprint_rows, separators=(",", ":"), sort_keys=True).encode("utf-8")
                    ).hexdigest()
                    cached = _load_lte_rrc_precompute(provided_cache_key)
                    if cached is not None:
                        _json(self, {"status": "success", "cacheKey": provided_cache_key, "cached": True, **cached})
                        return
                diagnostics = {
                    "candidateEvents": len(items),
                    "measurementReports": 0,
                    "reconfigurations": 0,
                    "decodedMeasurementReports": 0,
                    "decodedReconfigurations": 0,
                    "reconfigWithA3Resolvers": 0,
                    "exactA3Reports": 0,
                    "errors": [],
                }
                decoded_rows = []
                for index, item in enumerate(items):
                    if not isinstance(item, dict):
                        continue
                    row_id = item.get("rowId", index)
                    event_name = str(item.get("eventName") or item.get("event_name") or "").strip()
                    payload_hex = str(item.get("payloadHex") or item.get("payload_hex") or "").strip()
                    if not event_name or not payload_hex:
                        continue
                    properties = {}
                    try:
                        payload_bytes = bytes.fromhex(payload_hex)
                    except Exception:
                        diagnostics["errors"].append(f"{event_name}: invalid payload hex")
                        continue
                    name_lc = event_name.lower()
                    decoded = None
                    try:
                        if "measurementreport" in name_lc:
                            diagnostics["measurementReports"] += 1
                            decoded = decode_measurement_report_payload(payload_bytes)
                        elif "rrcconnectionreconfiguration" in name_lc and "complete" not in name_lc:
                            diagnostics["reconfigurations"] += 1
                            decoded = decode_rrc_reconfiguration_payload(payload_bytes)
                        else:
                            decoded = decode_rrc_event_payload(payload_bytes, event_name)
                    except Exception as exc:
                        diagnostics["errors"].append(f"{event_name}: {exc}")
                        continue
                    if not isinstance(decoded, dict) or not decoded.get("ok"):
                        continue
                    properties["RRC decoder"] = str(decoded.get("decoder") or "pycrate_rrclte")
                    if decoded.get("message_id"):
                        properties["rrc_message_id"] = str(decoded.get("message_id"))
                    row_meta = {
                        "rowId": row_id,
                        "ts": item.get("time"),
                        "eventName": event_name,
                        "properties": properties,
                        "servingPci": item.get("servingPci"),
                        "servingEarfcn": item.get("servingEarfcn"),
                        "decoded": decoded,
                        "resolver": None,
                    }
                    if str(decoded.get("message_id") or event_name).lower() == "measurement_report":
                        diagnostics["decodedMeasurementReports"] += 1
                        properties["measurement_report_full_decoded"] = "Yes"
                        properties["measurement_report_full_type"] = str(decoded.get("decoder_type") or "")
                        summary = decoded.get("summary") if isinstance(decoded.get("summary"), dict) else {}
                        if summary.get("measId") is not None:
                            properties["measurement_report_measid"] = str(summary.get("measId"))
                        if isinstance(decoded.get("serving"), dict):
                            properties["measurement_report_serving_json"] = json.dumps(decoded.get("serving"))
                        if isinstance(decoded.get("neighbors_lte"), list):
                            properties["measurement_report_neighbors_json"] = json.dumps(decoded.get("neighbors_lte"))
                        if isinstance(decoded.get("servfreq"), list):
                            properties["measurement_report_servfreq_json"] = json.dumps(decoded.get("servfreq"))
                        bits = []
                        if summary.get("measId") is not None:
                            bits.append(f"measId {summary.get('measId')}")
                        serving = decoded.get("serving") if isinstance(decoded.get("serving"), dict) else {}
                        serving_rsrp = _to_num_if_finite(serving.get("rsrp_dbm"))
                        if serving_rsrp is not None:
                            bits.append(f"serving RSRP {serving_rsrp} dBm")
                        if isinstance(decoded.get("neighbors_lte"), list):
                            bits.append(f"neighbors {len(decoded.get('neighbors_lte'))}")
                        if bits:
                            properties["rrc_message_summary"] = " | ".join(bits)
                    elif str(decoded.get("message_id") or "").lower() == "rrc_reconfiguration" or name_lc == "rrcconnectionreconfiguration":
                        diagnostics["decodedReconfigurations"] += 1
                        resolver = decoded.get("meas_resolver") if isinstance(decoded.get("meas_resolver"), dict) else None
                        row_meta["resolver"] = resolver
                        properties["rrc_recfg_full_decoded"] = "Yes"
                        properties["rrc_recfg_full_decoder"] = str(decoded.get("decoder") or "pycrate_rrclte")
                        properties["rrc_recfg_full_type"] = str(decoded.get("decoder_type") or "")
                        properties["rrc_recfg_full_json"] = json.dumps(decoded.get("decoded_json") or {})
                        summary = decoded.get("summary") if isinstance(decoded.get("summary"), dict) else {}
                        properties["rrc_recfg_meas_config_present"] = "Yes" if summary.get("has_measConfig") else "No"
                        if isinstance(decoded.get("meas_config"), dict):
                            properties["rrc_recfg_meas_config_json"] = json.dumps(decoded.get("meas_config"))
                        if resolver:
                            properties["rrc_recfg_meas_resolver_json"] = json.dumps(resolver)
                        a3_resolvers = resolver.get("a3Resolvers") if isinstance(resolver, dict) and isinstance(resolver.get("a3Resolvers"), list) else []
                        if a3_resolvers:
                            diagnostics["reconfigWithA3Resolvers"] += 1
                            first_a3 = a3_resolvers[0] if isinstance(a3_resolvers[0], dict) else None
                            if first_a3:
                                report_cfg = first_a3.get("reportConfig") if isinstance(first_a3.get("reportConfig"), dict) else {}
                                meas_object = first_a3.get("measObject") if isinstance(first_a3.get("measObject"), dict) else {}
                                if report_cfg.get("a3OffsetDb") is not None:
                                    properties["rrc_recfg_a3_offset_db"] = str(report_cfg.get("a3OffsetDb"))
                                if report_cfg.get("hysteresisDb") is not None:
                                    properties["rrc_recfg_hysteresis_db"] = str(report_cfg.get("hysteresisDb"))
                                if report_cfg.get("timeToTriggerMs") is not None:
                                    properties["rrc_recfg_ttt_ms"] = str(report_cfg.get("timeToTriggerMs"))
                                if first_a3.get("measId") is not None:
                                    properties["rrc_recfg_meas_id"] = str(first_a3.get("measId"))
                                if first_a3.get("measObjectId") is not None:
                                    properties["rrc_recfg_meas_object_id"] = str(first_a3.get("measObjectId"))
                                if first_a3.get("reportConfigId") is not None:
                                    properties["rrc_recfg_report_config_id"] = str(first_a3.get("reportConfigId"))
                                if report_cfg.get("eventType"):
                                    properties["rrc_recfg_event_type"] = str(report_cfg.get("eventType"))
                                if meas_object.get("offsetFreqDb") is not None:
                                    properties["rrc_recfg_offset_freq_db"] = str(meas_object.get("offsetFreqDb"))
                                properties["rrc_recfg_a3_resolver_summary"] = _render_a3_resolver_summary(resolver)
                        bits = []
                        if summary.get("has_measConfig"):
                            bits.append("measConfig")
                        if summary.get("has_mobilityControlInfo"):
                            bits.append("mobilityControlInfo")
                        if a3_resolvers:
                            bits.append(f"A3 resolvers {len(a3_resolvers)}")
                        if bits:
                            properties["rrc_message_summary"] = " | ".join(bits)
                    decoded_rows.append(row_meta)
                recfg_rows = [
                    row for row in decoded_rows
                    if row.get("eventName", "").lower() == "rrcconnectionreconfiguration" and isinstance(row.get("resolver"), dict)
                ]
                for row in decoded_rows:
                    if row.get("eventName", "").lower() != "measurementreport":
                        continue
                    resolved = _resolve_exact_a3_for_measurement_report(row, recfg_rows)
                    if not resolved:
                        continue
                    row["properties"]["measurement_report_a3_mapping_summary"] = resolved["mappingSummary"]
                    row["properties"]["measurement_report_a3_source_time"] = resolved["sourceTimeLabel"]
                    row["properties"]["measurement_report_a3_source_pci"] = str(resolved["sourcePci"]) if resolved.get("sourcePci") is not None else ""
                    row["properties"]["measurement_report_a3_eval_summary"] = resolved["evaluationSummary"]
                    row["properties"]["measurement_report_a3_eval_json"] = json.dumps(resolved)
                    diagnostics["exactA3Reports"] += 1
                result_payload = {
                    "diagnostics": diagnostics,
                    "items": [
                        {"rowId": row["rowId"], "properties": row["properties"]}
                        for row in decoded_rows
                        if row.get("properties")
                    ],
                }
                _store_lte_rrc_precompute(provided_cache_key, result_payload)
                _json(self, {"status": "success", "cacheKey": provided_cache_key, "cached": False, **result_payload})
                return

            if path == "/api/ho-analysis/run" or path == "/api/interfreq-ho-analysis/run":
                payload = _parse_json_body(self)
                dataset = payload.get("dataset")
                if dataset is None:
                    _json(self, {"status": "error", "message": "dataset is required"}, 400)
                    return
                mode = "interfreq" if path == "/api/interfreq-ho-analysis/run" else (payload.get("mode") or "intrafreq")
                result = _run_ho_analysis({
                    "dataset": dataset,
                    "options": payload.get("options") or {},
                    "mode": mode,
                })
                analysis_id = _store_ho_analysis(result, {
                    "label": payload.get("label"),
                    "source": payload.get("source"),
                    "mode": mode,
                })
                _json(self, {
                    "status": "success",
                    "analysisId": analysis_id,
                    "summary": result.get("kpis", {}).get("summary"),
                    "normalization": result.get("normalization"),
                    "eventCount": len(result.get("events") or []),
                })
                return

            if path == "/api/trp/import":
                os.makedirs(UPLOAD_DIR, exist_ok=True)

                ctype = self.headers.get("Content-Type", "")
                body = _read_body(self)

                filename, data = _parse_multipart_file(body, ctype)

                save_path = os.path.join(UPLOAD_DIR, os.path.basename(filename))
                with open(save_path, "wb") as f:
                    f.write(data)

                result = import_trp_file(save_path, DB_PATH, UPLOAD_DIR)
                if not result.get("ok", True):
                    report = result.get("importReport") or {}
                    _json(self, {
                        "status": "error",
                        "message": result.get("message") or "TRP import failed",
                        "importReport": report,
                        "metricsCount": int(result.get("kpi_count") or 0),
                        "eventTypesCount": int(result.get("event_count") or 0),
                        "trackCount": int(result.get("track_count") or 0),
                    }, 422)
                    return
                # Backward-compatible keys for legacy tests/clients.
                compat = {
                    "runId": result.get("runId"),
                    "metricsCount": int(result.get("kpi_count") or 0),
                    "eventTypesCount": int(result.get("event_count") or 0),
                    "importReport": (result.get("importReport") or {
                        "decodedSamples": int(result.get("kpi_count") or 0),
                        "decodedEvents": int(result.get("event_count") or 0),
                        "trackPoints": int(result.get("track_count") or 0),
                    }),
                }
                _json(self, {"status": "success", **result, **compat})
                return

            if path == "/api/nemo/lte/import":
                os.makedirs(UPLOAD_DIR, exist_ok=True)
                ctype = self.headers.get("Content-Type", "")
                body = _read_body(self)
                try:
                    parts = _parse_multipart_all_files(body, ctype)
                except Exception as exc:
                    _json(self, {"status": "error", "message": f"Multipart parse error: {exc}"}, 400)
                    return
                if not parts:
                    _json(self, {"status": "error", "message": "No files uploaded"}, 400)
                    return
                saved_paths = []
                for fname, data in parts:
                    safe_name = os.path.basename(fname) or "nemo_upload.txt"
                    save_path = os.path.join(UPLOAD_DIR, safe_name)
                    with open(save_path, "wb") as fh:
                        fh.write(data)
                    saved_paths.append(save_path)
                try:
                    parsed = parse_nemo_lte_files(saved_paths, skip_rrc=True)
                    if not parsed.get("kpi_samples"):
                        _json(self, {"status": "error", "message": "No LTE samples found in uploaded files"}, 422)
                        return
                    run_id = register_nemo_lte_run(parsed)
                    _json(self, {
                        "status": "success",
                        "runId": run_id,
                        "name": parsed.get("name", "Nemo LTE Export"),
                        "pointCount": parsed.get("point_count", 0),
                        "sampleCount": len(parsed.get("kpi_samples") or []),
                    })
                    # Parse RRC Layer-3 in background — populates L3 events after map loads
                    rrc_files = parsed.get("rrc_files") or []
                    if rrc_files:
                        from trp_importer import _RUNS
                        def _load_rrc(rid: int, paths: list) -> None:
                            try:
                                events = parse_rrc_files_deferred(paths)
                                if rid in _RUNS:
                                    _RUNS[rid]["events"] = events
                            except Exception as _e:
                                pass
                        threading.Thread(
                            target=_load_rrc, args=(run_id, rrc_files), daemon=True
                        ).start()
                except Exception as exc:
                    _json(self, {"status": "error", "message": str(exc)}, 500)
                return

            # BDD sectors-cache: store pre-parsed sectors from client-side import
            if path == "/api/bdd/sectors-cache":
                body = _parse_json_body(self)
                sectors = body.get("sectors")
                if isinstance(sectors, list) and len(sectors) > 0:
                    try:
                        with open(BDD_SECTORS_JSON_PATH, "w", encoding="utf-8") as f:
                            json.dump(sectors, f)
                    except Exception:
                        pass
                _json(self, {"status": "success"})
                return

            if path == "/api/benchmark/load":
                body = _parse_json_body(self)
                benchmark_path = str(body.get("path") or "").strip()
                result = _load_benchmark_workbook(benchmark_path)
                status_code = 200 if result.get("ok") else 400
                _json(self, {"status": "success" if result.get("ok") else "error", **result}, status_code)
                return

            if path == "/api/benchmark/upload":
                ct = self.headers.get("Content-Type", "")
                body_bytes = _read_body(self)
                try:
                    filename, data = _parse_multipart_file(body_bytes, ct)
                except Exception as exc:
                    _json(self, {"status": "error", "message": f"Multipart parse error: {exc}"}, 400)
                    return

                ext = os.path.splitext(filename or "")[1].lower()
                if ext not in (".xlsx", ".xls"):
                    _json(self, {"status": "error", "message": "Only .xlsx or .xls benchmark files are supported"}, 400)
                    return

                os.makedirs(UPLOAD_DIR, exist_ok=True)
                save_path = os.path.join(UPLOAD_DIR, "benchmark_current" + (ext or ".xlsx"))
                with open(save_path, "wb") as fh:
                    fh.write(data)

                result = _load_benchmark_workbook(save_path)
                status_code = 200 if result.get("ok") else 400
                _json(
                    self,
                    {
                        "status": "success" if result.get("ok") else "error",
                        "uploadedName": filename,
                        **result,
                    },
                    status_code,
                )
                return

            if path == "/api/benchmark-nemo/load":
                body = _parse_json_body(self) or {}
                result = _load_benchmark_nemo_files(
                    body.get("paths"),
                    dl_mode=body.get("dlMode"),
                    window_mode=body.get("windowMode"),
                )
                status_code = 200 if result.get("ok") else 400
                _json(self, {"status": "success" if result.get("ok") else "error", **result}, status_code)
                if result.get("ok"):
                    threading.Thread(
                        target=_benchmark_nemo_precompute_alt_window_background,
                        args=(body.get("dlMode"), body.get("windowMode")),
                        daemon=True,
                        name="benchmark-nemo-precompute-window",
                    ).start()
                return

            if path == "/api/benchmark-nemo/library/load":
                body = _parse_json_body(self) or {}
                dataset_id = body.get("datasetId")
                dataset_key = str(body.get("datasetKey") or "").strip()
                record = None
                if dataset_id is not None:
                    try:
                        record = _benchmark_nemo_library_load_dataset_by_id(int(dataset_id))
                    except (TypeError, ValueError):
                        _json(self, {"status": "error", "message": "Invalid dataset id"}, 400)
                        return
                elif dataset_key:
                    record = _benchmark_nemo_library_load_dataset_by_key(dataset_key)
                else:
                    _json(self, {"status": "error", "message": "Missing datasetId or datasetKey"}, 400)
                    return
                if not record:
                    _json(self, {"status": "error", "message": "Dataset not found"}, 404)
                    return
                dataset = _benchmark_nemo_library_load_into_memory(record)
                _json(
                    self,
                    {
                        "status": "success",
                        "ok": True,
                        "datasetId": record.get("id"),
                        "datasetKey": record.get("datasetKey") or "",
                        "dataset": dataset,
                        "paths": list(BENCHMARK_NEMO_DATASET.get("paths") or []),
                        "cached": True,
                        "persistent": True,
                    },
                    200,
                )
                threading.Thread(
                    target=_benchmark_nemo_precompute_alt_window_background,
                    args=((dataset or {}).get("dlMode"), (dataset or {}).get("windowMode")),
                    daemon=True,
                    name="benchmark-nemo-precompute-window",
                ).start()
                return

            if path == "/api/benchmark-nemo/dt-analysis":
                body = _parse_json_body(self) or {}
                try:
                    index = int(body.get("index"))
                except (TypeError, ValueError):
                    _json(self, {"status": "error", "message": "Missing or invalid DT index"}, 400)
                    return
                dataset = _benchmark_nemo_dt_dataset(
                    index,
                    dl_mode=body.get("dlMode"),
                    window_mode=body.get("windowMode"),
                )
                if not dataset:
                    _json(self, {"status": "error", "message": f"No data for DT index {index}. Load the benchmark files first."}, 400)
                    return
                _json(self, {"status": "success", "ok": True, "index": index, "dataset": dataset}, 200)
                return

            if path == "/api/benchmark-deep/generate-loaded":
                # Deep Benchmark analysis xlsx for the active scope (DT index, or -1 = cumulative).
                body = _parse_json_body(self) or {}
                try:
                    index = int(body.get("index", -1))
                except (TypeError, ValueError):
                    index = -1
                dl_mode = _benchmark_nemo_normalize_dl_mode(body.get("dlMode"))
                window_mode = _benchmark_nemo_normalize_window_mode(body.get("windowMode"))
                if index is not None and index >= 0:
                    dataset = _benchmark_nemo_dt_dataset(index, dl_mode=dl_mode, window_mode=window_mode)
                else:
                    dataset = BENCHMARK_NEMO_DATASET.get("data") or {}
                    if (
                        _benchmark_nemo_normalize_dl_mode(dataset.get("dlMode")) != dl_mode
                        or _benchmark_nemo_normalize_window_mode(dataset.get("windowMode")) != window_mode
                    ):
                        loaded = _load_benchmark_nemo_files(dl_mode=dl_mode, window_mode=window_mode)
                        dataset = loaded.get("dataset") or {}
                dataset = _ensure_deep_benchmark(dataset or {})
                deep = (dataset or {}).get("deepBenchmark")
                if not deep:
                    _json(self, {"status": "error", "message": "No Deep Benchmark analysis available. Load the benchmark files first (IAM required)."}, 400)
                    return
                try:
                    xlsx_bytes = generate_benchmark_deep_xlsx(deep, dataset)
                except Exception as exc:
                    import traceback as _tb
                    _tb.print_exc()
                    _json(self, {"status": "error", "message": f"XLSX generation failed: {exc}"}, 500)
                    return
                fname = "IAM_Deep_Benchmark.xlsx"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
                self.send_header("Content-Length", str(len(xlsx_bytes)))
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                self.wfile.write(xlsx_bytes)
                return

            if path == "/api/benchmark-mycom/load":
                body = _parse_json_body(self)
                result = _load_benchmark_mycom_file(str((body or {}).get("path") or "").strip())
                status_code = 200 if result.get("ok") else 400
                _json(self, {"status": "success" if result.get("ok") else "error", **result}, status_code)
                return

            if path == "/api/benchmark-optim/generate-loaded":
                # Reuse the benchmark already parsed in-memory (or restored from the SQLite
                # library on a prior load) — skips the multi-minute TXT re-parse.
                body = _parse_json_body(self) or {}
                dl_mode = _benchmark_nemo_normalize_dl_mode(body.get("dlMode"))
                window_mode = _benchmark_nemo_normalize_window_mode(body.get("windowMode"))
                operator_files = BENCHMARK_NEMO_DATASET.get("operator_files") or []
                loaded_dataset = BENCHMARK_NEMO_DATASET.get("data") or {}
                if (
                    not operator_files
                    or _benchmark_nemo_normalize_dl_mode(loaded_dataset.get("dlMode")) != dl_mode
                    or _benchmark_nemo_normalize_window_mode(loaded_dataset.get("windowMode")) != window_mode
                ):
                    try:
                        _load_benchmark_nemo_files(
                            dl_mode=dl_mode,
                            window_mode=window_mode,
                        )  # restores from SQLite cache when available
                    except Exception:
                        pass
                    operator_files = BENCHMARK_NEMO_DATASET.get("operator_files") or []
                if not operator_files:
                    _json(self, {"status": "error", "message": "No benchmark loaded. Load it via 'Benchmark Nemo TXT' first, or use the file picker."}, 409)
                    return
                _loaded = BENCHMARK_NEMO_DATASET.get("data") or {}
                global_serving = {
                    "IAM": _loaded.get("iamServingCells"),
                    "ORANGE": _loaded.get("orangeServingCells"),
                    "INWI": _loaded.get("inwiServingCells"),
                }
                try:
                    xlsx_bytes = generate_benchmark_optim_xlsx_from_operator_files(operator_files, global_serving, dl_mode=dl_mode)
                except Exception as exc:
                    import traceback as _tb
                    _tb.print_exc()
                    _json(self, {"status": "error", "message": f"XLSX generation failed: {exc}"}, 500)
                    return
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", 'attachment; filename="Benchmark-Optim.xlsx"')
                self.send_header("Content-Length", str(len(xlsx_bytes)))
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                self.wfile.write(xlsx_bytes)
                return

            if path == "/api/benchmark-optim/generate":
                ct = self.headers.get("Content-Type", "")
                body_bytes = _read_body(self)
                try:
                    parts = _parse_multipart_all_files(body_bytes, ct)
                except Exception as exc:
                    _json(self, {"status": "error", "message": f"Multipart parse error: {exc}"}, 400)
                    return
                txt_parts = [(fn, data) for fn, data in (parts or []) if os.path.splitext(fn or "")[1].lower() == ".txt"]
                if not txt_parts:
                    _json(self, {"status": "error", "message": "No .txt files uploaded"}, 400)
                    return
                os.makedirs(os.path.join(UPLOAD_DIR, "benchmark_optim"), exist_ok=True)
                saved_paths = []
                for filename, data in txt_parts:
                    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", os.path.basename(filename or "upload.txt")) or "upload.txt"
                    save_path = os.path.join(UPLOAD_DIR, "benchmark_optim", safe_name)
                    with open(save_path, "wb") as fh:
                        fh.write(data)
                    saved_paths.append(save_path)
                try:
                    xlsx_bytes = generate_benchmark_optim_xlsx(saved_paths)
                except Exception as exc:
                    import traceback as _tb
                    _tb.print_exc()
                    _json(self, {"status": "error", "message": f"XLSX generation failed: {exc}"}, 500)
                    return
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", 'attachment; filename="Benchmark-Optim.xlsx"')
                self.send_header("Content-Length", str(len(xlsx_bytes)))
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                self.wfile.write(xlsx_bytes)
                return

            if path == "/api/benchmark-nemo/upload":
                ct = self.headers.get("Content-Type", "")
                body_bytes = _read_body(self)
                try:
                    parts = _parse_multipart_all_files(body_bytes, ct)
                except Exception as exc:
                    _json(self, {"status": "error", "message": f"Multipart parse error: {exc}"}, 400)
                    return

                if not parts:
                    _json(self, {"status": "error", "message": "No files uploaded"}, 400)
                    return

                os.makedirs(os.path.join(UPLOAD_DIR, "benchmark_nemo"), exist_ok=True)
                saved_paths = []
                uploaded_names = []
                for filename, data in parts:
                    ext = os.path.splitext(filename or "")[1].lower()
                    if ext not in (".txt", ".csv", ".tsv"):
                        continue
                    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", os.path.basename(filename or "upload.txt")) or "upload.txt"
                    save_path = os.path.join(UPLOAD_DIR, "benchmark_nemo", safe_name)
                    should_write = True
                    try:
                        if os.path.isfile(save_path) and os.path.getsize(save_path) == len(data):
                            with open(save_path, "rb") as fh:
                                should_write = fh.read() != data
                    except Exception:
                        should_write = True
                    if should_write:
                        with open(save_path, "wb") as fh:
                            fh.write(data)
                    saved_paths.append(save_path)
                    uploaded_names.append(filename)

                if not saved_paths:
                    _json(self, {"status": "error", "message": "Only .txt, .csv, or .tsv Nemo benchmark files are supported"}, 400)
                    return

                uploaded_hashes = {}
                for filename, data in parts:
                    ext = os.path.splitext(filename or "")[1].lower()
                    if ext not in (".txt", ".csv", ".tsv"):
                        continue
                    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", os.path.basename(filename or "upload.txt")) or "upload.txt"
                    save_path = os.path.join(UPLOAD_DIR, "benchmark_nemo", safe_name)
                    uploaded_hashes[save_path] = hashlib.sha256(data).hexdigest()
                result = _load_benchmark_nemo_files(saved_paths, uploaded_hashes=uploaded_hashes)
                status_code = 200 if result.get("ok") else 400

                # Also register the IAM file as a full Nemo LTE run so the sidebar can load it
                iam_path = next(
                    (p for p in saved_paths if _nemo_guess_operator(p) == "IAM"),
                    None
                )

                _json(
                    self,
                    {
                        "status": "success" if result.get("ok") else "error",
                        "uploadedNames": uploaded_names,
                        "iamRunId": None,
                        "iamRunName": None,
                        "iamRunPending": bool(iam_path),
                        **result,
                    },
                    status_code,
                )
                if result.get("ok") and iam_path:
                    threading.Thread(
                        target=_benchmark_nemo_register_iam_run_background,
                        args=(iam_path,),
                        daemon=True,
                        name="benchmark-nemo-iam-register",
                    ).start()
                return

            if path == "/api/benchmark-mycom/upload":
                ct = self.headers.get("Content-Type", "")
                body_bytes = _read_body(self)
                try:
                    filename, data = _parse_multipart_file(body_bytes, ct)
                except Exception as exc:
                    _json(self, {"status": "error", "message": f"Multipart parse error: {exc}"}, 400)
                    return
                safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", os.path.basename(filename or "mycom_export.csv")) or "mycom_export.csv"
                save_dir = os.path.join(UPLOAD_DIR, "benchmark_mycom")
                os.makedirs(save_dir, exist_ok=True)
                save_path = os.path.join(save_dir, safe_name)
                with open(save_path, "wb") as fh:
                    fh.write(data)
                result = _load_benchmark_mycom_file(save_path)
                status_code = 200 if result.get("ok") else 400
                _json(
                    self,
                    {
                        "status": "success" if result.get("ok") else "error",
                        "uploadedName": filename,
                        **result,
                    },
                    status_code,
                )
                return

            if path == "/api/benchmark-mycom/correlate":
                result = _correlate_benchmark_mycom_context()
                status_code = 200 if result.get("ok") else 400
                _json(self, {"status": "success" if result.get("ok") else "error", **result}, status_code)
                return

            if path == "/api/statistics-mycom/load":
                body = _parse_json_body(self)
                result = _load_benchmark_mycom_file(str((body or {}).get("path") or "").strip())
                status_code = 200 if result.get("ok") else 400
                _json(self, {"status": "success" if result.get("ok") else "error", **result}, status_code)
                return

            if path == "/api/statistics-mycom/upload":
                ct = self.headers.get("Content-Type", "")
                body_bytes = _read_body(self)
                try:
                    filename, data = _parse_multipart_file(body_bytes, ct)
                except Exception as exc:
                    _json(self, {"status": "error", "message": f"Multipart parse error: {exc}"}, 400)
                    return
                safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", os.path.basename(filename or "statistics_mycom.txt")) or "statistics_mycom.txt"
                save_dir = os.path.join(UPLOAD_DIR, "statistics_mycom")
                os.makedirs(save_dir, exist_ok=True)
                save_path = os.path.join(save_dir, safe_name)
                with open(save_path, "wb") as fh:
                    fh.write(data)
                result = _load_benchmark_mycom_file(save_path)
                status_code = 200 if result.get("ok") else 400
                _json(
                    self,
                    {
                        "status": "success" if result.get("ok") else "error",
                        "uploadedName": filename,
                        **result,
                    },
                    status_code,
                )
                return

            if path == "/api/statistics-mycom/analyze":
                body = _parse_json_body(self)
                result = _build_statistics_mycom_analysis(body.get("thresholds"), str(body.get("query") or "").strip())
                status_code = 200 if result.get("ok") else 400
                _json(self, {"status": "success" if result.get("ok") else "error", **result}, status_code)
                return

            # BDD configure
            if path == "/api/bdd/configure":
                body = _parse_json_body(self)
                bdd_path = str(body.get("path") or "").strip()
                if not bdd_path:
                    _json(self, {"status": "error", "message": "path is required"}, 400)
                    return
                if not os.path.isfile(bdd_path):
                    _json(self, {"status": "error", "message": f"File not found: {bdd_path}"}, 400)
                    return
                result = _bdd.load_bdd(bdd_path)
                if result.get("ok"):
                    os.makedirs(os.path.dirname(BDD_CONFIG_PATH), exist_ok=True)
                    with open(BDD_CONFIG_PATH, "w", encoding="utf-8") as f:
                        json.dump({"path": bdd_path}, f)
                    _write_bdd_sectors_json(bdd_path)
                _json(self, {"status": "success" if result.get("ok") else "error", **result})
                return

            # BDD upload (multipart file → save + parse sectors)
            if path == "/api/bdd/upload":
                ct = self.headers.get("Content-Type", "")
                body_bytes = _read_body(self)
                try:
                    filename, data = _parse_multipart_file(body_bytes, ct)
                except Exception as exc:
                    _json(self, {"status": "error", "message": f"Multipart parse error: {exc}"}, 400)
                    return
                os.makedirs(_BDD_STORE_DIR, exist_ok=True)
                save_path = os.path.join(_BDD_STORE_DIR, "bdd_current.xlsx")
                with open(save_path, "wb") as fh:
                    fh.write(data)
                with open(BDD_CONFIG_PATH, "w", encoding="utf-8") as f:
                    json.dump({"path": save_path}, f)
                _bdd.load_bdd(save_path)
                threading.Thread(target=_write_bdd_sectors_json, args=(save_path,), daemon=True).start()
                result = _bdd.get_map_sectors(save_path)
                _json(self, {"status": "success" if result.get("ok") else "error", **result})
                return

            # BDD match (Phase 2)
            if path == "/api/bdd/match":
                body = _parse_json_body(self)
                try:
                    event_lat = float(body.get("eventLat") or 0)
                    event_lon = float(body.get("eventLon") or 0)
                    earfcn = int(body.get("earfcn") or 0)
                except Exception:
                    _json(self, {"status": "error", "message": "eventLat, eventLon, earfcn are required"}, 400)
                    return
                polluters = body.get("polluters") or []
                max_dist_m = float(body.get("maxDistM") or 15000)
                rat = str(body.get("rat") or "LTE")
                route_points = body.get("routePoints") or None
                serving_sequence = body.get("servingSequence") or None
                best_pci_sequence = body.get("bestPciSequence") or None
                result = _bdd.match_pollution_event(
                    event_lat, event_lon, polluters, earfcn,
                    max_dist_m=max_dist_m, rat=rat,
                    route_points=route_points,
                    serving_sequence=serving_sequence,
                    best_pci_sequence=best_pci_sequence,
                )
                _json(self, {"status": "success" if result.get("ok") else "error", **result})
                return

            _json(self, {"status": "error", "message": "Not found"}, 404)

        except LocalAIUploadError as e:
            message = str(e)
            status = 413 if "too large" in message.lower() else 400
            _json(self, {"status": "error", "requestId": self.request_id, "message": message}, status)
        except (LMStudioClientError, LMStudioModelUnavailableError) as e:
            _json(self, {"status": "error", "requestId": self.request_id, "message": str(e)}, 502)
        except Exception as e:
            traceback.print_exc()
            _json(self, {"status": "error", "requestId": self.request_id, "message": str(e)}, 500)
        finally:
            _request_log(self, "request_finished", method="POST", path=path, duration_ms=round((time.perf_counter() - started) * 1000, 1))


def main():
    _bdd_load_saved_path()
    port = int(os.environ.get("PORT", "8000"))
    httpd = ThreadingHTTPServer(("0.0.0.0", port), Handler)
    status = _bdd.bdd_status()
    if status["loaded"]:
        print(f"BDD loaded: {status['count']} cells from {status['path']}")
    else:
        print("BDD not loaded. Use POST /api/bdd/configure or set OPTIM_BDD_PATH to enable Phase 2.")
    print(f"Starting server on port {port}...")
    print("Use Ctrl+C to stop.")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping server...")
    finally:
        try:
            httpd.server_close()
        except Exception:
            pass
    print("Server stopped.")


if __name__ == "__main__":
    main()


# Legacy alias used in tests.
CustomHandler = Handler
