"""
BDD (Base Station Database) matcher for pilot pollution Phase 2.

Loads a cell reference database from an Excel workbook, matches pollution event
PCIs to physical sectors, and computes geometry (distance, bearing, azimuth error,
lobe classification, overshoot detection, root cause, recommendation).

Public API
----------
load_bdd(path)                      → load/reload the Excel BDD
bdd_status()                        → current load state
match_pollution_event(...)          → main Phase 2 entry point
"""

from __future__ import annotations

import os
import time as _time
from math import radians, sin, cos, asin, sqrt, atan2, degrees, atan
from typing import Any, Dict, List, Optional, Tuple

# ── Module-level BDD cache ────────────────────────────────────────────────────
_BDD_CELLS: List[Dict[str, Any]] = []
_BDD_PATH: str = ""
_BDD_LOAD_ERROR: str = ""
_BDD_CELL_COUNT: int = 0

# ── Band → expected dominance radius (urban/semi-urban, tuned for MA networks) ─
_BAND_RADIUS_M: Dict[str, float] = {
    "L800":  4000.0,
    "L900":  3000.0,
    "L1800": 2000.0,
    "L2100": 1500.0,
    "L2600": 1000.0,
}

# ── EARFCN ranges → band label ─────────────────────────────────────────────────
def _earfcn_to_band(earfcn: Optional[int]) -> str:
    if earfcn is None:
        return "L1800"
    e = int(earfcn)
    if 6150 <= e <= 6449:
        return "L800"
    if 2750 <= e <= 3449:
        return "L900"
    if 1200 <= e <= 1949:
        return "L1800"
    if 300 <= e <= 699:
        return "L2100"
    if 2400 <= e <= 2700:
        return "L2600"
    if 3400 <= e <= 3799:
        return "L2600"
    return "L1800"


def _nrarfcn_to_band(nrarfcn: Optional[int]) -> str:
    """Map NR-ARFCN to NR band label (most common MA/Morocco bands)."""
    if nrarfcn is None:
        return ""
    n = int(nrarfcn)
    if 151600 <= n <= 160600:
        return "n28"
    if 361000 <= n <= 376000:
        return "n3"
    if 384000 <= n <= 396000:
        return "n1"
    if 422000 <= n <= 440000:
        return "n1"
    if 499200 <= n <= 537999:
        return "n41"
    if 620000 <= n <= 680000:
        return "n78"
    if 693334 <= n <= 733333:
        return "n79"
    return ""


# ── Geometry helpers ──────────────────────────────────────────────────────────
def _haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R = 6_371_000.0
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    a = sin((lat2 - lat1) / 2) ** 2 + cos(lat1) * cos(lat2) * sin((lon2 - lon1) / 2) ** 2
    return R * 2.0 * asin(sqrt(max(0.0, min(1.0, a))))


def _bearing_deg(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    lat1r, lat2r = radians(lat1), radians(lat2)
    dlon = radians(lon2 - lon1)
    x = sin(dlon) * cos(lat2r)
    y = cos(lat1r) * sin(lat2r) - sin(lat1r) * cos(lat2r) * cos(dlon)
    return (degrees(atan2(x, y)) + 360.0) % 360.0


def _angle_diff(a: float, b: float) -> float:
    return abs((a - b + 180.0) % 360.0 - 180.0)


def _vertical_angle_deg(h_m: float, dist_m: float, ue_h: float = 1.5) -> Optional[float]:
    if dist_m <= 0:
        return None
    return degrees(atan(max(0.0, h_m - ue_h) / dist_m))


def _lobe_class(az_error: float, beamwidth: float = 65.0) -> str:
    half = beamwidth / 2.0
    if az_error <= half:
        return "main_lobe"
    if az_error <= beamwidth:
        return "edge_lobe"
    if az_error <= beamwidth + 50.0:
        return "side_lobe"
    return "back_lobe"


# ── Root cause + recommendation ───────────────────────────────────────────────
def _classify_root_cause(
    lobe: str,
    dist_m: float,
    avg_delta: float,
    non_serving_count: int,
    total_tilt: Optional[float],
    expected_radius_m: float,
) -> str:
    aligned = lobe in ("main_lobe", "edge_lobe")
    beyond_radius = dist_m > expected_radius_m
    well_beyond = dist_m > expected_radius_m * 1.5
    strong_signal = avg_delta <= 3.0
    persistent = non_serving_count >= 2

    # Overshooting: aligned, beyond expected radius, strong, low/moderate tilt
    if aligned and beyond_radius and strong_signal and persistent:
        if total_tilt is not None and total_tilt <= 4.0:
            return "overshooting_candidate"
        if total_tilt is not None and total_tilt <= 6.0 and well_beyond:
            return "overshooting_candidate"

    # Main lobe: distinguish close overlap from far overshoot
    if lobe == "main_lobe" and avg_delta <= 3.0:
        if beyond_radius:
            return "main_lobe_overshooting"
        return "main_lobe_overlap"

    # Edge lobe at significant distance: overlap or overshoot
    if lobe == "edge_lobe" and dist_m > expected_radius_m * 0.8 and avg_delta <= 4.0:
        return "edge_lobe_overlap_or_overshoot"

    if lobe == "edge_lobe" and avg_delta <= 3.0:
        return "main_lobe_overlap"

    if lobe == "side_lobe" and avg_delta <= 5.0:
        return "side_lobe_pollution_candidate"
    if lobe == "back_lobe" and avg_delta <= 5.0:
        return "back_lobe_or_bdd_error_candidate"
    return lobe


_RECOMMENDATIONS: Dict[str, str] = {
    "overshooting_candidate": (
        "Check RET/electrical tilt, mechanical tilt, antenna height, and RS power. "
        "Consider adding 1–2° downtilt after verifying coverage impact on adjacent cells."
    ),
    "main_lobe_overshooting": (
        "The sector is aligned with the event area but the distance suggests it may be overshooting. "
        "Verify whether this road segment is inside the intended coverage boundary. "
        "If not, check RET/electrical tilt and mechanical tilt to reduce footprint."
    ),
    "main_lobe_overlap": (
        "Create a clearer dominant server by balancing tilt, azimuth, and power between "
        "overlapping sectors. Review planned serving boundary and CIO if needed."
    ),
    "edge_lobe_overlap_or_overshoot": (
        "Check whether this sector is creating edge-lobe overlap into the event area. "
        "Review azimuth, tilt, and planned coverage boundary. "
        "If signal is strong beyond the intended footprint, also check for overshooting."
    ),
    "side_lobe_pollution_candidate": (
        "Check physical antenna azimuth, antenna pattern, feeder/antenna installation, "
        "and possible reflection or diffraction from local clutter."
    ),
    "back_lobe_or_bdd_error_candidate": (
        "Verify BDD azimuth accuracy and physical antenna installation. "
        "Check for back-lobe/reflection. If BDD azimuth is confirmed correct, "
        "review feeder connections and possible antenna swap."
    ),
}


def _recommend(root_cause: str) -> str:
    return _RECOMMENDATIONS.get(
        root_cause,
        "Validate cell mapping and compare with planned coverage before applying RF or mobility changes.",
    )


# ── Type helpers ──────────────────────────────────────────────────────────────
def _safe_int(v: Any) -> Optional[int]:
    try:
        return None if v is None else int(v)
    except Exception:
        return None


def _safe_float(v: Any) -> Optional[float]:
    try:
        return None if v is None else float(v)
    except Exception:
        return None


# ── BDD loader ────────────────────────────────────────────────────────────────
# Column name aliases: BDD file header → internal key
_COL_MAP = {
    "RAT": "rat",
    "Operator": "operator",
    "Site_Name": "site_name",
    "Cell_Name": "cell_name",
    "Sector_ID": "sector_id",
    "CGI/ECI": "cgi_eci",
    "eNB_ID": "enb_id",
    "Cell_ID": "cell_id",
    "EARFCN_DL": "earfcn",
    "PCI": "pci",
    "BAND": "band",
    "Lat": "lat",
    "Lon": "lon",
    "AZ": "azimuth",
    "Mechanical_Tilt": "mech_tilt",
    "Electrical_Tilt": "elec_tilt",
    "Total_Tilt": "total_tilt",
    "Antenna_Height": "antenna_height",
    "Horizontal_Beamwidth": "h_beamwidth",
    "Antenna_Model": "antenna_model",
    "RS_Power/": "rs_power",
    "Admin_State": "admin_state",
}


def load_bdd(path: str) -> Dict[str, Any]:
    """Load/reload BDD from an Excel workbook. Returns status dict."""
    global _BDD_CELLS, _BDD_PATH, _BDD_LOAD_ERROR, _BDD_CELL_COUNT
    t0 = _time.perf_counter()
    try:
        import openpyxl  # deferred import — not always required
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["BDD"] if "BDD" in wb.sheetnames else wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        raw_header = next(rows_iter, None)
        if raw_header is None:
            raise ValueError("Empty sheet — no header row found")
        header = [str(h or "").strip() for h in raw_header]
        # Map column header → column index
        col_idx: Dict[str, int] = {name: i for i, name in enumerate(header)}

        def _g(row: tuple, col_name: str, default=None):
            i = col_idx.get(col_name)
            if i is None or i >= len(row):
                return default
            return row[i]

        cells: List[Dict[str, Any]] = []
        for row in rows_iter:
            raw_lat = _safe_float(_g(row, "Lat"))
            raw_lon = _safe_float(_g(row, "Lon"))
            if raw_lat is None or raw_lon is None:
                continue
            raw_earfcn = _safe_int(_g(row, "EARFCN_DL"))
            raw_pci = _safe_int(_g(row, "PCI"))
            if raw_earfcn is None or raw_pci is None:
                continue
            cells.append(
                {
                    "rat": str(_g(row, "RAT") or "").strip().upper(),
                    "operator": str(_g(row, "Operator") or "").strip(),
                    "site_name": str(_g(row, "Site_Name") or "").strip(),
                    "cell_name": str(_g(row, "Cell_Name") or "").strip(),
                    "sector_id": _safe_int(_g(row, "Sector_ID")),
                    "cgi_eci": _g(row, "CGI/ECI"),
                    "enb_id": _safe_int(_g(row, "eNB_ID")),
                    "cell_id": _safe_int(_g(row, "Cell_ID")),
                    "earfcn": raw_earfcn,
                    "pci": raw_pci,
                    "band": str(_g(row, "BAND") or "").strip(),
                    "lat": raw_lat,
                    "lon": raw_lon,
                    "azimuth": _safe_float(_g(row, "AZ")),
                    "mech_tilt": _safe_float(_g(row, "Mechanical_Tilt")),
                    "elec_tilt": _safe_float(_g(row, "Electrical_Tilt")),
                    "total_tilt": _safe_float(_g(row, "Total_Tilt")),
                    "antenna_height": _safe_float(_g(row, "Antenna_Height")),
                    "h_beamwidth": _safe_float(_g(row, "Horizontal_Beamwidth")),
                    "rs_power": _safe_float(_g(row, "RS_Power/")),
                    "admin_state": str(_g(row, "Admin_State") or "").strip(),
                }
            )

        elapsed_ms = round((_time.perf_counter() - t0) * 1000)
        _BDD_CELLS = cells
        _BDD_PATH = path
        _BDD_LOAD_ERROR = ""
        _BDD_CELL_COUNT = len(cells)
        return {"ok": True, "count": len(cells), "path": path, "loadTimeMs": elapsed_ms}

    except Exception as exc:
        _BDD_LOAD_ERROR = str(exc)
        return {"ok": False, "error": str(exc)}


def bdd_status() -> Dict[str, Any]:
    return {
        "loaded": bool(_BDD_CELLS),
        "count": _BDD_CELL_COUNT,
        "path": _BDD_PATH or None,
        "error": _BDD_LOAD_ERROR or None,
    }


def get_map_sectors(path: Optional[str] = None) -> Dict[str, Any]:
    """Parse multi-sheet BDD Excel (2G/3G/4G/5G) and return frontend-compatible sector list."""
    actual_path = path or _BDD_PATH
    if not actual_path or not os.path.isfile(actual_path):
        return {"ok": False, "error": "BDD file not configured or not found", "sectors": []}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(actual_path, read_only=True, data_only=True)
    except Exception as exc:
        return {"ok": False, "error": str(exc), "sectors": []}

    _TECH_COLORS = {"2G": "#ef4444", "3G": "#eab308", "4G": "#3b82f6", "5G": "#a855f7"}

    def _nc(s: str) -> str:
        return str(s).lower().replace(" ", "").replace("_", "").replace(".", "").replace("/", "")

    rat_sheets = []
    for n in wb.sheetnames:
        clean_n = str(n).strip().upper().replace("BDD ", "").replace("BDD_", "")
        if clean_n in ("2G", "3G", "4G", "5G"):
            rat_sheets.append((n, clean_n))

    all_sectors: List[Dict[str, Any]] = []

    for sheet_name, tech in rat_sheets:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        raw_header = next(rows_iter, None)
        if raw_header is None:
            continue
        col_idx: Dict[str, int] = {}
        for i, h in enumerate(raw_header):
            if h is not None:
                col_idx[_nc(str(h))] = i

        def _gv(row: tuple, names: List[str]):
            for nm in names:
                i = col_idx.get(_nc(nm))
                if i is not None and i < len(row):
                    v = row[i]
                    if v is not None and str(v).strip() != "":
                        return v
            return None

        for row in rows_iter:
            lat = _safe_float(_gv(row, ["latitude", "lat"]))
            lng = _safe_float(_gv(row, ["longitude", "long", "lon", "lng"]))
            if lat is None or lng is None:
                continue
            azimuth = _safe_float(_gv(row, ["azimut", "azimuth"])) or 0.0

            if tech == "2G":
                name      = str(_gv(row, ["site name", "sitename"]) or "").strip()
                cell_name = str(_gv(row, ["cellname", "cell name"]) or "").strip()
                lac       = _safe_int(_gv(row, ["lac"]))
                tac       = None
                cid       = _safe_int(_gv(row, ["cellid", "cell id"]))
                freq      = _safe_int(_gv(row, ["bcch", "arfcn"]))
                ncc       = _safe_int(_gv(row, ["ncc"]))
                bcc       = _safe_int(_gv(row, ["bcc"]))
                pci       = (ncc * 8 + bcc) if (ncc is not None and bcc is not None) else None
                rnc       = str(_gv(row, ["bsc", "bscname"]) or "").strip() or None
            elif tech == "3G":
                name      = str(_gv(row, ["site name", "sitename"]) or "").strip()
                cell_name = str(_gv(row, ["cellname", "cell name"]) or "").strip()
                lac       = _safe_int(_gv(row, ["lac"]))
                tac       = None
                cid       = _safe_int(_gv(row, ["cellid", "cell id"]))
                freq      = _safe_int(_gv(row, ["downlink uarfcn", "uarfcn"]))
                pci       = _safe_int(_gv(row, ["psc", "primaryscramblingcode"]))
                rnc_raw   = _gv(row, ["rncid", "rnc", "rnc id"])
                rnc       = str(rnc_raw).strip() if rnc_raw is not None else None
            elif tech == "4G":
                name      = str(_gv(row, ["basestationname", "site name", "sitename", "nodeb name"]) or "").strip()
                cell_name = str(_gv(row, ["cellname", "cell name"]) or "").strip()
                lac       = None
                tac       = _safe_int(_gv(row, ["tracking area code", "tac"]))
                cid       = _safe_int(_gv(row, ["cell id", "cellid"]))
                freq      = _safe_int(_gv(row, ["downlink earfcn", "earfcn"]))
                pci       = _safe_int(_gv(row, ["physical cell id", "pci"]))
                rnc_raw   = _gv(row, ["enb id", "enbid", "enodebid", "enb id", "enodeb id"])
                rnc       = str(rnc_raw).strip() if rnc_raw is not None else None
            elif tech == "5G":
                name      = str(_gv(row, ["site name", "sitename", "gnb name", "gnbname", "base station name"]) or "").strip()
                cell_name = str(_gv(row, ["cellname", "cell name", "nr du cell name"]) or "").strip()
                lac       = None
                tac       = _safe_int(_gv(row, ["tracking area code", "tac", "tracking area id"]))
                cid       = _safe_int(_gv(row, ["cell id", "cellid", "nci", "nr du cell id"]))
                freq      = _safe_int(_gv(row, ["nrarfcn", "nr arfcn", "ssb arfcn", "arfcn", "downlink narfcn"]))
                pci       = _safe_int(_gv(row, ["pci", "physical cell id"]))
                rnc_raw   = _gv(row, ["gnb id", "gnbid", "gnodeb id"])
                rnc       = str(rnc_raw).strip() if rnc_raw is not None else None
            else:
                continue

            if not name:
                continue

            band_label = _nrarfcn_to_band(freq) if tech == "5G" else (_earfcn_to_band(freq) if tech == "4G" else "")
            all_sectors.append({
                "lat":      lat,
                "lng":      lng,
                "azimuth":  azimuth,
                "name":     name,
                "siteName": name,
                "cellName": cell_name or None,
                "cellId":   str(cell_name) if cell_name else None,
                "lac":      lac,
                "tac":      tac,
                "pci":      pci,
                "sc":       pci,
                "freq":     freq,
                "tech":     tech,
                "band":     band_label or None,
                "color":    _TECH_COLORS.get(tech, "#3b82f6"),
                "rnc":      rnc,
                "cid":      cid,
            })

    return {"ok": True, "sectors": all_sectors, "count": len(all_sectors)}


# ── Candidate search ──────────────────────────────────────────────────────────
def find_candidates(
    earfcn: int,
    pci: int,
    event_lat: float,
    event_lon: float,
    max_dist_m: float = 15_000.0,
    rat: str = "LTE",
) -> List[Dict[str, Any]]:
    """Return all BDD cells matching EARFCN+PCI within max_dist_m, sorted by distance."""
    rat_upper = rat.strip().upper()
    results = []
    for cell in _BDD_CELLS:
        if cell["rat"] != rat_upper:
            continue
        if cell["earfcn"] != earfcn or cell["pci"] != pci:
            continue
        dist = _haversine_m(event_lat, event_lon, cell["lat"], cell["lon"])
        if dist > max_dist_m:
            continue
        results.append({**cell, "_dist_m": dist})
    results.sort(key=lambda c: c["_dist_m"])
    return results


# ── Per-candidate geometry ────────────────────────────────────────────────────
def geometry_for_candidate(
    candidate: Dict[str, Any],
    event_lat: float,
    event_lon: float,
    avg_delta: float,
    non_serving_count: int,
    earfcn: int,
) -> Dict[str, Any]:
    dist_m = candidate["_dist_m"]
    brg = _bearing_deg(candidate["lat"], candidate["lon"], event_lat, event_lon)
    az = candidate.get("azimuth")
    az_err = _angle_diff(az, brg) if az is not None else None
    bw = candidate.get("h_beamwidth") or 65.0
    lobe = _lobe_class(az_err, bw) if az_err is not None else "unknown"
    total_tilt = candidate.get("total_tilt")
    h = candidate.get("antenna_height") or 0.0
    v_angle = _vertical_angle_deg(h, dist_m)
    band = candidate.get("band") or _earfcn_to_band(earfcn)
    expected_r = _BAND_RADIUS_M.get(band, 2000.0)
    root_cause = _classify_root_cause(lobe, dist_m, avg_delta, non_serving_count, total_tilt, expected_r)
    return {
        "distM": round(dist_m),
        "bearingDeg": round(brg, 1),
        "azimuthErrorDeg": round(az_err, 1) if az_err is not None else None,
        "lobeClass": lobe,
        "verticalAngleDeg": round(v_angle, 2) if v_angle is not None else None,
        "expectedRadiusM": expected_r,
        "rootCause": root_cause,
        "recommendation": _recommend(root_cause),
    }


# ── Mapping confidence ────────────────────────────────────────────────────────
def _mapping_confidence(n_candidates: int, best_lobe: str) -> float:
    if n_candidates == 0:
        return 0.0
    if n_candidates == 1:
        return 0.88 if best_lobe in ("main_lobe", "edge_lobe") else 0.72
    penalty = min(0.35, (n_candidates - 1) * 0.07)
    base = 0.82 if best_lobe in ("main_lobe", "edge_lobe") else 0.65
    return round(max(0.30, base - penalty), 2)


# ── Planned-server scoring helpers ────────────────────────────────────────────

def _azimuth_score(az_error: float) -> float:
    if az_error <= 20: return 100.0
    if az_error <= 35: return 85.0
    if az_error <= 60: return 60.0
    if az_error <= 90: return 30.0
    if az_error <= 120: return 15.0
    return 5.0


def _distance_score_planned(distance_m: float, expected_radius_m: float) -> float:
    ratio = distance_m / max(expected_radius_m, 1.0)
    if ratio <= 0.5: return 100.0
    if ratio <= 1.0: return 85.0
    if ratio <= 1.5: return 60.0
    if ratio <= 2.0: return 35.0
    return 10.0


def _lobe_score_planned(lobe: str) -> float:
    return {"main_lobe": 100.0, "edge_lobe": 70.0, "side_lobe": 25.0, "back_lobe": 5.0}.get(lobe, 30.0)


def _vertical_score_planned(cell: Dict[str, Any], distance_m: float) -> float:
    h = cell.get("antenna_height") or 0.0
    total_tilt = cell.get("total_tilt")
    if total_tilt is None or distance_m <= 0:
        return 50.0
    v_angle = degrees(atan(max(0.0, h - 1.5) / max(distance_m, 1.0)))
    v_err = abs(total_tilt - v_angle)
    if v_err <= 3: return 100.0
    if v_err <= 6: return 70.0
    if v_err <= 10: return 40.0
    return 15.0


def _score_cell_for_point(
    cell: Dict[str, Any], point_lat: float, point_lon: float, earfcn: int
) -> Dict[str, Any]:
    dist_m = _haversine_m(cell["lat"], cell["lon"], point_lat, point_lon)
    brg = _bearing_deg(cell["lat"], cell["lon"], point_lat, point_lon)
    az = cell.get("azimuth")
    az_err = _angle_diff(az, brg) if az is not None else 90.0
    bw = cell.get("h_beamwidth") or 65.0
    lobe = _lobe_class(az_err, bw)
    band = cell.get("band") or _earfcn_to_band(earfcn)
    expected_r = _BAND_RADIUS_M.get(band, 2000.0)
    a_score = _azimuth_score(az_err)
    d_score = _distance_score_planned(dist_m, expected_r)
    l_score = _lobe_score_planned(lobe)
    v_score = _vertical_score_planned(cell, dist_m)
    state_score = 100.0 if str(cell.get("admin_state", "")).upper() == "ON_AIR" else 0.0
    planned_score = (
        0.40 * a_score + 0.30 * d_score + 0.15 * l_score +
        0.10 * v_score + 0.05 * state_score
    )
    return {
        "pci": cell["pci"],
        "cellName": cell.get("cell_name"),
        "siteName": cell.get("site_name"),
        "plannedScore": round(planned_score, 1),
        "distM": round(dist_m),
        "azimuthErrorDeg": round(az_err, 1),
        "lobeClass": lobe,
        "expectedRadiusM": expected_r,
        "_cell_ref": cell,
    }


def _find_active_cells_on_earfcn(
    earfcn: int,
    event_lat: float,
    event_lon: float,
    max_dist_m: float = 15_000.0,
    rat: str = "LTE",
) -> List[Dict[str, Any]]:
    """Return all ON_AIR BDD cells matching EARFCN within max_dist_m (any PCI)."""
    rat_upper = rat.strip().upper()
    results = []
    for cell in _BDD_CELLS:
        if cell["rat"] != rat_upper:
            continue
        if cell["earfcn"] != earfcn:
            continue
        dist = _haversine_m(event_lat, event_lon, cell["lat"], cell["lon"])
        if dist > max_dist_m:
            continue
        results.append({**cell, "_dist_m": dist})
    results.sort(key=lambda c: c["_dist_m"])
    return results


_PLANNED_MIN_SCORE = 60.0
_PLANNED_OVERLAP_GAP = 10.0
_PLANNED_OVERLAP_MIN_SECOND = 55.0


def _classify_planned_conflict(
    main_planned_pci: Optional[int],
    polluter_pcis: List[int],
    serving_sequence: List[int],
    best_pci_sequence: List[int],
    serving_match_pct: int,
    overlap_pct: int,
) -> Tuple[str, str]:
    if main_planned_pci is None:
        return ("unknown", "No clear planned server identified from BDD geometry")
    planned_is_polluter = main_planned_pci in polluter_pcis
    planned_in_serving = main_planned_pci in serving_sequence
    planned_is_best = main_planned_pci in best_pci_sequence
    # Class A: planned server actually dominates
    if serving_match_pct >= 55 and planned_is_best and not planned_is_polluter:
        return ("A", "Planned server dominates but co-channel cells degrade quality")
    if serving_match_pct >= 55:
        return ("A", "Planned server is dominant; measured pollution is from nearby cells")
    # Class B/C: dominance failure caused by planned overlap (most severe combined case)
    if serving_match_pct < 35 and overlap_pct >= 55:
        return ("B/C", "Planned-server dominance failure with high planned overlap")
    # Class C: overlap is the primary problem, dominance is partial
    if overlap_pct >= 55:
        return ("C", "Planned overlap — BDD design does not create clean dominance at this location")
    # Class D: planned server absent from both serving and best sequences
    if serving_match_pct < 35 and not planned_in_serving and not planned_is_best:
        return ("D", "Planned server is not dominant — possible overshoot by another cell or coverage gap")
    # Class B: serving match is low but overlap is not the main driver
    if serving_match_pct < 50:
        return ("B", "Planned server does not consistently dominate — unstable dominance")
    return ("B", "Partial planned-server dominance — further analysis recommended")


def _polluter_planned_role(
    pci: int,
    main_planned_pci: Optional[int],
    ranked: List[Dict[str, Any]],
    polluter_pcis: List[int],
    planned_score: float,
) -> str:
    if pci == main_planned_pci:
        return "planned_server"
    # Second planned candidate: second in ranked list with score close to top
    if len(ranked) >= 2 and ranked[1]["pci"] == pci and planned_score >= _PLANNED_OVERLAP_MIN_SECOND:
        return "planned_overlap_candidate"
    if planned_score >= 70:
        return "planned_overlap_candidate"
    lobe = next((r.get("lobeClass") for r in ranked if r["pci"] == pci), None)
    is_measured_polluter = pci in polluter_pcis
    if planned_score < 35:
        if lobe == "back_lobe":
            return "back_lobe_or_bdd_error"
        if lobe == "side_lobe":
            return "unexpected_side_lobe" if is_measured_polluter else "side_lobe_candidate"
        return "unexpected_polluter" if is_measured_polluter else "low_score_candidate"
    return "secondary_candidate"


def _analyze_planned_server(
    route_points: List[Dict[str, Any]],
    bdd_candidates: List[Dict[str, Any]],
    earfcn: int,
    polluter_pcis: List[int],
    serving_sequence: List[int],
    best_pci_sequence: List[int],
) -> Dict[str, Any]:
    """Compute per-point planned-server scores and aggregate conflict classification."""
    if not bdd_candidates:
        return {"ok": False, "reason": "no_bdd_candidates_on_earfcn"}
    valid_pts = [
        p for p in route_points
        if _safe_float(p.get("lat")) is not None and _safe_float(p.get("lon")) is not None
    ]
    if not valid_pts:
        return {"ok": False, "reason": "no_valid_route_points"}

    per_point: List[Dict[str, Any]] = []
    for pt in valid_pts:
        lat = float(pt["lat"])
        lon = float(pt["lon"])
        serving_pci = _safe_int(pt.get("servingPci"))
        scored = [_score_cell_for_point(c, lat, lon, earfcn) for c in bdd_candidates]
        scored.sort(key=lambda x: -x["plannedScore"])
        top = scored[0] if scored else None
        second = scored[1] if len(scored) > 1 else None
        planned_overlap = bool(
            top and second
            and top["plannedScore"] >= _PLANNED_MIN_SCORE
            and (top["plannedScore"] - second["plannedScore"]) <= _PLANNED_OVERLAP_GAP
            and second["plannedScore"] >= _PLANNED_OVERLAP_MIN_SECOND
        )
        per_point.append({
            "tMs": pt.get("tMs"),
            "plannedPci": top["pci"] if top else None,
            "plannedScore": top["plannedScore"] if top else None,
            "secondPlannedPci": second["pci"] if second else None,
            "secondScore": second["plannedScore"] if second else None,
            "plannedOverlap": planned_overlap,
            "servingPci": serving_pci,
            "servingMatchesPlanned": (
                serving_pci is not None and top is not None and serving_pci == top["pci"]
            ),
        })

    total = len(per_point)
    planned_votes: Dict[int, int] = {}
    serving_match_count = 0
    overlap_count = 0
    for pt in per_point:
        pci = pt["plannedPci"]
        if pci is not None:
            planned_votes[pci] = planned_votes.get(pci, 0) + 1
        if pt["servingMatchesPlanned"]:
            serving_match_count += 1
        if pt["plannedOverlap"]:
            overlap_count += 1

    main_planned_pci = max(planned_votes, key=lambda p: planned_votes[p]) if planned_votes else None
    serving_match_pct = round(serving_match_count / total * 100) if total else 0
    overlap_pct = round(overlap_count / total * 100) if total else 0

    # Geometry of main planned cell at event centroid
    avg_lat = sum(float(p["tMs"] and p.get("tMs", 0) or 0) for p in per_point)  # dummy, recalc below
    avg_lat = sum(float(pt["lat"]) for pt in valid_pts) / len(valid_pts)
    avg_lon = sum(float(pt["lon"]) for pt in valid_pts) / len(valid_pts)

    main_info: Dict[str, Any] = {}
    if main_planned_pci is not None:
        for cell in bdd_candidates:
            if cell["pci"] == main_planned_pci:
                s = _score_cell_for_point(cell, avg_lat, avg_lon, earfcn)
                main_info = {
                    "mainPlannedCell": cell.get("cell_name"),
                    "mainPlannedSite": cell.get("site_name"),
                    "mainPlannedScore": s["plannedScore"],
                    "mainPlannedDistM": s["distM"],
                    "mainPlannedAzErr": s["azimuthErrorDeg"],
                    "mainPlannedLobe": s["lobeClass"],
                }
                break

    # Ranked candidate table at centroid (deduplicated by PCI, best score)
    best_per_pci: Dict[int, Dict[str, Any]] = {}
    for cell in bdd_candidates:
        pci = cell["pci"]
        s = _score_cell_for_point(cell, avg_lat, avg_lon, earfcn)
        if pci not in best_per_pci or s["plannedScore"] > best_per_pci[pci]["plannedScore"]:
            best_per_pci[pci] = s
    ranked = sorted(best_per_pci.values(), key=lambda x: -x["plannedScore"])
    for r in ranked:
        r["plannedRole"] = _polluter_planned_role(
            r["pci"], main_planned_pci, ranked, polluter_pcis, r["plannedScore"]
        )
        r.pop("_cell_ref", None)

    conflict_class, conflict_label = _classify_planned_conflict(
        main_planned_pci, polluter_pcis, serving_sequence, best_pci_sequence,
        serving_match_pct, overlap_pct,
    )

    return {
        "ok": True,
        "mainPlannedPci": main_planned_pci,
        "mainPlannedSamplePct": round(planned_votes.get(main_planned_pci, 0) / total * 100) if main_planned_pci and total else 0,
        "servingMatchPct": serving_match_pct,
        "plannedOverlapPct": overlap_pct,
        "conflictClass": conflict_class,
        "conflictLabel": conflict_label,
        "rankedCandidates": ranked[:6],
        **main_info,
    }


# ── Main Phase 2 entry point ─────────────────────────────────────────────────
def match_pollution_event(
    event_lat: float,
    event_lon: float,
    polluters: List[Dict[str, Any]],
    earfcn: int,
    max_dist_m: float = 15_000.0,
    rat: str = "LTE",
    route_points: Optional[List[Dict[str, Any]]] = None,
    serving_sequence: Optional[List[int]] = None,
    best_pci_sequence: Optional[List[int]] = None,
) -> Dict[str, Any]:
    """
    Phase 2: match pollution event polluters to BDD physical sectors.

    polluters: list of dicts from Phase 1 event.topPolluters, each containing at minimum:
        pci, averageDeltaToBest, nonServingPolluterSampleCount, corePolluterSampleCount

    Returns a dict with 'ok', 'polluters' (list of per-PCI attribution rows).
    """
    if not _BDD_CELLS:
        return {"ok": False, "error": "BDD not loaded. Use POST /api/bdd/configure to load a BDD file."}

    results: List[Dict[str, Any]] = []
    for polluter in polluters:
        pci = _safe_int(polluter.get("pci"))
        if pci is None:
            continue
        avg_delta = float(polluter.get("averageDeltaToBest") or 99.0)
        non_serving_count = int(polluter.get("nonServingPolluterSampleCount") or 0)
        core_hits = int(polluter.get("corePolluterSampleCount") or 0)
        score = float(polluter.get("polluterScore") or 0.0)

        candidates = find_candidates(earfcn, pci, event_lat, event_lon, max_dist_m, rat)

        if not candidates:
            results.append(
                {
                    "pci": pci,
                    "earfcn": earfcn,
                    "matched": False,
                    "candidateCount": 0,
                    "avgDeltaToBest": avg_delta,
                    "coreHits": core_hits,
                    "nonServingCount": non_serving_count,
                    "polluterScore": score,
                    "mappingConfidence": 0.0,
                    "error": f"No BDD cell found within {int(max_dist_m / 1000)} km for EARFCN {earfcn} PCI {pci}",
                }
            )
            continue

        geo_list = [
            geometry_for_candidate(c, event_lat, event_lon, avg_delta, non_serving_count, earfcn)
            for c in candidates[:5]
        ]
        best_lobe = geo_list[0].get("lobeClass", "unknown") if geo_list else "unknown"
        conf = _mapping_confidence(len(candidates), best_lobe)

        cand_rows = [
            {
                "siteName": c.get("site_name"),
                "cellName": c.get("cell_name"),
                "cgiEci": c.get("cgi_eci"),
                "lat": c.get("lat"),
                "lon": c.get("lon"),
                "azimuth": c.get("azimuth"),
                "totalTilt": c.get("total_tilt"),
                "antennaHeight": c.get("antenna_height"),
                "hBeamwidth": c.get("h_beamwidth"),
                "band": c.get("band"),
                **geo,
            }
            for c, geo in zip(candidates[:5], geo_list)
        ]

        results.append(
            {
                "pci": pci,
                "earfcn": earfcn,
                "matched": True,
                "candidateCount": len(candidates),
                "avgDeltaToBest": avg_delta,
                "coreHits": core_hits,
                "nonServingCount": non_serving_count,
                "polluterScore": score,
                "mappingConfidence": conf,
                "bestMatch": cand_rows[0] if cand_rows else None,
                "allCandidates": cand_rows,
            }
        )

    # ── Planned-server analysis (requires route_points) ───────────────────────
    planned_server: Optional[Dict[str, Any]] = None
    if route_points:
        earfcn_candidates = _find_active_cells_on_earfcn(
            earfcn, event_lat, event_lon, max_dist_m, rat
        )
        polluter_pcis = [int(p.get("pci") or 0) for p in polluters if p.get("pci") is not None]
        serving_seq = serving_sequence or []
        best_seq = best_pci_sequence or []
        planned_server = _analyze_planned_server(
            route_points, earfcn_candidates, earfcn,
            polluter_pcis, serving_seq, best_seq,
        )

    return {
        "ok": True,
        "eventLat": event_lat,
        "eventLon": event_lon,
        "earfcn": earfcn,
        "bddCellCount": _BDD_CELL_COUNT,
        "polluters": results,
        "plannedServer": planned_server,
    }
